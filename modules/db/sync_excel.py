"""modules/db/sync_excel.py — Carga el MASTER a la base (modo paralelo).

El Excel sigue siendo la fuente de verdad. Esta sincronización es de UNA
dirección (Excel → base) e idempotente: borra y recarga cada tabla, así nunca
quedan restos de una corrida anterior.
"""
import logging
import re
import unicodedata
from datetime import date, datetime

from openpyxl import load_workbook

from config import EXCEL_PATH
from modules.db.models import (
    crear_esquema, get_engine, get_session, SyncLog,
    Factura, FacturaLinea, MovimientoBanco, Conciliacion, Bitacora,
    Inventario, Personal, Proveedor,
)

logger = logging.getLogger(__name__)


# ── helpers ──
def _d(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for f in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(v.strip()[:10], f).date()
            except ValueError:
                continue
    return None


def _n(v):
    """Número o None (no 0: distinguir 'vacío' de 'cero')."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("$", "").replace(" ", "").strip()
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _s(v, limite=None):
    s = "" if v is None else str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s[:limite] if limite else s


def _norm(s):
    s = str(s or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    s = re.sub(r"\b(spa|ltda|s\.?a\.?|eirl|limitada|y cia|cia)\b", " ", s)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _nrokey(v):
    s = _s(v).upper().replace(" ", "").replace("-", "")
    return re.sub(r"^(F|FND|FNC|ND|NC)", "", s) or s


def sincronizar(excel_path: str = None, verbose: bool = True) -> dict:
    """Recarga la base completa desde el Excel. Devuelve conteos por tabla."""
    excel_path = excel_path or EXCEL_PATH
    engine = crear_esquema()
    ses = get_session(engine)
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    res = {}

    def log(msg):
        if verbose:
            print(msg)

    # ── limpiar (recarga completa) ──
    for modelo in (FacturaLinea, Factura, MovimientoBanco, Conciliacion,
                   Bitacora, Inventario, Personal, Proveedor):
        ses.query(modelo).delete()
    ses.commit()

    # ── Facturas (cabecera + líneas) ──
    if "Facturas" in wb.sheetnames:
        ws = wb["Facturas"]
        cabeceras = {}      # (prov_norm, nro_norm) -> Factura
        lineas = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:
                continue
            nro = _s(row[6])
            prov = _s(row[3], 160)
            if not nro:
                continue
            k = (_norm(prov), _nrokey(nro))
            f = cabeceras.get(k)
            if f is None:
                f = Factura(
                    proveedor=prov, proveedor_norm=k[0], rut=_s(row[4], 20),
                    documento=_s(row[5], 60), numero=nro, numero_norm=k[1],
                    fecha_emision=_d(row[0]), fecha_vencimiento=_d(row[1]),
                    fecha_pago=_d(row[2]), total=_n(row[15]),
                    categoria=_s(row[16], 60), cultivo=_s(row[17], 30),
                    n_archivo=int(row[20]) if len(row) > 20 and isinstance(row[20], (int, float)) else None,
                    fila_excel=i)
                cabeceras[k] = f
                ses.add(f)
            else:
                if f.fecha_pago is None:
                    f.fecha_pago = _d(row[2])
                t = _n(row[15])
                if t and (f.total is None or t > float(f.total)):
                    f.total = t
            lineas.append((k, FacturaLinea(
                glosa=_s(row[7], 300), glosa_detalle=_s(row[8]),
                cantidad=_n(row[10]), valor_unitario=_n(row[9]),
                neto=_n(row[11]), iva=_n(row[12]),
                impuesto_especifico=_n(row[13]), total_item=_n(row[14]),
                fila_excel=i)))
        ses.flush()                       # asigna IDs a las cabeceras
        for k, ln in lineas:
            ln.factura_id = cabeceras[k].id
            ses.add(ln)
        ses.commit()
        res["facturas"] = len(cabeceras)
        res["factura_lineas"] = len(lineas)
        log(f"  facturas: {len(cabeceras)} documentos / {len(lineas)} líneas")

    # ── Cuenta Banco ──
    if "Cuenta Banco" in wb.sheetnames:
        ws = wb["Cuenta Banco"]
        n = 0
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:
                continue
            f = _d(row[0])
            if not f:
                continue
            ses.add(MovimientoBanco(
                fecha=f, descripcion=_s(row[1], 300), referencia=_s(row[2], 40),
                cargo=_n(row[3]), abono=_n(row[4]), saldo=_n(row[5]),
                tipo=_s(row[6], 30), categoria=_s(row[7], 60),
                cultivo=_s(row[8], 30), fila_excel=i))
            n += 1
        ses.commit()
        res["cuenta_banco"] = n
        log(f"  cuenta_banco: {n}")

    # ── Conciliaciones ──
    if "Conciliaciones" in wb.sheetnames:
        ws = wb["Conciliaciones"]
        n = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            ses.add(Conciliacion(
                id=int(row[0]), fecha_conciliacion=_d(row[1]),
                fila_banco=int(row[2]) if row[2] else None,
                tipo_doc=_s(row[6], 20), numero_doc=_s(row[8], 40),
                proveedor=_s(row[9], 160), monto_asignado=_n(row[10]),
                criterio=_s(row[11], 30), usuario=_s(row[12], 40),
                nota=_s(row[13], 200)))
            n += 1
        ses.commit()
        res["conciliaciones"] = n
        log(f"  conciliaciones: {n}")

    # ── Bitácora ──
    if "Bitácora" in wb.sheetnames:
        ws = wb["Bitácora"]
        n = 0
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:
                continue
            f = _d(row[0])
            if not f:
                continue
            g = lambda idx: row[idx] if len(row) > idx else None
            ses.add(Bitacora(
                fecha=f, hora=_s(row[1], 8), tipo=_s(row[2], 20),
                actividad=_s(row[3], 160), cultivo=_s(row[4], 30),
                sector=_s(row[5], 80), jornadas_hombre=_n(row[6]),
                trabajadores=_s(row[7], 300), insumo=_s(row[8], 120),
                cantidad=_n(row[9]), unidad=_s(row[10], 12),
                registro=_s(row[11]), registrado_por=_s(row[12], 60),
                maquina=_s(g(13), 60), odometro=_n(g(14)),
                horas_dia=_n(g(15)), superficie_ha=_n(g(16)), fila_excel=i))
            n += 1
        ses.commit()
        res["bitacora"] = n
        log(f"  bitacora: {n}")

    # ── Inventario ──
    if "Inventario" in wb.sheetnames:
        ws = wb["Inventario"]
        n = 0
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:
                continue
            g = lambda idx: row[idx] if len(row) > idx else None
            ses.add(Inventario(
                producto=_s(row[0], 120), categoria=_s(row[1], 60),
                unidad=_s(row[2], 12), stock=_n(row[3]), stock_minimo=_n(row[4]),
                ultima_entrada=_d(row[5]), ultimo_uso=_d(row[6]),
                vencimiento=_d(g(7)), estado=_s(g(8), 20), fila_excel=i))
            n += 1
        ses.commit()
        res["inventario"] = n
        log(f"  inventario: {n}")

    # ── Proveedores ──
    if "Proveedores" in wb.sheetnames:
        ws = wb["Proveedores"]
        n = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[1]:
                continue
            ses.add(Proveedor(nombre=_s(row[1], 160), rut=_s(row[2], 20),
                              nombre_norm=_norm(row[1])))
            n += 1
        ses.commit()
        res["proveedores"] = n
        log(f"  proveedores: {n}")

    # ── Personal ──
    if "Personal" in wb.sheetnames:
        ws = wb["Personal"]
        n = 0
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:
                continue
            ses.add(Personal(nombre=_s(row[0], 120), rut=_s(row[1], 20),
                             cargo=_s(row[2], 80), fecha_ingreso=_d(row[3]),
                             fila_excel=i))
            n += 1
        ses.commit()
        res["personal"] = n
        log(f"  personal: {n}")

    wb.close()
    for tabla, filas in res.items():
        ses.add(SyncLog(tabla=tabla, filas=filas, detalle="sync desde Excel"))
    ses.commit()
    ses.close()
    logger.info(f"Sync Excel→DB: {res}")
    return res
