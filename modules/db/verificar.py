"""modules/db/verificar.py — Compara el Excel contra la base.

Es la red de seguridad del modo paralelo: mientras el Excel siga siendo la
fuente de verdad, esto confirma que la base dice exactamente lo mismo. Cuando
pase limpio de forma sostenida, se puede invertir la dirección.
"""
import logging
from datetime import date, datetime

from openpyxl import load_workbook
from sqlalchemy import func

from config import EXCEL_PATH
from modules.db.models import (get_session, Factura, FacturaLinea,
                                MovimientoBanco, Bitacora, Inventario)

logger = logging.getLogger(__name__)
TOL = 1.0     # tolerancia en pesos


def _d(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for f in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(v.strip()[:10], f).date()
            except ValueError:
                continue
    return None


def _n(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def comparar(excel_path: str = None) -> dict:
    excel_path = excel_path or EXCEL_PATH
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ses = get_session()
    checks = []

    def chk(nombre, excel_val, db_val, es_monto=True):
        dif = (db_val or 0) - (excel_val or 0)
        ok = abs(dif) <= (TOL if es_monto else 0)
        checks.append({"check": nombre, "excel": excel_val, "db": db_val,
                        "dif": dif, "ok": ok})

    # ── Facturas ──
    if "Facturas" in wb.sheetnames:
        ws = wb["Facturas"]
        docs, suma_items = set(), 0.0
        n_lineas = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            nro = str(row[6] or "").strip()
            if not nro:
                continue
            n_lineas += 1
            docs.add((str(row[3] or "").strip().lower(), nro))
            suma_items += _n(row[14])
        chk("Facturas · líneas", n_lineas,
            ses.query(func.count(FacturaLinea.id)).scalar(), es_monto=False)
        chk("Facturas · suma Total por Item", round(suma_items),
            round(_n(ses.query(func.sum(FacturaLinea.total_item)).scalar())))

    # ── Cuenta Banco ──
    if "Cuenta Banco" in wb.sheetnames:
        ws = wb["Cuenta Banco"]
        n, cargos, abonos = 0, 0.0, 0.0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not _d(row[0]):
                continue
            n += 1
            cargos += _n(row[3])
            abonos += _n(row[4])
        chk("Banco · movimientos", n,
            ses.query(func.count(MovimientoBanco.id)).scalar(), es_monto=False)
        chk("Banco · total cargos", round(cargos),
            round(_n(ses.query(func.sum(MovimientoBanco.cargo)).scalar())))
        chk("Banco · total abonos", round(abonos),
            round(_n(ses.query(func.sum(MovimientoBanco.abono)).scalar())))

    # ── Bitácora ──
    if "Bitácora" in wb.sheetnames:
        ws = wb["Bitácora"]
        n, jh = 0, 0.0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not _d(row[0]):
                continue
            n += 1
            jh += _n(row[6])
        chk("Bitácora · registros", n,
            ses.query(func.count(Bitacora.id)).scalar(), es_monto=False)
        chk("Bitácora · jornadas-hombre", round(jh, 1),
            round(_n(ses.query(func.sum(Bitacora.jornadas_hombre)).scalar()), 1))

    # ── Inventario ──
    if "Inventario" in wb.sheetnames:
        ws = wb["Inventario"]
        n = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0])
        chk("Inventario · productos", n,
            ses.query(func.count(Inventario.id)).scalar(), es_monto=False)

    wb.close()
    ses.close()
    fallidos = [c for c in checks if not c["ok"]]
    return {"checks": checks, "ok": not fallidos, "fallidos": len(fallidos)}


def formato(res: dict) -> str:
    lines = ["🔍 EXCEL vs BASE DE DATOS", ""]
    lines.append(f"{'Verificación':34} {'Excel':>16} {'Base':>16} {'Dif':>12}")
    lines.append("-" * 82)
    for c in res["checks"]:
        marca = "✅" if c["ok"] else "❌"
        lines.append(f"{marca} {c['check']:31} {c['excel']:>16,.0f} "
                      f"{c['db']:>16,.0f} {c['dif']:>12,.0f}")
    lines.append("")
    lines.append("✅ Todo calza — la base refleja el Excel." if res["ok"]
                 else f"❌ {res['fallidos']} verificación(es) no calzan.")
    return "\n".join(lines)
