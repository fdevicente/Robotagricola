"""modules/correlativo.py — Número correlativo de archivo físico (columna N de FXP).

El usuario numera cada factura en FXP (col D "N") y las archiva impresas con ese
número. Este módulo lleva ese correlativo también al Master, para poder buscar
el papel rápido:
  - facturas que ya están en FXP → se les asigna SU número de FXP;
  - facturas nuevas → se les asigna el siguiente número libre.
"""
import logging
import os
import re
import shutil
import tempfile
import unicodedata

from openpyxl import load_workbook

from config import EXCEL_PATH

logger = logging.getLogger(__name__)

FXP_PATH = os.getenv("FXP_PATH",
                     r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx")
COL_CORRELATIVO = 21          # columna U en Master.Facturas
HEADER_CORRELATIVO = "N° Archivo"


def norm_prov(s):
    s = str(s or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    s = re.sub(r"\b(spa|ltda|s\.?a\.?|eirl|limitada|y cia|cia)\b", " ", s)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def nrokey(v):
    """Normaliza el nº de documento. 'FND9792' y 'F9792' comparten raíz."""
    s = str(v or "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = s.upper().replace(" ", "").replace("-", "")
    return re.sub(r"^(F|FND|FNC|ND|NC)", "", s) or s


# Mismo proveedor escrito distinto en el Master y en FXP (verificado por monto
# idéntico en ambas planillas). Cada grupo = nombres equivalentes.
ALIAS_PROVEEDOR = [
    {"indelec", "hugo roco gajardo", "roco gajardo"},
    {"gmt", "lira graham", "lira graham y cia", "gmc chile"},
    {"irrifor", "irrifer"},
    {"inia", "instituto de investigacion", "instituto de investigaciones agropecuarias"},
    {"reylux", "servicios ambientales rene", "rene alejandro bravo novoa"},
    {"mercadolibre chile", "mercado libre", "mercadolibre"},
    {"comercializadora g y p", "comerc g y p"},
    {"roda asesorias profesionales", "rodasep", "roda asesorias"},
    {"copec", "soc comerc nueva loncomilla", "comercial dominga"},
    {"coragas", "cora"},
    {"ferreteria industrial talca", "ferrital", "ferreteriaindutrial talca"},
    {"certificacion de equipos ramirez", "certlab"},
    {"factoraurorania ductil", "imp pcplay"},
]
_ALIAS_IDX = {}
for _g in ALIAS_PROVEEDOR:
    for _n in _g:
        _ALIAS_IDX[_n] = id(_g)


def tokens_prov(s) -> set:
    """Palabras significativas del nombre de un proveedor."""
    return {t for t in re.split(r"[^a-z0-9]+", norm_prov(s)) if len(t) >= 4}


def _grupo_alias(nombre):
    n = norm_prov(nombre)
    if n in _ALIAS_IDX:
        return _ALIAS_IDX[n]
    for alias, gid in _ALIAS_IDX.items():
        if n.startswith(alias) or alias.startswith(n[:14]) and len(n) >= 5:
            return gid
    return None


def mismo_proveedor(a, b) -> bool:
    """True si dos nombres corresponden claramente al mismo proveedor."""
    ga, gb = _grupo_alias(a), _grupo_alias(b)
    if ga and gb:
        return ga == gb
    ta, tb = tokens_prov(a), tokens_prov(b)
    if ta and tb:
        return bool(ta & tb)
    na, nb = norm_prov(a), norm_prov(b)
    return bool(na and nb and (na.startswith(nb[:5]) or nb.startswith(na[:5])))


def buscar_en_fxp(proveedor, nro_documento, indice: dict):
    """Busca una factura en FXP por (proveedor, nº).

    `indice` = {nro_normalizado: [registros...]}. Un mismo número puede existir
    para varios proveedores (F94 de Misael y F94 de Contreras son distintas),
    por eso se elige el candidato cuyo PROVEEDOR calce; si ninguno calza y hay
    un solo candidato con nombre parecido, se devuelve None (mejor no adivinar).
    """
    candidatos = indice.get(nrokey(nro_documento)) or []
    if not candidatos:
        return None
    if len(candidatos) == 1 and mismo_proveedor(proveedor, candidatos[0].get("prov", "")):
        return candidatos[0]
    for c in candidatos:
        if mismo_proveedor(proveedor, c.get("prov", "")):
            return c
    return None


def leer_correlativos_fxp(path: str = None) -> tuple[dict, dict, int]:
    """Devuelve (por_prov_nro, por_nro, maximo) desde la hoja FXP."""
    path = path or FXP_PATH
    if not os.path.exists(path):
        raise FileNotFoundError(f"No encuentro FXP: {path}")
    tmp = os.path.join(tempfile.gettempdir(), "fxp_corr_read.xlsx")
    shutil.copy2(path, tmp)
    wb = load_workbook(tmp, read_only=True, data_only=True)
    ws = wb["FXP"]
    por_prov_nro, por_nro = {}, {}
    maximo = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        n = row[3]
        if not isinstance(n, (int, float)):
            continue
        n = int(n)
        maximo = max(maximo, n)
        nro = nrokey(row[7])
        if not nro:
            continue
        por_prov_nro[(norm_prov(row[6]), nro)] = n
        # varios proveedores pueden compartir el mismo nº de documento
        por_nro.setdefault(nro, []).append({"prov": str(row[6] or ""), "n": n})
    wb.close()
    return por_prov_nro, por_nro, maximo


def correlativo_en_indice(proveedor, nro, por_prov_nro, por_nro):
    """N° de archivo respetando el proveedor (no solo el nº de documento)."""
    n = por_prov_nro.get((norm_prov(proveedor), nrokey(nro)))
    if n:
        return n
    c = buscar_en_fxp(proveedor, nro, por_nro)
    return c["n"] if c else None


def _max_en_master(ws) -> int:
    m = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, COL_CORRELATIVO).value
        if isinstance(v, (int, float)):
            m = max(m, int(v))
    return m


def asegurar_columna(ws):
    if str(ws.cell(1, COL_CORRELATIVO).value or "").strip() != HEADER_CORRELATIVO:
        ws.cell(1, COL_CORRELATIVO, HEADER_CORRELATIVO)


def correlativo_para(proveedor, nro_documento) -> int | None:
    """N° de archivo de una factura, SOLO si ya está numerada en FXP.

    FXP es la fuente del correlativo: el número lo asigna el usuario al
    registrarla ahí. Si la factura todavía no está en FXP se devuelve None
    (queda en blanco) y luego `sincronizar_desde_fxp()` la completa.
    """
    try:
        por_prov_nro, por_nro, _ = leer_correlativos_fxp()
    except Exception as e:
        logger.warning(f"Correlativo: no pude leer FXP ({e})")
        return None
    return correlativo_en_indice(proveedor, nro_documento, por_prov_nro, por_nro)


def sincronizar_desde_fxp() -> dict:
    """Completa el N° de archivo en el Master para las facturas ya numeradas en FXP.

    Pensado para correr cada cierto tiempo: a medida que el usuario registra las
    facturas en FXP, el Master va tomando esos números.
    """
    por_prov_nro, por_nro, _ = leer_correlativos_fxp()
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Facturas"]
    asegurar_columna(ws)

    nuevos, corregidos, pendientes = 0, 0, 0
    detalle = []
    grupos = {}
    for r in range(2, ws.max_row + 1):
        if not ws.cell(r, 1).value:
            continue
        nro = nrokey(ws.cell(r, 7).value)
        if not nro:
            continue
        grupos.setdefault((norm_prov(ws.cell(r, 4).value), nro), []).append(r)

    for (prov, nro), filas in grupos.items():
        n = correlativo_en_indice(ws.cell(filas[0], 4).value, nro,
                                   por_prov_nro, por_nro)
        actual = ws.cell(filas[0], COL_CORRELATIVO).value
        if n is None:
            if actual not in (None, ""):
                for r in filas:
                    ws.cell(r, COL_CORRELATIVO).value = None
            pendientes += 1
            continue
        if actual == n:
            continue
        for r in filas:
            ws.cell(r, COL_CORRELATIVO).value = n
        if actual in (None, ""):
            nuevos += 1
            detalle.append((n, str(ws.cell(filas[0], 4).value or "")[:30],
                            str(ws.cell(filas[0], 7).value or "")))
        else:
            corregidos += 1

    from excel_manager import _save_wb
    _save_wb(wb)
    wb.close()
    logger.info(f"Correlativos sincronizados desde FXP: +{nuevos}, {corregidos} corregidos")
    return {"nuevos": nuevos, "corregidos": corregidos,
            "sin_numero": pendientes, "detalle": detalle}
