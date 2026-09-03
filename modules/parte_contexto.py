# -*- coding: utf-8 -*-
"""Lo que el bot ya sabe y la IA necesita para normalizar.

Se arma una vez por mensaje y se le pasa IGUAL al lector y al juez: si midieran
contra vocabularios distintos, el juez marcaria como duda lo que la IA hizo bien.

OJO CON LOS TRABAJADORES: no salen solo de la hoja Personal. Medido el
2-sep-2026, Personal tiene 6 filas con el nombre legal completo ("Felicito
Amigo Soto") y no incluye a Richard Padilla ni a su hijo, mientras la columna
Trabajadores de la bitacora usa los 8 nombres canonicos que el bot viene usando.
Manda el vocabulario de la bitacora; Personal solo agrega a los recien dados de
alta que todavia no tienen ninguna fila.
"""
import logging

logger = logging.getLogger(__name__)

BITACORA_SHEET = "Bitácora"
PERSONAL_SHEET = "Personal"


def construir(excel_path: str | None = None) -> dict:
    """Devuelve {"trabajadores": [...], "alias": {...}, "maquinas": [...]}."""
    from openpyxl import load_workbook

    from config import EXCEL_PATH
    from modules.bitacora_extractor import ALIAS, TRABAJADORES_CONOCIDOS
    from modules.maquinaria import maquinas_conocidas

    ruta = excel_path or EXCEL_PATH
    nombres, orden = set(), []

    def _sumar(n):
        n = str(n or "").strip()
        if n and n not in nombres:
            nombres.add(n)
            orden.append(n)

    try:
        wb = load_workbook(ruta, read_only=True, data_only=True)
        try:
            if BITACORA_SHEET in wb.sheetnames:
                ws = wb[BITACORA_SHEET]
                enc = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                if "Trabajadores" in enc:
                    i = enc.index("Trabajadores")
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if len(row) > i and row[i]:
                            for n in str(row[i]).split(","):
                                _sumar(n)
        finally:
            wb.close()
    except Exception as e:                    # un Excel raro no puede voltear esto
        logger.warning("parte_contexto: no pude leer %s: %s", ruta, e)

    for n in TRABAJADORES_CONOCIDOS:          # los de siempre, con sus apodos
        _sumar(n)

    try:
        wb = load_workbook(ruta, read_only=True, data_only=True)
        try:
            if PERSONAL_SHEET in wb.sheetnames:
                for row in wb[PERSONAL_SHEET].iter_rows(min_row=2, max_col=1,
                                                        values_only=True):
                    if row:
                        _sumar(row[0])
        finally:
            wb.close()
    except Exception as e:                    # un Excel raro no puede voltear esto
        logger.warning("parte_contexto: no pude leer Personal en %s: %s", ruta, e)

    try:
        maquinas = maquinas_conocidas(ruta)
    except Exception as e:
        logger.warning("parte_contexto: no pude leer las máquinas: %s", e)
        maquinas = []

    return {"trabajadores": orden, "alias": dict(ALIAS), "maquinas": maquinas}
