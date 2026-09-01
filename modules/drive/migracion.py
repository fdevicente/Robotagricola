# -*- coding: utf-8 -*-
"""Dónde va cada archivo en la migración única de los documentos a Drive.

La lógica vive acá y no en el script para poder probarla: el script se corre
una sola vez y a mano, así que es justo donde un error no se nota hasta que
ya movió 800 archivos.
"""
import os

# carpeta local -> carpeta en Drive
DESTINOS = {
    "Facturas Recibidas": "Facturas Recibidas",
    "Facturas Enviadas": "Facturas Enviadas",
    "BH": "Boletas Honorarios",
    "Guias de Despacho": "Guías de Despacho",
    "Rendiciones": "Rendiciones",
    "Legal": "Legal",
    "Carpeta Tributaria": "Tributario",
}

# `Facturas Recibidas` se ordena por año de emisión, no por la subcarpeta en
# que Telegram las fue dejando. Las demás conservan su estructura tal cual.
POR_ANIO = "Facturas Recibidas"
SIN_ANIO = "Sin año"

_BASURA = {".ds_store", "desktop.ini", "thumbs.db"}


def es_basura(nombre: str) -> bool:
    """Archivos que pone el sistema operativo y no son documentos de nadie."""
    return os.path.basename(nombre).casefold() in _BASURA


def destino_de(carpeta: str, ruta: str, origen: str, anio: str = None) -> str:
    """Ruta de la carpeta de Drive donde va este archivo.

    `carpeta` es la carpeta local que se está migrando, `ruta` el archivo y
    `origen` la raíz de esa carpeta en disco.
    """
    destino = DESTINOS[carpeta]
    if carpeta == POR_ANIO:
        return "%s/%s" % (destino, anio or SIN_ANIO)
    rel = os.path.relpath(os.path.dirname(ruta), origen)
    if rel == os.curdir:
        return destino
    return "%s/%s" % (destino, rel.replace(os.sep, "/"))


def indice_de_anios(excel_path: str) -> dict:
    """(nº de factura, proveedor) -> año de emisión, leído del Master.

    El año NO puede salir de la fecha del archivo: los 827 documentos tienen
    fecha de modificación 2026 porque Dropbox los resincronizó, mientras que
    las emisiones reales van de 2021 a 2026 — 393 son de 2025. Confiar en el
    mtime mandaba 404 documentos al año equivocado.

    Se indexa por número Y proveedor porque hay 19 números que usan más de un
    proveedor, y cada uno tiene su propia fecha.
    """
    from openpyxl import load_workbook
    from modules.drive.enlaces import _clave_proveedor, _normalizar
    wb = load_workbook(excel_path, read_only=True)
    try:
        indice = {}
        for fila in wb["Facturas"].iter_rows(min_row=2, values_only=True):
            numero = _normalizar(fila[6])
            if not numero:
                continue
            anio = _anio_de_celda(fila[0])
            if anio:
                indice[(numero, _clave_proveedor(fila[3]))] = anio
        return indice
    finally:
        wb.close()


def anio_de(indice: dict, nombre: str):
    """Año de emisión de este archivo, o None si el Master no lo conoce."""
    from modules.drive.enlaces import _clave_proveedor
    from modules.drive.subidor import _partes_del_nombre
    partes = _partes_del_nombre(nombre)
    if partes is None:
        return None
    proveedor, numero = partes
    return indice.get((numero, _clave_proveedor(proveedor)))


def _anio_de_celda(valor):
    """Año de una celda de fecha, venga como fecha o como texto.

    El Master tiene las dos formas mezcladas: 1.998 filas guardan un datetime y
    178 guardan la cadena '2026-05-05'. Leer solo `.year` mandaba esas 178 a
    "Sin año" sin motivo.
    """
    anio = getattr(valor, "year", None)
    if anio:
        return str(anio)
    texto = str(valor if valor is not None else "").strip()[:4]
    if texto.isdigit() and 1990 <= int(texto) <= 2100:
        return texto
    return None
