# -*- coding: utf-8 -*-
"""Escribe en la hoja Facturas el enlace del documento en Drive."""
import logging
import re

logger = logging.getLogger(__name__)

COL_PROVEEDOR, COL_NUMERO, COL_DRIVE = 4, 7, 22
URL = "https://drive.google.com/file/d/%s/view"


def guardar_enlace(excel_path: str, numero_factura: str, file_id: str,
                   proveedor: str = None) -> bool:
    """Pone el enlace en TODAS las líneas de esa factura. No pisa lo existente.

    `proveedor` desempata: hay 19 números de factura que usan dos o tres
    proveedores distintos, así que el número solo NO identifica una factura.
    Sin él, el enlace de ECOSMART terminaba en la fila de INV. SANTA VICTORIA.
    Cuando se pasa y no calza con ninguna fila, no se escribe nada: perder un
    enlace se repara con otra pasada, escribirlo en la factura equivocada no.
    """
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    try:
        ws = wb["Facturas"]
        objetivo = _normalizar(numero_factura)
        buscado = _clave_proveedor(proveedor) if proveedor else None
        tocadas = calzaron = 0
        for f in range(2, ws.max_row + 1):
            if _normalizar(ws.cell(f, COL_NUMERO).value) != objetivo:
                continue
            if buscado is not None and not _mismo_proveedor(
                    buscado, _clave_proveedor(ws.cell(f, COL_PROVEEDOR).value)):
                continue
            calzaron += 1
            if ws.cell(f, COL_DRIVE).value:
                continue
            ws.cell(f, COL_DRIVE).value = URL % file_id
            tocadas += 1
        if tocadas:
            wb.save(excel_path)          # ruta EXPLÍCITA siempre
        else:
            # Por qué no se escribió: son dos cosas muy distintas y el log
            # tiene que distinguirlas. 'Ya tenía enlace' es lo normal al
            # reenviar una foto; 'sin fila' es la factura que falta cargar.
            logger.info(
                "%s Nº%s: %s", proveedor or "(sin proveedor)",
                numero_factura,
                "la fila ya tenía enlace" if calzaron else
                "sin fila que calce en el Master")
        return tocadas > 0
    finally:
        wb.close()


def _normalizar(valor) -> str:
    """'2777', 2777 y 2777.0 son el mismo número de factura."""
    s = str(valor if valor is not None else "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _clave_proveedor(valor) -> str:
    """Deja comparables 'Silpa Sur Spa', 'Silpa_Sur_Spa' y 'SILPA SUR SPA'.

    El nombre del archivo se armó desde esta misma columna reemplazando los
    espacios por guiones bajos, así que para volver a calzarlos hay que
    deshacer ese reemplazo y olvidarse de las mayúsculas.
    """
    s = str(valor if valor is not None else "")
    return re.sub(r"[\s_]+", " ", s).strip().casefold()


# El nombre del ARCHIVO queda congelado el día que se sacó la foto; el del
# Master se corrige después. Al unificar "FERRETERIAINDUTRIAL TALCA LIMITADA"
# se arreglaron 7 archivos y se rompieron 8, que se llamaban con la grafía
# vieja. Por eso las dos grafías tienen que resolver a la misma clave.
# Cada par está comprobado con el RUT o con la factura, no por parecido.
_ALIAS = {
    "ferreteriaindutrial talca limitada": "ferreteria industrial talca limitada",
    "ferreteria industrial paghita spa": "ferreteria industrial pachita spa",
    "servicios y arriendos rotortec spa": "rotortec",
    "irrifer": "irrifor",
    "salina y fabres": "salinas y fabres",
}

# `handlers.facturas._limpiar` corta el nombre a 60 caracteres, así que un
# proveedor de nombre largo queda cortado en el archivo y entero en el Master.
_LARGO_MAXIMO = 60


def _mismo_proveedor(uno: str, otro: str) -> bool:
    """Si las dos claves son el mismo proveedor escrito distinto."""
    uno, otro = _ALIAS.get(uno, uno), _ALIAS.get(otro, otro)
    if uno == otro:
        return True
    # Un nombre de exactamente 60 viene cortado: puede ser el principio del
    # otro. Un prefijo más corto NO vale — ahí sí serían proveedores distintos.
    for corto, largo in ((uno, otro), (otro, uno)):
        if len(corto) == _LARGO_MAXIMO and largo.startswith(corto):
            return True
    return False
