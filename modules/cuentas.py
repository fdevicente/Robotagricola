"""Caja de la agrícola: cuenta corriente en pesos + cuenta en dólares.

A los exportadores se les cobra en USD, así que una parte del efectivo vive en
la cuenta dólar (Scotiabank ****9350) y no en la cuenta corriente (****2530).
Mirar solo la cuenta en pesos **subestima la caja** — eso llevó una vez a
concluir que un exportador no estaba pagando cuando sí lo hacía.

Regla: cualquier cálculo de caja o de flujo usa `caja_total()`.
"""
import logging
from datetime import date, datetime

from openpyxl import load_workbook

from config import CASH_FLOW_CONFIG, EXCEL_PATH

logger = logging.getLogger(__name__)

BANCO_SHEET = "Cuenta Banco"
DOLAR_SHEET = "Cuenta Dolar"
CONFIG_SHEET = "Config"

DOLAR_HEADERS = ["Fecha", "Descripcion", "Referencia", "Cargo USD",
                 "Abono USD", "Saldo USD", "Tipo", "Categoria", "Cultivo",
                 "Factura_linkeada"]

COL_FECHA, COL_SALDO = 0, 5


def _a_fecha(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(v)[:10], fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _ultimo_saldo(ws):
    """Saldo de la fila con fecha más reciente. Devuelve (saldo, fecha)."""
    saldo, fecha = 0.0, None
    for row in ws.iter_rows(min_row=2, max_col=COL_SALDO + 1, values_only=True):
        if not row or row[COL_FECHA] is None:
            continue
        f = _a_fecha(row[COL_FECHA])
        if not f:
            continue
        try:
            s = float(row[COL_SALDO] or 0)
        except (TypeError, ValueError):
            continue
        if fecha is None or f >= fecha:
            fecha, saldo = f, s
    return saldo, fecha


def _tipo_cambio(wb):
    """CLP por USD. Manda la hoja Config; si no está, el valor de config.py."""
    if CONFIG_SHEET in wb.sheetnames:
        for row in wb[CONFIG_SHEET].iter_rows(min_row=2, max_col=2, values_only=True):
            if row and str(row[0] or "").strip() == "usd_clp_estimado":
                try:
                    return float(row[1])
                except (TypeError, ValueError):
                    break
    return float(CASH_FLOW_CONFIG.get("usd_clp_estimado", 1000))


def caja_total(excel_path: str | None = None) -> dict:
    """Efectivo disponible sumando las dos cuentas.

    Devuelve montos, el tipo de cambio usado y la fecha del último movimiento
    de cada cuenta (sirve para avisar si una quedó desactualizada).
    """
    wb = load_workbook(excel_path or EXCEL_PATH, read_only=True, data_only=True)
    try:
        clp, fecha_clp = (_ultimo_saldo(wb[BANCO_SHEET])
                          if BANCO_SHEET in wb.sheetnames else (0.0, None))
        usd, fecha_usd = (_ultimo_saldo(wb[DOLAR_SHEET])
                          if DOLAR_SHEET in wb.sheetnames else (0.0, None))
        tc = _tipo_cambio(wb)
    finally:
        wb.close()

    usd_en_clp = usd * tc
    return {
        "clp": clp,
        "usd": usd,
        "tipo_cambio": tc,
        "usd_en_clp": usd_en_clp,
        "total": clp + usd_en_clp,
        "fecha_clp": fecha_clp,
        "fecha_usd": fecha_usd,
    }


DIAS_DESACTUALIZADA = 7


def desactualizadas(caja: dict | None = None, hoy: date | None = None) -> list[str]:
    """Cuentas cuyo último movimiento tiene más de `DIAS_DESACTUALIZADA` días.

    Una caja que se ve sana porque una cuenta quedó vieja es peor que no tener
    el dato: por eso se avisa en vez de sumar en silencio.
    """
    c = caja or caja_total()
    hoy = hoy or date.today()
    viejas = []
    for etiqueta, clave in (("cuenta corriente", "fecha_clp"),
                            ("cuenta dólar", "fecha_usd")):
        f = c.get(clave)
        if f is None:
            viejas.append(f"{etiqueta}: sin movimientos registrados")
        elif (hoy - f).days > DIAS_DESACTUALIZADA:
            viejas.append(f"{etiqueta}: último movimiento el {f} "
                           f"({(hoy - f).days} días)")
    return viejas


def formato(caja: dict | None = None) -> str:
    """Resumen legible de la caja, para Telegram o consola."""
    c = caja or caja_total()
    lineas = [
        f"💰 *Caja total: ${c['total']:,.0f}*",
        f"   Cuenta corriente  ${c['clp']:,.0f}"
        + (f"  (al {c['fecha_clp']})" if c["fecha_clp"] else ""),
        f"   Cuenta dólar      US${c['usd']:,.2f} = ${c['usd_en_clp']:,.0f}"
        + (f"  (al {c['fecha_usd']})" if c["fecha_usd"] else ""),
        f"   Tipo de cambio    {c['tipo_cambio']:,.0f} CLP/USD",
    ]
    viejas = desactualizadas(c)
    if viejas:
        lineas.append("")
        lineas.append("⚠️ Caja posiblemente desactualizada:")
        lineas += [f"   • {v}" for v in viejas]
    return "\n".join(lineas)
