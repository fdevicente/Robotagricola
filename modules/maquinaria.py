"""modules/maquinaria.py — Fichas, lecturas de horómetro y mantenciones.

Tres cosas que hasta ahora vivían sueltas:
  · la FICHA de cada máquina (marca, modelo, patente, serie) — no existía,
  · las LECTURAS de horómetro — se guardaban en la bitácora junto al trabajo,
  · las MANTENCIONES — solo estaban como facturas, sin saber a qué máquina.

Las lecturas siguen yendo a la Bitácora para no partir el historial de horas
(`bitacora_manager` calcula las horas por diferencia de odómetro).
"""
import logging
import re
import unicodedata
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from config import EXCEL_PATH
from excel_manager import _save_wb

logger = logging.getLogger(__name__)

FICHAS_SHEET = "Maquinaria"
FICHAS_HEADERS = ["Máquina", "Tipo", "Marca", "Modelo", "Año", "Patente",
                  "N° Serie", "Propiedad", "Mide", "Estado", "Notas",
                  "Actualizado"]
_FICHAS_ANCHOS = [30, 12, 16, 16, 7, 12, 20, 13, 10, 18, 34, 13]

MANT_SHEET = "Mantenciones"
MANT_HEADERS = ["ID", "Fecha", "Máquina", "Estado", "Tipo", "Descripción",
                "Odómetro", "Proveedor", "Costo", "N° Factura", "Próxima (h)",
                "Notas", "Registrado por"]
_MANT_ANCHOS = [6, 12, 28, 11, 18, 42, 11, 26, 13, 13, 12, 30, 16]

# HECHA = ya se realizó · PENDIENTE = hay que hacerla
ESTADOS_MANT = ("HECHA", "PENDIENTE")

# Tipos de mantención frecuentes (para normalizar lo que escriba Juan)
TIPOS_MANT = ["ACEITE MOTOR", "FILTROS", "HIDRAULICO", "TRANSMISION",
              "ENGRASE", "NEUMATICOS", "FRENOS", "REPARACION", "REVISION",
              "OTRO"]

# Las camionetas miden kilómetros; el resto, horas
MIDE_KM = ("SSANGYONG", "CAMIONETA")


def _sin_tildes(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", str(s or ""))
                   if unicodedata.category(c) != "Mn").upper()


def norm_maquina(nombre: str) -> str:
    """Nombre canónico de la máquina, tolerando cómo la escriba Juan."""
    t = " ".join(_sin_tildes(nombre).split())
    t = t.replace("JHON", "JOHN").replace("DEER ", "DEERE ")
    t = re.sub(r"\bJD\b", "JOHN DEERE", t)
    t = re.sub(r"\bMF\b", "MASSEY FERGUSON", t)
    t = re.sub(r"\bSSANG YONG\b", "SSANGYONG", t)
    return t.strip()


def unidad_de(maquina: str) -> str:
    return "km" if any(k in norm_maquina(maquina) for k in MIDE_KM) else "h"


def _crear(wb, nombre, headers, anchos, color="1F4E78"):
    if nombre in wb.sheetnames:
        return False
    ws = wb.create_sheet(nombre)
    fill = PatternFill("solid", fgColor=color)
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    return True


def crear_hojas(excel_path: str | None = None) -> None:
    """Crea las hojas Maquinaria y Mantenciones si faltan. Idempotente."""
    ruta = excel_path or EXCEL_PATH
    wb = load_workbook(ruta)
    n = 0
    n += _crear(wb, FICHAS_SHEET, FICHAS_HEADERS, _FICHAS_ANCHOS)
    n += _crear(wb, MANT_SHEET, MANT_HEADERS, _MANT_ANCHOS, "7D6608")
    if n:
        _save_wb(wb, ruta)
        logger.info(f"Hojas de maquinaria creadas ({n})")
    wb.close()


# ── Fichas ───────────────────────────────────────────────────────────────

def guardar_ficha(datos: dict, excel_path: str | None = None) -> str:
    """Crea o actualiza la ficha de una máquina. Devuelve el nombre canónico.

    Solo pisa los campos que vengan con valor: así se puede ir completando
    de a poco, a medida que Juan mande los datos.
    """
    nombre = norm_maquina(datos.get("maquina"))
    if not nombre:
        raise ValueError("falta el nombre de la máquina")
    ruta = excel_path or EXCEL_PATH
    crear_hojas(ruta)
    wb = load_workbook(ruta)
    ws = wb[FICHAS_SHEET]

    fila = None
    for r in range(2, ws.max_row + 1):
        if norm_maquina(ws.cell(r, 1).value) == nombre:
            fila = r
            break
    if fila is None:
        fila = ws.max_row + 1
        ws.cell(fila, 1).value = nombre
        ws.cell(fila, 9).value = unidad_de(nombre)

    campos = {2: "tipo", 3: "marca", 4: "modelo", 5: "anio", 6: "patente",
              7: "serie", 8: "propiedad", 10: "estado", 11: "notas"}
    for col, clave in campos.items():
        v = datos.get(clave)
        if v not in (None, "", []):
            ws.cell(fila, col).value = v
    ws.cell(fila, 12).value = date.today()

    _save_wb(wb, ruta)
    wb.close()
    logger.info(f"Ficha de maquinaria guardada: {nombre}")
    return nombre


def listar_fichas(excel_path: str | None = None) -> list:
    wb = load_workbook(excel_path or EXCEL_PATH, read_only=True, data_only=True)
    try:
        if FICHAS_SHEET not in wb.sheetnames:
            return []
        claves = ["maquina", "tipo", "marca", "modelo", "anio", "patente",
                  "serie", "propiedad", "mide", "estado", "notas", "actualizado"]
        out = []
        for row in wb[FICHAS_SHEET].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            fila = list(row) + [None] * (len(claves) - len(row))
            out.append(dict(zip(claves, fila)))
        return out
    finally:
        wb.close()


def campos_faltantes(ficha: dict) -> list:
    """Qué le falta a una ficha, en castellano, para poder pedírselo a Juan."""
    etiquetas = {"marca": "marca", "modelo": "modelo", "anio": "año",
                 "patente": "patente", "serie": "n° de serie",
                 "propiedad": "propia o arrendada"}
    return [t for k, t in etiquetas.items() if not ficha.get(k)]


# ── Mantenciones ─────────────────────────────────────────────────────────

def normalizar_tipo(texto: str) -> str:
    t = _sin_tildes(texto)
    if "ACEITE" in t and "MOTOR" in t: return "ACEITE MOTOR"
    if "ACEITE" in t: return "ACEITE MOTOR"
    if "FILTRO" in t: return "FILTROS"
    if "HIDRAUL" in t: return "HIDRAULICO"
    if "TRANSMI" in t or "CAJA" in t: return "TRANSMISION"
    if "ENGRAS" in t or "GRASA" in t: return "ENGRASE"
    if "NEUMAT" in t or "RUEDA" in t or "LLANTA" in t: return "NEUMATICOS"
    if "FRENO" in t: return "FRENOS"
    if "REPARA" in t or "ARREGL" in t: return "REPARACION"
    if "REVIS" in t or "CHEQUE" in t: return "REVISION"
    return "OTRO"


def registrar_mantencion(m: dict, registrado_por: str = "",
                         excel_path: str | None = None) -> int:
    """Guarda una mantención. Devuelve el ID."""
    nombre = norm_maquina(m.get("maquina"))
    if not nombre:
        raise ValueError("falta la máquina")
    ruta = excel_path or EXCEL_PATH
    crear_hojas(ruta)
    wb = load_workbook(ruta)
    ws = wb[MANT_SHEET]

    nid = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row and isinstance(row[0], (int, float)):
            nid = max(nid, int(row[0]))
    nid += 1

    f = m.get("fecha")
    if isinstance(f, str):
        try:
            f = datetime.strptime(f[:10], "%Y-%m-%d").date()
        except ValueError:
            f = None

    estado = str(m.get("estado") or "HECHA").upper()
    if estado not in ESTADOS_MANT:
        estado = "HECHA"
    # A lo pendiente no se le inventa fecha: todavía no ocurrió.
    fecha = f if f else (None if estado == "PENDIENTE" else date.today())

    ws.append([
        nid, fecha, nombre, estado,
        normalizar_tipo(m.get("tipo") or m.get("descripcion") or ""),
        str(m.get("descripcion") or "")[:200],
        m.get("odometro"), str(m.get("proveedor") or "")[:40],
        m.get("costo"), str(m.get("factura") or ""),
        m.get("proxima_h"), str(m.get("notas") or "")[:120],
        registrado_por,
    ])
    _save_wb(wb, ruta)
    wb.close()
    logger.info(f"Mantención {nid} registrada: {nombre} — {m.get('descripcion')}")
    return nid


def listar_mantenciones(maquina: str | None = None, estado: str | None = None,
                        excel_path: str | None = None) -> list:
    """Mantenciones, de la más reciente a la más antigua.

    `estado` filtra HECHA o PENDIENTE; sin él vienen todas.
    """
    wb = load_workbook(excel_path or EXCEL_PATH, read_only=True, data_only=True)
    try:
        if MANT_SHEET not in wb.sheetnames:
            return []
        objetivo = norm_maquina(maquina) if maquina else None
        filtro = str(estado).upper() if estado else None
        out = []
        for row in wb[MANT_SHEET].iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            if objetivo and norm_maquina(row[2]) != objetivo:
                continue
            if filtro and str(row[3] or "").upper() != filtro:
                continue
            out.append({
                "id": row[0], "fecha": row[1], "maquina": row[2],
                "estado": row[3], "tipo": row[4], "descripcion": row[5],
                "odometro": row[6], "proveedor": row[7], "costo": row[8],
                "factura": row[9], "proxima_h": row[10], "notas": row[11],
            })
        out.sort(key=lambda x: str(x["fecha"] or ""), reverse=True)
        return out
    finally:
        wb.close()


# ── Lecturas de horómetro (van a la Bitácora) ────────────────────────────

def maquinas_conocidas(excel_path: str | None = None) -> list:
    """Máquinas vistas en la bitácora o con ficha, con su última lectura."""
    from collections import defaultdict
    wb = load_workbook(excel_path or EXCEL_PATH, read_only=True, data_only=True)
    try:
        vistas = defaultdict(lambda: {"odometro": None, "fecha": None})
        if "Bitácora" in wb.sheetnames:
            ws = wb["Bitácora"]
            enc = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            try:
                iM, iO = enc.index("Máquina"), enc.index("Odómetro")
            except ValueError:
                iM = iO = None
            if iM is not None:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    m = norm_maquina(row[iM]) if len(row) > iM else ""
                    if not m:
                        continue
                    d = vistas[m]
                    o = row[iO] if len(row) > iO else None
                    f = row[0]
                    if isinstance(f, datetime):
                        f = f.date()
                    if o is not None and (d["fecha"] is None or str(f) >= str(d["fecha"])):
                        d["odometro"], d["fecha"] = o, f
        if FICHAS_SHEET in wb.sheetnames:
            for row in wb[FICHAS_SHEET].iter_rows(min_row=2, max_col=1, values_only=True):
                if row and row[0]:
                    vistas[norm_maquina(row[0])]
    finally:
        wb.close()
    return [{"maquina": m, "ultimo_odometro": d["odometro"],
             "fecha": d["fecha"], "unidad": unidad_de(m)}
            for m, d in sorted(vistas.items())]


def detectar_maquina(texto: str, conocidas: list | None = None) -> str | None:
    """Encuentra a qué máquina se refiere un texto libre."""
    t = norm_maquina(texto)
    if not t:
        return None
    nombres = [c["maquina"] for c in (conocidas or maquinas_conocidas())]
    # 1) nombre completo contenido en el texto
    for n in sorted(nombres, key=len, reverse=True):
        if n and n in t:
            return n
    # 2) por el número de modelo (5085, 6711, 4275…)
    for n in nombres:
        nums = re.findall(r"\d{3,5}", n)
        for num in nums:
            if re.search(rf"\b{num}\b", t):
                return n
    return None


# Un número como los escribe Juan: 3200 · 3.166 (miles) · 7240,7 (decimal)
_NUM = r"\d{1,3}(?:\.\d{3})+(?:,\d+)?|\d+(?:[.,]\d+)?"

_CLAVE = r"hor[oó]metro|od[oó]metro|kilometraje|kms?|hrs?"

# Juan califica la lectura: "Horometro inicio 2039" / "Horometro termino 2041".
# Cuando manda las dos, la que vale es la de término: es donde quedó la máquina.
_CAL_FINAL = r"t[eé]rmino|final|fin|llegada|salida|actual"
_CAL_INICIO = r"inicio|inicial|partida|comienzo|entrada"

# Lectura etiquetada, con el calificador opcional en medio.
_RE_ETIQUETADA = re.compile(
    rf"(?:{_CLAVE})\s*(?:de\s+)?(?P<cal>{_CAL_FINAL}|{_CAL_INICIO})?"
    rf"\s*:?\s*(?:en\s+)?(?P<num>{_NUM})", re.IGNORECASE)
_RE_ES_FINAL = re.compile(rf"(?:{_CAL_FINAL})\Z", re.IGNORECASE)

# La unidad también puede ir después: "1964 horas"
_RE_CON_UNIDAD = re.compile(
    rf"({_NUM})\s*(?:horas?|hrs?|kil[oó]metros?|kms?)\b", re.IGNORECASE)

# Etiquetas cuyo número NO es el horómetro. El parte de Juan trae "Equipo 4" y
# "Sector 1" al final: sin esta guarda, el fallback se llevaba el 1 del sector.
_RE_ETIQUETA_AJENA = re.compile(
    r"(?:sector|equipo|cuartel|hilera|potrero|lote|turno|horas?|jornadas?)"
    r"\s*(?:n[°ºo]?\s*)?:?\s*\Z", re.IGNORECASE)


def _a_float(crudo: str) -> float | None:
    """'3.166' → 3166 · '7240,7' → 7240.7 · '3.166,5' → 3166.5"""
    s = crudo.strip()
    if "," in s and "." in s:
        # el separador decimal es el que está más a la derecha
        s = (s.replace(".", "").replace(",", ".")
             if s.rfind(",") > s.rfind(".") else s.replace(",", ""))
    elif "," in s:
        s = s.replace(",", ".")
    elif s.count(".") >= 1 and all(len(p) == 3 for p in s.split(".")[1:]):
        s = s.replace(".", "")          # 3.166 son tres mil, no 3,166
    try:
        return float(s)
    except ValueError:
        return None


def extraer_odometro(texto: str) -> float | None:
    """Saca la lectura de horómetro/kilometraje de un texto libre.

    Orden de preferencia:
      1. lectura etiquetada de término/final ("Horometro termino 2041");
      2. cualquier otra lectura etiquetada ("Horómetro 3200", "…inicio 2039");
      3. número con la unidad detrás ("1964 horas");
      4. el último número suelto ("MF 6711 1980"), saltando los que pertenecen
         a otra etiqueta (Sector, Equipo, Total horas…).
    """
    t = (texto or "").strip()
    if not t:
        return None

    finales, etiquetadas = [], []
    for m in _RE_ETIQUETADA.finditer(t):
        v = _a_float(m.group("num"))
        if v is None:
            continue
        etiquetadas.append(v)
        if m.group("cal") and _RE_ES_FINAL.search(m.group("cal")):
            finales.append(v)
    if finales:
        return finales[-1]
    if etiquetadas:
        return etiquetadas[-1]

    m = _RE_CON_UNIDAD.search(t)
    if m:
        v = _a_float(m.group(1))
        if v is not None:
            return v

    # Sin unidad explícita, el último número del mensaje: "MF 6711 1980",
    # "JD 5085: 3200". El llamador ya verificó que se habla de una máquina.
    sueltos = [m for m in re.finditer(_NUM, t)
               if not _RE_ETIQUETA_AJENA.search(t[:m.start()])]
    return _a_float(sueltos[-1].group(0)) if sueltos else None
