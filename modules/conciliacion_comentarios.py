"""modules/conciliacion_comentarios.py — Notas sobre movimientos del banco.

Hay cargos que no se explican solos ("¿por qué este pago a Don Antonio?") y la
respuesta se pierde entre sesiones. Acá queda escrita, con fecha y autor.

Hoja `Comentarios Banco`: varias notas por movimiento, en orden cronológico.
"""
import logging
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from config import EXCEL_PATH
from excel_manager import _save_wb

logger = logging.getLogger(__name__)

SHEET = "Comentarios Banco"
HEADERS = ["ID", "Fecha", "Fila Banco", "Comentario", "Usuario"]
_WIDTHS = [6, 12, 10, 70, 16]
MAX_LARGO = 500


def crear_hoja(wb=None, excel_path: str | None = None) -> None:
    """Crea la hoja si no existe. Idempotente."""
    ruta = excel_path or EXCEL_PATH
    propio = wb is None
    if propio:
        wb = load_workbook(ruta)
    if SHEET not in wb.sheetnames:
        ws = wb.create_sheet(SHEET)
        fill = PatternFill("solid", fgColor="1F4E78")
        for i, h in enumerate(HEADERS, 1):
            c = ws.cell(1, i, h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = fill
            c.alignment = Alignment(horizontal="center")
        for i, w in enumerate(_WIDTHS, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = "A2"
        if propio:
            _save_wb(wb, ruta)
            logger.info(f"Hoja '{SHEET}' creada")
    if propio:
        wb.close()


def _next_id(ws) -> int:
    m = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row and isinstance(row[0], (int, float)):
            m = max(m, int(row[0]))
    return m + 1


def agregar(fila_banco: int, texto: str, usuario: str = "",
            excel_path: str | None = None) -> dict:
    """Guarda una nota. Devuelve {id, fecha}."""
    texto = " ".join(str(texto or "").split())[:MAX_LARGO]
    if not texto:
        raise ValueError("el comentario está vacío")
    ruta = excel_path or EXCEL_PATH
    wb = load_workbook(ruta)
    crear_hoja(wb, ruta)
    ws = wb[SHEET]
    nid = _next_id(ws)
    hoy = date.today()
    ws.append([nid, hoy, int(fila_banco), texto, usuario])
    _save_wb(wb, ruta)
    wb.close()
    logger.info(f"Comentario {nid} agregado al movimiento {fila_banco}")
    return {"id": nid, "fecha": hoy.isoformat()}


def de_movimiento(fila_banco: int, excel_path: str | None = None) -> list:
    """Notas de un movimiento, de la más vieja a la más nueva."""
    wb = load_workbook(excel_path or EXCEL_PATH, read_only=True, data_only=True)
    try:
        if SHEET not in wb.sheetnames:
            return []
        out = []
        for row in wb[SHEET].iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            try:
                if int(row[2]) != int(fila_banco):
                    continue
            except (TypeError, ValueError):
                continue
            out.append({"id": int(row[0]), "fecha": _iso(row[1]),
                        "texto": str(row[3] or ""), "usuario": str(row[4] or "")})
        out.sort(key=lambda x: x["id"])
        return out
    finally:
        wb.close()


def conteo_por_fila(excel_path: str | None = None) -> dict:
    """{fila_banco: nº de notas} — para pintar el 💬 en la tabla."""
    wb = load_workbook(excel_path or EXCEL_PATH, read_only=True, data_only=True)
    try:
        if SHEET not in wb.sheetnames:
            return {}
        out = {}
        for row in wb[SHEET].iter_rows(min_row=2, max_col=3, values_only=True):
            if not row or row[0] is None:
                continue
            try:
                fb = int(row[2])
            except (TypeError, ValueError):
                continue
            out[fb] = out.get(fb, 0) + 1
        return out
    finally:
        wb.close()


def eliminar(id_comentario: int, excel_path: str | None = None) -> bool:
    ruta = excel_path or EXCEL_PATH
    wb = load_workbook(ruta)
    if SHEET not in wb.sheetnames:
        wb.close()
        return False
    ws = wb[SHEET]
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, 1).value == id_comentario:
            ws.delete_rows(r)
            _save_wb(wb, ruta)
            wb.close()
            return True
    wb.close()
    return False


def _iso(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v or "")[:10]
