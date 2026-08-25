"""modules/conciliacion_export.py — Exportar el estado de conciliación a Excel.

Dos usos: mandarle al contador lo que está cuadrado, y tener a la vista lo que
falta. Se exporta lo que se ve, no un volcado crudo: cada movimiento con su
estado, lo asignado, el saldo y contra qué documentos quedó.
"""
import logging
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import EXCEL_PATH
from modules.conciliacion_store import SHEET as SHEET_CONC

logger = logging.getLogger(__name__)

BANCO_SHEET = "Cuenta Banco"
CABECERAS = ["Fecha", "Descripción", "Categoría", "Cargo", "Abono",
             "Monto", "Asignado", "Saldo", "Estado", "Documentos", "Fila"]
_ANCHOS = [12, 46, 24, 14, 14, 14, 14, 13, 14, 44, 7]


def _fecha(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def recolectar(estado: str = "todos", desde: date | None = None,
               hasta: date | None = None, excel_path: str | None = None) -> list:
    """Movimientos con su estado de conciliación.

    `estado`: 'conciliado' · 'parcial' · 'pendiente' · 'todos'.
    """
    wb = load_workbook(excel_path or EXCEL_PATH, read_only=True, data_only=True)
    try:
        # Vínculos por movimiento
        asignado, docs = {}, {}
        if SHEET_CONC in wb.sheetnames:
            for row in wb[SHEET_CONC].iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                try:
                    fb = int(row[2])
                    asignado[fb] = asignado.get(fb, 0.0) + float(row[10] or 0)
                except (TypeError, ValueError):
                    continue
                nro = str(row[8] or "").strip()
                etiqueta = f"{nro} {str(row[9] or '')[:22]}".strip() or str(row[6] or "")
                docs.setdefault(fb, []).append(etiqueta)

        filas = []
        for i, row in enumerate(wb[BANCO_SHEET].iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:
                continue
            f = _fecha(row[0])
            if not f or (desde and f < desde) or (hasta and f > hasta):
                continue
            try:
                cargo, abono = float(row[3] or 0), float(row[4] or 0)
            except (TypeError, ValueError):
                cargo = abono = 0.0
            monto = cargo if cargo > 0 else abono
            if monto <= 0:
                continue
            asig = asignado.get(i, 0.0)
            saldo = round(monto - asig)
            if asig <= 0:
                est = "pendiente"
            elif abs(saldo) <= 1:
                est = "conciliado"
            else:
                est = "parcial"
            if estado != "todos" and est != estado:
                continue
            filas.append({
                "fecha": f, "desc": str(row[1] or ""), "categoria": str(row[7] or ""),
                "cargo": cargo, "abono": abono, "monto": monto,
                "asignado": round(asig), "saldo": saldo, "estado": est,
                "docs": " + ".join(docs.get(i, [])), "fila": i,
            })
        filas.sort(key=lambda x: x["fecha"])
        return filas
    finally:
        wb.close()


def resumen(filas: list) -> dict:
    """Totales y % conciliado del conjunto."""
    total = sum(f["monto"] for f in filas)
    conc = sum(f["monto"] for f in filas if f["estado"] == "conciliado")
    parc = sum(f["asignado"] for f in filas if f["estado"] == "parcial")
    cubierto = conc + parc
    return {
        "movimientos": len(filas),
        "monto_total": round(total),
        "monto_conciliado": round(cubierto),
        "monto_pendiente": round(total - cubierto),
        "pct": round(cubierto / total * 100, 1) if total else 0.0,
        "n_conciliados": sum(1 for f in filas if f["estado"] == "conciliado"),
        "n_parciales": sum(1 for f in filas if f["estado"] == "parcial"),
        "n_pendientes": sum(1 for f in filas if f["estado"] == "pendiente"),
    }


def a_excel(filas: list, titulo: str = "Conciliación") -> BytesIO:
    """Genera el .xlsx en memoria y devuelve el buffer listo para descargar."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliación"

    r = resumen(filas)
    ws.append([titulo])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"{r['movimientos']} movimientos · {r['pct']}% conciliado · "
               f"pendiente ${r['monto_pendiente']:,.0f}"])
    ws["A2"].font = Font(color="7A8894")
    ws.append([])

    fila_hdr = 4
    fill = PatternFill("solid", fgColor="1F4E78")
    for i, h in enumerate(CABECERAS, 1):
        c = ws.cell(fila_hdr, i, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
    for i, w in enumerate(_ANCHOS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    colores = {"conciliado": "1F9D55", "parcial": "D98F00", "pendiente": "C0392B"}
    for n, f in enumerate(filas, fila_hdr + 1):
        ws.append([f["fecha"], f["desc"], f["categoria"],
                   f["cargo"] or None, f["abono"] or None, f["monto"],
                   f["asignado"], f["saldo"], f["estado"], f["docs"], f["fila"]])
        ws.cell(n, 9).font = Font(color=colores.get(f["estado"], "000000"), bold=True)
        for col in (4, 5, 6, 7, 8):
            ws.cell(n, col).number_format = '#,##0;[Red]-#,##0;-'
        ws.cell(n, 1).number_format = "dd-mm-yyyy"

    ws.freeze_panes = f"A{fila_hdr + 1}"
    ws.auto_filter.ref = f"A{fila_hdr}:{get_column_letter(len(CABECERAS))}{ws.max_row}"

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf


def pct_conciliado_mes(year: int, month: int,
                       excel_path: str | None = None) -> dict:
    """% conciliado de un mes, para el reporte mensual."""
    from calendar import monthrange
    ini = date(year, month, 1)
    fin = date(year, month, monthrange(year, month)[1])
    return resumen(recolectar("todos", ini, fin, excel_path=excel_path))
