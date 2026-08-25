"""modules/conciliacion_store.py — Registro de conciliaciones banco ↔ documentos.

Hoja `Conciliaciones`: una fila por VÍNCULO (movimiento ↔ documento, con monto
asignado). Esto permite lo que la col J (texto) no podía:
  - conciliación PARCIAL (un pago cubre parte de una factura),
  - N:M (un cargo paga varias facturas; una factura se paga en cuotas).

La col J de Cuenta Banco pasa a ser un resumen legible que este módulo mantiene.
"""
import logging
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import EXCEL_PATH
from excel_manager import _save_wb

logger = logging.getLogger(__name__)

SHEET = "Conciliaciones"
HEADERS = ["ID", "Fecha Conciliación", "Fila Banco", "Fecha Mov",
           "Descripción Mov", "Monto Mov", "Tipo Doc", "Fila Doc",
           "N° Doc", "Proveedor", "Monto Asignado", "Criterio",
           "Usuario", "Nota"]
_WIDTHS = [6, 15, 10, 12, 40, 13, 14, 9, 13, 30, 14, 16, 14, 30]

# Tipos de documento/destino válidos
TIPOS_DOC = ["FACTURA", "BOLETA", "BALANCE", "TERCEROS", "NO CONCILIABLE"]

COL_J_BANCO = 10   # Factura_linkeada (resumen legible)


def _open():
    return load_workbook(EXCEL_PATH)


def crear_hoja(wb=None) -> None:
    """Crea la hoja Conciliaciones si no existe. Idempotente."""
    propio = wb is None
    if propio:
        wb = _open()
    if SHEET not in wb.sheetnames:
        ws = wb.create_sheet(SHEET)
        fill = PatternFill("solid", fgColor="1F4E78")
        for i, h in enumerate(HEADERS, 1):
            c = ws.cell(1, i, h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = fill
            c.alignment = Alignment(horizontal="center")
        for i, w in enumerate(_WIDTHS, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = "A2"
        if propio:
            _save_wb(wb)
            logger.info("Hoja Conciliaciones creada")
    if propio:
        wb.close()


def _next_id(ws) -> int:
    m = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row and isinstance(row[0], (int, float)):
            m = max(m, int(row[0]))
    return m + 1


def _monto_mov(ws_banco, fila_banco: int) -> float:
    try:
        cargo = float(ws_banco.cell(fila_banco, 4).value or 0)
        abono = float(ws_banco.cell(fila_banco, 5).value or 0)
    except (TypeError, ValueError):
        return 0.0
    return cargo if cargo > 0 else abono


def _asignado_por_fila(ws_conc) -> dict:
    """{fila_banco: monto asignado total} desde la hoja Conciliaciones."""
    out = {}
    for row in ws_conc.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            fb = int(row[2])
            m = float(row[10] or 0)
        except (TypeError, ValueError):
            continue
        out[fb] = out.get(fb, 0.0) + m
    return out


def _actualizar_col_j(ws_banco, ws_conc, fila_banco: int) -> None:
    """Reescribe el resumen legible de la col J para un movimiento."""
    vinculos = []
    for row in ws_conc.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            if int(row[2]) != fila_banco:
                continue
        except (TypeError, ValueError):
            continue
        nro = str(row[8] or "").strip()
        prov = str(row[9] or "").strip()
        tipo = str(row[6] or "").strip()
        if tipo == "FACTURA" and nro:
            # Las notas de débito/crédito ya vienen con su propia sigla (ND, NC):
            # anteponerles "F" daba "FND506808".
            pre = "F" if nro[:1].isdigit() else ""
            vinculos.append(f"{pre}{nro} {prov[:24]}")
        elif tipo == "BOLETA" and nro:
            pre = "BH" if nro[:1].isdigit() else ""
            vinculos.append(f"{pre}{nro} {prov[:24]}")
        else:
            vinculos.append(tipo.capitalize() or "Conciliado")
    ws_banco.cell(fila_banco, COL_J_BANCO).value = (
        " + ".join(vinculos) if vinculos else None)


def registrar_vinculos(vinculos: list[dict], usuario: str = "") -> dict:
    """Registra una lista de vínculos en UN solo guardado.

    Cada vínculo: {fila_banco, tipo_doc, fila_doc, nro_doc, proveedor,
                   monto_asignado, criterio, nota, fecha_pago(optional),
                   filas_doc(list, para completar Fecha Pago en Facturas)}
    Devuelve {registrados, ids}.
    """
    if not vinculos:
        return {"registrados": 0, "ids": []}
    wb = _open()
    crear_hoja(wb)
    ws_conc = wb[SHEET]
    ws_banco = wb["Cuenta Banco"]
    ws_fact = wb["Facturas"]

    nid = _next_id(ws_conc)
    ids = []
    hoy = date.today().isoformat()
    filas_tocadas = set()

    for v in vinculos:
        fb = int(v["fila_banco"])
        fecha_mov = ws_banco.cell(fb, 1).value
        if isinstance(fecha_mov, datetime):
            fecha_mov = fecha_mov.date()
        desc_mov = str(ws_banco.cell(fb, 2).value or "")[:60]
        monto_mov = _monto_mov(ws_banco, fb)

        ws_conc.append([
            nid, hoy, fb,
            fecha_mov.isoformat() if isinstance(fecha_mov, date) else str(fecha_mov or ""),
            desc_mov, monto_mov,
            str(v.get("tipo_doc") or "FACTURA").upper(),
            v.get("fila_doc"),
            str(v.get("nro_doc") or ""),
            str(v.get("proveedor") or "")[:40],
            round(float(v.get("monto_asignado") or monto_mov)),
            str(v.get("criterio") or "manual"),
            usuario or str(v.get("usuario") or ""),
            str(v.get("nota") or ""),
        ])
        ids.append(nid)
        nid += 1
        filas_tocadas.add(fb)

        # Completar Fecha Pago en las líneas de la factura (si vacía)
        fecha_pago = v.get("fecha_pago") or fecha_mov
        for r in (v.get("filas_doc") or []):
            cell = ws_fact.cell(int(r), 3)
            if not (cell.value and str(cell.value).strip()):
                cell.value = fecha_pago

    for fb in filas_tocadas:
        _actualizar_col_j(ws_banco, ws_conc, fb)

    _save_wb(wb)
    wb.close()
    logger.info(f"Conciliaciones registradas: {len(ids)}")
    return {"registrados": len(ids), "ids": ids}


def desconciliar(id_vinculo: int) -> bool:
    """Elimina un vínculo por ID y actualiza el resumen del movimiento."""
    wb = _open()
    if SHEET not in wb.sheetnames:
        wb.close()
        return False
    ws_conc = wb[SHEET]
    ws_banco = wb["Cuenta Banco"]
    fila_borrar, fb = None, None
    for r in range(2, ws_conc.max_row + 1):
        if ws_conc.cell(r, 1).value == id_vinculo:
            fila_borrar = r
            try:
                fb = int(ws_conc.cell(r, 3).value)
            except (TypeError, ValueError):
                fb = None
            break
    if fila_borrar is None:
        wb.close()
        return False
    ws_conc.delete_rows(fila_borrar)
    if fb:
        _actualizar_col_j(ws_banco, ws_conc, fb)
    _save_wb(wb)
    wb.close()
    logger.info(f"Conciliación {id_vinculo} eliminada")
    return True


def _clave_doc(nro, proveedor) -> tuple:
    """Identifica un documento. Incluye el proveedor porque distintos
    proveedores repiten numeración de facturas."""
    n = str(nro or "").strip().upper()
    if n.endswith(".0"):
        n = n[:-2]
    return (" ".join(str(proveedor or "").upper().split()), n)


def asignado_por_documento(excel_path: str | None = None) -> dict:
    """{(proveedor, n° doc): monto ya asignado} sumando todos los movimientos.

    Es lo que permite pagar una factura en cuotas: cada abono suma contra el
    mismo documento y se sabe cuánto falta.
    """
    wb = load_workbook(excel_path or EXCEL_PATH, read_only=True, data_only=True)
    try:
        if SHEET not in wb.sheetnames:
            return {}
        out = {}
        for row in wb[SHEET].iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            nro = str(row[8] or "").strip()
            if not nro:
                continue
            try:
                m = float(row[10] or 0)
            except (TypeError, ValueError):
                continue
            k = _clave_doc(nro, row[9])
            out[k] = out.get(k, 0.0) + m
        return out
    finally:
        wb.close()


def estado_documento(nro, proveedor, total: float = 0.0,
                     excel_path: str | None = None) -> dict:
    """Cobertura de un documento: cuánto se pagó y cuánto falta.

    `estado`: 'pagado' (saldo ≤ $1) · 'parcial' · 'pendiente'.
    """
    asignado = asignado_por_documento(excel_path).get(
        _clave_doc(nro, proveedor), 0.0)
    total = float(total or 0)
    saldo = round(total - asignado)
    if asignado <= 0:
        estado = "pendiente"
    elif abs(saldo) <= 1:
        estado = "pagado"
    else:
        estado = "parcial"
    return {"total": total, "asignado": round(asignado),
            "saldo": saldo, "estado": estado}


def vinculos_de_documento(nro, proveedor,
                          excel_path: str | None = None) -> list:
    """Movimientos que pagaron ese documento (las cuotas), del más viejo al más nuevo."""
    wb = load_workbook(excel_path or EXCEL_PATH, read_only=True, data_only=True)
    try:
        if SHEET not in wb.sheetnames:
            return []
        objetivo = _clave_doc(nro, proveedor)
        out = []
        for row in wb[SHEET].iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            if _clave_doc(row[8], row[9]) != objetivo:
                continue
            out.append({
                "id": int(row[0]), "fila_banco": row[2], "fecha_mov": row[3],
                "desc_mov": row[4], "monto_mov": row[5],
                "monto_asignado": row[10], "criterio": row[11],
            })
        out.sort(key=lambda x: str(x["fecha_mov"] or ""))
        return out
    finally:
        wb.close()


def saldo_por_asignar(fila_banco: int, excel_path: str | None = None) -> dict:
    """Cuánto queda por asignar de un movimiento del banco."""
    wb = load_workbook(excel_path or EXCEL_PATH, read_only=True, data_only=True)
    try:
        ws_banco = wb["Cuenta Banco"]
        try:
            row = next(ws_banco.iter_rows(min_row=fila_banco, max_row=fila_banco,
                                           values_only=True))
        except StopIteration:
            return {"monto": 0, "asignado": 0, "saldo": 0, "estado": "por conciliar"}
        try:
            cargo, abono = float(row[3] or 0), float(row[4] or 0)
        except (TypeError, ValueError):
            cargo = abono = 0.0
        monto = cargo if cargo > 0 else abono
        asig = 0.0
        if SHEET in wb.sheetnames:
            for r in wb[SHEET].iter_rows(min_row=2, values_only=True):
                if not r or r[0] is None:
                    continue
                try:
                    if int(r[2]) == fila_banco:
                        asig += float(r[10] or 0)
                except (TypeError, ValueError):
                    continue
    finally:
        wb.close()
    saldo = round(monto - asig)
    if asig <= 0:
        estado = "por conciliar"
    elif abs(saldo) <= 1:
        estado = "conciliado"
    else:
        estado = "parcial"
    return {"monto": monto, "asignado": round(asig), "saldo": saldo,
            "estado": estado, "fecha": row[0], "desc": str(row[1] or "")}


def resumen_estados() -> dict:
    """{fila_banco: {monto, asignado, saldo, estado}} para pintar la UI.

    estado: 'conciliado' (saldo ≤ $1), 'parcial' (algo asignado),
    'por conciliar' (nada asignado — las filas sin entrada no aparecen).
    """
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        wb.close()
        return {}
    ws_conc = wb[SHEET]
    asignado = _asignado_por_fila(ws_conc)
    ws_banco = wb["Cuenta Banco"]
    out = {}
    for fb, asig in asignado.items():
        try:
            row = next(ws_banco.iter_rows(min_row=fb, max_row=fb, values_only=True))
        except StopIteration:
            continue
        try:
            cargo = float(row[3] or 0)
            abono = float(row[4] or 0)
        except (TypeError, ValueError):
            cargo = abono = 0
        monto = cargo if cargo > 0 else abono
        saldo = round(monto - asig)
        out[fb] = {"monto": monto, "asignado": round(asig), "saldo": saldo,
                   "estado": "conciliado" if abs(saldo) <= 1 else "parcial"}
    wb.close()
    return out
