"""modules/conciliador.py — Conciliación bancaria: cruza Cuenta Banco ↔ Facturas.

Formaliza el cruce manual que se hacía por scripts:
1. Match determinista: nº factura en la glosa del banco, o monto+proveedor.
2. Match con IA (Claude): casos dudosos (mismo monto, nombres distintos).
3. Aplicar: escribe el enlace en Cuenta Banco (col J "Factura Link") y
   completa Fecha Pago en Facturas si está vacía.

Todo el análisis es de solo-lectura; `aplicar_conciliacion` es la única
función que escribe, y se llama tras confirmación del usuario en Telegram.
"""
import json
import logging
import re
from datetime import date, datetime, timedelta
from collections import defaultdict

import requests
from openpyxl import load_workbook

from config import EXCEL_PATH, ANTHROPIC_API_KEY
from excel_manager import _save_wb, COL_BANCO_FACTURA_LINK, COL_BANCO_CATEGORIA

logger = logging.getLogger(__name__)

RE_FACT = re.compile(r"\bF?\s*N?[DC]?\s*(\d{2,})\b", re.IGNORECASE)

# Categorías del banco que normalmente NO tienen factura asociada
CATS_SIN_FACTURA = {
    "MANO DE OBRA PLANTA", "MANO DE OBRA TEMPORAL", "IMPUESTOS",
    "PRESTAMOS A OTRAS SOCIEDADES", "TRANSFERENCIA INTERNA", "CAMBIO DIVISA",
    "REINTEGROS Y DEVOLUCIONES", "GASTOS BANCARIOS", "PRE-2021 HISTORICO",
    "INGRESO OPERACIONAL", "INGRESO VENTAS", "INGRESO FINANCIERO",
    "CAJA CHICA / IMPREVISTOS", "BONO VENTA NUECES", "LEASING", "REVISAR",
}

TOLERANCIA_PCT = 0.01   # 1% de diferencia de monto aceptada
VENTANA_DIAS = 120      # banco vs emisión de factura


def _pd(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for f in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(v[:10], f).date()
            except Exception:
                pass
    return None


def _norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip().upper().replace(".", "")


def _nrokey(v):
    s = str(v or "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.upper().replace(" ", "")


def _tokens(nombre):
    stop = {"SPA", "LTDA", "SA", "S", "A", "Y", "DE", "DEL", "LA", "EL", "LOS",
            "CIA", "COMERCIAL", "SERVICIOS", "SERV", "AGRICOLA", "F", "EIRL"}
    return {t for t in re.split(r"[^A-ZÑ0-9]+", _norm(nombre)) if len(t) >= 3 and t not in stop}


def _monto_similar(a, b):
    if a <= 0 or b <= 0:
        return False
    return abs(a - b) <= max(1000.0, a * TOLERANCIA_PCT)


def cargar_datos(dias: int = 90):
    """Lee banco (cargos sin link) y facturas (agrupadas por prov+nro)."""
    desde = date.today() - timedelta(days=dias)
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    ws_b = wb["Cuenta Banco"]
    cargos = []
    for r_idx, row in enumerate(ws_b.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        f = _pd(row[0])
        if not f or f < desde:
            continue
        try:
            cargo = float(row[3] or 0)
        except Exception:
            continue
        if cargo <= 0:
            continue
        link = row[COL_BANCO_FACTURA_LINK - 1] if len(row) >= COL_BANCO_FACTURA_LINK else None
        if link and str(link).strip():
            continue  # ya conciliado
        cargos.append({
            "fila": r_idx, "fecha": f, "desc": str(row[1] or ""),
            "monto": cargo,
            "categoria": _norm(row[COL_BANCO_CATEGORIA - 1] if len(row) >= COL_BANCO_CATEGORIA else ""),
        })

    ws_f = wb["Facturas"]
    fact = {}   # (prov_norm, nro) -> info
    for r_idx, row in enumerate(ws_f.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        nro = _nrokey(row[6])
        prov = str(row[3] or "")
        if not nro:
            continue
        key = (_norm(prov), nro)
        try:
            total = float(row[15] or 0)
        except Exception:
            total = 0
        d = fact.setdefault(key, {
            "prov": prov, "nro": nro, "emision": _pd(row[0]),
            "fpago": None, "total": 0.0, "filas": [], "pagada": False,
        })
        d["filas"].append(r_idx)
        if total > d["total"]:
            d["total"] = total  # col P se repite por fila del grupo
        if row[2] and str(row[2]).strip():
            d["pagada"] = True
    wb.close()
    return cargos, fact


# Cómo se explica cada criterio en la vista de sugerencias
MOTIVOS = {
    "nro+monto": "N° de documento en la glosa y el monto calza",
    "monto+proveedor": "El monto calza y el nombre del proveedor aparece en la glosa",
}


def explicar(criterio: str) -> str:
    """Texto legible del porqué de una sugerencia."""
    c = str(criterio or "")
    if c.startswith("IA"):
        return f"Lo evaluó la IA — {c[3:].strip('() ')} de confianza"
    return MOTIVOS.get(c, c or "Sugerencia automática")


def analizar(dias: int = 90, excluir_rechazados: bool = True) -> dict:
    """Corre el matching. Devuelve dict con confirmados/dudosos/sin match.

    `excluir_rechazados` saca los pares que el usuario ya descartó, para que la
    vista de sugerencias no repita lo mismo en cada corrida.
    """
    cargos, fact = cargar_datos(dias)

    rechazados = set()
    if excluir_rechazados:
        try:
            from modules.conciliacion_rechazos import clave as _k, rechazados as _r
            rechazados = _r()
        except Exception as e:                       # nunca romper por esto
            logger.warning(f"No pude leer los rechazos: {e}")

    def _descartado(cargo, doc) -> bool:
        if not rechazados:
            return False
        from modules.conciliacion_rechazos import clave as _k
        return _k(cargo["fila"], doc["nro"], doc["prov"]) in rechazados

    # Índices auxiliares
    por_nro = defaultdict(list)
    for key, d in fact.items():
        por_nro[key[1]].append(d)

    confirmados = []   # [{cargo, factura, criterio}]
    dudosos = []
    usadas = set()     # (prov_norm, nro) ya asignadas

    for c in cargos:
        desc_tokens = _tokens(c["desc"])
        nros_en_desc = [m for m in RE_FACT.findall(c["desc"])]
        match = None
        criterio = ""

        # Paso 1: nº de factura en la glosa + monto similar
        for nro in nros_en_desc:
            for d in por_nro.get(_nrokey(nro), []):
                k = (_norm(d["prov"]), d["nro"])
                if k in usadas or _descartado(c, d):
                    continue
                if _monto_similar(c["monto"], d["total"]):
                    match, criterio = d, "nro+monto"
                    break
            if match:
                break

        # Paso 2: monto similar + proveedor por tokens + ventana de fechas
        if not match:
            candidatos = []
            for k, d in fact.items():
                if k in usadas or d["pagada"] or _descartado(c, d):
                    continue
                if not _monto_similar(c["monto"], d["total"]):
                    continue
                if d["emision"] and abs((c["fecha"] - d["emision"]).days) > VENTANA_DIAS:
                    continue
                overlap = desc_tokens & _tokens(d["prov"])
                candidatos.append((len(overlap), d))
            candidatos.sort(key=lambda x: -x[0])
            if candidatos and candidatos[0][0] >= 1:
                match, criterio = candidatos[0][1], "monto+proveedor"
            elif len(candidatos) == 1:
                # mismo monto, nombre no calza → dudoso (lo decide la IA)
                dudosos.append({"cargo": c, "factura": candidatos[0][1]})
                continue

        if match:
            usadas.add((_norm(match["prov"]), match["nro"]))
            confirmados.append({"cargo": c, "factura": match, "criterio": criterio})

    # Lo que queda sin match
    cargos_matcheados = {id(m["cargo"]) for m in confirmados} | {id(d["cargo"]) for d in dudosos}
    sin_factura = [c for c in cargos if id(c) not in cargos_matcheados
                   and c["categoria"] not in CATS_SIN_FACTURA]
    fact_sin_pago = [d for k, d in fact.items()
                     if k not in usadas and not d["pagada"] and d["total"] > 0
                     and d["emision"] and d["emision"] >= date.today() - timedelta(days=dias)]

    return {"confirmados": confirmados, "dudosos": dudosos,
            "sin_factura": sin_factura, "fact_sin_pago": fact_sin_pago,
            "total_cargos": len(cargos)}


def resolver_dudosos_ia(dudosos: list) -> list:
    """Pregunta a Claude por los matches dudosos. Devuelve los aceptados."""
    if not dudosos or not ANTHROPIC_API_KEY:
        return []
    casos = []
    for i, d in enumerate(dudosos[:40]):
        casos.append({
            "id": i,
            "banco": {"glosa": d["cargo"]["desc"][:80],
                      "monto": d["cargo"]["monto"],
                      "fecha": str(d["cargo"]["fecha"])},
            "factura": {"proveedor": d["factura"]["prov"][:60],
                        "nro": d["factura"]["nro"],
                        "total": d["factura"]["total"],
                        "emision": str(d["factura"]["emision"])},
        })
    prompt = (
        "Eres un conciliador bancario de una agrícola chilena. Para cada caso, "
        "decide si el cargo del banco corresponde al pago de esa factura "
        "(los montos ya calzan; evalúa si el nombre de la glosa y el proveedor "
        "pueden ser la misma entidad — abreviaciones, siglas, nombre de fantasía "
        "vs razón social, persona vs empresa). Devuelve SOLO JSON:\n"
        '{"matches": [{"id": <id>, "match": true|false, "confianza": 0.0-1.0}]}\n\n'
        f"Casos:\n{json.dumps(casos, ensure_ascii=False)}"
    )
    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": ANTHROPIC_API_KEY,
                     "anthropic-version": "2023-06-01",
                     "content-type": "application/json"},
            json={"model": "claude-haiku-4-5-20251001", "max_tokens": 1500,
                  "messages": [{"role": "user", "content": prompt}]},
            timeout=60)
        raw = resp.json()["content"][0]["text"].strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        data = json.loads(raw)
        aceptados = []
        for m in data.get("matches", []):
            if m.get("match") and float(m.get("confianza", 0)) >= 0.7:
                d = dudosos[int(m["id"])]
                d["criterio"] = f"IA ({float(m['confianza']):.0%})"
                aceptados.append(d)
        return aceptados
    except Exception as e:
        logger.warning(f"Conciliador IA: {e}")
        return []


def aplicar_conciliacion(matches: list, usuario: str = "") -> dict:
    """Registra los vínculos banco↔factura en la hoja Conciliaciones.

    El registro (una fila por vínculo con monto asignado) lo lleva
    `conciliacion_store`, que además mantiene el resumen de la col J y
    completa las fechas de pago vacías en Facturas.
    """
    if not matches:
        return {"links": 0, "fechas": 0}
    from modules.conciliacion_store import registrar_vinculos
    vinculos = []
    for m in matches:
        c, f = m["cargo"], m["factura"]
        vinculos.append({
            "fila_banco": c["fila"],
            "tipo_doc": "FACTURA",
            "fila_doc": (f.get("filas") or [None])[0],
            "filas_doc": f.get("filas") or [],
            "nro_doc": f["nro"],
            "proveedor": f["prov"],
            "monto_asignado": None,          # el store usa el monto del movimiento
            "criterio": m.get("criterio", "auto"),
            "fecha_pago": c.get("fecha"),
        })
    r = registrar_vinculos(vinculos, usuario=usuario)
    logger.info(f"Conciliación aplicada: {r['registrados']} vínculos")
    return {"links": r["registrados"], "fechas": r["registrados"]}


def formato_resumen(res: dict, ia_aceptados: list) -> str:
    """Texto del reporte para Telegram (sin Markdown frágil)."""
    conf = res["confirmados"]
    lines = [f"🏦 CONCILIACIÓN BANCARIA",
             f"Cargos analizados (sin conciliar): {res['total_cargos']}", ""]
    lines.append(f"✅ Matches seguros: {len(conf)}")
    for m in conf[:12]:
        c, f = m["cargo"], m["factura"]
        lines.append(f"  {c['fecha']} ${c['monto']:,.0f} → F{f['nro']} {f['prov'][:24]} [{m['criterio']}]")
    if len(conf) > 12:
        lines.append(f"  … y {len(conf)-12} más")
    lines.append("")
    lines.append(f"🤖 Aceptados por IA: {len(ia_aceptados)}")
    for m in ia_aceptados[:8]:
        c, f = m["cargo"], m["factura"]
        lines.append(f"  {c['fecha']} ${c['monto']:,.0f} → F{f['nro']} {f['prov'][:24]} [{m['criterio']}]")
    rechazados = len(res["dudosos"]) - len(ia_aceptados)
    if rechazados > 0:
        lines.append(f"  ({rechazados} dudosos descartados por la IA)")
    lines.append("")
    sf = res["sin_factura"]
    lines.append(f"❓ Cargos sin factura encontrada: {len(sf)}")
    for c in sorted(sf, key=lambda x: -x["monto"])[:8]:
        lines.append(f"  {c['fecha']} ${c['monto']:,.0f} {c['desc'][:34]}")
    if len(sf) > 8:
        lines.append(f"  … y {len(sf)-8} más")
    lines.append("")
    fsp = res["fact_sin_pago"]
    lines.append(f"📄 Facturas sin pago detectado: {len(fsp)}")
    for d in sorted(fsp, key=lambda x: -x["total"])[:8]:
        lines.append(f"  F{d['nro']} {d['prov'][:26]} ${d['total']:,.0f}")
    return "\n".join(lines)
