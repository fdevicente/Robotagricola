"""modules/bodega_check.py — Contrasta el Excel de bodega contra el Master.

El equipo lleva la bodega en su propio Excel (Dropbox). Este módulo calcula el
stock desde los movimientos (ingresos − salidas) y lo compara con la hoja
Inventario del Master, para avisar cuando la data deja de calzar.
"""
import glob
import logging
import os
import re
import shutil
import tempfile
import unicodedata
from collections import defaultdict

from openpyxl import load_workbook

from config import EXCEL_PATH

logger = logging.getLogger(__name__)

# Carpeta y patrón del Excel de bodega. Juan genera "copias en conflicto" de
# Dropbox, así que SIEMPRE se usa el archivo más reciente que calce el patrón.
BODEGA_DIR = os.getenv("BODEGA_DIR", r"C:\Users\Windows\Dropbox\CAMARICO 2023")
BODEGA_PATRON = os.getenv("BODEGA_PATRON", "BODEGA*ENTRADAS-SALIDAS*.xls*")


def archivo_bodega_vigente(carpeta: str = None, patron: str = None) -> str:
    """Devuelve el Excel de bodega MÁS RECIENTE (incluidas copias en conflicto)."""
    carpeta = carpeta or BODEGA_DIR
    patron = patron or BODEGA_PATRON
    candidatos = [p for p in glob.glob(os.path.join(carpeta, patron))
                  if not os.path.basename(p).startswith("~$")]
    if not candidatos:
        raise FileNotFoundError(
            f"No encuentro ningún Excel de bodega en {carpeta} (patrón: {patron})")
    elegido = max(candidatos, key=os.path.getmtime)
    if len(candidatos) > 1:
        logger.info(f"Bodega: {len(candidatos)} archivos, uso el más nuevo "
                    f"→ {os.path.basename(elegido)}")
    return elegido
# Hoja "STOCK 26": tabla dinámica de la temporada vigente (lo que revisa el equipo).
# Bloque derecho = Producto | STOCK | FECHA VENCIMIENTO
HOJA = "STOCK 26"
C_NOMBRE, C_STOCK, C_VENC = 7, 8, 9

TOLERANCIA = 0.01      # diferencia relativa aceptada
TOL_ABS = 0.5          # diferencia absoluta mínima para reportar

# Productos que en la planilla de bodega van en otra unidad que en el Master.
# Factor = cuánto multiplicar el valor de bodega para llegar a la unidad del Master.
# Splendor: la bodega lo lleva en cc (1.250 cc) y el Master en litros (1,25 L).
FACTOR_UNIDAD = {"splendor": 0.001}


def _norm(s: str) -> str:
    """Normaliza nombres de producto para poder cruzarlos entre planillas."""
    s = str(s or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    s = re.sub(r"\b(sc|ec|wp|wg|sp|sl|ew|cs|se|od)\b", " ", s)   # formulaciones
    s = re.sub(r"\d+\s*%", " ", s)                                # concentraciones
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(".", "").replace(",", "."))
    except (ValueError, AttributeError):
        return 0.0


def leer_stock_bodega(path: str = None) -> dict:
    """Stock por producto = ingresos − salidas, desde el Excel de bodega."""
    path = path or archivo_bodega_vigente()
    if not os.path.exists(path):
        raise FileNotFoundError(f"No encuentro el Excel de bodega: {path}")
    tmp = os.path.join(tempfile.gettempdir(), "bodega_chk.xlsx")
    shutil.copy2(path, tmp)          # copia: el original puede estar abierto/sincronizando
    wb = load_workbook(tmp, read_only=True, data_only=True)
    if HOJA not in wb.sheetnames:
        wb.close()
        raise ValueError(f"El Excel de bodega no tiene la hoja '{HOJA}'")
    ws = wb[HOJA]

    # Filas de producto = las que traen un número en la columna STOCK.
    # Las filas de categoría (ACARICIDA, FUNGICIDA…) y totales se saltan.
    OMITIR = ("etiquetas de", "total general", "suma de")
    out = {}
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < C_VENC:
            continue
        nombre = str(row[C_NOMBRE - 1] or "").strip()
        bruto = row[C_STOCK - 1]
        if not nombre or bruto is None or str(bruto).strip() == "":
            continue
        if any(o in nombre.lower() for o in OMITIR):
            continue
        try:
            cantidad = float(bruto)
        except (TypeError, ValueError):
            continue
        k = _norm(nombre)
        if not k:
            continue
        out[k] = {"nombre": nombre, "cantidad": round(cantidad, 3),
                  "unidad": "", "vencimiento": row[C_VENC - 1]}
    wb.close()
    return out


def movimientos_recientes(path: str = None, dias: int = 30) -> list:
    """Entradas/salidas registradas en la hoja TRABAJO en los últimos N días.

    IMPORTANTE: el stock se lee de la tabla dinámica 'STOCK 26', que solo se
    recalcula cuando alguien abre el Excel y la refresca. Por eso se listan
    aparte los movimientos crudos: si hay movimientos recientes, el pivote
    (y por lo tanto la comparación) puede estar desactualizado.
    """
    from datetime import date as _date, datetime as _dt, timedelta
    path = path or archivo_bodega_vigente()
    tmp = os.path.join(tempfile.gettempdir(), "bodega_mov_chk.xlsx")
    shutil.copy2(path, tmp)
    wb = load_workbook(tmp, read_only=True, data_only=True)
    if "TRABAJO" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["TRABAJO"]
    corte = _date.today() - timedelta(days=dias)
    out = []
    for row in ws.iter_rows(min_row=8, values_only=True):
        if not row or len(row) < 35:
            continue
        f = row[2]
        if isinstance(f, _dt):
            f = f.date()
        if not isinstance(f, _date) or f < corte:
            continue
        tipo = str(row[6] or "").strip().upper()
        if tipo not in ("INGRESOS", "SALIDAS"):
            continue
        try:
            q = float(row[34] or 0)
        except (TypeError, ValueError):
            q = 0
        out.append({"fecha": f, "tipo": tipo, "producto": str(row[14] or "")[:40],
                    "cantidad": q, "unidad": str(row[26] or "")[:8],
                    "destino": str(row[7] or "")[:20]})
    wb.close()
    out.sort(key=lambda m: m["fecha"])
    return out


def leer_inventario_master() -> dict:
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["Inventario"]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        k = _norm(row[0])
        if not k:
            continue
        out[k] = {"nombre": str(row[0]).strip(),
                  "cantidad": _num(row[3]),
                  "unidad": str(row[2] or "")}
    wb.close()
    return out


def comparar(path: str = None) -> dict:
    """Compara bodega vs Master. Devuelve diferencias y faltantes."""
    path = path or archivo_bodega_vigente()
    bodega = leer_stock_bodega(path)
    master = leer_inventario_master()
    import datetime as _dt
    info_archivo = (f"{os.path.basename(path)} "
                    f"({_dt.datetime.fromtimestamp(os.path.getmtime(path)):%d-%m-%Y %H:%M})")

    difs, solo_bodega, solo_master = [], [], []
    for k, b in bodega.items():
        m = master.get(k)
        if not m:
            if abs(b["cantidad"]) > TOL_ABS:
                solo_bodega.append(b)
            continue
        # Ajuste de unidad cuando bodega y Master usan escalas distintas
        cant_bodega = b["cantidad"]
        for clave, factor in FACTOR_UNIDAD.items():
            if clave in k:
                cant_bodega = round(cant_bodega * factor, 4)
                break
        b = {**b, "cantidad": cant_bodega}
        dif = b["cantidad"] - m["cantidad"]
        base = max(abs(b["cantidad"]), abs(m["cantidad"]), 1.0)
        if abs(dif) > TOL_ABS and abs(dif) / base > TOLERANCIA:
            difs.append({"producto": m["nombre"], "bodega": b["cantidad"],
                          "master": m["cantidad"], "dif": round(dif, 3),
                          "unidad": m["unidad"] or b["unidad"]})
    for k, m in master.items():
        if k not in bodega and abs(m["cantidad"]) > TOL_ABS:
            solo_master.append(m)

    difs.sort(key=lambda x: -abs(x["dif"]))
    try:
        recientes = movimientos_recientes(path, dias=30)
    except Exception as e:
        logger.warning(f"No pude leer movimientos recientes: {e}")
        recientes = []

    return {"diferencias": difs, "solo_bodega": solo_bodega,
            "solo_master": solo_master, "archivo": info_archivo,
            "movimientos_recientes": recientes,
            "n_bodega": len(bodega), "n_master": len(master)}


def formato_alerta(res: dict) -> str:
    d = res["diferencias"]
    recientes = res.get("movimientos_recientes") or []
    lines = ["📦 CHEQUEO INVENTARIO — bodega vs Master", ""]
    if res.get("archivo"):
        lines.append(f"Archivo: {res['archivo']}")
    lines.append(f"Productos: {res['n_bodega']} en bodega · {res['n_master']} en Master")

    if recientes:
        lines.append(f"\n⚠️ {len(recientes)} movimiento(s) recientes en la hoja TRABAJO.")
        lines.append("   El stock se lee de la tabla dinámica, que NO se actualiza sola:")
        lines.append("   ábrela en Excel y refréscala para que estos queden reflejados.")
        for m in recientes[-8:]:
            signo = "+" if m["tipo"] == "INGRESOS" else "−"
            lines.append(f"   {m['fecha']} {signo}{m['cantidad']:g} {m['unidad']} "
                          f"{m['producto'][:28]} → {m['destino']}")

    if not d and not res["solo_bodega"] and not res["solo_master"]:
        lines.append("\n✅ El stock de la planilla calza con el Master."
                     + ("  (ojo con los movimientos de arriba)" if recientes else ""))
        return "\n".join(lines)
    if d:
        lines.append(f"\n⚠️ Diferencias de stock: {len(d)}")
        for x in d[:15]:
            lines.append(f"  {x['producto'][:26]:26} bodega {x['bodega']:g} "
                          f"vs master {x['master']:g}  ({x['dif']:+g} {x['unidad']})")
        if len(d) > 15:
            lines.append(f"  … y {len(d) - 15} más")
    if res["solo_bodega"]:
        lines.append(f"\n➕ Solo en bodega (faltan en el Master): {len(res['solo_bodega'])}")
        for x in res["solo_bodega"][:8]:
            lines.append(f"  {x['nombre'][:30]} — {x['cantidad']:g} {x['unidad']}")
    if res["solo_master"]:
        lines.append(f"\n➖ Solo en el Master (no están en bodega): {len(res['solo_master'])}")
        for x in res["solo_master"][:8]:
            lines.append(f"  {x['nombre'][:30]} — {x['cantidad']:g} {x['unidad']}")
    return "\n".join(lines)
