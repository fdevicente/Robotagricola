"""modules/banco_import.py — Importa la cartola del banco desde archivo.

Alternativa manual al scraper: el usuario descarga la cartola del portal
(CSV/TXT o Excel) y la sube; el bot la revisa y agrega solo lo nuevo.

Formato esperado (Scotiabank "typeDesc"):
    Fecha,Descripcion,Sucursal,N° Doc.,Cargos,Abonos,Saldo
    24-07-2026,CARGO COMEX 50765765,NUEVA MORANDE,50765765,-11260,,110735196

Deduplicación:
  1) por N° de documento (único por movimiento) cuando el Master lo tiene;
  2) si no, por multiconjunto (fecha, cargo, abono) — cuenta repeticiones, de
     modo que 9 transferencias idénticas el mismo día se importan las 9.
"""
import csv
import io
import logging
import os
from collections import Counter
from datetime import date, datetime

from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb

logger = logging.getLogger(__name__)

BANCO_SHEET = "Cuenta Banco"
COL_FECHA, COL_DESC, COL_REF, COL_CARGO, COL_ABONO, COL_SALDO = 1, 2, 3, 4, 5, 6


def _pd(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for f in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(v.strip()[:10], f).date()
            except Exception:
                pass
    return None


def _num(v):
    """'-11260' / '1.234' / '' → float (0 si vacío)."""
    s = str(v or "").strip().replace("$", "").replace(" ", "")
    if not s:
        return 0.0
    # separador de miles con punto y decimal con coma (formato chileno)
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif s.count(".") > 1:
        s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _leer_texto(path: str) -> str:
    """Lee el archivo probando codificaciones típicas de exportes bancarios."""
    with open(path, "rb") as fh:
        raw = fh.read()
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1", errors="replace")


def parsear_cartola(path: str) -> list[dict]:
    """Devuelve [{fecha, desc, doc, cargo, abono, saldo}] desde CSV/TXT o XLSX."""
    ext = os.path.splitext(path)[1].lower()
    filas = []

    if ext in (".xlsx", ".xlsm", ".xls"):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True):
            filas.append(["" if c is None else c for c in row])
        wb.close()
    else:
        texto = _leer_texto(path)
        sample = texto[:2000]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            delim = dialect.delimiter
        except Exception:
            delim = ";" if sample.count(";") > sample.count(",") else ","
        filas = [r for r in csv.reader(io.StringIO(texto), delimiter=delim)]

    if not filas:
        return []

    # Localizar encabezado (puede haber líneas previas de título)
    idx_hdr = None
    for i, r in enumerate(filas[:15]):
        joined = " ".join(str(c).lower() for c in r)
        if "fecha" in joined and ("cargo" in joined or "abono" in joined):
            idx_hdr = i
            break
    if idx_hdr is None:
        idx_hdr = 0

    hdr = [str(c).strip().lower() for c in filas[idx_hdr]]

    def col(*claves, default=None):
        for i, h in enumerate(hdr):
            for k in claves:
                if k in h:
                    return i
        return default

    i_fecha = col("fecha", default=0)
    i_desc = col("descrip", "glosa", "detalle", default=1)
    i_doc = col("doc", "documento", "n°", "nro", default=None)
    i_cargo = col("cargo", "debe", default=None)
    i_abono = col("abono", "haber", "deposito", default=None)
    i_saldo = col("saldo", default=None)

    movs = []
    for r in filas[idx_hdr + 1:]:
        if not r or all(str(c).strip() == "" for c in r):
            continue
        f = _pd(r[i_fecha] if i_fecha < len(r) else None)
        if not f:
            continue
        cargo = abs(_num(r[i_cargo])) if i_cargo is not None and i_cargo < len(r) else 0.0
        abono = abs(_num(r[i_abono])) if i_abono is not None and i_abono < len(r) else 0.0
        if cargo == 0 and abono == 0:
            continue
        movs.append({
            "fecha": f,
            "desc": str(r[i_desc]).strip() if i_desc < len(r) else "",
            "doc": str(r[i_doc]).strip() if i_doc is not None and i_doc < len(r) else "",
            "cargo": cargo,
            "abono": abono,
            "saldo": _num(r[i_saldo]) if i_saldo is not None and i_saldo < len(r) else None,
        })
    return _en_orden_cronologico(movs)


def _en_orden_cronologico(movs: list[dict]) -> list[dict]:
    """Deja los movimientos del más viejo al más nuevo, respetando el orden
    DENTRO de cada día.

    El banco entrega la cartola del más nuevo al más viejo. Si solo se ordena
    por fecha (orden estable), los varios movimientos de un mismo día quedan
    invertidos y el saldo que se lee como "último" es el del primer movimiento
    del día, no el de cierre. Eso hacía que la caja mostrara $94,3M en vez de
    los $80,7M reales.
    """
    if len(movs) < 2:
        return movs
    fechas = [m["fecha"] for m in movs]
    descendente = sum(1 for a, b in zip(fechas, fechas[1:]) if a > b) > \
                  sum(1 for a, b in zip(fechas, fechas[1:]) if a < b)
    # Si el archivo viene del más nuevo al más viejo, se invierte el orden
    # original dentro de cada fecha; si ya venía ascendente, se conserva.
    orden = -1 if descendente else 1
    return [m for _, m in sorted(enumerate(movs),
                                  key=lambda p: (p[1]["fecha"], orden * p[0]))]


def _clave(fecha, cargo, abono, decimales=0):
    """Clave de deduplicación. En USD los montos traen centavos y redondear a
    entero juntaría movimientos distintos, así que se conserva el decimal."""
    return (fecha.isoformat(), round(cargo, decimales), round(abono, decimales))


def _indices_master(hoja: str = None, decimales: int = 0):
    """Devuelve (docs_existentes, conteo_por_clave, ultima_fila, ultima_fecha)."""
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[hoja or BANCO_SHEET]
    docs = set()
    conteo = Counter()
    ultima_fecha = None
    n = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        f = _pd(row[COL_FECHA - 1]) if row else None
        if not f:
            continue
        n += 1
        ref = str(row[COL_REF - 1] or "").strip()
        if ref:
            docs.add(ref)
        conteo[_clave(f, _num(row[COL_CARGO - 1]),
                       _num(row[COL_ABONO - 1]), decimales)] += 1
        if not ultima_fecha or f > ultima_fecha:
            ultima_fecha = f
    wb.close()
    return docs, conteo, n, ultima_fecha


def analizar_cartola(path: str, hoja: str = None, decimales: int = 0) -> dict:
    """Simulación: qué se agregaría (no escribe nada).

    `hoja` permite analizar la cartola de la cuenta dólar con la misma lógica;
    `decimales` conserva los centavos al deduplicar montos en USD.
    """
    movs = parsear_cartola(path)
    docs, conteo, _, ultima_fecha = _indices_master(hoja, decimales)

    nuevos, dups = [], []
    vistos = Counter()
    for m in movs:
        clave = _clave(m["fecha"], m["cargo"], m["abono"], decimales)
        if m["doc"] and m["doc"] in docs:
            dups.append(m)
            continue
        # multiconjunto: solo es duplicado si el Master ya tiene tantas copias
        vistos[clave] += 1
        if vistos[clave] <= conteo.get(clave, 0):
            dups.append(m)
        else:
            nuevos.append(m)

    # `movs` ya viene cronológico; no reordenar, se perdería el orden del día.
    # El saldo de cierre es el del ÚLTIMO movimiento, no el de `max(fecha)`:
    # con varios movimientos el mismo día, max() devuelve el primero.
    saldo_final = movs[-1].get("saldo") if movs else None
    return {
        "total_archivo": len(movs),
        "nuevos": nuevos,
        "duplicados": len(dups),
        "ultima_fecha_master": ultima_fecha,
        "saldo_archivo": saldo_final,
        "cargos": sum(m["cargo"] for m in nuevos),
        "abonos": sum(m["abono"] for m in nuevos),
    }


def importar_cartola(path: str, hoja: str = None, decimales: int = 0) -> dict:
    """Agrega al Master los movimientos nuevos. Devuelve resumen."""
    res = analizar_cartola(path, hoja, decimales)
    nuevos = res["nuevos"]
    if not nuevos:
        return {"agregados": 0, **res}

    wb = load_workbook(EXCEL_PATH)
    ws = wb[hoja or BANCO_SHEET]
    fila = ws.max_row
    while fila > 1 and not ws.cell(fila, COL_FECHA).value:
        fila -= 1

    for m in nuevos:
        fila += 1
        ws.cell(fila, COL_FECHA).value = m["fecha"]
        ws.cell(fila, COL_DESC).value = m["desc"]
        ws.cell(fila, COL_REF).value = m["doc"]      # N° doc → dedup futuro exacto
        ws.cell(fila, COL_CARGO).value = m["cargo"] or None
        ws.cell(fila, COL_ABONO).value = m["abono"] or None
        ws.cell(fila, COL_SALDO).value = m["saldo"]
        # Categoría/cultivo quedan vacíos → los completa la categorización
    _save_wb(wb)
    wb.close()
    logger.info(f"Cartola importada: {len(nuevos)} movimientos nuevos desde {os.path.basename(path)}")
    return {"agregados": len(nuevos), **res}


def formato_resumen(res: dict) -> str:
    lines = ["🏦 CARTOLA DEL BANCO", ""]
    lines.append(f"Movimientos en el archivo: {res['total_archivo']}")
    lines.append(f"Ya estaban en el Master: {res['duplicados']}")
    lines.append(f"NUEVOS a agregar: {len(res['nuevos'])}")
    if res.get("ultima_fecha_master"):
        lines.append(f"Master llegaba hasta: {res['ultima_fecha_master']}")
    if res["nuevos"]:
        lines.append(f"Rango nuevo: {res['nuevos'][0]['fecha']} → {res['nuevos'][-1]['fecha']}")
        lines.append(f"Cargos ${res['cargos']:,.0f} · Abonos ${res['abonos']:,.0f}")
    if res.get("saldo_archivo"):
        lines.append(f"Saldo según cartola: ${res['saldo_archivo']:,.0f}")
    lines.append("")
    for m in res["nuevos"][:25]:
        t = f"-${m['cargo']:,.0f}" if m["cargo"] else f"+${m['abono']:,.0f}"
        lines.append(f"  {m['fecha']} {t:>14}  {m['desc'][:40]}")
    if len(res["nuevos"]) > 25:
        lines.append(f"  … y {len(res['nuevos']) - 25} más")
    return "\n".join(lines)
