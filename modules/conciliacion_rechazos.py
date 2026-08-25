"""modules/conciliacion_rechazos.py — Sugerencias descartadas por el usuario.

Sin esto el conciliador vuelve a proponer el mismo par en cada corrida y la
vista de sugerencias se llena de ruido que ya se revisó. Un rechazo dice
"este cargo NO es el pago de ese documento" y se recuerda.

Hoja `Rechazos Conciliacion`: una fila por par descartado. Se puede deshacer.
"""
import logging
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from config import EXCEL_PATH
from excel_manager import _save_wb

logger = logging.getLogger(__name__)

SHEET = "Rechazos Conciliacion"
HEADERS = ["ID", "Fecha", "Fila Banco", "Fecha Mov", "Descripción Mov",
           "Monto Mov", "N° Doc", "Proveedor", "Monto Doc", "Criterio",
           "Motivo", "Usuario"]
_WIDTHS = [6, 12, 10, 12, 40, 13, 13, 30, 13, 16, 26, 14]


def _norm(s) -> str:
    return " ".join(str(s or "").upper().split())


def clave(fila_banco, nro_doc, proveedor="") -> tuple:
    """Identifica un par cargo↔documento.

    Se usa el proveedor además del número porque distintos proveedores repiten
    numeración de facturas (ya nos mordió al cruzar Master con FXP).
    """
    return (int(fila_banco or 0), _norm(nro_doc), _norm(proveedor))


def crear_hoja(wb=None, excel_path: str | None = None) -> None:
    """Crea la hoja si no existe. Idempotente."""
    ruta = excel_path or EXCEL_PATH
    propio = wb is None
    if propio:
        wb = load_workbook(ruta)
    if SHEET not in wb.sheetnames:
        ws = wb.create_sheet(SHEET)
        fill = PatternFill("solid", fgColor="7B241C")
        for i, h in enumerate(HEADERS, 1):
            c = ws.cell(1, i, h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = fill
            c.alignment = Alignment(horizontal="center")
        for i, w in enumerate(_WIDTHS, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = "A2"
        if propio:
            _save_wb(wb, ruta)
            logger.info(f"Hoja '{SHEET}' creada")
    if propio:
        wb.close()


def _next_id(ws) -> int:
    m = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row and isinstance(row[0], (int, float)):
            m = max(m, int(row[0]))
    return m + 1


def rechazados(excel_path: str | None = None) -> set:
    """Claves de los pares ya descartados."""
    try:
        wb = load_workbook(excel_path or EXCEL_PATH, read_only=True, data_only=True)
    except Exception as e:
        logger.warning(f"No pude leer rechazos: {e}")
        return set()
    try:
        if SHEET not in wb.sheetnames:
            return set()
        out = set()
        for row in wb[SHEET].iter_rows(min_row=2, max_col=8, values_only=True):
            if not row or row[2] is None:
                continue
            out.add(clave(row[2], row[6], row[7]))
        return out
    finally:
        wb.close()


def registrar(pares: list, usuario: str = "", motivo: str = "",
              excel_path: str | None = None) -> int:
    """Guarda pares descartados. Devuelve cuántos se agregaron.

    Cada par: {fila_banco, fecha_mov, desc_mov, monto_mov, nro, prov,
               total, criterio}
    """
    if not pares:
        return 0
    ruta = excel_path or EXCEL_PATH
    wb = load_workbook(ruta)
    crear_hoja(wb, ruta)
    ws = wb[SHEET]

    ya = set()
    for row in ws.iter_rows(min_row=2, max_col=8, values_only=True):
        if row and row[2] is not None:
            ya.add(clave(row[2], row[6], row[7]))

    nid = _next_id(ws)
    fila = ws.max_row
    agregados = 0
    for p in pares:
        k = clave(p.get("fila_banco"), p.get("nro"), p.get("prov"))
        if k in ya:
            continue          # ya estaba descartado: no duplicar
        ya.add(k)
        fila += 1
        ws.cell(fila, 1).value = nid
        ws.cell(fila, 2).value = date.today()
        ws.cell(fila, 3).value = int(p.get("fila_banco") or 0)
        ws.cell(fila, 4).value = _fecha(p.get("fecha_mov"))
        ws.cell(fila, 5).value = str(p.get("desc_mov") or "")[:120]
        ws.cell(fila, 6).value = float(p.get("monto_mov") or 0)
        ws.cell(fila, 7).value = str(p.get("nro") or "")
        ws.cell(fila, 8).value = str(p.get("prov") or "")
        ws.cell(fila, 9).value = float(p.get("total") or 0)
        ws.cell(fila, 10).value = str(p.get("criterio") or "")
        ws.cell(fila, 11).value = motivo or "Descartado por el usuario"
        ws.cell(fila, 12).value = usuario
        nid += 1
        agregados += 1

    if agregados:
        _save_wb(wb, ruta)
        logger.info(f"Rechazos registrados: {agregados}")
    wb.close()
    return agregados


def deshacer(id_rechazo: int, excel_path: str | None = None) -> bool:
    """Borra un rechazo para que la sugerencia vuelva a aparecer."""
    ruta = excel_path or EXCEL_PATH
    wb = load_workbook(ruta)
    if SHEET not in wb.sheetnames:
        wb.close()
        return False
    ws = wb[SHEET]
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, 1).value == id_rechazo:
            ws.delete_rows(r)
            _save_wb(wb, ruta)
            wb.close()
            logger.info(f"Rechazo {id_rechazo} deshecho")
            return True
    wb.close()
    return False


def listar(excel_path: str | None = None) -> list:
    """Rechazos guardados, del más reciente al más antiguo."""
    wb = load_workbook(excel_path or EXCEL_PATH, read_only=True, data_only=True)
    try:
        if SHEET not in wb.sheetnames:
            return []
        out = []
        for row in wb[SHEET].iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            out.append({
                "id": int(row[0]), "fecha": _iso(row[1]),
                "fila_banco": row[2], "fecha_mov": _iso(row[3]),
                "desc_mov": row[4], "monto_mov": row[5],
                "nro": row[6], "prov": row[7], "total": row[8],
                "criterio": row[9], "motivo": row[10], "usuario": row[11],
            })
        out.sort(key=lambda x: -(x["id"] or 0))
        return out
    finally:
        wb.close()


def _fecha(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _iso(v):
    f = _fecha(v)
    return f.isoformat() if f else None
