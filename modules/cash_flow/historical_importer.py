"""Onboarding: detecta patrones en Cuenta Banco para clasificar Tipo
sin llamar a Claude.
"""
import logging
import re
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import (
    CUENTA_BANCO_SHEET, COL_BANCO_TIPO, _save_wb,
)

logger = logging.getLogger(__name__)


PATTERNS = [
    (re.compile(r"venta.*dolares|venta.*dolar"), "venta_dolares", False, False),
    (re.compile(r"valbifrut|pacific\s*nuts|vitakai"), "ingreso_clp", True, False),
    (re.compile(r"remuneracion|sueldo|liquidacion|finiquito"), "sueldo", False, True),
    (re.compile(r"honorario|boleta\s+honorario"), "honorario", False, True),
]


def detect_income_patterns(excel_path: str | None = None) -> dict:
    """Clasifica Tipo en Cuenta Banco usando regex. No toca filas ya tagueadas."""
    excel_path = excel_path or EXCEL_PATH
    wb = load_workbook(excel_path)
    ws = wb[CUENTA_BANCO_SHEET]

    counts = {
        "venta_dolares": 0, "ingreso_clp": 0,
        "sueldo": 0, "honorario": 0,
        "skipped_tagged": 0, "no_match": 0,
    }

    for r in range(2, ws.max_row + 1):
        descripcion = str(ws.cell(r, 2).value or "")
        if not descripcion:
            continue
        if ws.cell(r, COL_BANCO_TIPO).value:
            counts["skipped_tagged"] += 1
            continue

        cargo = float(ws.cell(r, 4).value or 0)
        abono = float(ws.cell(r, 5).value or 0)
        desc_low = descripcion.lower()

        matched = False
        for rx, tipo, req_abono, req_cargo in PATTERNS:
            if rx.search(desc_low):
                if req_abono and abono <= 0:
                    continue
                if req_cargo and cargo <= 0:
                    continue
                ws.cell(r, COL_BANCO_TIPO, tipo)
                counts[tipo] += 1
                matched = True
                break

        if not matched:
            counts["no_match"] += 1

    _save_wb(wb, excel_path)
    wb.close()
    logger.info(f"detect_income_patterns: {counts}")
    return counts
