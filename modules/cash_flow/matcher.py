"""Match facturas pendientes con cargos bancarios."""
from datetime import date, datetime


def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(v[:10], fmt).date()
            except ValueError:
                pass
    return None


def _provider_match(provider: str, description: str) -> float:
    p = (provider or "").lower()
    d = (description or "").lower()
    words = [w for w in p.split() if len(w) > 3]
    if not words:
        return 0.0
    hits = sum(1 for w in words if w in d)
    return hits / len(words)


def match_score(factura: dict, bank_mov: dict) -> float:
    """Devuelve score 0-200. >=100 candidato fuerte. >=30 ambiguo."""
    cargo = float(bank_mov.get("cargo") or 0)
    if cargo <= 0:
        return 0.0

    total = float(factura.get("total") or 0)
    if total <= 0:
        return 0.0

    score = 0.0
    diff_pct = abs(cargo - total) / total
    if diff_pct < 0.001:
        score += 100
    elif diff_pct < 0.05:
        score += 30

    f_factura = _to_date(factura.get("fecha_emision"))
    f_banco = _to_date(bank_mov.get("fecha"))
    if f_factura and f_banco:
        diff_dias = abs((f_banco - f_factura).days)
        if diff_dias <= 5:
            score += 50
        elif diff_dias <= 15:
            score += 20

    score += 40 * _provider_match(factura.get("proveedor", ""),
                                   bank_mov.get("descripcion", ""))

    nro = str(factura.get("nro_factura") or "").strip()
    ref = str(bank_mov.get("referencia") or "")
    if nro and nro in ref:
        score += 50

    return score


MATCH_THRESHOLD = 30


def find_matches(bank_mov: dict, facturas_pendientes: list[dict]) -> list[dict]:
    """Devuelve candidatos sobre threshold, ordenados por score desc."""
    scored = []
    for f in facturas_pendientes:
        s = match_score(f, bank_mov)
        if s >= MATCH_THRESHOLD:
            scored.append({**f, "score": s})
    scored.sort(key=lambda x: x["score"], reverse=True)
    return scored


STRONG_THRESHOLD = 100
GAP_MIN = 30


def classify_match(candidates: list[dict]) -> dict:
    """Decide auto / ambiguo / no_match.

    - auto: top score >= STRONG_THRESHOLD y gap >= GAP_MIN con el 2do
    - ambiguo: top score >= MATCH_THRESHOLD pero no llega a auto
    - no_match: lista vacia
    """
    if not candidates:
        return {"status": "no_match"}

    top = candidates[0]
    if top["score"] < STRONG_THRESHOLD:
        return {"status": "ambiguo", "candidates": candidates}

    if len(candidates) == 1:
        return {"status": "auto", "fila": top["fila"], "score": top["score"]}

    gap = top["score"] - candidates[1]["score"]
    if gap < GAP_MIN:
        return {"status": "ambiguo", "candidates": candidates}

    return {"status": "auto", "fila": top["fila"], "score": top["score"]}


def match_new_bank_movements(excel_path=None, limit: int | None = None) -> dict:
    """Lee movimientos unlinked, busca matches, aplica auto-matches.

    Optimizado: abre Master 1 vez, escribe en memoria, guarda al final.
    Devuelve: {scanned, auto_matched, ambiguous, no_match, errors}
    """
    from openpyxl import load_workbook
    from config import EXCEL_PATH
    from excel_manager import (
        SHEET_NAME, CUENTA_BANCO_SHEET, _save_wb,
        COL_BANCO_TIPO, COL_BANCO_FACTURA_LINK,
        read_facturas_pendientes, read_bank_movements_unlinked,
    )
    excel_path = excel_path or EXCEL_PATH

    movs = read_bank_movements_unlinked(excel_path)
    pendientes = read_facturas_pendientes(excel_path)
    if limit is not None:
        movs = movs[:limit]

    report = {
        "scanned": len(movs), "auto_matched": 0,
        "ambiguous": 0, "no_match": 0, "errors": 0,
    }

    wb = load_workbook(excel_path)
    ws_f = wb[SHEET_NAME]
    ws_b = wb[CUENTA_BANCO_SHEET]
    dirty = False

    for mov in movs:
        try:
            candidates = find_matches(mov, pendientes)
            decision = classify_match(candidates)
            if decision["status"] == "auto":
                fila = decision["fila"]
                factura = next((f for f in pendientes if f["fila"] == fila), None)
                if not factura:
                    report["errors"] += 1
                    continue
                fecha = mov["fecha"]
                fecha_str = (fecha.strftime("%Y-%m-%d")
                              if hasattr(fecha, "strftime") else str(fecha)[:10])
                # Write in-memory
                if not ws_f.cell(fila, 3).value:
                    ws_f.cell(fila, 3, f"{fecha_str} (Banco)")
                ws_b.cell(mov["fila"], COL_BANCO_TIPO, "factura")
                ws_b.cell(mov["fila"], COL_BANCO_FACTURA_LINK,
                          str(factura.get("nro_factura") or ""))
                dirty = True
                pendientes = [f for f in pendientes if f["fila"] != fila]
                report["auto_matched"] += 1
            elif decision["status"] in ("ambiguous", "ambiguo"):
                report["ambiguous"] += 1
            else:
                report["no_match"] += 1
        except Exception:
            report["errors"] += 1

    if dirty:
        _save_wb(wb, excel_path)
    wb.close()
    return report
