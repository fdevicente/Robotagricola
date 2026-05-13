"""Cliente Claude para categorizar facturas y cargos bancarios.

Usa requests directo (mismo patron que processors/extractor.py).
"""
import logging
import os
import requests
from openpyxl import load_workbook

from config import ANTHROPIC_API_KEY, EXCEL_PATH, DROPBOX_BACKUP_PATH
from excel_manager import (
    SHEET_NAME, _save_wb,
    COL_CATEGORIA, COL_CULTIVO, COL_CONFIANZA, COL_CATEGORIZADO_POR,
    CUENTA_BANCO_SHEET, COL_BANCO_TIPO, COL_BANCO_CATEGORIA, COL_BANCO_CULTIVO,
)
from modules.cash_flow.prompt import (
    build_categorization_prompt,
    parse_categorization_response,
)
from modules.cash_flow.categorizer_cache import CategorizerCache

logger = logging.getLogger(__name__)

CLAUDE_URL = "https://api.anthropic.com/v1/messages"
CLAUDE_MODELS = ["claude-haiku-4-5-20251001", "claude-sonnet-4-6"]
MAX_TOKENS = 200
TIMEOUT_SEC = 30


def categorize_raw(proveedor: str, glosa: str, glosa_ii: str,
                    monto: float, fecha: str) -> dict:
    """Llama a Claude directo, sin cache. Devuelve dict con categoria/cultivo/confianza/razon."""
    prompt = build_categorization_prompt(proveedor, glosa, glosa_ii, monto, fecha)

    headers = {
        "x-api-key": ANTHROPIC_API_KEY,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    payload_base = {
        "max_tokens": MAX_TOKENS,
        "messages": [{"role": "user", "content": prompt}],
    }

    for model in CLAUDE_MODELS:
        payload = {**payload_base, "model": model}
        try:
            resp = requests.post(CLAUDE_URL, headers=headers,
                                  json=payload, timeout=TIMEOUT_SEC)
        except requests.RequestException as e:
            logger.warning(f"Claude {model} excepcion: {e}")
            continue

        if resp.status_code != 200:
            logger.warning(f"Claude {model} HTTP {resp.status_code}: {resp.text[:200]}")
            continue

        try:
            data = resp.json()
            raw_text = data["content"][0]["text"]
        except (KeyError, IndexError, ValueError) as e:
            logger.warning(f"Claude {model} respuesta inesperada: {e}")
            continue

        return parse_categorization_response(raw_text)

    logger.error(f"Todos los modelos Claude fallaron para {proveedor[:40]}")
    return {
        "categoria": "REVISAR",
        "cultivo": "GENERAL",
        "confianza": 0.0,
        "razon": "Claude API fallo",
    }


DEFAULT_CACHE_PATH = os.path.join(DROPBOX_BACKUP_PATH, "categorizer_cache.json")
CONFIANZA_REVISAR = 0.85


def _get_cache(cache_path=None) -> CategorizerCache:
    return CategorizerCache(cache_path or DEFAULT_CACHE_PATH)


def _read_invoice_row(ws, row: int) -> dict:
    return {
        "fecha": str(ws.cell(row, 1).value or ""),
        "proveedor": str(ws.cell(row, 4).value or ""),
        "documento": str(ws.cell(row, 6).value or ""),
        "glosa": str(ws.cell(row, 8).value or ""),
        "glosa_ii": str(ws.cell(row, 9).value or ""),
        "monto": float(ws.cell(row, 15).value or 0),
    }


def categorize_invoice(row: int, excel_path=None, cache_path=None) -> dict:
    """Categoriza la fila `row` de Facturas y escribe cols Q-T en el Master."""
    excel_path = excel_path or EXCEL_PATH
    cache = _get_cache(cache_path)

    wb = load_workbook(excel_path)
    ws = wb[SHEET_NAME]
    data = _read_invoice_row(ws, row)
    wb.close()

    cached = cache.get(data["proveedor"], data["glosa"])
    if cached:
        result = cached
        source = "cache"
    else:
        result = categorize_raw(
            proveedor=data["proveedor"], glosa=data["glosa"],
            glosa_ii=data["glosa_ii"], monto=data["monto"],
            fecha=data["fecha"],
        )
        cache.set(data["proveedor"], data["glosa"], result)
        source = "claude"

    cat_to_write = result["categoria"]
    if result["confianza"] < CONFIANZA_REVISAR:
        cat_to_write = "REVISAR"

    wb = load_workbook(excel_path)
    ws = wb[SHEET_NAME]
    ws.cell(row, COL_CATEGORIA, cat_to_write)
    ws.cell(row, COL_CULTIVO, result["cultivo"])
    ws.cell(row, COL_CONFIANZA, result["confianza"])
    ws.cell(row, COL_CATEGORIZADO_POR, source)
    _save_wb(wb, excel_path)
    wb.close()

    return result


def _read_bank_row(ws, row: int) -> dict:
    return {
        "fecha": str(ws.cell(row, 1).value or ""),
        "descripcion": str(ws.cell(row, 2).value or ""),
        "referencia": str(ws.cell(row, 3).value or ""),
        "cargo": float(ws.cell(row, 4).value or 0),
        "abono": float(ws.cell(row, 5).value or 0),
    }


def categorize_bank_movement(row: int, excel_path=None, cache_path=None) -> dict:
    """Categoriza fila de Cuenta Banco. Abono>0 sin Cargo -> ingreso, no llama Claude."""
    excel_path = excel_path or EXCEL_PATH
    cache = _get_cache(cache_path)

    wb = load_workbook(excel_path)
    ws = wb[CUENTA_BANCO_SHEET]
    data = _read_bank_row(ws, row)
    wb.close()

    if data["abono"] > 0 and data["cargo"] == 0:
        result = {
            "tipo": "ingreso", "categoria": "", "cultivo": "",
            "confianza": 1.0, "razon": "abono detectado",
        }
    else:
        cached = cache.get(data["descripcion"], data["referencia"])
        if cached:
            base = cached
        else:
            base = categorize_raw(
                proveedor=data["descripcion"], glosa=data["referencia"],
                glosa_ii="", monto=data["cargo"], fecha=data["fecha"],
            )
            cache.set(data["descripcion"], data["referencia"], base)
        result = {**base, "tipo": "egreso"}

    wb = load_workbook(excel_path)
    ws = wb[CUENTA_BANCO_SHEET]
    ws.cell(row, COL_BANCO_TIPO, result["tipo"])
    if result["tipo"] != "ingreso":
        cat = result["categoria"]
        if result["confianza"] < CONFIANZA_REVISAR:
            cat = "REVISAR"
        ws.cell(row, COL_BANCO_CATEGORIA, cat)
        ws.cell(row, COL_BANCO_CULTIVO, result["cultivo"])
    _save_wb(wb, excel_path)
    wb.close()

    return result


def batch_categorize_history(excel_path=None, cache_path=None,
                               limit: int | None = None,
                               progress_cb=None) -> dict:
    """Itera Master.Facturas, categoriza filas sin Categoria. Backup pre-batch."""
    from infrastructure.backups import backup_master
    excel_path = excel_path or EXCEL_PATH

    try:
        backup_master(reason="pre-batch-categorize", excel_path=excel_path)
    except Exception as e:
        logger.warning(f"Backup pre-batch fallo (continuo): {e}")

    wb = load_workbook(excel_path, read_only=True)
    ws = wb[SHEET_NAME]
    pending_rows = []
    skipped = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 4).value is None:
            continue
        if ws.cell(r, COL_CATEGORIA).value:
            skipped += 1
            continue
        pending_rows.append(r)
    wb.close()

    if limit is not None:
        pending_rows = pending_rows[:limit]

    report = {
        "total_pending": len(pending_rows),
        "processed": 0, "skipped": skipped,
        "low_confidence": 0, "errors": 0,
    }

    for idx, row in enumerate(pending_rows, 1):
        try:
            result = categorize_invoice(row, excel_path=excel_path,
                                          cache_path=cache_path)
            report["processed"] += 1
            if result["confianza"] < CONFIANZA_REVISAR:
                report["low_confidence"] += 1
        except Exception as e:
            logger.error(f"Error fila {row}: {e}")
            report["errors"] += 1
        if progress_cb:
            progress_cb(idx, len(pending_rows))

    logger.info(f"Batch categorize OK: {report}")
    return report
