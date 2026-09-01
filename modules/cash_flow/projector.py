"""Motor de proyeccion de flujo de caja.

Calcula proyeccion mes x categoria x cultivo escalando un ano base
por el factor de hectareas + aplica ajustes manuales del usuario.
"""
from collections import defaultdict
from datetime import date, datetime
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import (
    SHEET_NAME, CUENTA_BANCO_SHEET,
    COSECHAS_SHEET, HECTAREAS_SHEET, AJUSTES_SHEET, FLUJO_CAJA_SHEET,
    CATEGORIAS, CULTIVOS, IVA_RATE,
)


def _to_year_month(v) -> tuple[int, int] | None:
    if isinstance(v, datetime):
        return (v.year, v.month)
    if isinstance(v, date):
        return (v.year, v.month)
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                d = datetime.strptime(v[:10], fmt).date()
                return (d.year, d.month)
            except ValueError:
                pass
    return None


EXCLUIR_CATS_EGRESO = {
    "TRANSFERENCIA INTERNA", "PRE-2021 HISTORICO", "REVISAR",
    "INGRESO VENTAS", "INGRESO OPERACIONAL", "INGRESO FINANCIERO",
    "INGRESO NOGALES", "INGRESO CEREZOS", "INGRESO AVELLANOS", "INGRESO GENERAL",
    "CAMBIO DIVISA",                    # neutro contablemente
    "PRESTAMOS A OTRAS SOCIEDADES",     # no operacional
    "REINTEGROS Y DEVOLUCIONES",        # ajustes contables
    "BONO VENTA NUECES",                 # se calcula como 8% de ventas, via ajuste manual
}


# Categorías que provienen SOLO del banco (no tienen factura asociada):
# sueldos, impuestos, comisiones bancarias, TAG, S-Invest, etc.
CATS_SOLO_BANCO = {
    "MANO DE OBRA PLANTA",         # remuneraciones mensuales
    "MANO DE OBRA TEMPORAL",        # BH jornaleros (Alpabesa también)
    "SERVICIOS PROFESIONALES",      # BH Francisco Donoso + Capital Office
    "IMPUESTOS",                    # F29 mensual + contribuciones
    "GASTOS BANCARIOS",             # comisiones, intereses
    "GASTOS VEHICULOS",             # TAG, permiso circulación, SOAP
    "LEASING",                      # cuotas
    "COSTO ENERGETICO",             # S-Invest paneles solares
    "ENERGIA",                      # CGE
    "TRANSFERENCIA INTERNA",
    "PRESTAMOS A OTRAS SOCIEDADES",
    "CAMBIO DIVISA",
    "REINTEGROS Y DEVOLUCIONES",
    "BONO VENTA NUECES",
}


def load_historical_egresos(excel_path: str | None = None,
                              year: int | None = None) -> dict:
    """Agrupa egresos por (year, month, categoria, cultivo) -> total.

    Estrategia para EVITAR doble conteo:
    - Para categorías típicamente de Facturas (insumos, fertilizantes, etc.):
      usar SOLO la hoja Facturas (con o sin fecha_pago).
    - Para categorías típicamente del banco (sueldos, impuestos, TAG, etc.):
      usar SOLO la hoja Cuenta Banco.
    """
    excel_path = excel_path or EXCEL_PATH
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    agg: dict = defaultdict(float)

    # FACTURAS: usar fecha de pago si existe, sino fecha emisión
    ws = wb[SHEET_NAME]
    for row in ws.iter_rows(min_row=2, max_col=20, values_only=True):
        proveedor = row[3]
        if not proveedor: continue
        categoria = row[16]
        if not categoria: continue
        cat_up = str(categoria).strip().upper()
        if cat_up in EXCLUIR_CATS_EGRESO: continue
        if cat_up in CATS_SOLO_BANCO: continue  # estas se cuentan solo del banco
        # Saltar NN
        cat_por = str(row[19] or "") if len(row) > 19 else ""
        if "NN-no-pagar" in cat_por: continue
        cultivo = row[17] or "GENERAL"
        total = row[14]
        if not total: continue
        fecha_uso = row[2] if row[2] else row[0]
        ym = _to_year_month(fecha_uso)
        if not ym: continue
        if year is not None and ym[0] != year: continue
        agg[(ym[0], ym[1], cat_up, cultivo)] += float(total)

    # BANCO: solo cargos en CATS_SOLO_BANCO
    if CUENTA_BANCO_SHEET in wb.sheetnames:
        ws_b = wb[CUENTA_BANCO_SHEET]
        for row in ws_b.iter_rows(min_row=2, max_col=9, values_only=True):
            if not row[0]: continue
            try:
                cargo = float(row[3] or 0)
            except (TypeError, ValueError):
                continue
            if cargo <= 0: continue
            categoria = row[7]
            if not categoria: continue
            cat_up = str(categoria).strip().upper()
            if cat_up in EXCLUIR_CATS_EGRESO: continue
            if cat_up not in CATS_SOLO_BANCO: continue  # estas se cuentan de facturas
            ym = _to_year_month(row[0])
            if not ym: continue
            if year is not None and ym[0] != year: continue
            cultivo = "GENERAL"
            agg[(ym[0], ym[1], cat_up, cultivo)] += cargo

    wb.close()
    return dict(agg)


def load_hectareas(excel_path: str | None = None) -> dict:
    """Devuelve {year: {cultivo: hc}}. Cultivos en uppercase."""
    excel_path = excel_path or EXCEL_PATH
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[HECTAREAS_SHEET]
    hc: dict = {}
    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        year = row[0]
        if not isinstance(year, int):
            continue
        hc[year] = {
            "NOGALES": float(row[1] or 0),
            "CEREZOS": float(row[2] or 0),
            "AVELLANOS": float(row[3] or 0),
        }
    wb.close()
    return hc


def _parse_mes_str(v):
    if isinstance(v, datetime):
        return (v.year, v.month)
    if isinstance(v, date):
        return (v.year, v.month)
    if isinstance(v, str):
        try:
            parts = v.split("-")
            return (int(parts[0]), int(parts[1]))
        except (IndexError, ValueError):
            return None
    return None


def load_ajustes_manuales(excel_path: str | None = None) -> list:
    """Devuelve lista de ajustes activos."""
    excel_path = excel_path or EXCEL_PATH
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[AJUSTES_SHEET]
    ajustes = []
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        if not row[1]:
            continue
        if row[6] is False:
            continue
        ym = _parse_mes_str(row[1])
        if not ym:
            continue
        try:
            monto = float(row[4] or 0)
        except (TypeError, ValueError):
            continue
        ajustes.append({
            "mes_proyectado": ym,
            "categoria": row[2],
            "cultivo": row[3] or "GENERAL",
            "monto": monto,
            "razon": row[5] or "",
        })
    wb.close()
    return ajustes


def load_expected_ingresos(excel_path: str | None = None) -> list:
    """Lee Cosechas, devuelve ingresos proyectados convertidos a CLP.

    La columna 17 "Aplica IVA" (SI/NO) solo pesa en las filas `esperado`: si el
    comprador paga en pesos agrega 19% encima del neto, y esa plata SÍ entra a la
    caja. Valbifrut lo hace (mayo-2026: neto $223.596.523 + IVA $42.483.234);
    Pacific paga en USD por COMEX y no lo agrega. Las filas `recibido` no se tocan
    porque ya guardan el efectivo real, IVA incluido.
    """
    from modules.cuentas import tipo_cambio

    excel_path = excel_path or EXCEL_PATH
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    # MISMA fuente que usa la caja: la hoja Config manda y config.py es el
    # respaldo. Leer CASH_FLOW_CONFIG directo era lo que permitia que el
    # dashboard mostrara la caja a un tipo de cambio y los ingresos a otro.
    usd_clp = tipo_cambio(wb)
    ws = wb[COSECHAS_SHEET]
    ingresos = []
    for row in ws.iter_rows(min_row=2, max_col=17, values_only=True):
        if not row[0]:
            continue
        estado = row[11]
        aplica_iva = str(row[16] or "").strip().upper() in ("SI", "SÍ", "S", "TRUE", "1")
        if estado == "recibido":
            monto_real = row[13]
            fecha = row[12]
            moneda = (row[14] or "CLP").upper()
            try:
                m_val = float(monto_real or 0)
            except (TypeError, ValueError):
                m_val = 0
            if m_val <= 0:
                continue
            monto_clp = m_val if moneda == "CLP" else m_val * usd_clp
            ym = _to_year_month(fecha)
        else:
            monto_usd = row[9]
            try:
                m_val = float(monto_usd or 0)
            except (TypeError, ValueError):
                m_val = 0
            if m_val <= 0:
                continue
            monto_clp = m_val * usd_clp
            if aplica_iva:
                monto_clp *= 1 + IVA_RATE
            ym = _to_year_month(row[8])
        if not ym:
            continue
        ingresos.append({
            "year": ym[0], "month": ym[1],
            "cultivo": row[1] or "GENERAL",
            "exportadora": row[3] or "",
            "tipo_cuota": row[10] or "",
            "estado": estado or "esperado",
            "aplica_iva": aplica_iva,
            "monto_clp": float(monto_clp),
        })
    wb.close()
    return ingresos


def compute_factor_hc(hc: dict, cultivo: str, base_year: int, target_year: int) -> float:
    """Factor de escalamiento por hectareas."""
    if base_year == target_year:
        return 1.0
    if base_year not in hc or target_year not in hc:
        return 1.0

    if cultivo.upper() == "GENERAL":
        base = sum(hc[base_year].values())
        target = sum(hc[target_year].values())
    else:
        base = hc[base_year].get(cultivo.upper(), 0)
        target = hc[target_year].get(cultivo.upper(), 0)

    if base <= 0:
        return 1.0
    return target / base


EXCLUIR_PROYECCION = {
    "MANTENIMIENTO HELICOPTERO",   # operación descontinuada
}
# Categorías excluidas del histórico (no replicar del año pasado), pero ajustes
# manuales sí se aplican.
EXCLUIR_HISTORICO_SOLO = {
    "LEASING",                      # ya pagado en dic-2025
    "COSTO ENERGETICO",             # ajuste manual mensual según clima
    "GASTOS VEHICULOS",             # ajuste manual a valor fijo
    "MANO DE OBRA TEMPORAL",        # ajuste manual: $25M nueces + $4M cerezas
    "MATERIALES",                   # ajuste manual: $4M/año
    "MANO DE OBRA PLANTA",          # ajuste manual: líquidos + Previred + aguinaldos
}


def compute_egresos_proyectados(historicos: dict, ajustes: list,
                                  hc: dict, base_year: int,
                                  target_year: int) -> dict:
    """Proyecta egresos del target_year escalando base_year + sumando ajustes.
    Excluye categorías en EXCLUIR_PROYECCION (operaciones descontinuadas)."""
    proj: dict = defaultdict(float)

    for (y, m, cat, cul), monto in historicos.items():
        if y != base_year:
            continue
        cu = (cat or "").upper()
        if cu in EXCLUIR_PROYECCION or cu in EXCLUIR_HISTORICO_SOLO:
            continue
        factor = compute_factor_hc(hc, cul, base_year, target_year)
        proj[(target_year, m, cat, cul)] += monto * factor

    for a in ajustes:
        ym = a["mes_proyectado"]
        if ym[0] != target_year:
            continue
        if (a["categoria"] or "").upper() in EXCLUIR_PROYECCION:
            continue
        # Ajustes para cats EXCLUIR_HISTORICO_SOLO sí se aplican
        key = (target_year, ym[1], a["categoria"], a["cultivo"])
        proj[key] += a["monto"]

    return dict(proj)


def compute_saldo_mensual(saldo_inicial: float, ingresos: list,
                            egresos: dict, months: list) -> dict:
    """Running balance mes a mes."""
    ing_mes: dict = defaultdict(float)
    for i in ingresos:
        ing_mes[(i["year"], i["month"])] += i["monto_clp"]

    eg_mes: dict = defaultdict(float)
    for (y, m, _cat, _cul), monto in egresos.items():
        eg_mes[(y, m)] += monto

    result = {}
    saldo = saldo_inicial
    for ym in months:
        ing = ing_mes.get(ym, 0)
        eg = eg_mes.get(ym, 0)
        saldo_cierre = saldo + ing - eg
        result[ym] = {
            "saldo_inicio": saldo,
            "ingresos": ing,
            "egresos": eg,
            "saldo_cierre": saldo_cierre,
        }
        saldo = saldo_cierre
    return result


_MESES_ES = ["", "Ene", "Feb", "Mar", "Abr", "May", "Jun",
              "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


def _month_label(year: int, month: int) -> str:
    return f"{_MESES_ES[month]}-{str(year)[-2:]}"


def write_flujo_caja(saldo_data: dict, egresos: dict, ingresos: list,
                       months: list, excel_path: str | None = None):
    """Regenera la hoja Flujo Caja con la proyeccion."""
    from excel_manager import _save_wb
    excel_path = excel_path or EXCEL_PATH
    wb = load_workbook(excel_path)
    if FLUJO_CAJA_SHEET in wb.sheetnames:
        del wb[FLUJO_CAJA_SHEET]
    ws = wb.create_sheet(FLUJO_CAJA_SHEET)

    header = ["SECCION"] + [_month_label(y, m) for (y, m) in months]
    ws.append(header)

    ws.append(["SALDO INICIAL"] + [saldo_data.get(ym, {}).get("saldo_inicio", 0)
                                    for ym in months])

    ws.append(["INGRESOS"])
    exportadoras = sorted({i.get("exportadora", "") for i in ingresos})
    for exp in exportadoras:
        if not exp:
            continue
        row = [f"  {exp}"]
        for ym in months:
            total = sum(i["monto_clp"] for i in ingresos
                          if i.get("exportadora") == exp
                          and i["year"] == ym[0] and i["month"] == ym[1])
            row.append(total)
        ws.append(row)
    ws.append(["  TOTAL INGRESOS"] + [saldo_data.get(ym, {}).get("ingresos", 0)
                                       for ym in months])

    ws.append(["EGRESOS"])
    cats_set = sorted({k[2] for k in egresos.keys()})
    for cat in cats_set:
        row = [f"  {cat}"]
        for ym in months:
            total = sum(monto for (y, m, c, _cu), monto in egresos.items()
                          if y == ym[0] and m == ym[1] and c == cat)
            row.append(total)
        ws.append(row)
    ws.append(["  TOTAL EGRESOS"] + [saldo_data.get(ym, {}).get("egresos", 0)
                                      for ym in months])

    ws.append(["SALDO CIERRE MES"] + [saldo_data.get(ym, {}).get("saldo_cierre", 0)
                                       for ym in months])

    ws.column_dimensions["A"].width = 28
    for i in range(2, len(months) + 2):
        ws.column_dimensions[chr(64 + i)].width = 14

    _save_wb(wb, excel_path)
    wb.close()


def get_cash_flow(start: tuple, end: tuple,
                    saldo_inicial: float,
                    base_year: int = 2025,
                    excel_path: str | None = None) -> dict:
    """API principal del projector."""
    historicos = load_historical_egresos(excel_path, year=base_year)
    hc = load_hectareas(excel_path)
    ajustes = load_ajustes_manuales(excel_path)
    ingresos = load_expected_ingresos(excel_path)

    months = []
    y, mo = start
    while (y, mo) <= end:
        months.append((y, mo))
        mo += 1
        if mo > 12:
            mo = 1
            y += 1

    egresos_proj = {}
    years_target = {ym[0] for ym in months}
    for ty in years_target:
        e = compute_egresos_proyectados(
            historicos=historicos, ajustes=ajustes, hc=hc,
            base_year=base_year, target_year=ty,
        )
        egresos_proj.update(e)

    saldo = compute_saldo_mensual(saldo_inicial, ingresos, egresos_proj, months)

    return {
        "months": months, "saldo": saldo,
        "egresos": egresos_proj, "ingresos": ingresos,
    }
