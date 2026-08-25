"""bitacora_manager.py — Bitácora estructurada de labores del campo.

Hoja 'Bitácora' ampliada con campos para calcular jornadas-hombre,
eficiencias y trazabilidad de aplicaciones.
"""
import logging
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import EXCEL_PATH
from excel_manager import _save_wb  # guardado con reintentos si Excel está abierto

logger = logging.getLogger(__name__)

BITACORA_SHEET = "Bitácora"
BITACORA_HEADERS = [
    "Fecha", "Hora", "Tipo", "Actividad", "Cultivo", "Sector",
    "Jornadas Hombre", "Trabajadores", "Insumo", "Cantidad", "Unidad",
    "Registro", "Registrado por",
    "Máquina", "Odómetro", "Horas Día", "Superficie ha", "Días Cubiertos",
]
_WIDTHS = [12, 7, 12, 22, 12, 16, 14, 36, 18, 10, 8, 50, 16, 14, 12, 11, 13, 14]
# Índices (1-based) de las columnas de maquinaria
COL_MAQUINA, COL_ODOMETRO, COL_HORAS, COL_SUPERFICIE = 14, 15, 16, 17


def _open_wb():
    return load_workbook(EXCEL_PATH)


def _aplicar_estilo_header(ws):
    fill = PatternFill("solid", fgColor="5D4037")
    for i, h in enumerate(BITACORA_HEADERS, 1):
        c = ws.cell(1, i, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, w in enumerate(_WIDTHS, 1):
        ws.column_dimensions[chr(64 + i)].width = w


def crear_hoja_bitacora():
    """Crea o migra la hoja Bitácora al esquema ampliado.

    Si existe el esquema viejo (4 cols: Fecha, Hora, Registro, Categoría),
    migra los datos: el texto va a 'Registro', la categoría a 'Actividad'.
    """
    wb = _open_wb()
    if BITACORA_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(BITACORA_SHEET)
        _aplicar_estilo_header(ws)
        _save_wb(wb)
        wb.close()
        logger.info("Hoja Bitácora creada (esquema nuevo)")
        return

    ws = wb[BITACORA_SHEET]
    header3 = str(ws.cell(1, 3).value or "")
    # Si ya está en el esquema con Tipo, solo asegurar columnas de maquinaria
    if header3.strip().lower() == "tipo":
        header14 = str(ws.cell(1, COL_MAQUINA).value or "").strip().lower()
        if header14 != "máquina":
            # Agregar headers de maquinaria (col 14-17)
            from openpyxl.styles import Font, PatternFill, Alignment
            fill = PatternFill("solid", fgColor="5D4037")
            nuevos = {COL_MAQUINA: "Máquina", COL_ODOMETRO: "Odómetro",
                      COL_HORAS: "Horas Día", COL_SUPERFICIE: "Superficie ha"}
            for col, txt in nuevos.items():
                c = ws.cell(1, col, txt)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = fill
                c.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[chr(64 + col)].width = _WIDTHS[col - 1]
            _save_wb(wb)
            logger.info("Bitácora: columnas de maquinaria agregadas")
        wb.close()
        return

    # Migrar esquema viejo → nuevo
    datos_viejos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        datos_viejos.append({
            "fecha": row[0],
            "hora": row[1] if len(row) > 1 else "",
            "texto": row[2] if len(row) > 2 else "",
            "categoria": row[3] if len(row) > 3 else "",
        })

    # Recrear la hoja
    del wb[BITACORA_SHEET]
    ws = wb.create_sheet(BITACORA_SHEET)
    _aplicar_estilo_header(ws)
    for d in datos_viejos:
        ws.append([
            d["fecha"], d["hora"], "OTRO", d["categoria"] or "", "GENERAL", "",
            None, "", "", None, "", d["texto"], "",
        ])
    _save_wb(wb)
    wb.close()
    logger.info(f"Hoja Bitácora migrada al esquema nuevo ({len(datos_viejos)} registros)")


def _ultimo_odometro(ws, maquina: str, con_fecha: bool = False):
    """Último odómetro registrado de una máquina (o None).

    Con `con_fecha=True` devuelve (odómetro, fecha) — la fecha hace falta para
    saber cuántos días pasaron y validar que el salto sea posible.
    """
    maquina = (maquina or "").strip().upper()
    ultimo = None
    ultima_fecha = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < COL_ODOMETRO:
            continue
        m = str(row[COL_MAQUINA - 1] or "").strip().upper()
        if m != maquina:
            continue
        odo = row[COL_ODOMETRO - 1]
        try:
            odo = float(odo) if odo not in (None, "") else None
        except (TypeError, ValueError):
            odo = None
        if odo is None:
            continue
        f = row[0]
        if ultima_fecha is None or (f and str(f) >= str(ultima_fecha)):
            ultima_fecha = f
            ultimo = odo
    return (ultimo, ultima_fecha) if con_fecha else ultimo


# Un horómetro no puede correr más de 24 h por día ni retroceder. Para los
# vehículos se mide en km. Los márgenes son holgados a propósito: la idea es
# atajar disparates, no discutir por unas horas.
MAX_HORAS_DIA = 24
MAX_KM_DIA = 1500
MARGEN_HORAS = 50        # colchón cuando hay pocos días entre lecturas
MARGEN_KM = 2000


def _dias_entre(f1, f2) -> int:
    try:
        a = datetime.strptime(str(f1)[:10], "%Y-%m-%d").date()
        b = datetime.strptime(str(f2)[:10], "%Y-%m-%d").date()
        return abs((b - a).days)
    except (ValueError, TypeError):
        return 0


def validar_odometro(maquina: str, nuevo: float, previo: float | None,
                     dias: int = 0) -> str | None:
    """Devuelve el motivo por el que la lectura es inverosímil, o None.

    Existe porque el bot aceptó "Tractor jhon deere 50853200" (Juan escribió el
    modelo y el horómetro sin espacio) y guardó 50 millones de horas, y también
    aceptó que el 4292 bajara de 5.222 a 3.200.
    """
    if previo is None or nuevo is None:
        return None
    if nuevo < previo:
        return (f"el horómetro bajó: venía en {previo:,.1f} y llega "
                f"{nuevo:,.1f}. Un horómetro no retrocede.")
    es_km = "SSANGYONG" in maquina.upper() or "CAMIONETA" in maquina.upper()
    tope = (dias * MAX_KM_DIA + MARGEN_KM if es_km
            else dias * MAX_HORAS_DIA + MARGEN_HORAS)
    salto = nuevo - previo
    if salto > tope:
        u = "km" if es_km else "h"
        return (f"el salto es imposible: {salto:,.0f} {u} en {dias} días "
                f"(de {previo:,.1f} a {nuevo:,.1f}).")
    return None


def registrar_bitacora_estructurada(campos: dict, registrado_por: str = "",
                                     forzar: bool = False) -> dict:
    """Guarda un registro estructurado en la bitácora.

    Si es MAQUINARIA con odómetro, calcula las horas del día = odómetro
    actual − último odómetro de esa máquina. Devuelve dict con info de horas.

    Una lectura de odómetro inverosímil NO se guarda: vuelve con
    `{"error_odometro": <motivo>}`. Con `forzar=True` se guarda igual.
    """
    crear_hoja_bitacora()
    wb = _open_wb()
    ws = wb[BITACORA_SHEET]
    ahora = datetime.now()
    trabajadores = ", ".join(campos.get("trabajadores") or [])

    maquina = (campos.get("maquina") or "").strip().upper()
    try:
        odometro = float(campos["odometro"]) if campos.get("odometro") not in (None, "") else None
    except (TypeError, ValueError):
        odometro = None
    try:
        superficie = float(campos["superficie_ha"]) if campos.get("superficie_ha") not in (None, "") else None
    except (TypeError, ValueError):
        superficie = None

    # Fecha del TRABAJO: la que menciona el texto (Juan reporta días después);
    # si no viene, la de hoy.
    fecha_trabajo = str(campos.get("fecha") or "").strip()[:10]
    if not fecha_trabajo:
        fecha_trabajo = ahora.strftime("%Y-%m-%d")

    # Horas desde la lectura anterior de la misma máquina.
    # Si entre ambas pasaron varios días, esas horas NO son de un día: se
    # guardan igual y se anota cuántos días cubren, así el consumo queda
    # atribuido al MES de la lectura y nadie las lee como jornada diaria.
    horas_dia = None
    odo_previo = None
    dias_cubiertos = None
    if maquina and odometro is not None:
        odo_previo, fecha_previa = _ultimo_odometro(ws, maquina, con_fecha=True)
        dias = _dias_entre(fecha_previa, fecha_trabajo)
        motivo = None if forzar else validar_odometro(
            maquina, odometro, odo_previo, dias)
        if motivo:
            wb.close()
            logger.warning(f"Odómetro rechazado — {maquina}: {motivo}")
            return {"error_odometro": motivo, "odo_previo": odo_previo,
                    "horas_dia": None, "es_baseline": False}
        if odo_previo is not None and odometro >= odo_previo:
            horas_dia = round(odometro - odo_previo, 1)
            dias_cubiertos = dias if dias and dias > 0 else 1

    ws.append([
        fecha_trabajo,
        ahora.strftime("%H:%M"),
        campos.get("tipo") or "OTRO",
        campos.get("actividad") or "",
        campos.get("cultivo") or "GENERAL",
        campos.get("sector") or "",
        campos.get("jornadas_hombre"),
        trabajadores,
        campos.get("insumo") or "",
        campos.get("cantidad"),
        campos.get("unidad") or "",
        campos.get("texto_original") or campos.get("resumen") or "",
        registrado_por or "",
        maquina,
        odometro,
        horas_dia,
        superficie,
        dias_cubiertos,
    ])
    _save_wb(wb)
    wb.close()
    logger.info(f"Bitácora: {campos.get('actividad')} / {campos.get('cultivo')}"
                + (f" / {maquina} odo={odometro}" if maquina else ""))
    return {"horas_dia": horas_dia, "odo_previo": odo_previo,
            "dias_cubiertos": dias_cubiertos,
            "es_baseline": (maquina and odometro is not None and odo_previo is None)}


def listar_bitacora(dias: int = 7) -> list[dict]:
    """Lista entradas de los últimos N días (esquema nuevo)."""
    crear_hoja_bitacora()
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[BITACORA_SHEET]
    hoy = date.today()
    entradas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        try:
            if isinstance(row[0], str):
                fecha = datetime.strptime(row[0][:10], "%Y-%m-%d").date()
            elif isinstance(row[0], datetime):
                fecha = row[0].date()
            else:
                fecha = row[0]
            if (hoy - fecha).days > dias:
                continue
            entradas.append({
                "fecha": str(fecha), "hora": str(row[1] or ""),
                "tipo": str(row[2] or ""), "actividad": str(row[3] or ""),
                "cultivo": str(row[4] or ""), "sector": str(row[5] or ""),
                "jornadas_hombre": row[6], "trabajadores": str(row[7] or ""),
                "insumo": str(row[8] or ""), "cantidad": row[9],
                "unidad": str(row[10] or ""), "registro": str(row[11] or ""),
                "registrado_por": str(row[12] or "") if len(row) > 12 else "",
            })
        except Exception:
            continue
    wb.close()
    return entradas


def resumen_jornadas(temporada_inicio: date | None = None) -> dict:
    """Agrega jornadas-hombre por actividad y cultivo (para eficiencias)."""
    crear_hoja_bitacora()
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[BITACORA_SHEET]
    from collections import defaultdict
    jh_por_actividad_cultivo = defaultdict(float)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        try:
            jh = float(row[6] or 0)
        except (TypeError, ValueError):
            jh = 0
        if jh <= 0:
            continue
        actividad = str(row[3] or "?")
        cultivo = str(row[4] or "GENERAL")
        jh_por_actividad_cultivo[(actividad, cultivo)] += jh
    wb.close()
    return {f"{a} · {c}": jh for (a, c), jh in
            sorted(jh_por_actividad_cultivo.items(), key=lambda x: -x[1])}


def resumen_maquinaria(dias: int = 90) -> list[dict]:
    """Resumen de horas trabajadas por máquina (últimos N días).

    Suma las 'Horas Día' calculadas. Permite contrastar contra las horas
    facturadas por el proveedor de arriendo.
    """
    crear_hoja_bitacora()
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[BITACORA_SHEET]
    from collections import defaultdict
    hoy = date.today()
    por_maquina = defaultdict(lambda: {"horas": 0.0, "registros": 0,
                                        "superficie": 0.0, "ultimo_odo": None,
                                        "actividades": set()})
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or len(row) < COL_SUPERFICIE:
            continue
        maquina = str(row[COL_MAQUINA - 1] or "").strip().upper()
        if not maquina:
            continue
        # filtro por días
        try:
            if isinstance(row[0], str):
                f = datetime.strptime(row[0][:10], "%Y-%m-%d").date()
            elif isinstance(row[0], datetime):
                f = row[0].date()
            else:
                f = row[0]
            if (hoy - f).days > dias:
                continue
        except Exception:
            pass
        d = por_maquina[maquina]
        d["registros"] += 1
        try:
            h = float(row[COL_HORAS - 1] or 0)
            d["horas"] += h
        except (TypeError, ValueError):
            pass
        try:
            s = float(row[COL_SUPERFICIE - 1] or 0)
            d["superficie"] += s
        except (TypeError, ValueError):
            pass
        try:
            odo = float(row[COL_ODOMETRO - 1] or 0)
            if odo:
                d["ultimo_odo"] = odo
        except (TypeError, ValueError):
            pass
        act = str(row[3] or "").strip()
        if act:
            d["actividades"].add(act)
    wb.close()
    return [{
        "maquina": m, "horas_total": round(d["horas"], 1),
        "registros": d["registros"], "superficie": round(d["superficie"], 2),
        "ultimo_odometro": d["ultimo_odo"],
        "actividades": ", ".join(sorted(d["actividades"])),
    } for m, d in sorted(por_maquina.items(), key=lambda x: -x[1]["horas"])]
