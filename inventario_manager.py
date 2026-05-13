"""
inventario_manager.py — Gestión de Inventario y Uso por Cultivo
Hojas: Inventario, Aplicaciones
"""
import logging
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from config import EXCEL_PATH

logger = logging.getLogger(__name__)

INVENTARIO_SHEET = "Inventario"
APLICACIONES_SHEET = "Aplicaciones"

INVENTARIO_HEADERS = [
    "Producto", "Categoría", "Unidad", "Stock Actual",
    "Stock Mínimo", "Última Entrada", "Último Uso"
]
APLICACIONES_HEADERS = [
    "Fecha", "Producto", "Cantidad", "Unidad", "Cultivo",
    "Sector", "Responsable", "Observaciones"
]

CATEGORIAS = ["Fertilizante", "Fungicida", "Herbicida", "Insecticida", "Semilla", "Otro"]
CULTIVOS = ["Nogales", "Cerezos", "Avellanos"]


def _open_wb():
    return load_workbook(EXCEL_PATH)


def _create_sheet(wb, name, headers, widths, color):
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=11)
        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.fill = fill
            c.font = font
            c.alignment = Alignment(horizontal="center")
            c.border = thin
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = w
        logger.info(f"Hoja '{name}' creada")
    return wb[name]


def _ensure_inventario(wb):
    return _create_sheet(wb, INVENTARIO_SHEET, INVENTARIO_HEADERS,
                         [35, 16, 8, 14, 14, 12, 12], "1B5E20")


def _ensure_aplicaciones(wb):
    return _create_sheet(wb, APLICACIONES_SHEET, APLICACIONES_HEADERS,
                         [12, 30, 10, 8, 12, 15, 18, 35], "33691E")


def crear_hojas_inventario():
    """Crea las hojas Inventario y Aplicaciones."""
    wb = _open_wb()
    _ensure_inventario(wb)
    _ensure_aplicaciones(wb)
    wb.save(EXCEL_PATH)


def _find_producto(ws, producto: str):
    """Busca un producto en inventario. Retorna row_idx o None."""
    prod_lower = producto.strip().lower()
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val and str(val).strip().lower() == prod_lower:
            return row_idx
    return None


def agregar_stock(producto: str, cantidad: float, categoria: str = "Otro",
                  unidad: str = "L") -> dict:
    """Agrega stock a un producto existente o lo crea."""
    wb = _open_wb()
    ws = _ensure_inventario(wb)
    row_idx = _find_producto(ws, producto)
    hoy = date.today().strftime("%Y-%m-%d")

    if row_idx:
        stock_actual = float(ws.cell(row=row_idx, column=4).value or 0)
        nuevo_stock = stock_actual + cantidad
        ws.cell(row=row_idx, column=4).value = nuevo_stock
        ws.cell(row=row_idx, column=6).value = hoy
        wb.save(EXCEL_PATH)
        return {"producto": producto, "stock_anterior": stock_actual,
                "agregado": cantidad, "stock_nuevo": nuevo_stock, "nuevo": False}
    else:
        ws.append([producto, categoria, unidad, cantidad, 0, hoy, ""])
        wb.save(EXCEL_PATH)
        return {"producto": producto, "stock_anterior": 0,
                "agregado": cantidad, "stock_nuevo": cantidad, "nuevo": True}


def registrar_uso(producto: str, cantidad: float, cultivo: str,
                  sector: str = "", responsable: str = "",
                  observaciones: str = "") -> dict:
    """Registra uso de un producto y descuenta del inventario."""
    wb = _open_wb()
    ws_inv = _ensure_inventario(wb)
    ws_app = _ensure_aplicaciones(wb)

    row_idx = _find_producto(ws_inv, producto)
    hoy = date.today().strftime("%Y-%m-%d")

    if row_idx:
        stock_actual = float(ws_inv.cell(row=row_idx, column=4).value or 0)
        unidad = ws_inv.cell(row=row_idx, column=3).value or "L"
        nuevo_stock = stock_actual - cantidad
        ws_inv.cell(row=row_idx, column=4).value = nuevo_stock
        ws_inv.cell(row=row_idx, column=7).value = hoy
    else:
        unidad = "L"
        nuevo_stock = -cantidad
        ws_inv.append([producto, "Otro", unidad, nuevo_stock, 0, "", hoy])

    ws_app.append([
        hoy, producto, cantidad, unidad, cultivo,
        sector, responsable, observaciones
    ])

    wb.save(EXCEL_PATH)
    alerta = nuevo_stock <= float(ws_inv.cell(row=row_idx, column=5).value or 0) if row_idx else True
    return {
        "producto": producto, "cantidad": cantidad, "unidad": unidad,
        "cultivo": cultivo, "sector": sector,
        "stock_restante": nuevo_stock, "alerta_bajo": alerta
    }


def consultar_inventario() -> list[dict]:
    """Lista todo el inventario actual."""
    wb = _open_wb()
    ws = _ensure_inventario(wb)
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        stock = float(row[3] or 0)
        minimo = float(row[4] or 0)
        items.append({
            "producto": str(row[0]),
            "categoria": str(row[1] or ""),
            "unidad": str(row[2] or "L"),
            "stock": stock,
            "minimo": minimo,
            "alerta": stock <= minimo and minimo > 0,
        })
    wb.close()
    return items


def productos_bajo_stock() -> list[dict]:
    """Retorna productos con stock bajo el mínimo."""
    return [p for p in consultar_inventario() if p["alerta"]]


def consumo_por_cultivo(dias: int = 7) -> dict:
    """Resumen de consumo por cultivo en los últimos N días."""
    wb = _open_wb()
    ws = _ensure_aplicaciones(wb)
    hoy = date.today()
    resumen = {c: [] for c in CULTIVOS}
    resumen["Otro"] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        try:
            if isinstance(row[0], str):
                fecha = datetime.strptime(row[0][:10], "%Y-%m-%d").date()
            else:
                fecha = row[0] if isinstance(row[0], date) else row[0].date()
            if (hoy - fecha).days <= dias:
                cultivo = str(row[4] or "Otro")
                key = cultivo if cultivo in CULTIVOS else "Otro"
                resumen[key].append({
                    "producto": str(row[1] or ""),
                    "cantidad": float(row[2] or 0),
                    "unidad": str(row[3] or ""),
                    "fecha": str(fecha),
                })
        except Exception:
            continue
    wb.close()
    return resumen


def agregar_stock_desde_factura(items: list):
    """Agrega stock automáticamente desde items de factura procesada.
    Busca productos que parezcan insumos agrícolas."""
    keywords = {
        "fertilizante": "Fertilizante", "urea": "Fertilizante", "npk": "Fertilizante",
        "nitrato": "Fertilizante", "fosfato": "Fertilizante", "potasio": "Fertilizante",
        "abono": "Fertilizante", "compost": "Fertilizante",
        "fungicida": "Fungicida", "captan": "Fungicida", "cobre": "Fungicida",
        "azufre": "Fungicida", "mancozeb": "Fungicida",
        "herbicida": "Herbicida", "glifosato": "Herbicida", "roundup": "Herbicida",
        "insecticida": "Insecticida", "aceite": "Insecticida",
        "intrepid": "Insecticida", "clorpirifos": "Insecticida",
    }
    agregados = []
    for item in items:
        glosa = str(item.get("Detalle / Glosa") or "").lower()
        glosa2 = str(item.get("Glosa II") or "").lower()
        texto = f"{glosa} {glosa2}"

        categoria = None
        for kw, cat in keywords.items():
            if kw in texto:
                categoria = cat
                break

        if categoria:
            nombre = item.get("Detalle / Glosa") or item.get("Glosa II") or "Insumo"
            qty = float(item.get("Cantidad") or 1)
            unidad = "L" if any(u in texto for u in ("litro", "lt", "lts")) else "Kg" if any(
                u in texto for u in ("kilo", "kg")) else "Un"
            result = agregar_stock(nombre, qty, categoria, unidad)
            agregados.append(result)

    return agregados
