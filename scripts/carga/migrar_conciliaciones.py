#!/usr/bin/env python3
"""Migra los links de texto de la col J (Cuenta Banco) a la hoja Conciliaciones."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import re
from collections import defaultdict
from openpyxl import load_workbook
from config import EXCEL_PATH
from modules.conciliacion_store import registrar_vinculos, SHEET
from modules.correlativo import nrokey

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)

# Filas banco ya migradas (por si se corre dos veces)
ya = set()
if SHEET in wb.sheetnames:
    for row in wb[SHEET].iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            try:
                ya.add(int(row[2]))
            except (TypeError, ValueError):
                pass

# Links de texto en col J
ws_b = wb["Cuenta Banco"]
links = []
for i, row in enumerate(ws_b.iter_rows(min_row=2, values_only=True), 2):
    if not row or not row[0]:
        continue
    j = str(row[9] or "").strip() if len(row) > 9 else ""
    if j and i not in ya:
        links.append((i, j))

# Índice de facturas por nº
ws_f = wb["Facturas"]
fact = defaultdict(list)
for i, row in enumerate(ws_f.iter_rows(min_row=2, values_only=True), 2):
    if not row or not row[0]:
        continue
    nro = nrokey(row[6])
    if nro:
        fact[nro].append({"fila": i, "prov": str(row[3] or "")})
wb.close()

print(f"Links en col J pendientes de migrar: {len(links)}")
vinculos = []
for fila_banco, texto in links:
    m = re.match(r"F?([A-Z0-9]+)\s*(.*)", texto.strip().upper())
    if not m:
        print(f"  ⚠️ fila {fila_banco}: '{texto[:40]}' no parseable — se omite")
        continue
    nro, prov = m.group(1), m.group(2).strip()
    candidatos = fact.get(nrokey(nro), [])
    filas_doc = [f["fila"] for f in candidatos]
    if not prov and candidatos:
        prov = candidatos[0]["prov"]
    vinculos.append({
        "fila_banco": fila_banco, "tipo_doc": "FACTURA",
        "fila_doc": filas_doc[0] if filas_doc else None,
        "filas_doc": filas_doc, "nro_doc": nro, "proveedor": prov,
        "monto_asignado": None, "criterio": "migrado-colJ", "nota": texto[:40],
    })
    print(f"  fila {fila_banco}: F{nro} {prov[:30]} ({len(filas_doc)} líneas factura)")

if vinculos:
    r = registrar_vinculos(vinculos, usuario="migracion")
    print(f"\n✅ Migrados: {r['registrados']}")
else:
    print("\nNada que migrar.")

from modules.conciliacion_store import resumen_estados
est = resumen_estados()
print(f"\nEstados registrados: {len(est)}")
for fb, e in sorted(est.items()):
    print(f"  fila {fb}: ${e['monto']:,.0f} asignado ${e['asignado']:,.0f} "
          f"saldo ${e['saldo']:,.0f} → {e['estado']}")
