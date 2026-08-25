#!/usr/bin/env python3
"""1) Rellena el N° de archivo (correlativo de FXP) en Master.Facturas.
   2) Genera el listado de diferencias Master vs FXP con fechas.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import os
from collections import defaultdict
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import EXCEL_PATH
from excel_manager import _save_wb
from modules.correlativo import (
    leer_correlativos_fxp, norm_prov, nrokey,
    COL_CORRELATIVO, HEADER_CORRELATIVO, asegurar_columna, FXP_PATH,
)

SALIDA = r"C:\Users\Windows\Desktop\Workflow\Agricola Santa Elisa\Diferencias_Master_vs_FXP.xlsx"


def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for f in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try: return datetime.strptime(v[:10], f).date()
            except Exception: pass
    return None


def _f(v):
    try: return float(v or 0)
    except (TypeError, ValueError): return 0.0


# ── FXP completo (para montos, fechas y correlativo) ──
import shutil, tempfile
tmp = os.path.join(tempfile.gettempdir(), "fxp_full.xlsx")
shutil.copy2(FXP_PATH, tmp)
wbf = load_workbook(tmp, read_only=True, data_only=True)
wsf = wbf["FXP"]
fxp = {}
for row in wsf.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    nro = nrokey(row[7])
    if not nro:
        continue
    m = row[8]
    if isinstance(m, str):
        s = m.upper().replace("USD", "").replace("$", "").strip().replace(".", "").replace(",", ".")
        try: m = float(s)
        except ValueError: m = 0
    d = {"n": int(row[3]) if isinstance(row[3], (int, float)) else None,
         "prov": str(row[6] or ""), "nro": str(row[7] or ""), "monto": _f(m),
         "emision": _pd(row[0]), "venc": _pd(row[1]), "pago": _pd(row[2]),
         "estado": str(row[11] or "").strip(), "nota": str(row[12] or "")}
    fxp[(norm_prov(row[6]), nro)] = d
wbf.close()
fxp_por_nro = {}
for (p, n), d in fxp.items():
    fxp_por_nro.setdefault(n, d)

por_prov_nro, por_nro, max_fxp = leer_correlativos_fxp()
print(f"FXP: {len(fxp)} facturas, correlativo máximo {max_fxp}\n")

# ── Master ──
wb = load_workbook(EXCEL_PATH)
ws = wb["Facturas"]
asegurar_columna(ws)

grupos = defaultdict(list)
for r in range(2, ws.max_row + 1):
    if not ws.cell(r, 1).value:
        continue
    nro = nrokey(ws.cell(r, 7).value)
    if not nro:
        continue
    grupos[(norm_prov(ws.cell(r, 4).value), nro)].append(r)

# ── 1) Asignar correlativo ──
desde_fxp, nuevos = 0, 0
siguiente = max_fxp
sin_match = []
for (prov, nro), filas in sorted(grupos.items(),
                                  key=lambda x: min(x[1])):
    n = por_prov_nro.get((prov, nro)) or por_nro.get(nro)
    if n:
        desde_fxp += 1
    else:
        siguiente += 1
        n = siguiente
        nuevos += 1
        sin_match.append((ws.cell(filas[0], 1).value, ws.cell(filas[0], 4).value,
                           ws.cell(filas[0], 7).value, _f(ws.cell(filas[0], 16).value), n))
    for r in filas:
        ws.cell(r, COL_CORRELATIVO).value = n

_save_wb(wb)
print(f"✅ Correlativo asignado a {len(grupos)} facturas del Master")
print(f"   · {desde_fxp} tomaron su número desde FXP")
print(f"   · {nuevos} eran nuevas → números {max_fxp+1} a {siguiente}\n")

# ── 2) Listado de diferencias ──
difs, solo_master, solo_fxp = [], [], []
for (prov, nro), filas in grupos.items():
    r0 = filas[0]
    total = _f(ws.cell(r0, 16).value)
    corr = ws.cell(r0, COL_CORRELATIVO).value
    d = fxp.get((prov, nro)) or fxp_por_nro.get(nro)
    reg = {"n": corr, "prov": ws.cell(r0, 4).value, "nro": ws.cell(r0, 7).value,
           "emision_m": _pd(ws.cell(r0, 1).value), "venc_m": _pd(ws.cell(r0, 2).value),
           "pago_m": _pd(ws.cell(r0, 3).value), "total_m": total,
           "lineas": len(filas), "fila": r0}
    if not d:
        solo_master.append(reg)
        continue
    if total > 0 and d["monto"] > 0 and abs(total - d["monto"]) > max(1000, d["monto"] * 0.01):
        reg.update({"monto_f": d["monto"], "emision_f": d["emision"],
                    "pago_f": d["pago"], "estado_f": d["estado"], "nota_f": d["nota"],
                    "dif": total - d["monto"]})
        difs.append(reg)
for (p, n), d in fxp.items():
    if (p, n) not in grupos and n not in {k[1] for k in grupos}:
        solo_fxp.append(d)

wb.close()

# ── Escribir el archivo de salida ──
from openpyxl import Workbook
out = Workbook()
AZUL = PatternFill("solid", fgColor="1F4E78")


def hoja(wsx, headers, anchos):
    for i, h in enumerate(headers, 1):
        c = wsx.cell(1, i, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = AZUL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, a in enumerate(anchos, 1):
        wsx.column_dimensions[chr(64 + i)].width = a
    wsx.freeze_panes = "A2"


w1 = out.active
w1.title = "Diferencias de monto"
hoja(w1, ["N° Archivo", "Proveedor", "N° Documento", "Emisión Master", "Emisión FXP",
          "Fecha Pago Master", "Fecha Pago FXP", "Monto Master", "Monto FXP",
          "Diferencia", "Estado FXP", "Nota FXP", "Líneas", "Fila Master"],
     [11, 34, 16, 14, 13, 15, 14, 15, 15, 15, 12, 34, 8, 11])
for d in sorted(difs, key=lambda x: -abs(x["dif"])):
    w1.append([d["n"], d["prov"], d["nro"], d["emision_m"], d["emision_f"],
               d["pago_m"], d["pago_f"], d["total_m"], d["monto_f"], d["dif"],
               d["estado_f"], d["nota_f"], d["lineas"], d["fila"]])
for r in range(2, w1.max_row + 1):
    for c in (8, 9, 10):
        w1.cell(r, c).number_format = '#,##0'

w2 = out.create_sheet("Solo en Master")
hoja(w2, ["N° Archivo", "Proveedor", "N° Documento", "Emisión", "Vencimiento",
          "Fecha Pago", "Monto", "Líneas", "Fila"],
     [11, 34, 16, 13, 13, 13, 15, 8, 9])
for d in sorted(solo_master, key=lambda x: -x["total_m"]):
    w2.append([d["n"], d["prov"], d["nro"], d["emision_m"], d["venc_m"],
               d["pago_m"], d["total_m"], d["lineas"], d["fila"]])
for r in range(2, w2.max_row + 1):
    w2.cell(r, 7).number_format = '#,##0'

w3 = out.create_sheet("Solo en FXP")
hoja(w3, ["N° FXP", "Proveedor", "N° Documento", "Emisión", "Vencimiento",
          "Fecha Pago", "Monto", "Estado", "Nota"],
     [10, 34, 16, 13, 13, 13, 15, 12, 40])
for d in sorted(solo_fxp, key=lambda x: -x["monto"]):
    w3.append([d["n"], d["prov"], d["nro"], d["emision"], d["venc"], d["pago"],
               d["monto"], d["estado"], d["nota"]])
for r in range(2, w3.max_row + 1):
    w3.cell(r, 7).number_format = '#,##0'

w4 = out.create_sheet("Nuevas sin N° en FXP")
hoja(w4, ["N° Archivo asignado", "Emisión", "Proveedor", "N° Documento", "Monto"],
     [18, 13, 34, 16, 15])
for emi, prov, nro, tot, n in sorted(sin_match, key=lambda x: (x[4])):
    w4.append([n, _pd(emi), prov, nro, tot])
for r in range(2, w4.max_row + 1):
    w4.cell(r, 5).number_format = '#,##0'

out.save(SALIDA)
print(f"📄 Listado guardado: {SALIDA}")
print(f"   · Diferencias de monto: {len(difs)}")
print(f"   · Solo en Master: {len(solo_master)}")
print(f"   · Solo en FXP: {len(solo_fxp)}")
print(f"   · Nuevas con número asignado: {len(sin_match)}")
