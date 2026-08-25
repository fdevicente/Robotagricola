#!/usr/bin/env python3
"""Reconstruye las vacaciones desde la planilla oficial del usuario (ago-2026).

Deja UNA sola fuente de verdad:
  hoja `Vacaciones`  → todos los días tomados (2025 + 2026)
  hoja `Personal`    → saldo recalculado: base + acumulación − tomados
La acumulación es 1,25 días/mes (15 al año) desde la fecha del saldo base.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import EXCEL_PATH
from excel_manager import _save_wb

HOY = date(2026, 8, 4)
DIAS_MES = 15.0 / 12.0

# Nombres canónicos (los de la hoja Personal)
CANON = {
    "MORA PATRICIO":   "Luis Patricio Mora Amigo",
    "PARADA JUAN":     "Juan Parada Castillo",
    "AMIGO RAMIRO":    "Luis Ramiro Amigo Soto",
    "AMIGO FELICITO":  "Felicito Amigo Soto",
    "MORA AGUSTIN":    "Agustin Segundo Mora Hernandez",
    "GONZALES JAVIER": "Javier Gonzalez",
    "GONZALEZ JAVIER": "Javier Gonzalez",
}

# Planilla del usuario (foto): (nombre, desde, hasta, días, nota)
# OJO: la fila que dice "2027 AGOSTO" para Mora Agustín tiene fechas 17-07-2026;
# es un error de tipeo del año/mes en la planilla — se toma por la fecha real.
VACACIONES_2026 = [
    ("MORA PATRICIO",   "2026-01-19", "2026-01-30", 10, ""),
    ("PARADA JUAN",     "2026-01-23", "2026-01-24", 1,  ""),
    ("PARADA JUAN",     "2026-02-16", "2026-02-20", 5,  ""),
    ("MORA AGUSTIN",    "2026-03-16", "2026-03-20", 5,  ""),
    ("GONZALES JAVIER", "2026-05-18", "2026-05-19", 1,  ""),
    ("AMIGO RAMIRO",    "2026-05-19", "2026-05-22", 3,  ""),
    ("AMIGO FELICITO",  "2026-05-22", "2026-05-23", 1,  "Sándwich"),
    ("MORA AGUSTIN",    "2026-05-22", "2026-05-23", 1,  "Sándwich"),
    ("MORA PATRICIO",   "2026-05-22", "2026-05-23", 1,  "Sándwich"),
    ("PARADA JUAN",     "2026-05-22", "2026-05-23", 1,  "Sándwich"),
    ("GONZALES JAVIER", "2026-05-22", "2026-05-23", 1,  "Sándwich"),
    ("GONZALES JAVIER", "2026-06-01", "2026-06-05", 5,  ""),
    ("GONZALES JAVIER", "2026-06-19", "2026-06-20", 1,  ""),
    ("PARADA JUAN",     "2026-06-30", "2026-07-03", 4,  ""),
    ("AMIGO FELICITO",  "2026-07-17", "2026-07-18", 1,  "Sándwich"),
    ("AMIGO RAMIRO",    "2026-07-17", "2026-07-18", 1,  "Sándwich"),
    ("MORA PATRICIO",   "2026-07-17", "2026-07-18", 1,  "Sándwich"),
    ("GONZALES JAVIER", "2026-07-17", "2026-07-18", 1,  "Sándwich"),
    ("MORA AGUSTIN",    "2026-07-17", "2026-07-18", 1,  "Sándwich"),
    ("PARADA JUAN",     "2026-07-17", "2026-07-18", 1,  "Sándwich"),
    ("AMIGO RAMIRO",    "2026-08-05", "2026-08-06", 1,  ""),
]
# Totales declarados en la planilla, para validar
TOTALES_PLANILLA = {"MORA PATRICIO": 12, "PARADA JUAN": 12, "AMIGO RAMIRO": 5,
                    "AMIGO FELICITO": 2, "MORA AGUSTIN": 7, "GONZALES JAVIER": 9}

# Vacaciones anteriores a 2026 que ya estaban registradas (se conservan)
VACACIONES_PREVIAS = [
    ("MORA AGUSTIN",  "2025-03-03", "2025-03-14", 10, "Periodo 2025"),
    ("MORA PATRICIO", "2025-02-10", "2025-02-14", 5,  "Periodo 2025"),
]


def _d(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for f in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try: return datetime.strptime(v[:10], f).date()
            except ValueError: continue
    return None


# ── Validar contra los totales de la planilla ──
print("Validación contra los totales de tu planilla:")
suma = {}
for n, *_r in VACACIONES_2026:
    suma[n] = suma.get(n, 0) + _r[2]
ok = True
for n, t in TOTALES_PLANILLA.items():
    calc = suma.get(n, 0)
    marca = "✅" if calc == t else "❌"
    if calc != t: ok = False
    print(f"  {marca} {n:18} planilla={t:>3}  cargado={calc:>3}")
if not ok:
    print("\n⚠️ No calzan los totales — reviso antes de escribir.")
    sys.exit(1)

wb = load_workbook(EXCEL_PATH)

# ── Reescribir hoja Vacaciones ──
ws = wb["Vacaciones"]
ws.delete_rows(2, max(ws.max_row, 2))
todas = [(CANON[n], d, h, dd, nt or f"Periodo {d[:4]}")
         for n, d, h, dd, nt in VACACIONES_PREVIAS + VACACIONES_2026]
todas.sort(key=lambda x: x[1])
for nombre, desde, hasta, dias, nota in todas:
    ws.append([nombre, _d(desde), _d(hasta), dias, "Aprobado", nota])
print(f"\n✅ Hoja Vacaciones: {len(todas)} registros "
      f"({len(VACACIONES_PREVIAS)} de 2025 + {len(VACACIONES_2026)} de 2026)")

# ── Saldos base (hoja Vacaciones Pendientes) ──
base = {}
if "Vacaciones Pendientes" in wb.sheetnames:
    wsp = wb["Vacaciones Pendientes"]
    for row in wsp.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        nom = str(row[0]).strip()
        try: saldo = float(row[3] or 0)
        except (TypeError, ValueError): saldo = 0.0
        base[nom] = {"saldo": saldo, "fecha": _d(row[4]), "contrato": _d(row[2])}

# ── Recalcular hoja Personal ──
wsp = wb["Personal"]
tomados = {}
for nombre, desde, hasta, dias, nota in todas:
    tomados[nombre] = tomados.get(nombre, 0) + dias

print("\n=== Saldos recalculados ===")
print(f"{'Trabajador':32} {'base':>7} {'+acum':>8} {'−tom':>6} {'=saldo':>8}")
print("-" * 68)
existentes = set()
for r in range(2, wsp.max_row + 1):
    nombre = str(wsp.cell(r, 1).value or "").strip()
    if not nombre:
        continue
    existentes.add(nombre)
    b = base.get(nombre, {})
    saldo_base = b.get("saldo", 0.0)
    f_saldo = b.get("fecha") or _d(wsp.cell(r, 4).value) or HOY
    meses = (HOY.year - f_saldo.year) * 12 + (HOY.month - f_saldo.month)
    acum = meses * DIAS_MES
    tom = tomados.get(nombre, 0)
    total = round(saldo_base + acum - tom, 2)
    wsp.cell(r, 5).value = total
    wsp.cell(r, 6).value = tom
    ult = max((v[2] for v in todas if v[0] == nombre), default=None)
    if ult:
        wsp.cell(r, 7).value = _d(ult)
    wsp.cell(r, 8).value = (f"Saldo base {saldo_base:g} al {f_saldo} "
                             f"+ {meses} meses × 1,25 − {tom} tomados")
    print(f"{nombre[:32]:32} {saldo_base:>7.2f} {acum:>8.2f} {tom:>6} {total:>8.2f}")

# ── Agregar quien falte (Javier) ──
faltan = {v[0] for v in todas} - existentes
for nombre in sorted(faltan):
    b = base.get(nombre, {})
    tom = tomados.get(nombre, 0)
    # Sin saldo base conocido: se acumula desde su fecha de ingreso si la hay
    f_ini = b.get("fecha") or date(2026, 1, 1)
    meses = (HOY.year - f_ini.year) * 12 + (HOY.month - f_ini.month)
    acum = meses * DIAS_MES
    total = round(b.get("saldo", 0.0) + acum - tom, 2)
    fila = wsp.max_row + 1
    wsp.cell(fila, 1).value = nombre
    wsp.cell(fila, 5).value = total
    wsp.cell(fila, 6).value = tom
    ult = max((v[2] for v in todas if v[0] == nombre), default=None)
    if ult:
        wsp.cell(fila, 7).value = _d(ult)
    wsp.cell(fila, 8).value = ("⚠️ FALTA fecha de ingreso y saldo base — "
                                f"acumulado provisorio desde {f_ini}")
    print(f"{nombre[:32]:32} {'?':>7} {acum:>8.2f} {tom:>6} {total:>8.2f}  ← NUEVO")

_save_wb(wb)
wb.close()
print("\n✅ Hoja Personal actualizada.")
