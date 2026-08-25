"""Borra la fila duplicada de Copeval 6432660 (la que se subió dos veces).

Es la ÚLTIMA fila de la hoja Facturas y ninguna conciliación la referencia, así
que borrarla no corre ningún número de fila guardado en otra parte.
Dejarla marcada no servía: al compartir número con la original, el grupo entero
se leía como pagado y la factura desaparecía de los pendientes.
"""
import shutil
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb

FILA = 2165
ESPERADO = ("COPEVAL", "6432660", 347_733)

resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"Respaldo: {resp}\n")

wb = load_workbook(EXCEL_PATH)
ws = wb["Facturas"]

if ws.max_row != FILA:
    print(f"❌ La fila {FILA} ya no es la última (max_row={ws.max_row}). No se toca.")
    wb.close()
    sys.exit(1)

prov = str(ws.cell(FILA, 4).value or "").upper()
nro = str(ws.cell(FILA, 7).value or "").strip()
try:
    total = float(ws.cell(FILA, 16).value or 0)
except (TypeError, ValueError):
    total = 0

if ESPERADO[0] not in prov or nro != ESPERADO[1] or abs(total - ESPERADO[2]) > 1:
    print(f"❌ La fila {FILA} no es la esperada: {prov[:20]} N°{nro} ${total:,.0f}")
    wb.close()
    sys.exit(1)

print(f"Borrando fila {FILA}: {prov[:20]} N°{nro} ${total:,.0f}")
ws.delete_rows(FILA)
_save_wb(wb)
wb.close()
print(f"✅ Borrada. Facturas queda con {FILA - 1} filas.")
