# -*- coding: utf-8 -*-
"""Borra del Master las 7 facturas que quedaron cargadas DOS VECES.

Se corre A MANO. Sin --aplicar solo simula.

Detectadas el 2-sep-2026 por una señal que no admite discusion: la suma de los
"Total por Item" del grupo da EXACTAMENTE el doble del "TOTAL FACTURA" (ratio
2,00 en las 7). No son dos lineas iguales de una misma factura: es el documento
entero cargado de nuevo.

Cinco son doble lectura del mismo documento. Dos (Ferreteria Talca e Irrifor)
son otra cosa: el documento esta una vez como UNA linea con el total y otra vez
desglosado por producto — ahi se conserva el desglose.

⚠️ LO DELICADO NO ES BORRAR, ES LO QUE ARRASTRA. `Conciliaciones` referencia a
`Facturas` POR NUMERO DE FILA. Al borrar, todo lo de abajo sube uno y esos
vinculos quedan apuntando a otra factura, en silencio. Por eso se recalculan
las referencias en la misma operacion, con `modules.filas.ajustar_referencia`.
"""
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(
    os.path.dirname(os.path.abspath(__file__)))))

from config import EXCEL_PATH
from modules.filas import ajustar_referencia

# fila -> por que se borra. Se conserva SIEMPRE la version pagada/conciliada,
# o el desglose por producto cuando la otra copia es una sola linea.
BORRAR = {
    501:  "27987 Irrifor — copia en UNA linea; queda el desglose 499-500",
    954:  "302052 Ferreteria Talca — copia en UNA linea; queda el desglose 962-968",
    2117: "107228 Llaneza — copia sin valor unitario; queda la 2148",
    2139: "23683864 Admin. Ventas — identica a la 2091",
    2152: "607 Albino Fuentealba — identica a la 2151",
    2181: "78322 Efrain Morales — 2a lectura, sin Fecha Pago; queda 2156-2159",
    2182: "78322 Efrain Morales — 2a lectura, sin Fecha Pago",
    2183: "78322 Efrain Morales — 2a lectura, sin Fecha Pago",
    2184: "78322 Efrain Morales — 2a lectura, sin Fecha Pago",
    2192: "6950 Ferreteria M y G — identica a la 2171, que esta conciliada",
}

COL_FILA_DOC = 8          # 'Fila Doc' en Conciliaciones


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--aplicar", action="store_true")
    args = ap.parse_args()

    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_PATH)
    try:
        ws = wb["Facturas"]
        conc = wb["Conciliaciones"]

        print("=== FILAS A BORRAR (%d) ===" % len(BORRAR))
        for f in sorted(BORRAR):
            print("  %-6s Nº%-10s %-30s  $%-10s %s"
                  % (f, ws.cell(f, 7).value, str(ws.cell(f, 4).value or "")[:30],
                     ws.cell(f, 15).value,
                     "SIN PAGO" if not ws.cell(f, 3).value else ""))
            print("         %s" % BORRAR[f])

        print("\n=== REFERENCIAS DE Conciliaciones ===")
        cambios = []
        for r in range(2, conc.max_row + 1):
            v = conc.cell(r, COL_FILA_DOC).value
            if v is None:
                continue
            viejo = int(v)
            nuevo = ajustar_referencia(viejo, BORRAR)   # lanza si apunta a una borrada
            if nuevo != viejo:
                cambios.append((r, viejo, nuevo, conc.cell(r, 9).value))
        for r, viejo, nuevo, nro in cambios:
            print("  vinculo fila %-4s Nº%-10s  Facturas %s -> %s"
                  % (r, nro, viejo, nuevo))
        print("  (%d referencias se corren, %d quedan igual)"
              % (len(cambios), conc.max_row - 1 - len(cambios)))

        if not args.aplicar:
            print("\n(simulacion: no se escribio nada)")
            return

        # De abajo hacia arriba, para que los numeros de arriba sigan validos
        for f in sorted(BORRAR, reverse=True):
            ws.delete_rows(f, 1)
        for r, _, nuevo, _ in cambios:
            conc.cell(r, COL_FILA_DOC).value = nuevo

        wb.save(EXCEL_PATH)          # ruta EXPLICITA
        print("\nBorradas %d filas y corregidas %d referencias."
              % (len(BORRAR), len(cambios)))
    finally:
        wb.close()


if __name__ == "__main__":
    main()
