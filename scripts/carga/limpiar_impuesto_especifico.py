# -*- coding: utf-8 -*-
"""Borra del Master los impuestos especificos que la aritmetica desmiente.

Se corre A MANO. Sin --aplicar solo lista.

Usa EXACTAMENTE la misma regla que `sanear_impuesto_especifico` del extractor,
que desde el 1-sep evita que se sigan generando: si neto x IVA ya da el total,
el impuesto sobra. Lo que se previene y lo que se limpia son la misma regla.
"""
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(
    os.path.dirname(os.path.abspath(__file__)))))

from config import EXCEL_PATH
from processors.extractor import TOLERANCIA_CUADRE

COL_DOC, COL_UNIT, COL_CANT, COL_IMP, COL_TOTAL = 6, 10, 11, 14, 16


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--aplicar", action="store_true")
    args = ap.parse_args()

    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_PATH)
    try:
        ws = wb["Facturas"]
        tocadas = []
        for f in range(2, ws.max_row + 1):
            try:
                imp = float(ws.cell(f, COL_IMP).value or 0)
                if imp <= 0:
                    continue
                neto = (float(ws.cell(f, COL_UNIT).value or 0)
                        * float(ws.cell(f, COL_CANT).value or 1))
                total = float(ws.cell(f, COL_TOTAL).value or 0)
            except (TypeError, ValueError):
                continue
            if total <= 0:
                continue
            doc = str(ws.cell(f, COL_DOC).value or "").lower()
            sin_iva = any(k in doc for k in ("exenta", "exento", "no afecta",
                                             "no afecto", "boleta de honorario"))
            iva = 1.0 if sin_iva else 1.19
            sin_imp = abs(neto * iva - total)
            con_imp = abs(neto * iva + imp - total)
            if sin_imp <= TOLERANCIA_CUADRE and sin_imp < con_imp:
                tocadas.append((f, ws.cell(f, 4).value, ws.cell(f, 7).value,
                                neto, imp, total))
                if args.aplicar:
                    ws.cell(f, COL_IMP).value = 0

        print("%-7s %-30s %-10s %-11s %-10s %s"
              % ("fila", "proveedor", "nro", "neto", "imp.esp", "TOTAL"))
        print("-" * 84)
        for f, prov, nro, neto, imp, total in tocadas:
            print("%-7s %-30s %-10s %-11s %-10s %s"
                  % (f, str(prov)[:30], nro, int(neto), int(imp), int(total)))
        print("\n%d fila(s), $%s de impuesto inventado"
              % (len(tocadas), format(int(sum(t[4] for t in tocadas)), ",")))

        if not args.aplicar:
            print("(simulacion: no se escribio nada)")
            return
        if tocadas:
            wb.save(EXCEL_PATH)          # ruta EXPLICITA
            print("Guardado.")
    finally:
        wb.close()


if __name__ == "__main__":
    main()
