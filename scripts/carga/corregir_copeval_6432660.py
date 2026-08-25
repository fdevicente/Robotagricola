"""Copeval $347.733 del 10-06-2026: se deja con el número de FXP (6432660).

El bot la leyó como 6432360 y además quedó duplicada (se subió dos veces).
El dueño confirmó que el número bueno es el de FXP.
  · fila 2122 → pasa a 6432660 y queda como la factura pendiente
  · fila 2165 → es el duplicado, se cierra y se marca
"""
import shutil
import sys
from datetime import date, datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb

NRO_MALO, NRO_BUENO = "6432360", "6432660"
MONTO = 347_733
COL_PROV, COL_NRO, COL_TOTAL = 4, 7, 16
COL_PAGO, COL_NOTA = 3, 20

resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"Respaldo: {resp}\n")

wb = load_workbook(EXCEL_PATH)
ws = wb["Facturas"]

filas = []
for r in range(2, ws.max_row + 1):
    if "COPEVAL" not in str(ws.cell(r, COL_PROV).value or "").upper():
        continue
    nro = str(ws.cell(r, COL_NRO).value or "").strip()
    if nro.endswith(".0"):
        nro = nro[:-2]
    if nro != NRO_MALO:
        continue
    try:
        total = float(ws.cell(r, COL_TOTAL).value or 0)
    except (TypeError, ValueError):
        total = 0
    if abs(total - MONTO) > 1:
        continue
    filas.append(r)

if not filas:
    print(f"No encontré filas de Copeval {NRO_MALO} por ${MONTO:,}")
    wb.close()
    sys.exit(0)

print(f"Filas encontradas: {filas}\n")
principal, duplicados = filas[0], filas[1:]

ws.cell(principal, COL_NRO).value = NRO_BUENO
print(f"  fila {principal}: N° {NRO_MALO} → {NRO_BUENO}   (queda PENDIENTE)")

for r in duplicados:
    ws.cell(r, COL_NRO).value = NRO_BUENO
    ws.cell(r, COL_PAGO).value = date.today()
    previo = str(ws.cell(r, COL_NOTA).value or "")
    ws.cell(r, COL_NOTA).value = (
        (previo + " · " if previo else "")
        + f"DUPLICADA de la fila {principal} — se subió dos veces, no es deuda aparte")
    print(f"  fila {r}: marcada como duplicada de la {principal} y cerrada")

_save_wb(wb)
wb.close()
print("\n✅ Listo.")
