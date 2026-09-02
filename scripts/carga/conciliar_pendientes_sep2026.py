# -*- coding: utf-8 -*-
"""Marca pagadas las facturas que el banco muestra pagadas y el Master no.

Se corre A MANO. Sin --aplicar solo muestra lo que haria.

Contexto (2-sep-2026): tras importar la cartola al 2-sep, el conciliador
encontro 9 facturas sin Fecha Pago cuyo cargo esta en el banco. El dueño pidio
dejarlas pagadas usando FXP como referencia ("el FXP esta al dia y correcto").

La fecha de pago sale de FXP cuando FXP la trae; si no, del cargo del banco.
Se usa `registrar_vinculos` y no se escribe la Fecha Pago a mano: asi queda el
vinculo banco<->factura, el estado del documento y la columna J del banco.
"""
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(
    os.path.dirname(os.path.abspath(__file__)))))

from config import FXP_PATH
from modules.conciliador import _nrokey, analizar
from modules.drive.enlaces import _clave_proveedor


def _fecha_fxp():
    """(nro, proveedor) -> fecha de pago que trae FXP, si la trae."""
    from openpyxl import load_workbook
    wb = load_workbook(FXP_PATH, read_only=True, data_only=True)
    try:
        out = {}
        for row in wb["FXP"].iter_rows(min_row=2, values_only=True):
            n = _nrokey(row[7])
            if n and row[2]:
                out.setdefault((n, _clave_proveedor(row[6])), row[2])
        return out
    finally:
        wb.close()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--aplicar", action="store_true")
    args = ap.parse_args()

    res = analizar(dias=3650)
    pend = [m for m in res["confirmados"] if not m["factura"]["pagada"]]
    fxp = _fecha_fxp()

    vinculos = []
    print("%-9s %-30s %-11s %-11s %s" % ("Nº", "proveedor", "monto", "pago", "origen"))
    print("-" * 82)
    for m in sorted(pend, key=lambda x: x["cargo"]["fecha"]):
        f, c = m["factura"], m["cargo"]
        clave = (f["nro"], _clave_proveedor(f["prov"]))
        pago = fxp.get(clave)
        origen = "FXP"
        if pago is None:
            pago, origen = c["fecha"], "banco"
        pago_str = str(pago)[:10]
        print("%-9s %-30s $%-10s %-11s %s"
              % (f["nro"], f["prov"][:30], format(int(f["total"]), ","),
                 pago_str, origen))
        vinculos.append({
            "fila_banco": c["fila"], "tipo_doc": "factura",
            "fila_doc": f["filas"][0], "nro_doc": f["nro"],
            "proveedor": f["prov"], "monto_asignado": c["monto"],
            "criterio": m["criterio"], "filas_doc": f["filas"],
            "fecha_pago": pago_str,
            "nota": "Conciliado 2026-09-02 con la cartola al 2-sep; "
                    "fecha de pago segun %s" % origen,
        })

    print("\n%d factura(s), $%s"
          % (len(vinculos), format(int(sum(v["monto_asignado"] for v in vinculos)), ",")))
    if not args.aplicar:
        print("(simulacion: no se escribio nada)")
        return

    from modules.conciliacion_store import registrar_vinculos
    r = registrar_vinculos(vinculos, usuario="conciliacion sep-2026")
    print("Registrados %d vinculos: %s" % (r["registrados"], r["ids"]))


if __name__ == "__main__":
    main()
