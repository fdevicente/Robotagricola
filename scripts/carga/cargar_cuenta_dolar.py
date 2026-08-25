"""Crea la hoja `Cuenta Dolar` y carga el saldo de apertura.

El saldo viene de la captura del portal del 2026-08-05. Los movimientos que lo
formaron (pagos de exportadores en USD) no están cargados uno a uno: se registra
un saldo de apertura y de ahí en adelante se van agregando movimientos.
"""
import shutil
import sys
from datetime import date, datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb
from modules.cuentas import DOLAR_HEADERS, DOLAR_SHEET, caja_total

FECHA = date(2026, 8, 5)
SALDO_USD = 141_701.84
DETALLE = "Saldo de apertura — cuenta dólar Scotiabank ****9350 (captura del portal)"

resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"Respaldo: {resp}\n")

wb = load_workbook(EXCEL_PATH)
if DOLAR_SHEET not in wb.sheetnames:
    ws = wb.create_sheet(DOLAR_SHEET)
    ws.append(DOLAR_HEADERS)
    print(f"Hoja '{DOLAR_SHEET}' creada.")
else:
    ws = wb[DOLAR_SHEET]
    print(f"Hoja '{DOLAR_SHEET}' ya existía ({ws.max_row - 1} movimientos).")

ya = any(str(ws.cell(r, 2).value or "").startswith("Saldo de apertura")
         for r in range(2, ws.max_row + 1))
if ya:
    print("El saldo de apertura ya estaba cargado — no se duplica.")
else:
    # Fecha · Descripcion · Referencia · Cargo · Abono · Saldo · Tipo · Categoria · Cultivo · Link
    ws.append([FECHA, DETALLE, "", None, SALDO_USD, SALDO_USD,
               "APERTURA", "INGRESO VENTAS", "NOGALES", ""])
    print(f"Cargado: US$ {SALDO_USD:,.2f} al {FECHA}")

_save_wb(wb)
wb.close()

print()
from modules.cuentas import formato
print(formato(caja_total()))
