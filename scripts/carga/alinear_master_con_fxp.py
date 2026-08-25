"""Deja pendientes en el Master solo las facturas que FXP tiene por pagar.

Todo lo demás que figuraba pendiente se cierra ("cancelada"), buscando primero
el cargo real en el banco para usar SU fecha; si no aparece, se usa la fecha de
vencimiento y queda anotado que se cerró al alinear con FXP.

Las filas duplicadas NO se borran (la hoja Conciliaciones guarda números de
fila): se marcan y se cierran.

Uso:  python scripts/carga/alinear_master_con_fxp.py [--aplicar]
"""
import os
import re
import shutil
import sys
import tempfile
from datetime import date, datetime, timedelta

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb
from modules.correlativo import ALIAS_PROVEEDOR

APLICAR = "--aplicar" in sys.argv
FXP = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"
HOY = date.today()
COL_PAGO, COL_NOTA = 3, 20        # Fecha Pago · Categorizado_por


def _pd(v):
    """Fecha desde datetime, date o TEXTO.

    Sin el caso texto, las filas con la fecha como string daban None y el
    filtro por ventana de fechas dejaba pasar cargos de 2018 para facturas
    de 2026.
    """
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for f in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(v.strip()[:10], f).date()
            except ValueError:
                continue
    return None


def nro_key(n):
    s = str(n or "").strip().upper().replace(" ", "")
    if s.endswith(".0"):
        s = s[:-2]
    return re.sub(r"\D", "", s) or s


def prov_key(p):
    s = " ".join(str(p or "").upper().split())
    for g in ALIAS_PROVEEDOR:
        if any(a.upper() in s for a in g):
            return sorted(g)[0].upper()
    s = re.sub(r"\b(LTDA|SPA|S\.A\.|SA|LIMITADA|Y CIA|E HIJOS|SUR|CHILE)\b", "", s)
    s = re.sub(r"[^A-ZÑ ]", " ", s)
    return " ".join([w for w in s.split() if len(w) > 2][:2])


# Palabras que aparecen en medio mundo y no identifican a nadie
GENERICAS = {"LIMITADA", "LTDA", "COMERCIAL", "SERVICIOS", "SOCIEDAD",
             "AGRICOLA", "EMPRESAS", "DISTRIBUIDORA", "INDUSTRIAL",
             "VENTAS", "DETALLE", "ADMINISTRADORA", "GENERAL", "CHILE",
             "NACIONAL", "REGIONAL", "TRANSPORTES", "MAQUINARIAS"}


def tokens(s):
    return {w for w in re.sub(r"[^A-ZÑ ]", " ", str(s or "").upper()).split()
            if len(w) > 3 and w not in GENERICAS}


def coincide_proveedor(prov, desc) -> bool:
    """El proveedor aparece en la glosa del banco.

    Se compara por prefijo para tolerar cómo lo abrevia el banco
    ("EMPRESAS LIPIGA" ↔ "LIPIGAS"). Se exige un token distintivo, no
    cualquiera: si no, "Comercial Álamos" calzaba con "Riego Control
    Limitada" por la palabra "limitada".
    """
    tp, td = tokens(prov), tokens(desc)
    if not tp or not td:
        return False
    for a in tp:
        for b in td:
            if a == b:
                return True
            # Uno tiene que ser prefijo COMPLETO del otro: "LIPIGAS"/"LIPIGA"
            # sí, pero "FACTURIA"/"FACTURAS" no (solo comparten la raíz).
            corto, largo = (a, b) if len(a) <= len(b) else (b, a)
            if len(corto) >= 5 and largo.startswith(corto):
                return True
    return False


# ── Las que FXP tiene por pagar ──
tmp = os.path.join(tempfile.gettempdir(), "fxp_al.xlsx")
shutil.copy2(FXP, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
por_pagar = set()
for row in wb["FXP"].iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    s = str(row[11]).strip().upper() if row[11] is not None else ""
    if s in ("PAGADA", "NN", ""):
        continue
    por_pagar.add((prov_key(row[6]), nro_key(row[7])))
wb.close()

# Mismo documento con el número escrito distinto en cada planilla. Cuando se
# resuelve cuál es el bueno, se corrige en el Master y se saca de acá.
EQUIVALENCIAS: dict = {}
for master_k, fxp_k in EQUIVALENCIAS.items():
    if fxp_k in por_pagar:
        por_pagar.add(master_k)
        print(f"  Equivalencia: Master {master_k[1]} ≡ FXP {fxp_k[1]}")

print(f"FXP tiene {len(por_pagar) - len(EQUIVALENCIAS)} documentos por pagar\n")

# ── Cargos del banco, para fechar los pagos ──
wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
cargos = []
for row in wb["Cuenta Banco"].iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    try:
        c = float(row[3] or 0)
    except (TypeError, ValueError):
        continue
    if c > 0:
        cargos.append({"fecha": _pd(row[0]), "desc": str(row[1] or ""), "monto": c})

# ── Pendientes del Master ──
ws = wb["Facturas"]
grupos = {}
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    if not row or not row[0] or not row[6]:
        continue
    k = (prov_key(row[3]), nro_key(row[6]))
    g = grupos.setdefault(k, {"prov": str(row[3] or ""), "nro": str(row[6] or ""),
                               "filas": [], "total": 0.0, "pago": None,
                               "emision": _pd(row[0]), "venc": _pd(row[1]),
                               "doc": str(row[5] or ""), "nn": False})
    g["filas"].append(i)
    try:
        g["total"] = max(g["total"], float(row[15] or 0))
    except (TypeError, ValueError):
        pass
    if row[2] and str(row[2]).strip():
        g["pago"] = _pd(row[2]) or row[2]
    if len(row) > 19 and "NN" in str(row[19] or "").upper():
        g["nn"] = True
wb.close()

pendientes = {k: g for k, g in grupos.items() if not g["pago"] and not g["nn"]}
quedan = {k: g for k, g in pendientes.items() if k in por_pagar}
cerrar = {k: g for k, g in pendientes.items() if k not in por_pagar}


def buscar_pago(g):
    """Cargo(s) del banco que pagan esta factura.

    Devuelve (fecha, detalle) o None. Exige proveedor + monto + ventana de
    fechas entre la emisión y hoy: un pago no puede ser anterior a la factura
    ni posterior a hoy.
    """
    desde = (g["emision"] or date(2000, 1, 1)) - timedelta(days=5)
    candidatos = [c for c in cargos
                  if c["fecha"] and desde <= c["fecha"] <= HOY
                  and coincide_proveedor(g["prov"], c["desc"])]
    if not candidatos:
        return None

    tol = max(1, g["total"] * 0.01)
    # 1) un solo cargo por el total
    exactos = [c for c in candidatos if abs(c["monto"] - g["total"]) <= tol]
    if exactos:
        c = min(exactos, key=lambda x: x["fecha"])
        return c["fecha"], f"banco {c['fecha']}: {c['desc'][:34]}"

    # 2) pagada en cuotas: varios cargos del mismo proveedor que suman el total
    #    (el caso de las plantas de avellano, 5 transferencias)
    candidatos.sort(key=lambda x: x["fecha"])
    for i in range(len(candidatos)):
        suma = 0.0
        for j in range(i, len(candidatos)):
            suma += candidatos[j]["monto"]
            if abs(suma - g["total"]) <= tol and j > i:
                return (candidatos[j]["fecha"],
                        f"banco: {j - i + 1} cuotas hasta {candidatos[j]['fecha']}")
            if suma > g["total"] + tol:
                break
    return None


print("=" * 92)
print(f"QUEDAN PENDIENTES ({len(quedan)}) — las que FXP tiene por pagar")
print("=" * 92)
for k, g in sorted(quedan.items(), key=lambda x: -x[1]["total"]):
    print(f"  {g['prov'][:30]:30} N°{g['nro']:>11} ${g['total']:>12,.0f}")

print("\n" + "=" * 92)
print(f"SE CIERRAN ({len(cerrar)}) — no figuran por pagar en FXP")
print("=" * 92)
print(f"  {'proveedor':28}{'documento':>12}{'monto':>13}  {'fecha de pago':22} origen")
print("  " + "-" * 88)
acciones, total = [], 0.0
for k, g in sorted(cerrar.items(), key=lambda x: -x[1]["total"]):
    total += g["total"]
    pago = buscar_pago(g)
    if pago:
        fecha, origen = pago
        nota = f"Pagada — {origen}"
    else:
        # Sin evidencia de la fecha real: se usa la de alineación y se dice.
        # Nunca el vencimiento, que puede ser futuro.
        fecha = HOY
        origen = "sin cargo identificado → se cierra con fecha de hoy"
        nota = f"Cerrada al alinear con FXP el {HOY} (FXP no la tiene por pagar)"
    acciones.append((g, fecha, nota))
    print(f"  {g['prov'][:28]:28}{g['nro']:>12}${g['total']:>12,.0f}  "
          f"{str(fecha):12} {origen[:48]}")
print("  " + "-" * 88)
print(f"  {'TOTAL':28}{len(cerrar):>12}{total:>13,.0f}")

con_banco = sum(1 for _g, _f, n in acciones if n.startswith("Pagada"))
print(f"\n  {con_banco} con su cargo identificado en el banco · "
      f"{len(acciones) - con_banco} cerradas por alineación")

if not APLICAR:
    print("\n(simulación — nada se escribió; agrega --aplicar)")
    sys.exit(0)

resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"\nRespaldo: {resp}")

wb = load_workbook(EXCEL_PATH)
ws = wb["Facturas"]
n = 0
for g, fecha, nota in acciones:
    for f in g["filas"]:
        ws.cell(f, COL_PAGO).value = fecha
        previo = str(ws.cell(f, COL_NOTA).value or "")
        ws.cell(f, COL_NOTA).value = (previo + " · " if previo else "") + nota
        n += 1
_save_wb(wb)
wb.close()
print(f"✅ {n} líneas cerradas ({len(acciones)} facturas).")
print(f"   Quedan {len(quedan)} pendientes, alineadas con FXP.")
