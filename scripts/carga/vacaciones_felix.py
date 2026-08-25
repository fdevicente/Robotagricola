"""Registra las vacaciones del dueño (Felix De Vicente) en 2026.

Se cuentan DÍAS HÁBILES (lunes a viernes, sin feriados), que es el criterio
con que están registradas las del resto: p.ej. "19-01 al 30-01 = 10 días".

Uso:  python scripts/carga/vacaciones_felix.py [--aplicar]
"""
import shutil
import sys
from datetime import date, datetime, timedelta

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb
from modules.feriados import anio_cubierto, es_feriado, nombre_feriado

APLICAR = "--aplicar" in sys.argv
NOMBRE = "Felix De Vicente"
DIAS = ["lun", "mar", "mié", "jue", "vie", "sáb", "dom"]

PERIODOS = [
    (date(2026, 2, 1), date(2026, 2, 28), "Febrero completo"),
    (date(2026, 6, 13), date(2026, 7, 14), ""),
]


def habiles(desde, hasta):
    """Días hábiles del período, y el detalle de lo que se descontó."""
    n, feriados, findes = 0, [], 0
    d = desde
    while d <= hasta:
        if d.weekday() >= 5:
            findes += 1
        elif es_feriado(d):
            feriados.append(d)
        else:
            n += 1
        d += timedelta(days=1)
    return n, feriados, findes


for a in {p[0].year for p in PERIODOS} | {p[1].year for p in PERIODOS}:
    if not anio_cubierto(a):
        print(f"⚠️ Faltan los feriados de {a} — el conteo puede estar mal.\n")

print("=" * 68)
print(f"VACACIONES DE {NOMBRE.upper()} — 2026")
print("=" * 68)
total = 0
filas = []
for desde, hasta, nota in PERIODOS:
    n, fer, findes = habiles(desde, hasta)
    total += n
    corridos = (hasta - desde).days + 1
    print(f"\n▸ {desde:%d-%m-%Y} ({DIAS[desde.weekday()]}) → "
          f"{hasta:%d-%m-%Y} ({DIAS[hasta.weekday()]})")
    print(f"    {corridos} días corridos − {findes} de fin de semana"
          + (f" − {len(fer)} feriado(s)" if fer else "")
          + f"  =  {n} días hábiles")
    for f in fer:
        print(f"      · {f:%d-%m} {nombre_feriado(f)}")
    filas.append((desde, hasta, n, nota))

print("\n" + "=" * 68)
print(f"TOTAL: {total} días hábiles")
print("=" * 68)

# Impacto en el saldo
wb = load_workbook(EXCEL_PATH, data_only=True)
base = None
for row in wb["Vacaciones Pendientes"].iter_rows(min_row=2, values_only=True):
    if row and str(row[0] or "").strip() == NOMBRE:
        f = row[4]
        base = {"saldo": float(row[3] or 0),
                "fecha": f.date() if isinstance(f, datetime) else f}
ya = 0.0
for row in wb["Vacaciones"].iter_rows(min_row=2, values_only=True):
    if row and str(row[0] or "").strip() == NOMBRE:
        ya += float(row[3] or 0)
wb.close()

if base:
    hoy = date.today()
    meses = (hoy.year - base["fecha"].year) * 12 + (hoy.month - base["fecha"].month)
    ganados = base["saldo"] + meses * 15 / 12
    print(f"\n  Ganados desde {base['fecha']}   : {ganados:>7.2f}")
    print(f"  Ya registradas                : {ya:>7.0f}")
    print(f"  Estas vacaciones              : {total:>7.0f}")
    print(f"  {'PENDIENTES QUEDARÍAN':30}: {ganados - ya - total:>7.2f}"
          f"   (antes {ganados - ya:.2f})")

if not APLICAR:
    print("\n(simulación — nada se escribió; agrega --aplicar)")
    sys.exit(0)

resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"\nRespaldo: {resp}")

wb = load_workbook(EXCEL_PATH)
ws = wb["Vacaciones"]
existentes = {(str(r[0] or "").strip(), str(r[1])[:10])
              for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]}
n = 0
for desde, hasta, dias, nota in filas:
    if (NOMBRE, str(desde)) in existentes:
        print(f"  Ya estaba: {desde} — no se duplica")
        continue
    ws.append([NOMBRE, desde, hasta, dias, "Aprobado", nota or "Periodo 2026"])
    n += 1
_save_wb(wb)
wb.close()
print(f"✅ {n} períodos agregados a la hoja Vacaciones.")
