"""Importa las dos cartolas del 5-ago-2026: cuenta corriente y cuenta dólar.

Uso:  python scripts/carga/importar_cartolas_ago.py [--aplicar]
Sin --aplicar solo muestra qué haría.
"""
import shutil
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")
from config import EXCEL_PATH
from modules.banco_import import analizar_cartola, importar_cartola
from modules.cuentas import DOLAR_SHEET, formato, caja_total

APLICAR = "--aplicar" in sys.argv
DESCARGAS = r"C:\Users\Windows\Downloads"
CARTOLAS = [
    ("Cuenta corriente (CLP)", rf"{DESCARGAS}\typeDesc (1).txt", None, 0),
    ("Cuenta dólar (USD)",      rf"{DESCARGAS}\typeDesc (2).txt", DOLAR_SHEET, 2),
]


def _resumen(titulo, res, simbolo):
    print(f"\n{'=' * 62}\n{titulo}\n{'=' * 62}")
    print(f"  Movimientos en el archivo : {res['total_archivo']}")
    print(f"  Ya estaban en el Master   : {res['duplicados']}")
    print(f"  NUEVOS                    : {len(res['nuevos'])}")
    if res.get("ultima_fecha_master"):
        print(f"  El Master llegaba hasta   : {res['ultima_fecha_master']}")
    if res.get("saldo_archivo") is not None:
        print(f"  Saldo según la cartola    : {simbolo}{res['saldo_archivo']:,.2f}")
    if res["nuevos"]:
        print(f"  Cargos {simbolo}{res['cargos']:,.2f} · Abonos {simbolo}{res['abonos']:,.2f}\n")
        for m in res["nuevos"]:
            signo = -m["cargo"] if m["cargo"] else m["abono"]
            print(f"    {m['fecha']}  {signo:>16,.2f}  {m['desc'][:44]}")


def _quitar_saldo_apertura():
    """Borra el stub de apertura de la cuenta dólar.

    Se cargó cuando solo teníamos la captura del portal; ahora llega el
    historial real y dejarlo contaría el saldo dos veces.
    """
    from openpyxl import load_workbook
    from excel_manager import _save_wb
    wb = load_workbook(EXCEL_PATH)
    ws = wb[DOLAR_SHEET]
    borradas = 0
    for r in range(ws.max_row, 1, -1):
        if str(ws.cell(r, 2).value or "").startswith("Saldo de apertura"):
            ws.delete_rows(r)
            borradas += 1
    if borradas:
        _save_wb(wb)
        print(f"  Quitado el saldo de apertura provisorio ({borradas} fila).")
    wb.close()


if APLICAR:
    resp = shutil.copy2(EXCEL_PATH,
                        EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
    print(f"Respaldo: {resp}")
    _quitar_saldo_apertura()

for titulo, path, hoja, dec in CARTOLAS:
    simbolo = "US$" if hoja else "$"
    if APLICAR:
        res = importar_cartola(path, hoja, dec)
        _resumen(titulo, res, simbolo)
        print(f"\n  ✅ {res['agregados']} movimientos agregados.")
    else:
        _resumen(titulo, analizar_cartola(path, hoja, dec), simbolo)

if not APLICAR:
    print("\n\n(simulación — nada se escribió; agrega --aplicar para importar)")
else:
    print("\n" + formato(caja_total()).replace("*", ""))
