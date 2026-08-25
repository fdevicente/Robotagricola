#!/usr/bin/env python3
"""Deja los montos del Master iguales a los de FXP (FXP = fuente correcta).

Las diferencias venían de IVA u otros ajustes. Como los reportes de costos suman
la columna "Total por Item" (O), no basta con corregir "TOTAL FACTURA" (P):
se reescalan las líneas proporcionalmente para que sumen el monto de FXP.

Se ajustan: Valor unitario (J), IVA (M), Total por Item (O) y TOTAL FACTURA (P).
El Impuesto Específico (N) NO se toca (es un monto fijo por litro).
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import os, shutil, tempfile
from collections import defaultdict
from openpyxl import load_workbook
from config import EXCEL_PATH
from excel_manager import _save_wb
from modules.correlativo import norm_prov, nrokey, FXP_PATH


def _f(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


# ── FXP ──
tmp = os.path.join(tempfile.gettempdir(), "fxp_ajuste.xlsx")
shutil.copy2(FXP_PATH, tmp)
wbf = load_workbook(tmp, read_only=True, data_only=True)
fxp, fxp_nro = {}, {}
for row in wbf["FXP"].iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    nro = nrokey(row[7])
    if not nro:
        continue
    m = row[8]
    if isinstance(m, str):
        s = m.upper().replace("USD", "").replace("$", "").strip().replace(".", "").replace(",", ".")
        try:
            m = float(s)
        except ValueError:
            m = 0
    m = _f(m)
    if m <= 0:
        continue
    fxp[(norm_prov(row[6]), nro)] = m
    # un mismo nº puede existir para varios proveedores → guardar todos
    fxp_nro.setdefault(nro, []).append({"prov": str(row[6] or ""), "monto": m})
wbf.close()

# ── Master ──
wb = load_workbook(EXCEL_PATH)
ws = wb["Facturas"]
grupos = defaultdict(list)
for r in range(2, ws.max_row + 1):
    if not ws.cell(r, 1).value:
        continue
    nro = nrokey(ws.cell(r, 7).value)
    if not nro:
        continue
    grupos[(norm_prov(ws.cell(r, 4).value), nro)].append(r)

from modules.correlativo import buscar_en_fxp

ajustadas, sin_lineas, dudosas = [], [], []
for (prov, nro), filas in grupos.items():
    prov_master = str(ws.cell(filas[0], 4).value or "")
    ref = fxp.get((prov, nro))
    if ref is None:
        # varios proveedores pueden compartir el nº: elegir por proveedor
        cand = buscar_en_fxp(prov_master, nro, fxp_nro)
        if not cand:
            otros = fxp_nro.get(nro) or []
            if otros:
                dudosas.append((prov_master, str(ws.cell(filas[0], 7).value or ""),
                                 _f(ws.cell(filas[0], 16).value), otros[0]["monto"],
                                 otros[0]["prov"], "proveedor distinto"))
            continue
        ref = cand["monto"]
    if not ref:
        continue
    suma = sum(_f(ws.cell(r, 15).value) for r in filas)
    total_actual = _f(ws.cell(filas[0], 16).value)
    if abs(total_actual - ref) <= max(1000, ref * 0.01):
        continue                       # ya calza

    nombre = str(ws.cell(filas[0], 4).value or "")
    ndoc = str(ws.cell(filas[0], 7).value or "")
    if suma <= 0:
        # sin líneas con monto: solo se corrige el total
        for r in filas:
            ws.cell(r, 16).value = round(ref)
        sin_lineas.append((nombre, ndoc, total_actual, ref))
        continue

    factor = ref / suma
    # Un factor muy alejado de 1 (o de 0.5 por duplicación) sugiere que no es
    # la misma factura: se deja para revisión manual en vez de reescribirla.
    if not (0.45 <= factor <= 2.2):
        dudosas.append((nombre, ndoc, total_actual, ref, "", f"factor {factor:.2f}"))
        continue
    for r in filas:
        for col in (10, 13, 15):        # valor unitario, IVA, total por ítem
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                ws.cell(r, col).value = round(v * factor, 2)
        ws.cell(r, 16).value = round(ref)
    ajustadas.append((nombre, ndoc, total_actual, ref, len(filas), factor))

_save_wb(wb)
wb.close()

ajustadas.sort(key=lambda x: -abs(x[2] - x[3]))
print(f"✅ Facturas ajustadas al monto de FXP: {len(ajustadas)}")
print(f"   (+ {len(sin_lineas)} en las que solo se corrigió el total)\n")
print(f"{'Proveedor':32} {'Documento':13} {'Master':>13} {'→ FXP':>13} {'Dif':>13}  x")
print("-" * 92)
for prov, nro, antes, ref, n, factor in ajustadas[:40]:
    print(f"{prov[:32]:32} F{nro[:12]:12} ${antes:>12,.0f} ${ref:>12,.0f} "
          f"${ref-antes:>+12,.0f}  {factor:.3f}")
if len(ajustadas) > 40:
    print(f"… y {len(ajustadas)-40} más")
if sin_lineas:
    print("\nSolo se corrigió el TOTAL (no tenían montos por línea):")
    for prov, nro, antes, ref in sin_lineas:
        print(f"  {prov[:32]:32} F{nro[:12]:12} ${antes:>12,.0f} → ${ref:>12,.0f}")

if dudosas:
    print(f"\n⚠️ NO ajustadas — requieren revisión manual: {len(dudosas)}")
    print(f"{'Proveedor (Master)':32} {'Doc':13} {'Master':>13} {'FXP':>13}  motivo")
    print("-" * 100)
    for prov, nro, antes, ref, prov_fxp, motivo in sorted(
            dudosas, key=lambda x: -abs(x[2] - x[3])):
        extra = f" · FXP dice: {prov_fxp[:26]}" if prov_fxp else ""
        print(f"{prov[:32]:32} F{nro[:12]:12} ${antes:>12,.0f} ${ref:>12,.0f}  {motivo}{extra}")

tot_antes = sum(a[2] for a in ajustadas)
tot_ahora = sum(a[3] for a in ajustadas)
print(f"\nEfecto neto en los costos: ${tot_antes:,.0f} → ${tot_ahora:,.0f} "
      f"({tot_ahora-tot_antes:+,.0f})")
