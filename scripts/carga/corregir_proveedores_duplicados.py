# -*- coding: utf-8 -*-
"""Unifica proveedores que estaban cargados con dos grafias y dos RUT.

Se corre A MANO y UNA VEZ. Con --simular no escribe nada.

Salio de la migracion a Drive (31-ago-2026): 20 documentos no enlazaban porque
el nombre del proveedor en el Master no coincidia con el del archivo. Al
revisarlos aparecio algo peor que un typo — RUT invalidos:

  FERRETERIAINDUTRIAL TALCA LIMITADA    31 filas  78.045.980-8  <- DV correcto
  Ferreteria Industrial Talca Limitada   9 filas  78.045.980-6  <- DV INVALIDO

O sea el nombre bien escrito llevaba el RUT malo y viceversa. El digito
verificador se calcula, no se opina: 78045980 -> 8.

  IRRIFER  6 filas  '76,300,530-3' (con COMAS)
  IRRIFOR / Irrifor / Irrifer  9 filas  76155160-4

Decision del dueño (31-ago): son el mismo proveedor, se unifica a Irrifor con
76155160-4.

Y un typo simple: 'Salina y fabres' (falta la s) -> 'Salinas y Fabres'.
"""
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(
    os.path.dirname(os.path.abspath(__file__)))))

from config import EXCEL_PATH

COL_PROVEEDOR, COL_RUT = 4, 5

# nombre viejo (comparado sin mayusculas ni espacios de mas) -> (nombre, rut)
# rut None = no tocar el RUT
REGLAS = {
    "ferreteriaindutrial talca limitada": ("Ferreteria Industrial Talca Limitada", "78.045.980-8"),
    "ferreteria industrial talca limitada": ("Ferreteria Industrial Talca Limitada", "78.045.980-8"),
    "irrifer": ("Irrifor", "76155160-4"),
    "irrifor": ("Irrifor", "76155160-4"),
    "salina y fabres": ("Salinas y Fabres", None),
    "salinas y fabres": ("Salinas y Fabres", None),
}


def _clave(v):
    return " ".join(str(v or "").split()).casefold()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--simular", action="store_true")
    args = ap.parse_args()

    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_PATH)
    try:
        ws = wb["Facturas"]
        cambios = []
        for f in range(2, ws.max_row + 1):
            actual = ws.cell(f, COL_PROVEEDOR).value
            regla = REGLAS.get(_clave(actual))
            if not regla:
                continue
            nombre, rut = regla
            rut_actual = str(ws.cell(f, COL_RUT).value or "").strip()
            toca_nombre = str(actual or "").strip() != nombre
            toca_rut = rut is not None and rut_actual != rut
            if not (toca_nombre or toca_rut):
                continue
            cambios.append((f, actual, nombre, rut_actual, rut if toca_rut else None))
            if not args.simular:
                if toca_nombre:
                    ws.cell(f, COL_PROVEEDOR).value = nombre
                if toca_rut:
                    ws.cell(f, COL_RUT).value = rut

        print("Filas a cambiar: %d" % len(cambios))
        for f, viejo, nuevo, rv, rn in cambios:
            print("   fila %-6s %-38s -> %s" % (f, viejo, nuevo))
            if rn:
                print("              RUT %-16s -> %s" % (rv or "(vacio)", rn))
        if args.simular:
            print("\n(simulacion: no se escribio nada)")
            return
        wb.save(EXCEL_PATH)          # ruta EXPLICITA
        print("\nGuardado en %s" % EXCEL_PATH)
    finally:
        wb.close()


if __name__ == "__main__":
    main()
