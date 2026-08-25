"""Marca como recibidas las cuotas de Pacific Nuts que llegaron a la cuenta dólar.

La cartola USD (5-ago-2026) muestra los ABONO COMEX en las fechas exactas de las
cuotas 1, 2 y 3. Estaban como "esperado" porque el modelo solo miraba la cuenta
en pesos.
"""
import shutil
import sys
from datetime import date, datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb

# cuota → (fecha del abono, USD recibidos) según la cartola de la cuenta dólar
RECIBIDAS = {
    1: (date(2026, 5, 22), 80_000.00),
    2: (date(2026, 6, 26), 31_807.00),
    3: (date(2026, 7, 24), 29_815.00),
}
EXPORTADORA = "Pacific Nuts"
COL_EXP, COL_CUOTA, COL_USD, COL_ESTADO = 4, 8, 10, 12
COL_FECHA_REAL, COL_MONTO_REAL, COL_MONEDA, COL_NOTAS = 13, 14, 15, 16

resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"Respaldo: {resp}\n")

wb = load_workbook(EXCEL_PATH)
ws = wb["Cosechas"]

print(f"{'fila':>4} {'cuota':>5} {'esperado':>10} {'recibido':>10} {'fecha':12} dif")
print("-" * 56)
tocadas = 0
for r in range(2, ws.max_row + 1):
    if str(ws.cell(r, COL_EXP).value or "").strip() != EXPORTADORA:
        continue
    try:
        cuota = int(ws.cell(r, COL_CUOTA).value or 0)
    except (TypeError, ValueError):
        continue
    if cuota not in RECIBIDAS:
        continue
    if str(ws.cell(r, COL_ESTADO).value or "").lower() == "recibido":
        print(f"{r:>4} {cuota:>5}   (ya estaba marcada)")
        continue

    fecha, usd = RECIBIDAS[cuota]
    esperado = float(ws.cell(r, COL_USD).value or 0)
    ws.cell(r, COL_ESTADO).value = "recibido"
    ws.cell(r, COL_FECHA_REAL).value = fecha
    ws.cell(r, COL_MONTO_REAL).value = usd
    ws.cell(r, COL_MONEDA).value = "USD"
    ws.cell(r, COL_NOTAS).value = (
        f"ABONO COMEX a la cuenta dólar el {fecha} — US$ {usd:,.2f}")
    tocadas += 1
    print(f"{r:>4} {cuota:>5} {esperado:>10,.0f} {usd:>10,.2f} {str(fecha):12} "
          f"{usd - esperado:+,.0f}")

if tocadas:
    _save_wb(wb)
    print(f"\n✅ {tocadas} cuotas marcadas como recibidas.")
else:
    print("\nNada que cambiar.")
wb.close()
