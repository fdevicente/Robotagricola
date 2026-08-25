"""Quita las jornadas-hombre de los días marcados como ausencia.

Juan escribe "Aucente" (por ausente) y el filtro viejo no lo reconocía, así que
esas filas quedaron sumando 1 JH cada una. Corrige lo ya guardado.
"""
import shutil
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb
from modules.bitacora_asistencia import _sin_tildes

import re
AUSENCIA = re.compile(r"au[cs]+ente|vaca[cs]ion|licen[cs]ia|permiso|inasist|falt|no vino")

COL_FECHA, COL_ACT, COL_JH, COL_TRAB = 1, 4, 7, 8

resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx",
                                       f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"Respaldo: {resp}\n")

wb = load_workbook(EXCEL_PATH)
ws = wb["Bitácora"]

cambios = []
for r in range(2, ws.max_row + 1):
    act = ws.cell(r, COL_ACT).value
    jh = ws.cell(r, COL_JH).value
    if not act or jh in (None, ""):
        continue
    if AUSENCIA.search(_sin_tildes(str(act))):
        cambios.append((r, ws.cell(r, COL_FECHA).value, act, jh,
                        ws.cell(r, COL_TRAB).value))
        ws.cell(r, COL_JH).value = None

if not cambios:
    print("Nada que corregir.")
    sys.exit(0)

print(f"{'fila':>5}  {'fecha':10} {'actividad':22} {'JH':>3}  trabajadores")
print("-" * 66)
for r, f, act, jh, trab in cambios:
    fecha = f.date() if hasattr(f, "date") else f
    print(f"{r:>5}  {str(fecha):10} {str(act)[:22]:22} {jh:>3}  {trab}")

_save_wb(wb)
wb.close()
print(f"\n✅ {len(cambios)} filas corregidas "
      f"(−{sum(c[3] for c in cambios)} jornadas-hombre mal contadas).")
