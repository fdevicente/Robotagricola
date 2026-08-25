"""Javier González: fecha de contrato real 2025-11-01 (la anterior era provisoria).

Recalcula su saldo de vacaciones y lo incorpora a `Vacaciones Pendientes`, que es
de donde leen el dashboard y el reporte mensual (antes no aparecía).
"""
import shutil
import sys
from datetime import date, datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb

HOY = date(2026, 8, 4)
DIAS_MES = 15.0 / 12.0
NOMBRE = "Javier Gonzalez"
RUT = "20.230.894-5"
CARGO = "OPERARIO AGRICOLA"
INGRESO = date(2025, 11, 1)

resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"Respaldo: {resp}\n")

wb = load_workbook(EXCEL_PATH)

# Días ya tomados según la hoja Vacaciones (solo aprobadas)
tomados = 0.0
for row in wb["Vacaciones"].iter_rows(min_row=2, values_only=True):
    if row[0] and str(row[0]).strip().upper() == NOMBRE.upper():
        if "aprobad" in str(row[4] or "").lower():
            tomados += float(row[3] or 0)

meses = (HOY.year - INGRESO.year) * 12 + (HOY.month - INGRESO.month)
acumulados = meses * DIAS_MES
saldo = round(acumulados - tomados, 2)

print(f"Contrato        : {INGRESO}")
print(f"Meses cumplidos : {meses}  →  {acumulados:.2f} días acumulados (1,25/mes)")
print(f"Días tomados    : {tomados:g}")
print(f"SALDO           : {saldo:.2f} días\n")

# ── Hoja Personal ──
ws = wb["Personal"]
fila = next((r for r in range(2, ws.max_row + 1)
             if str(ws.cell(r, 1).value or "").strip().upper() == NOMBRE.upper()), None)
if not fila:
    print("❌ Javier no está en la hoja Personal"); sys.exit(1)

antes = ws.cell(fila, 5).value
ws.cell(fila, 4).value = INGRESO
ws.cell(fila, 5).value = saldo
ws.cell(fila, 6).value = tomados
ws.cell(fila, 8).value = (f"Contrato {INGRESO} · {meses} meses × 1,25 "
                           f"− {tomados:g} tomados")
print(f"Personal fila {fila}: días pendientes {antes} → {saldo}")

# ── Hoja Vacaciones Pendientes (fuente del dashboard y del reporte) ──
wsp = wb["Vacaciones Pendientes"]
ya = next((r for r in range(2, wsp.max_row + 1)
           if str(wsp.cell(r, 1).value or "").strip().upper() == NOMBRE.upper()), None)
destino = ya or wsp.max_row + 1
wsp.cell(destino, 1).value = NOMBRE
wsp.cell(destino, 2).value = RUT
wsp.cell(destino, 3).value = INGRESO      # Fecha Contrato
wsp.cell(destino, 4).value = 0            # Saldo base: parte de cero al contratarse
wsp.cell(destino, 5).value = INGRESO      # Fecha del Saldo
wsp.cell(destino, 6).value = "Contrato 01-11-2025 confirmado por el dueño"
wsp.cell(destino, 7).value = CARGO
print(f"Vacaciones Pendientes fila {destino}: "
      f"{'actualizado' if ya else 'AGREGADO'} (antes no estaba → no salía en el dashboard)")

_save_wb(wb)
wb.close()
print("\n✅ Listo.")
