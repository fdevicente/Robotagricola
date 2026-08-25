#!/usr/bin/env python3
"""Reconstruye el Inventario desde el conteo físico de Juan (11-jun-2026).

- La hoja anterior (86 filas auto-generadas desde facturas, con unidades y
  cantidades erróneas: 5000 L, 35000 L…) se preserva como 'Inventario ANTIGUO'.
- Base = conteo físico de Juan del 11-jun.
- Se aplican los movimientos posteriores: compras (+) y aplicaciones (−).
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import EXCEL_PATH
from excel_manager import _save_wb

FECHA_CONTEO = "2026-06-11"

# (producto, categoría, unidad, stock contado 11-jun)
CONTEO = [
    ("Espirodiclofen 240 SC", "Insecticida", "L", 0.5),
    ("Konan", "Insecticida", "L", 0.1),
    ("Defender Zn", "Fertilizante", "L", 10),
    ("Defender Calcio", "Fertilizante", "L", 10),
    ("Defender K", "Fertilizante", "L", 12),
    ("Manvert Avanza", "Bioestimulante", "L", 3),
    ("Defender Boro", "Fertilizante", "L", 2.5),
    ("Black K", "Fertilizante", "L", 18),
    ("Cito Four", "Bioestimulante", "L", 0.5),
    ("Macroquel Magnesio", "Fertilizante", "L", 17),
    ("Nitrato de Potasio", "Fertilizante", "Kg", 25),
    ("Urea", "Fertilizante", "Kg", 75),
    ("Procal", "Fertilizante", "L", 1),
    ("Talentus", "Fungicida", "L", 17),
    ("Eterfhon", "Regulador", "L", 2.5),
    ("Nexus 50 SL", "Insumo", "L", 20),
    ("Perlan", "Regulador", "L", 0.5),
    ("Promalina", "Regulador", "L", 0.6),
    ("Esplendor 5%", "Regulador", "L", 1.25),   # Juan anotó "1.250 litros"
    ("Cytoplus", "Bioestimulante", "Kg", 1.3),
    ("Fontelis", "Fungicida", "L", 2),
    ("Comet", "Fungicida", "L", 0.2),
    ("Altivo", "Insumo", "L", 0.3),
    ("Agrocupper", "Fungicida", "Kg", 4.9),
    ("Elmuss", "Coadyuvante", "L", 3),
    ("Pomarsol Forte", "Fungicida", "Kg", 15),
    ("Silitec Poda", "Cicatrizante", "L", 10),
    ("Nordox Super 75 WP", "Fungicida", "Kg", 11.5),
    ("Ripper Full", "Herbicida", "L", 70),
    ("Reglone", "Herbicida", "L", 5),
    ("Valor 50 WP", "Herbicida", "Kg", 0.4),
    ("Aliado", "Herbicida", "g", 160),
    ("Abamectin 18 EC", "Insecticida", "L", 22),
    ("Avaunt 30 WG", "Insecticida", "L", 3),
    ("Stong", "Insumo", "L", 0.4),
    ("Closer", "Insecticida", "L", 1),
    ("Diazol", "Insecticida", "L", 5),
    ("Acetamiprid", "Insecticida", "L", 3),
    ("Bull", "Insumo", "L", 0.1),
    ("Selecron", "Insecticida", "L", 1),
    ("Succes 48", "Insecticida", "L", 0.1),
    ("Wispray", "Coadyuvante", "L", 190),
    ("Acomplish", "Bioestimulante", "L", 110),
    ("Bioadvance", "Bioestimulante", "L", 48),
    ("Podexal", "Cicatrizante", "L", 20),
    ("Stopit", "Fertilizante", "L", 80),
    ("Petróleo", "Combustible", "L", 1248),     # saldo reportado 10-jun
]

# Compras posteriores al conteo: (producto, cantidad, fecha, origen)
COMPRAS = [
    ("Nordox Super 75 WP", 25, "2026-07-10", "Factura Copeval"),
]

# Consumos posteriores al conteo: (producto, cantidad, fecha)
CONSUMOS = [
    ("Ripper Full", 20, "2026-06-12"),
    ("Ripper Full", 15, "2026-06-15"),
    ("Aliado", 0.48, "2026-06-16"),
    ("Ripper Full", 13, "2026-06-17"),
    ("Ripper Full", 5, "2026-06-17"),
    ("Nordox Super 75 WP", 5.4, "2026-06-17"),
    ("Nordox Super 75 WP", 3.6, "2026-06-17"),
    ("Nordox Super 75 WP", 2.5, "2026-06-17"),
    ("Ripper Full", 15, "2026-06-22"),
    ("Nordox Super 75 WP", 5.4, "2026-07-13"),
    ("Nordox Super 75 WP", 3.6, "2026-07-13"),
]

stock = {p: {"cat": c, "uni": u, "qty": float(q),
             "entrada": FECHA_CONTEO, "uso": ""} for p, c, u, q in CONTEO}

for prod, cant, fecha, _origen in COMPRAS:
    stock[prod]["qty"] += cant
    stock[prod]["entrada"] = fecha
for prod, cant, fecha in CONSUMOS:
    stock[prod]["qty"] -= cant
    if fecha > (stock[prod]["uso"] or ""):
        stock[prod]["uso"] = fecha

wb = load_workbook(EXCEL_PATH)

# Preservar la hoja vieja
if "Inventario ANTIGUO" in wb.sheetnames:
    del wb["Inventario ANTIGUO"]
if "Inventario" in wb.sheetnames:
    wb["Inventario"].title = "Inventario ANTIGUO"

ws = wb.create_sheet("Inventario")
HDR = ["Producto", "Categoría", "Unidad", "Stock Actual", "Stock Mínimo",
       "Última Entrada", "Último Uso"]
fill = PatternFill("solid", fgColor="2E7D32")
for i, h in enumerate(HDR, 1):
    c = ws.cell(1, i, h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = fill
    c.alignment = Alignment(horizontal="center")
for i, w in enumerate([28, 15, 8, 13, 13, 15, 13], 1):
    ws.column_dimensions[chr(64 + i)].width = w

for prod, d in sorted(stock.items()):
    qty = round(d["qty"], 3)
    ws.append([prod, d["cat"], d["uni"], qty, 0, d["entrada"], d["uso"]])

_save_wb(wb)
wb.close()

print(f"✅ Inventario reconstruido: {len(stock)} productos")
print("   (hoja anterior preservada como 'Inventario ANTIGUO')\n")
print("=== Productos con movimiento desde el conteo ===")
tocados = {p for p, *_ in COMPRAS} | {p for p, *_ in CONSUMOS}
for p in sorted(tocados):
    base = next(q for pr, _, _, q in CONTEO if pr == p)
    ent = sum(c for pr, c, _, _ in COMPRAS if pr == p)
    sal = sum(c for pr, c, _ in CONSUMOS if pr == p)
    print(f"  {p:24} {base:>7g} {stock[p]['uni']:<3} +{ent:g} −{sal:g}  →  "
          f"{round(stock[p]['qty'], 3):g} {stock[p]['uni']}")
print("\n=== Stock más bajo (revisar reposición) ===")
bajos = sorted(stock.items(), key=lambda x: x[1]["qty"])[:8]
for p, d in bajos:
    print(f"  {p:24} {round(d['qty'], 3):>8g} {d['uni']}")
