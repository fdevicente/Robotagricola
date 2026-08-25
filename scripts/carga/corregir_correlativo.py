#!/usr/bin/env python3
"""Deja el correlativo solo donde tiene sentido físico.

- Facturas que están en FXP → su número de FXP (ya asignado).
- Facturas recientes que aún no están en FXP → números nuevos desde max(FXP)+1,
  para poder archivarlas impresas (y que el usuario use el mismo N° en FXP).
- Facturas antiguas fuera de FXP → se deja EN BLANCO (nunca tuvieron número).
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from collections import defaultdict
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH
from excel_manager import _save_wb
from modules.correlativo import (
    leer_correlativos_fxp, norm_prov, nrokey, COL_CORRELATIVO,
)

DESDE = date(2026, 6, 1)     # lo subido en esta tanda


def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for f in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try: return datetime.strptime(v[:10], f).date()
            except Exception: pass
    return None


por_prov_nro, por_nro, max_fxp = leer_correlativos_fxp()
wb = load_workbook(EXCEL_PATH)
ws = wb["Facturas"]

grupos = defaultdict(list)
for r in range(2, ws.max_row + 1):
    if not ws.cell(r, 1).value:
        continue
    nro = nrokey(ws.cell(r, 7).value)
    if not nro:
        continue
    grupos[(norm_prov(ws.cell(r, 4).value), nro)].append(r)

en_fxp, recientes, borrados = 0, [], 0
sin_numero = []
for (prov, nro), filas in grupos.items():
    n = por_prov_nro.get((prov, nro)) or por_nro.get(nro)
    if n:
        for r in filas:
            ws.cell(r, COL_CORRELATIVO).value = n
        en_fxp += 1
        continue
    emi = _pd(ws.cell(filas[0], 1).value)
    if emi and emi >= DESDE:
        recientes.append((emi, filas, ws.cell(filas[0], 4).value,
                           ws.cell(filas[0], 7).value,
                           float(ws.cell(filas[0], 16).value or 0)))
    else:
        for r in filas:
            ws.cell(r, COL_CORRELATIVO).value = None
        borrados += 1
        sin_numero.append(prov)

# Numerar las recientes por fecha de emisión, continuando la serie de FXP
recientes.sort(key=lambda x: (x[0], min(x[1])))
siguiente = max_fxp
asignados = []
for emi, filas, prov, nro, total in recientes:
    siguiente += 1
    for r in filas:
        ws.cell(r, COL_CORRELATIVO).value = siguiente
    asignados.append((siguiente, emi, prov, nro, total))

_save_wb(wb)
wb.close()

print(f"✅ Correlativo desde FXP: {en_fxp} facturas")
print(f"✅ Recientes numeradas (desde {max_fxp+1}): {len(asignados)}")
print(f"➖ Antiguas fuera de FXP dejadas en blanco: {borrados}\n")
print("=== Números asignados a las facturas nuevas ===")
for n, emi, prov, nro, total in asignados:
    print(f"  N°{n:<5} {emi}  {str(prov)[:34]:34} F{str(nro):<12} ${total:>11,.0f}")
