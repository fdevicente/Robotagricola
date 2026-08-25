#!/usr/bin/env python3
"""Carga en la Bitácora todo lo que Juan reportó por Telegram (09-jun → 27-jul 2026).

Fuente: export del chat (messages.html). Escritura en UN solo batch.
Reemplaza las 3 filas auto-guardadas del 27-jul (quedaron con fecha del mensaje,
no del trabajo, y mal estructuradas).
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from collections import defaultdict
from datetime import date
from openpyxl import load_workbook
from config import EXCEL_PATH
from excel_manager import _save_wb

REGISTRADO_POR = "Juan Parada"

F, R, A, P, RP, J = ("Felicito Amigo", "Ramiro Amigo", "Agustin Mora",
                     "Patricio Mora", "Richard Padilla", "Javier Gonzalez")

# (fecha, fracción de jornada, {trabajador: actividad})
ASISTENCIA = [
    ("2026-06-08", 1.0, {F: "Sacar plantas cortadas", R: "Sacar riego nogales",
                         A: "Sacar plantas cortadas", P: "Sacar plantas cortadas",
                         RP: "Aplicación herbicida avellanos", J: "Cortar nogales"}),
    ("2026-06-09", 1.0, {F: "Sacar plantas cortadas", R: "Sacar riego",
                         A: "Sacar restos poda", P: "Sacar plantas cortadas",
                         RP: "Aplicación herbicida avellanos", J: "Cortar nogales"}),
    ("2026-06-10", 1.0, {F: "Mantención planta", R: "Aseo general",
                         A: "Mantención planta", P: "Mantención maquinaria",
                         RP: "Mantención maquinaria", J: "Mantención planta"}),
    ("2026-06-11", 1.0, {F: "Mantención general", R: "Aplicación herbicida",
                         A: "Sacar plantas cortadas", P: "Sacar plantas cortadas",
                         RP: "Aplicación herbicida", J: "Sacar plantas cortadas"}),
    ("2026-06-12", 1.0, {F: "Mantención general", A: "Mantención planta",
                         P: "Mantención maquinaria", R: "Mantención planta",
                         J: "Mantención general", RP: "Mantención planta"}),
    ("2026-06-15", 1.0, {F: "Mantención general", R: "Aplicación herbicida",
                         A: "Sacar plantas cortadas", P: "Aplicación herbicida",
                         RP: "Aplicación herbicida", J: "Mantención general"}),
    ("2026-06-16", 1.0, {F: "Poda nogales", R: "Aplicación herbicida",
                         A: "Pasar rastra", P: "Poda nogales",
                         RP: "Poda avellanos", J: "Aplicación herbicida"}),
    ("2026-06-17", 1.0, {F: "Poda nogales", R: "Pintar poda nogales",
                         A: "Aplicación cerezos", P: "Poda nogales",
                         RP: "Poda avellanos", J: "Aplicación herbicida cerezos"}),
    ("2026-06-18", 1.0, {F: "Poda nogales", R: "Pintar poda nogales",
                         A: "Pasar rastra", P: "Poda nogales",
                         RP: "Poda avellanos", J: "Aplicación herbicida"}),
    ("2026-06-19", 1.0, {F: "Poda nogales", R: "Pintar poda nogales",
                         A: "Pasar pala cola", P: "Poda nogales",
                         RP: "Poda avellanos", J: "Vacaciones"}),
    ("2026-06-22", 1.0, {F: "Poda nogales", R: "Pintar poda nogales",
                         A: "Pasar rastra", P: "Poda nogales",
                         RP: "Poda avellanos", J: "Aplicación herbicida"}),
    ("2026-06-23", 1.0, {F: "Poda nogales", R: "Pintar poda nogales",
                         A: "Pasar pala cola", P: "Poda nogales",
                         RP: "Poda avellanos", J: "Aplicación herbicida"}),
    ("2026-06-30", 1.0, {F: "Poda nogales", R: "Bajar ramas nogales",
                         A: "Pasar rastra", P: "Poda nogales",
                         RP: "Poda avellanos", J: "Pintar poda nogales"}),
    ("2026-07-01", 1.0, {F: "Poda nogales", R: "Bajar ramas nogales",
                         A: "Pasar pata fija", P: "Poda nogales",
                         RP: "Poda avellanos", J: "Pintar poda nogales"}),
    ("2026-07-02", 1.0, {F: "Poda nogales", R: "Bajar ramas nogales",
                         A: "Pasar rastra", P: "Poda nogales",
                         RP: "Poda avellanos", J: "Pintar poda nogales"}),
    ("2026-07-03", 1.0, {F: "Poda nogales", R: "Bajar ramas nogales",
                         A: "Pasar rastra", P: "Poda nogales",
                         RP: "Poda avellanos", J: "Pintar poda nogales"}),
    # 6-jul: dos medias jornadas
    ("2026-07-06", 0.5, {F: "Barbecho plantas", R: "Barbecho plantas",
                         A: "Barbecho plantas", P: "Barbecho plantas",
                         RP: "Barbecho plantas", J: "Barbecho plantas"}),
    ("2026-07-06", 0.5, {F: "Poda nogales", R: "Bajar ramas nogales",
                         A: "Poda nogales", P: "Trabajos varios",
                         RP: "Poda avellanos", J: "Pintar poda nogales"}),
    ("2026-07-07", 1.0, {F: "Poda nogales", R: "Bajar ramas nogales",
                         A: "Pasar pala niveladora", P: "Poda nogales",
                         RP: "Poda avellanos", J: "Pintar poda nogales"}),
    ("2026-07-08", 1.0, {F: "Poda nogales", R: "Bajar ramas nogales",
                         A: "Pasar pala cola", P: "Poda nogales",
                         RP: "Poda avellanos", J: "Pintar poda nogales"}),
    ("2026-07-09", 1.0, {F: "Mantención riego", R: "Aseo general",
                         A: "Mantención maquinaria", P: "Mantención maquinaria",
                         RP: "Mantención maquinaria", J: "Aseo general"}),
    ("2026-07-10", 1.0, {F: "Mantención riego", R: "Aseo general",
                         A: "Mantención maquinaria", P: "Mantención maquinaria",
                         RP: "Mantención maquinaria", J: "Aseo general"}),
    ("2026-07-13", 1.0, {F: "Poda nogales", R: "Bajar ramas",
                         A: "Aplicación cerezos", P: "Poda nogales",
                         RP: "Poda avellanos", J: "Pintar poda nogales"}),
    ("2026-07-14", 1.0, {F: "Poda nogales", R: "Bajar ramas",
                         A: "Mantención caminos", P: "Poda nogales",
                         RP: "Poda avellanos", J: "Pintar poda nogales"}),
    ("2026-07-15", 1.0, {F: "Poda nogales", R: "Bajar ramas",
                         A: "Poda avellanos",
                         RP: "Poda avellanos", J: "Pintar poda nogales"}),
    ("2026-07-17", 1.0, {F: "Vacaciones", R: "Vacaciones", A: "Vacaciones",
                         RP: "Desaguar cerezos", J: "Vacaciones"}),
    ("2026-07-20", 1.0, {F: "Mantención riego", R: "Mantención riego",
                         P: "Mantención maquinaria", RP: "Mantención maquinaria",
                         J: "Mantención riego", A: "Mantención maquinaria"}),
    ("2026-07-21", 1.0, {F: "Mantención riego", R: "Mantención riego",
                         P: "Mantención maquinaria", RP: "Mantención maquinaria",
                         J: "Mantención riego"}),
    ("2026-07-22", 1.0, {F: "Desaguar avellanos", R: "Desaguar avellanos",
                         P: "Desaguar avellanos", J: "Desaguar avellanos",
                         RP: "Desaguar avellanos", A: "Desaguar avellanos"}),
    ("2026-07-23", 1.0, {F: "Mantención riego", R: "Mantención riego",
                         P: "Mantención maquinaria", J: "Mantención riego",
                         RP: "Mantención maquinaria", A: "Mantención maquinaria"}),
    ("2026-07-24", 1.0, {F: "Desaguar", R: "Desaguar", P: "Desaguar",
                         J: "Desaguar", RP: "Desaguar", A: "Mantención maquinaria"}),
    ("2026-07-27", 1.0, {F: "Mantención general", R: "Mantención general",
                         P: "Mantención maquinaria", J: "Ausente",
                         RP: "Desaguar", A: "Desaguar"}),
]

# (fecha, máquina, odo_inicio, odo_fin, labor)   odo None = horómetro malo
MAQUINARIA = [
    ("2026-06-08", "TRACTOR MASSEY FERGUSON 4275", 3415, 3416, "Trabajos de campo"),
    ("2026-06-08", "TRACTOR JOHN DEERE 5085", 3150, 3154, "Trabajos de campo"),
    ("2026-06-08", "TRACTOR JOHN DEERE 5425", None, None, "Horómetro en mal estado"),
    ("2026-06-08", "TRACTOR MASSEY FERGUSON 6711", 1950, 1950, "Sin uso"),
    ("2026-06-08", "TRACTOR MASSEY FERGUSON 4292", 5222, 5222, "Sin uso"),
    ("2026-06-09", "TRACTOR MASSEY FERGUSON 4275", 3416, 3416, "Sin uso"),
    ("2026-06-09", "TRACTOR JOHN DEERE 5085", 3154, 3158, "Trabajos de campo"),
    ("2026-06-09", "TRACTOR JOHN DEERE 5425", None, None, "Horómetro en mal estado"),
    ("2026-06-09", "TRACTOR MASSEY FERGUSON 6711", 1950, 1958, "Trabajos de campo"),
    ("2026-06-09", "TRACTOR MASSEY FERGUSON 4292", 5222, 5222, "Sin uso"),
    ("2026-06-11", "TRACTOR MASSEY FERGUSON 6711", 1958, 1964, "Trabajos de campo"),
    ("2026-06-11", "TRACTOR JOHN DEERE 5085", 3158, 3166, "Trabajos de campo"),
    ("2026-06-11", "TRACTOR MASSEY FERGUSON 4275", 3416, 3417, "Trabajos de campo"),
    ("2026-06-11", "TRACTOR JOHN DEERE 5425", None, None, "Horómetro malo (~5 h aprox)"),
    # Excavadora / retroexcavadora — destronque de nogales
    ("2026-06-08", "EXCAVADORA", 7205.9, 7214.3, "Destroncar nogales"),
    ("2026-06-09", "EXCAVADORA", 7214.3, 7223.3, "Destroncar nogales"),
    ("2026-06-10", "EXCAVADORA", 7223.3, 7232.4, "Destroncar nogales"),
    ("2026-06-11", "EXCAVADORA", 7232.4, 7240.7, "Destroncar nogales"),
]

# (fecha, tipo, producto, cantidad, unidad, cultivo, sector, ha, detalle)
APLICACIONES = [
    ("2026-06-12", "Aplicación herbicida", "Ripper Full", 20, "L", "AVELLANOS", "", None,
     "Dosis 2.5 L/100 L agua; mojamiento 100 L/ha; total 800 L caldo"),
    ("2026-06-15", "Aplicación herbicida", "Ripper Full", 15, "L", "AVELLANOS", "", None,
     "Dosis 2.5 L/100; total 600 L caldo"),
    ("2026-06-16", "Aplicación herbicida", "Aliado", 0.48, "g", "GENERAL", "Canales y cercos", None,
     "Dosis 0.08 g/100 L; total 600 L caldo (Juan anotó 0.048 g — revisar)"),
    ("2026-06-17", "Aplicación herbicida", "Ripper Full", 13, "L", "CEREZOS", "Producción y formación", None,
     "Dosis 2.5 L/100; total 520 L caldo"),
    ("2026-06-17", "Aplicación herbicida", "Ripper Full", 5, "L", "GENERAL", "CSSA y orilleros", None,
     "Dosis 1.5 L/100; total 333 L caldo"),
    ("2026-06-17", "Aplicación fungicida", "Nordox Super 75 WP", 5.4, "kg", "CEREZOS", "Producción", 2,
     "Dosis 180 g/100; mojamiento 1500 L/ha; total 3000 L caldo"),
    ("2026-06-17", "Aplicación fungicida", "Nordox Super 75 WP", 3.6, "kg", "CEREZOS", "Formación", 2,
     "Dosis 180 g/100; mojamiento 1000 L/ha; total 2000 L caldo"),
    ("2026-06-17", "Aplicación fungicida", "Nordox Super 75 WP", 2.5, "kg", "AVELLANOS", "Formación", 3.4,
     "Dosis 150 g/100; mojamiento 500 L/ha; total 1666 L caldo"),
    ("2026-06-22", "Aplicación herbicida", "Ripper Full", 15, "L", "NOGALES", "", None,
     "Dosis 2.5 L/100; mojamiento 100 L/ha; total 600 L caldo"),
    ("2026-07-13", "Aplicación fungicida", "Nordox Super 75 WP", 5.4, "kg", "CEREZOS", "Producción", 2,
     "Dosis 180 g/100; mojamiento 1500 L/ha; total 3000 L caldo"),
    ("2026-07-13", "Aplicación fungicida", "Nordox Super 75 WP", 3.6, "kg", "CEREZOS", "Formación", 2,
     "Dosis 180 g/100; mojamiento 1000 L/ha; total 2000 L caldo"),
]

# (fecha, máquina, litros)  — saldo reportado 1248 L al 10-jun
COMBUSTIBLE = [
    ("2026-06-02", "SSANGYONG 2", 69), ("2026-06-02", "TRACTOR JOHN DEERE 5085", 30),
    ("2026-06-02", "TRACTOR JOHN DEERE 5425", 40), ("2026-06-02", "TRACTOR MASSEY FERGUSON 6711", 60),
    ("2026-06-02", "SSANGYONG 1", 69), ("2026-06-02", "TRACTOR MASSEY FERGUSON 4275", 50),
    ("2026-06-07", "SSANGYONG 2", 55),
    ("2026-06-08", "TRACTOR JOHN DEERE 5085", 31), ("2026-06-08", "TRACTOR MASSEY FERGUSON 6711", 50),
    ("2026-06-08", "TRACTOR MASSEY FERGUSON 4275", 20),
    ("2026-06-09", "TRACTOR JOHN DEERE 5425", 40), ("2026-06-09", "SSANGYONG 2", 47),
]


def cultivo_de(act: str) -> str:
    a = act.lower()
    if "nogal" in a:
        return "NOGALES"
    if "avellano" in a or "barbecho" in a:
        return "AVELLANOS"
    if "cerezo" in a:
        return "CEREZOS"
    return "GENERAL"


def tipo_de(act: str) -> str:
    a = act.lower()
    if "aplicación" in a or "aplicacion" in a:
        return "APLICACION"
    if "vacacion" in a or "ausente" in a:
        return "OTRO"
    if "riego" in a or "desagu" in a:
        return "RIEGO"
    return "LABOR"


filas = []

# ── Asistencia → una fila por (día, actividad) ──
for fecha, frac, asig in ASISTENCIA:
    por_act = defaultdict(list)
    for trab, act in asig.items():
        por_act[act].append(trab)
    for act, trabs in sorted(por_act.items()):
        no_trabajada = ("vacacion" in act.lower() or "ausente" in act.lower())
        jh = None if no_trabajada else round(len(trabs) * frac, 2)
        sufijo = " (media jornada)" if frac == 0.5 else ""
        filas.append([
            fecha, "08:00", tipo_de(act), act, cultivo_de(act), "",
            jh, ", ".join(sorted(trabs)), "", None, "",
            f"Asistencia {fecha}{sufijo} — reportado por Juan", REGISTRADO_POR,
            "", None, None, None,
        ])

# ── Maquinaria ──
for fecha, maq, ini, fin, labor in MAQUINARIA:
    horas = round(fin - ini, 1) if (ini is not None and fin is not None) else None
    detalle = (f"Horómetro {ini} → {fin}" if ini is not None else "Horómetro en mal estado")
    filas.append([
        fecha, "08:00", "MAQUINARIA", labor, cultivo_de(labor), "",
        None, "", "", None, "",
        f"{detalle} — reportado por Juan", REGISTRADO_POR,
        maq, fin, horas, None,
    ])

# ── Aplicaciones ──
for fecha, tipo_app, prod, cant, uni, cult, sector, ha, detalle in APLICACIONES:
    filas.append([
        fecha, "08:00", "APLICACION", tipo_app, cult, sector,
        None, "", prod, cant, uni,
        f"{detalle} — reportado por Juan", REGISTRADO_POR,
        "", None, None, ha,
    ])

# ── Combustible ──
for fecha, maq, litros in COMBUSTIBLE:
    filas.append([
        fecha, "08:00", "MAQUINARIA", "Carga de combustible", "GENERAL", "",
        None, "", "Petróleo", litros, "L",
        f"Carga {litros} L a {maq} — reportado por Juan", REGISTRADO_POR,
        maq, None, None, None,
    ])

filas.sort(key=lambda r: (r[0], r[2]))

# ── Escritura en UN batch ──
wb = load_workbook(EXCEL_PATH)
ws = wb["Bitácora"]

# Quitar las 3 filas auto-guardadas del 27-jul (fecha del mensaje ≠ fecha del trabajo)
a_borrar = []
for r in range(2, ws.max_row + 1):
    fecha = str(ws.cell(r, 1).value or "")[:10]
    reg = str(ws.cell(r, 12).value or "")
    if fecha == "2026-07-27" and ("Asistencia 23 julio" in reg or
                                   "24 de julio 2026" in reg or
                                   "Lunes 27 de julio 2026" in reg):
        a_borrar.append(r)
for r in reversed(a_borrar):
    ws.delete_rows(r)
print(f"Filas auto-guardadas reemplazadas: {len(a_borrar)}")

fila = ws.max_row
while fila > 1 and not ws.cell(fila, 1).value:
    fila -= 1
for f in filas:
    fila += 1
    for c, val in enumerate(f, 1):
        ws.cell(fila, c).value = val

_save_wb(wb)
wb.close()

print(f"\n✅ Bitácora cargada: {len(filas)} registros")
tipos = defaultdict(int)
for f in filas:
    tipos[f[2]] += 1
for t, n in sorted(tipos.items(), key=lambda x: -x[1]):
    print(f"   {n:3} | {t}")
jh_total = sum(f[6] for f in filas if isinstance(f[6], (int, float)))
print(f"\n   Jornadas-hombre totales: {jh_total:g}")
print(f"   Rango: {filas[0][0]} → {filas[-1][0]}")
