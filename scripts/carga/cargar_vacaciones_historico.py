"""Carga el histórico COMPLETO de vacaciones desde el Excel de Juan.

Fuente: CAMARICO 2023 / "ASISTENCIA TEMP 2023-2024 fda JUAN PARADA (Copia en
conflicto...)" pestaña `VACAC 26` — el archivo más nuevo (6-ago-2026).

Antes solo se habían cargado las de 2026 (venían de una foto), así que los
saldos estaban sobreestimados en ~101 días en total.

Ojo con el archivo:
  · varias filas tienen el AÑO mal tipeado (2026 dentro del bloque 2024).
    Se corrige tomando el año del encabezado de bloque ("AÑO 24").
  · en 2023 el nombre de Ramiro está invertido ("RAMIRO AMIGO").
  · la tabla resumen de la derecha es la que mantiene Juan: se usa para
    VALIDAR que el detalle cargado calce.

Uso:  python scripts/carga/cargar_vacaciones_historico.py [--aplicar]
"""
import re
import shutil
import sys
from collections import defaultdict
from datetime import date, datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb

APLICAR = "--aplicar" in sys.argv
HOY = date.today()
DIAS_MES = 15.0 / 12.0

ORIGEN = (r"C:\Users\Windows\Dropbox\CAMARICO 2023"
          r"\ASISTENCIA TEMP 2023-2024 fda JUAN PARADA "
          r"(Copia en conflicto de juan parada 2025-03-06).xlsx")

# Nombres del archivo de Juan → nombres del Master
CANON = {
    "MORA PATRICIO": "Luis Patricio Mora Amigo",
    "PARADA JUAN": "Juan Parada Castillo",
    "AMIGO RAMIRO": "Luis Ramiro Amigo Soto",
    "RAMIRO AMIGO": "Luis Ramiro Amigo Soto",      # invertido en 2023
    "AMIGO FELICITO": "Felicito Amigo Soto",
    "FELICITO AMIGO": "Felicito Amigo Soto",
    "MORA AGUSTIN": "Agustin Segundo Mora Hernandez",
    "AGUSTIN MORA": "Agustin Segundo Mora Hernandez",
    "GONZALES JAVIER": "Javier Gonzalez",
    "GONZALEZ JAVIER": "Javier Gonzalez",
}


def norm(n):
    return " ".join(str(n or "").upper().split())


def _corregir_anio(f, anio_bloque):
    """El año del bloque manda: en el archivo hay varios mal tipeados."""
    if not hasattr(f, "year"):
        return None, False
    f = f.date() if isinstance(f, datetime) else f
    if f.year == anio_bloque:
        return f, False
    try:
        return f.replace(year=anio_bloque), True
    except ValueError:                       # 29-feb en año no bisiesto
        return f, False


# ── Leer el origen ──
wb = load_workbook(ORIGEN, read_only=True, data_only=True)
filas = list(wb["VACAC 26"].iter_rows(values_only=True))
wb.close()

registros, corregidas = [], 0
resumen, contratos = {}, {}
anio_bloque = None
for row in filas:
    if not row:
        continue
    # OJO: la fila del encabezado de bloque ("AÑO 24") también trae datos del
    # resumen a la derecha. No se puede saltar entera.
    m = re.match(r"AÑO\s*(\d{2})", str(row[0] or "").strip().upper())
    if m:
        anio_bloque = 2000 + int(m.group(1))

    # Tabla resumen de la derecha
    if len(row) >= 17 and norm(row[10]) and norm(row[10]) != "NOMBRES":
        n = CANON.get(norm(row[10]))
        if n:
            try:
                resumen[n] = float(row[15] or 0)
                contratos[n] = row[16]
            except (TypeError, ValueError):
                pass

    # Detalle (la fila de encabezado no trae vacaciones en las cols 1-6)
    if m:
        continue
    nombre = CANON.get(norm(row[2])) if len(row) > 2 else None
    if not nombre or anio_bloque is None:
        continue
    try:
        dias = float(row[5] or 0)
    except (TypeError, ValueError):
        continue
    if dias <= 0:
        continue
    desde, c1 = _corregir_anio(row[3], anio_bloque)
    hasta, c2 = _corregir_anio(row[4], anio_bloque)
    corregidas += int(c1 or c2)
    nota = str(row[6] or "").strip()
    registros.append((nombre, desde, hasta, dias, anio_bloque, nota))

# ── Validar contra el resumen de Juan ──
suma = defaultdict(float)
for n, _d, _h, dias, _a, _nt in registros:
    suma[n] += dias

print("=" * 74)
print("VALIDACIÓN — detalle cargado vs la tabla resumen de Juan")
print("=" * 74)
ok_todo = True
for n in sorted(resumen):
    calc, decl = suma.get(n, 0), resumen[n]
    ok = abs(calc - decl) < 0.01
    ok_todo &= ok
    print(f"  {'✅' if ok else '❌'} {n[:34]:34} cargado={calc:>5.0f}  "
          f"resumen={decl:>5.0f}")
if not ok_todo:
    print("\n⚠️ No calza — no se escribe nada.")
    sys.exit(1)
print(f"\n  {len(registros)} registros · {corregidas} con el año corregido "
      f"según su bloque")

# ── Saldos ──
wb = load_workbook(EXCEL_PATH, data_only=True)
base = {}
for row in wb["Vacaciones Pendientes"].iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    n = str(row[0]).strip()
    try:
        saldo = float(row[3] or 0)
    except (TypeError, ValueError):
        saldo = 0.0
    f = row[4]
    if isinstance(f, datetime):
        f = f.date()
    base[n] = {"saldo": saldo, "fecha": f}
personal_actual = {}
for row in wb["Personal"].iter_rows(min_row=2, values_only=True):
    if row and row[0]:
        personal_actual[str(row[0]).strip()] = row[4]
wb.close()

print("\n" + "=" * 74)
print("SALDOS RECALCULADOS")
print("=" * 74)
print(f"{'trabajador':32}{'base':>7}{'+acum':>8}{'−tom':>6}{'=saldo':>9}"
      f"{'antes':>9}{'dif':>8}")
print("-" * 74)
nuevos = {}
for n in sorted(set(list(base) + list(suma))):
    b = base.get(n, {})
    f0 = b.get("fecha")
    if not f0:
        continue
    meses = (HOY.year - f0.year) * 12 + (HOY.month - f0.month)
    acum = meses * DIAS_MES
    tom = suma.get(n, 0)
    total = round(b.get("saldo", 0) + acum - tom, 2)
    nuevos[n] = (total, tom)
    antes = personal_actual.get(n)
    dif = (total - float(antes)) if isinstance(antes, (int, float)) else 0
    print(f"{n[:32]:32}{b.get('saldo', 0):>7.2f}{acum:>8.2f}{tom:>6.0f}"
          f"{total:>9.2f}{(antes if antes is not None else 0):>9}{dif:>8.2f}")

if not APLICAR:
    print("\n(simulación — nada se escribió; agrega --aplicar)")
    sys.exit(0)

# ── Escribir ──
resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"\nRespaldo: {resp}")

wb = load_workbook(EXCEL_PATH)
ws = wb["Vacaciones"]
ws.delete_rows(2, max(ws.max_row, 2))
for n, desde, hasta, dias, anio, nota in sorted(registros, key=lambda x: (x[1] or date.min)):
    ws.append([n, desde, hasta, dias, "Aprobado",
               nota or f"Periodo {anio}"])
print(f"✅ Hoja Vacaciones: {len(registros)} registros ({sorted({r[4] for r in registros})})")

wsp = wb["Personal"]
for r in range(2, wsp.max_row + 1):
    n = str(wsp.cell(r, 1).value or "").strip()
    if n not in nuevos:
        continue
    total, tom = nuevos[n]
    wsp.cell(r, 5).value = total
    wsp.cell(r, 6).value = tom
    ult = max((x[2] for x in registros if x[0] == n and x[2]), default=None)
    if ult:
        wsp.cell(r, 7).value = ult
    wsp.cell(r, 8).value = (f"Histórico completo 2023-2026 desde el Excel de "
                             f"Juan (VACAC 26) · {tom:.0f} días tomados")
    # Fecha de contrato según el archivo de Juan
    c = contratos.get(n)
    if isinstance(c, datetime):
        c = c.date()
    if isinstance(c, date):
        wsp.cell(r, 4).value = c
_save_wb(wb)
wb.close()
print("✅ Hoja Personal actualizada (saldos, tomados y fecha de contrato).")
