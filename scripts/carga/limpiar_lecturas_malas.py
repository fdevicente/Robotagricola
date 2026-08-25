"""Borra de la bitácora las lecturas de horómetro imposibles del 10-ago-2026.

  · fila 230: "Tractor jhon deere 50853200" — Juan escribió el modelo y el
    horómetro pegados. Quedó odómetro 50.853.200 y 50.850.034 "horas día".
  · fila 231: MF 4292 con 3.200 cuando venía en 5.222 (el horómetro bajó).
  · filas 227-228: mensajes sueltos de Juan ("Bitácora", "Maquinaria") que se
    guardaron como registros vacíos.

Se borran de abajo hacia arriba para no correr los números de fila.

Uso:  python scripts/carga/limpiar_lecturas_malas.py [--aplicar]
"""
import shutil
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb

APLICAR = "--aplicar" in sys.argv

# fila → (qué esperamos encontrar para no borrar la equivocada, motivo)
BORRAR = {
    231: ("TRACTOR MASSEY FERGUSON 4292", "el horómetro bajó de 5.222 a 3.200"),
    230: ("TRACTOR JOHN DEERE 5085", "odómetro 50.853.200 (modelo y horómetro pegados)"),
    228: (None, "mensaje suelto «Maquinaria», sin contenido"),
    227: (None, "mensaje suelto «Bitácora», sin contenido"),
}

wb = load_workbook(EXCEL_PATH)
ws = wb["Bitácora"]
COL_MAQ, COL_ODO, COL_ACT = 14, 15, 4

print("Filas a borrar:\n")
ok = True
for fila in sorted(BORRAR, reverse=True):
    esperado, motivo = BORRAR[fila]
    maq = str(ws.cell(fila, COL_MAQ).value or "").strip()
    act = str(ws.cell(fila, COL_ACT).value or "").strip()
    odo = ws.cell(fila, COL_ODO).value
    if esperado and maq != esperado:
        print(f"  ❌ fila {fila}: esperaba «{esperado}» y encontré «{maq}» — NO se toca")
        ok = False
        continue
    print(f"  fila {fila} | {act[:26]:26} | {maq[:30]:30} | odo={odo}")
    print(f"           └─ {motivo}")

if not ok:
    print("\n⚠️ Algo no calza. No se borra nada.")
    sys.exit(1)

if not APLICAR:
    wb.close()
    print("\n(simulación — nada se borró; agrega --aplicar)")
    sys.exit(0)

resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"\nRespaldo: {resp}")

for fila in sorted(BORRAR, reverse=True):
    ws.delete_rows(fila)
_save_wb(wb)
wb.close()
print(f"✅ {len(BORRAR)} filas borradas.")
