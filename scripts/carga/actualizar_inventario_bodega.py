#!/usr/bin/env python3
"""Actualiza Inventario + Vencimientos con la planilla de bodega (28-jul-2026).

Fuente: planilla Excel de bodega que envía el equipo (stock + fecha vencimiento
+ categoría real). Reemplaza el stock cargado desde el conteo de Juan y llena
la hoja Vencimientos, que hasta ahora estaba vacía.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import calendar
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import EXCEL_PATH
from excel_manager import _save_wb

HOY = date(2026, 7, 28)

# (producto, categoría, unidad, stock, vencimiento "mmm-aa" o "")
BODEGA = [
    ("Espirodiclofen 240 SC", "Acaricida", "L", 0.5, "oct-24"),
    ("Konan", "Acaricida", "L", 0.1, "sept-23"),
    ("Defender Zn", "Fertilizante foliar", "L", 9, "feb-27"),
    ("Defender Calcio", "Fertilizante foliar", "L", 10, "may-30"),
    ("Defender K (Potasio)", "Fertilizante foliar", "L", 12, "dic-28"),
    ("Manvert Avanza", "Fertilizante foliar", "L", 3, "oct-27"),
    ("Defender Boro", "Fertilizante foliar", "L", 2, "mar-27"),
    ("Black K", "Fertilizante foliar", "L", 17, "may-23"),
    ("Cito Four", "Fertilizante foliar", "L", 0.5, "mar-26"),
    ("Macroquel Magnesio", "Fertilizante foliar", "L", 16, "sept-24"),
    ("Petróleo Diesel", "Combustibles y lubricantes", "L", 0, "jun-26"),
    ("Bioestabilizado", "Enmiendas", "Kg", 20000, "jun-26"),
    ("Nitrato de Potasio", "Fertilizante", "Kg", 25, ""),
    ("Urea", "Fertilizante", "Kg", 75, ""),
    ("Pro Cal", "Fertilizante", "L", 1, ""),            # S/FECHA
    ("Talentus", "Fertilizante", "L", 17, "jul-19"),
    ("Eterfon 500", "Fitohormona", "L", 2.5, "ago-27"),
    ("Nexus 50 SL", "Fitohormona", "L", 20, "ene-15"),
    ("Perlan", "Fitohormona", "L", 0.5, "oct-24"),
    ("Promalina", "Fitohormona", "L", 0.6, "feb-26"),
    ("Splendor 5% SC", "Fitohormona", "L", 1250, "jul-23"),   # ⚠ unidad a confirmar
    ("Cytoplus", "Fungicida", "Kg", 1.3, "abr-25"),
    ("Fontelis", "Fungicida", "L", 2, "may-27"),
    ("Comet", "Fungicida", "L", 0.2, "ene-27"),
    ("Altivo", "Fungicida", "L", 0.3, "abr-27"),
    ("Agrocupper SP", "Fungicida", "Kg", 4.9, "sept-28"),
    ("Elmuss", "Fungicida", "L", 3, "sept-27"),
    ("Pomarsol Forte", "Fungicida", "Kg", 15, ""),      # S/F
    ("Silitec Poda", "Fungicida", "L", 10, "dic-30"),
    ("Nordox Super 75 WP", "Fungicida", "Kg", 16, "ago-26"),
    ("Ripper Full", "Herbicida", "L", 0, "feb-28"),
    ("Reglone", "Herbicida", "L", 0, "ago-28"),
    ("Valor 50 WP", "Herbicida", "Kg", 0.4, "dic-27"),
    ("Aliado", "Herbicida", "g", 0, "may-27"),
    ("Abamectin 18 EC", "Insecticida", "L", 22, "nov-27"),
    ("Avaunt 30 WG", "Insecticida", "Kg", 3, "jun-17"),
    ("Stong", "Insecticida", "L", 0.4, "jun-27"),
    ("Closer", "Insecticida", "L", 1, "nov-18"),
    ("Diazol 50 EW", "Insecticida", "L", 5, "feb-24"),
    ("Acetamiprid 20%", "Insecticida", "L", 3, "ago-27"),
    ("Bull", "Insecticida", "L", 0.1, "jun-26"),
    ("Selecron 720 EC", "Insecticida", "L", 1, "feb-21"),
    ("Success 48", "Insecticida", "L", 0.1, "jul-25"),
    ("Winspray", "Insecticida", "L", 190, "jul-24"),
    ("Acomplish", "Otros", "L", 110, "feb-19"),
    ("Bioadvance", "Otros", "L", 48, "abr-23"),
    ("Podexal", "Otros", "L", 20, "may-17"),
    ("Stopit", "Otros", "L", 80, "sept-26"),
]

MESES = {"ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
         "jul": 7, "ago": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dic": 12}


def parse_venc(s):
    """'oct-24' → date(2024,10,31) (último día del mes)."""
    if not s:
        return None
    try:
        mes_txt, anio_txt = s.strip().lower().split("-")
        mes = MESES.get(mes_txt)
        if not mes:
            return None
        anio = 2000 + int(anio_txt)
        return date(anio, mes, calendar.monthrange(anio, mes)[1])
    except Exception:
        return None


def estado_de(venc):
    if venc is None:
        return "SIN FECHA"
    dias = (venc - HOY).days
    if dias < 0:
        return "VENCIDO"
    if dias <= 90:
        return "POR VENCER"
    return "VIGENTE"


# ── Stock previo (para reportar diferencias) ──
wb = load_workbook(EXCEL_PATH)
previo = {}
if "Inventario" in wb.sheetnames:
    for row in wb["Inventario"].iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            try:
                previo[str(row[0]).strip().lower()] = float(row[3] or 0)
            except (TypeError, ValueError):
                pass
    del wb["Inventario"]

ws = wb.create_sheet("Inventario")
HDR = ["Producto", "Categoría", "Unidad", "Stock Actual", "Stock Mínimo",
       "Última Entrada", "Último Uso", "Vencimiento", "Estado"]
fill = PatternFill("solid", fgColor="2E7D32")
for i, h in enumerate(HDR, 1):
    c = ws.cell(1, i, h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = fill
    c.alignment = Alignment(horizontal="center")
for i, w in enumerate([26, 24, 8, 13, 13, 15, 13, 13, 13], 1):
    ws.column_dimensions[chr(64 + i)].width = w

ROJO = PatternFill("solid", fgColor="FFC7CE")
NARANJA = PatternFill("solid", fgColor="FFEB9C")

vencidos, por_vencer, difs = [], [], []
for prod, cat, uni, stock, venc_txt in BODEGA:
    venc = parse_venc(venc_txt)
    est = estado_de(venc)
    fila = [prod, cat, uni, stock, 0, "2026-07-28", "",
            venc.isoformat() if venc else "", est]
    ws.append(fila)
    r = ws.max_row
    if est == "VENCIDO":
        for c in range(1, 10):
            ws.cell(r, c).fill = ROJO
        vencidos.append((prod, venc, stock, uni))
    elif est == "POR VENCER":
        for c in range(1, 10):
            ws.cell(r, c).fill = NARANJA
        por_vencer.append((prod, venc, stock, uni))
    ant = previo.get(prod.strip().lower())
    if ant is not None and abs(ant - stock) > 0.001:
        difs.append((prod, ant, stock, uni))

# ── Hoja Vencimientos ──
if "Vencimientos" in wb.sheetnames:
    wsv = wb["Vencimientos"]
    wsv.delete_rows(2, wsv.max_row)
    for prod, cat, uni, stock, venc_txt in BODEGA:
        venc = parse_venc(venc_txt)
        if not venc:
            continue
        wsv.append([prod, "", "", "", venc.isoformat(), "", "",
                    estado_de(venc), HOY.isoformat()])

_save_wb(wb)
wb.close()

print(f"✅ Inventario actualizado: {len(BODEGA)} productos (con vencimientos)\n")
print(f"🔴 VENCIDOS: {len(vencidos)}")
for p, v, s, u in sorted(vencidos, key=lambda x: x[1]):
    print(f"   {v:%b-%Y}  {p:26} {s:>8g} {u}")
print(f"\n🟠 POR VENCER (≤90 días): {len(por_vencer)}")
for p, v, s, u in sorted(por_vencer, key=lambda x: x[1]):
    print(f"   {v:%b-%Y}  {p:26} {s:>8g} {u}")
print(f"\n📊 Diferencias vs stock anterior: {len(difs)}")
for p, ant, nue, u in difs:
    print(f"   {p:26} {ant:>9g} → {nue:g} {u}")
