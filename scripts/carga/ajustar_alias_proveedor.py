#!/usr/bin/env python3
"""Ajusta los últimos casos: mismo proveedor escrito distinto en Master y FXP.

Verificados a mano (Irrifor=Irrifer, Coragas=Cora, etc.). El resto de los
"proveedor distinto" son calces falsos por número y NO se tocan.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook
from config import EXCEL_PATH
from excel_manager import _save_wb
from modules.correlativo import nrokey

# (nro documento, proveedor esperado en Master, monto correcto de FXP)
CASOS = [
    ("27987", "irrifor",     171360),
    ("16466", "coragas",     50800),
    ("17046", "coragas",     350000),
    ("263",   "gruasmestre", 285600),
]


def _f(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


wb = load_workbook(EXCEL_PATH)
ws = wb["Facturas"]
hechos = []
for nro, prov_esp, ref in CASOS:
    filas = [r for r in range(2, ws.max_row + 1)
             if nrokey(ws.cell(r, 7).value) == nrokey(nro)
             and prov_esp in str(ws.cell(r, 4).value or "").lower()]
    if not filas:
        print(f"  ⚠️ no encontré F{nro} de {prov_esp}")
        continue
    suma = sum(_f(ws.cell(r, 15).value) for r in filas)
    antes = _f(ws.cell(filas[0], 16).value)
    if suma <= 0:
        for r in filas:
            ws.cell(r, 16).value = round(ref)
        hechos.append((prov_esp, nro, antes, ref, 0))
        continue
    factor = ref / suma
    for r in filas:
        for col in (10, 13, 15):
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                ws.cell(r, col).value = round(v * factor, 2)
        ws.cell(r, 16).value = round(ref)
    hechos.append((str(ws.cell(filas[0], 4).value or ""), nro, antes, ref, factor))

_save_wb(wb)
wb.close()
for prov, nro, antes, ref, f in hechos:
    print(f"  ✔ {prov[:30]:30} F{nro:<9} ${antes:>11,.0f} → ${ref:>11,.0f}  (x{f:.3f})")
print(f"\n✅ {len(hechos)} facturas ajustadas.")
