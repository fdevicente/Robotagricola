#!/usr/bin/env python3
"""Genera el Excel con las diferencias Master vs FXP (solo lectura)."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import os, shutil, tempfile
from collections import defaultdict
from datetime import date, datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import EXCEL_PATH
from modules.correlativo import (norm_prov, nrokey, COL_CORRELATIVO, FXP_PATH,
                                 buscar_en_fxp)

SALIDA = r"C:\Users\Windows\Desktop\Workflow\Agricola Santa Elisa\Diferencias_Master_vs_FXP.xlsx"


def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for f in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try: return datetime.strptime(v[:10], f).date()
            except Exception: pass
    return None


def _f(v):
    try: return float(v or 0)
    except (TypeError, ValueError): return 0.0


tmp = os.path.join(tempfile.gettempdir(), "fxp_list.xlsx")
shutil.copy2(FXP_PATH, tmp)
wbf = load_workbook(tmp, read_only=True, data_only=True)
fxp = {}
for row in wbf["FXP"].iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    nro = nrokey(row[7])
    if not nro:
        continue
    m = row[8]
    if isinstance(m, str):
        s = m.upper().replace("USD", "").replace("$", "").strip().replace(".", "").replace(",", ".")
        try: m = float(s)
        except ValueError: m = 0
    fxp[(norm_prov(row[6]), nro)] = {
        "n": int(row[3]) if isinstance(row[3], (int, float)) else None,
        "prov": str(row[6] or ""), "nro": str(row[7] or ""), "monto": _f(m),
        "emision": _pd(row[0]), "venc": _pd(row[1]), "pago": _pd(row[2]),
        "estado": str(row[11] or "").strip(), "nota": str(row[12] or "")}
wbf.close()
fxp_nro = {}
for (p, n), d in fxp.items():
    fxp_nro.setdefault(n, []).append(d)

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Facturas"]
grupos = defaultdict(list)
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    if not row or not row[0]:
        continue
    nro = nrokey(row[6])
    if not nro:
        continue
    grupos[(norm_prov(row[3]), nro)].append({
        "fila": i, "emision": _pd(row[0]), "venc": _pd(row[1]), "pago": _pd(row[2]),
        "prov": str(row[3] or ""), "nro": str(row[6] or ""), "total": _f(row[15]),
        "corr": row[COL_CORRELATIVO - 1] if len(row) >= COL_CORRELATIVO else None,
        "cat": str(row[16] or "")})
wb.close()

difs, solo_master, solo_fxp = [], [], []
for (prov, nro), filas in grupos.items():
    f0 = filas[0]
    d = fxp.get((prov, nro)) or buscar_en_fxp(f0["prov"], nro, fxp_nro)
    base = {"n": f0["corr"], "prov": f0["prov"], "nro": f0["nro"],
            "emision_m": f0["emision"], "venc_m": f0["venc"], "pago_m": f0["pago"],
            "total_m": f0["total"], "lineas": len(filas), "fila": f0["fila"],
            "cat": f0["cat"]}
    if not d:
        solo_master.append(base)
        continue
    if f0["total"] > 0 and d["monto"] > 0 and abs(f0["total"] - d["monto"]) > max(1000, d["monto"] * 0.01):
        base.update({"monto_f": d["monto"], "emision_f": d["emision"], "pago_f": d["pago"],
                     "estado_f": d["estado"], "nota_f": d["nota"],
                     "dif": f0["total"] - d["monto"]})
        difs.append(base)
claves = set(grupos.keys())
nros_master = {k[1] for k in claves}
for (p, n), d in fxp.items():
    if (p, n) not in claves and n not in nros_master:
        solo_fxp.append(d)

out = Workbook()
AZUL = PatternFill("solid", fgColor="1F4E78")


def cab(wsx, headers, anchos):
    for i, h in enumerate(headers, 1):
        c = wsx.cell(1, i, h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = AZUL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, a in enumerate(anchos, 1):
        wsx.column_dimensions[chr(64 + i)].width = a
    wsx.freeze_panes = "A2"


w1 = out.active
w1.title = "Diferencias de monto"
cab(w1, ["N° Archivo", "Proveedor", "N° Documento", "Emisión Master", "Emisión FXP",
         "Pago Master", "Pago FXP", "Monto Master", "Monto FXP", "Diferencia",
         "Estado FXP", "Nota FXP", "Categoría", "Líneas", "Fila Master"],
    [11, 34, 16, 14, 13, 13, 13, 15, 15, 15, 12, 34, 22, 8, 11])
for d in sorted(difs, key=lambda x: -abs(x["dif"])):
    w1.append([d["n"], d["prov"], d["nro"], d["emision_m"], d["emision_f"], d["pago_m"],
               d["pago_f"], d["total_m"], d["monto_f"], d["dif"], d["estado_f"],
               d["nota_f"], d["cat"], d["lineas"], d["fila"]])
for r in range(2, w1.max_row + 1):
    for c in (8, 9, 10):
        w1.cell(r, c).number_format = '#,##0'

w2 = out.create_sheet("Solo en Master")
cab(w2, ["N° Archivo", "Proveedor", "N° Documento", "Emisión", "Vencimiento",
         "Pago", "Monto", "Categoría", "Líneas", "Fila"],
    [11, 34, 16, 13, 13, 13, 15, 22, 8, 9])
for d in sorted(solo_master, key=lambda x: (x["emision_m"] or date(1900, 1, 1)), reverse=True):
    w2.append([d["n"], d["prov"], d["nro"], d["emision_m"], d["venc_m"], d["pago_m"],
               d["total_m"], d["cat"], d["lineas"], d["fila"]])
for r in range(2, w2.max_row + 1):
    w2.cell(r, 7).number_format = '#,##0'

w3 = out.create_sheet("Solo en FXP")
cab(w3, ["N° FXP", "Proveedor", "N° Documento", "Emisión", "Vencimiento", "Pago",
         "Monto", "Estado", "Nota"], [10, 34, 16, 13, 13, 13, 15, 12, 42])
for d in sorted(solo_fxp, key=lambda x: -x["monto"]):
    w3.append([d["n"], d["prov"], d["nro"], d["emision"], d["venc"], d["pago"],
               d["monto"], d["estado"], d["nota"]])
for r in range(2, w3.max_row + 1):
    w3.cell(r, 7).number_format = '#,##0'

out.save(SALIDA)
print(f"📄 {SALIDA}")
print(f"   Diferencias de monto: {len(difs)}")
print(f"   Solo en Master: {len(solo_master)}")
print(f"   Solo en FXP: {len(solo_fxp)}")
