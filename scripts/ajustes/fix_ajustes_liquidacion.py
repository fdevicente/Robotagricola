#!/usr/bin/env python3
"""Corrige ajustes: elimina INGRESO VENTAS de Ajustes Manuales, agrega liquidación a Cosechas."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from config import EXCEL_PATH

wb = load_workbook(EXCEL_PATH)

# 1) Eliminar el ajuste INGRESO VENTAS de Ajustes Manuales
ws_a = wb["Ajustes Manuales"]
removed = 0
for row in range(ws_a.max_row, 1, -1):
    cat = ws_a.cell(row, 3).value
    if cat == "INGRESO VENTAS":
        ws_a.delete_rows(row)
        removed += 1
print(f"Removidos {removed} ajustes INGRESO VENTAS de Ajustes Manuales\n")

# 2) Actualizar liquidaciones diciembre en Cosechas
ws_c = wb["Cosechas"]

# Distribución del $0.4 USD × 240,000 kg = $96,000 USD
# Valbifrut: 140k kg × $0.4 = $56,000 USD
# Pacific Nuts: 100k kg × $0.4 = $40,000 USD

actualizados = []
for row in range(2, ws_c.max_row + 1):
    exportadora = ws_c.cell(row, 4).value
    tipo = ws_c.cell(row, 11).value
    if tipo != "liquidacion final":
        continue
    if exportadora == "Valbifrut":
        ws_c.cell(row, 10).value = 56000
        ws_c.cell(row, 16).value = "Liquidación dic 2026: $0.4 USD × 140k kg ($904 CLP/USD)"
        actualizados.append(f"Fila {row} Valbifrut: $56,000 USD")
    elif exportadora == "Pacific Nuts":
        ws_c.cell(row, 10).value = 40000
        ws_c.cell(row, 16).value = "Liquidación dic 2026: $0.4 USD × 100k kg ($904 CLP/USD)"
        actualizados.append(f"Fila {row} Pacific Nuts: $40,000 USD")

print("Liquidaciones actualizadas:")
for a in actualizados:
    print(f"  {a}")
print(f"\nTotal: $96,000 USD = ${96000 * 904:,.0f} CLP\n")

wb.save(EXCEL_PATH)
wb.close()
print("Done!")
