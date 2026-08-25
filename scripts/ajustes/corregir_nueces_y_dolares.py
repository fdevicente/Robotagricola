"""Corrige la venta de nueces con lo que aclaró el dueño (5-ago-2026).

1. Valbifrut adelantó SOLO 1,8 USD/kg. Su liquidación de diciembre cierra el
   precio en 2,1 → 0,3 USD/kg = 41.700 USD.
2. Pacific Nuts queda como estaba: su estructura ya suma 2,1 USD/kg
   (0,8+0,3+0,3+0,2+0,5). Se revierte la baja que le había aplicado.
3. Bono de venta recalculado con la nueva base.
"""
import shutil
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb

PRECIO_TOTAL = 2.1
ADELANTO_VALBIFRUT = 1.8
USD_CLP_BONO = 904
PCT_BONO = 0.08

COL_EXP, COL_KG, COL_MONTO_USD, COL_TIPO, COL_ESTADO, COL_NOTAS = 4, 5, 10, 11, 12, 16

resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"Respaldo: {resp}\n")

wb = load_workbook(EXCEL_PATH)
ws = wb["Cosechas"]

# Valores correctos de la liquidación final por exportadora
NUEVO = {
    "Valbifrut":    round(139_000 * (PRECIO_TOTAL - ADELANTO_VALBIFRUT)),  # 0,3 USD/kg
    "Pacific Nuts": 50_000,                                                # como estaba
}

print("Liquidación final de diciembre:")
print(f"  {'fila':>4} {'exportadora':14} {'kg':>9} {'antes':>9} {'queda':>9}  USD/kg")
print("  " + "-" * 58)
esperado_usd = 0
for r in range(2, ws.max_row + 1):
    if str(ws.cell(r, 2).value or "").upper() != "NOGALES":
        continue
    exp = str(ws.cell(r, COL_EXP).value or "")
    recibido = str(ws.cell(r, COL_ESTADO).value or "").lower() == "recibido"
    monto = float(ws.cell(r, COL_MONTO_USD).value or 0)
    if not recibido:
        esperado_usd += monto
    if "liquidacion" not in str(ws.cell(r, COL_TIPO).value or "").lower() or recibido:
        continue
    if exp not in NUEVO:
        continue
    kg = float(ws.cell(r, COL_KG).value or 0)
    nuevo = NUEVO[exp]
    esperado_usd += nuevo - monto        # corrige lo ya sumado
    ws.cell(r, COL_MONTO_USD).value = nuevo
    ws.cell(r, COL_NOTAS).value = (f"Liquidación dic 2026 a {nuevo / kg:.2f} USD/kg "
                                    f"— cierra el precio en {PRECIO_TOTAL} USD/kg")
    print(f"  {r:>4} {exp[:14]:14} {kg:>9,.0f} {monto:>9,.0f} {nuevo:>9,.0f}  {nuevo / kg:.2f}")

print(f"\n  Total por recibir en USD: {esperado_usd:,.0f}")

# ── Bono venta nueces ──
RECIBIDO_CLP = 266_079_757     # los dos abonos de Valbifrut (1,8 USD/kg)
base = RECIBIDO_CLP + esperado_usd * USD_CLP_BONO
bono = round(base * PCT_BONO)
wsa = wb["Ajustes Manuales"]
for r in range(2, wsa.max_row + 1):
    if "BONO" not in str(wsa.cell(r, 3).value or "").upper():
        continue
    antes = float(wsa.cell(r, 5).value or 0)
    wsa.cell(r, 5).value = bono
    wsa.cell(r, 6).value = f"Bono 8% × ventas nueces ${base:,.0f} (precio 2,1 USD/kg)"
    print(f"\nBono venta nueces (fila {r}):")
    print(f"  base ventas ${base:,.0f}")
    print(f"  bono        ${antes:,.0f} → ${bono:,.0f}  ({bono - antes:+,.0f})")
    break

_save_wb(wb)
wb.close()
print("\n✅ Aplicado.")
