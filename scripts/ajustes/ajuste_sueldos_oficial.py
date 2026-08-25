#!/usr/bin/env python3
"""Ajuste de MANO DE OBRA PLANTA basado en liquidaciones oficiales:
- Líquidos mensuales del personal fijo
- Previred (cotización empleador)
- Aguinaldo de fiestas patrias (septiembre) y navidad (diciembre)
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
from openpyxl import load_workbook
from config import EXCEL_PATH

# Líquidos mensuales según liquidación abril 2026
LIQUIDOS = {
    "Felicito Amigo Soto":            989_640,
    "Luis Ramiro Amigo Soto":         631_637,
    "Felix De Vicente":             2_023_980,
    "Luis Patricio Mora Amigo":       679_556,
    "Agustin Segundo Mora Hernandez": 748_834,
    "Juan Parada Castillo":         1_350_880,
    "Javier Gonzalez":                584_733,  # REAL: liquidación julio 2026, mes normal reconstruido
}                                               # (base 553.553 + gratif. 25% − AFP 11,16% − Fonasa 7% − cesantía 0,6% + carga familiar)

# Bases de los 6 originales (para aguinaldo = medio sueldo base c/u)
BASES_ORIGINALES = (871_600 + 539_000 + 2_315_367 + 575_000 + 586_989 + 1_470_365)
LIQUIDO_ORIGINALES = (989_640 + 631_637 + 2_023_980 + 679_556 + 748_834 + 1_350_880)

# Previred (cotización empleador mensual) - histórico abril 2026 = $2,006,342 (6 personas)
PREVIRED_BASE = 2_006_342
# Javier González — liquidación julio 2026:
#   base y líquido son REALES; el Previred sigue prorrateado (no tenemos su planilla)
JAVIER_LIQUIDO = 584_733
JAVIER_BASE = 553_553                                                          # REAL
PREVIRED_JAVIER = round(JAVIER_LIQUIDO * PREVIRED_BASE / LIQUIDO_ORIGINALES)   # ≈ $183k ESTIMADO

PREVIRED_MENSUAL = PREVIRED_BASE + PREVIRED_JAVIER
LIQUIDO_MENSUAL = sum(LIQUIDOS.values())
SUELDO_BASE_TOTAL = BASES_ORIGINALES + JAVIER_BASE
AGUINALDO = SUELDO_BASE_TOTAL / 2  # medio sueldo cada uno

print(f"Líquido mensual: ${LIQUIDO_MENSUAL:,}")
print(f"Previred mensual: ${PREVIRED_MENSUAL:,}")
print(f"Total mensual: ${LIQUIDO_MENSUAL + PREVIRED_MENSUAL:,}")
print(f"Aguinaldo (medio sueldo base): ${AGUINALDO:,.0f}")
print()

MESES_TEMP = [
    (2026, 6), (2026, 7), (2026, 8), (2026, 9), (2026, 10), (2026, 11),
    (2026, 12), (2027, 1), (2027, 2), (2027, 3), (2027, 4), (2027, 5),
]

wb = load_workbook(EXCEL_PATH)
ws_a = wb["Ajustes Manuales"]

# Eliminar ajustes anteriores de MANO DE OBRA PLANTA
removed = 0
for r in range(ws_a.max_row, 1, -1):
    if ws_a.cell(r, 3).value == "MANO DE OBRA PLANTA":
        ws_a.delete_rows(r)
        removed += 1
print(f"Removidos {removed} ajustes anteriores\n")

next_row = ws_a.max_row + 1
while ws_a.cell(next_row, 1).value:
    next_row += 1

hoy = date.today().isoformat()


def add(y, m, monto, razon):
    global next_row
    ws_a.cell(next_row, 1).value = hoy
    ws_a.cell(next_row, 2).value = f"{y}-{m:02d}"
    ws_a.cell(next_row, 3).value = "MANO DE OBRA PLANTA"
    ws_a.cell(next_row, 4).value = "GENERAL"
    ws_a.cell(next_row, 5).value = monto
    ws_a.cell(next_row, 6).value = razon
    ws_a.cell(next_row, 7).value = True
    next_row += 1


total_anual = 0
for (y, m) in MESES_TEMP:
    monto_base = LIQUIDO_MENSUAL + PREVIRED_MENSUAL
    extra = 0
    extra_label = ""
    # Aguinaldo Fiestas Patrias (septiembre)
    if m == 9:
        extra = AGUINALDO
        extra_label = " + aguinaldo Fiestas Patrias"
    # Aguinaldo Navidad (diciembre)
    elif m == 12:
        extra = AGUINALDO
        extra_label = " + aguinaldo Navidad"

    monto = monto_base + extra
    total_anual += monto
    add(y, m, monto, f"Sueldos líquidos ${LIQUIDO_MENSUAL:,} + Previred ${PREVIRED_MENSUAL:,}{extra_label}")
    print(f"  {y}-{m:02d}: ${monto:,.0f}")

print(f"\nTotal anual: ${total_anual:,.0f}")

wb.save(EXCEL_PATH)
wb.close()
print("\nDone!")
