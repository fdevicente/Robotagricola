#!/usr/bin/env python3
"""Reclasifica los 6 traspasos $5M del 2026-01-28 como préstamo de Felix De Vicente."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]

reclass = 0
for row in range(2, ws.max_row + 1):
    fecha_val = ws.cell(row, 1).value
    if not fecha_val: continue
    fecha = fecha_val.date() if isinstance(fecha_val, datetime) else fecha_val
    if not isinstance(fecha, date): continue
    if fecha != date(2026, 1, 28): continue

    desc = str(ws.cell(row, 2).value or "")
    if "agricola santa elisa bco chile" not in desc.lower():
        continue
    try:
        abono = float(ws.cell(row, 5).value or 0)
    except: continue
    if abono <= 0: continue

    ws.cell(row, 8).value = "PRESTAMOS A OTRAS SOCIEDADES"
    reclass += 1

print(f"Reclasificados {reclass} → PRESTAMOS A OTRAS SOCIEDADES (préstamo Felix De Vicente, deuda pendiente)")
wb.save(EXCEL_PATH)
wb.close()
print("Done!")
