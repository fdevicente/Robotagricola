#!/usr/bin/env python3
"""Actualiza Cosechas: Villa Las Delicias 40kg ya pagado + Valbifrut adicional."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
from openpyxl import load_workbook
from config import EXCEL_PATH

wb = load_workbook(EXCEL_PATH)
ws = wb["Cosechas"]

print("Actualizando Cosechas...\n")
for r in range(2, ws.max_row + 1):
    if not ws.cell(r, 1).value: continue
    exportadora = str(ws.cell(r, 4).value or "")
    kg = ws.cell(r, 3).value or 0

    # Villa Las Delicias 40kg → marcar como recibido
    if "Dulces la Villa" in exportadora and kg == 40:
        ws.cell(r, 12).value = "recibido"
        ws.cell(r, 13).value = date(2026, 5, 14)
        ws.cell(r, 14).value = 240000
        ws.cell(r, 15).value = "CLP"
        ws.cell(r, 16).value = "40 kg × $6000 c/IVA (F217 - 2026-05-14)"
        print(f"  Fila {r}: Dulces la Villa 40kg → recibido $240,000")

# Agregar Valbifrut adicional $42.5M (probablemente cuota 2 adelanto)
next_row = ws.max_row + 1
while ws.cell(next_row, 1).value:
    next_row += 1

ws.cell(next_row, 1).value = 2026
ws.cell(next_row, 2).value = "NOGALES"
ws.cell(next_row, 3).value = 140000
ws.cell(next_row, 4).value = "Valbifrut"
ws.cell(next_row, 5).value = 139000
ws.cell(next_row, 6).value = 1.8
ws.cell(next_row, 7).value = 3  # Ahora son 3 cuotas (estaba como 2)
ws.cell(next_row, 8).value = 2  # cuota 2 de 3
ws.cell(next_row, 9).value = "2026-05-18"
ws.cell(next_row, 10).value = 47000  # USD estimado (42.5M / 904)
ws.cell(next_row, 11).value = "adelanto"
ws.cell(next_row, 12).value = "recibido"
ws.cell(next_row, 13).value = date(2026, 5, 18)
ws.cell(next_row, 14).value = 42483234
ws.cell(next_row, 15).value = "CLP"
ws.cell(next_row, 16).value = "2do adelanto Valbifrut 2026-05-18"
print(f"  Fila {next_row}: Valbifrut 2do adelanto $42,483,234 (2026-05-18)")

wb.save(EXCEL_PATH)
wb.close()
print("\nDone!")
