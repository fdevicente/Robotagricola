#!/usr/bin/env python3
"""Correcciones a ingresos 2026:
1. Reclasificar 6 traspasos $5M (2026-01-28) como TRANSFERENCIA INTERNA
2. Agregar Cerezos 2026 a Cosechas: 24,000 kg × $1.2 USD = $28,800 USD (recibido marzo)
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

USD_CLP = 904

wb = load_workbook(EXCEL_PATH)

# ─── 1) Reclasificar traspasos ──────────────────────────────────
ws_b = wb["Cuenta Banco"]
reclass = 0
for row in range(2, ws_b.max_row + 1):
    fecha_val = ws_b.cell(row, 1).value
    if not fecha_val: continue
    fecha = fecha_val.date() if isinstance(fecha_val, datetime) else fecha_val
    if not isinstance(fecha, date): continue
    if fecha != date(2026, 1, 28): continue

    desc = str(ws_b.cell(row, 2).value or "")
    if "agricola santa elisa bco chile" not in desc.lower():
        continue
    try:
        abono = float(ws_b.cell(row, 5).value or 0)
    except: continue
    if abono <= 0: continue

    cat_old = ws_b.cell(row, 8).value
    ws_b.cell(row, 8).value = "TRANSFERENCIA INTERNA"
    print(f"  Fila {row}: {cat_old} → TRANSFERENCIA INTERNA (${abono:,.0f})")
    reclass += 1

print(f"\nReclasificados: {reclass} traspasos\n")

# ─── 2) Agregar Cerezos a Cosechas ───────────────────────────────
ws_c = wb["Cosechas"]

# Encontrar primera fila libre
next_row = ws_c.max_row + 1
while ws_c.cell(next_row, 1).value:
    next_row += 1

# Datos cerezos
kg = 24000
precio_usd = 1.2
monto_usd = kg * precio_usd
monto_clp = monto_usd * USD_CLP

ws_c.cell(next_row, 1).value = 2026
ws_c.cell(next_row, 2).value = "CEREZOS"
ws_c.cell(next_row, 3).value = kg          # Kg total
ws_c.cell(next_row, 4).value = "Exportadora cerezas"
ws_c.cell(next_row, 5).value = kg          # Kg asignados
ws_c.cell(next_row, 6).value = precio_usd  # Precio USD/kg
ws_c.cell(next_row, 7).value = 1           # N° cuotas
ws_c.cell(next_row, 8).value = 1           # Cuota #
ws_c.cell(next_row, 9).value = "2026-03-15" # Fecha estimada
ws_c.cell(next_row, 10).value = monto_usd  # Monto USD estimado
ws_c.cell(next_row, 11).value = "pago único" # Tipo cuota
ws_c.cell(next_row, 12).value = "recibido"  # Estado
ws_c.cell(next_row, 13).value = date(2026, 3, 15) # Fecha real
ws_c.cell(next_row, 14).value = monto_clp   # Monto real recibido
ws_c.cell(next_row, 15).value = "CLP"       # Moneda
ws_c.cell(next_row, 16).value = f"24,000 kg × $1.2 USD ($904 CLP/USD)"

print(f"Agregado a Cosechas fila {next_row}:")
print(f"  CEREZOS 2026: 24,000 kg × $1.2 USD = $28,800 USD")
print(f"  CLP: ${monto_clp:,.0f}")
print(f"  Estado: recibido (2026-03-15)\n")

wb.save(EXCEL_PATH)
wb.close()
print("Done!")
