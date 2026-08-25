#!/usr/bin/env python3
"""Marca todas las facturas de Administradora de Ventas al Detalle (Copec)
como pagadas al momento - fecha de pago = fecha de emisión."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

KEYWORDS = ["adm", "ventas al detalle"]  # ambos deben aparecer


def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try: return datetime.strptime(v[:10], fmt).date()
            except: pass
    return None


def es_adm_ventas(prov):
    p = (prov or "").lower()
    return all(kw in p for kw in KEYWORDS)


wb = load_workbook(EXCEL_PATH)
ws = wb["Facturas"]

# Col 1 = fecha_emi, col 2 = vencimiento, col 3 = fecha_pago, col 4 = proveedor
updated = 0
total_monto = 0
for r in range(2, ws.max_row + 1):
    prov = ws.cell(r, 4).value
    if not es_adm_ventas(prov): continue
    fecha_pago = ws.cell(r, 3).value
    if fecha_pago and str(fecha_pago).strip():
        continue  # ya pagada
    fecha_emi = _pd(ws.cell(r, 1).value)
    if not fecha_emi: continue
    # Pagada al momento = fecha de pago = fecha de emisión
    ws.cell(r, 3).value = fecha_emi
    try:
        monto = float(ws.cell(r, 15).value or 0)
    except: monto = 0
    total_monto += monto
    updated += 1

print(f"Facturas Adm Ventas al Detalle marcadas como pagadas: {updated}")
print(f"Monto total: ${total_monto:,.0f}")

wb.save(EXCEL_PATH)
wb.close()
print("Done!")
