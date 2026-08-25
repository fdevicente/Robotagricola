#!/usr/bin/env python3
"""Agrega venta Villa Las Delicias / Dulces la Villa a Cosechas.
Buscar también pago de 40kg en el banco."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

print("Buscando pagos cerca a $240,000 (40kg × $6000) en banco abril-mayo 2026...\n")
wb = load_workbook(EXCEL_PATH)
ws_b = wb["Cuenta Banco"]

def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None

candidatos_40kg = []
for r in range(2, ws_b.max_row + 1):
    fecha = _pd(ws_b.cell(r, 1).value)
    if not fecha or fecha < date(2026, 4, 1) or fecha > date(2026, 5, 31):
        continue
    try:
        abono = float(ws_b.cell(r, 5).value or 0)
    except: continue
    if abono <= 0: continue
    desc = str(ws_b.cell(r, 2).value or "")
    # Cualquier abono entre 200k y 300k
    if 200000 <= abono <= 300000:
        candidatos_40kg.append((r, fecha, abono, desc))

print("Candidatos pago 40kg ($200K-$300K) abril-mayo 2026:")
for r, f, a, d in candidatos_40kg:
    print(f"  Fila {r} | {f} | ${a:,.0f} | {d}")

# Agregar a Cosechas: 50kg total (10kg ya pagado, 40kg por pagar)
print("\nAgregando ventas Villa Las Delicias a Cosechas...")
ws_c = wb["Cosechas"]
next_row = ws_c.max_row + 1
while ws_c.cell(next_row, 1).value:
    next_row += 1

# Pago 1: 10kg recibido 2026-05-11 ($60K) — F218 Dulces la Villa
ws_c.cell(next_row, 1).value = 2026
ws_c.cell(next_row, 2).value = "NOGALES"
ws_c.cell(next_row, 3).value = 10
ws_c.cell(next_row, 4).value = "Dulces la Villa"
ws_c.cell(next_row, 5).value = 10
ws_c.cell(next_row, 6).value = 0  # no USD, venta local CLP
ws_c.cell(next_row, 7).value = 1
ws_c.cell(next_row, 8).value = 1
ws_c.cell(next_row, 9).value = "2026-05-11"
ws_c.cell(next_row, 10).value = 0
ws_c.cell(next_row, 11).value = "venta local"
ws_c.cell(next_row, 12).value = "recibido"
ws_c.cell(next_row, 13).value = date(2026, 5, 11)
ws_c.cell(next_row, 14).value = 60000  # CLP
ws_c.cell(next_row, 15).value = "CLP"
ws_c.cell(next_row, 16).value = "10 kg × $6000 c/IVA (F218)"
print(f"  Fila {next_row}: 10kg recibido $60,000 (2026-05-11)")
next_row += 1

# Pago 2: 40kg ($240K) — buscar fecha, si no, ponerla como esperado/pendiente
fecha_40 = None
monto_40 = 240000
for r, f, a, d in candidatos_40kg:
    if "villa" in d.lower() or "delicias" in d.lower() or abs(a - 240000) < 5000:
        fecha_40 = f
        monto_40 = a
        break

estado = "recibido" if fecha_40 else "esperado"
ws_c.cell(next_row, 1).value = 2026
ws_c.cell(next_row, 2).value = "NOGALES"
ws_c.cell(next_row, 3).value = 40
ws_c.cell(next_row, 4).value = "Dulces la Villa"
ws_c.cell(next_row, 5).value = 40
ws_c.cell(next_row, 6).value = 0
ws_c.cell(next_row, 7).value = 1
ws_c.cell(next_row, 8).value = 1
ws_c.cell(next_row, 9).value = fecha_40.isoformat() if fecha_40 else "2026-05-31"
ws_c.cell(next_row, 10).value = 0
ws_c.cell(next_row, 11).value = "venta local"
ws_c.cell(next_row, 12).value = estado
if fecha_40:
    ws_c.cell(next_row, 13).value = fecha_40
    ws_c.cell(next_row, 14).value = monto_40
    ws_c.cell(next_row, 15).value = "CLP"
ws_c.cell(next_row, 16).value = "40 kg × $6000 c/IVA"
print(f"  Fila {next_row}: 40kg {estado} {f'${monto_40:,.0f}' if fecha_40 else 'PENDIENTE'} ({fecha_40 or 'fecha por confirmar'})")

wb.save(EXCEL_PATH)
wb.close()
print("\nDone!")
