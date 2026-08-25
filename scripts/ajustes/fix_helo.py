#!/usr/bin/env python3
"""Reclasifica facturas/cargos de helicoptero a MANTENIMIENTO HELICOPTERO
y actualiza ajuste manual para anular el total real."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

HELO_KEYWORDS = [
    "arrayan-aeromar", "arrayan aeromar", "aeromar",
    "helicoptero", "helicóptero", "vision air", "hangar",
    "smartways", "control heladas", "control de heladas",
    "francisco pena", "miguel marin",  # piloto y mantención helo
    "manga viento", "manga de viento",
    "vuelo control heladas", "vuelo helicoptero",
]


def has_helo(text):
    t = (text or "").lower()
    return any(kw in t for kw in HELO_KEYWORDS)


def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try: return datetime.strptime(v[:10], fmt).date()
            except: pass
    return None


wb = load_workbook(EXCEL_PATH)
total_helo = 0

# Facturas
print("=== RECLASIFICANDO FACTURAS HELICOPTERO ===\n")
ws_f = wb["Facturas"]
fact_changed = 0
fact_total = 0
for r in range(2, ws_f.max_row + 1):
    if not ws_f.cell(r, 1).value: continue
    prov = str(ws_f.cell(r, 4).value or "")
    detalle = str(ws_f.cell(r, 8).value or "")
    glosa_ii = str(ws_f.cell(r, 9).value or "")
    cat = str(ws_f.cell(r, 17).value or "").upper()

    if cat == "MANTENIMIENTO HELICOPTERO":
        try:
            monto = float(ws_f.cell(r, 15).value or 0)
        except: monto = 0
        fact_total += monto
        continue

    full_text = f"{prov} {detalle} {glosa_ii}"
    if has_helo(full_text):
        try:
            monto = float(ws_f.cell(r, 15).value or 0)
        except: monto = 0
        old = ws_f.cell(r, 17).value
        ws_f.cell(r, 17).value = "MANTENIMIENTO HELICOPTERO"
        fact_changed += 1
        fact_total += monto
        if monto > 100000:
            print(f"  Fila {r} | ${monto:>12,.0f} | {old:25} → MANT HELO | {prov[:30]} | {detalle[:40]}")

print(f"\nFacturas reclasificadas: {fact_changed}, total facturas helo: ${fact_total:,.0f}\n")

# Banco
print("=== RECLASIFICANDO CARGOS BANCO HELICOPTERO ===\n")
ws_b = wb["Cuenta Banco"]
banco_changed = 0
banco_total = 0
for r in range(2, ws_b.max_row + 1):
    if not ws_b.cell(r, 1).value: continue
    cat = str(ws_b.cell(r, 8).value or "").upper()
    desc = str(ws_b.cell(r, 2).value or "")
    ref = str(ws_b.cell(r, 3).value or "")

    if cat == "MANTENIMIENTO HELICOPTERO":
        try:
            cargo = float(ws_b.cell(r, 4).value or 0)
        except: cargo = 0
        banco_total += cargo
        continue

    if has_helo(desc + " " + ref):
        try:
            cargo = float(ws_b.cell(r, 4).value or 0)
        except: cargo = 0
        if cargo <= 0: continue
        old = ws_b.cell(r, 8).value
        ws_b.cell(r, 8).value = "MANTENIMIENTO HELICOPTERO"
        banco_changed += 1
        banco_total += cargo
        if cargo > 100000:
            print(f"  Fila {r} | ${cargo:>12,.0f} | {old or '(vacío)':25} → MANT HELO | {desc[:55]}")

print(f"\nCargos banco reclasificados: {banco_changed}, total banco helo: ${banco_total:,.0f}\n")

# Total
total_helo = fact_total + banco_total
print(f"=== TOTAL HELICOPTERO HISTORICO: ${total_helo:,.0f} ===\n")

# Actualizar Ajustes Manuales
print("=== ACTUALIZANDO AJUSTES MANUALES ===\n")
ws_a = wb["Ajustes Manuales"]
# Eliminar ajustes existentes de MANTENIMIENTO HELICOPTERO
removed = 0
for r in range(ws_a.max_row, 1, -1):
    if ws_a.cell(r, 3).value == "MANTENIMIENTO HELICOPTERO":
        ws_a.delete_rows(r)
        removed += 1
print(f"  Removidos {removed} ajustes anteriores de helicoptero\n")

# Calcular nuevo gasto histórico mensual (último año) para anular
# Vamos a calcular gasto real últimos 12 meses
print("Calculando gasto mensual a anular (base últimos 12 meses)...")
from collections import defaultdict
helo_mensual = defaultdict(float)
for r in range(2, ws_b.max_row + 1):
    cat = str(ws_b.cell(r, 8).value or "")
    if cat != "MANTENIMIENTO HELICOPTERO": continue
    fecha = _pd(ws_b.cell(r, 1).value)
    if not fecha or fecha < date(2025, 5, 1): continue
    try:
        cargo = float(ws_b.cell(r, 4).value or 0)
    except: continue
    helo_mensual[(fecha.year, fecha.month)] += cargo

for r in range(2, ws_f.max_row + 1):
    cat = str(ws_f.cell(r, 17).value or "")
    if cat != "MANTENIMIENTO HELICOPTERO": continue
    fecha = _pd(ws_f.cell(r, 1).value)
    if not fecha or fecha < date(2025, 5, 1): continue
    try:
        monto = float(ws_f.cell(r, 15).value or 0)
    except: continue
    helo_mensual[(fecha.year, fecha.month)] += monto

total_year = sum(helo_mensual.values())
print(f"  Total últimos 12m: ${total_year:,.0f}")
for (y, m), v in sorted(helo_mensual.items()):
    print(f"    {y}-{m:02d}: ${v:,.0f}")

# Agregar ajuste negativo por cada mes del año proyectado
print(f"\nAgregando ajustes de anulación...")
hoy = date.today()
next_row = ws_a.max_row + 1
while ws_a.cell(next_row, 1).value:
    next_row += 1

# Para cada mes 2026-05 a 2027-04, crear ajuste -X que anule
# Pero usar el monto histórico del mismo mes para anular punto a punto
for (y, m), v in sorted(helo_mensual.items()):
    if v <= 0: continue
    # Proyectar al mismo mes del año siguiente
    target_y = y + 1
    target_m = m
    if (target_y, target_m) >= (hoy.year, hoy.month):
        ws_a.cell(next_row, 1).value = hoy.isoformat()
        ws_a.cell(next_row, 2).value = f"{target_y}-{target_m:02d}"
        ws_a.cell(next_row, 3).value = "MANTENIMIENTO HELICOPTERO"
        ws_a.cell(next_row, 4).value = "GENERAL"
        ws_a.cell(next_row, 5).value = -v
        ws_a.cell(next_row, 6).value = "Anular - se descontinuó helicoptero"
        ws_a.cell(next_row, 7).value = True
        print(f"  {target_y}-{target_m:02d}: ajuste -${v:,.0f}")
        next_row += 1

wb.save(EXCEL_PATH)
wb.close()
print("\nDone!")
