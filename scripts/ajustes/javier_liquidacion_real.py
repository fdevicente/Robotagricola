"""Javier González: reemplaza los ESTIMADOS por su liquidación real (julio 2026).

La liquidación de julio NO sirve tal cual como costo mensual: trae 2 días de
inasistencia (−$36.904) y un anticipo de $150.000. Acá se reconstruye el mes
normal desde el sueldo base y los % de la propia liquidación.
"""
import shutil
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb

# ── Datos de la liquidación (julio 2026) ──
BASE = 553_553          # sueldo base contractual (30 días × $18.451,77)
CARGA_FAMILIAR = 22_601  # no imponible
PCT_GRATIF = 0.25        # gratificación 25% de remuneraciones (bajo el tope legal)
PCT_AFP = 0.1116         # PlanVital
PCT_SALUD = 0.07         # Fonasa
PCT_CESANTIA_TRAB = 0.006

# Estimados que se reemplazan
EST_LIQUIDO = 532_872

# Para prorratear el aporte del empleador con el mismo criterio del resto
LIQUIDO_ORIGINALES = 6_424_527   # 6 trabajadores, liquidación abril 2026
BASES_ORIGINALES = 6_358_321
PREVIRED_BASE = 2_006_342        # pago Previred de esos 6

# ── Mes normal (sin inasistencia ni anticipo) ──
gratificacion = round(BASE * PCT_GRATIF)
imponible = BASE + gratificacion
haberes = imponible + CARGA_FAMILIAR
afp = round(imponible * PCT_AFP)
salud = round(imponible * PCT_SALUD)
cesantia = round(imponible * PCT_CESANTIA_TRAB)
descuentos = afp + salud + cesantia
liquido = haberes - descuentos

ratio_previred = PREVIRED_BASE / LIQUIDO_ORIGINALES
previred = round(liquido * ratio_previred)

print("=" * 62)
print("JAVIER GONZÁLEZ — mes normal reconstruido desde la liquidación")
print("=" * 62)
print(f"  Sueldo base                       ${BASE:>12,}")
print(f"  Gratificación 25%                 ${gratificacion:>12,}")
print(f"  {'':34}{'—' * 13}")
print(f"  Total imponible                   ${imponible:>12,}")
print(f"  Carga familiar (no imponible)     ${CARGA_FAMILIAR:>12,}")
print(f"  TOTAL HABERES                     ${haberes:>12,}")
print()
print(f"  AFP PlanVital 11,16%              ${-afp:>12,}")
print(f"  Fonasa 7%                         ${-salud:>12,}")
print(f"  Seguro cesantía 0,6%              ${-cesantia:>12,}")
print(f"  {'':34}{'—' * 13}")
print(f"  LÍQUIDO mensual normal            ${liquido:>12,}")
print()
print(f"  Previred (prorrateado)            ${previred:>12,}   ← estimado")
print(f"  COSTO EMPRESA / MES               ${liquido + previred:>12,}")
print()

est_previred = round(EST_LIQUIDO * ratio_previred)
est_costo = EST_LIQUIDO + est_previred
real_costo = liquido + previred
dif_mes = real_costo - est_costo

est_base = round(EST_LIQUIDO * BASES_ORIGINALES / LIQUIDO_ORIGINALES)
dif_aguinaldo = (BASE - est_base) / 2 * 2   # dos aguinaldos al año (medio sueldo c/u)

print("=" * 62)
print("IMPACTO vs lo que estaba proyectado")
print("=" * 62)
print(f"{'':22}{'ESTIMADO':>14}{'REAL':>14}{'DIF':>12}")
print(f"  {'Sueldo base':20}{est_base:>14,}{BASE:>14,}{BASE - est_base:>+12,}")
print(f"  {'Líquido':20}{EST_LIQUIDO:>14,}{liquido:>14,}{liquido - EST_LIQUIDO:>+12,}")
print(f"  {'Costo empresa/mes':20}{est_costo:>14,}{real_costo:>14,}{dif_mes:>+12,}")
print()
print(f"  12 meses de la temporada        {dif_mes * 12:>+12,}")
print(f"  2 aguinaldos (medio sueldo)     {dif_aguinaldo:>+12,.0f}")
print(f"  {'':32}{'—' * 12}")
print(f"  IMPACTO EN LA PROYECCIÓN 26/27  {dif_mes * 12 + dif_aguinaldo:>+12,.0f}")
print()

# ── Guardar el sueldo base en la hoja que lo registra ──
resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
wb = load_workbook(EXCEL_PATH)
ws = wb["Vacaciones Pendientes"]
fila = next((r for r in range(2, ws.max_row + 1)
             if str(ws.cell(r, 1).value or "").strip().upper() == "JAVIER GONZALEZ"), None)
if fila:
    ws.cell(fila, 7).value = "TRABAJADOR AGRICOLA"
    ws.cell(fila, 8).value = BASE
    ws.cell(fila, 6).value = ("Liquidación julio 2026 · contrato 01-11-2025 · "
                               "AFP PlanVital · Fonasa · 1 carga familiar")
    _save_wb(wb)
    print(f"✅ Hoja 'Vacaciones Pendientes' fila {fila}: sueldo base ${BASE:,} guardado")
else:
    print("❌ No encontré a Javier en 'Vacaciones Pendientes'")
wb.close()
print(f"   Respaldo: {resp}")
