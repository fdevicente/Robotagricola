#!/usr/bin/env python3
"""Actualiza superficies 2026 y ajuste replante (9→6 ha)."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
from openpyxl import load_workbook
from config import EXCEL_PATH

# Datos actuales (post-replante el total avellanos será 10.6 + 6 = 16.6)
SUPERFICIES_2026 = {
    "NOGALES": 46.82,
    "CEREZOS": 3.9,
    "AVELLANOS": 16.6,  # 10.6 actual + 6 replante
}
REPLANTE_HA = 6
COSTO_HA = 5_000_000

wb = load_workbook(EXCEL_PATH)

# Actualizar hoja Hectareas
print("[1/2] Actualizando hoja Hectareas...")
if "Hectareas" in wb.sheetnames:
    ws = wb["Hectareas"]
    # Mostrar contenido actual
    print("  Headers:")
    for c in range(1, ws.max_column + 1):
        print(f"    Col {c}: {ws.cell(1, c).value}")
    print()
    print("  Filas actuales:")
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 6)]
        if any(vals):
            print(f"    R{r}: {vals}")

    # Buscar fila 2026 y actualizar
    updated = False
    for r in range(2, ws.max_row + 1):
        anio = ws.cell(r, 1).value
        try:
            anio_int = int(anio) if anio else 0
        except: anio_int = 0
        if anio_int == 2026:
            # Asumimos: col 1=Año, col 2=NOGALES, col 3=CEREZOS, col 4=AVELLANOS
            ws.cell(r, 2).value = SUPERFICIES_2026["NOGALES"]
            ws.cell(r, 3).value = SUPERFICIES_2026["CEREZOS"]
            ws.cell(r, 4).value = SUPERFICIES_2026["AVELLANOS"]
            print(f"\n  ✓ Fila 2026 actualizada: NOGALES={SUPERFICIES_2026['NOGALES']} CEREZOS={SUPERFICIES_2026['CEREZOS']} AVELLANOS={SUPERFICIES_2026['AVELLANOS']}")
            updated = True
            break

    if not updated:
        # Agregar fila 2026
        next_r = ws.max_row + 1
        ws.cell(next_r, 1).value = 2026
        ws.cell(next_r, 2).value = SUPERFICIES_2026["NOGALES"]
        ws.cell(next_r, 3).value = SUPERFICIES_2026["CEREZOS"]
        ws.cell(next_r, 4).value = SUPERFICIES_2026["AVELLANOS"]
        print(f"\n  ✓ Agregada fila {next_r} 2026")

# Actualizar ajuste manual del replante
print("\n[2/2] Actualizando ajuste replante (9→6 ha)...")
ws_a = wb["Ajustes Manuales"]
nuevo_monto = REPLANTE_HA * COSTO_HA
for r in range(2, ws_a.max_row + 1):
    cat = ws_a.cell(r, 3).value
    cultivo = ws_a.cell(r, 4).value
    if cat == "INVERSION / REPLANTE" and cultivo == "AVELLANOS":
        monto_old = ws_a.cell(r, 5).value
        ws_a.cell(r, 5).value = nuevo_monto
        ws_a.cell(r, 6).value = f"Replante {REPLANTE_HA} ha avellanos ({REPLANTE_HA} × ${COSTO_HA:,.0f}/ha)"
        print(f"  Fila {r}: ${monto_old:,.0f} → ${nuevo_monto:,.0f}")

wb.save(EXCEL_PATH)
wb.close()
print("\nDone!")
