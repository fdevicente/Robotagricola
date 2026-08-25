#!/usr/bin/env python3
"""Crea hoja 'Vacaciones Pendientes' con saldos del personal fijo."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
from openpyxl import load_workbook
from config import EXCEL_PATH

# Datos del personal con saldo histórico conocido
PERSONAL = [
    # (Nombre, RUT, Fecha contrato, Saldo último conocido, Fecha del saldo)
    ("Luis Ramiro Amigo Soto",         "11.768.374-5", date(2020, 11, 1), 24.04, date(2023, 5, 15)),
    ("Luis Patricio Mora Amigo",       "21.331.792-k", date(2022, 3, 1),  14.88, date(2023, 2, 28)),
    ("Felicito Amigo Soto",            "9.850.887-2",  date(2021, 11, 1), 19.83, date(2023, 2, 28)),
    ("Agustin Segundo Mora Hernandez", "12.318.508-0", date(2021, 12, 1), 19.96, date(2023, 3, 31)),
    # Personal sin datos de vacaciones documentado (estimar desde fecha contrato)
    ("Felix De Vicente",               "—",            date(2022, 1, 1),  0.0,   date(2022, 1, 1)),
    ("Juan Parada",                    "—",            date(2022, 1, 1),  0.0,   date(2022, 1, 1)),
]

wb = load_workbook(EXCEL_PATH)

# Crear o reemplazar hoja
SHEET = "Vacaciones Pendientes"
if SHEET in wb.sheetnames:
    del wb[SHEET]
ws = wb.create_sheet(SHEET)

# Headers
ws.cell(1, 1).value = "Nombre"
ws.cell(1, 2).value = "RUT"
ws.cell(1, 3).value = "Fecha Contrato"
ws.cell(1, 4).value = "Saldo Último Conocido (días)"
ws.cell(1, 5).value = "Fecha del Saldo"
ws.cell(1, 6).value = "Notas"

# Datos
for i, (nombre, rut, fc, saldo, fs) in enumerate(PERSONAL, start=2):
    ws.cell(i, 1).value = nombre
    ws.cell(i, 2).value = rut
    ws.cell(i, 3).value = fc
    ws.cell(i, 4).value = saldo
    ws.cell(i, 5).value = fs
    ws.cell(i, 6).value = "Faltan documentos" if saldo == 0 else "Según control de vacaciones"

# Ajustar anchos
for col, w in enumerate([35, 15, 14, 22, 14, 28], start=1):
    ws.column_dimensions[chr(64+col)].width = w

print(f"Hoja '{SHEET}' creada con {len(PERSONAL)} trabajadores\n")

# Mostrar cálculo en tiempo real
hoy = date.today()
print(f"Cálculo de vacaciones al {hoy}:\n")
print(f"{'NOMBRE':<35} {'CONTRATO':<12} {'SALDO BASE':<10} {'MESES NUEVO':<11} {'DÍAS ACUM':<10} {'TOTAL PEND':<10}")
print("=" * 100)

DIAS_X_MES = 15.0 / 12.0  # 1.25 días/mes legal Chile

for nombre, rut, fc, saldo, fs in PERSONAL:
    meses_desde_saldo = (hoy.year - fs.year) * 12 + (hoy.month - fs.month)
    dias_acumulados = meses_desde_saldo * DIAS_X_MES
    total = saldo + dias_acumulados
    print(f"{nombre[:34]:<35} {fc.isoformat():<12} {saldo:>9.2f} {meses_desde_saldo:>11} {dias_acumulados:>10.2f} {total:>10.2f}")

wb.save(EXCEL_PATH)
wb.close()
print("\nDone!")
