"""Alinea Pacific Nuts con Valbifrut según la cláusula comparativa (±5%).

Tras nivelar el adelanto de octubre, Pacific quedaba en 2,300 USD/kg contra
2,100 de Valbifrut: 9,5% de diferencia, fuera de la cláusula. Se ajusta la
liquidación de diciembre para que ambos cierren en el mismo precio.
"""
import shutil
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb

PRECIO_OBJETIVO = 2.1        # el que cierra Valbifrut
KG = 100_000
COL_EXP, COL_CUOTA, COL_USD, COL_TIPO, COL_ESTADO, COL_NOTAS = 4, 8, 10, 11, 12, 16

resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"Respaldo: {resp}\n")

wb = load_workbook(EXCEL_PATH)
ws = wb["Cosechas"]

filas, total_otras, fila_liq = [], 0.0, None
for r in range(2, ws.max_row + 1):
    if str(ws.cell(r, COL_EXP).value or "").strip() != "Pacific Nuts":
        continue
    usd = float(ws.cell(r, COL_USD).value or 0)
    tipo = str(ws.cell(r, COL_TIPO).value or "").lower()
    filas.append((r, usd, tipo))
    if "liquidacion" in tipo:
        fila_liq = r
    else:
        total_otras += usd

objetivo = KG * PRECIO_OBJETIVO
nueva_liq = round(objetivo - total_otras)

print(f"  Objetivo ({PRECIO_OBJETIVO} USD/kg) : {objetivo:>9,.0f} USD")
print(f"  Adelantos + octubre           : {total_otras:>9,.0f} USD")
print(f"  → liquidación de diciembre    : {nueva_liq:>9,.0f} USD")

if fila_liq is None:
    print("\n  ❌ No encontré la liquidación final de Pacific.")
else:
    antes = float(ws.cell(fila_liq, COL_USD).value or 0)
    ws.cell(fila_liq, COL_USD).value = nueva_liq
    ws.cell(fila_liq, COL_NOTAS).value = (
        f"⚠️ ESTIMADO. Ajustada para cerrar en {PRECIO_OBJETIVO} USD/kg, igual "
        f"que Valbifrut, por la cláusula comparativa (±5%). CONFIRMAR con la "
        f"revisión de adelantos de octubre.")
    _save_wb(wb)
    print(f"\n  Diciembre: US$ {antes:,.0f} → US$ {nueva_liq:,.0f}  "
          f"({antes - nueva_liq:+,.0f})")

wb.close()
