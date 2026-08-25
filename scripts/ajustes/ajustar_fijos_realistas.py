#!/usr/bin/env python3
"""Ajusta proyección de gastos fijos según información real:
1. LEASING = $0 (ya pagado en diciembre)
2. COSTO ENERGETICO: $5M verano (dic-mar), $4M otoño/primavera, $3M invierno (jun-ago)
3. GASTOS VEHICULOS = $2M/año (ajustar al alza)
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
from openpyxl import load_workbook
from config import EXCEL_PATH


MESES_TEMP = [
    (2026, 6), (2026, 7), (2026, 8), (2026, 9), (2026, 10), (2026, 11),
    (2026, 12), (2027, 1), (2027, 2), (2027, 3), (2027, 4), (2027, 5),
]

# Costo energético S-Invest por mes (clima Chile)
COSTO_ENERGETICO_X_MES = {
    1: 5_000_000,   # ene - verano
    2: 5_000_000,   # feb - verano
    3: 5_000_000,   # mar - verano
    4: 4_000_000,   # abr - otoño
    5: 4_000_000,   # may - otoño
    6: 3_000_000,   # jun - invierno
    7: 3_000_000,   # jul - invierno
    8: 3_000_000,   # ago - invierno
    9: 4_000_000,   # sep - primavera
    10: 4_000_000,  # oct - primavera
    11: 4_000_000,  # nov - primavera
    12: 5_000_000,  # dic - verano
}


wb = load_workbook(EXCEL_PATH)
ws_a = wb["Ajustes Manuales"]

# Eliminar ajustes anteriores de las 3 categorías
removed = 0
for r in range(ws_a.max_row, 1, -1):
    cat = ws_a.cell(r, 3).value
    if cat in ("LEASING", "COSTO ENERGETICO", "GASTOS VEHICULOS"):
        ws_a.delete_rows(r)
        removed += 1
print(f"Removidos {removed} ajustes anteriores\n")

# Buscar próxima fila libre
next_row = ws_a.max_row + 1
while ws_a.cell(next_row, 1).value:
    next_row += 1

hoy = date.today().isoformat()

# Necesitamos saber cuánto hay en histórico para cada categoría/mes
# y agregar ajustes que lo lleven al valor deseado
from collections import defaultdict
from datetime import datetime

def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try: return datetime.strptime(v[:10], fmt).date()
            except: pass
    return None

# Calcular histórico mes a mes (TEMP 25/26: jun-25 a may-26)
hist = defaultdict(lambda: defaultdict(float))

# Banco
ws_b = wb["Cuenta Banco"]
for row in ws_b.iter_rows(min_row=2, max_col=9, values_only=True):
    if not row[0]: continue
    f = _pd(row[0])
    if not f or f < date(2025, 6, 1) or f >= date(2026, 6, 1): continue
    cat = str(row[7] or "").strip().upper()
    if cat not in ("LEASING", "COSTO ENERGETICO", "GASTOS VEHICULOS"): continue
    try: cargo = float(row[3] or 0)
    except: continue
    if cargo <= 0: continue
    hist[cat][(f.year, f.month)] += cargo

# Facturas
ws_f = wb["Facturas"]
for row in ws_f.iter_rows(min_row=2, max_col=20, values_only=True):
    if not row[0]: continue
    cat = str(row[16] or "").strip().upper()
    if cat not in ("LEASING", "COSTO ENERGETICO", "GASTOS VEHICULOS"): continue
    cat_por = str(row[19] or "") if len(row) > 19 else ""
    if "NN-no-pagar" in cat_por: continue
    fecha = _pd(row[2]) if row[2] else _pd(row[0])
    if not fecha or fecha < date(2025, 6, 1) or fecha >= date(2026, 6, 1): continue
    try: total = float(row[14] or 0)
    except: continue
    hist[cat][(fecha.year, fecha.month)] += total

print("Histórico TEMP 25/26 (que el proyector replicaría):")
for cat in ("LEASING", "COSTO ENERGETICO", "GASTOS VEHICULOS"):
    tot = sum(hist[cat].values())
    print(f"  {cat}: ${tot:,.0f} ({len(hist[cat])} meses)")
print()


# Crear ajustes
def add_ajuste(y, m, cat, monto, razon):
    global next_row
    ws_a.cell(next_row, 1).value = hoy
    ws_a.cell(next_row, 2).value = f"{y}-{m:02d}"
    ws_a.cell(next_row, 3).value = cat
    ws_a.cell(next_row, 4).value = "GENERAL"
    ws_a.cell(next_row, 5).value = monto
    ws_a.cell(next_row, 6).value = razon
    ws_a.cell(next_row, 7).value = True
    next_row += 1


# 1) LEASING = $0 (anular histórico)
print("[1] LEASING (ya pagado, anular):")
for (y, m) in MESES_TEMP:
    hist_y, hist_m = y - 1, m
    hist_val = hist["LEASING"].get((hist_y, hist_m), 0)
    if hist_val > 0:
        add_ajuste(y, m, "LEASING", -hist_val, "Leasing ya pagado en dic-2025")
        print(f"  {y}-{m:02d}: -${hist_val:,.0f}")

# 2) COSTO ENERGETICO según mes
print("\n[2] COSTO ENERGETICO (S-Invest según clima):")
for (y, m) in MESES_TEMP:
    hist_y, hist_m = y - 1, m
    hist_val = hist["COSTO ENERGETICO"].get((hist_y, hist_m), 0)
    target = COSTO_ENERGETICO_X_MES[m]
    diff = target - hist_val
    if abs(diff) > 1000:
        razon = f"Ajuste S-Invest mes {m}: target ${target:,.0f} (real hist ${hist_val:,.0f})"
        add_ajuste(y, m, "COSTO ENERGETICO", diff, razon)
        print(f"  {y}-{m:02d}: hist=${hist_val:,.0f} → target=${target:,.0f} (ajuste {diff:+,.0f})")

# 3) GASTOS VEHICULOS = $2M/año = $167K/mes
print("\n[3] GASTOS VEHICULOS ($2M/año = $167K/mes):")
target_mensual = 2_000_000 / 12
for (y, m) in MESES_TEMP:
    hist_y, hist_m = y - 1, m
    hist_val = hist["GASTOS VEHICULOS"].get((hist_y, hist_m), 0)
    diff = target_mensual - hist_val
    if abs(diff) > 1000:
        add_ajuste(y, m, "GASTOS VEHICULOS", diff,
                    f"Vehiculos ${target_mensual:,.0f}/mes (TAG + seg + mant)")

print(f"  → ~$167K/mes durante 12 meses\n")

wb.save(EXCEL_PATH)
wb.close()
print("Done!")
