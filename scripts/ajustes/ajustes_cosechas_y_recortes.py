#!/usr/bin/env python3
"""Ajustes finales:
1. Agregar cerezas marzo 2027 (30,000 kg × $1.4 USD)
2. MANO DE OBRA TEMPORAL absoluto: $25M cosecha nueces (mar-may 2027) + $4M cerezas (dic 2026)
3. MATERIALES absoluto: $4M/año
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
from openpyxl import load_workbook
from config import EXCEL_PATH

USD_CLP = 904

wb = load_workbook(EXCEL_PATH)

# ─── 1) Agregar cerezas marzo 2027 a Cosechas ─────
print("[1] Agregando cerezas marzo 2027 a Cosechas...")
ws_c = wb["Cosechas"]
next_row = ws_c.max_row + 1
while ws_c.cell(next_row, 1).value:
    next_row += 1

kg = 30000
precio_usd = 1.4
monto_usd = kg * precio_usd  # 42,000 USD
monto_clp = monto_usd * USD_CLP  # 37,968,000 CLP

ws_c.cell(next_row, 1).value = 2027
ws_c.cell(next_row, 2).value = "CEREZOS"
ws_c.cell(next_row, 3).value = kg
ws_c.cell(next_row, 4).value = "Exportadora cerezas"
ws_c.cell(next_row, 5).value = kg
ws_c.cell(next_row, 6).value = precio_usd
ws_c.cell(next_row, 7).value = 1
ws_c.cell(next_row, 8).value = 1
ws_c.cell(next_row, 9).value = "2027-03-15"
ws_c.cell(next_row, 10).value = monto_usd
ws_c.cell(next_row, 11).value = "venta única"
ws_c.cell(next_row, 12).value = "esperado"
ws_c.cell(next_row, 16).value = f"Cosecha cerezas 2027: 30,000 kg × $1.4 USD"
print(f"   {kg} kg × ${precio_usd} USD = ${monto_usd:,.0f} USD = ${monto_clp:,.0f} CLP")
print(f"   Programado para 2027-03-15\n")

# ─── 2) Ajustes manuales ─────
ws_a = wb["Ajustes Manuales"]

# Eliminar ajustes anteriores relacionados (limpiar duplicados)
removed = 0
for r in range(ws_a.max_row, 1, -1):
    cat = ws_a.cell(r, 3).value
    if cat in ("MANO DE OBRA TEMPORAL", "MATERIALES"):
        ws_a.delete_rows(r)
        removed += 1
print(f"[2] Removidos {removed} ajustes anteriores de MANO OBRA TEMPORAL y MATERIALES\n")

next_row = ws_a.max_row + 1
while ws_a.cell(next_row, 1).value:
    next_row += 1

hoy = date.today().isoformat()

def add(y, m, cat, monto, razon, cultivo="GENERAL"):
    global next_row
    ws_a.cell(next_row, 1).value = hoy
    ws_a.cell(next_row, 2).value = f"{y}-{m:02d}"
    ws_a.cell(next_row, 3).value = cat
    ws_a.cell(next_row, 4).value = cultivo
    ws_a.cell(next_row, 5).value = monto
    ws_a.cell(next_row, 6).value = razon
    ws_a.cell(next_row, 7).value = True
    print(f"   {y}-{m:02d} | {cat:25} {cultivo:10} ${monto:>+15,.0f} | {razon}")
    next_row += 1

# MANO DE OBRA TEMPORAL: $25M cosecha nueces (mar-may 2027) + $4M cerezas (dic 2026)
print("[3] MANO DE OBRA TEMPORAL (cosecha):")
# Cerezas cosecha dic-2026: $4M (3,333 totes × $1,200)
totes_cerezas = 30000 / 9
costo_cerezas = totes_cerezas * 1200
add(2026, 12, "MANO DE OBRA TEMPORAL", costo_cerezas,
    f"Cosecha cerezas: {totes_cerezas:.0f} totes × $1,200 = ${costo_cerezas:,.0f}",
    cultivo="CEREZOS")
# Cosecha nueces: $25M distribuido mar-may 2027
add(2027, 3, "MANO DE OBRA TEMPORAL", 8_000_000,
    "Cosecha nueces marzo 2027 (parte 1)", cultivo="NOGALES")
add(2027, 4, "MANO DE OBRA TEMPORAL", 10_000_000,
    "Cosecha nueces abril 2027 (parte 2 - pico)", cultivo="NOGALES")
add(2027, 5, "MANO DE OBRA TEMPORAL", 7_000_000,
    "Cosecha nueces mayo 2027 (parte 3)", cultivo="NOGALES")

print(f"\n[4] MATERIALES ($4M/año = $333K/mes):")
target_mensual = 4_000_000 / 12
MESES = [(2026, m) for m in range(6, 13)] + [(2027, m) for m in range(1, 6)]
for (y, m) in MESES:
    add(y, m, "MATERIALES", target_mensual,
        "Materiales/ferretería target $4M/año")

wb.save(EXCEL_PATH)
wb.close()
print("\nDone!")
