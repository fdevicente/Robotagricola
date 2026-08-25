"""Pacific Nuts: el adelanto de agosto no se deposita (aviso del 10-ago-2026).

Los adelantos se revisan en octubre y ahí se paga el mínimo garantizado. El
monto de octubre está POR CONFIRMAR: se conserva el que había para no borrar
ingreso del modelo, pero queda marcado.
"""
import shutil
import sys
from datetime import date, datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb

EXPORTADORA = "Pacific Nuts"
CUOTA = 4
NUEVA_FECHA = date(2026, 10, 15)
COL_EXP, COL_CUOTA, COL_FECHA, COL_USD = 4, 8, 9, 10
COL_TIPO, COL_ESTADO, COL_NOTAS = 11, 12, 16

resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"Respaldo: {resp}\n")

wb = load_workbook(EXCEL_PATH)
ws = wb["Cosechas"]

for r in range(2, ws.max_row + 1):
    if str(ws.cell(r, COL_EXP).value or "").strip() != EXPORTADORA:
        continue
    try:
        if int(ws.cell(r, COL_CUOTA).value or 0) != CUOTA:
            continue
    except (TypeError, ValueError):
        continue
    if str(ws.cell(r, COL_ESTADO).value or "").lower() == "recibido":
        print(f"La cuota {CUOTA} ya figura recibida — no se toca.")
        break

    antes = ws.cell(r, COL_FECHA).value
    usd = float(ws.cell(r, COL_USD).value or 0)
    ws.cell(r, COL_FECHA).value = NUEVA_FECHA
    ws.cell(r, COL_TIPO).value = "minimo garantizado"
    ws.cell(r, COL_NOTAS).value = (
        "El adelanto de agosto NO se deposita. Los adelantos se revisan en "
        "octubre y ahí se paga el mínimo garantizado. ⚠️ MONTO POR CONFIRMAR")
    _save_wb(wb)
    print(f"Fila {r} · cuota {CUOTA} de {EXPORTADORA}")
    print(f"  fecha  {str(antes)[:10]}  →  {NUEVA_FECHA}")
    print(f"  monto  US$ {usd:,.0f}  ⚠️ por confirmar según la revisión de octubre")
    break
else:
    print(f"No encontré la cuota {CUOTA} de {EXPORTADORA}.")

wb.close()
