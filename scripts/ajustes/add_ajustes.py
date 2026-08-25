#!/usr/bin/env python3
"""Agrega ajustes manuales a la proyección."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

USD_CLP = 904
print(f"Tipo cambio: ${USD_CLP} CLP/USD\n")

# Calcular ajustes
helo_mensual = -22_291_532 / 12  # anular gasto mensual promedio
replante_total = 9 * 5_000_000   # 9 ha × $5M
liquidacion = 240_000 * 0.4 * USD_CLP  # 240k kg × $0.4 USD

print(f"Helicóptero (anulación mensual): ${helo_mensual:,.0f}")
print(f"Replante 9 ha avellanos:        ${replante_total:,.0f}")
print(f"Liquidación diciembre:           ${liquidacion:,.0f}")
print()

# Generar lista de ajustes (12 meses: jun-2026 a may-2027)
meses_helo = [
    (2026, 6), (2026, 7), (2026, 8), (2026, 9), (2026, 10), (2026, 11),
    (2026, 12), (2027, 1), (2027, 2), (2027, 3), (2027, 4), (2027, 5),
]

ajustes = []
# 1) Anular helicóptero
for y, m in meses_helo:
    ajustes.append({
        "mes": f"{y}-{m:02d}",
        "categoria": "MANTENIMIENTO HELICOPTERO",
        "cultivo": "GENERAL",
        "monto": helo_mensual,
        "razon": "Sin helicóptero - se descontinuó la operación",
    })

# 2) Replante 9 ha avellanos en agosto 2026
ajustes.append({
    "mes": "2026-08",
    "categoria": "INVERSION / REPLANTE",
    "cultivo": "AVELLANOS",
    "monto": replante_total,
    "razon": "Replante 9 ha avellanos (9 × $5M/ha)",
})

# 3) Liquidación diciembre 2026
ajustes.append({
    "mes": "2026-12",
    "categoria": "INGRESO VENTAS",
    "cultivo": "NOGALES",
    "monto": liquidacion,
    "razon": "Liquidación cierre dic 2026: 240,000 kg × $0.4 USD × $904",
})

# Escribir al Excel
wb = load_workbook(EXCEL_PATH)
ws = wb["Ajustes Manuales"]

hoy = date.today().isoformat()
start_row = 2
for i, aj in enumerate(ajustes):
    r = start_row + i
    ws.cell(r, 1).value = hoy                  # Fecha agregado
    ws.cell(r, 2).value = aj["mes"]            # Mes proyectado
    ws.cell(r, 3).value = aj["categoria"]      # Categoria
    ws.cell(r, 4).value = aj["cultivo"]        # Cultivo
    ws.cell(r, 5).value = aj["monto"]          # Monto
    ws.cell(r, 6).value = aj["razon"]          # Razón
    ws.cell(r, 7).value = True                  # Activo

print(f"Escribiendo {len(ajustes)} ajustes en hoja 'Ajustes Manuales'...")
wb.save(EXCEL_PATH)
wb.close()
print("Done!\n")

# Resumen
helo_total = helo_mensual * 12
print("=== IMPACTO NETO PROYECCIÓN ===")
print(f"  Anular helicóptero (12 meses): +${-helo_total:,.0f} (reduce egresos)")
print(f"  Replante avellanos:            -${replante_total:,.0f} (aumenta egresos)")
print(f"  Liquidación diciembre:         +${liquidacion:,.0f} (aumenta ingresos)")
print(f"  Neto:                          ${-helo_total - replante_total + liquidacion:+,.0f}")
