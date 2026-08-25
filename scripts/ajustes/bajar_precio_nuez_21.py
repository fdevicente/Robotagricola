"""Precio de la nuez 2,3 → 2,1 USD/kg: se descuenta de la liquidación de diciembre.

Los adelantos ya pagados no se tocan. La liquidación final baja 0,2 USD/kg sobre
los kilos de cada exportadora (queda en 0,3 USD/kg en vez de 0,5), y el bono de
venta (8% de las ventas) baja en consecuencia.
"""
import shutil
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb

BAJA_USD_KG = 0.2
USD_CLP_BONO = 904      # el tipo de cambio con que se calculó el bono original
PCT_BONO = 0.08

COL_KG_ASIG, COL_MONTO_USD, COL_TIPO, COL_ESTADO, COL_NOTAS = 5, 10, 11, 12, 16

resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"Respaldo: {resp}\n")

wb = load_workbook(EXCEL_PATH)
ws = wb["Cosechas"]

print("Liquidaciones finales de nueces (dic-2026):")
print(f"  {'fila':>4} {'exportadora':20} {'kg':>9} {'antes':>10} {'baja':>10} {'queda':>10}")
print("  " + "-" * 66)
baja_total_usd = 0
for r in range(2, ws.max_row + 1):
    if str(ws.cell(r, 2).value or "").upper() != "NOGALES":
        continue
    if "liquidacion" not in str(ws.cell(r, COL_TIPO).value or "").lower():
        continue
    if str(ws.cell(r, COL_ESTADO).value or "").lower() == "recibido":
        continue
    kg = float(ws.cell(r, COL_KG_ASIG).value or 0)
    antes = float(ws.cell(r, COL_MONTO_USD).value or 0)
    baja = kg * BAJA_USD_KG
    queda = max(antes - baja, 0)
    ws.cell(r, COL_MONTO_USD).value = queda
    ws.cell(r, COL_NOTAS).value = (
        f"Liquidación dic 2026 a {queda / kg:.2f} USD/kg "
        f"(precio total 2,1 en vez de 2,3)")
    baja_total_usd += antes - queda
    print(f"  {r:>4} {str(ws.cell(r, 4).value)[:20]:20} {kg:>9,.0f} "
          f"{antes:>10,.0f} {-(antes - queda):>10,.0f} {queda:>10,.0f}")

print(f"\n  Menor ingreso: {baja_total_usd:,.0f} USD "
      f"= ${baja_total_usd * USD_CLP_BONO:,.0f} CLP (a {USD_CLP_BONO})")

# ── Bono venta nueces: 8% de las ventas, baja con la venta ──
wsa = wb["Ajustes Manuales"]
for r in range(2, wsa.max_row + 1):
    if "BONO" not in str(wsa.cell(r, 3).value or "").upper():
        continue
    antes = float(wsa.cell(r, 5).value or 0)
    base_antes = antes / PCT_BONO
    base_nueva = base_antes - baja_total_usd * USD_CLP_BONO
    nuevo = round(base_nueva * PCT_BONO)
    wsa.cell(r, 5).value = nuevo
    wsa.cell(r, 6).value = f"Bono 8% × ventas nueces ${base_nueva:,.0f} (precio 2,1 USD/kg)"
    print(f"\nBono venta nueces (fila {r}):")
    print(f"  base ventas ${base_antes:,.0f} → ${base_nueva:,.0f}")
    print(f"  bono        ${antes:,.0f} → ${nuevo:,.0f}   ({nuevo - antes:+,.0f} menos egreso)")
    break

_save_wb(wb)
wb.close()
print("\n✅ Aplicado.")
