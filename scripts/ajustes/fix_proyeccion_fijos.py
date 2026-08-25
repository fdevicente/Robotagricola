#!/usr/bin/env python3
"""Arregla proyección:
1. Calcula sueldo mensual real del personal fijo (promedio de meses pagados).
2. Agrega ajustes manuales para MANO DE OBRA PLANTA mes a mes.
3. Calcula BONO VENTA NUECES = 8% de ingresos nueces proyectados.
4. Excluye CAMBIO DIVISA / PRESTAMOS / REINTEGROS del proyector (via projector.py).
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from collections import defaultdict
from openpyxl import load_workbook
from config import EXCEL_PATH

PERSONAL_FIJO = {
    "felix de vicente": "Felix De Vicente",
    "juan parada": "Juan Parada",
    "felicito amigo": "Felicito Amigo",
    "agustin mora": "Agustin Mora",
    "patricio mora": "Patricio Mora",
    "ramiro amigo": "Ramiro Amigo",
}


def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try: return datetime.strptime(v[:10], fmt).date()
            except: pass
    return None


# ─── 1) Calcular sueldo real promedio por persona ────────
print("[1/4] Calculando sueldos reales por persona (TEMP 25/26)...")
tmp = os.path.join(tempfile.gettempdir(), "fix_fijos.xlsx")
shutil.copy2(EXCEL_PATH, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["Cuenta Banco"]

sueldos = defaultdict(lambda: defaultdict(float))  # persona -> mes -> monto
for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
    if not row[0]: continue
    f = _pd(row[0])
    if not f or f < date(2025, 6, 1) or f >= date(2026, 6, 1): continue
    cat = str(row[7] or "").upper()
    if cat != "MANO DE OBRA PLANTA": continue
    try: cargo = float(row[3] or 0)
    except: continue
    if cargo <= 0: continue
    desc = str(row[1] or "").lower()
    for clave, nombre in PERSONAL_FIJO.items():
        if clave in desc:
            sueldos[nombre][(f.year, f.month)] += cargo
            break

wb.close()

print()
sueldo_mensual_total = 0
for persona, meses in sorted(sueldos.items()):
    total = sum(meses.values())
    n_meses = len(meses)
    prom = total / max(n_meses, 1)
    sueldo_mensual_total += prom
    print(f"   {persona:25} {n_meses} meses pagados | promedio mensual: ${prom:>12,.0f}")

print(f"\n   SUELDO MENSUAL TOTAL: ${sueldo_mensual_total:,.0f}")
print(f"   ANUAL × 13 (con aguinaldo): ${sueldo_mensual_total * 13:,.0f}")


# ─── 2) Calcular ingresos proyectados de nueces para BONO ────
print("\n[2/4] Calculando bono nueces (8% ventas)...")
wb = load_workbook(tmp, read_only=True, data_only=True)
ws_c = wb["Cosechas"]
USD_CLP = 904
total_venta_nueces = 0
for row in ws_c.iter_rows(min_row=2, max_col=16, values_only=True):
    if not row[0]: continue
    cultivo = str(row[1] or "").upper()
    if cultivo != "NOGALES": continue
    # Considerar lo de TEMP 26/27 = jun-26 a may-27
    fecha = _pd(row[12]) if row[11] == "recibido" else _pd(row[8])
    if not fecha or fecha < date(2026, 6, 1): continue
    if row[11] == "recibido":
        monto = float(row[13] or 0)
        moneda = str(row[14] or "CLP").upper()
        clp = monto if moneda == "CLP" else monto * USD_CLP
    else:
        clp = float(row[9] or 0) * USD_CLP
    total_venta_nueces += clp

bono_total = total_venta_nueces * 0.08
print(f"   Ventas nueces TEMP 26/27 proyectadas: ${total_venta_nueces:,.0f}")
print(f"   Bono 8%: ${bono_total:,.0f}")

# Bono se paga típicamente en octubre (post-cosecha completa)
wb.close()


# ─── 3) Limpiar ajustes anteriores y crear nuevos ─────
print("\n[3/4] Actualizando ajustes manuales...")
wb = load_workbook(EXCEL_PATH)
ws_a = wb["Ajustes Manuales"]

# Eliminar ajustes anteriores de MANO DE OBRA PLANTA y BONO VENTA NUECES
removed = 0
for r in range(ws_a.max_row, 1, -1):
    cat = ws_a.cell(r, 3).value
    if cat in ("MANO DE OBRA PLANTA", "BONO VENTA NUECES"):
        ws_a.delete_rows(r)
        removed += 1
print(f"   Removidos {removed} ajustes anteriores")

# Buscar próxima fila libre
next_row = ws_a.max_row + 1
while ws_a.cell(next_row, 1).value:
    next_row += 1

hoy = date.today().isoformat()
MESES_TEMP = [
    (2026, 6), (2026, 7), (2026, 8), (2026, 9), (2026, 10), (2026, 11),
    (2026, 12), (2027, 1), (2027, 2), (2027, 3), (2027, 4), (2027, 5),
]

# Ajuste: completar MANO DE OBRA PLANTA al sueldo real
# Estrategia: para cada mes, agregar ajuste = sueldo_mensual_total
# Previred ~$2M aparte ya está en histórico
# Aguinaldo en septiembre y diciembre (medio sueldo extra cada uno)
aguinaldo_sep_dic = sueldo_mensual_total * 0.5  # medio mes en sep y dic

# El proyector ya tiene algo de histórico - usaremos ajustes adicionales para llegar
# al sueldo real. Mejor: limpiar primero histórico y poner ajustes completos.
# Pero como histórico viene del banco con datos reales, mejor agregar la diferencia.

# Calcular faltante por mes
# Mes histórico (mes-12 atrás) → cuánto se pagó
sueldos_hist = defaultdict(float)
for persona, meses in sueldos.items():
    for ym, monto in meses.items():
        sueldos_hist[ym] += monto

# Determinar el "ideal" mensual = sueldo_mensual_total
# Para cada mes futuro, ajustar la diferencia
ajustes_creados = 0
for (y, m) in MESES_TEMP:
    # Mes histórico equivalente: mes anterior
    hist_y, hist_m = y - 1, m
    hist_pagado = sueldos_hist.get((hist_y, hist_m), 0)
    ideal = sueldo_mensual_total
    # Sep y dic: ajustar para aguinaldo
    if m in (9, 12):
        ideal += aguinaldo_sep_dic
    diff = ideal - hist_pagado
    if abs(diff) > 1000:
        ws_a.cell(next_row, 1).value = hoy
        ws_a.cell(next_row, 2).value = f"{y}-{m:02d}"
        ws_a.cell(next_row, 3).value = "MANO DE OBRA PLANTA"
        ws_a.cell(next_row, 4).value = "GENERAL"
        ws_a.cell(next_row, 5).value = diff
        ws_a.cell(next_row, 6).value = f"Ajuste sueldos personal fijo (real ${ideal:,.0f} vs hist ${hist_pagado:,.0f})"
        ws_a.cell(next_row, 7).value = True
        next_row += 1
        ajustes_creados += 1

print(f"   Ajustes MANO DE OBRA PLANTA creados: {ajustes_creados}")

# Bono nueces: aplicar 100% en octubre (post-cosecha)
if bono_total > 0:
    ws_a.cell(next_row, 1).value = hoy
    ws_a.cell(next_row, 2).value = "2026-10"
    ws_a.cell(next_row, 3).value = "BONO VENTA NUECES"
    ws_a.cell(next_row, 4).value = "NOGALES"
    ws_a.cell(next_row, 5).value = bono_total
    ws_a.cell(next_row, 6).value = f"Bono 8% sobre ventas nueces TEMP 26/27 (${total_venta_nueces:,.0f})"
    ws_a.cell(next_row, 7).value = True
    print(f"   Ajuste BONO VENTA NUECES creado: +${bono_total:,.0f} en oct-2026")

wb.save(EXCEL_PATH)
wb.close()
print("\n[4/4] Done!")
