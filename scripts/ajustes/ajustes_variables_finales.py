#!/usr/bin/env python3
"""Aplica ajustes finales para variables:
1. Re-categorizar Lipigas mal puesto en MANO DE OBRA TEMPORAL → COMBUSTIBLE
2. Anular duplicados banco+factura (Alpabesa, Jorge Bravo)
3. Excluir arriendo excavadora 2025 único ($30M)
4. Agregar arriendo excavadora para avellanos (180 hrs × $55K + IVA = $11.78M)
5. Agregar gas anual de cosecha (~$14M en mayo 2027)
"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH


def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None


wb = load_workbook(EXCEL_PATH)

# ─── 1) Re-categorizar Lipigas mal puesto ─────
print("[1] Re-categorizando Lipigas como COMBUSTIBLE...")
ws = wb["Cuenta Banco"]
changed = 0
for r in range(2, ws.max_row + 1):
    if not ws.cell(r, 1).value: continue
    desc = str(ws.cell(r, 2).value or "")
    cat = str(ws.cell(r, 8).value or "").strip().upper()
    if "lipigas" in desc.lower() and cat == "MANO DE OBRA TEMPORAL":
        ws.cell(r, 8).value = "COMBUSTIBLE"
        changed += 1
        print(f"  Fila {r}: {desc[:50]} → COMBUSTIBLE")
print(f"  Total: {changed}\n")

# ─── 2) Ajustes manuales ─────
print("[2-5] Agregando ajustes manuales...")
ws_a = wb["Ajustes Manuales"]

# Eliminar ajustes anteriores relacionados
removed = 0
for r in range(ws_a.max_row, 1, -1):
    razon = str(ws_a.cell(r, 6).value or "")
    if "arriendo excavadora" in razon.lower() or \
       "alpabesa dup" in razon.lower() or \
       "jorge bravo dup" in razon.lower() or \
       "gas anual cosecha" in razon.lower():
        ws_a.delete_rows(r)
        removed += 1
print(f"  Removidos {removed} ajustes anteriores\n")

next_row = ws_a.max_row + 1
while ws_a.cell(next_row, 1).value:
    next_row += 1

hoy = date.today().isoformat()

def add(y, m, cat, monto, razon, cultivo="GENERAL"):
    global next_row
    ws_a.cell(next_row, 1).value = hoy
    ws_a.cell(next_row, 2).value = f"{y}-{m:02d}"
    ws_a.cell(next_row, 3).value = cat
    ws_a.cell(next_row, 4).value = cultivo
    ws_a.cell(next_row, 5).value = monto
    ws_a.cell(next_row, 6).value = razon
    ws_a.cell(next_row, 7).value = True
    print(f"  {y}-{m:02d} | {cat:30} {cultivo:10} ${monto:>+15,.0f} | {razon}")
    next_row += 1

# 2) Anular duplicado Alpabesa F677 mayo-2026 → proyectado a mayo-2027
# La factura $23.9M aparece duplicada con los cargos banco $29.6M
add(2027, 5, "MANO DE OBRA TEMPORAL", -23876160,
    "Anular duplicado Alpabesa F677 (factura + cargos banco)")

# Anular duplicado Jorge Bravo F2493 sep-2025 → proyectado a sep-2026
# Factura $13M + cargos banco $15M
add(2026, 9, "MAQUINARIA - MANTENCION", -13051280,
    "Anular duplicado Jorge Bravo F2493 (factura + cargos banco)")

# 3) Anular arriendo excavadora 2025 (único, no se repite)
# $30M total = 3 cargos banco $5M + factura $13M (= $28M aprox, $30M con duplicado)
# El recorte que mencionó el usuario: ese arriendo fue de 2025 único
# Sumando los 3 cargos banco que NO fueron duplicados con factura: hay $15M cargos extra
# Restamos -$15M de MAQUINARIA en sep-2026
add(2026, 9, "MAQUINARIA - MANTENCION", -15000000,
    "Anular arriendo excavadora 2025 (no se repite - era único)")

# 4) Nuevo arriendo excavadora para avellanos (replante)
# 180 hrs × $55,000/hr + IVA 19%
hrs = 180
precio_hr = 55000
neto = hrs * precio_hr  # 9,900,000
iva = neto * 0.19  # 1,881,000
total_excav = neto + iva  # 11,781,000
add(2026, 8, "INVERSION / REPLANTE", total_excav,
    f"Arriendo excavadora avellanos {hrs}h × ${precio_hr:,} + IVA = ${total_excav:,.0f}",
    cultivo="AVELLANOS")

# 5) Gas anual cosecha ~$14M al final temporada 26/27
add(2027, 5, "COMBUSTIBLE", 14000000,
    "Gas anual cosecha TEMP 26/27 (final temporada)")

wb.save(EXCEL_PATH)
wb.close()
print("\nDone!")
