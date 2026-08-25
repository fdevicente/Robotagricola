#!/usr/bin/env python3
"""Ajustes finales basados en comparación con flujo 2024:
1. +$5M COMBUSTIBLE (gas adicional)
2. +$X CAJA CHICA / IMPREVISTOS (aumentar)
3. Mantener NOGALTEC SUR (asesoria nogales) tal como está
4. NO agregar: Vitakai, Biopinon, Caballos, Importación, Wiseconn, Seguro casa
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
from openpyxl import load_workbook
from config import EXCEL_PATH

wb = load_workbook(EXCEL_PATH)
ws_a = wb["Ajustes Manuales"]

next_row = ws_a.max_row + 1
while ws_a.cell(next_row, 1).value:
    next_row += 1

hoy = date.today().isoformat()


def add(y, m, cat, monto, razon, cultivo="GENERAL"):
    global next_row
    ws_a.cell(next_row, 1).value = hoy
    ws_a.cell(next_row, 2).value = f"{y}-{m:02d}"
    ws_a.cell(next_row, 3).value = cat
    ws_a.cell(next_row, 4).value = cultivo
    ws_a.cell(next_row, 5).value = monto
    ws_a.cell(next_row, 6).value = razon
    ws_a.cell(next_row, 7).value = True
    print(f"  {y}-{m:02d} | {cat:30} ${monto:>12,.0f} | {razon}")
    next_row += 1


# 1) +$5M COMBUSTIBLE adicional (gas) — agregar a mayo 2027 (cosecha)
print("[1] +$5M COMBUSTIBLE adicional (gas):")
add(2027, 5, "COMBUSTIBLE", 5_000_000,
    "Gas adicional cosecha TEMP 26/27 ($5M extra)")

# 2) Aumentar CAJA CHICA / IMPREVISTOS
# Histórico: $4M proyectado. Usuario dice "capaz habria que considerar mas".
# Sugerencia: $7M/año = ~$583K/mes
print("\n[2] CAJA CHICA / IMPREVISTOS aumentado a $7M/año:")
target_mensual = 7_000_000 / 12
MESES = [(2026, m) for m in range(6, 13)] + [(2027, m) for m in range(1, 6)]
# Antes histórico era ~$4M. Agregamos diferencia +$3M anual = +$250K/mes
diff = (7_000_000 - 4_000_000) / 12
for (y, m) in MESES:
    add(y, m, "CAJA CHICA / IMPREVISTOS", diff,
        f"Ajuste caja chica gastos menores irregulares (target $7M/año)")

print(f"\n  +${diff:,.0f}/mes × 12 = +$3,000,000 anual")

wb.save(EXCEL_PATH)
wb.close()
print("\nDone!")
