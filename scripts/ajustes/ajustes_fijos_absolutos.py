#!/usr/bin/env python3
"""Reemplaza ajustes con valores absolutos para LEASING, COSTO ENERGETICO, GASTOS VEHICULOS.
Ahora que estas cats están excluidas del histórico, el ajuste es el monto final."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
from openpyxl import load_workbook
from config import EXCEL_PATH

MESES_TEMP = [
    (2026, 6), (2026, 7), (2026, 8), (2026, 9), (2026, 10), (2026, 11),
    (2026, 12), (2027, 1), (2027, 2), (2027, 3), (2027, 4), (2027, 5),
]

COSTO_ENERGETICO_X_MES = {
    1: 5_000_000, 2: 5_000_000, 3: 5_000_000,   # verano
    4: 4_000_000, 5: 4_000_000,                  # otoño
    6: 3_000_000, 7: 3_000_000, 8: 3_000_000,   # invierno
    9: 4_000_000, 10: 4_000_000, 11: 4_000_000, # primavera
    12: 5_000_000,                                # verano
}

GASTOS_VEHICULOS_X_MES = 2_000_000 / 12  # $166K/mes

wb = load_workbook(EXCEL_PATH)
ws_a = wb["Ajustes Manuales"]

# Eliminar ajustes anteriores
removed = 0
for r in range(ws_a.max_row, 1, -1):
    cat = ws_a.cell(r, 3).value
    if cat in ("LEASING", "COSTO ENERGETICO", "GASTOS VEHICULOS"):
        ws_a.delete_rows(r)
        removed += 1
print(f"Removidos {removed} ajustes anteriores\n")

next_row = ws_a.max_row + 1
while ws_a.cell(next_row, 1).value:
    next_row += 1

hoy = date.today().isoformat()

# COSTO ENERGETICO: valor absoluto por mes
print("COSTO ENERGETICO (S-Invest valor absoluto):")
for (y, m) in MESES_TEMP:
    monto = COSTO_ENERGETICO_X_MES[m]
    ws_a.cell(next_row, 1).value = hoy
    ws_a.cell(next_row, 2).value = f"{y}-{m:02d}"
    ws_a.cell(next_row, 3).value = "COSTO ENERGETICO"
    ws_a.cell(next_row, 4).value = "GENERAL"
    ws_a.cell(next_row, 5).value = monto
    ws_a.cell(next_row, 6).value = f"S-Invest mes {m} (clima)"
    ws_a.cell(next_row, 7).value = True
    next_row += 1
    print(f"  {y}-{m:02d}: ${monto:,.0f}")

# GASTOS VEHICULOS
print(f"\nGASTOS VEHICULOS: ${GASTOS_VEHICULOS_X_MES:,.0f}/mes × 12 = $2M/año")
for (y, m) in MESES_TEMP:
    ws_a.cell(next_row, 1).value = hoy
    ws_a.cell(next_row, 2).value = f"{y}-{m:02d}"
    ws_a.cell(next_row, 3).value = "GASTOS VEHICULOS"
    ws_a.cell(next_row, 4).value = "GENERAL"
    ws_a.cell(next_row, 5).value = GASTOS_VEHICULOS_X_MES
    ws_a.cell(next_row, 6).value = "TAG + seguros + mantención mínima"
    ws_a.cell(next_row, 7).value = True
    next_row += 1

# LEASING = $0 (no agregar nada, la exclusión del histórico ya lo deja en 0)
print("\nLEASING: $0 (excluido del proyectado, ya pagado dic-2025)")

wb.save(EXCEL_PATH)
wb.close()
print("\nDone!")
