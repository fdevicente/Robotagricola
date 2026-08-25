#!/usr/bin/env python3
"""Ajustes finales:
1. Francisco Donoso BH mensual $2.5M × 12 = $30M anual
2. PATENTES / ARRIENDOS / SEGUROS: $1.2M anual
3. BONO VENTA NUECES: 8% × total ventas TEMP 26/27
4. Inversión avellanos: usar excedente para llegar a saldo 0
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

USD_CLP = 904

wb = load_workbook(EXCEL_PATH)
ws_a = wb["Ajustes Manuales"]

# Eliminar ajustes anteriores afectados
removed = 0
for r in range(ws_a.max_row, 1, -1):
    cat = ws_a.cell(r, 3).value
    razon = str(ws_a.cell(r, 6).value or "")
    if cat in ("BONO VENTA NUECES", "ARRIENDOS / PATENTES / SEGUROS") or \
       "francisco donoso" in razon.lower():
        ws_a.delete_rows(r)
        removed += 1
print(f"[1] Removidos {removed} ajustes anteriores\n")

next_row = ws_a.max_row + 1
while ws_a.cell(next_row, 1).value:
    next_row += 1

hoy = date.today().isoformat()

def add(y, m, cat, monto, razon, cultivo="GENERAL"):
    global next_row
    ws_a.cell(next_row, 1).value = hoy
    ws_a.cell(next_row, 2).value = f"{y}-{m:02d}"
    ws_a.cell(next_row, 3).value = cat
    ws_a.cell(next_row, 4).value = cultivo
    ws_a.cell(next_row, 5).value = monto
    ws_a.cell(next_row, 6).value = razon
    ws_a.cell(next_row, 7).value = True
    next_row += 1


# 1) Francisco Donoso BH $2.5M mensual = $30M anual
# El histórico actual es ~$20M. Falta $10M / 12 = $833K/mes adicional
print("[2] Francisco Donoso BH mensual $2.5M:")
target_donoso_mensual = 2_500_000
# Histórico ya proyecta ~$1.7M/mes. Agregar diferencia para llegar a $2.5M
# Pero como SERVICIOS PROFESIONALES viene del BANCO (CATS_SOLO_BANCO), incluye el histórico
# Mejor: agregar ajuste por la diferencia mensual
diff_mensual = target_donoso_mensual - (20_000_000 / 12)  # ~$833K
MESES = [(2026, m) for m in range(6, 13)] + [(2027, m) for m in range(1, 6)]
for (y, m) in MESES:
    add(y, m, "SERVICIOS PROFESIONALES", diff_mensual,
        f"Ajuste Francisco Donoso BH $2.5M/mes (target)")
print(f"   +${diff_mensual:,.0f}/mes × 12 = +${diff_mensual * 12:,.0f}")

# 2) Patentes / Arriendos / Seguros: $1.2M anual = $100K/mes
print("\n[3] ARRIENDOS / PATENTES / SEGUROS = $1.2M anual:")
target_patentes = 1_200_000 / 12
for (y, m) in MESES:
    add(y, m, "ARRIENDOS / PATENTES / SEGUROS", target_patentes,
        "Patentes/seguros target $1.2M/año")
print(f"   ${target_patentes:,.0f}/mes × 12 = $1,200,000")

# 3) BONO VENTA NUECES: 8% × total ventas nueces TEMP 26/27
# Ventas nueces TEMP 26/27:
#   Valbifrut adelanto 1: $223.6M (recibido)
#   Valbifrut adelanto 2: $42.5M (recibido)
#   Pacific Nuts cuotas adelanto: 80k+30k+30k+20k USD × $904 = $144.6M
#   Valbifrut liquidación: $70k USD × $904 = $63.3M
#   Pacific Nuts liquidación: $50k USD × $904 = $45.2M
total_ventas_nueces = 223_596_523 + 42_483_234 + 144_640_000 + 63_280_000 + 45_200_000
bono = total_ventas_nueces * 0.08
print(f"\n[4] BONO VENTA NUECES:")
print(f"   Total ventas nueces TEMP 26/27: ${total_ventas_nueces:,.0f}")
print(f"   Bono 8% = ${bono:,.0f}")
# Bono se paga típicamente post-liquidación final (enero 2027)
add(2027, 1, "BONO VENTA NUECES", bono,
    f"Bono 8% × ventas nueces ${total_ventas_nueces:,.0f}", cultivo="NOGALES")

wb.save(EXCEL_PATH)
wb.close()
print("\nDone!")
