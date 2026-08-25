#!/usr/bin/env python3
"""Agrega vacaciones tomadas a la hoja 'Vacaciones' del Master."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date
from openpyxl import load_workbook
from config import EXCEL_PATH

# Mapeo nombre corto → nombre canónico (igual al de Vacaciones Pendientes)
CANONICO = {
    "RAMIRO AMIGO":   "Luis Ramiro Amigo Soto",
    "AGUSTIN MORA":   "Agustin Segundo Mora Hernandez",
    "FELICITO AMIGO": "Felicito Amigo Soto",
    "PATRICIO MORA":  "Luis Patricio Mora Amigo",
    "JUAN PARADA":    "Juan Parada",
    "FELIX DE VICENTE": "Felix De Vicente",
}

# Vacaciones tomadas (persona, fecha_inicio, fecha_fin, días)
VACACIONES = [
    ("RAMIRO AMIGO",   date(2026, 5, 19), date(2026, 5, 22), 3),
    ("AGUSTIN MORA",   date(2025, 3, 3),  date(2025, 3, 14), 10),
    ("FELICITO AMIGO", date(2026, 5, 22), date(2026, 5, 22), 1),
    ("PATRICIO MORA",  date(2025, 2, 10), date(2025, 2, 14), 5),
    ("AGUSTIN MORA",   date(2026, 5, 22), date(2026, 5, 22), 1),
    ("PATRICIO MORA",  date(2026, 5, 22), date(2026, 5, 22), 1),
    ("JUAN PARADA",    date(2026, 5, 22), date(2026, 5, 22), 1),
    ("AGUSTIN MORA",   date(2026, 5, 16), date(2026, 5, 20), 5),   # típo: 16-03 → 16-05
    ("JUAN PARADA",    date(2026, 2, 16), date(2026, 2, 20), 5),
    ("PATRICIO MORA",  date(2026, 1, 19), date(2026, 1, 30), 10),
    ("JUAN PARADA",    date(2026, 1, 23), date(2026, 1, 23), 1),
    ("PATRICIO MORA",  date(2025, 7, 7),  date(2025, 7, 21), 11),
    ("RAMIRO AMIGO",   date(2025, 7, 15), date(2025, 7, 22), 27),  # nota: 8 días calendario pero usuario dice 27
]

wb = load_workbook(EXCEL_PATH)
ws = wb["Vacaciones"]

# Limpiar registros previos (excepto headers)
# Pero solo si ya hay registros con CANONICO names para no duplicar
print("Eliminando registros previos...")
for r in range(ws.max_row, 1, -1):
    nombre = str(ws.cell(r, 1).value or "")
    if nombre in CANONICO.values() or nombre == "Juan Pérez":
        ws.delete_rows(r)

# Encontrar próxima fila libre
next_row = ws.max_row + 1
while ws.cell(next_row, 1).value:
    next_row += 1

print(f"\nAgregando {len(VACACIONES)} registros de vacaciones tomadas...\n")
total_por_persona = {}
for persona, fi, ff, dias in VACACIONES:
    canonico = CANONICO[persona]
    ws.cell(next_row, 1).value = canonico
    ws.cell(next_row, 2).value = fi
    ws.cell(next_row, 3).value = ff
    ws.cell(next_row, 4).value = dias
    ws.cell(next_row, 5).value = "Aprobado"
    ws.cell(next_row, 6).value = f"Periodo {fi.year}"
    total_por_persona[canonico] = total_por_persona.get(canonico, 0) + dias
    print(f"  {canonico[:35]:<35} {fi} → {ff} ({dias}d)")
    next_row += 1

print("\nTotal por persona:")
for p, d in sorted(total_por_persona.items(), key=lambda x: -x[1]):
    print(f"  {p}: {d} días")

wb.save(EXCEL_PATH)
wb.close()
print("\nDone!")
