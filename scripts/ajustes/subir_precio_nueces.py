#!/usr/bin/env python3
"""Sube $0.1 USD/kg al precio de las nueces (Valbifrut + Pacific Nuts liquidación final)."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from config import EXCEL_PATH

USD_CLP = 904
PRECIO_EXTRA_USD = 0.1

# Liquidaciones actuales: 140k Valbifrut × $0.4 + 100k Pacific × $0.4
# Nuevo: $0.5 USD/kg de liquidación = $0.4 + $0.1
VALBIFRUT_KG = 140_000
PACIFIC_KG = 100_000

valbi_new_usd = VALBIFRUT_KG * (0.4 + PRECIO_EXTRA_USD)  # 70,000 USD
pacific_new_usd = PACIFIC_KG * (0.4 + PRECIO_EXTRA_USD)   # 50,000 USD

wb = load_workbook(EXCEL_PATH)
ws = wb["Cosechas"]

actualizadas = 0
for r in range(2, ws.max_row + 1):
    if not ws.cell(r, 1).value: continue
    anio = ws.cell(r, 1).value
    exp = str(ws.cell(r, 4).value or "")
    tipo = str(ws.cell(r, 11).value or "").lower()
    estado = str(ws.cell(r, 12).value or "").lower()

    if anio != 2026 or "liquidacion" not in tipo: continue
    if estado != "esperado": continue

    if "valbifrut" in exp.lower():
        ws.cell(r, 10).value = valbi_new_usd
        ws.cell(r, 16).value = f"Liquidación dic 2026: 140k × $0.5 USD ($904 CLP/USD)"
        print(f"  Fila {r} Valbifrut: ${valbi_new_usd:,.0f} USD = ${valbi_new_usd * USD_CLP:,.0f} CLP")
        actualizadas += 1
    elif "pacific" in exp.lower():
        ws.cell(r, 10).value = pacific_new_usd
        ws.cell(r, 16).value = f"Liquidación dic 2026: 100k × $0.5 USD ($904 CLP/USD)"
        print(f"  Fila {r} Pacific Nuts: ${pacific_new_usd:,.0f} USD = ${pacific_new_usd * USD_CLP:,.0f} CLP")
        actualizadas += 1

print(f"\nActualizadas: {actualizadas} liquidaciones")
print(f"Ingreso adicional vs antes: ${(VALBIFRUT_KG + PACIFIC_KG) * PRECIO_EXTRA_USD:,.0f} USD = ${(VALBIFRUT_KG + PACIFIC_KG) * PRECIO_EXTRA_USD * USD_CLP:,.0f} CLP")

wb.save(EXCEL_PATH)
wb.close()
print("\nDone!")
