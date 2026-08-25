"""Dos correcciones sobre la cosecha 2026 (aviso del dueño, 10-ago-2026).

1. VALBIFRUT adelantó SOLO 1,8 USD/kg. La hoja tenía dos filas de adelanto
   (250.200 + 47.000 USD = 2,14/kg): son la MISMA plata en dos transferencias.
   Los $266.079.757 recibidos son los 250.200 USD a ~1.063 CLP/USD.
   Se deja la segunda fila con 0 USD para que no duplique el precio por kilo;
   el efectivo recibido no se toca.

2. PACIFIC NUTS: el adelanto de agosto no se deposita. En octubre se revisa y
   se paga el mínimo garantizado. Hay **cláusula comparativa con Valbifrut**:
   el precio no puede diferir más de ~5%. Valbifrut va en 1,8 USD/kg adelantado
   y Pacific en 1,416 → octubre debería nivelarlo.
"""
import shutil
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb

COL_EXP, COL_KG, COL_CUOTA, COL_FECHA = 4, 5, 8, 9
COL_USD, COL_TIPO, COL_ESTADO, COL_NOTAS = 10, 11, 12, 16

ADELANTO_VALBIFRUT = 1.8      # USD/kg confirmado por el dueño
KG_PACIFIC = 100_000

resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"Respaldo: {resp}\n")

wb = load_workbook(EXCEL_PATH)
ws = wb["Cosechas"]

# ── 1. Valbifrut: quitar el adelanto duplicado ──
print("=" * 62)
print("VALBIFRUT — adelanto real 1,8 USD/kg")
print("=" * 62)
adelantos = []
for r in range(2, ws.max_row + 1):
    if str(ws.cell(r, COL_EXP).value or "").strip() != "Valbifrut":
        continue
    if "adelanto" not in str(ws.cell(r, COL_TIPO).value or "").lower():
        continue
    adelantos.append((r, float(ws.cell(r, COL_USD).value or 0)))

total_antes = sum(u for _, u in adelantos)
if len(adelantos) > 1 and total_antes > 250_200:
    principal = max(adelantos, key=lambda x: x[1])
    for r, usd in adelantos:
        if r == principal[0]:
            continue
        ws.cell(r, COL_USD).value = 0
        ws.cell(r, COL_NOTAS).value = (
            "Segunda transferencia del MISMO adelanto de 1,8 USD/kg "
            "(el efectivo recibido está en la columna de monto real)")
        print(f"  fila {r}: {usd:,.0f} USD → 0   (era doble conteo)")
    print(f"  Adelanto total: {total_antes:,.0f} → {principal[1]:,.0f} USD "
          f"= {principal[1] / 139_000:.2f} USD/kg ✅")
else:
    print("  Ya estaba correcto.")

# ── 2. Pacific: nivelar octubre con Valbifrut ──
print("\n" + "=" * 62)
print("PACIFIC NUTS — mínimo garantizado de octubre")
print("=" * 62)
recibido = 0.0
fila_oct = None
for r in range(2, ws.max_row + 1):
    if str(ws.cell(r, COL_EXP).value or "").strip() != "Pacific Nuts":
        continue
    usd = float(ws.cell(r, COL_USD).value or 0)
    if str(ws.cell(r, COL_ESTADO).value or "").lower() == "recibido":
        recibido += usd
    elif "minimo garantizado" in str(ws.cell(r, COL_TIPO).value or "").lower():
        fila_oct = r

objetivo = KG_PACIFIC * ADELANTO_VALBIFRUT
falta = round(objetivo - recibido)
print(f"  Adelantado por Pacific : {recibido:>9,.0f} USD  "
      f"({recibido / KG_PACIFIC:.3f} USD/kg)")
print(f"  Nivel de Valbifrut     : {objetivo:>9,.0f} USD  "
      f"({ADELANTO_VALBIFRUT:.3f} USD/kg)")
print(f"  Diferencia a nivelar   : {falta:>9,.0f} USD")

if fila_oct:
    antes = float(ws.cell(fila_oct, COL_USD).value or 0)
    ws.cell(fila_oct, COL_USD).value = falta
    ws.cell(fila_oct, COL_NOTAS).value = (
        f"⚠️ ESTIMADO. El adelanto de agosto no se depositó; en octubre se "
        f"revisan los adelantos y se paga el mínimo garantizado. Se estima "
        f"que nivela con Valbifrut ({ADELANTO_VALBIFRUT} USD/kg) por la "
        f"cláusula comparativa (±5%). CONFIRMAR el monto real.")
    _save_wb(wb)
    print(f"\n  Octubre: US$ {antes:,.0f} → US$ {falta:,.0f}  ⚠️ estimado")
else:
    print("\n  ❌ No encontré la fila de octubre.")

wb.close()
