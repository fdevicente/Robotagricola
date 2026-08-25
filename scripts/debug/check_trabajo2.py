import sys, shutil, tempfile, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

src = r"C:\Users\Windows\Dropbox\CAMARICO 2023\BODEGA  ENTRADAS-SALIDAS fda .xlsb.xlsx"
tmp = os.path.join(tempfile.gettempdir(), "trabajo2.xlsx")
shutil.copy2(src, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["TRABAJO"]

print("HEADERS fila 7 (TODAS las 41 columnas):")
for c in range(1, 42):
    v = ws.cell(7, c).value
    if v is not None and str(v).strip():
        print(f"   Col {c}: {v}")

print("\nFila 8 (datos) cols 13-41:")
for c in range(13, 42):
    v = ws.cell(8, c).value
    if v is not None and str(v).strip():
        print(f"   Col {c}: {v}")

print("\nUltimas 15 filas con datos (para ver stock actual):")
last_data = ws.max_row
for r in range(max(2, last_data-30), last_data+1):
    prod = ws.cell(r, 13).value  # tentativo producto
    vals = [ws.cell(r, c).value for c in [1,3,7,13,14,15,16,17]]
    if any(v is not None and str(v).strip() for v in vals):
        print(f"   R{r}: {[str(v)[:16] if v is not None else '' for v in vals]}")
wb.close()
