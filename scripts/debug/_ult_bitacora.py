"""Últimos registros de bitácora: fecha del trabajo vs fecha de ingreso."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from datetime import datetime, date
from openpyxl import load_workbook
from config import EXCEL_PATH

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Bitácora"]
enc = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
print("Columnas:", enc, "\n")

filas = []
for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if r and r[0]:
        filas.append((i, r))
wb.close()

# Fecha·Hora·Tipo·Actividad·Cultivo·Sector·JH·Trabajadores ... Registrado por
print(f"{'fila':>5}  {'trabajo':10} {'hora':6} {'tipo':11} {'actividad':32} {'JH':>3}  trabajadores")
print("-" * 108)
for i, r in filas[-18:]:
    f = r[0].date() if isinstance(r[0], datetime) else r[0]
    hora = str(r[1] or "")[:5]
    trab = str(r[7] or "")
    trab = ", ".join(n.strip().split()[0] for n in trab.split(",") if n.strip())
    print(f"{i:>5}  {str(f):10} {hora:6} {str(r[2] or ''):11} "
          f"{str(r[3] or '')[:32]:32} {str(r[6] if r[6] is not None else '-'):>3}  {trab[:38]}")
