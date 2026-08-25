import sys, shutil, tempfile, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

src = r"C:\Users\Windows\Dropbox\CAMARICO 2023\BODEGA  ENTRADAS-SALIDAS fda .xlsb.xlsx"
tmp = os.path.join(tempfile.gettempdir(), "trabajo.xlsx")
shutil.copy2(src, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["TRABAJO"]
print(f"TRABAJO: {ws.max_row} filas x {ws.max_column} cols\n")
print("Primeras 40 filas (cols 1-12):")
for r in range(1, min(41, ws.max_row + 1)):
    vals = [ws.cell(r, c).value for c in range(1, min(13, ws.max_column + 1))]
    if any(v is not None and str(v).strip() for v in vals):
        print(f"  R{r}: {[str(v)[:18] if v is not None else '' for v in vals]}")
wb.close()
