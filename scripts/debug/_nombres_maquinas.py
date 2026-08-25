"""Nombres completos de las máquinas y su última lectura registrada."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from collections import defaultdict
from datetime import date, datetime

from openpyxl import load_workbook

from config import EXCEL_PATH

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Bitácora"]
enc = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
iMaq, iOdo = enc.index("Máquina"), enc.index("Odómetro")

maq = defaultdict(list)
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[0]:
        continue
    m = str(r[iMaq] or "").strip().upper()
    if not m:
        continue
    f = r[0]
    if isinstance(f, datetime):
        f = f.date()
    maq[m].append((str(f)[:10], r[iOdo], str(r[3] or "")[:30]))
wb.close()

for m in sorted(maq):
    regs = maq[m]
    odos = [(f, o) for f, o, _ in regs if o is not None]
    print(f"\n▸ {m}   ({len(regs)} registros)")
    if odos:
        ult = max(odos)
        print(f"    última lectura: {ult[1]:,.1f}  el {ult[0]}")
    else:
        print("    ⚠️ SIN NINGUNA LECTURA DE ODÓMETRO")
    for f, o, act in regs[-3:]:
        print(f"      {f}  odo={o if o is not None else '—':>10}  {act}")
