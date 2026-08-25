"""Días hábiles sin registro en la bitácora (descontando feriados)."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from datetime import date, datetime, timedelta
from openpyxl import load_workbook
from config import EXCEL_PATH
from modules.feriados import anio_cubierto, es_feriado, nombre_feriado

HOY = date.today()
DESDE = date(2026, 6, 8)          # primer día que reportó Juan
DIAS = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Bitácora"]
con_datos = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[0]:
        continue
    f = r[0]
    if isinstance(f, datetime):
        f = f.date()
    elif isinstance(f, str):
        try:
            f = datetime.strptime(f[:10], "%Y-%m-%d").date()
        except ValueError:
            continue
    if isinstance(f, date):
        con_datos[f] = con_datos.get(f, 0) + 1
wb.close()

for a in {DESDE.year, HOY.year}:
    if not anio_cubierto(a):
        print(f"⚠️ Faltan cargar los feriados de {a} en modules/feriados.py\n")

faltan, feriados, ok = [], [], 0
d = DESDE
while d <= HOY:
    if d.weekday() < 5:
        if es_feriado(d):
            feriados.append(d)
        elif d in con_datos:
            ok += 1
        else:
            faltan.append(d)
    d += timedelta(days=1)

print(f"Bitácora entre {DESDE} y {HOY}")
print(f"  Días hábiles CON registro : {ok}")
print(f"  Días hábiles SIN registro : {len(faltan)}")
print(f"  Feriados (no se cuentan)  : {len(feriados)}")
for f in feriados:
    print(f"      {f} {DIAS[f.weekday()]:9} — {nombre_feriado(f)}")

if not faltan:
    print("\n✅ No falta ningún día hábil.")
    sys.exit(0)

print("\n📋 Pedirle a Juan estos días:")
for f in faltan:
    print(f"   • {DIAS[f.weekday()]} {f.day} de "
          f"{['','enero','febrero','marzo','abril','mayo','junio','julio','agosto','septiembre','octubre','noviembre','diciembre'][f.month]}"
          f" ({f})")
