"""Busca una factura en el Master por número, proveedor o monto."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from datetime import datetime

from openpyxl import load_workbook

from config import EXCEL_PATH

CLAVES = [a.lower() for a in sys.argv[1:]] or ["6432660"]

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Facturas"]
enc = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
filas = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

print("Columnas:", [f"{i}:{v}" for i, v in enumerate(enc) if v][:22], "\n")

hits = []
for i, r in enumerate(filas, 2):
    if not r or not r[0]:
        continue
    txt = " ".join(str(x or "") for x in r).lower()
    if any(k in txt for k in CLAVES):
        hits.append((i, r))

print(f"{len(hits)} coincidencias\n")
for i, r in hits[:20]:
    f = r[0].date() if isinstance(r[0], datetime) else r[0]
    pago = r[2] or "—"
    print(f"  fila {i:>4} | emi {str(f)[:10]} | {str(r[3])[:30]:30} | "
          f"N°{str(r[6]):>12} | ${float(r[15] or 0):>12,.0f} | pago {str(pago)[:10]}")

print("\n--- Últimas 6 facturas cargadas en el Master ---")
for i, r in list(enumerate(filas, 2))[-6:]:
    if not r or not r[0]:
        continue
    f = r[0].date() if isinstance(r[0], datetime) else r[0]
    print(f"  fila {i:>4} | emi {str(f)[:10]} | {str(r[3])[:30]:30} | "
          f"N°{str(r[6]):>12} | ${float(r[15] or 0):>12,.0f}")
