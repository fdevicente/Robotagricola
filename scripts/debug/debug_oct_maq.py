#!/usr/bin/env python3
"""Investiga el $90.3M proyectado para MAQUINARIA en octubre 2026."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from collections import defaultdict
from openpyxl import load_workbook
from config import EXCEL_PATH

def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try: return datetime.strptime(v[:10], fmt).date()
            except: pass
    return None

tmp = os.path.join(tempfile.gettempdir(), "debug_oct.xlsx")
shutil.copy2(EXCEL_PATH, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)

# Octubre 2025 - de dónde se proyecta octubre 2026
print("=== MAQUINARIA octubre 2025 (base de proyección) ===\n")
ws_b = wb["Cuenta Banco"]
items = []
for row in ws_b.iter_rows(min_row=2, max_col=9, values_only=True):
    if not row[0]: continue
    fecha = _pd(row[0])
    if not fecha or fecha.year != 2025 or fecha.month != 10: continue
    cat = str(row[7] or "").upper()
    if "MAQUINARIA" not in cat: continue
    try: cargo = float(row[3] or 0)
    except: continue
    if cargo <= 0: continue
    desc = str(row[1] or "")
    items.append((fecha, cargo, desc, cat))

items.sort(key=lambda x: -x[1])
total = sum(x[1] for x in items)
print(f"Total MAQUINARIA banco oct-2025: ${total:,.0f} ({len(items)} items)\n")
for f, c, d, cat in items[:15]:
    print(f"  {f} | ${c:>13,.0f} | {cat:25} | {d[:55]}")

print("\n=== MAQUINARIA Facturas oct-2025 ===\n")
ws_f = wb["Facturas"]
fitems = []
for row in ws_f.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    fecha = _pd(row[0])
    if not fecha or fecha.year != 2025 or fecha.month != 10: continue
    cat = str(row[16] or "").upper()
    if "MAQUINARIA" not in cat: continue
    monto = float(row[14] or 0)
    fitems.append((fecha, monto, str(row[3] or ""), str(row[7] or "")[:50]))

fitems.sort(key=lambda x: -x[1])
ftotal = sum(x[1] for x in fitems)
print(f"Total facturas MAQUINARIA oct-2025: ${ftotal:,.0f} ({len(fitems)} items)\n")
for f, m, p, d in fitems[:15]:
    print(f"  {f} | ${m:>13,.0f} | {p[:25]:25} | {d}")

print(f"\nTOTAL HISTORICO oct-2025 MAQUINARIA = ${total + ftotal:,.0f}")
print(f"(Proyectado oct-2026 toma este valor como base)")

wb.close()
