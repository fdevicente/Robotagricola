"""Movimientos del banco sin categoría, y qué hizo el historial con ese proveedor."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import re
from collections import Counter
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Cuenta Banco"]

def _f(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    try: return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError): return None

def _tokens(desc):
    d = re.sub(r"TEF|REDCOMPRA|PROVEEDORE|NUEVA MORANDE|\d{7,}|[-.]", " ", desc.upper())
    return [t for t in d.split() if len(t) > 3]

filas, hist = [], {}
for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if not r or not r[0]:
        continue
    desc = str(r[1] or ""); cat = str(r[7] or "").strip()
    cargo = float(r[3] or 0); abono = float(r[4] or 0)
    if cat:
        for t in _tokens(desc):
            hist.setdefault(t, Counter())[cat] += 1
    else:
        filas.append((i, _f(r[0]), desc, cargo, abono))
wb.close()

print(f"{len(filas)} movimientos sin categoría\n")
for i, f, desc, cargo, abono in filas:
    monto = -cargo if cargo else abono
    sug = Counter()
    for t in _tokens(desc):
        sug.update(hist.get(t, {}))
    mejor = sug.most_common(2)
    pista = "  →  " + " · ".join(f"{c} ({n}×)" for c, n in mejor) if mejor else "  →  (sin antecedente)"
    print(f"fila {i:>4} | {f} | {monto:>14,.0f} | {desc[:42]:42}{pista}")
