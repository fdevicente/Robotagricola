"""¿Llegaron los adelantos de Pacific Nuts? ¿Y CRAVE duplica a Felix?"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Cuenta Banco"]
enc = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
print("Columnas:", [f"{i}:{v}" for i, v in enumerate(enc)], "\n")
iA = enc.index("Abono") if "Abono" in enc else 5
iC = enc.index("Cargo") if "Cargo" in enc else 4
iD = 1
iCat = enc.index("Categoria") if "Categoria" in enc else 7


def _f(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    try: return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError): return None


ventas, crave, felix_rem = [], [], []
for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if not r or not r[0]:
        continue
    f = _f(r[0]); desc = str(r[iD] or ""); d = desc.upper()
    ab = float(r[iA] or 0); cg = float(r[iC] or 0)
    cat = str(r[iCat] or "")
    if ab > 0 and f and f >= date(2026, 4, 1):
        ventas.append((i, f, desc[:46], ab, cat))
    if "CRAVE" in d:
        crave.append((i, f, desc[:42], cg, cat))
    if "VICENTE" in d and "REMUNERAC" in d:
        felix_rem.append((i, f, desc[:42], cg))
wb.close()

print("=== Abonos desde abril 2026 ===")
for i, f, d, a, c in sorted(ventas, key=lambda x: x[1]):
    print(f"  fila {i:>4} | {f} | ${a:>14,.0f} | {d:46} | {c}")

print("\n=== CRAVE SPA (sueldo del dueño desde jul-2026) ===")
for i, f, d, cg, c in sorted(crave, key=lambda x: (x[1] or date.min)):
    print(f"  fila {i:>4} | {f} | ${cg:>11,.0f} | {d:42} | {c}")

print("\n=== Últimas 'Remuneración ... Felix De Vicente' pagadas directo ===")
for i, f, d, cg in sorted(felix_rem, key=lambda x: (x[1] or date.min))[-6:]:
    print(f"  fila {i:>4} | {f} | ${cg:>11,.0f} | {d}")
