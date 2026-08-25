"""Listado de vacaciones pendientes por trabajador."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, "src")
from datetime import date, datetime

from dashboard_data import get_vacaciones_pendientes

HOY = date.today()
MES = ["", "enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
       "agosto", "septiembre", "octubre", "noviembre", "diciembre"]

v = sorted(get_vacaciones_pendientes(), key=lambda x: -x["total_pendiente"])

print(f"VACACIONES PENDIENTES · al {HOY.strftime('%d-%m-%Y')}\n")
print(f"{'trabajador':32}{'contrato':>12}{'años':>6}{'ganados':>9}"
      f"{'tomados':>9}{'PENDIENTES':>12}")
print("-" * 80)
tot_pend = tot_tom = tot_gan = 0.0
for x in v:
    ganados = x["saldo_base"] + x["dias_acumulados"]
    tot_gan += ganados
    tot_tom += x["dias_tomadas"]
    tot_pend += x["total_pendiente"]
    fc = x["fecha_contrato"]
    print(f"{x['nombre'][:32]:32}{str(fc)[:10]:>12}{x['anos_trabajados']:>6.1f}"
          f"{ganados:>9.2f}{x['dias_tomadas']:>9.0f}{x['total_pendiente']:>12.2f}")
print("-" * 80)
print(f"{'TOTAL':32}{'':12}{'':6}{tot_gan:>9.2f}{tot_tom:>9.0f}{tot_pend:>12.2f}")

# Cuánto cuesta ese pasivo, con el sueldo de la hoja Vacaciones Pendientes
from openpyxl import load_workbook

from config import EXCEL_PATH
wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
sueldos = {}
for row in wb["Vacaciones Pendientes"].iter_rows(min_row=2, values_only=True):
    if row and row[0] and len(row) > 7 and isinstance(row[7], (int, float)):
        sueldos[str(row[0]).strip()] = float(row[7])
wb.close()

if sueldos:
    print(f"\n{'Costo estimado del pasivo (sueldo base / 30 × días)':56}")
    print("-" * 80)
    total = 0.0
    for x in v:
        s = sueldos.get(x["nombre"].strip())
        if not s:
            print(f"  {x['nombre'][:32]:32} (sin sueldo base cargado)")
            continue
        costo = s / 30 * x["total_pendiente"]
        total += costo
        print(f"  {x['nombre'][:32]:32} ${s:>10,.0f}/mes → ${costo:>12,.0f}")
    print(f"  {'TOTAL':32} {'':11}   ${total:>12,.0f}")

print("\nNota: se acumulan 1,25 días por mes trabajado (15 al año).")
