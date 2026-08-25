#!/usr/bin/env python3
"""Inspecciona estructura de Ajustes Manuales."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from openpyxl import load_workbook
from config import EXCEL_PATH

tmp = os.path.join(tempfile.gettempdir(), "check_ajustes.xlsx")
shutil.copy2(EXCEL_PATH, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["Ajustes Manuales"]

print("Headers (fila 1):")
for col in range(1, 10):
    print(f"  Col {col}: {ws.cell(1, col).value}")

print(f"\nFilas totales: {ws.max_row}")
print("\nFilas existentes:")
for row in range(2, min(ws.max_row + 1, 12)):
    vals = [ws.cell(row, c).value for c in range(1, 8)]
    print(f"  Fila {row}: {vals}")

wb.close()
