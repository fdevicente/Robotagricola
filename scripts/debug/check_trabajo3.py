import sys, shutil, tempfile, os
sys.stdout.reconfigure(encoding='utf-8')
from collections import Counter
from openpyxl import load_workbook

src = r"C:\Users\Windows\Dropbox\CAMARICO 2023\BODEGA  ENTRADAS-SALIDAS fda .xlsb.xlsx"
tmp = os.path.join(tempfile.gettempdir(), "trabajo3.xlsx")
shutil.copy2(src, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["TRABAJO"]

# Tipos de movimiento (col 7) y temporadas (col 2)
tipos = Counter()
temporadas = Counter()
filas_con_producto = 0
for r in range(8, ws.max_row + 1):
    tipo = ws.cell(r, 7).value
    temp = ws.cell(r, 2).value
    prod = ws.cell(r, 15).value
    if tipo: tipos[str(tipo).strip()] += 1
    if temp: temporadas[str(temp).strip()] += 1
    if prod and str(prod).strip(): filas_con_producto += 1

print(f"Total filas con producto: {filas_con_producto}\n")
print("Tipos ENTRADA/SALIDA (col 7):")
for k, v in tipos.most_common():
    print(f"   {k}: {v}")
print("\nTemporadas (col 2):")
for k, v in temporadas.most_common():
    print(f"   {k}: {v}")

# Ejemplos de SALIDA y COMPRA (no STOCK) con producto y cantidad
print("\n\nEjemplos de movimientos NO-STOCK (producto + cantidades):")
count = 0
for r in range(8, ws.max_row + 1):
    tipo = str(ws.cell(r, 7).value or "").strip().upper()
    prod = ws.cell(r, 15).value
    if not prod or not str(prod).strip(): continue
    if tipo in ("STOCK", ""): continue
    fecha = ws.cell(r, 3).value
    unidad = ws.cell(r, 27).value
    cant_env = ws.cell(r, 28).value
    vol_env = ws.cell(r, 29).value
    total_kg = ws.cell(r, 34).value
    gran_total = ws.cell(r, 36).value
    destino = ws.cell(r, 8).value
    print(f"   R{r} | {tipo:8} | {str(prod)[:25]:25} | env={cant_env} vol={vol_env} {unidad} | totKG={total_kg} granTot={gran_total} | dest={destino}")
    count += 1
    if count >= 15: break

wb.close()
