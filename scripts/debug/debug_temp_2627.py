#!/usr/bin/env python3
"""Debug por qué TEMP 26/27 no muestra ingresos."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

tmp = os.path.join(tempfile.gettempdir(), "debug_temp.xlsx")
shutil.copy2(EXCEL_PATH, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)

def _parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    return None

print("=== ABONOS BANCO POST 2026-05-01 ===\n")
ws = wb["Cuenta Banco"]
for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
    if not row[0]: continue
    fecha = _parse_date(row[0])
    if not fecha or fecha < date(2026, 5, 1): continue
    try:
        cargo = float(row[3] or 0)
        abono = float(row[4] or 0)
    except: continue
    if abono <= 0: continue
    desc = str(row[1] or "")[:60]
    cat = row[7] or "(SIN CAT)"
    print(f"  {fecha} | cargo=${cargo:>12,.0f} abono=${abono:>12,.0f} | {cat[:30]} | {desc}")

print("\n=== COSECHAS RECIBIDAS (todas) ===\n")
ws_c = wb["Cosechas"]
for row in ws_c.iter_rows(min_row=2, max_col=16, values_only=True):
    if not row[0]: continue
    estado = row[11]
    if estado != "recibido": continue
    fecha = _parse_date(row[12])
    cultivo = row[1]
    exp = row[3]
    monto = row[13] or 0
    moneda = row[14] or "CLP"
    print(f"  {fecha} | {cultivo:8} | {exp:25} | ${monto:>15,.0f} {moneda}")

print("\n=== BUSCAR 'VILLA LAS DELICIAS' o 'VALBIFRUT' en banco ===\n")
for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
    if not row[0]: continue
    desc = str(row[1] or "") + " " + str(row[2] or "")
    if "valbifrut" in desc.lower() or "villa las delicias" in desc.lower() or "delicias" in desc.lower():
        fecha = _parse_date(row[0])
        try:
            cargo = float(row[3] or 0)
            abono = float(row[4] or 0)
        except: continue
        print(f"  {fecha} | cargo=${cargo:>12,.0f} abono=${abono:>12,.0f} | {desc[:80]}")

wb.close()
