"""Historial de categorías por proveedor (busca por RUT y por nombre)."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from collections import Counter
from openpyxl import load_workbook
from config import EXCEL_PATH

BUSCAR = [
    ("SmartWays",       ["77696047", "smartway"]),
    ("Agricola Don An", ["76806957", "don an"]),
    ("S-Invest 2",      ["77270528", "s invest", "s-invest"]),
    ("Comercial Alamos", ["76055408", "alamo"]),
    ("Lipigas",         ["96928510", "lipiga"]),
]

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
filas = [(str(r[1] or ""), str(r[7] or "").strip(), float(r[3] or 0), r[0])
         for r in wb["Cuenta Banco"].iter_rows(min_row=2, values_only=True)
         if r and r[0]]
wb.close()

for nombre, claves in BUSCAR:
    print(f"\n=== {nombre} ===")
    cats = Counter()
    ejemplos = []
    for desc, cat, cargo, fecha in filas:
        d = desc.lower()
        if any(k in d for k in claves):
            if cat:
                cats[cat] += 1
            ejemplos.append((fecha, desc[:44], cargo, cat or "(sin categoría)"))
    if not ejemplos:
        print("  Sin antecedentes en el banco.")
        continue
    for cat, n in cats.most_common():
        print(f"  {n:>3}×  {cat}")
    print("  Últimos:")
    for f, d, c, cat in ejemplos[-4:]:
        print(f"    {str(f)[:10]} {c:>12,.0f}  {d:44} {cat}")
