"""Posición neta con Full Ice SpA: cuánto salió y cuánto volvió."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from collections import defaultdict
from datetime import datetime

from openpyxl import load_workbook

from config import EXCEL_PATH

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
por_anio = defaultdict(lambda: {"salio": 0.0, "volvio": 0.0, "n": 0})
for r in wb["Cuenta Banco"].iter_rows(min_row=2, values_only=True):
    if not r or not r[0]:
        continue
    d = str(r[1] or "").lower()
    if "full ice" not in d and "76984635" not in str(r[2] or ""):
        continue
    f = r[0].date() if isinstance(r[0], datetime) else r[0]
    a = por_anio[f.year]
    a["salio"] += float(r[3] or 0)
    a["volvio"] += float(r[4] or 0)
    a["n"] += 1
wb.close()

print(f"{'año':>6}{'movs':>6}{'salió (cargos)':>18}{'volvió (abonos)':>18}{'neto':>16}")
print("-" * 64)
ts = tv = 0.0
for y in sorted(por_anio):
    a = por_anio[y]
    ts += a["salio"]; tv += a["volvio"]
    print(f"{y:>6}{a['n']:>6}{a['salio']:>18,.0f}{a['volvio']:>18,.0f}"
          f"{a['volvio'] - a['salio']:>16,.0f}")
print("-" * 64)
print(f"{'TOTAL':>6}{'':>6}{ts:>18,.0f}{tv:>18,.0f}{tv - ts:>16,.0f}")
neto = tv - ts
print()
if abs(neto) < 1000:
    print("  ⚖️  Está en cero: todo lo prestado volvió.")
elif neto > 0:
    print(f"  📥 Volvió ${neto:,.0f} MÁS de lo que salió → Santa Elisa recibió neto.")
else:
    print(f"  📤 Falta que vuelvan ${-neto:,.0f} → Full Ice todavía le debe a Santa Elisa.")
