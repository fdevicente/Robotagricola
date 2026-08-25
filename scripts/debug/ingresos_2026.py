#!/usr/bin/env python3
"""Lista todos los ingresos 2026 ordenados por fecha (Cosechas + Banco)."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

USD_CLP = 904

def _parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try: return datetime.strptime(val[:10], fmt).date()
            except: pass
    return None


tmp = os.path.join(tempfile.gettempdir(), "ingresos_v2.xlsx")
shutil.copy2(EXCEL_PATH, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)

ingresos = []

# Cosechas
ws_c = wb["Cosechas"]
for row in ws_c.iter_rows(min_row=2, max_col=16, values_only=True):
    if not row[0]: continue
    estado = row[11]
    cultivo = row[1]
    exportadora = row[3]
    tipo = row[10]

    if estado == "recibido":
        fecha = _parse_date(row[12])
        monto = row[13] or 0
        moneda = (row[14] or "CLP").upper()
        monto_clp = float(monto) if moneda == "CLP" else float(monto) * USD_CLP
    else:
        fecha = _parse_date(row[8])
        monto_usd = row[9] or 0
        monto_clp = float(monto_usd) * USD_CLP

    if not fecha or fecha.year != 2026 or monto_clp <= 0:
        continue
    ingresos.append({
        "fecha": fecha,
        "monto": monto_clp,
        "fuente": "COSECHA",
        "estado": estado or "esperado",
        "descripcion": f"{cultivo} - {exportadora} ({tipo})",
    })

# Banco abonos
ws_b = wb["Cuenta Banco"]
for row in ws_b.iter_rows(min_row=2, max_col=9, values_only=True):
    if not row[0]: continue
    fecha = _parse_date(row[0])
    if not fecha or fecha.year != 2026: continue
    try:
        cargo = float(row[3] or 0)
        abono = float(row[4] or 0)
    except: continue
    if abono <= 0 or cargo > 0: continue
    desc = str(row[1] or "")
    cat = row[7] or ""
    ingresos.append({
        "fecha": fecha,
        "monto": abono,
        "fuente": "BANCO",
        "estado": str(cat),
        "descripcion": desc[:60],
    })

wb.close()

ingresos.sort(key=lambda x: x["fecha"])

total_recibido = 0
total_esperado = 0

print("=" * 115)
print(f"{'FECHA':12} {'FUENTE':8} {'ESTADO':25} {'MONTO CLP':>18}  DESCRIPCION")
print("=" * 115)

for i in ingresos:
    fecha_str = i["fecha"].isoformat()
    fuente = i["fuente"]
    estado = i["estado"][:24]
    monto = i["monto"]
    desc = i["descripcion"]
    print(f"{fecha_str:12} {fuente:8} {estado:25} ${monto:>16,.0f}  {desc}")

    if "recibido" in i["estado"].lower() or fuente == "BANCO":
        total_recibido += monto
    else:
        total_esperado += monto

print("=" * 115)
print(f"\nTotal RECIBIDO (ya en cuenta):  ${total_recibido:>15,.0f}")
print(f"Total ESPERADO (proyectado):    ${total_esperado:>15,.0f}")
print(f"TOTAL INGRESOS 2026:            ${total_recibido + total_esperado:>15,.0f}")
