"""Las facturas por pagar según FXP, y si están o no en el Master.

Regla: en la columna `Saldo` de FXP, "Pagada" está lista, "NN" no se paga, y
cualquier otra cosa (un monto o una fórmula rota) es por pagar.
"""
import os
import re
import shutil
import sys
import tempfile
from datetime import date, datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from modules.correlativo import ALIAS_PROVEEDOR

FXP = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"
HOY = date.today()


def _pd(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def nro_key(n):
    s = str(n or "").strip().upper().replace(" ", "")
    if s.endswith(".0"):
        s = s[:-2]
    return re.sub(r"\D", "", s) or s


def prov_key(p):
    s = " ".join(str(p or "").upper().split())
    for g in ALIAS_PROVEEDOR:
        if any(a.upper() in s for a in g):
            return sorted(g)[0].upper()
    s = re.sub(r"\b(LTDA|SPA|S\.A\.|SA|LIMITADA|Y CIA|E HIJOS|SUR|CHILE)\b", "", s)
    s = re.sub(r"[^A-ZÑ ]", " ", s)
    return " ".join([w for w in s.split() if len(w) > 2][:2])


tmp = os.path.join(tempfile.gettempdir(), "fxp_pp.xlsx")
shutil.copy2(FXP, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
por_pagar = []
for row in wb["FXP"].iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    saldo = row[11]
    s = str(saldo).strip().upper() if saldo is not None else ""
    if s in ("PAGADA", "NN") or s == "":
        continue
    monto = float(row[8] or 0) if isinstance(row[8], (int, float)) else 0
    por_pagar.append({
        "n": row[3], "prov": str(row[6] or "").strip(),
        "nro": str(row[7] or "").strip(),
        "monto": float(saldo) if isinstance(saldo, (int, float)) else monto,
        "roto": not isinstance(saldo, (int, float)),
        "venc": _pd(row[1]), "nota": str(row[12] or "")[:38],
    })
wb.close()

# ¿Está en el Master y cómo?
wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
master = {}
for row in wb["Facturas"].iter_rows(min_row=2, values_only=True):
    if not row or not row[0] or not row[6]:
        continue
    k = (prov_key(row[3]), nro_key(row[6]))
    d = master.setdefault(k, {"pago": None, "nn": False})
    if row[2] and str(row[2]).strip():
        d["pago"] = _pd(row[2]) or row[2]
    if len(row) > 19 and "NN" in str(row[19] or "").upper():
        d["nn"] = True
wb.close()

por_pagar.sort(key=lambda x: -x["monto"])
print(f"FACTURAS POR PAGAR SEGÚN FXP · al {HOY:%d-%m-%Y}\n")
print(f"{'N°':>5} {'proveedor':26} {'documento':>12} {'monto':>13} "
      f"{'vencimiento':>13}  en el Master")
print("-" * 100)
total = vencido = 0.0
for p in por_pagar:
    total += p["monto"]
    k = (prov_key(p["prov"]), nro_key(p["nro"]))
    m = master.get(k)
    if m is None:
        est = "❌ no está"
    elif m["pago"]:
        est = f"⚠️ figura pagada {m['pago']}"
    elif m["nn"]:
        est = "⚠️ marcada NN"
    else:
        est = "✅ pendiente"
    v = ""
    if p["venc"]:
        d = (HOY - p["venc"]).days
        v = f"{p['venc']}" + (f" ({d}d)" if d > 0 else "")
        if d > 0:
            vencido += p["monto"]
    print(f"{str(p['n']):>5} {p['prov'][:26]:26} {p['nro']:>12} "
          f"${p['monto']:>12,.0f} {v:>13}  {est}"
          + ("  ⚠️ saldo con fórmula rota" if p["roto"] else ""))
print("-" * 100)
print(f"{'':5} {'TOTAL':26} {len(por_pagar):>12} facturas ${total:>12,.0f}")
print(f"{'':5} {'de las cuales vencidas':26} {'':12} ${vencido:>12,.0f}")
