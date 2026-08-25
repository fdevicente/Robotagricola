#!/usr/bin/env python3
"""Busca movimientos con Rotortec y AYV."""
import shutil, tempfile, os
from openpyxl import load_workbook
from config import EXCEL_PATH

tmp = os.path.join(tempfile.gettempdir(), "master_check.xlsx")
shutil.copy2(EXCEL_PATH, tmp)
print("File copied, opening...")

wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["Cuenta Banco"]
print("Sheet opened, iterating...")

rotortec = []
ayv = []
for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=9, values_only=True), start=2):
    if not row[0]:
        continue
    desc = str(row[1] or "")
    ref = str(row[2] or "")
    text = f"{desc} {ref}".lower()
    if "rotortec" in text:
        rotortec.append((idx, row[7], row[0], desc, row[3]))
    if "ayv" in text:
        ayv.append((idx, row[7], row[0], desc, row[3]))

wb.close()
print(f"\nROTORTEC: {len(rotortec)} movimientos")
for r in rotortec[:20]:
    cargo = float(r[4] or 0)
    print(f"  Fila {r[0]} | {str(r[2])[:10]} | Cat={r[1]} | ${cargo:>15,.0f} | {r[3][:60]}")

print(f"\nAYV: {len(ayv)} movimientos")
for r in ayv[:20]:
    cargo = float(r[4] or 0)
    print(f"  Fila {r[0]} | {str(r[2])[:10]} | Cat={r[1]} | ${cargo:>15,.0f} | {r[3][:60]}")
