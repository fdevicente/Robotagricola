"""Las notas de débito Copeval ND506808..ND506813 que agrupa un solo cargo."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook
from config import EXCEL_PATH

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Facturas"]
vistos = {}
for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    nro = str(r[6] or "").strip()
    if not nro.upper().startswith("ND5068"):
        continue
    k = (str(r[3] or ""), nro)
    d = vistos.setdefault(k, {"filas": [], "total": 0.0})
    d["filas"].append(i)
    try:
        d["total"] = max(d["total"], float(r[15] or 0))
    except (TypeError, ValueError):
        pass
wb.close()

total = 0.0
for (prov, nro), d in sorted(vistos.items(), key=lambda x: x[0][1]):
    total += d["total"]
    print(f"  {nro:12} {prov[:22]:22} ${d['total']:>10,.0f}  filas {d['filas']}")
print(f"\n  SUMA: ${total:,.0f}    ·  cargo del banco: $713.590")
