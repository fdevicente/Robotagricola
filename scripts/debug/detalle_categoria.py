#!/usr/bin/env python3
"""Detalla MAQUINARIA - MANTENCION y MANO DE OBRA TEMPORAL del histórico TEMP 25/26."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from collections import defaultdict
from openpyxl import load_workbook
from config import EXCEL_PATH

def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try: return datetime.strptime(v[:10], fmt).date()
            except: pass
    return None


tmp = os.path.join(tempfile.gettempdir(), "detalle_cat.xlsx")
shutil.copy2(EXCEL_PATH, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)


def detalle(categoria):
    print(f"\n{'='*100}")
    print(f"DETALLE: {categoria} - TEMP 25/26 (jun-2025 a may-2026)")
    print(f"{'='*100}\n")

    items = []

    # Banco
    ws = wb["Cuenta Banco"]
    for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
        if not row[0]: continue
        f = _pd(row[0])
        if not f or f < date(2025, 6, 1) or f >= date(2026, 6, 1): continue
        cat = str(row[7] or "").strip().upper()
        if cat != categoria: continue
        try: cargo = float(row[3] or 0)
        except: continue
        if cargo <= 0: continue
        items.append((f, cargo, str(row[1] or ""), "banco"))

    # Facturas
    ws_f = wb["Facturas"]
    for row in ws_f.iter_rows(min_row=2, max_col=20, values_only=True):
        if not row[0]: continue
        cat = str(row[16] or "").strip().upper()
        if cat != categoria: continue
        cat_por = str(row[19] or "") if len(row) > 19 else ""
        if "NN-no-pagar" in cat_por: continue
        f = _pd(row[2]) if row[2] else _pd(row[0])
        if not f or f < date(2025, 6, 1) or f >= date(2026, 6, 1): continue
        try: total = float(row[14] or 0)
        except: continue
        prov = str(row[3] or "")
        det = str(row[7] or "")
        items.append((f, total, f"{prov} - {det}", "factura"))

    items.sort(key=lambda x: -x[1])
    total = sum(x[1] for x in items)
    print(f"Total: ${total:,.0f} ({len(items)} items)\n")

    # Por mes
    por_mes = defaultdict(float)
    for f, m, _, _ in items:
        por_mes[(f.year, f.month)] += m
    print("Por mes:")
    for ym in sorted(por_mes.keys()):
        print(f"  {ym[0]}-{ym[1]:02d}: ${por_mes[ym]:>12,.0f}")
    print()

    # Top 25 items
    print("Top 25 items:")
    for f, m, desc, fuente in items[:25]:
        print(f"  {f} | ${m:>11,.0f} | {fuente:7} | {desc[:80]}")


detalle("MAQUINARIA - MANTENCION")
detalle("MANO DE OBRA TEMPORAL")
wb.close()
