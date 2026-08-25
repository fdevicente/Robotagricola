#!/usr/bin/env python3
"""Separa gastos proyectados TEMP 26/27 en FIJOS vs VARIABLES.
Muestra detalle de sueldos (MANO DE OBRA PLANTA) y Boletas Honorarios."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from collections import defaultdict
from openpyxl import load_workbook
from config import EXCEL_PATH
from modules.cash_flow.projector import (
    load_historical_egresos, load_ajustes_manuales, load_hectareas,
    compute_factor_hc, EXCLUIR_PROYECCION, EXCLUIR_HISTORICO_SOLO,
)

# Categorías consideradas FIJAS (recurrentes mensuales)
FIJOS = {
    "MANO DE OBRA PLANTA",        # sueldos personal fijo
    "SERVICIOS PROFESIONALES",     # asesorías recurrentes
    "ARRIENDOS / PATENTES / SEGUROS",
    "ENERGIA",                     # CGE mensual
    "COSTO ENERGETICO",            # S-Invest paneles solares
    "LEASING",
    "GASTOS VEHICULOS",            # TAG mensual + seguros
    "GASTOS BANCARIOS",
    "IMPUESTOS",                   # F29 mensual + contribuciones
}

def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try: return datetime.strptime(v[:10], fmt).date()
            except: pass
    return None


MESES = [
    (2026, 6), (2026, 7), (2026, 8), (2026, 9), (2026, 10), (2026, 11),
    (2026, 12), (2027, 1), (2027, 2), (2027, 3), (2027, 4), (2027, 5),
]

# Calcular proyectado por categoría
historicos = load_historical_egresos()
ajustes = load_ajustes_manuales()
hc = load_hectareas()
base_year = 2025

proj = defaultdict(float)
proj_mes = defaultdict(lambda: defaultdict(float))

for (y_h, m_h, cat, cul), monto in historicos.items():
    if y_h != base_year: continue
    cu = (cat or "").upper()
    if cu in EXCLUIR_PROYECCION or cu in EXCLUIR_HISTORICO_SOLO: continue
    factor = compute_factor_hc(hc, cul, base_year, base_year + 1)
    target = (base_year + 1, m_h)
    if target in MESES:
        proj[cat] += monto * factor
        proj_mes[cat][target] += monto * factor

for a in ajustes:
    ym = a["mes_proyectado"]
    if ym in MESES:
        if (a["categoria"] or "").upper() in EXCLUIR_PROYECCION: continue
        proj[a["categoria"]] += a["monto"]
        proj_mes[a["categoria"]][ym] += a["monto"]


# Clasificar
print("=" * 100)
print("GASTOS FIJOS vs VARIABLES - TEMP 26/27 (jun-2026 → may-2027)")
print("=" * 100)
print()

fijos_total = 0
fijos_list = []
var_total = 0
var_list = []
for cat, total in proj.items():
    if total == 0: continue
    if cat.upper() in FIJOS:
        fijos_total += total
        fijos_list.append((cat, total))
    else:
        var_total += total
        var_list.append((cat, total))

fijos_list.sort(key=lambda x: -x[1])
var_list.sort(key=lambda x: -x[1])

print(f"💼 GASTOS FIJOS (recurrentes mensuales):")
print(f"   {'CATEGORIA':45} {'TOTAL':>15} {'/MES':>15}")
print(f"   {'-'*45} {'-'*15} {'-'*15}")
for cat, total in fijos_list:
    print(f"   {cat:45} ${total:>13,.0f} ${total/12:>13,.0f}")
print(f"   {'-'*45} {'-'*15} {'-'*15}")
print(f"   {'TOTAL FIJOS':45} ${fijos_total:>13,.0f} ${fijos_total/12:>13,.0f}")
print()

print(f"🔄 GASTOS VARIABLES (depende temporada/cosecha):")
print(f"   {'CATEGORIA':45} {'TOTAL':>15} {'/MES PROM':>15}")
print(f"   {'-'*45} {'-'*15} {'-'*15}")
for cat, total in var_list:
    print(f"   {cat:45} ${total:>13,.0f} ${total/12:>13,.0f}")
print(f"   {'-'*45} {'-'*15} {'-'*15}")
print(f"   {'TOTAL VARIABLES':45} ${var_total:>13,.0f} ${var_total/12:>13,.0f}")
print()

print(f"💰 RESUMEN:")
print(f"   TOTAL FIJOS:     ${fijos_total:>13,.0f} ({fijos_total/(fijos_total+var_total)*100:.0f}%)")
print(f"   TOTAL VARIABLES: ${var_total:>13,.0f} ({var_total/(fijos_total+var_total)*100:.0f}%)")
print(f"   TOTAL EGRESOS:   ${fijos_total + var_total:>13,.0f}")
print()

# ─── Detalle sueldos: lista del personal fijo ─────────
print("=" * 100)
print("DETALLE: SUELDOS Y BOLETAS HONORARIOS (TEMP 25/26 últimos 12 meses)")
print("=" * 100)
print()

tmp = os.path.join(tempfile.gettempdir(), "sueldos.xlsx")
shutil.copy2(EXCEL_PATH, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["Cuenta Banco"]

sueldos = defaultdict(float)  # persona -> total año
sueldos_mes = defaultdict(lambda: defaultdict(float))  # persona -> mes -> monto
bh = defaultdict(float)
bh_count = defaultdict(int)

for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
    if not row[0]: continue
    f = _pd(row[0])
    if not f or f < date(2025, 6, 1) or f >= date(2026, 6, 1): continue
    cat = str(row[7] or "").upper()
    if cat not in ("MANO DE OBRA PLANTA", "MANO DE OBRA TEMPORAL"): continue
    desc = str(row[1] or "")
    try: cargo = float(row[3] or 0)
    except: continue
    if cargo <= 0: continue

    desc_lo = desc.lower()
    # Identificar persona
    if "remuneracion" in desc_lo or "remuneración" in desc_lo:
        # "Remuneracion Abril Felix De Vicente" -> "Felix De Vicente"
        parts = desc.split()
        # Asume formato "Remuneracion <Mes> <Nombre...>"
        if len(parts) >= 3:
            persona = " ".join(parts[2:])
        else:
            persona = desc
        sueldos[persona] += cargo
        sueldos_mes[persona][(f.year, f.month)] += cargo
    elif "bh " in desc_lo or " bh" in desc_lo or desc_lo.startswith("bh"):
        # "BH 185 Juan Parada" -> "Juan Parada"
        # "BH Francisco Donoso" -> "Francisco Donoso"
        # Quitar prefijo BH y números
        import re
        m = re.search(r"BH\s*\d*\s+(.+)", desc, re.IGNORECASE)
        persona = m.group(1).strip() if m else desc
        bh[persona] += cargo
        bh_count[persona] += 1

wb.close()

print(f"📋 SUELDOS (Remuneraciones mensuales) — Categoría: MANO DE OBRA PLANTA")
print(f"   {'PERSONA':40} {'ANUAL':>15} {'/MES':>12}")
print(f"   {'-'*40} {'-'*15} {'-'*12}")
total_sueldos = 0
for p, t in sorted(sueldos.items(), key=lambda x: -x[1]):
    meses_pagados = len(sueldos_mes[p])
    promedio = t / max(meses_pagados, 1)
    total_sueldos += t
    print(f"   {p[:40]:40} ${t:>13,.0f} ${promedio:>10,.0f} ({meses_pagados}m)")
print(f"   {'-'*40} {'-'*15} {'-'*12}")
print(f"   {'TOTAL SUELDOS ANUAL':40} ${total_sueldos:>13,.0f} ${total_sueldos/12:>10,.0f}")
print()

print(f"📝 BOLETAS HONORARIOS (BH XXX)")
print(f"   {'PERSONA':40} {'TOTAL':>15} {'#':>5}")
print(f"   {'-'*40} {'-'*15} {'-'*5}")
total_bh = 0
for p, t in sorted(bh.items(), key=lambda x: -x[1])[:30]:
    total_bh += t
    print(f"   {p[:40]:40} ${t:>13,.0f} {bh_count[p]:>5}")
total_bh_all = sum(bh.values())
print(f"   {'-'*40} {'-'*15} {'-'*5}")
print(f"   {'TOTAL BH (top 30)':40} ${total_bh:>13,.0f}")
print(f"   {'TOTAL BH (todos)':40} ${total_bh_all:>13,.0f}")
print()

print("=" * 100)
print(f"📊 RESUMEN GENERAL:")
print(f"   Sueldos anuales (personal fijo):  ${total_sueldos:>15,.0f}")
print(f"   BH (boletas honorarios):           ${total_bh_all:>15,.0f}")
print(f"   TOTAL MANO DE OBRA TEMP 25/26:    ${total_sueldos + total_bh_all:>15,.0f}")
