import sys, shutil, tempfile, os
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None

tmp = os.path.join(tempfile.gettempdir(), "rev_may2.xlsx")
shutil.copy2(EXCEL_PATH, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["Cuenta Banco"]

rev = []
for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
    if not row[0]: continue
    fecha = _pd(row[0])
    if not fecha or fecha < date(2026, 5, 1): continue
    if row[7] != "REVISAR": continue
    try: cargo = float(row[3] or 0)
    except: continue
    if cargo <= 0: continue
    rev.append((fecha, cargo, str(row[1] or "")))

rev.sort(key=lambda x: -x[1])
print(f"REVISAR mayo 2026: {len(rev)}\n")
for f, c, d in rev:
    print(f"  {f} | ${c:>12,.0f} | {d[:65]}")
wb.close()
