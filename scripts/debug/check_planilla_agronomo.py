#!/usr/bin/env python3
"""Inspecciona planilla gastos del agronomo Camarico."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from openpyxl import load_workbook

src = r"C:\Users\Windows\Dropbox\CAMARICO 2023\PLANILLA GASTOS CAMARICO 2023-2026.xlsx"
tmp = os.path.join(tempfile.gettempdir(), "planilla_agronomo.xlsx")
shutil.copy2(src, tmp)
print(f"Copiado a {tmp}\n")

wb = load_workbook(tmp, read_only=True, data_only=True)
print(f"Sheets ({len(wb.sheetnames)}):")
for s in wb.sheetnames:
    ws = wb[s]
    print(f"  - {s}: {ws.max_row} filas x {ws.max_column} cols")
print()

# Para cada hoja, mostrar headers
for sheet_name in wb.sheetnames[:10]:
    ws = wb[sheet_name]
    print(f"=== {sheet_name} ===")
    # Buscar header row en primeras 10 filas
    for r in range(1, min(8, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, min(15, ws.max_column + 1))]
        if any(vals):
            print(f"  Fila {r}: {vals[:12]}")
    print()

wb.close()
