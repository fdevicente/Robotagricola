#!/usr/bin/env python3
"""Inspecciona la pestaña 'FXP' del archivo FXP.xlsx - facturas pendientes."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from openpyxl import load_workbook

src = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"
tmp = os.path.join(tempfile.gettempdir(), "fxp_tab.xlsx")
shutil.copy2(src, tmp)

wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["FXP"]

print(f"Pestaña FXP: {ws.max_row} filas x {ws.max_column} cols\n")

# Buscar headers en primeras 15 filas
print("Primeras 15 filas:")
for r in range(1, min(16, ws.max_row + 1)):
    vals = [ws.cell(r, c).value for c in range(1, min(20, ws.max_column + 1))]
    if any(vals):
        vals_str = [str(v)[:18] if v else '' for v in vals]
        print(f"  R{r}: {vals_str}")

# Buscar fila con headers reales
print("\nBuscando fila headers (FECHA, MONTO, etc.)...")
for r in range(1, min(15, ws.max_row + 1)):
    for c in range(1, min(20, ws.max_column + 1)):
        v = str(ws.cell(r, c).value or "").upper()
        if any(kw in v for kw in ["FECHA", "EMISION", "EMISIÓN", "PROVEEDOR", "MONTO"]):
            print(f"  Hint R{r} C{c}: {ws.cell(r, c).value}")

wb.close()
