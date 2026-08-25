"""Calcula el stock consolidado por producto desde la pestaña TRABAJO.

Stock = Σ(STOCK inicial) + Σ(INGRESOS) - Σ(SALIDAS), agrupado por
(producto, unidad). Solo PREVIEW — no escribe nada.
"""
import sys, shutil, tempfile, os
sys.stdout.reconfigure(encoding='utf-8')
from collections import defaultdict
from openpyxl import load_workbook

src = r"C:\Users\Windows\Dropbox\CAMARICO 2023\BODEGA  ENTRADAS-SALIDAS fda .xlsb.xlsx"
tmp = os.path.join(tempfile.gettempdir(), "stock_calc.xlsx")
shutil.copy2(src, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["TRABAJO"]

# Columnas: 7=tipo, 15=producto, 16=ingrediente, 18=grupo, 27=unidad, 34=total kg
stock = defaultdict(lambda: {"entradas": 0.0, "salidas": 0.0, "ingrediente": "", "grupo": "", "unidad": ""})

for r in range(8, ws.max_row + 1):
    prod = ws.cell(r, 15).value
    if not prod or not str(prod).strip():
        continue
    prod = str(prod).strip()
    tipo = str(ws.cell(r, 7).value or "").strip().upper()
    try:
        cant = float(ws.cell(r, 34).value or 0)
    except (TypeError, ValueError):
        cant = 0
    unidad = str(ws.cell(r, 27).value or "").strip()
    ingr = ws.cell(r, 16).value
    grupo = ws.cell(r, 18).value

    key = prod
    if unidad and not stock[key]["unidad"]:
        stock[key]["unidad"] = unidad
    if ingr and "REF" not in str(ingr) and not stock[key]["ingrediente"]:
        stock[key]["ingrediente"] = str(ingr)[:30]
    if grupo and "REF" not in str(grupo) and not stock[key]["grupo"]:
        stock[key]["grupo"] = str(grupo)[:20]

    if tipo in ("STOCK", "INGRESOS"):
        stock[key]["entradas"] += cant
    elif tipo == "SALIDAS":
        stock[key]["salidas"] += cant

wb.close()

# Calcular saldo
resultado = []
for prod, d in stock.items():
    saldo = d["entradas"] - d["salidas"]
    resultado.append({
        "producto": prod, "unidad": d["unidad"],
        "entradas": d["entradas"], "salidas": d["salidas"],
        "saldo": saldo, "ingrediente": d["ingrediente"], "grupo": d["grupo"],
    })

# Ordenar por saldo desc
resultado.sort(key=lambda x: -x["saldo"])

print(f"PRODUCTOS ÚNICOS: {len(resultado)}\n")
print(f"{'PRODUCTO':32} {'UNID':10} {'ENTRADAS':>12} {'SALIDAS':>12} {'SALDO':>12}")
print("=" * 85)
con_saldo = [r for r in resultado if r["saldo"] > 0.01]
sin_saldo = [r for r in resultado if r["saldo"] <= 0.01]
for r in con_saldo:
    print(f"{r['producto'][:31]:32} {r['unidad'][:9]:10} {r['entradas']:>12,.2f} {r['salidas']:>12,.2f} {r['saldo']:>12,.2f}")
print(f"\n--- {len(con_saldo)} productos CON saldo positivo ---")
print(f"--- {len(sin_saldo)} productos con saldo 0 o negativo (no se mostrarían) ---")

if sin_saldo:
    print("\nProductos con saldo <= 0 (posible inconsistencia de unidades o agotados):")
    for r in sin_saldo[:15]:
        print(f"  {r['producto'][:31]:32} ent={r['entradas']:.1f} sal={r['salidas']:.1f} saldo={r['saldo']:.1f}")
