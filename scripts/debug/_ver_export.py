"""Revisa los .xlsx exportados por el conciliador."""
import glob
import os
import sys

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

patron = sys.argv[1] if len(sys.argv) > 1 else "*.xlsx"
for f in sorted(glob.glob(patron)):
    wb = load_workbook(f)
    ws = wb.active
    print(f"\n=== {os.path.basename(f)} ===")
    print(" ", ws["A1"].value)
    print(" ", ws["A2"].value)
    print(f"  {max(0, ws.max_row - 4)} filas de datos · filtro {'sí' if ws.auto_filter.ref else 'no'}")
    for r in ws.iter_rows(min_row=5, max_row=min(8, ws.max_row), values_only=True):
        docs = str(r[9] or "")[:30]
        print(f"    {str(r[0])[:10]} | {str(r[1])[:32]:32} | {r[5]:>12,.0f} | "
              f"{r[8]:11} | {docs}")
    wb.close()
