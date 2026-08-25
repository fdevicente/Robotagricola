"""¿Calzan las facturas pendientes del Master con las de FXP?

Regla del dueño: en FXP la columna `Saldo` dice "Pagada", "NN" (no se va a
pagar) o un MONTO. **Solo las que tienen monto están pendientes de pago.**
En el Master, pendiente = sin Fecha Pago y sin la marca NN.

El match es por proveedor + número (con alias): hacerlo solo por número
produce falsos positivos — distintos proveedores repiten numeración.
"""
import os
import re
import shutil
import sys
import tempfile
from datetime import date, datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from modules.correlativo import ALIAS_PROVEEDOR

FXP = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"
HOY = date.today()


def _pd(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for f in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(v[:10], f).date()
            except ValueError:
                pass
    return None


def nro_key(n):
    """Deja solo los dígitos del documento.

    En FXP el mismo papel aparece como 'ND116572' y en el Master como '116572';
    también hay 'F 6231521' y '6231521.0'. Comparar el número pelado evita
    falsos faltantes. El proveedor es el que desambigua.
    """
    s = str(n or "").strip().upper().replace(" ", "")
    if s.endswith(".0"):
        s = s[:-2]
    digitos = re.sub(r"\D", "", s)
    return digitos or s


def prov_key(p):
    """Primera palabra significativa del proveedor, con alias."""
    s = " ".join(str(p or "").upper().split())
    for grupo in ALIAS_PROVEEDOR:
        if any(a.upper() in s for a in grupo):
            return sorted(grupo)[0].upper()
    s = re.sub(r"\b(LTDA|SPA|S\.A\.|SA|LIMITADA|Y CIA|E HIJOS|SUR|CHILE)\b", "", s)
    s = re.sub(r"[^A-ZÑ ]", " ", s)
    palabras = [w for w in s.split() if len(w) > 2]
    # Con las 2 primeras palabras alcanza: "NOGALTEC SUR" ≡ "NOGALTEC",
    # "GONZALO YAN PIERR BOBADILLA" ≡ "GONZALO YAN"
    return " ".join(palabras[:2])


# ── FXP ──
tmp = os.path.join(tempfile.gettempdir(), "fxp_cmp.xlsx")
shutil.copy2(FXP, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
fxp_pend, fxp_nn, fxp_error, fxp_todas = [], [], [], {}
for row in wb["FXP"].iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    saldo = row[11]
    reg = {
        "n": row[3], "emision": _pd(row[0]), "venc": _pd(row[1]),
        "pago": _pd(row[2]), "prov": str(row[6] or "").strip(),
        "nro": str(row[7] or "").strip(),
        "monto": float(row[8] or 0) if isinstance(row[8], (int, float)) else 0,
        "nota": str(row[12] or "")[:60],
    }
    fxp_todas[(prov_key(reg["prov"]), nro_key(reg["nro"]))] = reg
    s = str(saldo).strip().upper() if saldo is not None else ""
    if isinstance(saldo, (int, float)) and saldo != 0:
        reg["saldo"] = float(saldo)
        fxp_pend.append(reg)
    elif s == "NN":
        fxp_nn.append(reg)
    elif "VALUE" in s or "REF" in s:
        fxp_error.append(reg)
wb.close()

# ── Master ──
wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Facturas"]
grupos = {}
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    if not row or not row[0]:
        continue
    prov, nro = str(row[3] or "").strip(), str(row[6] or "").strip()
    if not nro:
        continue
    k = (prov_key(prov), nro_key(nro))
    g = grupos.setdefault(k, {"prov": prov, "nro": nro, "filas": [],
                               "total": 0.0, "pago": None, "nn": False,
                               "emision": _pd(row[0])})
    g["filas"].append(i)
    try:
        g["total"] = max(g["total"], float(row[15] or 0))
    except (TypeError, ValueError):
        pass
    if row[2] and str(row[2]).strip():
        g["pago"] = _pd(row[2]) or row[2]
    if len(row) > 19 and "NN" in str(row[19] or "").upper():
        g["nn"] = True
wb.close()
master_pend = {k: g for k, g in grupos.items() if not g["pago"] and not g["nn"]}

print("=" * 78)
print("FXP")
print("=" * 78)
print(f"  Pendientes de pago (con monto) : {len(fxp_pend):>4}"
      f"   ${sum(p['saldo'] for p in fxp_pend):>14,.0f}")
print(f"  NN (no se van a pagar)         : {len(fxp_nn):>4}")
print(f"  Con la fórmula rota (#VALUE!)  : {len(fxp_error):>4}  ← hay que revisarlas")
print(f"  Total de facturas en FXP       : {len(fxp_todas):>4}")

print("\n" + "=" * 78)
print("MASTER")
print("=" * 78)
print(f"  Sin fecha de pago y sin NN     : {len(master_pend):>4}"
      f"   ${sum(g['total'] for g in master_pend.values()):>14,.0f}")
print(f"  Total de facturas en Master    : {len(grupos):>4}")

# ── Cruce ──
print("\n" + "=" * 78)
print("PENDIENTES EN FXP — ¿cómo están en el Master?")
print("=" * 78)
calzan, ya_pagada, faltan = [], [], []
for p in fxp_pend + fxp_error:
    k = (prov_key(p["prov"]), nro_key(p["nro"]))
    g = grupos.get(k)
    if not g:
        faltan.append(p)
    elif g["pago"]:
        ya_pagada.append((p, g))
    else:
        calzan.append((p, g))
print(f"  ✅ Pendiente en ambos            : {len(calzan)}")
print(f"  ⚠️ FXP pendiente / Master pagada : {len(ya_pagada)}")
print(f"  ❌ En FXP y NO está en el Master : {len(faltan)}")

if calzan:
    print("\n  --- Pendientes en los dos ---")
    for p, g in sorted(calzan, key=lambda x: -(x[0].get("saldo") or x[0]["monto"])):
        m = p.get("saldo") or p["monto"]
        venc = f"venc {p['venc']}" if p["venc"] else "sin venc"
        atraso = ""
        if p["venc"] and p["venc"] < HOY:
            atraso = f" ⚠️ {(HOY - p['venc']).days}d vencida"
        print(f"    N°{p['n']:>4} {p['prov'][:26]:26} {p['nro']:>10} "
              f"${m:>12,.0f}  {venc}{atraso}")
if ya_pagada:
    print("\n  --- FXP dice pendiente pero el Master la tiene pagada ---")
    for p, g in ya_pagada:
        print(f"    N°{p['n']:>4} {p['prov'][:26]:26} {p['nro']:>10} "
              f"${(p.get('saldo') or p['monto']):>12,.0f}  Master pagó {g['pago']}")
if faltan:
    print("\n  --- En FXP pero NO están en el Master ---")
    for p in faltan:
        print(f"    N°{p['n']:>4} {p['prov'][:26]:26} {p['nro']:>10} "
              f"${(p.get('saldo') or p['monto']):>12,.0f}  {p['nota'][:34]}")

print("\n" + "=" * 78)
print("PENDIENTES EN EL MASTER — ¿están en FXP?")
print("=" * 78)
sin_fxp, fxp_dice_pagada = [], []
for k, g in master_pend.items():
    f = fxp_todas.get(k)
    if not f:
        sin_fxp.append(g)
    elif f["pago"]:
        fxp_dice_pagada.append((g, f))
print(f"  ❌ No están en FXP               : {len(sin_fxp)}")
print(f"  ⚠️ Master pendiente / FXP pagada : {len(fxp_dice_pagada)}")

if sin_fxp:
    print("\n  --- Pendientes del Master que NO están en FXP ---")
    for g in sorted(sin_fxp, key=lambda x: -x["total"])[:25]:
        print(f"    {str(g['emision'])[:10]:10} {g['prov'][:30]:30} "
              f"{g['nro']:>12} ${g['total']:>12,.0f}")
    if len(sin_fxp) > 25:
        print(f"    … y {len(sin_fxp) - 25} más")
if fxp_dice_pagada:
    print("\n  --- El Master las tiene pendientes pero FXP dice pagada ---")
    for g, f in sorted(fxp_dice_pagada, key=lambda x: -x[0]["total"])[:20]:
        print(f"    {g['prov'][:28]:28} {g['nro']:>12} ${g['total']:>12,.0f}"
              f"   FXP pagó {f['pago']}")
    if len(fxp_dice_pagada) > 20:
        print(f"    … y {len(fxp_dice_pagada) - 20} más")
