"""Muestra filas de la bitácora completas, con el texto original."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH

DESDE = int(sys.argv[1]) if len(sys.argv) > 1 else 225

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Bitácora"]
enc = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    if i < DESDE:
        continue
    if not r or not r[0]:
        continue
    print(f"=== fila {i} ===")
    for h, v in zip(enc, r):
        if v is None or str(v).strip() == "":
            continue
        print(f"   {str(h)[:22]:22} : {str(v)[:80]}")
    print()
wb.close()
