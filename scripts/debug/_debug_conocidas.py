"""¿Por qué maquinas_conocidas() devuelve odómetros absurdos?"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from datetime import datetime

from openpyxl import load_workbook

from config import EXCEL_PATH
from modules.maquinaria import maquinas_conocidas, norm_maquina

print("=== maquinas_conocidas() ===")
for m in maquinas_conocidas():
    print(f"  {m['maquina'][:36]:36} {m['ultimo_odometro']}  {m['fecha']}")

print("\n=== Filas crudas de la Bitácora con odómetro ===")
wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Bitácora"]
enc = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
iM, iO = enc.index("Máquina"), enc.index("Odómetro")
print(f"  índices: Máquina={iM} Odómetro={iO}\n")
for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    if not r or len(r) <= iO:
        continue
    m = r[iM] if len(r) > iM else None
    o = r[iO]
    if not m or o is None:
        continue
    f = r[0]
    print(f"  fila {i:>3} | {str(f)[:10]:10} tipo={type(f).__name__:8} | "
          f"{str(m)[:32]:32} | odo={o!r} ({type(o).__name__})")
wb.close()
