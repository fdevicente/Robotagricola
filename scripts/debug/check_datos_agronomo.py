#!/usr/bin/env python3
"""Inspecciona la hoja DATOS de la planilla del agronomo."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from openpyxl import load_workbook

src = r"C:\Users\Windows\Dropbox\CAMARICO 2023\PLANILLA GASTOS CAMARICO 2023-2026.xlsx"
tmp = os.path.join(tempfile.gettempdir(), "datos_agronomo.xlsx")
shutil.copy2(src, tmp)

wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["DATOS"]

print("Headers (fila 1):")
for c in range(1, 37):
    print(f"  Col {c}: {ws.cell(1, c).value}")

print("\nMuestra primeras 5 filas con datos:")
for r in range(2, 8):
    vals = [ws.cell(r, c).value for c in range(1, 37)]
    if any(vals):
        print(f"\n  Fila {r}:")
        for c in range(1, 37):
            v = ws.cell(r, c).value
            if v is not None:
                header = ws.cell(1, c).value
                print(f"    {header}: {v}")

# Contar valores únicos de columnas categóricas claves
print("\n\nValores únicos por columna (samples):")
from collections import Counter
for col in range(1, 37):
    header = ws.cell(1, col).value
    if not header: continue
    vals = []
    for r in range(2, min(500, ws.max_row + 1)):
        v = ws.cell(r, col).value
        if v is not None:
            vals.append(str(v))
    if not vals: continue
    cnt = Counter(vals)
    if len(cnt) < 30:
        print(f"\n  {header} (col {col}) - {len(cnt)} únicos:")
        for v, n in cnt.most_common():
            print(f"    {v}: {n}")

wb.close()
