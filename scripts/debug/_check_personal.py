"""¿Coinciden los dos sistemas de vacaciones? ¿Falta alguien en las hojas?"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, "src")
from openpyxl import load_workbook
from config import EXCEL_PATH
from dashboard_data import get_vacaciones_pendientes

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
personal = {}
for r in wb["Personal"].iter_rows(min_row=2, values_only=True):
    if r[0]:
        personal[str(r[0]).strip()] = r[4]
pendientes = {str(r[0]).strip() for r in
              wb["Vacaciones Pendientes"].iter_rows(min_row=2, values_only=True) if r[0]}
# Quién aparece trabajando en la bitácora
en_bitacora = set()
for r in wb["Bitácora"].iter_rows(min_row=2, values_only=True):
    for n in str(r[7] or "").split(","):
        if n.strip():
            en_bitacora.add(n.strip())
wb.close()

tiempo_real = {v["nombre"].strip(): v for v in get_vacaciones_pendientes()}

print(f"{'Trabajador':34} {'Personal':>9} {'TiempoReal':>11}  ¿calza?")
print("-" * 70)
for n in sorted(set(personal) | set(tiempo_real)):
    p = personal.get(n)
    t = tiempo_real.get(n, {}).get("total_pendiente")
    if p is None or t is None:
        marca = "❌ falta en una hoja"
    else:
        marca = "✅" if abs(float(p) - float(t)) < 0.02 else f"⚠️ dif {float(p)-float(t):+.2f}"
    print(f"{n[:34]:34} {str(p if p is not None else '—'):>9} "
          f"{(f'{t:.2f}' if t is not None else '—'):>11}  {marca}")

print("\nGente que aparece en la BITÁCORA pero no en Personal:")
faltan = {n for n in en_bitacora if n not in personal}
print("  " + (", ".join(sorted(faltan)) if faltan else "(ninguna)"))
print("\nEn Personal pero NO en 'Vacaciones Pendientes' (invisibles al dashboard):")
inv = [n for n in personal if n not in pendientes]
print("  " + (", ".join(sorted(inv)) if inv else "(ninguna)"))
