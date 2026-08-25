"""Estructura de la pestaña FXP y qué valores toma la columna de estado."""
import os
import shutil
import sys
import tempfile
from collections import Counter

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

FXP = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"

tmp = os.path.join(tempfile.gettempdir(), "fxp_ver.xlsx")
shutil.copy2(FXP, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
print("Hojas:", ", ".join(wb.sheetnames), "\n")

ws = wb["FXP"]
filas = list(ws.iter_rows(values_only=True))
wb.close()

print(f"=== FXP · {len(filas)} filas ===\n")
enc = filas[0]
for i, v in enumerate(enc):
    if v is not None:
        print(f"  col {i:>2} : {v}")

print("\n--- Primeras 4 filas de datos ---")
for row in filas[1:5]:
    print("  ", [str(v)[:18] if v is not None else "" for v in row[:16]])

print("\n--- Últimas 4 filas con datos ---")
utiles = [r for r in filas[1:] if r and r[0]]
for row in utiles[-4:]:
    print("  ", [str(v)[:18] if v is not None else "" for v in row[:16]])

# ¿Qué valores toma cada columna candidata a "estado"?
for col in (11, 12, 13):
    if col >= len(enc):
        continue
    vals = Counter()
    for r in utiles:
        if col < len(r):
            v = r[col]
            vals[type(v).__name__ if not isinstance(v, str) else v.strip().upper()[:16]] += 1
    print(f"\n--- Valores de la col {col} ({enc[col]}) ---")
    for v, n in vals.most_common(12):
        print(f"    {str(v)[:28]:28} {n}")
print(f"\nFilas con datos: {len(utiles)}")
