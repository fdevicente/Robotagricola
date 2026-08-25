"""Cruza el detalle de vacaciones de la pestaña VACAC 26 contra su resumen.

El detalle trae varios años mal tipeados (2026 dentro del bloque 2024, etc.),
así que el año se toma del ENCABEZADO DE BLOQUE ("AÑO 24"), no de las fechas.
La tabla resumen de la derecha es la que mantiene Juan.
"""
import re
import sys
from collections import defaultdict

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

RUTA = (r"C:\Users\Windows\Dropbox\CAMARICO 2023"
        r"\ASISTENCIA TEMP 2023-2024 fda JUAN PARADA "
        r"(Copia en conflicto de juan parada 2025-03-06).xlsx")


def norm(n):
    return " ".join(str(n or "").upper().split())


wb = load_workbook(RUTA, read_only=True, data_only=True)
ws = wb["VACAC 26"]
filas = list(ws.iter_rows(values_only=True))
wb.close()

# ── Detalle: año del bloque + nombre + días ──
detalle = defaultdict(lambda: defaultdict(float))
registros = defaultdict(list)
anio_bloque = None
for row in filas:
    if not row:
        continue
    c0 = str(row[0] or "").strip().upper()
    m = re.match(r"AÑO\s*(\d{2})", c0)
    if m:
        anio_bloque = 2000 + int(m.group(1))
        continue
    if not row[2] or c0 in ("", "NOMBRE"):
        continue
    try:
        dias = float(row[5] or 0)
    except (TypeError, ValueError):
        continue
    if dias <= 0:
        continue
    n = norm(row[2])
    anio = anio_bloque
    detalle[n][anio] += dias
    registros[n].append((anio, row[3], row[4], dias, str(row[0] or "")))

# ── Resumen de la derecha (cols 11-17) ──
resumen, contrato = {}, {}
for row in filas:
    if not row or len(row) < 17:
        continue
    n = norm(row[10])
    if not n or n == "NOMBRES":
        continue
    try:
        resumen[n] = {a: float(row[11 + i] or 0)
                      for i, a in enumerate((2023, 2024, 2025, 2026))}
        resumen[n]["TOTAL"] = float(row[15] or 0)
    except (TypeError, ValueError):
        continue
    contrato[n] = row[16]

print("=" * 84)
print("DETALLE vs RESUMEN (la tabla que mantiene Juan)")
print("=" * 84)
print(f"{'trabajador':18}{'2023':>7}{'2024':>7}{'2025':>7}{'2026':>7}"
      f"{'total':>8}{'resumen':>9}  ¿calza?")
print("-" * 84)
for n in sorted(resumen):
    d = detalle.get(n, {})
    fila = [d.get(a, 0) for a in (2023, 2024, 2025, 2026)]
    suma = sum(fila)
    rt = resumen[n]["TOTAL"]
    ok = "✅" if abs(suma - rt) < 0.01 else f"❌ dif {suma - rt:+.0f}"
    print(f"{n[:18]:18}" + "".join(f"{v:>7.0f}" for v in fila)
          + f"{suma:>8.0f}{rt:>9.0f}  {ok}")

print(f"\n{'':18}{'':28}{'  por año según el resumen:'}")
for n in sorted(resumen):
    r = resumen[n]
    print(f"  {n[:18]:18} 2023={r[2023]:>3.0f}  2024={r[2024]:>3.0f}  "
          f"2025={r[2025]:>3.0f}  2026={r[2026]:>3.0f}  →  TOTAL {r['TOTAL']:>3.0f}"
          f"   contrato {contrato.get(n)}")

# Filas con el año de las fechas distinto al del bloque
print("\n" + "=" * 84)
print("FILAS CON AÑO SOSPECHOSO (la fecha no calza con su bloque)")
print("=" * 84)
for n, regs in sorted(registros.items()):
    for anio, desde, hasta, dias, _c in regs:
        for f, et in ((desde, "desde"), (hasta, "hasta")):
            if hasattr(f, "year") and f.year != anio:
                print(f"  {n[:16]:16} bloque {anio} · {et} {f:%d-%m-%Y} · {dias:g} días")
                break
