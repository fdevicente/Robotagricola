"""Busca una entidad en banco, facturas y cosechas."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH

CLAVES = [c.lower() for c in (sys.argv[1:] or ["full ice", "76984635"])]

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)


def _hit(*campos):
    txt = " ".join(str(c or "") for c in campos).lower()
    return any(k in txt for k in CLAVES)


print("=== Cuenta Banco ===")
n = 0
for i, r in enumerate(wb["Cuenta Banco"].iter_rows(min_row=2, values_only=True), 2):
    if not r or not r[0] or not _hit(r[1], r[2]):
        continue
    n += 1
    monto = -(float(r[3] or 0)) if r[3] else float(r[4] or 0)
    print(f"  fila {i} | {str(r[0])[:10]} | {monto:>14,.0f} | {str(r[1])[:44]:44} | {r[7] or '(sin cat)'}")
print(f"  → {n} movimientos")

print("\n=== Facturas ===")
n = 0
for i, r in enumerate(wb["Facturas"].iter_rows(min_row=2, values_only=True), 2):
    if not r or not r[0] or not _hit(r[3], r[6]):
        continue
    n += 1
    if n <= 8:
        print(f"  fila {i} | {str(r[0])[:10]} | {str(r[3])[:30]:30} | N°{r[6]} | {r[15]}")
print(f"  → {n} líneas")

print("\n=== Cosechas (exportadoras) ===")
for r in wb["Cosechas"].iter_rows(min_row=2, values_only=True):
    if r and r[0] and _hit(r[3]):
        print(f"  {r[0]} {r[1]} {r[3]} — {r[9]} USD {r[11]}")
wb.close()
