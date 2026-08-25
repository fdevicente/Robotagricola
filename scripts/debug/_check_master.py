import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook
from config import EXCEL_PATH
wb = load_workbook(EXCEL_PATH, read_only=True)
print("Hojas:", len(wb.sheetnames))
for n in wb.sheetnames:
    print(f"   {n}: {wb[n].max_row} filas")
wb.close()
