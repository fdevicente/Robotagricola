#!/usr/bin/env python3
"""Calcula gasto promedio mensual del helicoptero (últimos 12 meses)."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from collections import defaultdict
from openpyxl import load_workbook
from config import EXCEL_PATH

tmp = os.path.join(tempfile.gettempdir(), "check_helo.xlsx")
shutil.copy2(EXCEL_PATH, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)

# Buscar gastos helicóptero en facturas y banco
total_facturas = 0
count_facturas = 0
total_banco = 0
count_banco = 0
por_mes = defaultdict(float)

def _parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    return None

# Facturas
ws_f = wb["Facturas"]
for row in ws_f.iter_rows(min_row=2, max_col=20, values_only=True):
    if not row[0]: continue
    cat = row[16] if len(row) > 16 else None  # Col Q = 17
    if cat != "MANTENIMIENTO HELICOPTERO":
        continue
    fecha = _parse_date(row[0])
    if not fecha or fecha < date(2024, 1, 1):
        continue
    try:
        monto = float(row[15] or 0)
    except: monto = 0
    total_facturas += monto
    count_facturas += 1

# Banco
ws_b = wb["Cuenta Banco"]
for row in ws_b.iter_rows(min_row=2, max_col=9, values_only=True):
    if not row[0]: continue
    cat = row[7] if len(row) > 7 else None
    if cat != "MANTENIMIENTO HELICOPTERO":
        continue
    fecha = _parse_date(row[0])
    if not fecha or fecha < date(2024, 1, 1):
        continue
    try:
        cargo = float(row[3] or 0)
    except: cargo = 0
    if cargo <= 0: continue
    total_banco += cargo
    count_banco += 1
    ym = (fecha.year, fecha.month)
    por_mes[ym] += cargo

wb.close()

print(f"=== GASTOS HELICÓPTERO 2024+ ===")
print(f"\nFacturas:")
print(f"  Items: {count_facturas}")
print(f"  Total: ${total_facturas:,.0f}")
print(f"\nBanco (Cargos):")
print(f"  Items: {count_banco}")
print(f"  Total: ${total_banco:,.0f}")
print(f"\nDistribución mensual banco:")
for ym in sorted(por_mes.keys()):
    print(f"  {ym[0]}-{ym[1]:02d}: ${por_mes[ym]:>12,.0f}")

# Promedio mensual
total_meses = len(por_mes) or 1
promedio = total_banco / total_meses
print(f"\nPromedio mensual: ${promedio:,.0f}")
