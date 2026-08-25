import sys
sys.stdout.reconfigure(encoding="utf-8")
import os, re, shutil, tempfile, unicodedata
from datetime import date, datetime
from openpyxl import load_workbook
from modules.correlativo import FXP_PATH

# Las 10 que quedaron con diferencia (proveedor Master, nº doc, monto Master)
BUSCAR = [
    ("MISAEL ALEJANDRO HENRIQUEZ CHAVEZ", "94", 11186000),
    ("AGRICOLA STA ELISA", "109", 74203),
    ("FERRETERIA INDUSTRIAL PACHITA SPA", "2493", 16257),
    ("SERVICIO AUTOMOTRIZ B Y J SPA", "210", 95000),
    ("CORA SERVICIOS SPA", "355", 1904000),
    ("CHRISTIAN FRANCISCO RODRIGUEZ TORRES", "119", 964913),
    ("GONZALO YAN PIERR BOBADILLA FARIAS", "271", 493850),
    ("AGRICOLA STA ELISA", "116", 600000),
    ("CALS", "2135611", 722449),
    ("COPEVAL", "5559215", 23205),
]


def sin_tildes(s):
    return "".join(c for c in unicodedata.normalize("NFD", str(s or ""))
                   if unicodedata.category(c) != "Mn").lower()


def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None


def _f(v):
    if isinstance(v, str):
        s = v.upper().replace("USD", "").replace("$", "").strip().replace(".", "").replace(",", ".")
        try: return float(s)
        except ValueError: return 0.0
    try: return float(v or 0)
    except (TypeError, ValueError): return 0.0


tmp = os.path.join(tempfile.gettempdir(), "fxp_busca.xlsx")
shutil.copy2(FXP_PATH, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
filas = []
for i, row in enumerate(wb["FXP"].iter_rows(min_row=2, values_only=True), 2):
    if not row or not row[0]:
        continue
    filas.append({"fila": i, "n": row[3], "prov": str(row[6] or ""),
                  "nro": str(row[7] or ""), "monto": _f(row[8]),
                  "emi": _pd(row[0]), "estado": str(row[11] or ""),
                  "nota": str(row[12] or "")})
wb.close()
print(f"FXP: {len(filas)} filas\n")

for prov_m, nro_m, monto_m in BUSCAR:
    print("=" * 78)
    print(f"MASTER: {prov_m}  F{nro_m}  ${monto_m:,.0f}")
    tokens = {t for t in re.split(r"[^a-z]+", sin_tildes(prov_m)) if len(t) >= 4}
    # candidatos por nombre
    por_nombre = [f for f in filas
                  if tokens & {t for t in re.split(r"[^a-z]+", sin_tildes(f["prov"])) if len(t) >= 4}]
    # candidatos por monto exacto (±1%)
    por_monto = [f for f in filas if f["monto"] > 0
                 and abs(f["monto"] - monto_m) <= max(500, monto_m * 0.01)]
    # por monto = master × 1.19 (IVA)
    iva = monto_m * 1.19
    por_iva = [f for f in filas if f["monto"] > 0
               and abs(f["monto"] - iva) <= max(500, iva * 0.01)]

    if por_nombre:
        print(f"  → por PROVEEDOR ({len(por_nombre)}):")
        for f in por_nombre[:6]:
            print(f"      N°{str(f['n']):>4} fila {f['fila']:>4} | {f['emi']} | "
                  f"{f['prov'][:30]:30} F{f['nro']:<10} ${f['monto']:>12,.0f} [{f['estado'][:9]}]")
    if por_monto:
        print(f"  → por MONTO igual ({len(por_monto)}):")
        for f in por_monto[:4]:
            print(f"      N°{str(f['n']):>4} fila {f['fila']:>4} | {f['emi']} | "
                  f"{f['prov'][:30]:30} F{f['nro']:<10} ${f['monto']:>12,.0f}")
    if por_iva and not por_monto:
        print(f"  → por MONTO ×1.19 ({len(por_iva)}):")
        for f in por_iva[:4]:
            print(f"      N°{str(f['n']):>4} fila {f['fila']:>4} | {f['emi']} | "
                  f"{f['prov'][:30]:30} F{f['nro']:<10} ${f['monto']:>12,.0f}")
    if not (por_nombre or por_monto or por_iva):
        print("  ✗ sin candidatos")
