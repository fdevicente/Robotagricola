import sys, shutil, tempfile, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

DROPBOX = r"C:\Users\Windows\Dropbox\CAMARICO 2023"
candidatos = [
    "BODEGA  ENTRADAS-SALIDAS fda .xlsb.xlsx",
    "BODEGA  ENTRADAS-SALIDAS fda.xlsx",
    "BODEGA  ENTRADAS-SALIDAS fda 2025 PRIVADO.xlsx",
]
for nombre in candidatos:
    src = os.path.join(DROPBOX, nombre)
    if not os.path.exists(src):
        print(f"{nombre}: NO EXISTE\n")
        continue
    tmp = os.path.join(tempfile.gettempdir(), "bod_check.xlsx")
    try:
        shutil.copy2(src, tmp)
        wb = load_workbook(tmp, read_only=True, data_only=True)
        print(f"{nombre}:")
        print(f"   Pestañas: {wb.sheetnames}")
        print(f"   Tiene TRABAJO: {'TRABAJO' in wb.sheetnames}")
        wb.close()
    except Exception as e:
        print(f"{nombre}: ERROR {e}")
    print()
