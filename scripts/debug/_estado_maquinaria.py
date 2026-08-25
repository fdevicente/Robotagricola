"""Qué sabe hoy el sistema de cada máquina, y qué le falta."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from collections import defaultdict
from datetime import date, datetime

from openpyxl import load_workbook

from config import EXCEL_PATH


def _f(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)

# ── Bitácora: máquinas, odómetros y horas ──
ws = wb["Bitácora"]
enc = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
iMaq = enc.index("Máquina"); iOdo = enc.index("Odómetro"); iHrs = enc.index("Horas Día")

maq = defaultdict(lambda: {"regs": 0, "odos": [], "horas": 0.0,
                            "sin_odo": 0, "primera": None, "ultima": None})
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[0]:
        continue
    m = str(r[iMaq] or "").strip().upper()
    if not m:
        continue
    f = _f(r[0])
    d = maq[m]
    d["regs"] += 1
    if r[iOdo] is not None:
        try:
            d["odos"].append((f, float(r[iOdo])))
        except (TypeError, ValueError):
            d["sin_odo"] += 1
    else:
        d["sin_odo"] += 1
    try:
        d["horas"] += float(r[iHrs] or 0)
    except (TypeError, ValueError):
        pass
    if f:
        d["primera"] = min(d["primera"] or f, f)
        d["ultima"] = max(d["ultima"] or f, f)

# ── Facturas: mantenciones y arriendos ──
prov_mant = defaultdict(lambda: {"n": 0, "monto": 0.0, "ultima": None})
ws_f = wb["Facturas"]
for r in ws_f.iter_rows(min_row=2, values_only=True):
    if not r or not r[0]:
        continue
    cat = str(r[16] or "").upper() if len(r) > 16 else ""
    if "MAQUINARIA" not in cat:
        continue
    prov = str(r[3] or "").strip()[:32]
    f = _f(r[0])
    d = prov_mant[prov]
    d["n"] += 1
    try:
        d["monto"] += float(r[15] or 0)
    except (TypeError, ValueError):
        pass
    if f:
        d["ultima"] = max(d["ultima"] or f, f)
wb.close()

print("=" * 74)
print("MÁQUINAS QUE APARECEN EN LA BITÁCORA")
print("=" * 74)
print(f"{'máquina':16}{'regs':>5}{'c/odóm':>8}{'horas':>8}  {'rango de odómetro':26} período")
print("-" * 74)
for m, d in sorted(maq.items(), key=lambda x: -x[1]["horas"]):
    odos = sorted(d["odos"])
    if odos:
        rango = f"{odos[0][1]:,.1f} → {odos[-1][1]:,.1f}"
        delta = odos[-1][1] - odos[0][1]
        rango += f"  (Δ{delta:,.1f})"
    else:
        rango = "— sin lecturas —"
    print(f"{m[:16]:16}{d['regs']:>5}{len(d['odos']):>8}{d['horas']:>8.1f}  {rango:26} "
          f"{d['primera']} a {d['ultima']}")

print("\n" + "=" * 74)
print("PROVEEDORES DE MANTENCIÓN / MAQUINARIA EN FACTURAS")
print("=" * 74)
for p, d in sorted(prov_mant.items(), key=lambda x: -x[1]["monto"])[:12]:
    print(f"  {p:34} {d['n']:>3} fact  ${d['monto']:>12,.0f}   última {d['ultima']}")

print("\n" + "=" * 74)
print("LO QUE FALTA")
print("=" * 74)
print("  · Ninguna máquina tiene ficha: marca, modelo, año, patente, n° de serie.")
print("  · No hay hoja de mantenciones (qué se hizo, a qué odómetro, quién, cuánto).")
print("  · No hay plan de mantención (cada cuántas horas toca cada servicio).")
sin = [m for m, d in maq.items() if not d["odos"]]
if sin:
    print(f"  · Sin NINGUNA lectura de odómetro: {', '.join(sorted(sin))}")
