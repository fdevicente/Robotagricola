#!/usr/bin/env python3
"""Detalla qué se está contando como egreso en TEMP 26/27."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from collections import defaultdict
from openpyxl import load_workbook
from config import EXCEL_PATH

def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try: return datetime.strptime(v[:10], fmt).date()
            except: pass
    return None

def temp_de(f):
    if not f: return None
    if f.month >= 5: return f"TEMP {f.year % 100:02d}/{(f.year + 1) % 100:02d}"
    return f"TEMP {(f.year - 1) % 100:02d}/{f.year % 100:02d}"

tmp = os.path.join(tempfile.gettempdir(), "debug_egr.xlsx")
shutil.copy2(EXCEL_PATH, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)

# Facturas TEMP 26/27
print("=== FACTURAS TEMP 26/27 ===\n")
ws = wb["Facturas"]
fact_total = 0
fact_count = 0
fact_por_cat = defaultdict(lambda: [0, 0])
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    fecha = _pd(row[0])
    if temp_de(fecha) != "TEMP 26/27": continue
    monto = float(row[14] or 0)
    cat = row[16] or "(SIN)"
    fact_total += monto
    fact_count += 1
    fact_por_cat[cat][0] += 1
    fact_por_cat[cat][1] += monto

print(f"Facturas total: {fact_count} = ${fact_total:,.0f}\n")
for cat, (n, m) in sorted(fact_por_cat.items(), key=lambda x: -x[1][1]):
    print(f"  {cat:30} {n:3} items  ${m:>15,.0f}")

# Cargos banco TEMP 26/27
print("\n\n=== CARGOS BANCO TEMP 26/27 ===\n")
ws_b = wb["Cuenta Banco"]
carg_total = 0
carg_count = 0
carg_por_cat = defaultdict(lambda: [0, 0])
for row in ws_b.iter_rows(min_row=2, max_col=9, values_only=True):
    if not row[0]: continue
    fecha = _pd(row[0])
    if temp_de(fecha) != "TEMP 26/27": continue
    try:
        cargo = float(row[3] or 0)
    except: continue
    if cargo <= 0: continue
    cat = str(row[7] or "(SIN)").upper()
    # Aplicar mismo filtro que dashboard_data.get_resumen_temporada
    if "INGRESO" in cat or "TRANSFERENCIA" in cat or "PRE-2021" in cat:
        continue
    carg_total += cargo
    carg_count += 1
    carg_por_cat[cat][0] += 1
    carg_por_cat[cat][1] += cargo

print(f"Cargos banco (filtrados) total: {carg_count} = ${carg_total:,.0f}\n")
for cat, (n, m) in sorted(carg_por_cat.items(), key=lambda x: -x[1][1]):
    print(f"  {cat:30} {n:3} items  ${m:>15,.0f}")

print(f"\n\nTOTAL EGRESOS TEMP 26/27 = ${fact_total + carg_total:,.0f}")

wb.close()
