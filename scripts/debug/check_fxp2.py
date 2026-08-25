#!/usr/bin/env python3
"""Encuentra columna NOTAS en ScotiaBCO."""
import shutil, tempfile, os
from openpyxl import load_workbook

src = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"
tmp = os.path.join(tempfile.gettempdir(), "fxp_check2.xlsx")
shutil.copy2(src, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["ScotiaBCO"]

# Read all header columns
print("Todas las columnas de fila 5 (headers):")
for col in range(1, 30):
    cell_val = ws.cell(5, col).value
    if cell_val:
        print(f"  Col {col}: {cell_val}")

# Sample row with all columns
print("\nFila 6 completa:")
for col in range(1, 30):
    val = ws.cell(6, col).value
    if val is not None:
        print(f"  Col {col}: {val}")

# Look for "NOTAS" header
print("\nBuscando NOTAS en cualquier fila/columna...")
for r in range(1, 10):
    for c in range(1, 30):
        v = ws.cell(r, c).value
        if v and "NOTA" in str(v).upper():
            print(f"  Encontrado 'NOTA' en fila {r}, col {c}: {v}")

wb.close()
