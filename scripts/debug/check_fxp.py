#!/usr/bin/env python3
"""Inspecciona FXP.xlsx pestaña ScotiaBCO."""
import shutil, tempfile, os
from openpyxl import load_workbook

src = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"
tmp = os.path.join(tempfile.gettempdir(), "fxp_check.xlsx")
shutil.copy2(src, tmp)
print("FXP copiado, abriendo...")

wb = load_workbook(tmp, read_only=True, data_only=True)
print(f"Sheets: {wb.sheetnames}\n")

if "ScotiaBCO" in wb.sheetnames:
    ws = wb["ScotiaBCO"]
    print(f"ScotiaBCO: {ws.max_row} filas\n")

    # Headers
    print("Headers:")
    for col, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), 1):
        print(f"  Col {col}: {cell}")

    # Sample rows
    print("\nPrimeras 5 filas:")
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), 2):
        print(f"  Fila {idx}: {row[:10]}")

    # Count rows with NOTAS
    print("\nBuscando columna NOTAS...")
wb.close()
