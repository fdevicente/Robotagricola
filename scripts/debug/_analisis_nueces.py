"""Desarma la venta de nueces: qué está recibido, qué esperado, y a cuánto el kg."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from collections import defaultdict
from openpyxl import load_workbook
from config import EXCEL_PATH

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)

usd_clp = 1000
for r in wb["Config"].iter_rows(min_row=2, values_only=True):
    if r[0] == "usd_clp_estimado":
        usd_clp = float(r[1])
print(f"Tipo de cambio del modelo: {usd_clp:,.0f} CLP/USD\n")

ws = wb["Cosechas"]
exp = defaultdict(lambda: {"kg": 0, "rec_usd": 0.0, "rec_clp": 0.0,
                            "esp_usd": 0.0, "filas": []})
for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if not r[0] or str(r[1]).upper() != "NOGALES":
        continue
    nombre = str(r[3])
    kg_asig = float(r[4] or 0)
    monto_usd = float(r[9] or 0)
    estado = str(r[11] or "").lower()
    e = exp[nombre]
    e["kg"] = max(e["kg"], kg_asig)
    if estado == "recibido":
        e["rec_usd"] += monto_usd
        e["rec_clp"] += float(r[13] or 0)
    else:
        e["esp_usd"] += monto_usd
    e["filas"].append((i, r[7], r[8], monto_usd, estado, r[10]))
wb.close()

tot_kg = tot_rec_clp = tot_esp_usd = tot_rec_usd = 0
for nombre, e in exp.items():
    if e["kg"] < 100:          # ventas locales sueltas
        continue
    print(f"── {nombre} — {e['kg']:,.0f} kg ──")
    for i, cuota, fecha, usd, estado, tipo in e["filas"]:
        marca = "✅" if estado == "recibido" else "⏳"
        pkg = usd / e["kg"] if e["kg"] else 0
        print(f"   {marca} fila {i:>2} cuota {cuota} {str(fecha)[:10]:10} "
              f"{usd:>9,.0f} USD  ({pkg:.3f} USD/kg)  {tipo}")
    total_usd = e["rec_usd"] + e["esp_usd"]
    print(f"   {'':6}{'TOTAL':>24} {total_usd:>9,.0f} USD  "
          f"({total_usd / e['kg']:.3f} USD/kg)")
    print(f"   {'':6}{'recibido':>24} {e['rec_usd']:>9,.0f} USD = ${e['rec_clp']:>13,.0f} CLP"
          f"   ({e['rec_clp'] / e['rec_usd']:.0f} CLP/USD real)" if e["rec_usd"] else "")
    print(f"   {'':6}{'por recibir':>24} {e['esp_usd']:>9,.0f} USD = ${e['esp_usd'] * usd_clp:>13,.0f} CLP\n")
    tot_kg += e["kg"]; tot_rec_clp += e["rec_clp"]
    tot_esp_usd += e["esp_usd"]; tot_rec_usd += e["rec_usd"]

tot_usd = tot_rec_usd + tot_esp_usd
print("=" * 66)
print(f"NUECES 2026 — {tot_kg:,.0f} kg")
print(f"  Total contratado      {tot_usd:>10,.0f} USD  → {tot_usd / tot_kg:.3f} USD/kg")
print(f"  Ya recibido           {tot_rec_usd:>10,.0f} USD  = ${tot_rec_clp:,.0f} CLP  (no se puede cambiar)")
print(f"  Por recibir           {tot_esp_usd:>10,.0f} USD  = ${tot_esp_usd * usd_clp:,.0f} CLP")
print()
for objetivo in (2.1, 2.3):
    meta_usd = tot_kg * objetivo
    ajuste = meta_usd - tot_usd
    print(f"  Si el TOTAL fuera {objetivo} USD/kg → {meta_usd:,.0f} USD "
          f"({ajuste:+,.0f} USD = {ajuste * usd_clp:+,.0f} CLP)")
