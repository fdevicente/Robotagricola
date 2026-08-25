"""Lee la pestaña VACAS del Excel de asistencia de Juan."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

RUTA = (r"C:\Users\Windows\Dropbox\CAMARICO 2023"
        r"\ASISTENCIA TEMP 2023-2024 fda JUAN PARADA "
        r"(Copia en conflicto de juan parada 2025-03-06).xlsx")

wb = load_workbook(RUTA, read_only=True, data_only=True)
print("Hojas:", ", ".join(wb.sheetnames), "\n")

nombre = next((h for h in wb.sheetnames if "VACA" in h.upper()), None)
if not nombre:
    print("❌ No encontré una pestaña VACAS")
    sys.exit(1)

ws = wb[nombre]
print(f"=== {nombre} · {ws.max_row} filas × {ws.max_column} columnas ===\n")
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if not row or not any(v is not None and str(v).strip() != "" for v in row):
        continue
    celdas = []
    for v in row:
        if v is None:
            celdas.append("")
        elif hasattr(v, "strftime"):
            celdas.append(v.strftime("%d-%m-%Y"))
        else:
            celdas.append(str(v).strip()[:26])
    while celdas and celdas[-1] == "":
        celdas.pop()
    print(f"{i:>3} | " + " | ".join(celdas))
    if i > 120:
        print("   … (cortado)")
        break
wb.close()
