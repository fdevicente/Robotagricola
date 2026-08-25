#!/usr/bin/env python3
"""Resumen del estado actual de Cuenta Banco."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from openpyxl import load_workbook
from config import EXCEL_PATH

tmp = os.path.join(tempfile.gettempdir(), "master_summary.xlsx")
shutil.copy2(EXCEL_PATH, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["Cuenta Banco"]

from collections import Counter
categorias = Counter()
montos = {}
revisar_post2021 = 0

from datetime import date, datetime

def _parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None

for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
    if not row[0]:
        continue
    cat = row[7] or "(SIN CATEGORIA)"
    cargo = float(row[3] or 0)
    categorias[cat] += 1
    montos[cat] = montos.get(cat, 0) + cargo

    if cat == "REVISAR":
        fecha = _parse_date(row[0])
        if fecha and fecha >= date(2021, 1, 1):
            revisar_post2021 += 1

wb.close()

print("=" * 70)
print("ESTADO ACTUAL: Cuenta Banco")
print("=" * 70)
print(f"\nREVISAR post-2021 (relevantes): {revisar_post2021}\n")

for cat, count in sorted(categorias.items(), key=lambda x: -montos.get(x[0], 0)):
    monto = montos.get(cat, 0)
    print(f"  {cat:50s} {count:5d} items  ${monto:>18,.0f}")
