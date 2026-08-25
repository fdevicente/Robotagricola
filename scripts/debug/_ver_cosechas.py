import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook
from config import EXCEL_PATH

wb = load_workbook(EXCEL_PATH, data_only=False)
print("Hojas:", ", ".join(wb.sheetnames), "\n")

for hoja in ("Cosechas", "Ingresos Proyectados", "Config"):
    if hoja not in wb.sheetnames:
        continue
    ws = wb[hoja]
    print(f"=== {hoja} ({ws.max_row} filas × {ws.max_column} cols) ===")
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if not any(v is not None for v in r):
            continue
        print(f"  {i:>3}", [str(v)[:22] if v is not None else "" for v in r])
        if i > 30:
            print("   ...")
            break
    print()

# ¿Dónde aparece 2.3 / 2,3 como precio?
print("=== Buscando el precio de la nuez (2.3) en todo el libro ===")
for hoja in wb.sheetnames:
    ws = wb[hoja]
    if ws.max_row > 3000:
        continue
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, (int, float)) and 2.0 <= float(v) <= 2.6:
                print(f"  {hoja}!{c.coordinate} = {v}")
wb.close()
