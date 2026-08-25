import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook
from config import EXCEL_PATH

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
print("Hojas:", [s for s in wb.sheetnames])
for nombre in wb.sheetnames:
    if "inventario" in nombre.lower() or "insumo" in nombre.lower() or "stock" in nombre.lower():
        ws = wb[nombre]
        print(f"\n=== {nombre} ===")
        hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        for i, h in enumerate(hdr, 1):
            if h:
                print(f"  col {i}: {h}")
        filas = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
        print(f"  filas con datos: {len(filas)}")
        for r in filas[:15]:
            print("   ", [str(x)[:22] if x is not None else "" for x in r[:8]])
wb.close()
