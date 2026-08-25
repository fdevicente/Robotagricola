"""Últimos movimientos del banco y su categoría."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from datetime import date, datetime

from openpyxl import load_workbook

from config import EXCEL_PATH

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
filas = []
for i, r in enumerate(wb["Cuenta Banco"].iter_rows(min_row=2, values_only=True), 2):
    if not r or not r[0]:
        continue
    f = r[0].date() if isinstance(r[0], datetime) else r[0]
    filas.append((i, f, str(r[1] or ""), float(r[3] or 0), float(r[4] or 0),
                  r[5], str(r[7] or "")))
wb.close()

print(f"{'fila':>5} {'fecha':11} {'monto':>15} {'saldo':>15}  categoría")
print("-" * 92)
for i, f, desc, cargo, abono, saldo, cat in filas[-10:]:
    monto = -cargo if cargo else abono
    marca = "⚠️ SIN CATEGORÍA" if not cat else cat
    print(f"{i:>5} {str(f):11} {monto:>15,.0f} "
          f"{(saldo if saldo is not None else 0):>15,.0f}  {marca}")
    print(f"{'':5} {'':11} {desc[:60]}")

sin_cat = [x for x in filas if not x[6]]
print(f"\nSin categorizar: {len(sin_cat)}")
for i, f, desc, cargo, abono, saldo, cat in sin_cat:
    print(f"   fila {i} | {f} | {(-cargo if cargo else abono):>14,.0f} | {desc[:50]}")
