import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook
from config import EXCEL_PATH
wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
print("Col J (resumen legible) fila 4873:")
print("  ", wb["Cuenta Banco"].cell(4873, 10).value)
wb.close()
