"""Gasto real por mes en el banco vs lo que proyecta el modelo."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from collections import defaultdict
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

# Categorías que NO son operación (préstamos, divisas, reintegros)
FUERA = {"PRESTAMOS A OTRAS SOCIEDADES", "CAMBIO DIVISA",
         "REINTEGROS Y DEVOLUCIONES"}
MES = ["", "ene", "feb", "mar", "abr", "may", "jun",
       "jul", "ago", "sep", "oct", "nov", "dic"]

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Cuenta Banco"]


def _f(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    try: return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError): return None


por_mes = defaultdict(float)
por_mes_todo = defaultdict(float)
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[0]:
        continue
    f = _f(r[0])
    if not f or f < date(2025, 6, 1):
        continue
    cargo = float(r[3] or 0)
    if cargo <= 0:
        continue
    cat = str(r[7] or "").upper()
    por_mes_todo[(f.year, f.month)] += cargo
    if cat not in FUERA:
        por_mes[(f.year, f.month)] += cargo
wb.close()

print(f"{'mes':10}{'egreso operacional':>22}{'todo (incl. préstamos)':>26}")
print("-" * 58)
meses = sorted(por_mes_todo)
for ym in meses:
    print(f"{MES[ym[1]]}-{str(ym[0])[-2:]:5}{por_mes[ym]:>22,.0f}{por_mes_todo[ym]:>26,.0f}")

# Promedio de los últimos 12 meses completos (excluye el mes en curso)
completos = [ym for ym in meses if ym < (date.today().year, date.today().month)]
ult12 = completos[-12:]
prom = sum(por_mes[ym] for ym in ult12) / len(ult12)
print("-" * 58)
print(f"\nPromedio operacional últimos {len(ult12)} meses: ${prom:,.0f}/mes")
print(f"  → 10 meses a ese ritmo:            ${prom * 10:,.0f}")
print(f"  El modelo proyecta para ago-may:    $648,132,564  "
      f"(${648_132_564 / 10:,.0f}/mes)")
print(f"  Diferencia:                        ${648_132_564 - prom * 10:+,.0f}")
