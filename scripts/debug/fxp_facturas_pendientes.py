#!/usr/bin/env python3
"""Compara facturas pendientes en FXP vs estado en Master."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from collections import defaultdict
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP_PATH = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"


def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try: return datetime.strptime(v[:10], fmt).date()
            except: pass
    return None


def norm(s):
    return (s or "").strip().upper().replace(".", "").replace("  ", " ")


# ─── Cargar FXP pestaña FXP ──────────────────────────────
print("[1/3] Cargando FXP.FXP (facturas)...")
tmp = os.path.join(tempfile.gettempdir(), "fxp_pend.xlsx")
shutil.copy2(FXP_PATH, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["FXP"]

# Cols: 1=emi, 2=venc, 3=pago, 7=proveedor, 8=nro, 9=monto, 12=estado (Saldo), 13=notas
pendientes = []  # Estado != Pagada
pagadas = 0
hoy = date.today()

for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    estado = str(row[11] or "").strip().upper()
    if estado == "PAGADA":
        pagadas += 1
        continue

    # Pendiente
    f_emi = _pd(row[0])
    f_venc = _pd(row[1])
    f_pago = _pd(row[2])
    prov = str(row[6] or "")
    nro = str(row[7] or "").strip()
    try:
        monto = float(row[8] or 0)
    except: monto = 0
    nota = str(row[12] or "")

    if not nro or monto <= 0: continue

    dias_venc = (hoy - f_venc).days if f_venc else None
    estado_calc = "VENCIDA" if (f_venc and f_venc < hoy) else "POR VENCER"

    pendientes.append({
        "fecha_emi": f_emi,
        "fecha_venc": f_venc,
        "fecha_pago": f_pago,
        "proveedor": prov,
        "nro": nro,
        "monto": monto,
        "estado_fxp": estado,
        "estado_calc": estado_calc,
        "dias_venc": dias_venc,
        "nota": nota,
    })

wb.close()
print(f"   FXP Pagadas: {pagadas}")
print(f"   FXP Pendientes (no Pagada): {len(pendientes)}")
vencidas_fxp = [p for p in pendientes if p["estado_calc"] == "VENCIDA"]
por_vencer = [p for p in pendientes if p["estado_calc"] == "POR VENCER"]
print(f"   - Vencidas (venc < hoy):  {len(vencidas_fxp)}")
print(f"   - Por vencer (venc ≥ hoy): {len(por_vencer)}\n")

# ─── Cargar Master.Facturas ──────────────────────────────
print("[2/3] Cargando Master.Facturas...")
wb_m = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws_m = wb_m["Facturas"]
master_idx = {}
for r_idx, row in enumerate(ws_m.iter_rows(min_row=2, values_only=True), start=2):
    if not row[0]: continue
    prov = norm(row[3])
    nro = str(row[6] or "").strip()
    if not prov or not nro: continue
    fpago = row[2]
    master_idx[(prov, nro)] = {
        "fila": r_idx,
        "fecha_pago": fpago,
        "monto": row[14],
        "prov_orig": row[3],
    }
wb_m.close()
print(f"   Master: {len(master_idx)} facturas indexadas\n")

# ─── Comparar ────────────────────────────────────────────
print("[3/3] Comparando pendientes FXP vs Master...\n")

en_master_no_pagada_master = 0  # FXP dice pendiente, Master también
en_master_pagada_master = 0  # FXP dice pendiente, Master dice pagada
no_en_master = 0
no_en_master_list = []

for p in pendientes:
    key = (norm(p["proveedor"]), p["nro"])
    # Buscar también por nro solo
    if key in master_idx:
        m = master_idx[key]
    else:
        match_nro = [k for k in master_idx if k[1] == p["nro"]]
        m = master_idx[match_nro[0]] if match_nro else None

    if not m:
        no_en_master += 1
        no_en_master_list.append(p)
        continue

    if m["fecha_pago"] and str(m["fecha_pago"]).strip():
        en_master_pagada_master += 1
    else:
        en_master_no_pagada_master += 1

print(f"  FXP pendiente + Master pendiente:  {en_master_no_pagada_master}")
print(f"  FXP pendiente + Master ya pagada:  {en_master_pagada_master}")
print(f"  FXP pendiente y NO en Master:      {no_en_master}\n")

# Mostrar las vencidas FXP
print("=== VENCIDAS según FXP (top 30) ===")
vencidas_fxp.sort(key=lambda x: -(x["dias_venc"] or 0))
total_venc = sum(p["monto"] for p in vencidas_fxp)
print(f"Total {len(vencidas_fxp)} facturas vencidas = ${total_venc:,.0f}\n")
for p in vencidas_fxp[:30]:
    print(f"  Venc {p['fecha_venc']} (-{p['dias_venc']:>3}d) | {p['proveedor'][:30]:30} F{p['nro']:<10} | ${p['monto']:>11,.0f}")
    if p["nota"]: print(f"    Nota: {p['nota'][:80]}")

print(f"\n=== POR VENCER según FXP (próximas 15) ===")
por_vencer.sort(key=lambda x: (x["fecha_venc"] or date(2099,1,1)))
total_pv = sum(p["monto"] for p in por_vencer)
print(f"Total {len(por_vencer)} = ${total_pv:,.0f}\n")
for p in por_vencer[:15]:
    dias = -p["dias_venc"] if p["dias_venc"] else 0
    print(f"  Venc {p['fecha_venc']} (+{dias:>3}d) | {p['proveedor'][:30]:30} F{p['nro']:<10} | ${p['monto']:>11,.0f}")
