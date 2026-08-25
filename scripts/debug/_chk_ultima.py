import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook
from config import EXCEL_PATH
wb = load_workbook(EXCEL_PATH, read_only=True)
ws = wb["Facturas"]
print("max_row de Facturas:", ws.max_row)
wb.close()
# ¿Alguna conciliación apunta a filas > 2164?
wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
if "Conciliaciones" in wb.sheetnames:
    altas = [r[0] for r in wb["Conciliaciones"].iter_rows(min_row=2, values_only=True)
             if r and r[7] and isinstance(r[7], (int, float)) and r[7] >= 2165]
    print("Conciliaciones que apuntan a filas >= 2165:", len(altas))
wb.close()
