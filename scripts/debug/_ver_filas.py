"""Muestra filas concretas del Master completas."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from datetime import datetime

from openpyxl import load_workbook

from config import EXCEL_PATH

FILAS = [int(a) for a in sys.argv[1:]] or [2085, 2122, 2165]

wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb["Facturas"]
enc = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
datos = {}
for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    if i in FILAS:
        datos[i] = r
wb.close()

for f in FILAS:
    r = datos.get(f)
    if not r:
        print(f"fila {f}: no existe\n")
        continue
    print(f"=== fila {f} ===")
    for i, (h, v) in enumerate(zip(enc, r)):
        if v is None or str(v).strip() == "":
            continue
        if isinstance(v, datetime):
            v = v.date()
        print(f"   {str(h)[:34]:34} : {str(v)[:56]}")
    print()
