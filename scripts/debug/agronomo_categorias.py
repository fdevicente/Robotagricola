#!/usr/bin/env python3
"""Extrae las categorías del agronomo (AREA, cargo, cargo ll)."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from collections import Counter
from openpyxl import load_workbook

src = r"C:\Users\Windows\Dropbox\CAMARICO 2023\PLANILLA GASTOS CAMARICO 2023-2026.xlsx"
tmp = os.path.join(tempfile.gettempdir(), "agronomo_cat.xlsx")
shutil.copy2(src, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["DATOS"]

# Headers fila 9, datos desde fila 10
areas = Counter()
sub_areas = Counter()
cargos = Counter()
cargos_ll = Counter()
combos = Counter()  # (cargo, cargo_ll)
proveedores = Counter()

total_facturas = 0
monto_total = 0

for row in ws.iter_rows(min_row=10, max_col=27, values_only=True):
    if not row[0]: continue  # TEMPORADA vacía
    total_facturas += 1
    area = str(row[2] or "").strip()
    sub = str(row[3] or "").strip()
    cargo = str(row[4] or "").strip()
    cargo_ll = str(row[5] or "").strip()
    prov = str(row[7] or "").strip()
    try:
        total = float(row[25] or 0)
    except: total = 0
    monto_total += abs(total)

    if area: areas[area] += 1
    if sub: sub_areas[sub] += 1
    if cargo: cargos[cargo] += 1
    if cargo_ll: cargos_ll[cargo_ll] += 1
    if cargo and cargo_ll: combos[(cargo, cargo_ll)] += 1
    if prov: proveedores[prov] += 1

print(f"Total filas: {total_facturas}")
print(f"Monto total: ${monto_total:,.0f}\n")

print(f"=== AREAS ({len(areas)}) ===")
for k, n in areas.most_common():
    print(f"  {k}: {n}")

print(f"\n=== SUB AREAS ({len(sub_areas)}) ===")
for k, n in sub_areas.most_common():
    print(f"  {k}: {n}")

print(f"\n=== CARGOS ({len(cargos)}) ===")
for k, n in cargos.most_common():
    print(f"  {k}: {n}")

print(f"\n=== CARGO II - subcategorías ({len(cargos_ll)}) ===")
for k, n in cargos_ll.most_common():
    print(f"  {k}: {n}")

print(f"\n=== TOP COMBOS cargo + cargo ll ===")
for (c, c2), n in combos.most_common(25):
    print(f"  {c} / {c2}: {n}")

print(f"\n=== TOP 20 PROVEEDORES ===")
for k, n in proveedores.most_common(20):
    print(f"  {k}: {n}")

wb.close()
