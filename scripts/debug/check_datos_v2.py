#!/usr/bin/env python3
"""Busca headers en hoja DATOS - escanea primeras 20 filas."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from openpyxl import load_workbook

src = r"C:\Users\Windows\Dropbox\CAMARICO 2023\PLANILLA GASTOS CAMARICO 2023-2026.xlsx"
tmp = os.path.join(tempfile.gettempdir(), "datos_v2.xlsx")
shutil.copy2(src, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["DATOS"]

print(f"Total filas: {ws.max_row}, cols: {ws.max_column}\n")
print("Primeras 30 filas (cols 1-12):")
for r in range(1, 31):
    vals = [ws.cell(r, c).value for c in range(1, 13)]
    if any(vals):
        # Truncar strings largos
        vals_str = [str(v)[:20] if v else '' for v in vals]
        print(f"  R{r}: {vals_str}")

print("\nUna fila de datos completa (fila 50):")
for c in range(1, 37):
    v = ws.cell(50, c).value
    if v is not None:
        print(f"  Col {c}: {v}")

wb.close()
