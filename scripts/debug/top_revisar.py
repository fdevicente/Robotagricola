#!/usr/bin/env python3
"""Top 30 REVISAR post-2021 por monto, con notas FXP."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP_PATH = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"


def _parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    return None


# Cargar FXP
tmp_f = os.path.join(tempfile.gettempdir(), "fxp_top.xlsx")
shutil.copy2(FXP_PATH, tmp_f)
wb_f = load_workbook(tmp_f, read_only=True, data_only=True)
ws_f = wb_f["ScotiaBCO"]
fxp_idx = {}
for row in ws_f.iter_rows(min_row=6, values_only=True):
    if not row or len(row) < 11: continue
    fecha = _parse_date(row[2])
    if not fecha: continue
    try:
        monto = int(round(float(row[6] or 0)))
    except (TypeError, ValueError):
        continue
    fxp_idx[(fecha.isoformat(), monto)] = (
        str(row[5] or ""), str(row[9] or ""), str(row[10] or "")
    )
wb_f.close()

# Master
tmp_m = os.path.join(tempfile.gettempdir(), "master_top.xlsx")
shutil.copy2(EXCEL_PATH, tmp_m)
wb = load_workbook(tmp_m, read_only=True, data_only=True)
ws = wb["Cuenta Banco"]

revisar = []
for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=9, values_only=True), start=2):
    if not row[0]: continue
    cat = row[7]
    if cat != "REVISAR": continue
    fecha = _parse_date(row[0])
    if not fecha or fecha < date(2021, 1, 1): continue
    try:
        cargo = float(row[3] or 0)
    except (TypeError, ValueError):
        continue
    if cargo <= 0: continue

    cargo_int = int(round(cargo))
    fxp = fxp_idx.get((fecha.isoformat(), cargo_int), ("", "", ""))
    revisar.append({
        "fila": idx,
        "fecha": fecha.isoformat(),
        "desc": str(row[1] or "")[:50],
        "ref": str(row[2] or "")[:30],
        "cargo": cargo,
        "fxp_desc": fxp[0][:40],
        "asig": fxp[1][:30],
        "notas": fxp[2][:60],
    })

wb.close()
revisar.sort(key=lambda x: -x["cargo"])

print(f"Total REVISAR post-2021: {len(revisar)}\n")
print("TOP 30 por monto:\n")
for i, r in enumerate(revisar[:30], 1):
    print(f"{i:2d}. Fila {r['fila']} | {r['fecha']} | ${r['cargo']:>15,.0f}")
    print(f"    Desc: {r['desc']}")
    if r['ref']:    print(f"    Ref:  {r['ref']}")
    if r['fxp_desc']: print(f"    FXP:  {r['fxp_desc']}")
    if r['asig']:   print(f"    Asig: {r['asig']}")
    if r['notas'] and r['notas'].strip() != '-': print(f"    Nota: {r['notas']}")
    print()
