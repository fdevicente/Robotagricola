"""
fix_unitario.py
Para cada fila existente en la hoja Facturas:
  1. Lee el TOTAL NETO actual (valor correcto) y la Cantidad
  2. Calcula Valor unitario = TOTAL NETO / Cantidad y lo escribe como valor
  3. Reemplaza TOTAL NETO con la fórmula =J*K (dinámica, sin IVA)
Ejecutar una sola vez para corregir filas existentes.
"""
from openpyxl import load_workbook
import os

BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE, "..", "MASTER Agricola Santa Elisa.xlsx")
SHEET_NAME = "Facturas"
COL_UNITARIO = 10   # J
COL_CANTIDAD = 11   # K
COL_NETO     = 12   # L

print(f"Abriendo: {EXCEL_PATH}")
wb = load_workbook(EXCEL_PATH)
ws = wb[SHEET_NAME]

fixed = 0
for row in range(2, ws.max_row + 1):
    cantidad = ws.cell(row=row, column=COL_CANTIDAD).value
    neto     = ws.cell(row=row, column=COL_NETO).value

    if cantidad is None and neto is None:
        continue

    try:
        qty  = float(cantidad or 0)
        neto_val = float(neto or 0)
    except (TypeError, ValueError):
        continue

    # 1. Escribir Valor unitario como valor numérico (neto / cantidad)
    cell_j = ws.cell(row=row, column=COL_UNITARIO)
    cell_j.value = round(neto_val / qty, 6) if qty else 0
    cell_j.number_format = '#,##0.000'

    # 2. Reemplazar TOTAL NETO con fórmula dinámica =J*K
    cell_l = ws.cell(row=row, column=COL_NETO)
    cell_l.value = f"=J{row}*K{row}"
    cell_l.number_format = '#,##0'

    fixed += 1

wb.save(EXCEL_PATH)
print(f"Listo — {fixed} filas corregidas: J=valor unitario, L=J*K (fórmula)")
