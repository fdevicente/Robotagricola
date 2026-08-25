#!/usr/bin/env python3
"""Lee datos, recategoriza, y escribe nuevo archivo Excel limpio."""
import shutil
from openpyxl import load_workbook
from config import EXCEL_PATH

backup_path = EXCEL_PATH + '.clean_backup'
shutil.copy2(EXCEL_PATH, backup_path)

print('Reading source file...')
src_wb = load_workbook(backup_path, data_only=False)
src_ws = src_wb['Facturas']

# Almacenar todos los datos en memoria
all_rows = []
for row_idx in range(1, src_ws.max_row + 1):
    row_data = []
    for col_idx in range(1, src_ws.max_column + 1):
        cell = src_ws.cell(row_idx, col_idx)
        row_data.append({
            'value': cell.value,
            'formula': cell.data_type == 'f'
        })
    all_rows.append(row_data)

src_wb.close()

# Aplicar categorías
print('Categorizing...')
def get_cat(p, d):
    p = (p or '').upper()
    d = (d or '').upper()
    if 'ECOSMART' in p: return 'SERVICIOS DE EXPORTACION'
    if 'S-INVEST' in p or 'S.INVEST' in p: return 'COSTO ENERGETICO'
    if 'SCOTIABANK' in p: return 'CAMBIO DIVISA'
    if 'LUIS QUIROZ' in p or 'LUIS ALBERTO' in p: return 'TRANSPORTE'
    if 'CGE' in p or 'ELECTRICIDAD' in p: return 'ENERGIA'
    if 'AGUA' in p or 'RIEGO' in p: return 'AGUA'
    if 'TRANSPORTE' in p or 'FLETE' in p or 'DHL' in p or 'LATAM' in p: return 'TRANSPORTE'
    if 'AGROKIMUN' in p or 'FERTILI' in p or 'ABONO' in p: return 'INSUMOS AGRICOLAS'
    if 'FERRETERIA' in p or 'AGROEQUIPOS' in p: return 'HERRAMIENTAS'
    if 'ARIDOS' in p or 'GRAVA' in p or 'ARENA' in p: return 'MATERIALES'
    if 'LIBRERIA' in p or 'TUCAN' in p or 'MERCADOLIBRE' in p: return 'SUMINISTROS'
    return 'SERVICIOS'

count = 0
for row_idx in range(2, len(all_rows) + 1):
    if all_rows[row_idx - 1][16]['value'] == 'REVISAR':
        prov = all_rows[row_idx - 1][3]['value']
        detail = all_rows[row_idx - 1][7]['value']
        all_rows[row_idx - 1][16]['value'] = get_cat(prov, detail)
        count += 1

print(f'Categorized {count} invoices. Writing to file...')

# Crear nuevo workbook
from openpyxl import Workbook
new_wb = Workbook()
new_ws = new_wb.active
new_ws.title = 'Facturas'

# Escribir todos los datos
for row_idx, row_data in enumerate(all_rows, 1):
    for col_idx, cell_data in enumerate(row_data, 1):
        new_ws.cell(row_idx, col_idx).value = cell_data['value']

# Copiar otras hojas
src_wb = load_workbook(backup_path)
for sheet_name in src_wb.sheetnames:
    if sheet_name != 'Facturas':
        new_ws2 = new_wb.create_sheet(sheet_name)
        src_sheet = src_wb[sheet_name]
        for row in src_sheet.iter_rows():
            for cell in row:
                new_cell = new_ws2.cell(cell.row, cell.column)
                new_cell.value = cell.value

src_wb.close()

print('Saving new file...')
new_wb.save(EXCEL_PATH)
new_wb.close()
print('Done!')
