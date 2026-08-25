#!/usr/bin/env python3
"""Aplicación segura de categorías con verificación de integridad."""
import os
import shutil
from openpyxl import load_workbook
from config import EXCEL_PATH

def categorize():
    print(f'Backing up {EXCEL_PATH}...')
    backup_path = EXCEL_PATH + '.backup2'
    shutil.copy2(EXCEL_PATH, backup_path)

    print('Opening workbook (write mode)...')
    wb = load_workbook(EXCEL_PATH)
    ws = wb['Facturas']

    def get_cat(p, d):
        p = (p or '').upper()
        d = (d or '').upper()
        if 'ECOSMART' in p: return 'SERVICIOS DE EXPORTACION'
        if 'S-INVEST' in p or 'S.INVEST' in p: return 'COSTO ENERGETICO'
        if 'SCOTIABANK' in p: return 'CAMBIO DIVISA'
        if 'LUIS QUIROZ' in p or 'LUIS ALBERTO' in p: return 'TRANSPORTE'
        if 'CGE' in p or 'ELECTRICIDAD' in p: return 'ENERGIA'
        if 'AGUA' in p or 'RIEGO' in p: return 'AGUA'
        if 'TRANSPORTE' in p or 'FLETE' in p or 'DHL' in p or 'LATAM' in p: return 'TRANSPORTE'
        if 'AGROKIMUN' in p or 'FERTILI' in p or 'ABONO' in p: return 'INSUMOS AGRICOLAS'
        if 'FERRETERIA' in p or 'AGROEQUIPOS' in p: return 'HERRAMIENTAS'
        if 'ARIDOS' in p or 'GRAVA' in p or 'ARENA' in p: return 'MATERIALES'
        if 'LIBRERIA' in p or 'TUCAN' in p or 'MERCADOLIBRE' in p: return 'SUMINISTROS'
        return 'SERVICIOS'

    count = 0
    for row in range(2, 2000):
        cat = ws.cell(row, 17).value
        if cat == 'REVISAR':
            prov = ws.cell(row, 4).value
            detail = ws.cell(row, 8).value
            new_cat = get_cat(prov, detail)
            ws.cell(row, 17).value = new_cat
            count += 1

    print(f'Updated {count} invoices. Saving carefully...')

    # Save to temp first
    temp_path = EXCEL_PATH + '.tmp'
    wb.save(temp_path)
    wb.close()

    print('Verifying saved file...')
    try:
        test_wb = load_workbook(temp_path, read_only=True, data_only=True)
        test_wb.close()
        print('✅ File integrity verified')

        # Move temp to real location
        os.replace(temp_path, EXCEL_PATH)
        print('✅ File replaced successfully')
        return True
    except Exception as e:
        print(f'❌ Verification failed: {e}')
        # Restore from backup
        shutil.copy2(backup_path, EXCEL_PATH)
        print('File restored from backup')
        return False

if __name__ == '__main__':
    success = categorize()
    exit(0 if success else 1)
