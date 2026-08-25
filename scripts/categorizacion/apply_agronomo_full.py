#!/usr/bin/env python3
"""Aplica las 3 acciones desde la planilla del agronomo:
1. Normalizar categorías a MAYÚSCULAS
2. Asignar Cultivo del agronomo (NOGALES/CEREZOS/etc)
3. Reclasificar discrepancias reales
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from openpyxl import load_workbook
from config import EXCEL_PATH

AGRO_PATH = r"C:\Users\Windows\Dropbox\CAMARICO 2023\PLANILLA GASTOS CAMARICO 2023-2026.xlsx"

# Normalización MAYÚSCULAS para categorías existentes
NORMALIZE_MAP = {
    "Fertilizantes": "FERTILIZANTES",
    "Fitosanitarios": "FITOSANITARIOS",
    "Maquinaria - mantencion": "MAQUINARIA - MANTENCION",
    "Mano de obra temporal": "MANO DE OBRA TEMPORAL",
    "Mano de obra planta": "MANO DE OBRA PLANTA",
    "Combustible": "COMBUSTIBLE",
    "Riego": "RIEGO",
    "Inversion / Replante": "INVERSION / REPLANTE",
    "Servicios profesionales": "SERVICIOS PROFESIONALES",
    "Arriendos / Patentes / Seguros": "ARRIENDOS / PATENTES / SEGUROS",
    "Caja chica / Imprevistos": "CAJA CHICA / IMPREVISTOS",
}

# Mapeo sub_area → cultivo del MASTER
SUBAREA_TO_CULTIVO = {
    "NOGALES": "NOGALES",
    "CEREZOS": "CEREZOS",
    "VIVERO": "AVELLANOS",  # vivero es para replante
    "NUEVOS": "AVELLANOS",  # replante
    "EXPORTA": "GENERAL",
    "GENERAL": "GENERAL",
    "PLANTA": "GENERAL",
    "REMUNERACIONES": "GENERAL",
    "LAS CASAS": "GENERAL",
}


def map_agronomo_to_master(cargo, cargo_ll, sub_area):
    c = (cargo or "").strip().upper()
    cl = (cargo_ll or "").strip().upper()

    if c == "INSUMOS":
        if "PESTICIDAS" in cl or "FERTILIZANTES" in cl: return "INSUMOS AGRICOLAS"
        if "COMBUSTIBLE" in cl: return "COMBUSTIBLE"
        if "ELETRI" in cl or "ELECTRI" in cl: return "ENERGIA"
        if "RIEGO" in cl: return "RIEGO"
        return "INSUMOS AGRICOLAS"

    if c == "REMUNERACIONES":
        if "CAMPO" in cl or "CONTRATISTA" in cl: return "MANO DE OBRA TEMPORAL"
        if "ADMINISTRACION" in cl or "OFICINA" in cl: return "MANO DE OBRA PLANTA"
        if "CASA CAMPO" in cl: return "MANO DE OBRA PLANTA"
        return "MANO DE OBRA TEMPORAL"

    if c == "MANTENCION MAQUINARIA": return "MAQUINARIA - MANTENCION"

    if c == "MANTENCIONES":
        if "RIEGO" in cl: return "RIEGO"
        if "PLANTA" in cl: return "MANTENCION PLANTA"
        if "TERRENO" in cl: return "MANTENIMIENTO INFRAESTRUCTURA"
        if "OFICINA" in cl: return "MANTENIMIENTO INFRAESTRUCTURA"
        return "MAQUINARIA - MANTENCION"

    if c == "COSTOS FIJOS":
        if "ASESORIAS" in cl or "ASESORIA" in cl: return "SERVICIOS PROFESIONALES"
        return "ARRIENDOS / PATENTES / SEGUROS"

    if c == "ACTIVO FIJO":
        if "PLANTAS" in cl: return "INVERSION / REPLANTE"
        return "INVERSION ACTIVO PLANTA"

    if c == "EXPORTA": return "SERVICIOS DE EXPORTACION"
    if c == "ARRIENDOS": return "ARRIENDOS / PATENTES / SEGUROS"
    if c == "ESTUDIOS Y EVALUACIONES": return "SERVICIOS PROFESIONALES"
    if c == "HABILITACION CAMPO": return "MANTENIMIENTO INFRAESTRUCTURA"
    if c == "VIVERO": return "INVERSION / REPLANTE"

    if c == "VARIOS":
        if "HERRAMIENTAS" in cl: return "HERRAMIENTAS"
        if "UTILES OFICINA" in cl: return "MATERIALES"
        if "FLETES" in cl: return "TRANSPORTE"
        return "MATERIALES"

    return None


def _norm_str(s):
    if s is None: return ""
    return str(s).strip().upper().replace(".", "").replace("  ", " ")


def should_override(cat_master_norm, cat_agronomo):
    """Decide si reemplazar la categoría del MASTER con la del agronomo.

    Reglas:
    - Si categoría MASTER es más granular que agronomo, mantener (FERTILIZANTES vs INSUMOS AGRICOLAS).
    - Si categoría MASTER es 'SERVICIOS' (genérica), reemplazar.
    - Si categoría MASTER es muy diferente de agronomo, reemplazar.
    """
    if not cat_agronomo: return False
    if cat_master_norm == cat_agronomo: return False

    # Granularidad MASTER mayor (mantener)
    if cat_master_norm in ("FERTILIZANTES", "FITOSANITARIOS") and cat_agronomo == "INSUMOS AGRICOLAS":
        return False
    if cat_master_norm == "MANTENCION PLANTA" and cat_agronomo == "MAQUINARIA - MANTENCION":
        return False
    if cat_master_norm == "INVERSION ACTIVO PLANTA" and cat_agronomo == "INVERSION / REPLANTE":
        return False
    if cat_master_norm == "RIEGO" and cat_agronomo == "MAQUINARIA - MANTENCION":
        return False

    # Caso especial: HERRAMIENTAS vs MATERIALES — mantener la del agronomo (MATERIALES más amplio)
    # SERVICIOS genérico → siempre reemplazar
    return True


# ─── Cargar agronomo ──────────────────────────────────────────
print("[1/4] Cargando planilla agronomo...")
tmp_a = os.path.join(tempfile.gettempdir(), "agro_full.xlsx")
shutil.copy2(AGRO_PATH, tmp_a)
wb_a = load_workbook(tmp_a, read_only=True, data_only=True)
ws_a = wb_a["DATOS"]

agro_idx = {}
for row in ws_a.iter_rows(min_row=10, max_col=27, values_only=True):
    if not row[0]: continue
    prov = _norm_str(row[7])
    nro = str(row[18] or "").strip()
    if not prov or not nro: continue
    sub_area = row[3] or ""
    cat = map_agronomo_to_master(row[4], row[5], sub_area)
    cultivo = SUBAREA_TO_CULTIVO.get(str(sub_area).strip().upper(), "GENERAL")
    agro_idx[(prov, nro)] = {"cat": cat, "cultivo": cultivo, "sub_area": sub_area}
wb_a.close()
print(f"   {len(agro_idx)} facturas agronomo indexadas\n")

# ─── Procesar MASTER ──────────────────────────────────────────
print("[2/4] Procesando MASTER.Facturas...")
wb = load_workbook(EXCEL_PATH)
ws = wb["Facturas"]

stats = {
    "normalizadas": 0,
    "cultivo_asignado": 0,
    "cultivo_ya_correcto": 0,
    "categoria_reemplazada": 0,
    "categoria_mantenida_granular": 0,
    "no_match_agronomo": 0,
}
reemplazos = []

for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=20, values_only=False), start=2):
    if not row[0].value: continue
    prov = _norm_str(ws.cell(r_idx, 4).value)
    nro = str(ws.cell(r_idx, 7).value or "").strip()
    cat_master = ws.cell(r_idx, 17).value
    cultivo_master = ws.cell(r_idx, 18).value

    # 1) Normalizar categoría a MAYÚSCULAS
    if cat_master in NORMALIZE_MAP:
        cat_norm = NORMALIZE_MAP[cat_master]
        ws.cell(r_idx, 17).value = cat_norm
        cat_master = cat_norm
        stats["normalizadas"] += 1

    # Cruce con agronomo
    key = (prov, nro)
    if key not in agro_idx:
        stats["no_match_agronomo"] += 1
        continue

    agro = agro_idx[key]

    # 2) Asignar cultivo si está vacío o es GENERAL
    nuevo_cult = agro["cultivo"]
    if nuevo_cult and nuevo_cult != "GENERAL":
        if not cultivo_master or str(cultivo_master).strip().upper() == "GENERAL":
            ws.cell(r_idx, 18).value = nuevo_cult
            stats["cultivo_asignado"] += 1
        else:
            stats["cultivo_ya_correcto"] += 1

    # 3) Reclasificar si discrepancia real
    if should_override(cat_master, agro["cat"]):
        old = cat_master
        ws.cell(r_idx, 17).value = agro["cat"]
        stats["categoria_reemplazada"] += 1
        reemplazos.append((r_idx, old, agro["cat"], agro["sub_area"]))
    elif agro["cat"] and cat_master in ("FERTILIZANTES", "FITOSANITARIOS"):
        stats["categoria_mantenida_granular"] += 1

print(f"[3/4] Resumen:")
for k, v in stats.items():
    print(f"   {k}: {v}")

print(f"\nTop reemplazos (muestra 15):")
from collections import Counter
pares = Counter()
for r, old, new, sa in reemplazos:
    pares[(old, new)] += 1
for (o, n), c in pares.most_common(15):
    print(f"   {c:3d}× '{o}' → '{n}'")

print(f"\n[4/4] Guardando...")
wb.save(EXCEL_PATH)
wb.close()
print("Done!")
