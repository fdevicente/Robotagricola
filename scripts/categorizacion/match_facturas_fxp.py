#!/usr/bin/env python3
"""Match facturas Master ↔ FXP:
1. Si Master tiene una factura sin fecha_pago y aparece pagada en FXP → completar fecha_pago.
2. Si FXP tiene una factura (cargo con Fnnnnn) que NO existe en Master → agregarla.
"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP_PATH = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"

# Regex para extraer nro de factura de descripciones tipo "Copeval F6233577"
RE_FACT = re.compile(r"\bF\s*(\d{2,})\b", re.IGNORECASE)
# Patrones que NO son facturas (descartar)
NO_FACT_KEYWORDS = ["bh ", " bh", "remuneracion", "remuneración", "f29", "previred",
                     "pago tc", "cargo pac", "pago web", "comex", "tag",
                     "traspaso", "transferencia santa elisa", "rendicion",
                     "rendición", "aguinaldo", "saldo", "deposito",
                     "comision", "comisión", "pago impt"]


def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try: return datetime.strptime(v[:10], fmt).date()
            except: pass
    return None


def es_factura(desc, nota):
    text = (desc + " " + nota).lower()
    for kw in NO_FACT_KEYWORDS:
        if kw in text:
            return False
    return bool(RE_FACT.search(desc))


def extract_nro(desc):
    m = RE_FACT.search(desc)
    return m.group(1) if m else ""


def normalize(s):
    return (s or "").strip().upper().replace(".", "").replace("  ", " ")


# ─── Cargar FXP ──────────────────────────────────────────
print("[1/4] Cargando FXP.ScotiaBCO...")
tmp = os.path.join(tempfile.gettempdir(), "fxp_match.xlsx")
shutil.copy2(FXP_PATH, tmp)
wb_f = load_workbook(tmp, read_only=True, data_only=True)
ws_f = wb_f["ScotiaBCO"]

fxp_facturas = []  # cargos del banco que parecen facturas
for row in ws_f.iter_rows(min_row=6, values_only=True):
    if not row or len(row) < 11: continue
    fecha = _pd(row[2])
    if not fecha: continue
    try:
        cargo = float(row[6] or 0)
    except: continue
    if cargo <= 0: continue
    desc = str(row[5] or "")
    nota = str(row[10] or "")
    if not es_factura(desc, nota): continue

    # Extraer proveedor (antes del "F\d+")
    m = RE_FACT.search(desc)
    prov_raw = desc[:m.start()].strip(" -,")
    nro = m.group(1)
    fxp_facturas.append({
        "fecha_pago": fecha,
        "monto": cargo,
        "proveedor": prov_raw,
        "nro": nro,
        "desc_full": desc,
        "nota": nota.strip(" -"),
    })

wb_f.close()
print(f"   FXP cargos con # de factura: {len(fxp_facturas)}\n")

# ─── Cargar Master.Facturas ──────────────────────────────
print("[2/4] Cargando Master.Facturas...")
wb_m = load_workbook(EXCEL_PATH)
ws_m = wb_m["Facturas"]

master_idx = {}  # (proveedor_norm, nro) -> fila
master_last_row = 1
for r in range(2, ws_m.max_row + 1):
    if not ws_m.cell(r, 1).value: continue
    master_last_row = r
    prov = normalize(ws_m.cell(r, 4).value)
    nro = str(ws_m.cell(r, 7).value or "").strip()
    if not prov or not nro: continue
    master_idx[(prov, nro)] = r

print(f"   Master: {len(master_idx)} facturas indexadas\n")

# ─── Match: completar fecha_pago + agregar faltantes ─────
print("[3/4] Procesando matches...")
completadas_pago = 0
agregadas = 0
ya_pagadas = 0
no_match_prov = 0
agregadas_list = []

for f in fxp_facturas:
    prov_norm = normalize(f["proveedor"])
    nro = f["nro"]

    # Buscar match exacto
    key = (prov_norm, nro)
    if key in master_idx:
        # Existe: completar fecha_pago si está vacía
        r = master_idx[key]
        fecha_pago_actual = ws_m.cell(r, 3).value
        if not fecha_pago_actual or not str(fecha_pago_actual).strip():
            ws_m.cell(r, 3).value = f["fecha_pago"]
            completadas_pago += 1
        else:
            ya_pagadas += 1
        continue

    # No existe: buscar por nro solo (proveedor podría estar abreviado)
    matches_nro = [k for k in master_idx.keys() if k[1] == nro]
    if matches_nro:
        # Tomar primer match - probablemente mismo proveedor con grafía distinta
        r = master_idx[matches_nro[0]]
        fecha_pago_actual = ws_m.cell(r, 3).value
        if not fecha_pago_actual or not str(fecha_pago_actual).strip():
            ws_m.cell(r, 3).value = f["fecha_pago"]
            completadas_pago += 1
        else:
            ya_pagadas += 1
        continue

    # No existe: agregar al Master
    master_last_row += 1
    r = master_last_row
    ws_m.cell(r, 1).value = f["fecha_pago"]   # Fecha emisión (no la tenemos, usamos pago)
    ws_m.cell(r, 2).value = f["fecha_pago"]   # Vencimiento (no la tenemos)
    ws_m.cell(r, 3).value = f["fecha_pago"]   # Fecha pago
    ws_m.cell(r, 4).value = f["proveedor"]    # Proveedor
    ws_m.cell(r, 5).value = ""                # RUT
    ws_m.cell(r, 6).value = "Factura"
    ws_m.cell(r, 7).value = f["nro"]          # Nro factura
    ws_m.cell(r, 8).value = f["nota"] or f["desc_full"]  # Detalle
    ws_m.cell(r, 9).value = ""
    ws_m.cell(r, 15).value = f["monto"]       # Total por item
    ws_m.cell(r, 16).value = f["monto"]       # TOTAL FACTURA
    ws_m.cell(r, 17).value = ""                # Categoría (se asigna después con reglas)
    ws_m.cell(r, 18).value = ""                # Cultivo
    ws_m.cell(r, 19).value = ""                # Confianza
    ws_m.cell(r, 20).value = "FXP-import"     # Categorizado_por

    # Agregar al índice
    master_idx[(prov_norm, nro)] = r
    agregadas += 1
    agregadas_list.append((f["fecha_pago"], f["monto"], f["proveedor"][:35], f["nro"], f["nota"][:40]))

print(f"\n[4/4] Guardando...")
wb_m.save(EXCEL_PATH)
wb_m.close()

print(f"\n=== RESUMEN ===")
print(f"  Facturas FXP con cargo en banco: {len(fxp_facturas)}")
print(f"  Match con fecha_pago completada: {completadas_pago}")
print(f"  Ya tenían fecha_pago:            {ya_pagadas}")
print(f"  Agregadas al Master:             {agregadas}")

print(f"\n=== TOP 20 AGREGADAS ===")
agregadas_list.sort(key=lambda x: -x[1])
for f, m, p, n, nt in agregadas_list[:20]:
    print(f"  {f} | ${m:>10,.0f} | F{n:<10} | {p:35} | {nt}")
