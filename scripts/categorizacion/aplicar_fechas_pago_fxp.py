#!/usr/bin/env python3
"""Completa Fecha Pago en Master.Facturas desde FXP.FXP (pestaña de facturas).

Fuente de verdad = FXP.FXP (col 3 Fecha Pago, col 12 estado 'Pagada').
- Para cada factura PAGADA en FXP con fecha, busca en Master por nro (y proveedor)
  y completa la Fecha Pago SOLO en las líneas que la tienen vacía (no sobrescribe).
- Si la factura pagada no existe en Master, la agrega (con fechas/proveedor/monto de FXP).
Categoría se deja vacía: se asigna en el paso de categorización.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import shutil, tempfile, os
from datetime import date, datetime
from collections import defaultdict
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"


def _pd(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(v[:10], fmt).date()
            except Exception:
                pass
    return None


def norm(s):
    return (s or "").strip().upper().replace(".", "").replace("  ", " ")


def nrokey(v):
    s = str(v or "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.upper().replace(" ", "")


def _monto(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.upper().replace("USD", "").replace("$", "").strip()
        s = s.replace(".", "").replace(",", ".")  # 16.972,38 -> 16972.38
        try:
            return float(s)
        except ValueError:
            return 0.0
    return 0.0


# ── Cargar FXP.FXP (facturas pagadas) ──
print("[1/4] Cargando FXP.FXP (facturas pagadas)...")
tmp = os.path.join(tempfile.gettempdir(), "fxp_aplicar.xlsx")
shutil.copy2(FXP, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["FXP"]
pagadas = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    if str(row[11] or "").strip().upper() != "PAGADA":
        continue
    fpago = _pd(row[2])
    if not fpago:
        continue
    nro = nrokey(row[7])
    if not nro:
        continue
    pagadas.append({
        "f_emi": _pd(row[0]), "f_venc": _pd(row[1]), "f_pago": fpago,
        "prov": str(row[6] or ""), "nro": nro,
        "monto": _monto(row[8]),
        "nota": str(row[12] or ""),
    })
wb.close()
print(f"   Pagadas con fecha: {len(pagadas)}\n")

# ── Cargar Master.Facturas ──
print("[2/4] Cargando Master.Facturas...")
wb_m = load_workbook(EXCEL_PATH)
ws_m = wb_m["Facturas"]
por_nro = defaultdict(list)
last_row = 1
for r in range(2, ws_m.max_row + 1):
    if not ws_m.cell(r, 1).value:
        continue
    last_row = r
    nro = nrokey(ws_m.cell(r, 7).value)
    prov = norm(ws_m.cell(r, 4).value)
    if not nro:
        continue
    por_nro[nro].append({"fila": r, "prov": prov})
print(f"   Master última fila: {last_row}\n")

# ── Completar fechas / agregar faltantes ──
print("[3/4] Completando fechas de pago...")
lineas_completadas = 0
facturas_completadas = set()
agregadas = 0
agregadas_list = []

for f in pagadas:
    filas = por_nro.get(f["nro"])
    if filas:
        # Preferir líneas del mismo proveedor; si no, todas las del nro
        mismas = [x for x in filas if x["prov"] == norm(f["prov"])] or filas
        completo_alguna = False
        for x in mismas:
            cell = ws_m.cell(x["fila"], 3)
            if not (cell.value and str(cell.value).strip()):
                cell.value = f["f_pago"]
                lineas_completadas += 1
                completo_alguna = True
        if completo_alguna:
            facturas_completadas.add(f["nro"])
    else:
        # Agregar factura faltante
        last_row += 1
        r = last_row
        ws_m.cell(r, 1).value = f["f_emi"] or f["f_pago"]
        ws_m.cell(r, 2).value = f["f_venc"] or f["f_pago"]
        ws_m.cell(r, 3).value = f["f_pago"]
        ws_m.cell(r, 4).value = f["prov"]
        ws_m.cell(r, 6).value = "Factura"
        ws_m.cell(r, 7).value = f["nro"]
        ws_m.cell(r, 8).value = f["nota"]
        ws_m.cell(r, 15).value = f["monto"]
        ws_m.cell(r, 16).value = f["monto"]
        ws_m.cell(r, 20).value = "FXP-import"
        por_nro[f["nro"]].append({"fila": r, "prov": norm(f["prov"])})
        agregadas += 1
        agregadas_list.append((f["f_pago"], f["prov"][:30], f["nro"], f["monto"]))

print(f"[4/4] Guardando...")
wb_m.save(EXCEL_PATH)
wb_m.close()

print("\n=== RESUMEN ===")
print(f"  Líneas con fecha de pago completada: {lineas_completadas}")
print(f"  Facturas afectadas:                  {len(facturas_completadas)}")
print(f"  Facturas agregadas al Master:        {agregadas}")
for fp, p, n, m in agregadas_list:
    print(f"    + {fp} | {p:30} F{n:<10} | ${m:,.0f}")
