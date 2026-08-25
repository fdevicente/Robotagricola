#!/usr/bin/env python3
"""Aplica categorias usando la columna 'Asig Cta' del FXP por cruce fecha+monto."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP_PATH = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"


def _parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    return None


def map_asig_to_categoria(asig: str, notas: str = "") -> str | None:
    """Mapea 'Asig Cta' del FXP a categoría del MASTER."""
    a = (asig or "").lower().strip()
    n = (notas or "").lower()

    if not a or a == "-":
        return None

    # Prestamos y aportes
    if "préstamo" in a or "prestamo" in a: return "PRESTAMOS A OTRAS SOCIEDADES"
    if "aporte" in a: return "PRESTAMOS A OTRAS SOCIEDADES"
    if "inversión" in a or "inversion" in a:
        # Inversion en planta vs inversion financiera
        if "activo" in a or "planta" in a: return "INVERSION ACTIVO PLANTA"
        return "PRESTAMOS A OTRAS SOCIEDADES"

    # Impuestos
    if "impuesto" in a: return "IMPUESTOS"

    # Transferencias internas (entre cuentas propias)
    if "transfencia" in a or "transferencia" in a: return "TRANSFERENCIA INTERNA"

    # Cambio divisa
    if "comex" in a or "dolares" in a or "usd" in a: return "CAMBIO DIVISA"

    # Activos planta
    if "activo" in a and ("camarico" in a or "planta" in a):
        return "INVERSION ACTIVO PLANTA"

    # Gastos cultivos / operativos
    if "gto" in a or "gasto" in a:
        if "cerezo" in a: return "MANO DE OBRA TEMPORAL"
        if "nogal" in a or "nuez" in a: return "MANO DE OBRA TEMPORAL"
        if "avellano" in a: return "MANO DE OBRA TEMPORAL"
        if "camarico" in a: return "MANO DE OBRA TEMPORAL"
        return "MANO DE OBRA TEMPORAL"

    # Maquinaria
    if "barredor" in a or "soplador" in a or "sorter" in a:
        return "MAQUINARIA - MANTENCION"

    # Helicóptero
    if "helo" in a or "helic" in a or "aeronave" in a:
        return "MANTENIMIENTO HELICOPTERO"

    # Saldo inicial / saldo
    if "saldo" in a: return None  # ignorar

    return None


# Cargar FXP
print("[1/3] Cargando FXP...")
tmp_f = os.path.join(tempfile.gettempdir(), "fxp_asig.xlsx")
shutil.copy2(FXP_PATH, tmp_f)
wb_f = load_workbook(tmp_f, read_only=True, data_only=True)
ws_f = wb_f["ScotiaBCO"]

# Indexar FXP por (fecha, monto): (asig_cta, notas)
fxp_idx = {}
for row in ws_f.iter_rows(min_row=6, values_only=True):
    if not row or len(row) < 11: continue
    fecha = _parse_date(row[2])
    if not fecha: continue
    try:
        monto = int(round(float(row[6] or 0)))
    except (TypeError, ValueError):
        continue
    asig = str(row[9] or "")
    notas = str(row[10] or "")
    fxp_idx[(fecha.isoformat(), monto)] = (asig, notas)
wb_f.close()
print(f"   FXP indexado: {len(fxp_idx)} movimientos\n")

# Procesar Master
print("[2/3] Procesando Master...")
wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]

stats = {"procesados": 0, "sin_fxp_match": 0, "asig_no_mapeable": 0, "actualizadas": 0}
cat_counts = {}
asig_no_mapeadas = {}

for row in range(2, ws.max_row + 1):
    cat_actual = ws.cell(row, 8).value
    if cat_actual != "REVISAR":
        continue

    fecha_val = ws.cell(row, 1).value
    fecha = _parse_date(fecha_val)
    if not fecha or fecha < date(2021, 1, 1):
        continue

    try:
        cargo = int(round(float(ws.cell(row, 4).value or 0)))
    except (TypeError, ValueError):
        continue

    stats["procesados"] += 1

    fxp_data = fxp_idx.get((fecha.isoformat(), cargo))
    if not fxp_data:
        stats["sin_fxp_match"] += 1
        continue

    asig, notas = fxp_data
    nueva_cat = map_asig_to_categoria(asig, notas)
    if not nueva_cat:
        stats["asig_no_mapeable"] += 1
        asig_no_mapeadas[asig] = asig_no_mapeadas.get(asig, 0) + 1
        continue

    ws.cell(row, 8).value = nueva_cat
    ws.cell(row, 9).value = "GENERAL"
    stats["actualizadas"] += 1
    cat_counts[nueva_cat] = cat_counts.get(nueva_cat, 0) + 1

print("[3/3] Guardando...")
wb.save(EXCEL_PATH)
wb.close()

print("\n=== RESUMEN ===")
for k, v in stats.items():
    print(f"  {k}: {v}")

print("\n=== CATEGORIAS APLICADAS ===")
for cat, n in sorted(cat_counts.items(), key=lambda x: -x[1]):
    print(f"  {cat}: {n}")

print("\n=== TOP ASIG SIN MAPEAR (no se aplicaron) ===")
for asig, n in sorted(asig_no_mapeadas.items(), key=lambda x: -x[1])[:20]:
    print(f"  '{asig}': {n}")
