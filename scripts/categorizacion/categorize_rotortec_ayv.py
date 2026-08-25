#!/usr/bin/env python3
"""Categoriza todos los movimientos Rotortec/AYV como PRESTAMOS A OTRAS SOCIEDADES."""
from openpyxl import load_workbook
from config import EXCEL_PATH

CATEGORIA = "PRESTAMOS A OTRAS SOCIEDADES"

print("Abriendo Master...")
wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]
print(f"Iterando {ws.max_row} filas...")

candidatos = []
for row in range(2, ws.max_row + 1):
    desc = str(ws.cell(row, 2).value or "")
    ref = str(ws.cell(row, 3).value or "")
    text = f"{desc} {ref}".lower()
    if "rotortec" in text or "ayv" in text:
        cat_actual = ws.cell(row, 8).value
        cargo = float(ws.cell(row, 4).value or 0)
        candidatos.append((row, cat_actual, cargo, desc[:60]))

print(f"\nTotal Rotortec/AYV: {len(candidatos)} filas")
print(f"Monto total cargos: ${sum(c[2] for c in candidatos):,.0f} CLP\n")

# Aplicar a TODOS (independiente de categoria actual)
for row, cat_old, cargo, desc in candidatos:
    ws.cell(row, 8).value = CATEGORIA
    ws.cell(row, 9).value = "GENERAL"

print(f"Aplicado '{CATEGORIA}' a {len(candidatos)} filas")
print("Guardando...")
wb.save(EXCEL_PATH)
wb.close()
print("Done!")

print("\nDistribución de categorías reemplazadas:")
from collections import Counter
counter = Counter(c[1] for c in candidatos)
for cat, count in counter.most_common():
    print(f"  {cat}: {count}")
