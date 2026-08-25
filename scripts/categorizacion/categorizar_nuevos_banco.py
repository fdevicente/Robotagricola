#!/usr/bin/env python3
"""Categoriza los movimientos nuevos de Cuenta Banco (desde una fecha).

Solo toca filas sin categoría o en REVISAR. Lo que no calza con una regla
clara queda en REVISAR para que el usuario lo defina en /banco/revisar.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from collections import Counter
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH
from excel_manager import _save_wb

import os
DESDE = date.fromisoformat(os.getenv("CATEG_DESDE", "2026-06-10"))

# RUTs del personal fijo → sueldos
RUTS_PERSONAL = {
    "13373052-4": "Juan Parada", "12318508-0": "Agustin Mora",
    "21331792-K": "Patricio Mora", "20230894-5": "Javier Gonzalez",
    "11768374-5": "Ramiro Amigo", "9850887-2": "Felicito Amigo",
    "17407271-K": "Felix De Vicente", "9359341-3": "Felix De Vicente",
}

REGLAS = [
    (["previred", "cotiz.previred"], "MANO DE OBRA PLANTA"),
    (["pago impto sii", "impto sii", "f29"], "IMPUESTOS"),
    (["cargo comex", "comision", "gasto banc"], "GASTOS BANCARIOS"),
    (["copec", "petrobras", "shell", "lipigas", "abastible"], "COMBUSTIBLE"),
    (["sodimac", "easy", "construmart", "ferreteria", "imperial"], "MATERIALES"),
    (["bridgestone", "venom", "neumatico", "vulcaniz"], "GASTOS VEHICULOS"),
    (["gruasmestre", "grua", "maquinaria", "maestranza"], "MAQUINARIA - MANTENCION"),
    (["agri for", "copeval", "martinez y vald", "agrocampo", "anasac",
      "agrokimun"], "INSUMOS AGRICOLAS"),
    (["s invest", "s-invest", "sinvest", "cge", "enel"], "ENERGIA"),
    (["francisco donos", "9424305-k"], "SERVICIOS PROFESIONALES"),
    (["pacific", "valbifrut"], "INGRESO VENTAS"),
]


def _pd(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for f in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(v[:10], f).date()
            except Exception:
                pass
    return None


def clasificar(desc, fecha, cargo, abono):
    t = " " + desc.lower() + " "

    # Abonos (ingresos)
    if abono > 0:
        for kws, cat in REGLAS:
            if cat.startswith("INGRESO") and any(k in t for k in kws):
                return cat
        if "proveedore" in t and "81290800" in t:      # nota de crédito Copeval
            return "REINTEGROS Y DEVOLUCIONES"
        return None  # ingreso desconocido → REVISAR

    # Sueldos: RUT de personal + monto de sueldo en el cambio de mes.
    # Se pagan el día 1-3 o los últimos días del mes anterior (ej: 31-jul).
    for rut in RUTS_PERSONAL:
        if rut.lower() in t:
            fin_de_mes = fecha.day >= 28
            if (fecha.day <= 3 or fin_de_mes) and cargo >= 400_000:
                return "MANO DE OBRA PLANTA"
            return None  # otro pago a la misma persona → revisar (rendición, etc.)
    # Sueldo del dueño vía su sociedad
    if "77912665" in t or "crave spa" in t:
        return "MANO DE OBRA PLANTA"

    for kws, cat in REGLAS:
        if cat.startswith("INGRESO"):
            continue
        if any(k in t for k in kws):
            return cat
    return None


wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]
aplicadas = Counter()
revisar = []
for r in range(2, ws.max_row + 1):
    f = _pd(ws.cell(r, 1).value)
    if not f or f < DESDE:
        continue
    cat_actual = str(ws.cell(r, 8).value or "").strip()
    if cat_actual and cat_actual != "REVISAR":
        continue
    desc = str(ws.cell(r, 2).value or "")
    try:
        cargo = float(ws.cell(r, 4).value or 0)
        abono = float(ws.cell(r, 5).value or 0)
    except Exception:
        cargo = abono = 0
    nueva = clasificar(desc, f, cargo, abono)
    if nueva:
        ws.cell(r, 8).value = nueva
        if not ws.cell(r, 9).value:
            ws.cell(r, 9).value = "GENERAL"
        aplicadas[nueva] += 1
    else:
        ws.cell(r, 8).value = "REVISAR"
        revisar.append((f, cargo or -abono, desc))

_save_wb(wb)
wb.close()

print("=== Categorías aplicadas ===")
for c, n in aplicadas.most_common():
    print(f"  {n:3} | {c}")
print(f"\n=== Quedan en REVISAR: {len(revisar)} ===")
for f, monto, d in sorted(revisar, key=lambda x: -abs(x[1])):
    signo = "-" if monto > 0 else "+"
    print(f"  {f} {signo}${abs(monto):>12,.0f}  {d[:46]}")
