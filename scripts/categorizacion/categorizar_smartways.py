"""SmartWays 4-ago-2026: gastos de mantención del helicóptero que se devolvieron.

El dueño confirmó (6-ago) que son mantenciones del helicóptero que hubo y se
están pagando, y que **ya no queda deuda por el helicóptero**. Van a
MANTENIMIENTO HELICOPTERO, que está en EXCLUIR_PROYECCION (operación
descontinuada) → no inflan la proyección de la temporada.
"""
import shutil
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb
from modules.cash_flow.projector import EXCLUIR_PROYECCION

CATEGORIA = "MANTENIMIENTO HELICOPTERO"
NOTA = "Mantenciones del helicóptero — saldado, no queda deuda (confirmado 6-ago-2026)"
COL_DESC, COL_CARGO, COL_CAT, COL_CULT = 2, 4, 8, 9

assert CATEGORIA in EXCLUIR_PROYECCION, "la categoría debe estar excluida del flujo"

resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"Respaldo: {resp}\n")

wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]

total = 0.0
n = 0
for r in range(2, ws.max_row + 1):
    desc = str(ws.cell(r, COL_DESC).value or "")
    if "smartway" not in desc.lower():
        continue
    if str(ws.cell(r, COL_CAT).value or "").strip():
        continue
    monto = float(ws.cell(r, COL_CARGO).value or 0)
    ws.cell(r, COL_CAT).value = CATEGORIA
    ws.cell(r, COL_CULT).value = "GENERAL"
    total += monto
    n += 1
    print(f"  fila {r:>4} | {-monto:>13,.0f} | {desc[:42]}")

if n:
    _save_wb(wb)
    print(f"\n✅ {n} movimientos → {CATEGORIA}  (total ${total:,.0f})")
    print(f"   {NOTA}")
    print("\n   Esta categoría NO entra en la proyección "
          "(operación descontinuada), así que el colchón no se mueve.")
else:
    print("Nada pendiente de categorizar en SmartWays.")
wb.close()
