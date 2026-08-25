#!/usr/bin/env python3
"""Reglas v9 - patrones que faltan en mayo 2026."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

RULES = [
    (["remuneracion abril", "remuneracion marzo", "remuneracion mayo",
       "remuneracion junio", "remuneracion julio", "remuneracion agosto",
       "remuneracion septiembre", "remuneracion octubre", "remuneracion noviembre",
       "remuneracion diciembre", "remuneracion enero", "remuneracion febrero",
       "remuneracion ", "remuneración"], "MANO DE OBRA PLANTA"),
    (["previred"], "MANO DE OBRA PLANTA"),
    (["lipigas"], "COMBUSTIBLE"),
    (["copec"], "COMBUSTIBLE"),
    (["leonardo ivan nunez", "leonardo nunez", "leonardo nuñez"],
     "SERVICIOS PROFESIONALES"),
    (["nogalte sur", "nogaltec sur", "nogaltec"], "SERVICIOS PROFESIONALES"),
    (["central quemadores", "quemador"], "MAQUINARIA - MANTENCION"),
    (["implementos f", "implementos para"], "MATERIALES"),
    (["bridgestone"], "GASTOS VEHICULOS"),
]


def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None


def match(text):
    t = text.lower()
    for kws, cat in RULES:
        for kw in kws:
            if kw in t:
                return cat
    return None


wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]

actualizadas = 0
counts = {}
for r in range(2, ws.max_row + 1):
    cat = ws.cell(r, 8).value
    if cat != "REVISAR": continue
    fecha = _pd(ws.cell(r, 1).value)
    if not fecha or fecha < date(2021, 1, 1): continue
    desc = str(ws.cell(r, 2).value or "")
    ref = str(ws.cell(r, 3).value or "")
    new = match(desc + " " + ref)
    if new:
        ws.cell(r, 8).value = new
        ws.cell(r, 9).value = "GENERAL"
        counts[new] = counts.get(new, 0) + 1
        actualizadas += 1

print(f"Actualizadas: {actualizadas}")
for c, n in sorted(counts.items(), key=lambda x: -x[1]):
    print(f"  {c}: {n}")

wb.save(EXCEL_PATH)
wb.close()
print("Done!")
