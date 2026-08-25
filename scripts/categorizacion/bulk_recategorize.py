#!/usr/bin/env python3
"""Recategorización masiva del banco con todas las reglas del usuario."""
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP_PATH = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"

# ─── Reglas del usuario (orden de prioridad) ───────────────────────
RULES = [
    # (keywords_lowercase, categoria)
    (["rotortec", "ayv"], "PRESTAMOS A OTRAS SOCIEDADES"),
    (["smartways"], "PRESTAMOS A OTRAS SOCIEDADES"),
    (["copeval", "martinez", "valdivieso"], "INSUMOS AGRICOLAS"),
    (["cge"], "ENERGIA"),
    (["packman", "sortek"], "MAQUINARIA - MANTENCION"),
    (["cals"], "INSUMOS AGRICOLAS"),
    (["astara"], "INVERSION VEHICULOS"),
    (["alpabesa"], "MANO DE OBRA TEMPORAL"),
    (["arrayan", "aeromar"], "MANTENIMIENTO HELICOPTERO"),
    (["pago cuota", "leasing"], "LEASING"),
    (["fondo mutuo", "fondos mutuos", "credito", "crédito", "navidad",
       "préstamo", "prestamo", "inversión", "inversion"],
     "PRESTAMOS A OTRAS SOCIEDADES"),
]

CATEGORIA_HISTORICA = "PRE-2021 HISTORICO"
CUTOFF = date(2021, 1, 1)


def _parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(val[:10], fmt).date()
            except ValueError:
                pass
    return None


def match_rule(text: str) -> str | None:
    t = text.lower()
    for kws, cat in RULES:
        for kw in kws:
            if kw in t:
                return cat
    return None


# ─── Paso 1: Cargar FXP para cruce ─────────────────────────────────
print("[1/3] Cargando FXP para cruce de notas...")
tmp_fxp = os.path.join(tempfile.gettempdir(), "fxp_xref.xlsx")
shutil.copy2(FXP_PATH, tmp_fxp)
wb_fxp = load_workbook(tmp_fxp, read_only=True, data_only=True)
ws_fxp = wb_fxp["ScotiaBCO"]

# Index: (fecha_iso, monto_int) -> (descripcion, asig_cta, notas)
fxp_idx = {}
for row in ws_fxp.iter_rows(min_row=6, values_only=True):
    if not row or len(row) < 11:
        continue
    fecha = _parse_date(row[2])
    monto = row[6]
    if not fecha or monto is None:
        continue
    try:
        monto_int = int(round(float(monto)))
    except (TypeError, ValueError):
        continue
    desc = str(row[5] or "")
    asig = str(row[9] or "")
    notas = str(row[10] or "")
    key = (fecha.isoformat(), monto_int)
    fxp_idx[key] = (desc, asig, notas)

wb_fxp.close()
print(f"   FXP indexado: {len(fxp_idx)} movimientos")

# ─── Paso 2: Procesar Master ───────────────────────────────────────
print("\n[2/3] Procesando Master...")
wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]

stats = {
    "pre_2021_marcadas": 0,
    "reglas_aplicadas": 0,
    "fxp_enriquecidas": 0,
    "fxp_matches": 0,
    "ya_categorizadas_ok": 0,
    "sin_cambio": 0,
}
rule_counts = {}

for row in range(2, ws.max_row + 1):
    fecha_val = ws.cell(row, 1).value
    fecha = _parse_date(fecha_val)
    if not fecha:
        continue

    cat_actual = ws.cell(row, 8).value
    desc = str(ws.cell(row, 2).value or "")
    ref = str(ws.cell(row, 3).value or "")
    cargo = ws.cell(row, 4).value
    try:
        cargo_int = int(round(float(cargo or 0)))
    except (TypeError, ValueError):
        cargo_int = 0

    # Marcar pre-2021 como histórico
    if fecha < CUTOFF:
        if cat_actual not in (CATEGORIA_HISTORICA, "PRESTAMOS A OTRAS SOCIEDADES"):
            ws.cell(row, 8).value = CATEGORIA_HISTORICA
            ws.cell(row, 9).value = "GENERAL"
            stats["pre_2021_marcadas"] += 1
        continue

    # Solo aplicar reglas/FXP a REVISAR o vacías post-2021
    if cat_actual and cat_actual not in ("REVISAR", None, ""):
        stats["ya_categorizadas_ok"] += 1
        continue

    # Enriquecer con FXP
    fxp_data = fxp_idx.get((fecha.isoformat(), cargo_int))
    enriched = desc + " " + ref
    if fxp_data:
        fxp_desc, fxp_asig, fxp_notas = fxp_data
        enriched += " " + fxp_desc + " " + fxp_notas
        stats["fxp_matches"] += 1

    # Aplicar reglas
    cat = match_rule(enriched)
    if cat:
        ws.cell(row, 8).value = cat
        ws.cell(row, 9).value = "GENERAL"
        stats["reglas_aplicadas"] += 1
        if fxp_data:
            stats["fxp_enriquecidas"] += 1
        rule_counts[cat] = rule_counts.get(cat, 0) + 1
    else:
        stats["sin_cambio"] += 1

print(f"\n[3/3] Guardando...")
wb.save(EXCEL_PATH)
wb.close()

print("\n═══ RESUMEN ═══")
for k, v in stats.items():
    print(f"  {k}: {v}")

print("\n═══ POR CATEGORIA APLICADA ═══")
for cat, n in sorted(rule_counts.items(), key=lambda x: -x[1]):
    print(f"  {cat}: {n}")
