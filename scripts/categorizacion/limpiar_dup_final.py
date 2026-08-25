#!/usr/bin/env python3
"""Tercera pasada: líneas redundantes que repiten el total de la factura."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook
from config import EXCEL_PATH
from excel_manager import _save_wb

# (fila, nro esperado, glosa esperada (fragmento), motivo)
CASOS = [
    (1325, "23361436",  "diesel",       "Adm. Ventas: línea Diesel repetida"),
    (1305, "145829553", "solerillas",   "Sodimac: línea resumen que duplica el detalle"),
    (2131, "6432363",   "despacho",     "Copeval: despacho quedó con el total completo"),
]


def nrokey(v):
    s = str(v or "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.upper().replace(" ", "").replace("-", "")


wb = load_workbook(EXCEL_PATH)
ws = wb["Facturas"]
ok = []
for fila, nro, frag, motivo in CASOS:
    nro_actual = nrokey(ws.cell(fila, 7).value)
    glosa = str(ws.cell(fila, 8).value or "").lower()
    if nro_actual == nrokey(nro) and frag in glosa:
        ok.append((fila, nro, motivo, ws.cell(fila, 8).value))
    else:
        print(f"  ⚠️ OMITIDA fila {fila}: esperaba F{nro}/'{frag}', "
              f"hay F{nro_actual}/'{glosa[:30]}'")

for fila, nro, motivo, glosa in sorted(ok, key=lambda x: -x[0]):
    print(f"  ✔ fila {fila:>5} borrada | F{nro:<11} | {str(glosa)[:34]:34} | {motivo}")
    ws.delete_rows(fila)

_save_wb(wb)
wb.close()
print(f"\n✅ {len(ok)} líneas eliminadas.")
