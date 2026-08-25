#!/usr/bin/env python3
"""Propaga glosas editadas de FXP.ScotiaBCO -> Master.CuentaBanco.

FXP es la fuente de verdad. Para cada movimiento con clave única (fecha, cargo,
abono), si la glosa del Master difiere de la de FXP, se actualiza con la de FXP.
- Solo claves ÚNICAS (si el mismo día+monto aparece varias veces, se omite: ambiguo).
- Se omiten claves dudosas marcadas a mano (entidad distinta).
También limpia 2 fechas de pago falsas en Facturas (F94 Contreras, F6776 Indelec, NN).
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import shutil, tempfile, os
from datetime import date, datetime
from collections import defaultdict, Counter
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"

# Claves (fecha_iso, cargo, abono) a NO tocar (entidad distinta, revisar a mano)
SKIP = {
    ("2026-01-28", 257190, 0),
    ("2026-03-23", 500000, 0),
    ("2026-03-23", 785200, 0),
}


def _pd(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for f in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(v[:10], f).date()
            except Exception:
                pass
    return None


def i(v):
    try:
        return int(round(float(v or 0)))
    except Exception:
        return 0


def cln(s):
    return " ".join(str(s or "").split()).strip()


def nrokey(v):
    s = str(v or "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.upper().replace(" ", "")


# ── FXP.ScotiaBCO: glosa por clave única ──
print("[1/4] Cargando FXP.ScotiaBCO...")
tmp = os.path.join(tempfile.gettempdir(), "fxp_glosas.xlsx")
shutil.copy2(FXP, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["ScotiaBCO"]
desc_por_clave = {}
cuenta_clave = Counter()
for row in ws.iter_rows(min_row=6, values_only=True):
    if not row or len(row) < 11:
        continue
    f = _pd(row[2])
    if not f:
        continue
    c, d = i(row[6]), i(row[7])
    if c == 0 and d == 0:
        continue
    k = (f.isoformat(), c, d)
    cuenta_clave[k] += 1
    desc_por_clave[k] = cln(row[5])
wb.close()
# Solo claves únicas
unicas = {k: v for k, v in desc_por_clave.items() if cuenta_clave[k] == 1}
print(f"   Claves únicas en FXP: {len(unicas)}\n")

# ── Master.CuentaBanco: actualizar glosa ──
print("[2/4] Actualizando glosas en Master.CuentaBanco...")
wb_m = load_workbook(EXCEL_PATH)
ws_m = wb_m["Cuenta Banco"]
cambiadas = 0
por_anio = Counter()
ejemplos = []
for r in range(2, ws_m.max_row + 1):
    f = _pd(ws_m.cell(r, 1).value)
    if not f:
        continue
    c, d = i(ws_m.cell(r, 4).value), i(ws_m.cell(r, 5).value)
    k = (f.isoformat(), c, d)
    if k in SKIP or k not in unicas:
        continue
    actual = cln(ws_m.cell(r, 2).value)
    nueva = unicas[k]
    if nueva and nueva.upper() != actual.upper():
        ws_m.cell(r, 2).value = nueva
        cambiadas += 1
        por_anio[f.year] += 1
        if len(ejemplos) < 12:
            ejemplos.append((f, actual[:34], nueva[:34]))
print(f"   Glosas actualizadas: {cambiadas}")

# ── Limpiar 2 fechas de pago falsas (NN) en Facturas ──
print("\n[3/4] Limpiando 2 fechas de pago falsas (NN)...")
ws_f = wb_m["Facturas"]
limpiadas = 0
objetivo = [("CONTRERAS", "94"), ("INDELEC", "6776")]
for r in range(2, ws_f.max_row + 1):
    prov = str(ws_f.cell(r, 4).value or "").upper()
    nro = nrokey(ws_f.cell(r, 7).value)
    for prov_kw, nro_t in objetivo:
        if prov_kw in prov and nro == nro_t:
            if ws_f.cell(r, 3).value:
                ws_f.cell(r, 3).value = None
                limpiadas += 1
                print(f"   - fila {r}: {prov[:30]} F{nro} -> fecha pago borrada")
print(f"   Fechas limpiadas: {limpiadas}")

print("\n[4/4] Guardando...")
wb_m.save(EXCEL_PATH)
wb_m.close()

print("\n=== RESUMEN ===")
print(f"  Glosas propagadas: {cambiadas}")
for y in sorted(por_anio):
    print(f"     {y}: {por_anio[y]}")
print(f"  Fechas falsas limpiadas: {limpiadas}")
print("\n  Ejemplos:")
for f, a, n in ejemplos:
    print(f"   {f} | '{a}'  ->  '{n}'")
