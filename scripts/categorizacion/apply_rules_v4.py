#!/usr/bin/env python3
"""Reglas v4 + reclasificación de MAQUINARIA con 'planta' → MANTENCION PLANTA."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP_PATH = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"


def _parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    return None


KEYWORD_RULES = [
    # MANTENCION PLANTA (prioridad alta, antes de MAQUINARIA - MANTENCION)
    (["mantenimiento planta", "mantención planta", "reparaciones planta",
       "reparacion planta", "mantencion planta", "planta proceso",
       "planta cosecha", "reparaciones y mantención planta",
       "reparaciones y mantencion planta"], "MANTENCION PLANTA"),
    (["sorter", "despelonadora"], "MANTENCION PLANTA"),

    # MANTENIMIENTO INFRAESTRUCTURA - trabajos menores construcción
    (["jose luis menares", "menares", "trabajos menores construccion",
       "trabajos menores construcción"], "MANTENIMIENTO INFRAESTRUCTURA"),

    # GASTOS VEHICULOS
    (["pago total tag", "tag total", " tag ", "pago tag", "tag tbpw",
       "tag camioneta"], "GASTOS VEHICULOS"),

    # MANTENIMIENTO HELICOPTERO
    (["repuestos helicoptero", "repuestos helicóptero",
       "francisco pena"], "MANTENIMIENTO HELICOPTERO"),

    # SERVICIOS DE EXPORTACION
    (["servicios logisticos ecosmart", "servicios logísticos ecosmart",
       "mov contenedor", "movimiento contenedor", "handling", "gateout",
       "gate out", "sps sag", "sag exportacion", "sag exportación",
       "visita sag", "visitas sag"], "SERVICIOS DE EXPORTACION"),
    (["ecosmart"], "SERVICIOS DE EXPORTACION"),

    # MANO DE OBRA TEMPORAL
    (["bh ", " bh", "boleta honorarios"], "MANO DE OBRA TEMPORAL"),
    (["jornal", "jornalero", "jornaleros"], "MANO DE OBRA TEMPORAL"),
    (["trabajos cosecha", "apoyo poda", "trabajo enero", "trabajo cosecha"],
     "MANO DE OBRA TEMPORAL"),

    # MAQUINARIA - MANTENCION (después de MANTENCION PLANTA)
    (["agrocampo", "felipe noguera", "ricardo mahla", "kohen"],
     "MAQUINARIA - MANTENCION"),
    (["repuestos tractor", "filtros tractor", "filtro tractor", "aceite tractor",
       "cambio aceite", "mantencion camioneta", "mantención camioneta",
       "mantencion tractor", "mantención tractor", "podadora"],
     "MAQUINARIA - MANTENCION"),

    # RIEGO
    (["mantenimiento riego", "mantención riego", "mantencion riego",
       "wissecon", "instalacion riego", "instalación riego"], "RIEGO"),

    # INSUMOS AGRICOLAS
    (["fosfonat"], "INSUMOS AGRICOLAS"),
    (["gmt f"], "INSUMOS AGRICOLAS"),  # GMT facturas son insumos

    # SERVICIOS PROFESIONALES
    (["miguel marin", "miguel marín", "piloto"], "SERVICIOS PROFESIONALES"),
    (["capital office", "servicio postal", "domicilio tributario"],
     "SERVICIOS PROFESIONALES"),
    (["rendicion viajes", "rendición viajes", "rendicion de viaje",
       "rendición de viaje"], "SERVICIOS PROFESIONALES"),

    # REINTEGROS
    (["pago diferencias facturas", "pago diferencias factura",
       "diferencia facturas"], "REINTEGROS Y DEVOLUCIONES"),

    # Reglas anteriores - mantenemos por completitud
    (["traspaso banco", "traspaso bcos", "traspaso de banco",
       "traspaso agricola santa elisa", "traspaso entre banco"],
     "TRANSFERENCIA INTERNA"),
    (["s-invest", "s invest", "sinvest"], "ENERGIA"),
    (["agencia aduana", "aduana"], "SERVICIOS DE EXPORTACION"),
    (["intralog"], "SERVICIOS DE EXPORTACION"),
    (["carmen gloria martinez", "carmengloria martinez", "carmen gloria"],
     "SERVICIOS DE EXPORTACION"),
    (["francisco donoso"], "SERVICIOS PROFESIONALES"),
]


def match_keywords(text: str) -> str | None:
    t = text.lower()
    for kws, cat in KEYWORD_RULES:
        for kw in kws:
            if kw in t:
                return cat
    return None


# Cargar FXP
print("[1/4] FXP...")
tmp_f = os.path.join(tempfile.gettempdir(), "fxp_v4.xlsx")
shutil.copy2(FXP_PATH, tmp_f)
wb_f = load_workbook(tmp_f, read_only=True, data_only=True)
ws_f = wb_f["ScotiaBCO"]
fxp_idx = {}
for row in ws_f.iter_rows(min_row=6, values_only=True):
    if not row or len(row) < 11: continue
    fecha = _parse_date(row[2])
    if not fecha: continue
    try:
        monto = int(round(float(row[6] or 0)))
    except: continue
    fxp_idx[(fecha.isoformat(), monto)] = (
        str(row[5] or ""), str(row[10] or "")
    )
wb_f.close()
print(f"   {len(fxp_idx)}\n")

print("[2/4] Aplicando v4 a REVISAR post-2021...")
wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]

cat_counts = {}
actualizadas = 0
total = 0

for row in range(2, ws.max_row + 1):
    cat = ws.cell(row, 8).value
    if cat != "REVISAR": continue
    fecha = _parse_date(ws.cell(row, 1).value)
    if not fecha or fecha < date(2021, 1, 1): continue
    total += 1

    desc = str(ws.cell(row, 2).value or "")
    ref = str(ws.cell(row, 3).value or "")
    try:
        cargo = int(round(float(ws.cell(row, 4).value or 0)))
    except: cargo = 0

    fxp_desc, fxp_notas = fxp_idx.get((fecha.isoformat(), cargo), ("", ""))
    enriched = f"{desc} {ref} {fxp_desc} {fxp_notas}"

    nueva = match_keywords(enriched)
    if nueva:
        ws.cell(row, 8).value = nueva
        ws.cell(row, 9).value = "GENERAL"
        cat_counts[nueva] = cat_counts.get(nueva, 0) + 1
        actualizadas += 1

print(f"   Procesados: {total}, Actualizados: {actualizadas}\n")

# Reclasificar MAQUINARIA - MANTENCION con 'planta' → MANTENCION PLANTA
print("[3/4] Reclasificando MAQUINARIA - MANTENCION con 'planta'...")
reclassed = 0
for row in range(2, ws.max_row + 1):
    cat = ws.cell(row, 8).value
    if cat != "MAQUINARIA - MANTENCION":
        continue
    desc = str(ws.cell(row, 2).value or "").lower()
    ref = str(ws.cell(row, 3).value or "").lower()
    try:
        cargo = int(round(float(ws.cell(row, 4).value or 0)))
    except: cargo = 0
    fecha = _parse_date(ws.cell(row, 1).value)
    fxp_d, fxp_n = ("", "")
    if fecha:
        fxp_d, fxp_n = fxp_idx.get((fecha.isoformat(), cargo), ("", ""))
    text = f"{desc} {ref} {fxp_d} {fxp_n}".lower()
    if "planta" in text or "sorter" in text or "despelonadora" in text:
        ws.cell(row, 8).value = "MANTENCION PLANTA"
        reclassed += 1

print(f"   Reclasificadas: {reclassed}\n")

print("[4/4] Guardando...")
wb.save(EXCEL_PATH)
wb.close()

print("\n=== NUEVAS APLICADAS ===")
for cat, n in sorted(cat_counts.items(), key=lambda x: -x[1]):
    print(f"  {cat}: {n}")
