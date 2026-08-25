#!/usr/bin/env python3
"""Mapeo final por keywords (descripción + FXP notas) + normalización mayúsculas."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP_PATH = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"


def _parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    return None


# Reglas (orden importa, más específicas primero)
KEYWORD_RULES = [
    (["gestora y tecnolog", "devolucion gestora"], "REINTEGROS Y DEVOLUCIONES"),
    (["reintegro"], "REINTEGROS Y DEVOLUCIONES"),
    (["devolucion", "devolución"], "REINTEGROS Y DEVOLUCIONES"),

    (["comex", "compra usd", "compra dolar"], "CAMBIO DIVISA"),
    ([" f29 ", "f29 ", " f29"], "IMPUESTOS"),  # cuidado: F29 puede aparecer

    (["bono nueces", "bono venta nueces"], "BONO VENTA NUECES"),

    (["miniexcabadora", "miniexcavadora"], "INVERSION ACTIVO PLANTA"),
    (["trueno"], "INVERSION ACTIVO PLANTA"),
    (["jorge bravo"], "MAQUINARIA - MANTENCION"),  # arriendo excavadora
    (["excabadora", "excavadora", "camion tolva"], "MAQUINARIA - MANTENCION"),

    (["luis quiroz"], "COMBUSTIBLE"),  # gas granel quemador
    (["gas granel", "gas a granel", "gas glp"], "COMBUSTIBLE"),
    (["quemador"], "COMBUSTIBLE"),

    (["control humedad"], "SERVICIOS PROFESIONALES"),
    (["frasol"], "SERVICIOS PROFESIONALES"),
    (["serv. rad", "serv rad"], "SERVICIOS PROFESIONALES"),

    (["compra cerezos"], "INVERSION / REPLANTE"),
    (["entrada campo"], "INVERSION / REPLANTE"),

    (["aminopower", "bioestimulante", "vals bio"], "INSUMOS AGRICOLAS"),
    (["hidalga"], "INSUMOS AGRICOLAS"),
    (["frutical emilia"], "INSUMOS AGRICOLAS"),
    (["fertilizante"], "INSUMOS AGRICOLAS"),
    (["herbicida"], "INSUMOS AGRICOLAS"),
    (["fitosanitario"], "INSUMOS AGRICOLAS"),

    (["remuneracion", "remuneración", "sueldo", "nomina", "nómina"], "MANO DE OBRA PLANTA"),

    (["copeval", "martinez", "valdivieso"], "INSUMOS AGRICOLAS"),
    (["cge"], "ENERGIA"),
    (["packman", "sortek"], "MAQUINARIA - MANTENCION"),
    (["cals"], "INSUMOS AGRICOLAS"),
    (["astara"], "INVERSION VEHICULOS"),
    (["alpabesa"], "MANO DE OBRA TEMPORAL"),
    (["arrayan", "aeromar"], "MANTENIMIENTO HELICOPTERO"),
    (["pago cuota", "leasing"], "LEASING"),
    (["smartways"], "PRESTAMOS A OTRAS SOCIEDADES"),
    (["rotortec", "ayv"], "PRESTAMOS A OTRAS SOCIEDADES"),
]


def match_keywords(text: str) -> str | None:
    t = text.lower()
    for kws, cat in KEYWORD_RULES:
        for kw in kws:
            if kw in t:
                return cat
    return None


# Mapeo de categorías existentes inconsistentes → formato estándar (UPPERCASE)
NORMALIZE_MAP = {
    "Fertilizantes": "FERTILIZANTES",
    "Fitosanitarios": "FITOSANITARIOS",
    "Maquinaria - mantencion": "MAQUINARIA - MANTENCION",
    "Mano de obra temporal": "MANO DE OBRA TEMPORAL",
    "Mano de obra planta": "MANO DE OBRA PLANTA",
    "Combustible": "COMBUSTIBLE",
    "Riego": "RIEGO",
    "Inversion / Replante": "INVERSION / REPLANTE",
    "Servicios profesionales": "SERVICIOS PROFESIONALES",
    "Arriendos / Patentes / Seguros": "ARRIENDOS / PATENTES / SEGUROS",
    "Caja chica / Imprevistos": "CAJA CHICA / IMPREVISTOS",
}


# ─── Cargar FXP ────────────────────────────────────────────────────
print("[1/4] Cargando FXP...")
tmp_f = os.path.join(tempfile.gettempdir(), "fxp_final.xlsx")
shutil.copy2(FXP_PATH, tmp_f)
wb_f = load_workbook(tmp_f, read_only=True, data_only=True)
ws_f = wb_f["ScotiaBCO"]
fxp_idx = {}
for row in ws_f.iter_rows(min_row=6, values_only=True):
    if not row or len(row) < 11: continue
    fecha = _parse_date(row[2])
    if not fecha: continue
    try:
        monto = int(round(float(row[6] or 0)))
    except: continue
    fxp_idx[(fecha.isoformat(), monto)] = (
        str(row[5] or ""), str(row[10] or "")
    )
wb_f.close()
print(f"   FXP indexado: {len(fxp_idx)}\n")

# ─── Procesar Master ────────────────────────────────────────────────
print("[2/4] Aplicando reglas a REVISAR post-2021...")
wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]

cat_counts = {}
actualizadas = 0
total_revisar = 0

for row in range(2, ws.max_row + 1):
    cat = ws.cell(row, 8).value
    if cat != "REVISAR": continue
    fecha = _parse_date(ws.cell(row, 1).value)
    if not fecha or fecha < date(2021, 1, 1): continue
    total_revisar += 1

    desc = str(ws.cell(row, 2).value or "")
    ref = str(ws.cell(row, 3).value or "")
    try:
        cargo = int(round(float(ws.cell(row, 4).value or 0)))
    except: cargo = 0

    fxp_desc, fxp_notas = fxp_idx.get((fecha.isoformat(), cargo), ("", ""))
    enriched = f"{desc} {ref} {fxp_desc} {fxp_notas}"

    nueva = match_keywords(enriched)
    if nueva:
        ws.cell(row, 8).value = nueva
        ws.cell(row, 9).value = "GENERAL"
        cat_counts[nueva] = cat_counts.get(nueva, 0) + 1
        actualizadas += 1

print(f"   Procesadas: {total_revisar}, Actualizadas: {actualizadas}\n")

# ─── Normalizar categorías existentes ──────────────────────────────
print("[3/4] Normalizando categorías a formato uniforme (MAYÚSCULAS)...")
normalize_counts = {}
for row in range(2, ws.max_row + 1):
    cat = ws.cell(row, 8).value
    if cat in NORMALIZE_MAP:
        nueva = NORMALIZE_MAP[cat]
        ws.cell(row, 8).value = nueva
        normalize_counts[f"{cat} → {nueva}"] = normalize_counts.get(f"{cat} → {nueva}", 0) + 1

print(f"   Normalizadas: {sum(normalize_counts.values())} filas")

# ─── Guardar ───────────────────────────────────────────────────────
print("\n[4/4] Guardando...")
wb.save(EXCEL_PATH)
wb.close()

print("\n=== NUEVAS CATEGORIAS APLICADAS ===")
for cat, n in sorted(cat_counts.items(), key=lambda x: -x[1]):
    print(f"  {cat}: {n}")

print("\n=== NORMALIZACIONES ===")
for change, n in sorted(normalize_counts.items(), key=lambda x: -x[1]):
    print(f"  {change}: {n}")

print("\nDONE!")
