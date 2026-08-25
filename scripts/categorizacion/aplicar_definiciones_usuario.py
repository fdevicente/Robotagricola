#!/usr/bin/env python3
"""Aplica las categorías que definió el usuario (2026-07-27) a Cuenta Banco.

- Inversiones San (77266652-7) + Gestora E → PRESTAMOS A OTRAS SOCIEDADES
  (devolución a Santa Elisa; NO es gasto operacional)
- Misael Henrique (15725130-9)  → INVERSION / REPLANTE (plantas de avellano)
- CRAVE SPA (77912665-K)        → MANO DE OBRA PLANTA (sueldo del dueño)
- Albino Fuenteal (10505528-5)  → SERVICIOS PROFESIONALES (contador)
- CARGO PAC VISA                → CAJA CHICA / IMPREVISTOS (gastos varios campo)
- Juan Parada $700.000          → CAJA CHICA / IMPREVISTOS (rendición; mismo
  patrón que marzo "RENDICION CAJA CHICA"). Inferido: avisar al usuario.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from collections import Counter
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH
from excel_manager import _save_wb

REGLAS = [
    (["77266652-7", "inversiones san", "gestora e"], "PRESTAMOS A OTRAS SOCIEDADES"),
    (["15725130-9", "misael henrique"],              "INVERSION / REPLANTE"),
    (["77912665-k", "crave spa"],                    "MANO DE OBRA PLANTA"),
    (["10505528-5", "albino fuenteal"],              "SERVICIOS PROFESIONALES"),
    (["cargo pac visa"],                             "CAJA CHICA / IMPREVISTOS"),
]


def _pd(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for f in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(v[:10], f).date()
            except Exception:
                pass
    return None


wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]
aplicadas = Counter()
detalle = []
quedan = []

for r in range(2, ws.max_row + 1):
    f = _pd(ws.cell(r, 1).value)
    if not f:
        continue
    cat = str(ws.cell(r, 8).value or "").strip()
    if cat and cat != "REVISAR":
        continue
    desc = str(ws.cell(r, 2).value or "")
    t = " " + desc.lower() + " "
    try:
        cargo = float(ws.cell(r, 4).value or 0)
        abono = float(ws.cell(r, 5).value or 0)
    except Exception:
        cargo = abono = 0

    nueva = None
    for kws, c in REGLAS:
        if any(k in t for k in kws):
            nueva = c
            break
    # Juan Parada $700.000 = rendición de caja chica (patrón conocido)
    if not nueva and "13373052-4" in t and abs(cargo - 700_000) < 1:
        nueva = "CAJA CHICA / IMPREVISTOS"

    if nueva:
        ws.cell(r, 8).value = nueva
        if not ws.cell(r, 9).value:
            ws.cell(r, 9).value = "GENERAL"
        aplicadas[nueva] += 1
        detalle.append((f, cargo or -abono, desc[:42], nueva))
    elif cat == "REVISAR":
        quedan.append((f, cargo or -abono, desc[:46]))

_save_wb(wb)
wb.close()

print("=== Categorías aplicadas ===")
for c, n in aplicadas.most_common():
    print(f"  {n:3} | {c}")
print("\n=== Detalle ===")
for f, monto, d, c in sorted(detalle, key=lambda x: (x[0], -abs(x[1]))):
    signo = "-" if monto > 0 else "+"
    print(f"  {f} {signo}${abs(monto):>12,.0f}  {d:42} → {c}")
print(f"\n=== Siguen en REVISAR: {len(quedan)} ===")
for f, monto, d in sorted(quedan, key=lambda x: -abs(x[1]))[:15]:
    signo = "-" if monto > 0 else "+"
    print(f"  {f} {signo}${abs(monto):>12,.0f}  {d}")
