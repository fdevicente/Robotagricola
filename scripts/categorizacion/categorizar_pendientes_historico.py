#!/usr/bin/env python3
"""Categorización best-guess de lo pendiente (histórico) en el Master.

- Cuenta Banco: filas en REVISAR o sin categoría -> categoría por keyword.
  (Se OMITE 'Gestora E': es préstamo, lo define el usuario.)
- Facturas: filas sin categoría -> categoría por proveedor/glosa.
- Lo que no matchea ningún keyword -> default MATERIALES (reportado).
Solo toca celdas vacías/REVISAR; nunca pisa categorías ya asignadas.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from collections import Counter
from openpyxl import load_workbook
from config import EXCEL_PATH

# Orden: específico -> genérico. Primer match gana.
KEYWORD_MAP = [
    (["remuneracion", "remuneración", "previred", "finiquito", "aguinaldo"], "MANO DE OBRA PLANTA"),
    (["f29", "impto", "impuesto", " sii", "tesoreria", "tgr", "patente comercial"], "IMPUESTOS"),
    (["reintegro", "devolucion", "devolución"], "REINTEGROS Y DEVOLUCIONES"),
    (["aduana", "exportac", "comex", "naviera", "embarque"], "SERVICIOS DE EXPORTACION"),
    (["irrifer", "irrifor", "wiseconn", "equipos de riego", "fertiriego", "indelec", "riego", "irrigat"], "RIEGO"),
    (["s-invest", "s invest", "sinvest", " cge", "enel", "energia generada", "energía generada"], "ENERGIA"),
    (["copec", "lipigas", "abastible", "gasco", "coragas", "estacion de servicio",
      "estacion servicio", "bencina", "petroleo", "petróleo", "combustible"], "COMBUSTIBLE"),
    (["fertiliz", "anagra", "fosfato", "urea", "nitrato", "salitre"], "FERTILIZANTES"),
    (["fitosanit", "fungicida", "herbicida", "pesticida", "abamectin", "abametin",
      "silitec", "ripper"], "FITOSANITARIOS"),
    (["copeval", "martinez y valdiv", "agrokimun", "cals f", "cooperativa agr",
      "vals bio", "anasac", "agrocampo", "aliagrototal", "facaz", "agromaq",
      "qlf labs", "qualified", "inveragro", "disal", "anagra"], "INSUMOS AGRICOLAS"),
    (["maqagri", "jl maquinaria", "promaq", "repuesto", "rodamendez", "rodamiento",
      "rodasep", "quemador", "implementos", "maquinaria", "central quemador",
      "importacion y venta de repuestos", "roda asesorias"], "MAQUINARIA - MANTENCION"),
    (["ferreteria", "ferretería", "sodimac", "casa del perno", "perno", "construmart",
      "easy", "imperial", "casa musa", "resorte gomez", "treck", "comercial flemec",
      "carvallo", "plastitec", "importadora santa teresita", "rodamendez",
      "ferreteria industrial pachita", "cementos bio", "cemento", "casa musa"], "MATERIALES"),
    (["bridgestone", "neumatico", "neumático", "vulcaniz", "automotriz",
      "revision tecnica", "permiso circulacion", "autopista"], "GASTOS VEHICULOS"),
    (["arriendo", "inmobiliaria", "yelcho", "seguro", "patente"], "ARRIENDOS / PATENTES / SEGUROS"),
    (["flete", "transporte", "cantu"], "TRANSPORTE"),
    (["caja chica", "gasto menor", "rendicion", "rendición", "mercado pago",
      "mercado libre", "mercadoli", "cargo pac", "redcompra", "ecommerce",
      "vuelo", "deposito vuelo"], "CAJA CHICA / IMPREVISTOS"),
    (["facava", "electricidad", "construccion", "construcción", "infraestructura"], "MANTENIMIENTO INFRAESTRUCTURA"),
    (["asesoria", "asesoría", "fgh", "auditor", "figueroa", "contador", "ingenieria",
      "ingeniería", "certificacion", "certificación", "universidad de talca",
      "laboratorio", "nogaltec", "nogalte", "leonardo", "agrologica", "agrológica",
      "consult", "bh ", "boleta honorario", "globaltecno", "intercom"], "SERVICIOS PROFESIONALES"),
    (["servicio", "serv ", "serv.", "cora servicios", "samuel vargas"], "SERVICIOS PROFESIONALES"),
]

DEFAULT = "MATERIALES"


def match(text):
    t = " " + text.lower() + " "
    for kws, cat in KEYWORD_MAP:
        for kw in kws:
            if kw in t:
                return cat
    return None


wb = load_workbook(EXCEL_PATH)

# ── Cuenta Banco ──
ws = wb["Cuenta Banco"]
banco_cnt = Counter()
banco_def = []
banco_skip_gestora = 0
for r in range(2, ws.max_row + 1):
    if not ws.cell(r, 1).value:
        continue
    cat = str(ws.cell(r, 8).value or "").strip()
    if cat and cat != "REVISAR":
        continue
    desc = str(ws.cell(r, 2).value or "")
    if "gestora e" in desc.lower():
        banco_skip_gestora += 1
        continue
    ref = str(ws.cell(r, 3).value or "")
    nueva = match(desc + " " + ref)
    if not nueva:
        nueva = DEFAULT
        banco_def.append(desc[:45])
    ws.cell(r, 8).value = nueva
    if not ws.cell(r, 9).value:
        ws.cell(r, 9).value = "GENERAL"
    banco_cnt[nueva] += 1

# ── Facturas ──
ws_f = wb["Facturas"]
fact_cnt = Counter()
fact_def = []
for r in range(2, ws_f.max_row + 1):
    if not ws_f.cell(r, 1).value:
        continue
    cat = str(ws_f.cell(r, 17).value or "").strip()
    if cat:
        continue
    prov = str(ws_f.cell(r, 4).value or "")
    glosa = str(ws_f.cell(r, 8).value or "") + " " + str(ws_f.cell(r, 9).value or "")
    nueva = match(prov + " " + glosa)
    if not nueva:
        nueva = DEFAULT
        fact_def.append(prov[:35])
    ws_f.cell(r, 17).value = nueva
    if not ws_f.cell(r, 20).value:
        ws_f.cell(r, 20).value = "auto-historico"
    fact_cnt[nueva] += 1

wb.save(EXCEL_PATH)
wb.close()

print("=== BANCO categorizado ===")
for c, n in banco_cnt.most_common():
    print(f"  {n:4} | {c}")
print(f"  (omitidos Gestora E -> REVISAR: {banco_skip_gestora})")

print("\n=== FACTURAS categorizadas ===")
for c, n in fact_cnt.most_common():
    print(f"  {n:4} | {c}")

print(f"\n=== Asignados por DEFAULT ({DEFAULT}) — revisar ===")
print(f"  Banco ({len(banco_def)}): " + ", ".join(banco_def[:15]))
dd = Counter(fact_def)
print(f"  Facturas ({len(fact_def)} líneas):")
for p, n in dd.most_common(20):
    print(f"     {n:3} | {p}")
