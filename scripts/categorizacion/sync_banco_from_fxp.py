#!/usr/bin/env python3
"""Busca Villa Las Delicias en FXP y reporta diferencias entre FXP banco y Master.CuentaBanco."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP_PATH = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"


def _parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try: return datetime.strptime(val[:10], fmt).date()
            except: pass
    return None


# Cargar FXP
print("[1/3] Buscando 'VILLA' o 'DELICIAS' en FXP ScotiaBCO...\n")
tmp = os.path.join(tempfile.gettempdir(), "fxp_villa.xlsx")
shutil.copy2(FXP_PATH, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["ScotiaBCO"]

villa_rows = []
for row in ws.iter_rows(min_row=6, values_only=True):
    if not row or len(row) < 11: continue
    desc = str(row[5] or "")
    notas = str(row[10] or "")
    text = (desc + " " + notas).lower()
    if "las delicias" in text or "vld " in text or "villa las" in text or " nueces" in text:
        fecha = _parse_date(row[2])
        try:
            m = float(row[6] or 0)
            d = float(row[7] or 0)
        except: m, d = 0, 0
        villa_rows.append({
            "fecha": fecha,
            "desc": desc,
            "monto": m,
            "deposito": d,
            "asig": str(row[9] or ""),
            "notas": notas,
        })

for v in villa_rows:
    print(f"  {v['fecha']} | cargo=${v['monto']:>12,.0f} abono=${v['deposito']:>12,.0f} | {v['desc'][:60]}")
    if v['notas'] and v['notas'].strip() != '-':
        print(f"    Nota: {v['notas']}")
    if v['asig']:
        print(f"    Asig: {v['asig']}")

# Contar diferencias FXP vs Master post-2024
print(f"\n[2/3] Contando movimientos FXP vs Master 2024+...\n")
fxp_count = 0
fxp_recent = 0
for row in ws.iter_rows(min_row=6, values_only=True):
    if not row or len(row) < 11: continue
    fecha = _parse_date(row[2])
    if not fecha: continue
    fxp_count += 1
    if fecha >= date(2024, 1, 1):
        fxp_recent += 1
wb.close()

print(f"   FXP total: {fxp_count}")
print(f"   FXP 2024+: {fxp_recent}")

# Master.CuentaBanco
print("\n[3/3] Master.CuentaBanco...\n")
tmp_m = os.path.join(tempfile.gettempdir(), "master_check.xlsx")
shutil.copy2(EXCEL_PATH, tmp_m)
wb_m = load_workbook(tmp_m, read_only=True, data_only=True)
ws_m = wb_m["Cuenta Banco"]
master_count = 0
master_recent = 0
master_max_date = None
for row in ws_m.iter_rows(min_row=2, max_col=2, values_only=True):
    fecha = _parse_date(row[0])
    if not fecha: continue
    master_count += 1
    if fecha >= date(2024, 1, 1):
        master_recent += 1
    if not master_max_date or fecha > master_max_date:
        master_max_date = fecha
wb_m.close()
print(f"   Master total: {master_count}")
print(f"   Master 2024+: {master_recent}")
print(f"   Master última fecha: {master_max_date}")
