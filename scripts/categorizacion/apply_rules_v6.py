#!/usr/bin/env python3
"""Reglas v6 con todas las confirmaciones."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP_PATH = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"


def _parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    return None


KEYWORD_RULES = [
    # SEGURIDAD (cámaras, switches, electrocerraduras)
    (["fjtech", "sp digital", "dacam", "jca supplies",
       "switch poe", "switch tplink", "tplink eap", "camara seguridad",
       "cámara seguridad", "camaras bullet", "cámaras bullet",
       "punto acceso", "electrocerradura", "porton campo", "portón campo"],
     "SEGURIDAD"),

    # MANTENCION PLANTA (conveyer, planta, etc.)
    (["cinta conveyer", "conveyer recogedora", "motor aceite conveyer",
       "instalacion electrica planta", "instalación eléctrica planta",
       "arreglo y instalacion electrica planta",
       "fabricacion parrilla", "fabricación parrilla",
       "fabricacion mesa y enganches", "fabricación mesa y enganches"],
     "MANTENCION PLANTA"),

    # MAQUINARIA - MANTENCION
    (["mantencion generador", "mantención generador", "marsof",
       "neumatico tractor", "neumático tractor", "patricio carvallo"],
     "MAQUINARIA - MANTENCION"),

    # GASTOS VEHICULOS
    (["parachoque camioneta", "parachoque camion",
       "jc services f17"], "GASTOS VEHICULOS"),

    # COMBUSTIBLE
    (["combustible lslg", "rivo", "comerc. y serv. rivo"], "COMBUSTIBLE"),
    (["coragas", "gas vma", "45kg"], "COMBUSTIBLE"),

    # INSUMOS AGRICOLAS
    (["tetrapack", "insumos albornoz"], "INSUMOS AGRICOLAS"),

    # ARRIENDOS / PATENTES / SEGUROS
    (["aspor poliza", "aspor f", "seguro campo", "pago seguro casa",
       "seguro casa"], "ARRIENDOS / PATENTES / SEGUROS"),
    (["bano quimico", "baño químico", "banos quimicos", "baños químicos",
       "reylux", "ambipar"], "ARRIENDOS / PATENTES / SEGUROS"),

    # SERVICIOS PROFESIONALES
    (["servicios juridicos", "servicios jurídicos", "luis machuca",
       "subdivision del campo", "subdivisión del campo"],
     "SERVICIOS PROFESIONALES"),

    # SERVICIOS DE EXPORTACION
    (["pago impto tgr", "tgr", "visita inspeccion fitosanitaria",
       "visita inspección fitosanitaria", "inspeccion fitosanitaria",
       "inspección fitosanitaria"], "SERVICIOS DE EXPORTACION"),
    (["compra regalo", "regalo espana", "regalo españa"],
     "SERVICIOS DE EXPORTACION"),

    # IMPUESTOS
    (["pago impt", "pago imp ", " imp sii"], "IMPUESTOS"),

    # MATERIALES
    (["taladro percusion", "taladro percusión", "baumart"], "MATERIALES"),
    (["cama saltarina", "sillon casa", "sillón casa", "material riego casa",
       "sercoriego"], "MATERIALES"),
    (["dreamtech", "impresora", "comerc. verazzi", "com. verazzi"],
     "MATERIALES"),

    # MANO DE OBRA TEMPORAL
    (["juan parada", "adelanto javier gonzalez", "adelanto"],
     "MANO DE OBRA TEMPORAL"),

    # GASTOS BANCARIOS (cargos administrativos del banco)
    (["reg rev de pago", "pago web", "comision banco", "comisión banco",
       "cargo administrativo"], "GASTOS BANCARIOS"),
]


def match_keywords(text: str) -> str | None:
    t = text.lower()
    for kws, cat in KEYWORD_RULES:
        for kw in kws:
            if kw in t:
                return cat
    return None


print("[1/3] FXP...")
tmp_f = os.path.join(tempfile.gettempdir(), "fxp_v6.xlsx")
shutil.copy2(FXP_PATH, tmp_f)
wb_f = load_workbook(tmp_f, read_only=True, data_only=True)
ws_f = wb_f["ScotiaBCO"]
fxp_idx = {}
for row in ws_f.iter_rows(min_row=6, values_only=True):
    if not row or len(row) < 11: continue
    fecha = _parse_date(row[2])
    if not fecha: continue
    try:
        monto = int(round(float(row[6] or 0)))
    except: continue
    fxp_idx[(fecha.isoformat(), monto)] = (
        str(row[5] or ""), str(row[10] or "")
    )
wb_f.close()
print(f"   {len(fxp_idx)}\n")

print("[2/3] Aplicando v6...")
wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]

cat_counts = {}
actualizadas = 0
total = 0

for row in range(2, ws.max_row + 1):
    cat = ws.cell(row, 8).value
    if cat != "REVISAR": continue
    fecha = _parse_date(ws.cell(row, 1).value)
    if not fecha or fecha < date(2021, 1, 1): continue
    total += 1

    desc = str(ws.cell(row, 2).value or "")
    ref = str(ws.cell(row, 3).value or "")
    try:
        cargo = int(round(float(ws.cell(row, 4).value or 0)))
    except: cargo = 0

    fxp_desc, fxp_notas = fxp_idx.get((fecha.isoformat(), cargo), ("", ""))
    enriched = f"{desc} {ref} {fxp_desc} {fxp_notas}"

    nueva = match_keywords(enriched)
    if nueva:
        ws.cell(row, 8).value = nueva
        ws.cell(row, 9).value = "GENERAL"
        cat_counts[nueva] = cat_counts.get(nueva, 0) + 1
        actualizadas += 1

print(f"   Procesados: {total}, Actualizados: {actualizadas}\n")

print("[3/3] Guardando...")
wb.save(EXCEL_PATH)
wb.close()

print("\n=== APLICADAS ===")
for cat, n in sorted(cat_counts.items(), key=lambda x: -x[1]):
    print(f"  {cat}: {n}")
