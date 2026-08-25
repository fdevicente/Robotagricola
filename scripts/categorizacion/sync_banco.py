#!/usr/bin/env python3
"""Sincroniza Master.CuentaBanco con FXP.ScotiaBCO.
Agrega movimientos faltantes (por fecha+monto+desc únicos)."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP_PATH = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"


def _parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try: return datetime.strptime(val[:10], fmt).date()
            except: pass
    return None


# Cargar Master existente
print("[1/4] Cargando Master.CuentaBanco existente...")
wb_m = load_workbook(EXCEL_PATH)
ws_m = wb_m["Cuenta Banco"]

# Indexar Master por (fecha_iso, cargo_int, abono_int)
master_idx = set()
master_max_date = None
master_last_row = 1
for r in range(2, ws_m.max_row + 1):
    fecha = _parse_date(ws_m.cell(r, 1).value)
    if not fecha: continue
    master_last_row = r
    try:
        cargo = int(round(float(ws_m.cell(r, 4).value or 0)))
        abono = int(round(float(ws_m.cell(r, 5).value or 0)))
    except: cargo, abono = 0, 0
    master_idx.add((fecha.isoformat(), cargo, abono))
    if not master_max_date or fecha > master_max_date:
        master_max_date = fecha

print(f"   Master: {len(master_idx)} movimientos, última fecha: {master_max_date}\n")

# Cargar FXP
print("[2/4] Cargando FXP.ScotiaBCO...")
tmp = os.path.join(tempfile.gettempdir(), "fxp_sync.xlsx")
shutil.copy2(FXP_PATH, tmp)
wb_f = load_workbook(tmp, read_only=True, data_only=True)
ws_f = wb_f["ScotiaBCO"]

# Headers fila 5: Emisión, Vencimiento, Pago, Año, N, Descripción, Monto, Deposito, Saldo, Asig Cta, Notas
nuevos = []
for row in ws_f.iter_rows(min_row=6, values_only=True):
    if not row or len(row) < 11: continue
    fecha = _parse_date(row[2])  # col 3 = Pago
    if not fecha: continue
    try:
        monto = int(round(float(row[6] or 0)))      # cargo
        deposito = int(round(float(row[7] or 0)))   # abono
    except: continue
    desc = str(row[5] or "")
    asig = str(row[9] or "")
    notas = str(row[10] or "")
    saldo = row[8]

    key = (fecha.isoformat(), monto, deposito)
    if key in master_idx:
        continue
    nuevos.append({
        "fecha": fecha, "desc": desc,
        "cargo": monto, "abono": deposito,
        "saldo": saldo, "asig": asig, "notas": notas,
    })

wb_f.close()
print(f"   Nuevos a agregar: {len(nuevos)}\n")

if nuevos:
    print("[3/4] Agregando al Master...")
    # Ordenar por fecha
    nuevos.sort(key=lambda x: x["fecha"])

    # Columnas Cuenta Banco: A=fecha, B=descripcion, C=referencia, D=cargo, E=abono, F=saldo, G=tipo, H=categoria, I=cultivo
    start = master_last_row + 1
    for i, n in enumerate(nuevos):
        r = start + i
        ws_m.cell(r, 1).value = n["fecha"]
        ws_m.cell(r, 2).value = n["desc"]
        ws_m.cell(r, 3).value = ""  # referencia
        ws_m.cell(r, 4).value = n["cargo"] if n["cargo"] > 0 else None
        ws_m.cell(r, 5).value = n["abono"] if n["abono"] > 0 else None
        ws_m.cell(r, 6).value = n["saldo"] if isinstance(n["saldo"], (int, float)) else None
        # G=tipo, H=categoria, I=cultivo se dejan vacíos (los va a categorizar el bot)

    print(f"   Agregadas {len(nuevos)} filas (desde {nuevos[0]['fecha']} hasta {nuevos[-1]['fecha']})\n")

    print("[4/4] Guardando Master...")
    wb_m.save(EXCEL_PATH)
    print("Done!\n")

    # Resumen post-2026-04
    print("Nuevos post-2026-04:")
    for n in nuevos:
        if n["fecha"] >= date(2026, 4, 1):
            tipo = "abono" if n["abono"] > 0 else "cargo"
            monto = n["abono"] if n["abono"] > 0 else n["cargo"]
            print(f"  {n['fecha']} | {tipo}=${monto:>12,.0f} | {n['desc'][:60]}")
else:
    print("[3/4] Sin novedades. Master ya está al día.\n")

wb_m.close()
