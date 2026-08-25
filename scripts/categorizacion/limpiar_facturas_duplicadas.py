#!/usr/bin/env python3
"""Elimina líneas duplicadas en Master.Facturas.

A) Duplicados de la carga de hoy: filas explícitas ya verificadas.
B) Duplicados históricos: facturas donde el Master quedó con el monto duplicado
   respecto de FXP. Se buscan los bloques de filas contiguas cuya suma calce
   con el monto de FXP; el resto se elimina y se corrige TOTAL FACTURA.
   Si no hay una solución inequívoca, la factura se OMITE (no se toca).
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
import os, re, shutil, tempfile, unicodedata
from collections import defaultdict
from datetime import date, datetime
from itertools import combinations
from openpyxl import load_workbook
from config import EXCEL_PATH
from excel_manager import _save_wb

FXP = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"

# A) Filas duplicadas de hoy (verificadas una a una)
FILAS_HOY = [2113, 2114, 2125, 2126, 2127, 2128, 2129, 2146]


def norm(s):
    s = str(s or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = re.sub(r"\b(spa|ltda|s\.?a\.?|eirl|limitada|y cia|cia)\b", " ", s)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def nrokey(v):
    s = str(v or "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.upper().replace(" ", "").replace("-", "")


def _f(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


# ── FXP como referencia de monto ──
tmp = os.path.join(tempfile.gettempdir(), "fxp_limpieza.xlsx")
shutil.copy2(FXP, tmp)
wbf = load_workbook(tmp, read_only=True, data_only=True)
fxp = {}
for row in wbf["FXP"].iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    nro = nrokey(row[7])
    if not nro:
        continue
    m = row[8]
    if isinstance(m, str):
        s = m.upper().replace("USD", "").replace("$", "").strip().replace(".", "").replace(",", ".")
        try:
            m = float(s)
        except ValueError:
            m = 0
    fxp[nro] = _f(m)
wbf.close()

# ── Master ──
wb = load_workbook(EXCEL_PATH)
ws = wb["Facturas"]
grupos = defaultdict(list)
for r in range(2, ws.max_row + 1):
    if not ws.cell(r, 1).value:
        continue
    nro = nrokey(ws.cell(r, 7).value)
    if not nro:
        continue
    grupos[(norm(ws.cell(r, 4).value), nro)].append({
        "fila": r, "item": _f(ws.cell(r, 15).value),
        "total": _f(ws.cell(r, 16).value),
        "prov": str(ws.cell(r, 4).value or ""), "nro": str(ws.cell(r, 7).value or ""),
    })

# ── B) Plan para históricos ──
plan_borrar, plan_total, omitidas = [], [], []
for (prov, nro), filas in grupos.items():
    if any(f["fila"] in FILAS_HOY for f in filas):
        continue                      # ya cubierto por la lista A
    ref = fxp.get(nro)
    if not ref or ref <= 0:
        continue
    total = max(f["total"] for f in filas)
    if total <= ref * 1.5:            # no está duplicado
        continue

    # bloques de filas contiguas
    nums = sorted(f["fila"] for f in filas)
    bloques, ini, ant = [], nums[0], nums[0]
    for n in nums[1:]:
        if n == ant + 1:
            ant = n
        else:
            bloques.append((ini, ant)); ini = ant = n
    bloques.append((ini, ant))
    if len(bloques) < 2:
        omitidas.append((prov, nro, total, ref, "un solo bloque"))
        continue

    por_bloque = []
    for (a, b) in bloques:
        sub = [f for f in filas if a <= f["fila"] <= b]
        por_bloque.append({"rango": (a, b), "suma": sum(x["item"] for x in sub),
                            "filas": [x["fila"] for x in sub]})

    # combinación de bloques cuya suma calce con FXP
    tol = max(1000, ref * 0.01)
    sol = None
    for k in range(1, len(por_bloque)):
        for combo in combinations(range(len(por_bloque)), k):
            s = sum(por_bloque[i]["suma"] for i in combo)
            if abs(s - ref) <= tol:
                sol = combo
                break
        if sol:
            break
    if not sol:
        omitidas.append((prov, nro, total, ref, "no calza ninguna combinación"))
        continue

    quedan = [i for i in sol]
    fuera = [i for i in range(len(por_bloque)) if i not in quedan]
    filas_fuera = [f for i in fuera for f in por_bloque[i]["filas"]]
    filas_quedan = [f for i in quedan for f in por_bloque[i]["filas"]]
    plan_borrar.extend(filas_fuera)
    plan_total.append((filas_quedan, ref))
    print(f"  {prov[:30]:30} F{nro:<11} master ${total:>12,.0f} → ${ref:>12,.0f}  "
          f"borra {len(filas_fuera)} línea(s) {filas_fuera}")

print(f"\n--- Resumen del plan ---")
print(f"A) Duplicados de hoy: {len(FILAS_HOY)} líneas")
print(f"B) Duplicados históricos: {len(plan_total)} facturas, {len(plan_borrar)} líneas")
print(f"   Omitidas por ambigüedad (NO se tocan): {len(omitidas)}")
for prov, nro, total, ref, motivo in omitidas[:12]:
    print(f"     {prov[:28]:28} F{nro:<11} ${total:>11,.0f} vs FXP ${ref:>11,.0f}  ({motivo})")
if len(omitidas) > 12:
    print(f"     … y {len(omitidas)-12} más")

# ── Aplicar ──
# 1) corregir TOTAL FACTURA en las filas que quedan
for filas_quedan, ref in plan_total:
    for r in filas_quedan:
        ws.cell(r, 16).value = round(ref)
# 2) borrar filas (de mayor a menor para no correr índices)
todas = sorted(set(FILAS_HOY) | set(plan_borrar), reverse=True)
for r in todas:
    ws.delete_rows(r)

_save_wb(wb)
wb.close()
print(f"\n✅ Eliminadas {len(todas)} líneas duplicadas en total.")
print(f"✅ Corregido TOTAL FACTURA en {len(plan_total)} facturas históricas.")
