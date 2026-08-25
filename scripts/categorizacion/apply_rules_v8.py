#!/usr/bin/env python3
"""Reglas v8 - últimos patrones simples (bencina, ferretería, gas)."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP_PATH = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"


def _parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    return None


KEYWORD_RULES = [
    # COMBUSTIBLE (bencina G93/G95)
    (["lts de g93", "lts de g95", "lts g93", "lts g95",
       "g93 ", "g95 ", "litros g93", "litros g95",
       "bencina 93", "bencina 95", "facaz", "adm vtal",
       "adm. de ventas al detalle", "adm de ventas al dte",
       "ventas al detalle"], "COMBUSTIBLE"),
    (["gas casa", "gas balones", "balones gas"], "COMBUSTIBLE"),

    # SEGURIDAD
    (["swicht dahua", "switch dahua", "transformadores camaras",
       "transformadores cámaras", "pcplay", "electrotecnologia",
       "electrotecnología"], "SEGURIDAD"),

    # SERVICIOS PROFESIONALES (telefonía como WOM)
    (["movistar", "entel", "claro pago", "telefonia",
       "telefonía"], "SERVICIOS PROFESIONALES"),

    # MAQUINARIA - MANTENCION
    (["mantencion motocierra", "mantención motocierra",
       "motosierra", "motocierra"], "MAQUINARIA - MANTENCION"),
    (["rodamientos", "rodacar"], "MAQUINARIA - MANTENCION"),
    (["arreglo maquina cosechadora", "arreglo máquina cosechadora",
       "maquina cosechadora", "máquina cosechadora"],
     "MAQUINARIA - MANTENCION"),

    # MATERIALES (ferretería, herramientas, EPP, oficina casa, materiales varios)
    (["mascarilla", "buzo", "guantes", "epp"], "MATERIALES"),
    (["pernos", "tuercas", "colillas", "colilla", "tuerca",
       "adrian barrios", "adrián barrios", "resortes gomez",
       "resortes gómez"], "MATERIALES"),
    (["pala hoyera", "pala hoyer", "pedro alberto rojo"],
     "MATERIALES"),
    (["cloro piscina", "comercializadora yukon"], "MATERIALES"),
    (["pita y vinilit", "pita", "vinilit",
       "ferreteria pachita", "ferretería pachita",
       "ferreteria", "ferretería"], "MATERIALES"),
    (["luces led casa", "studio group", "luces led",
       "enchufe, conector", "harry plotter", "mapa campo",
       "panchita", "electricidad talca",
       "comercial electricidad talca"], "MATERIALES"),
    (["cargo pac visa", "pac visa", "pago automatico tc",
       "pago automático tc"], "MATERIALES"),

    # INSUMOS AGRICOLAS
    (["semillas huerta", "plusagro", "aliagrototal",
       "aliagrotótal"], "INSUMOS AGRICOLAS"),

    # ALIMENTACION Y ALOJAMIENTO
    (["alimento gato", "alimento perro", "alimento mascota",
       "verduras y bebidas asado", "verduras y bebidas",
       "jumbo f", "la vega chica"], "ALIMENTACION Y ALOJAMIENTO"),

    # REINTEGROS Y DEVOLUCIONES
    (["compra juan", "error j. parada", "error j parada"],
     "REINTEGROS Y DEVOLUCIONES"),

    # MANTENCION PLANTA - additional
    (["arreglo cosechadora", "cosechadora"], "MAQUINARIA - MANTENCION"),
]


def match_keywords(text: str) -> str | None:
    t = text.lower()
    for kws, cat in KEYWORD_RULES:
        for kw in kws:
            if kw in t:
                return cat
    return None


print("[1/3] FXP...")
tmp_f = os.path.join(tempfile.gettempdir(), "fxp_v8.xlsx")
shutil.copy2(FXP_PATH, tmp_f)
wb_f = load_workbook(tmp_f, read_only=True, data_only=True)
ws_f = wb_f["ScotiaBCO"]
fxp_idx = {}
for row in ws_f.iter_rows(min_row=6, values_only=True):
    if not row or len(row) < 11: continue
    fecha = _parse_date(row[2])
    if not fecha: continue
    try:
        monto = int(round(float(row[6] or 0)))
    except: continue
    fxp_idx[(fecha.isoformat(), monto)] = (
        str(row[5] or ""), str(row[10] or "")
    )
wb_f.close()
print(f"   {len(fxp_idx)}\n")

print("[2/3] Aplicando v8...")
wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]

cat_counts = {}
actualizadas = 0
total = 0

for row in range(2, ws.max_row + 1):
    cat = ws.cell(row, 8).value
    if cat != "REVISAR": continue
    fecha = _parse_date(ws.cell(row, 1).value)
    if not fecha or fecha < date(2021, 1, 1): continue
    total += 1

    desc = str(ws.cell(row, 2).value or "")
    ref = str(ws.cell(row, 3).value or "")
    try:
        cargo = int(round(float(ws.cell(row, 4).value or 0)))
    except: cargo = 0

    fxp_desc, fxp_notas = fxp_idx.get((fecha.isoformat(), cargo), ("", ""))
    enriched = f"{desc} {ref} {fxp_desc} {fxp_notas}"

    nueva = match_keywords(enriched)
    if nueva:
        ws.cell(row, 8).value = nueva
        ws.cell(row, 9).value = "GENERAL"
        cat_counts[nueva] = cat_counts.get(nueva, 0) + 1
        actualizadas += 1

print(f"   Procesados: {total}, Actualizados: {actualizadas}\n")

print("[3/3] Guardando...")
wb.save(EXCEL_PATH)
wb.close()

print("\n=== APLICADAS ===")
for cat, n in sorted(cat_counts.items(), key=lambda x: -x[1]):
    print(f"  {cat}: {n}")
