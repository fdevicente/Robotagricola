#!/usr/bin/env python3
"""Reglas v5 con confirmaciones del usuario."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP_PATH = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"


def _parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    return None


KEYWORD_RULES = [
    # MANTENCION PLANTA (alta prioridad antes que MAQUINARIA)
    (["variadores de frecuencia", "variadores frecuencia",
       "balanza de piso para planta", "balanza planta",
       "filtro elevacion agua planta", "filtro elevación agua planta",
       "instalación filtro", "instalacion filtro"], "MANTENCION PLANTA"),

    # RIEGO
    (["wiseconn", "indelec", "mantencion bombas", "mantención bombas",
       "bombas riego", "arreglo bomba", "bomba de agua",
       "hugo roco gajardo", "ana maria sanchez"], "RIEGO"),

    # INVERSION / REPLANTE
    (["adelanto compra avellanos", "diferencia reserva del 30%",
       "agricola el huingan", "agrícola el huingan"], "INVERSION / REPLANTE"),

    # INSUMOS AGRICOLAS
    (["calcio", "magnesio", "potasio", "cna f"], "INSUMOS AGRICOLAS"),
    (["fardo", "fardos"], "INSUMOS AGRICOLAS"),
    (["trabajos cal", "rene hernan meneses", "rene meneses",
       "rené hernán meneses"], "INSUMOS AGRICOLAS"),
    (["silitec", "nasjac"], "INSUMOS AGRICOLAS"),

    # SERVICIOS DE EXPORTACION
    (["parrila fdv", "parrilla fdv", "parrillas fdv",
       "parrila para espana", "parrillas para espana",
       "para espana", "para españa"], "SERVICIOS DE EXPORTACION"),
    (["hotel espanoles", "hotel españoles", "pullman vitacura",
       "examenes labs", "exámenes labs", "qflabs",
       "labs nueces"], "SERVICIOS DE EXPORTACION"),

    # ALIMENTACION Y ALOJAMIENTO
    (["pension juan parada", "pensión juan parada", "platillo volador",
       "alojamiento"], "ALIMENTACION Y ALOJAMIENTO"),

    # TRANSPORTE
    (["gastos de trans. op", "gastos trans op", "gasto de transporte",
       "gasto trans op"], "TRANSPORTE"),

    # GASTOS VEHICULOS
    (["permiso de circulacion", "permiso de circulación", "soap",
       "neumaticos camioneta", "neumáticos camioneta",
       "mantenimiento camioneta", "mantenimiento camion",
       "aventura motors", "mantencion camioneta",
       "mantención camioneta"], "GASTOS VEHICULOS"),

    # COMBUSTIBLE
    (["combustible y movilizacion", "combustible y movilización"], "COMBUSTIBLE"),

    # MATERIALES
    (["starlink"], "MATERIALES"),
    (["ferreteria myg", "ferreteria m y g", "ferretería myg",
       "ferreteria marsella", "ferretería marsella",
       "materiales campo", "soplador espalda", "soplador a espalda",
       "malla, llaves", "clavos", "cemento", "arena y gravilla",
       "arena", "gravilla", "base para cemento", "aridos"], "MATERIALES"),

    # MAQUINARIA - MANTENCION
    (["agroequipos", "agromaq", "vicas spa", "hidrolavadora",
       "martillo trituradora", "martillo triturador"],
     "MAQUINARIA - MANTENCION"),

    # SERVICIOS PROFESIONALES
    (["analisis de suelo", "análisis de suelo", "juan hirzel"],
     "SERVICIOS PROFESIONALES"),
    (["arreglo judicial", "abogado"], "SERVICIOS PROFESIONALES"),

    # IMPUESTOS
    (["pago impto sii", "pago imp sii", "contribuciones",
       "impto contribuciones"], "IMPUESTOS"),

    # TRANSFERENCIA INTERNA
    (["traspaso cuenta bco chile", "traspaso cuenta banco chile",
       "agricola santa elisa bco chile", "agrícola santa elisa bco chile"],
     "TRANSFERENCIA INTERNA"),

    # ─── Reglas ya aplicadas (para nuevos cargos) ───
    (["traspaso banco", "traspaso bcos", "traspaso de banco"],
     "TRANSFERENCIA INTERNA"),
    (["ecosmart"], "SERVICIOS DE EXPORTACION"),
    (["bh ", " bh", "boleta honorarios"], "MANO DE OBRA TEMPORAL"),
    (["jornal", "jornalero"], "MANO DE OBRA TEMPORAL"),
    (["agrocampo", "felipe noguera", "ricardo mahla", "kohen"],
     "MAQUINARIA - MANTENCION"),
    (["repuestos tractor", "filtros tractor", "filtro tractor"],
     "MAQUINARIA - MANTENCION"),
    (["pago total tag", " tag ", "pago tag"], "GASTOS VEHICULOS"),
    (["pago diferencias facturas"], "REINTEGROS Y DEVOLUCIONES"),
    (["miguel marin", "miguel marín", "piloto"], "SERVICIOS PROFESIONALES"),
    (["capital office"], "SERVICIOS PROFESIONALES"),
    (["rendicion viajes", "rendición viajes"], "SERVICIOS PROFESIONALES"),
    (["francisco donoso"], "SERVICIOS PROFESIONALES"),
]


def match_keywords(text: str) -> str | None:
    t = text.lower()
    for kws, cat in KEYWORD_RULES:
        for kw in kws:
            if kw in t:
                return cat
    return None


print("[1/3] FXP...")
tmp_f = os.path.join(tempfile.gettempdir(), "fxp_v5.xlsx")
shutil.copy2(FXP_PATH, tmp_f)
wb_f = load_workbook(tmp_f, read_only=True, data_only=True)
ws_f = wb_f["ScotiaBCO"]
fxp_idx = {}
for row in ws_f.iter_rows(min_row=6, values_only=True):
    if not row or len(row) < 11: continue
    fecha = _parse_date(row[2])
    if not fecha: continue
    try:
        monto = int(round(float(row[6] or 0)))
    except: continue
    fxp_idx[(fecha.isoformat(), monto)] = (
        str(row[5] or ""), str(row[10] or "")
    )
wb_f.close()
print(f"   {len(fxp_idx)}\n")

print("[2/3] Aplicando v5...")
wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]

cat_counts = {}
actualizadas = 0
total = 0

for row in range(2, ws.max_row + 1):
    cat = ws.cell(row, 8).value
    if cat != "REVISAR": continue
    fecha = _parse_date(ws.cell(row, 1).value)
    if not fecha or fecha < date(2021, 1, 1): continue
    total += 1

    desc = str(ws.cell(row, 2).value or "")
    ref = str(ws.cell(row, 3).value or "")
    try:
        cargo = int(round(float(ws.cell(row, 4).value or 0)))
    except: cargo = 0

    fxp_desc, fxp_notas = fxp_idx.get((fecha.isoformat(), cargo), ("", ""))
    enriched = f"{desc} {ref} {fxp_desc} {fxp_notas}"

    nueva = match_keywords(enriched)
    if nueva:
        ws.cell(row, 8).value = nueva
        ws.cell(row, 9).value = "GENERAL"
        cat_counts[nueva] = cat_counts.get(nueva, 0) + 1
        actualizadas += 1

print(f"   Procesados: {total}, Actualizados: {actualizadas}\n")

print("[3/3] Guardando...")
wb.save(EXCEL_PATH)
wb.close()

print("\n=== APLICADAS ===")
for cat, n in sorted(cat_counts.items(), key=lambda x: -x[1]):
    print(f"  {cat}: {n}")
