#!/usr/bin/env python3
from openpyxl import load_workbook
from config import EXCEL_PATH

def get_category(proveedor, detalle):
    if not proveedor: proveedor = ''
    if not detalle: detalle = ''

    p = (proveedor or '').upper()
    d = (detalle or '').upper()

    # Known cases from user input
    if 'ECOSMART' in p: return 'SERVICIOS DE EXPORTACION'
    if 'S-INVEST' in p or 'S.INVEST' in p: return 'COSTO ENERGETICO'
    if 'SCOTIABANK' in p: return 'CAMBIO DIVISA'
    if 'LUIS QUIROZ' in p or 'LUIS ALBERTO' in p: return 'TRANSPORTE'

    # Energy
    if 'CGE' in p or 'ELECTRICIDAD' in p: return 'ENERGIA'
    if 'AGUA' in p or 'RIEGO' in p: return 'AGUA'

    # Transport
    if 'TRANSPORTE' in p or 'FLETE' in p or 'DHL' in p or 'LATAM' in p: return 'TRANSPORTE'

    # Agriculture
    if 'AGROKIMUN' in p or 'FERTILI' in p or 'ABONO' in p: return 'INSUMOS AGRICOLAS'

    # Hardware
    if 'FERRETERIA' in p or 'AGROEQUIPOS' in p: return 'HERRAMIENTAS'
    if 'ARIDOS' in p or 'GRAVA' in p or 'ARENA' in p: return 'MATERIALES'

    # Office
    if 'LIBRERIA' in p or 'TUCAN' in p or 'MERCADOLIBRE' in p: return 'SUMINISTROS'

    # Default
    return 'SERVICIOS'

print('Opening file with openpyxl...')
wb = load_workbook(EXCEL_PATH)
ws = wb['Facturas']

update_count = 0
for row in range(2, 2000):
    cat = ws.cell(row, 17).value
    if cat == 'REVISAR':
        prov = ws.cell(row, 4).value
        detail = ws.cell(row, 8).value
        new_cat = get_category(prov, detail)
        ws.cell(row, 17).value = new_cat
        update_count += 1

        if row % 200 == 0:
            print(f'  Row {row}...')

print(f'Updated {update_count} invoices')
print('Saving...')
wb.save(EXCEL_PATH)
wb.close()
print('Done!')
