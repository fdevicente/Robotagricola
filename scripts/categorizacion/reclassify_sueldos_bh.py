#!/usr/bin/env python3
"""Reclasifica pagos de sueldos y BH:
- 'Remuneracion <mes> <Persona Fija>' → MANO DE OBRA PLANTA
- 'BH Francisco Donoso' → SERVICIOS PROFESIONALES
- 'BH <jornalero>' → MANO DE OBRA TEMPORAL
"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from config import EXCEL_PATH

# Personal fijo (sueldos mensuales)
PERSONAL_FIJO = {
    "felix de vicente", "juan parada", "felicito amigo",
    "agustin mora", "patricio mora", "ramiro amigo",
    "patricio amigo",
}

# Asesores/profesionales (BH recurrente, pero servicio profesional)
SERVICIOS_PROFESIONALES_BH = {
    "francisco donoso",  # agrónomo
}


def detectar_tipo(desc):
    """Devuelve la categoría correcta o None si no aplica."""
    d = desc.lower().strip()

    # Remuneracion mensual
    m = re.search(r"remuneraci[oó]n\s+\w+\s+(.+)", d)
    if m:
        persona = m.group(1).strip()
        # Verificar si está en personal fijo
        for fijo in PERSONAL_FIJO:
            if fijo in persona:
                return "MANO DE OBRA PLANTA"
        return "MANO DE OBRA PLANTA"  # default: cualquier remuneración es fijo

    # Boletas Honorarios - "BH XXX Persona" o "BH Persona"
    m = re.search(r"\bbh\s*\d*\s+(.+)", d)
    if not m:
        m = re.search(r"^bh\s+(.+)", d)
    if m:
        persona = m.group(1).strip()
        for prof in SERVICIOS_PROFESIONALES_BH:
            if prof in persona:
                return "SERVICIOS PROFESIONALES"
        # Personal fijo via BH (raro pero posible)
        for fijo in PERSONAL_FIJO:
            if fijo in persona:
                return "MANO DE OBRA PLANTA"
        return "MANO DE OBRA TEMPORAL"

    # Previred / aguinaldos / cotizaciones → fijo planta
    if "previred" in d or "aguinaldo" in d:
        return "MANO DE OBRA PLANTA"

    return None


wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]

stats = {
    "MANO DE OBRA PLANTA": 0,
    "MANO DE OBRA TEMPORAL": 0,
    "SERVICIOS PROFESIONALES": 0,
}
sin_cambio = 0

for r in range(2, ws.max_row + 1):
    if not ws.cell(r, 1).value: continue
    desc = str(ws.cell(r, 2).value or "")
    cat_actual = str(ws.cell(r, 8).value or "").strip().upper()

    nueva = detectar_tipo(desc)
    if not nueva: continue

    if cat_actual == nueva:
        sin_cambio += 1
        continue

    ws.cell(r, 8).value = nueva
    if not ws.cell(r, 9).value:
        ws.cell(r, 9).value = "GENERAL"
    stats[nueva] += 1

print(f"Reclasificadas:")
for cat, n in stats.items():
    print(f"  {cat}: {n}")
print(f"Sin cambio (ya estaban OK): {sin_cambio}")

wb.save(EXCEL_PATH)
wb.close()
print("Done!")
