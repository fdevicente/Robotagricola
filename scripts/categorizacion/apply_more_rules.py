#!/usr/bin/env python3
"""Reglas adicionales del usuario."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP_PATH = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"


def _parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    return None


# Nuevas reglas
KEYWORD_RULES = [
    (["traspaso de banco", "traspaso entre banco", "traspaso bcos", "traspaso bancario"],
     "TRANSFERENCIA INTERNA"),
    (["transferencia santa elisa scotia a santa", "transferencia santa elisa scotia a santander"],
     "TRANSFERENCIA INTERNA"),

    (["s-invest", "s invest", "sinvest"], "ENERGIA"),

    (["agencia aduana", "agencia aduanal", "aduana"], "SERVICIOS DE EXPORTACION"),
    (["intralog"], "SERVICIOS DE EXPORTACION"),
    (["carmen gloria martinez", "carmengloria martinez", "carmen gloria"],
     "SERVICIOS DE EXPORTACION"),
    (["maxisaco", "maxi saco"], "SERVICIOS DE EXPORTACION"),

    (["francisco donoso"], "SERVICIOS PROFESIONALES"),
]


def match_keywords(text: str) -> str | None:
    t = text.lower()
    for kws, cat in KEYWORD_RULES:
        for kw in kws:
            if kw in t:
                return cat
    return None


# Cargar FXP
print("[1/3] Cargando FXP...")
tmp_f = os.path.join(tempfile.gettempdir(), "fxp_more.xlsx")
shutil.copy2(FXP_PATH, tmp_f)
wb_f = load_workbook(tmp_f, read_only=True, data_only=True)
ws_f = wb_f["ScotiaBCO"]
fxp_idx = {}
for row in ws_f.iter_rows(min_row=6, values_only=True):
    if not row or len(row) < 11: continue
    fecha = _parse_date(row[2])
    if not fecha: continue
    try:
        monto = int(round(float(row[6] or 0)))
    except: continue
    fxp_idx[(fecha.isoformat(), monto)] = (
        str(row[5] or ""), str(row[10] or "")
    )
wb_f.close()
print(f"   FXP: {len(fxp_idx)}\n")

# Aplicar reglas
print("[2/3] Aplicando reglas a REVISAR post-2021...")
wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]

cat_counts = {}
actualizadas = 0
total = 0

for row in range(2, ws.max_row + 1):
    cat = ws.cell(row, 8).value
    if cat != "REVISAR": continue
    fecha = _parse_date(ws.cell(row, 1).value)
    if not fecha or fecha < date(2021, 1, 1): continue
    total += 1

    desc = str(ws.cell(row, 2).value or "")
    ref = str(ws.cell(row, 3).value or "")
    try:
        cargo = int(round(float(ws.cell(row, 4).value or 0)))
    except: cargo = 0

    fxp_desc, fxp_notas = fxp_idx.get((fecha.isoformat(), cargo), ("", ""))
    enriched = f"{desc} {ref} {fxp_desc} {fxp_notas}"

    nueva = match_keywords(enriched)
    if nueva:
        ws.cell(row, 8).value = nueva
        ws.cell(row, 9).value = "GENERAL"
        cat_counts[nueva] = cat_counts.get(nueva, 0) + 1
        actualizadas += 1

print(f"   Procesados: {total}, Actualizados: {actualizadas}\n")

print("[3/3] Guardando...")
wb.save(EXCEL_PATH)
wb.close()

print("\n=== APLICADAS ===")
for cat, n in sorted(cat_counts.items(), key=lambda x: -x[1]):
    print(f"  {cat}: {n}")
