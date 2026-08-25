#!/usr/bin/env python3
"""Marca las facturas con estado NN (no se van a pagar) en Master.
Usa col 20 (Categorizado_por) con valor 'NN-no-pagar'."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP_PATH = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"


def norm(s):
    return (s or "").strip().upper().replace(".", "").replace("  ", " ")


# Cargar FXP - identificar NN
print("[1/3] Identificando facturas NN en FXP...")
tmp = os.path.join(tempfile.gettempdir(), "fxp_nn.xlsx")
shutil.copy2(FXP_PATH, tmp)
wb = load_workbook(tmp, read_only=True, data_only=True)
ws = wb["FXP"]

nn_keys = set()
nn_list = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    estado = str(row[11] or "").strip().upper()
    if estado != "NN": continue
    prov = str(row[6] or "")
    nro = str(row[7] or "").strip()
    if not prov or not nro: continue
    nn_keys.add((norm(prov), nro))
    nn_list.append((prov, nro, row[8]))
wb.close()
print(f"   FXP NN: {len(nn_keys)} facturas\n")

# Cargar Master, marcar NN
print("[2/3] Marcando en Master...")
wb_m = load_workbook(EXCEL_PATH)
ws_m = wb_m["Facturas"]

marcadas = 0
ya_marcadas = 0
no_encontradas = 0
nn_master = {(k[0], k[1]): False for k in nn_keys}

for r in range(2, ws_m.max_row + 1):
    if not ws_m.cell(r, 1).value: continue
    prov = norm(ws_m.cell(r, 4).value)
    nro = str(ws_m.cell(r, 7).value or "").strip()
    key = (prov, nro)
    if key not in nn_keys:
        # También probar por nro solo
        key_alt = next((k for k in nn_keys if k[1] == nro), None)
        if not key_alt: continue
        key = key_alt

    estado_actual = str(ws_m.cell(r, 20).value or "")
    if "NN-no-pagar" in estado_actual:
        ya_marcadas += 1
    else:
        ws_m.cell(r, 20).value = "NN-no-pagar"
        marcadas += 1
    nn_master[key] = True

no_encontradas = sum(1 for v in nn_master.values() if not v)

print(f"   Marcadas: {marcadas}")
print(f"   Ya marcadas: {ya_marcadas}")
print(f"   NN no encontradas en Master: {no_encontradas}\n")

if no_encontradas:
    print("Top NN no encontradas (agregar luego):")
    for (prov_n, nro), encontrada in nn_master.items():
        if encontrada: continue
        # Buscar info original
        for p, n, m in nn_list:
            if norm(p) == prov_n and n == nro:
                print(f"   {p:30} F{n:<10} ${m or 0:,.0f}")
                break

wb_m.save(EXCEL_PATH)
wb_m.close()
print("\nDone!")
