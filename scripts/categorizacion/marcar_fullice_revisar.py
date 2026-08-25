"""Deja el abono de Full Ice (7-ago-2026) en REVISAR hasta saber qué fue.

El dueño no sabe todavía a qué corresponde. REVISAR está en
`EXCLUIR_CATS_EGRESO`, así que no entra a ningún cálculo, y aparece en
/banco/revisar para no olvidarlo.
"""
import shutil
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb

FILA = 4952
COL_DESC, COL_ABONO, COL_CAT, COL_CULT = 2, 5, 8, 9

resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"Respaldo: {resp}\n")

wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]
desc = str(ws.cell(FILA, COL_DESC).value or "")
if "FULL ICE" not in desc.upper():
    print(f"❌ La fila {FILA} no es la de Full Ice, es: {desc[:60]}")
    wb.close()
    sys.exit(1)

actual = str(ws.cell(FILA, COL_CAT).value or "").strip()
if actual and actual != "REVISAR":
    print(f"Ya tenía categoría ({actual}) — no se toca.")
else:
    ws.cell(FILA, COL_CAT).value = "REVISAR"
    ws.cell(FILA, COL_CULT).value = "GENERAL"
    _save_wb(wb)
    monto = float(ws.cell(FILA, COL_ABONO).value or 0)
    print(f"✅ Fila {FILA} → REVISAR")
    print(f"   {desc[:60]}  +${monto:,.0f}")
    print("\n   Aparece en /banco/revisar. Al saber qué fue, se recategoriza ahí.")
wb.close()
