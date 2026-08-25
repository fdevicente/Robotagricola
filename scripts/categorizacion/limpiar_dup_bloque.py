#!/usr/bin/env python3
"""Elimina líneas repetidas dentro de un mismo bloque (2ª pasada).

Casos verificados uno a uno: la misma línea aparece dos veces (una con la glosa
cruda y otra reescrita), o una línea de despacho quedó con el total completo.
Se conserva la primera aparición y se corrige TOTAL FACTURA al monto de FXP.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook
from config import EXCEL_PATH
from excel_manager import _save_wb

# (fila_a_borrar, nro_factura, total_correcto, motivo)
CASOS = [
    (245,  "130492563", 277121,   "Sodimac: kit soldar repetido"),
    (256,  "5546669",   2055999,  "Copeval: PETROLEO repetido"),
    (827,  "6231207",   107928,   "Copeval: línea de despacho con el total completo"),
    (1157, "627",       5069400,  "Alpabesa: jornadas corte sierpe repetido"),
    (1205, "6",         595000,   "Inveragro: arriendo de jote repetido"),
]


def nrokey(v):
    s = str(v or "").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.upper().replace(" ", "").replace("-", "")


wb = load_workbook(EXCEL_PATH)
ws = wb["Facturas"]

# Verificar que las filas sigan siendo las esperadas antes de borrar
ok, abortar = [], []
for fila, nro, total_ok, motivo in CASOS:
    actual = nrokey(ws.cell(fila, 7).value)
    if actual == nrokey(nro):
        ok.append((fila, nro, total_ok, motivo))
    else:
        abortar.append((fila, nro, actual, motivo))

for fila, nro, actual, motivo in abortar:
    print(f"  ⚠️ OMITIDA fila {fila}: esperaba F{nro} pero hay F{actual} — no se toca")

# Corregir TOTAL FACTURA en todas las líneas de esas facturas
for fila, nro, total_ok, motivo in ok:
    for r in range(2, ws.max_row + 1):
        if nrokey(ws.cell(r, 7).value) == nrokey(nro):
            ws.cell(r, 16).value = total_ok

for fila, nro, total_ok, motivo in sorted(ok, key=lambda x: -x[0]):
    print(f"  ✔ fila {fila:>5} borrada | F{nro:<11} total → ${total_ok:>11,.0f} | {motivo}")
    ws.delete_rows(fila)

_save_wb(wb)
wb.close()
print(f"\n✅ {len(ok)} líneas eliminadas, {len(ok)} facturas con total corregido.")
