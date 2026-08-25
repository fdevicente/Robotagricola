#!/usr/bin/env python3
"""Reglas v7 - últimos patrones."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP_PATH = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"


def _parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    return None


KEYWORD_RULES = [
    # COMBUSTIBLE (camioneta LSLG62 + gas grua)
    (["combustible camioneta lslg", "combustible lslg62",
       "combustible lslg", "gas grua", "gas grúa"], "COMBUSTIBLE"),

    # MANTENCION PLANTA
    (["pasador planta", "pasadores plastico planta",
       "pasadores plástico planta", "fierros mesa sag",
       "mesa sag"], "MANTENCION PLANTA"),

    # SEGURIDAD (cámaras, wifi, electronica)
    (["camaras exteriores", "cámaras exteriores",
       "antena wifi", "repetidor wifi", "modulo porton",
       "módulo portón", "huellero digital", "huellero",
       "cable red", "conector y cubridor", "porton electrico",
       "portón eléctrico"], "SEGURIDAD"),
    (["electro tecnologias", "electro tecnologías",
       "technology & bits", "technology y bits",
       "comercializadora todoclick", "ikseg"], "SEGURIDAD"),

    # SERVICIOS PROFESIONALES
    (["notaria diez", "notaría diez", "notaria", "notaría",
       "alzamiento hipoteca", "alzamiento de hipoteca"],
     "SERVICIOS PROFESIONALES"),
    (["wom", "internet wom", "wom pago total", "internet campo"],
     "SERVICIOS PROFESIONALES"),

    # GASTOS BANCARIOS
    (["intereses por mora", "interes por mora", "comision transf",
       "comisión transf", "comsision transf", "comision transferencia",
       "comisión transferencia"], "GASTOS BANCARIOS"),

    # MANTENIMIENTO INFRAESTRUCTURA
    (["mantencion electrica campo", "mantención eléctrica campo",
       "mantenimiento electrico campo", "mantenimiento eléctrico campo"],
     "MANTENIMIENTO INFRAESTRUCTURA"),

    # ALIMENTACION Y ALOJAMIENTO
    (["cordero asado", "cordero 18 de septiembre",
       "cordero 18 septiembre"], "ALIMENTACION Y ALOJAMIENTO"),

    # SERVICIOS DE EXPORTACION
    (["envio doc espana", "envío doc españa", "envio doc españa",
       "envio documentos espana", "envío documentos españa",
       "dhl envio espana", "dhl envío españa", "dhl espana",
       "dhl españa"], "SERVICIOS DE EXPORTACION"),

    # MANO DE OBRA PLANTA (aguinaldos)
    (["aguinaldo septiembre", "aguinaldo navidad",
       "aguinaldo fiestas"], "MANO DE OBRA PLANTA"),

    # MANTENIMIENTO HELICOPTERO
    (["manga viento", "manga de viento", "servicios aeroespi",
       "aeroespi"], "MANTENIMIENTO HELICOPTERO"),

    # MATERIALES
    (["equipo proteccion trabajadores", "equipo protección trabajadores",
       "epp", "material oficina", "compra herramientas",
       "librerias tucan", "librería tucán", "librerías tucán",
       "easy ", "arreglo rifle"], "MATERIALES"),
    (["mercado libre", "mercadolibre"], "MATERIALES"),

    # MAQUINARIA - MANTENCION
    (["rueda carro campo", "rueda carro", "inv. santa victoria",
       "inv santa victoria"], "MAQUINARIA - MANTENCION"),
]


def match_keywords(text: str) -> str | None:
    t = text.lower()
    for kws, cat in KEYWORD_RULES:
        for kw in kws:
            if kw in t:
                return cat
    return None


print("[1/3] FXP...")
tmp_f = os.path.join(tempfile.gettempdir(), "fxp_v7.xlsx")
shutil.copy2(FXP_PATH, tmp_f)
wb_f = load_workbook(tmp_f, read_only=True, data_only=True)
ws_f = wb_f["ScotiaBCO"]
fxp_idx = {}
for row in ws_f.iter_rows(min_row=6, values_only=True):
    if not row or len(row) < 11: continue
    fecha = _parse_date(row[2])
    if not fecha: continue
    try:
        monto = int(round(float(row[6] or 0)))
    except: continue
    fxp_idx[(fecha.isoformat(), monto)] = (
        str(row[5] or ""), str(row[10] or "")
    )
wb_f.close()
print(f"   {len(fxp_idx)}\n")

print("[2/3] Aplicando v7...")
wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]

cat_counts = {}
actualizadas = 0
total = 0

for row in range(2, ws.max_row + 1):
    cat = ws.cell(row, 8).value
    if cat != "REVISAR": continue
    fecha = _parse_date(ws.cell(row, 1).value)
    if not fecha or fecha < date(2021, 1, 1): continue
    total += 1

    desc = str(ws.cell(row, 2).value or "")
    ref = str(ws.cell(row, 3).value or "")
    try:
        cargo = int(round(float(ws.cell(row, 4).value or 0)))
    except: cargo = 0

    fxp_desc, fxp_notas = fxp_idx.get((fecha.isoformat(), cargo), ("", ""))
    enriched = f"{desc} {ref} {fxp_desc} {fxp_notas}"

    nueva = match_keywords(enriched)
    if nueva:
        ws.cell(row, 8).value = nueva
        ws.cell(row, 9).value = "GENERAL"
        cat_counts[nueva] = cat_counts.get(nueva, 0) + 1
        actualizadas += 1

print(f"   Procesados: {total}, Actualizados: {actualizadas}\n")

print("[3/3] Guardando...")
wb.save(EXCEL_PATH)
wb.close()

print("\n=== APLICADAS ===")
for cat, n in sorted(cat_counts.items(), key=lambda x: -x[1]):
    print(f"  {cat}: {n}")
