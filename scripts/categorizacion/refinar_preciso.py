#!/usr/bin/env python3
"""Refina con precisión: re-evalúa SOLO las filas que esta sesión categorizó.

Identifica esas filas comparando contra el backup pre-categorización (col Categoría
vacía allí). Aplica el mapa mejorado (reglas documentadas del usuario). No toca
ninguna fila que ya tenía categoría antes de esta sesión.
"""
import sys, glob, os
sys.stdout.reconfigure(encoding="utf-8")
from collections import Counter
from openpyxl import load_workbook
from config import EXCEL_PATH

BKDIR = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\Backups\Master\snapshots"
bk = sorted(glob.glob(os.path.join(BKDIR, "MASTER_pre_categoriza_hist_*.xlsx")))[-1]
print(f"Backup de referencia: {os.path.basename(bk)}\n")

KEYWORD_MAP = [
    (["carlos torres", "comec", "pack man"], "INVERSION ACTIVO PLANTA"),
    (["remuneracion", "remuneración", "previred", "aguinaldo"], "MANO DE OBRA PLANTA"),
    (["finiquito", "jornal", "cuadrilla", "temporada", "gto cerezos",
      "gto camarico", "gto nogales"], "MANO DE OBRA TEMPORAL"),
    (["f29", "impto", "impuesto", " sii", "tesoreria", "tgr"], "IMPUESTOS"),
    (["reintegro", "devolucion", "devolución"], "REINTEGROS Y DEVOLUCIONES"),
    (["aduana", "exportac", "comex", "naviera", "maxisaco", "pacifor"], "SERVICIOS DE EXPORTACION"),
    (["irrifer", "irrifor", "wiseconn", "equipos de riego", "fertiriego",
      "indelec", "riego", "irrigat"], "RIEGO"),
    (["s-invest", "s invest", "sinvest", " cge", "enel", "energia generada",
      "energía generada"], "ENERGIA"),
    (["copec", "lipigas", "abastible", "gasco", "coragas", "estacion de servicio",
      "estacion servicio", "bencina", "petroleo", "petróleo"], "COMBUSTIBLE"),
    (["fertiliz", "anagra", "fosfato", "urea", "nitrato", "salitre"], "FERTILIZANTES"),
    (["fitosanit", "fungicida", "herbicida", "pesticida", "abamectin", "abametin",
      "silitec", "ripper"], "FITOSANITARIOS"),
    (["copeval", "martinez y valdiv", "agrokimun", "cals f", "cooperativa agr",
      "vals bio", "anasac", "agrocampo", "aliagrototal", "facaz", "agromaq",
      "qlf labs", "qualified", "inveragro", "disal"], "INSUMOS AGRICOLAS"),
    (["maqagri", "jl maquinaria", "promaq", "agroequipos", "arvimac", "maestranza",
      "repuesto", "rodamendez", "rodamiento", "rodasep", "quemador", "implementos",
      "maquinaria", "importacion y venta de repuestos", "roda asesorias"], "MAQUINARIA - MANTENCION"),
    (["ferreteria", "ferretería", "sodimac", "casa del perno", "perno", "construmart",
      "easy", "imperial", "casa musa", "resorte gomez", "treck", "comercial flemec",
      "carvallo", "plastitec", "importadora santa teresita", "cementos bio",
      "cemento", "alambres"], "MATERIALES"),
    (["bridgestone", "neumatico", "neumático", "vulcaniz", "automotriz",
      "revision tecnica", "permiso circulacion", "autopista", "tag "], "GASTOS VEHICULOS"),
    (["arriendo", "inmobiliaria", "yelcho", "seguro", "patente"], "ARRIENDOS / PATENTES / SEGUROS"),
    (["flete", "transporte", "cantu"], "TRANSPORTE"),
    (["caja chica", "gasto menor", "rendicion", "rendición", "mercado pago",
      "mercado libre", "mercadoli", "cargo pac", "redcompra", "ecommerce",
      "mandato agricola"], "CAJA CHICA / IMPREVISTOS"),
    (["facava", "electricidad", "construccion", "construcción", "menares"], "MANTENIMIENTO INFRAESTRUCTURA"),
    (["asesoria", "asesoría", "fgh", "auditor", "figueroa", "contador", "ingenieria",
      "ingeniería", "certificacion", "certificación", "universidad de talca",
      "laboratorio", "nogaltec", "nogalte", "leonardo", "agrologica", "agrológica",
      "consult", "globaltecno", "intercom", "wom", "movistar", "bh "], "SERVICIOS PROFESIONALES"),
]
DEFAULT_FACT = "SERVICIOS PROFESIONALES"
DEFAULT_BANCO = "MATERIALES"


def match(text):
    t = " " + text.lower() + " "
    for kws, cat in KEYWORD_MAP:
        for kw in kws:
            if kw in t:
                return cat
    return None


# ── Filas que estaban SIN categoría en el backup ──
wb_b = load_workbook(bk, read_only=True, data_only=True)
fact_vacias = set()
for idx, row in enumerate(wb_b["Facturas"].iter_rows(min_row=2, values_only=True), start=2):
    if row[0] and not str(row[16] or "").strip():
        fact_vacias.add(idx)
banco_vacias = set()
for idx, row in enumerate(wb_b["Cuenta Banco"].iter_rows(min_row=2, values_only=True), start=2):
    if not row[0]:
        continue
    c = str(row[7] or "").strip()
    if c == "" or c == "REVISAR":
        banco_vacias.add(idx)
wb_b.close()
print(f"Filas a re-evaluar -> Facturas: {len(fact_vacias)} | Banco: {len(banco_vacias)}\n")

# ── Re-evaluar en el Master actual ──
wb = load_workbook(EXCEL_PATH)
ws_f = wb["Facturas"]
cf = Counter()
chf = 0
for r in fact_vacias:
    prov = str(ws_f.cell(r, 4).value or "")
    glosa = str(ws_f.cell(r, 8).value or "") + " " + str(ws_f.cell(r, 9).value or "")
    nueva = match(prov + " " + glosa) or DEFAULT_FACT
    if nueva != str(ws_f.cell(r, 17).value or ""):
        ws_f.cell(r, 17).value = nueva
        chf += 1
    cf[nueva] += 1

ws = wb["Cuenta Banco"]
cb = Counter()
chb = 0
skip = 0
for r in banco_vacias:
    desc = str(ws.cell(r, 2).value or "")
    if "gestora e" in desc.lower():
        ws.cell(r, 8).value = "REVISAR"  # se mantiene
        skip += 1
        continue
    nueva = match(desc + " " + str(ws.cell(r, 3).value or "")) or DEFAULT_BANCO
    if nueva != str(ws.cell(r, 8).value or ""):
        ws.cell(r, 8).value = nueva
        chb += 1
    if not ws.cell(r, 9).value:
        ws.cell(r, 9).value = "GENERAL"
    cb[nueva] += 1

wb.save(EXCEL_PATH)
wb.close()

print("=== FACTURAS (re-evaluadas) ===")
for c, n in cf.most_common():
    print(f"  {n:4} | {c}")
print(f"  cambiadas vs pase anterior: {chf}")
print("\n=== BANCO (re-evaluadas, Gestora E queda REVISAR) ===")
for c, n in cb.most_common():
    print(f"  {n:4} | {c}")
print(f"  Gestora E en REVISAR: {skip} | cambiadas: {chb}")
