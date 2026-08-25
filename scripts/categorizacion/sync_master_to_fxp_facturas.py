#!/usr/bin/env python3
"""Sincroniza Master.Facturas con FXP.FXP:
1. Quita fecha_pago en facturas que FXP marca como pendientes/NN pero Master cree pagadas.
2. Agrega las que están en FXP y no en Master.
3. Elimina duplicados (misma factura con y sin prefijo FND).
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP_PATH = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"


def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try: return datetime.strptime(v[:10], fmt).date()
            except: pass
    return None


def norm(s):
    return (s or "").strip().upper().replace(".", "").replace("  ", " ")


# ─── Cargar FXP.FXP ──────────────────────────────────────
print("[1/5] Cargando FXP.FXP...")
tmp = os.path.join(tempfile.gettempdir(), "fxp_sync_fact.xlsx")
shutil.copy2(FXP_PATH, tmp)
wb_f = load_workbook(tmp, read_only=True, data_only=True)
ws_f = wb_f["FXP"]

fxp_facts = []
for row in ws_f.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    prov = str(row[6] or "")
    nro = str(row[7] or "").strip()
    if not nro: continue
    try:
        monto = float(row[8] or 0)
    except: monto = 0
    estado = str(row[11] or "").strip()
    fxp_facts.append({
        "fecha_emi": _pd(row[0]),
        "fecha_venc": _pd(row[1]),
        "fecha_pago": _pd(row[2]),
        "proveedor": prov,
        "nro": nro,
        "monto": monto,
        "estado": estado,
        "nota": str(row[12] or "").strip(),
        "key": (norm(prov), nro),
    })
wb_f.close()
print(f"   FXP facturas: {len(fxp_facts)}")
pagadas_fxp = sum(1 for f in fxp_facts if f["estado"].upper() == "PAGADA")
print(f"   Pagadas FXP: {pagadas_fxp}")
print(f"   Pendientes/NN FXP: {len(fxp_facts) - pagadas_fxp}\n")

# Indexar por key + por nro solo (fallback)
fxp_by_key = {f["key"]: f for f in fxp_facts}
fxp_by_nro = {}
for f in fxp_facts:
    fxp_by_nro.setdefault(f["nro"], []).append(f)


# ─── Cargar Master ──────────────────────────────────────
print("[2/5] Cargando Master.Facturas...")
wb_m = load_workbook(EXCEL_PATH)
ws_m = wb_m["Facturas"]
master_rows = {}  # (prov_norm, nro) -> [fila, fila_2, ...]
master_last_row = 1
for r in range(2, ws_m.max_row + 1):
    if not ws_m.cell(r, 1).value: continue
    master_last_row = r
    prov = norm(ws_m.cell(r, 4).value)
    nro = str(ws_m.cell(r, 7).value or "").strip()
    if not prov or not nro: continue
    master_rows.setdefault((prov, nro), []).append(r)
print(f"   Master: {len(master_rows)} llaves únicas, última fila {master_last_row}\n")


# ─── 1) Quitar fecha_pago en las que FXP dice pendiente ──
print("[3/5] Sincronizando fechas de pago...")
quitadas = 0
puestas = 0
for f in fxp_facts:
    # Buscar en master por key, fallback por nro
    rows = master_rows.get(f["key"]) or []
    if not rows:
        rows_nro = []
        for x in fxp_by_nro.get(f["nro"], []): pass  # solo para tener una referencia
        # Buscar en master por nro únicamente
        for k, rs in master_rows.items():
            if k[1] == f["nro"]:
                rows.extend(rs)
                break

    for r in rows:
        fecha_pago_master = ws_m.cell(r, 3).value
        tiene_pago = fecha_pago_master and str(fecha_pago_master).strip()

        if f["estado"].upper() == "PAGADA":
            if not tiene_pago and f["fecha_pago"]:
                ws_m.cell(r, 3).value = f["fecha_pago"]
                puestas += 1
        else:
            # Pendiente o NN → quitar fecha_pago si Master la tiene
            if tiene_pago:
                ws_m.cell(r, 3).value = None
                quitadas += 1
print(f"   Fechas de pago quitadas (FXP pendiente): {quitadas}")
print(f"   Fechas de pago agregadas (FXP pagada):  {puestas}\n")


# ─── 2) Agregar facturas FXP que no están en Master ─────
print("[4/5] Agregando facturas faltantes...")
agregadas = 0
agregadas_list = []
for f in fxp_facts:
    if f["key"] in master_rows: continue
    # Fallback por nro
    if any(k[1] == f["nro"] for k in master_rows.keys()): continue

    master_last_row += 1
    r = master_last_row
    fpago = f["fecha_pago"] if f["estado"].upper() == "PAGADA" else None
    fecha_emi = f["fecha_emi"] or f["fecha_venc"] or f["fecha_pago"]
    fecha_venc = f["fecha_venc"] or fecha_emi

    ws_m.cell(r, 1).value = fecha_emi
    ws_m.cell(r, 2).value = fecha_venc
    ws_m.cell(r, 3).value = fpago
    ws_m.cell(r, 4).value = f["proveedor"]
    ws_m.cell(r, 5).value = ""
    ws_m.cell(r, 6).value = "Factura"
    ws_m.cell(r, 7).value = f["nro"]
    ws_m.cell(r, 8).value = f["nota"] or ""
    ws_m.cell(r, 15).value = f["monto"]
    ws_m.cell(r, 16).value = f["monto"]
    ws_m.cell(r, 20).value = "FXP-pendiente" if f["estado"].upper() != "PAGADA" else "FXP-import"

    master_rows.setdefault(f["key"], []).append(r)
    agregadas += 1
    if f["estado"].upper() != "PAGADA":
        agregadas_list.append((f["fecha_venc"], f["monto"], f["proveedor"][:30], f["nro"], f["estado"], f["nota"][:40]))

print(f"   Agregadas: {agregadas}\n")


# ─── 3) Eliminar duplicados (F y FND/ND con mismo nro base) ─────
# CONSERVADOR: solo eliminar si UNA fila usa "F<num>" y OTRA "FND<num>" (o "ND<num>")
# con MISMO proveedor y MISMO monto. Las facturas con múltiples items (mismo nro,
# mismo prefijo) NO se tocan.
print("[5/5] Buscando duplicados F vs FND...")
filas_a_eliminar = set()

import re
def split_prefix_num(s):
    """Devuelve ('FND', '506812') o ('F', '506812') o (None, '506812')."""
    s = str(s or "").strip()
    m = re.match(r"^(FND|ND|F)(\d+)$", s, re.IGNORECASE)
    if m:
        return (m.group(1).upper(), m.group(2))
    return (None, s)

# Indexar por (prov, num_base) y guardar todas las (fila, prefijo, monto)
base_idx = {}
for (prov, nro), rows in master_rows.items():
    pref, num = split_prefix_num(nro)
    for r in rows:
        try: m = float(ws_m.cell(r, 16).value or 0)
        except: m = 0
        base_idx.setdefault((prov, num), []).append((r, pref, m, nro))

duplicados_log = []
for (prov, num), entries in base_idx.items():
    # Buscar pares donde hay F y FND/ND con el MISMO monto
    by_monto = {}
    for r, pref, monto, nro_orig in entries:
        by_monto.setdefault(round(monto, 2), []).append((r, pref, nro_orig))

    for monto, group in by_monto.items():
        if len(group) < 2: continue
        # Solo dedup si hay tanto "F" como "FND" o "ND" en el grupo
        prefs = {p for _, p, _ in group}
        has_F = "F" in prefs
        has_ND = "FND" in prefs or "ND" in prefs
        if not (has_F and has_ND): continue

        # Mantener la fila con FND/ND (viene de FXP con nota); eliminar las "F" puras
        for r, pref, nro_o in group:
            if pref == "F":
                filas_a_eliminar.add(r)
                duplicados_log.append((r, prov, nro_o, monto))

print(f"   Duplicados F+FND encontrados: {len(filas_a_eliminar)}")
for r, prov, nro_o, monto in duplicados_log[:20]:
    print(f"   Eliminar fila {r}: {prov[:30]:30} {nro_o:<12} ${monto:>10,.0f}")

for r in sorted(filas_a_eliminar, reverse=True):
    ws_m.delete_rows(r)

print(f"\nGuardando...")
wb_m.save(EXCEL_PATH)
wb_m.close()

print(f"\n=== RESUMEN ===")
print(f"  Fechas de pago quitadas: {quitadas}")
print(f"  Fechas de pago agregadas: {puestas}")
print(f"  Facturas agregadas:       {agregadas}")
print(f"  Duplicados eliminados:    {len(filas_a_eliminar)}")

if agregadas_list:
    print(f"\n=== NUEVAS PENDIENTES AGREGADAS ===")
    for venc, m, p, n, e, nt in sorted(agregadas_list, key=lambda x: (x[0] or date(2099,1,1))):
        print(f"  {venc} | {e:5} | {p:30} F{n:<10} | ${m:>10,.0f} | {nt}")
