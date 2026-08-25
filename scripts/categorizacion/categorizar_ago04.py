"""Categoriza los movimientos del 4-ago-2026 según el historial del proveedor.

SmartWays queda fuera a propósito: su historial está etiquetado como
MANTENIMIENTO HELICOPTERO pero las glosas dicen "Devolución préstamos".
Hay que preguntarle al dueño antes de asignarlo.
"""
import shutil
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import _save_wb

# fila → (categoría, cultivo, por qué)
ASIGNAR = {
    4948: ("INVERSION / REPLANTE", "AVELLANOS",
           "Agrícola Don Antonio — igual que el pago del 31-jul"),
    4949: ("ENERGIA", "GENERAL",
           "S-Invest 2 — 11 pagos previos como ENERGIA"),
    4950: ("MAQUINARIA - MANTENCION", "GENERAL",
           "Comercial Álamos — 4 pagos previos como MAQUINARIA"),
    4951: ("COMBUSTIBLE", "GENERAL",
           "Lipigas — 33 pagos previos como COMBUSTIBLE (gas)"),
}
COL_DESC, COL_CARGO, COL_CAT, COL_CULT = 2, 4, 8, 9

resp = shutil.copy2(EXCEL_PATH,
                    EXCEL_PATH.replace(".xlsx", f"_bak_{datetime.now():%Y%m%d_%H%M%S}.xlsx"))
print(f"Respaldo: {resp}\n")

wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]

print(f"{'fila':>5} {'monto':>13}  {'categoría':24} {'proveedor'}")
print("-" * 78)
n = 0
for fila, (cat, cult, razon) in sorted(ASIGNAR.items()):
    desc = str(ws.cell(fila, COL_DESC).value or "")
    actual = str(ws.cell(fila, COL_CAT).value or "").strip()
    if actual:
        print(f"{fila:>5}  ya tenía categoría ({actual}) — no se toca")
        continue
    monto = float(ws.cell(fila, COL_CARGO).value or 0)
    ws.cell(fila, COL_CAT).value = cat
    ws.cell(fila, COL_CULT).value = cult
    n += 1
    print(f"{fila:>5} {-monto:>13,.0f}  {cat[:24]:24} {desc[:34]}")
    print(f"{'':5} {'':13}  └─ {razon}")

_save_wb(wb)
wb.close()
print(f"\n✅ {n} movimientos categorizados.")
print("\n⚠️ Pendiente de confirmar: las 2 transferencias a SmartWays SpA "
      "(filas 4946-4947, $7.258.468).")
