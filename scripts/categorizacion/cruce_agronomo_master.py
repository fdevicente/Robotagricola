#!/usr/bin/env python3
"""Cruza planilla agronomo Camarico vs MASTER.Facturas y reporta discrepancias.

Mapeo agronomo → MASTER (propuesto):
- INSUMOS + PESTICIDAS Y FERTILIZANTES → FERTILIZANTES / FITOSANITARIOS / INSUMOS AGRICOLAS
- INSUMOS + COMBUSTIBLES Y LUBRICANTES → COMBUSTIBLE
- INSUMOS + ELETRICIDAD → ENERGIA
- REMUNERACIONES + CAMPO → MANO DE OBRA TEMPORAL
- REMUNERACIONES + ADMINISTRACION → MANO DE OBRA PLANTA
- MANTENCION MAQUINARIA → MAQUINARIA - MANTENCION
- MANTENCIONES + RIEGO → RIEGO
- COSTOS FIJOS + ASESORIAS → SERVICIOS PROFESIONALES
- ACTIVO FIJO + MAQUINAS Y EQUIPOS → INVERSION ACTIVO PLANTA
- ARRIENDOS → ARRIENDOS / PATENTES / SEGUROS
- EXPORTA / NUECES ESPAÑA / FLETES → SERVICIOS DE EXPORTACION
- VARIOS / HERRAMIENTAS MENORES → HERRAMIENTAS o MATERIALES
- VARIOS / UTILES OFICINA → MATERIALES
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from collections import defaultdict
from openpyxl import load_workbook
from config import EXCEL_PATH

AGRO_PATH = r"C:\Users\Windows\Dropbox\CAMARICO 2023\PLANILLA GASTOS CAMARICO 2023-2026.xlsx"


def map_agronomo_to_master(cargo: str, cargo_ll: str, sub_area: str) -> str:
    """Mapea categoría agronomo (cargo, cargo ll, sub_area) → categoría MASTER."""
    c = (cargo or "").strip().upper()
    cl = (cargo_ll or "").strip().upper()
    sa = (sub_area or "").strip().upper()

    if c == "INSUMOS":
        if "PESTICIDAS" in cl or "FERTILIZANTES" in cl: return "INSUMOS AGRICOLAS"
        if "COMBUSTIBLE" in cl: return "COMBUSTIBLE"
        if "ELETRI" in cl or "ELECTRI" in cl: return "ENERGIA"
        if "RIEGO" in cl: return "RIEGO"
        return "INSUMOS AGRICOLAS"

    if c == "REMUNERACIONES":
        if "CAMPO" in cl or "CONTRATISTA" in cl: return "MANO DE OBRA TEMPORAL"
        if "ADMINISTRACION" in cl or "OFICINA" in cl: return "MANO DE OBRA PLANTA"
        if "CASA CAMPO" in cl: return "MANO DE OBRA PLANTA"
        return "MANO DE OBRA TEMPORAL"

    if c == "MANTENCION MAQUINARIA":
        return "MAQUINARIA - MANTENCION"

    if c == "MANTENCIONES":
        if "RIEGO" in cl: return "RIEGO"
        if "PLANTA" in cl: return "MANTENCION PLANTA"
        if "TERRENO" in cl: return "MANTENIMIENTO INFRAESTRUCTURA"
        if "OFICINA" in cl: return "MANTENIMIENTO INFRAESTRUCTURA"
        return "MAQUINARIA - MANTENCION"

    if c == "COSTOS FIJOS":
        if "ASESORIAS" in cl or "ASESORIA" in cl: return "SERVICIOS PROFESIONALES"
        return "ARRIENDOS / PATENTES / SEGUROS"

    if c == "ACTIVO FIJO":
        if "PLANTAS" in cl: return "INVERSION / REPLANTE"
        return "INVERSION ACTIVO PLANTA"

    if c == "EXPORTA":
        return "SERVICIOS DE EXPORTACION"

    if c == "ARRIENDOS":
        return "ARRIENDOS / PATENTES / SEGUROS"

    if c == "ESTUDIOS Y EVALUACIONES":
        return "SERVICIOS PROFESIONALES"

    if c == "HABILITACION CAMPO":
        return "MANTENIMIENTO INFRAESTRUCTURA"

    if c == "VIVERO":
        return "INVERSION / REPLANTE"

    if c == "VARIOS":
        if "HERRAMIENTAS" in cl: return "HERRAMIENTAS"
        if "UTILES OFICINA" in cl: return "MATERIALES"
        if "FLETES" in cl: return "TRANSPORTE"
        return "MATERIALES"

    return None


def _parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    return None


def _norm_str(s):
    if s is None: return ""
    return str(s).strip().upper().replace(".", "").replace("  ", " ")


# Cargar agronomo
print("[1/3] Cargando planilla agronomo...")
tmp_a = os.path.join(tempfile.gettempdir(), "agro_cruce.xlsx")
shutil.copy2(AGRO_PATH, tmp_a)
wb_a = load_workbook(tmp_a, read_only=True, data_only=True)
ws_a = wb_a["DATOS"]

agro_idx = {}  # (proveedor_norm, nro_doc) -> dict
for row in ws_a.iter_rows(min_row=10, max_col=27, values_only=True):
    if not row[0]: continue
    prov = _norm_str(row[7])
    nro = str(row[18] or "").strip()
    cargo = row[4]
    cargo_ll = row[5]
    sub_area = row[3]
    try:
        total = float(row[25] or 0)
    except: total = 0
    if not prov or not nro: continue

    key = (prov, nro)
    cat_master = map_agronomo_to_master(cargo, cargo_ll, sub_area)

    agro_idx[key] = {
        "cargo": cargo, "cargo_ll": cargo_ll, "sub_area": sub_area,
        "cat_master": cat_master, "total": abs(total),
    }

wb_a.close()
print(f"   {len(agro_idx)} facturas agronomo indexadas\n")

# Cargar MASTER y cruzar
print("[2/3] Cruzando con MASTER.Facturas...")
tmp_m = os.path.join(tempfile.gettempdir(), "master_cruce.xlsx")
shutil.copy2(EXCEL_PATH, tmp_m)
wb_m = load_workbook(tmp_m, read_only=True, data_only=True)
ws_m = wb_m["Facturas"]

matches = 0
no_match = 0
coinciden = 0
discrepan = 0
sin_mapeo_agro = 0
discrepancias = []

for r_idx, row in enumerate(ws_m.iter_rows(min_row=2, max_col=20, values_only=True), start=2):
    if not row[0]: continue
    prov = _norm_str(row[3])  # col 4 = proveedor
    nro = str(row[6] or "").strip()  # col 7 = Nro factura
    cat_master = row[16]  # col 17 = categoria

    key = (prov, nro)
    if key not in agro_idx:
        no_match += 1
        continue
    matches += 1

    agro = agro_idx[key]
    cat_propuesta = agro["cat_master"]
    if not cat_propuesta:
        sin_mapeo_agro += 1
        continue

    if cat_master == cat_propuesta:
        coinciden += 1
    else:
        discrepan += 1
        discrepancias.append({
            "fila": r_idx, "prov": row[3], "nro": nro,
            "master": cat_master, "agronomo": agro,
            "propuesta": cat_propuesta,
        })

wb_m.close()

print(f"   Facturas en MASTER cruzadas con agronomo: {matches}")
print(f"   Sin match en agronomo:                    {no_match}")
print(f"   Coinciden categorías:                     {coinciden}")
print(f"   Discrepan:                                {discrepan}")
print(f"   Sin mapeo agronomo:                       {sin_mapeo_agro}\n")

# Resumen discrepancias por par master vs propuesta
print("[3/3] Discrepancias agrupadas (top 20):")
from collections import Counter
pares = Counter()
for d in discrepancias:
    pares[(d["master"], d["propuesta"])] += 1
for (m, p), n in pares.most_common(20):
    print(f"  {n:3d}× MASTER='{m}' → AGRONOMO_PROPONE='{p}'")

# Detalle primeras 15 discrepancias
print("\n\nPrimeras 15 discrepancias (detalle):")
for d in discrepancias[:15]:
    print(f"  Fila {d['fila']} | {d['prov']} | F{d['nro']}")
    print(f"    MASTER:     {d['master']}")
    print(f"    AGRONOMO:   {d['agronomo']['cargo']} / {d['agronomo']['cargo_ll']} (sub_area={d['agronomo']['sub_area']})")
    print(f"    PROPUESTA:  {d['propuesta']}\n")
