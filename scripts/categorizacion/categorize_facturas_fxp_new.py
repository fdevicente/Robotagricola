#!/usr/bin/env python3
"""Categoriza las facturas recién agregadas desde FXP (campo Categorizado_por='FXP-import')."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

# Reglas consolidadas v9 + reglas del agronomo
RULES = [
    # Conocidas operativas
    (["rotortec", "ayv"], "PRESTAMOS A OTRAS SOCIEDADES"),
    (["smartways"], "MANTENIMIENTO HELICOPTERO"),
    (["arrayan-aero", "aeromar", "helicoptero"], "MANTENIMIENTO HELICOPTERO"),
    (["miguel marin", "miguel marín"], "MANTENIMIENTO HELICOPTERO"),
    (["vision air", "hangar"], "MANTENIMIENTO HELICOPTERO"),
    (["s-invest", "s invest", "sinvest"], "ENERGIA"),
    (["cge", "lipigas"], "ENERGIA"),
    (["ecosmart", "intralog", "carmen gloria", "carmengloria"], "SERVICIOS DE EXPORTACION"),
    (["alpabesa"], "MANO DE OBRA TEMPORAL"),
    (["copeval", "martinez y valdivieso", "martinez valdivieso", "cals", "gmt",
       "agroquimic", "fitosanitar", "fertilizan", "pesticida", "fungicida",
       "herbicida", "insecticida", "acaricida", "fosfonat", "abamectin",
       "ethrel", "etherfon", "bioaminol", "aminopower", "bioestabilizado",
       "calcio", "magnesio", "potasio", "cna f", "hidalga", "agrokimun",
       "frutical emilia", "frasol"], "INSUMOS AGRICOLAS"),
    (["luis quiroz", "luis alberto quiroz", "flete", "transporte"], "TRANSPORTE"),
    (["scotiabank", "comex", "compra usd", "compra dolar"], "CAMBIO DIVISA"),
    (["francisco donoso", "donoso"], "SERVICIOS PROFESIONALES"),
    (["figueroa", "asesoria contable", "asesoría contable"], "SERVICIOS PROFESIONALES"),
    (["nogalte", "nogaltec", "asesoria nogal"], "SERVICIOS PROFESIONALES"),
    (["wiseconn", "wissecon"], "RIEGO"),
    (["agencia aduana", "aduana"], "SERVICIOS DE EXPORTACION"),
    (["agencia pardo", "francisco pardo"], "SERVICIOS DE EXPORTACION"),
    (["pack man", "packman", "sortek", "sorter", "despelonadora",
       "despelonador", "cilindro lavador"], "MANTENCION PLANTA"),
    (["pago cuota", "leasing", "rentaleas"], "LEASING"),
    (["pago tc", "tarjeta credito", "tarjeta crédito"], "MATERIALES"),
    # Pre-2021 marcador
    # Combustible
    (["copec", "facaz", "petrobras", "shell", "lubricante", "aceite motor",
       "diesel", "bencina"], "COMBUSTIBLE"),
    (["ventas al detalle"], "COMBUSTIBLE"),
    # Maquinaria
    (["agrocampo", "agromaq", "kohen", "felipe noguera", "ricardo mahla",
       "filtro tractor", "filtro de aire", "filtro petroleo", "filtro hidraulico",
       "neumatico", "neumático", "rodamiento", "tractor", "marsof",
       "central quemadores", "quemador", "patricio carvallo"], "MAQUINARIA - MANTENCION"),
    # Riego
    (["sercoriego", "indelec", "instalacion riego", "instalación riego",
       "bombas riego"], "RIEGO"),
    # Seguros
    (["aspor", "seguro"], "ARRIENDOS / PATENTES / SEGUROS"),
    (["municipalidad", "permiso edificacion", "permiso edificación",
       "permiso municipal"], "ARRIENDOS / PATENTES / SEGUROS"),
    # Materiales / herramientas / ferretería
    (["ferreteria", "ferretería", "ferreteria m y g", "ferreteria pachita",
       "ferreteria bascuñ", "ferreteria san", "sodimac", "easy ",
       "comercial alamos", "comercial alamosk", "comercializadora de aridos",
       "aridos san rafael", "el mimbral", "libreria tucan", "librería tucán",
       "panchita", "mercado libre", "mercadolibre", "starlink"], "MATERIALES"),
    # Energia / electricidad
    (["electricidad talca", "comercial electricidad", "leonardo nunez",
       "leonardo ivan nunez", "leonardo nuñez"], "SERVICIOS PROFESIONALES"),
    # Seguridad
    (["dacam", "fjtech", "sp digital", "jca supplies", "ikseg",
       "todoclick", "electrotecnologia", "electrotecnología",
       "technology & bits", "technology y bits", "camara seguridad",
       "cámara seguridad", "switch poe", "tplink", "porton electrico",
       "portón eléctrico"], "SEGURIDAD"),
    # Impuestos
    (["pago impto sii", "pago imp sii", "pago impt", " imp sii",
       "contribuciones", "f29", "pago impto tgr", "tgr"], "IMPUESTOS"),
    # Gastos vehiculos
    (["tag total", "pago total tag", "pago tag", " tag ", "neumaticos camioneta",
       "neumáticos camioneta", "mantenimiento camioneta",
       "mantenimiento camion", "aventura motors", "permiso de circulacion",
       "permiso de circulación", "soap", "bridgestone"], "GASTOS VEHICULOS"),
    # Mano de obra
    (["remuneracion", "remuneración", "previred", "aguinaldo",
       "boleta honorarios", "boletas honorarios"], "MANO DE OBRA PLANTA"),
    (["bh "], "MANO DE OBRA TEMPORAL"),
    # Reintegros
    (["gestora y tecnolog", "reintegro", "devolucion", "devolución"],
     "REINTEGROS Y DEVOLUCIONES"),
    # Transferencias internas
    (["traspaso banco", "traspaso bcos", "traspaso de banco",
       "traspaso agricola santa elisa", "traspaso cuenta bco"],
     "TRANSFERENCIA INTERNA"),
    # Pacifor / pallets
    (["pacifor", "pallet"], "INSUMOS AGRICOLAS"),
    # Alimentacion y alojamiento
    (["cordero asado", "alimento gato", "verduras y bebidas",
       "platillo volador", "pullman vitacura", "hotel"], "ALIMENTACION Y ALOJAMIENTO"),
    # Inversiones
    (["agricola el huingan", "agrícola el huingan", "compra avellanos",
       "compra cerezos"], "INVERSION / REPLANTE"),
    (["activo camarico planta", "balanza de piso", "trueno"],
     "INVERSION ACTIVO PLANTA"),
    (["astara"], "INVERSION VEHICULOS"),
    # Servicios profesionales
    (["rendicion viajes", "rendición viajes", "rendicion gg",
       "capital office", "movistar", "entel", "wom"],
     "SERVICIOS PROFESIONALES"),
    # Gastos bancarios
    (["intereses por mora", "intereses mora", "comision banco",
       "comisión banco", "comision transferencia",
       "comisión transferencia"], "GASTOS BANCARIOS"),
    # Mantencion infraestructura
    (["jose luis menares", "menares", "casa adm", "casa administracion"],
     "MANTENIMIENTO INFRAESTRUCTURA"),
    # Préstamos
    (["fondo mutuo", "fondos mutuos", "credito navidad", "navidad ltda",
       "inversiones navidad", "credito scotia", "préstamo",
       "prestamo rotortec"], "PRESTAMOS A OTRAS SOCIEDADES"),
    (["aporte el salto"], "PRESTAMOS A OTRAS SOCIEDADES"),
]


def match(text):
    t = (text or "").lower()
    for kws, cat in RULES:
        for kw in kws:
            if kw in t:
                return cat
    return None


def _pd(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try: return datetime.strptime(v[:10], fmt).date()
            except: pass
    return None


CUTOFF_HISTORICO = date(2021, 1, 1)

wb = load_workbook(EXCEL_PATH)
ws = wb["Facturas"]

actualizadas = 0
pre_2021 = 0
sin_match = 0
counts = {}

for r in range(2, ws.max_row + 1):
    if not ws.cell(r, 1).value: continue
    cat = ws.cell(r, 17).value
    # Solo procesar facturas sin categoría o REVISAR
    if cat and str(cat).strip() not in ("", "REVISAR"): continue

    fecha = _pd(ws.cell(r, 1).value)
    prov = str(ws.cell(r, 4).value or "")
    detalle = str(ws.cell(r, 8).value or "")
    glosa2 = str(ws.cell(r, 9).value or "")

    # Pre-2021 → HISTORICO
    if fecha and fecha < CUTOFF_HISTORICO:
        ws.cell(r, 17).value = "PRE-2021 HISTORICO"
        ws.cell(r, 18).value = "GENERAL"
        pre_2021 += 1
        continue

    text = f"{prov} {detalle} {glosa2}"
    new = match(text)
    if new:
        ws.cell(r, 17).value = new
        if not ws.cell(r, 18).value:
            ws.cell(r, 18).value = "GENERAL"
        counts[new] = counts.get(new, 0) + 1
        actualizadas += 1
    else:
        sin_match += 1

print(f"Actualizadas: {actualizadas}")
print(f"Pre-2021 → HISTORICO: {pre_2021}")
print(f"Sin match (quedan REVISAR/vacías): {sin_match}\n")
for c, n in sorted(counts.items(), key=lambda x: -x[1]):
    print(f"  {c}: {n}")

wb.save(EXCEL_PATH)
wb.close()
print("\nDone!")
