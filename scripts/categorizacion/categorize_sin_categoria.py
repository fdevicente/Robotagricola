#!/usr/bin/env python3
"""Categoriza los 605 movimientos sin Categoria - mayormente abonos (ingresos)."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import shutil, tempfile, os
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

FXP_PATH = r"C:\Users\Windows\Dropbox\Agricola Santa Elisa\FXP.xlsx"


def _parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    return None


def categorize_ingreso(desc: str, ref: str, fxp_desc: str, fxp_notas: str) -> str:
    """Categoriza ingresos (abonos) por keywords."""
    text = f"{desc} {ref} {fxp_desc} {fxp_notas}".lower()

    # Ventas / Exportaciones
    if any(k in text for k in ["valbifrut", "exportacion espana", "exportación españa",
                                "exportacion españa", "venta nueces", "venta espana",
                                "pacific nuts", "pacific nut", "factor", "facturing"]):
        return "INGRESO VENTAS"

    # Cosecha
    if any(k in text for k in ["cosecha", "venta cosecha"]):
        return "INGRESO VENTAS"

    # Transferencias internas / aportes recibidos
    if any(k in text for k in ["traspaso", "transferencia santa elisa",
                                "scotia a santander", "santander a scotia",
                                "aporte santa elisa"]):
        return "TRANSFERENCIA INTERNA"

    # Reintegros / devoluciones recibidas
    if any(k in text for k in ["devolucion", "devolución", "reintegro",
                                "gestora y tecnolog"]):
        return "REINTEGROS Y DEVOLUCIONES"

    # Préstamos recibidos / devoluciones de préstamos
    if any(k in text for k in ["prestamo", "préstamo", "rotortec", "sedos",
                                "smartways", "el salto"]):
        return "PRESTAMOS A OTRAS SOCIEDADES"

    # Intereses bancarios / depósitos a plazo
    if any(k in text for k in ["interes", "interés", "deposito plazo",
                                "depósito plazo", "rescate"]):
        return "INGRESO FINANCIERO"

    # Default: INGRESO OPERACIONAL
    return "INGRESO OPERACIONAL"


print("[1/3] FXP...")
tmp_f = os.path.join(tempfile.gettempdir(), "fxp_sincat.xlsx")
shutil.copy2(FXP_PATH, tmp_f)
wb_f = load_workbook(tmp_f, read_only=True, data_only=True)
ws_f = wb_f["ScotiaBCO"]
fxp_idx = {}
for row in ws_f.iter_rows(min_row=6, values_only=True):
    if not row or len(row) < 11: continue
    fecha = _parse_date(row[2])
    if not fecha: continue
    try:
        monto = int(round(float(row[6] or 0)))
        deposito = int(round(float(row[7] or 0)))
    except: continue
    # FXP usa monto para cargos y deposito para abonos
    key_monto = monto if monto > 0 else deposito
    if key_monto > 0:
        fxp_idx[(fecha.isoformat(), key_monto)] = (
            str(row[5] or ""), str(row[10] or "")
        )
wb_f.close()
print(f"   {len(fxp_idx)}\n")

print("[2/3] Procesando SIN CATEGORIA...")
wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]

cat_counts = {}
actualizadas = 0
pre_2021 = 0
post_2021 = 0
abonos = 0
cargos_sin_cat = 0

for row in range(2, ws.max_row + 1):
    fecha_val = ws.cell(row, 1).value
    if not fecha_val: continue
    fecha = _parse_date(fecha_val)
    if not fecha: continue

    cat = ws.cell(row, 8).value
    if cat:  # ya categorizado
        continue

    try:
        cargo = float(ws.cell(row, 4).value or 0)
    except: cargo = 0
    try:
        abono = float(ws.cell(row, 5).value or 0)
    except: abono = 0

    # Pre-2021 → HISTORICO
    if fecha < date(2021, 1, 1):
        ws.cell(row, 8).value = "PRE-2021 HISTORICO"
        ws.cell(row, 9).value = "GENERAL"
        pre_2021 += 1
        continue

    post_2021 += 1
    desc = str(ws.cell(row, 2).value or "")
    ref = str(ws.cell(row, 3).value or "")
    monto_clave = int(round(abono if abono > 0 else cargo))
    fxp_desc, fxp_notas = fxp_idx.get((fecha.isoformat(), monto_clave), ("", ""))

    if abono > 0 and cargo == 0:
        # Es un ingreso
        nueva = categorize_ingreso(desc, ref, fxp_desc, fxp_notas)
        abonos += 1
    elif cargo > 0:
        # Es un cargo sin categoría — marcar como REVISAR
        nueva = "REVISAR"
        cargos_sin_cat += 1
    else:
        # Sin cargo ni abono
        continue

    ws.cell(row, 8).value = nueva
    ws.cell(row, 9).value = "GENERAL"
    cat_counts[nueva] = cat_counts.get(nueva, 0) + 1
    actualizadas += 1

print(f"   Pre-2021 → HISTORICO: {pre_2021}")
print(f"   Post-2021 abonos: {abonos}")
print(f"   Post-2021 cargos sin cat → REVISAR: {cargos_sin_cat}")
print(f"   Total actualizadas: {actualizadas + pre_2021}\n")

print("[3/3] Guardando...")
wb.save(EXCEL_PATH)
wb.close()

print("\n=== CATEGORIAS APLICADAS ===")
for cat, n in sorted(cat_counts.items(), key=lambda x: -x[1]):
    print(f"  {cat}: {n}")
