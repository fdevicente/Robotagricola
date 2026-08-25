#!/usr/bin/env python3
"""Categoriza cargos REVISAR pre-2021 que sean prestamos/inversiones/transferencias
a otras sociedades como PRESTAMOS A OTRAS SOCIEDADES."""
import re
from datetime import date, datetime
from openpyxl import load_workbook
from config import EXCEL_PATH

# Keywords del usuario para identificar movimientos pre-presencia
KEYWORDS = [
    r"\bcredito\b", r"\bcr[eé]dito\b",
    r"fondo\s*mutuo", r"fondos\s*mutuos",
    r"\bnavidad\b",
    r"\bpr[eé]stamo", r"\bprestamo",
    r"\binversi[oó]n\b", r"\binversion\b",
    r"transferncia\s+alto\s+valor",
    r"transferencia\s+alto\s+valor",
]
PATTERN = re.compile("|".join(KEYWORDS), re.IGNORECASE)
CATEGORIA_NUEVA = "PRESTAMOS A OTRAS SOCIEDADES"
CUTOFF_DATE = date(2021, 1, 1)


def _parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(val[:10], fmt).date()
            except ValueError:
                pass
    return None


print(f"Abriendo {EXCEL_PATH}...")
wb = load_workbook(EXCEL_PATH)
ws = wb["Cuenta Banco"]

candidatos = []
for row in range(2, ws.max_row + 1):
    fecha_val = ws.cell(row, 1).value
    fecha = _parse_date(fecha_val)
    if not fecha or fecha >= CUTOFF_DATE:
        continue

    categoria = ws.cell(row, 8).value
    if categoria != "REVISAR":
        continue

    descripcion = str(ws.cell(row, 2).value or "")
    referencia = str(ws.cell(row, 3).value or "")
    cargo = float(ws.cell(row, 4).value or 0)

    if cargo <= 0:
        continue

    texto = f"{descripcion} {referencia}".lower()
    if PATTERN.search(texto):
        candidatos.append({
            "fila": row, "fecha": fecha.isoformat(),
            "descripcion": descripcion[:60], "cargo": cargo
        })

print(f"\nCandidatos: {len(candidatos)} cargos pre-2021 con keywords prestamo/inversion")
print(f"Monto total: ${sum(c['cargo'] for c in candidatos):,.0f} CLP\n")

print("Top 10:")
for c in sorted(candidatos, key=lambda x: -x["cargo"])[:10]:
    print(f"  Fila {c['fila']} | {c['fecha']} | ${c['cargo']:>15,.0f} | {c['descripcion']}")

# Aplicar categoria
print(f"\nAplicando '{CATEGORIA_NUEVA}' a {len(candidatos)} filas...")
for c in candidatos:
    ws.cell(c["fila"], 8).value = CATEGORIA_NUEVA
    ws.cell(c["fila"], 9).value = "GENERAL"

print("Guardando...")
wb.save(EXCEL_PATH)
wb.close()
print("Done!")
