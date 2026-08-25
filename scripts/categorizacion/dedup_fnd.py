#!/usr/bin/env python3
"""Elimina duplicados: facturas con el mismo número que tienen una versión 'normal'
y otra con prefijo FND/ND - solo elimina si MISMO proveedor y MISMO monto.
"""
import sys, re
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from config import EXCEL_PATH


def split_nro(s):
    """Devuelve (prefijo_tipo, nro_base).
    Tipo: 'ND' si tiene FND/ND, 'NORMAL' si no.
    """
    s = str(s or "").strip()
    m = re.match(r"^(FND|ND)(\d+)$", s, re.IGNORECASE)
    if m: return ("ND", m.group(2))
    m = re.match(r"^F?(\d+)$", s, re.IGNORECASE)
    if m: return ("NORMAL", m.group(1))
    return ("OTRO", s)


def norm(s):
    return (s or "").strip().upper().replace(".", "").replace("  ", " ")


wb = load_workbook(EXCEL_PATH)
ws = wb["Facturas"]

# Contar pendientes antes
pendientes_antes = 0
for r in range(2, ws.max_row + 1):
    if not ws.cell(r, 1).value: continue
    fpago = ws.cell(r, 3).value
    if not fpago or not str(fpago).strip():
        pendientes_antes += 1
print(f"Pendientes antes: {pendientes_antes}\n")

# Indexar por (proveedor_norm, nro_base) → [(fila, tipo, monto, nro_original), ...]
idx = {}
for r in range(2, ws.max_row + 1):
    if not ws.cell(r, 1).value: continue
    prov = norm(ws.cell(r, 4).value)
    nro_orig = ws.cell(r, 7).value
    if not prov or not nro_orig: continue
    tipo, base = split_nro(nro_orig)
    try: monto = float(ws.cell(r, 16).value or 0)
    except: monto = 0
    idx.setdefault((prov, base), []).append((r, tipo, monto, str(nro_orig)))

# Buscar pares NORMAL + ND con mismo monto
filas_eliminar = set()
ejemplos = []
for (prov, base), entries in idx.items():
    tipos = {t for _, t, _, _ in entries}
    if "ND" not in tipos or "NORMAL" not in tipos: continue

    # Agrupar por monto
    by_monto = {}
    for r, t, m, no in entries:
        by_monto.setdefault(round(m, 0), []).append((r, t, no))

    for monto, group in by_monto.items():
        tipos_g = {t for _, t, _ in group}
        if "ND" not in tipos_g or "NORMAL" not in tipos_g: continue
        # Mantener ND, eliminar NORMAL
        normales = [(r, no) for r, t, no in group if t == "NORMAL"]
        nds = [(r, no) for r, t, no in group if t == "ND"]
        for r, no in normales:
            filas_eliminar.add(r)
            ejemplos.append((r, prov[:30], no, monto, nds[0][1] if nds else ""))

print(f"Duplicados a eliminar (NORMAL con par ND mismo monto): {len(filas_eliminar)}\n")
print("Primeros 20 ejemplos (a eliminar):")
for r, p, no, m, nd in ejemplos[:20]:
    print(f"  Fila {r}: {p:30} {no:<12} ${m:>10,.0f} (queda {nd})")

# Eliminar (desc para no romper índices)
for r in sorted(filas_eliminar, reverse=True):
    ws.delete_rows(r)

# Contar pendientes después
pendientes_despues = 0
for r in range(2, ws.max_row + 1):
    if not ws.cell(r, 1).value: continue
    fpago = ws.cell(r, 3).value
    if not fpago or not str(fpago).strip():
        pendientes_despues += 1

print(f"\nPendientes antes:   {pendientes_antes}")
print(f"Pendientes después: {pendientes_despues}")
print(f"Diferencia:         {pendientes_antes - pendientes_despues}")

wb.save(EXCEL_PATH)
wb.close()
print("\nDone!")
