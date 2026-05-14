"""
main.py — Bot Agrícola Santa Elisa
"""
import os
import re
import sys
import logging
import asyncio
from datetime import datetime, time as dtime

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, CommandHandler,
    MessageHandler, CallbackQueryHandler,
    filters, ContextTypes
)

from config import TELEGRAM_TOKEN, DOWNLOAD_DIR, BOLETAS_DIR, EXCEL_PATH, TELEGRAM_CHAT_ID
from processors.extractor import process_file
from excel_manager import (append_to_excel, delete_last_rows, buscar_factura, registrar_pago,
                          append_boleta, delete_last_boletas, registrar_deposito_caja, consultar_saldo_caja,
                          reporte_diario, reporte_semanal, reporte_mensual, crear_hoja_banco,
                          guardar_movimientos_banco, obtener_resumen_banco)
from tareas_manager import (crear_hojas_tareas, crear_tarea, listar_tareas, actualizar_tarea,
                            registrar_bitacora, listar_bitacora, resumen_tareas_semana)
from inventario_manager import (crear_hojas_inventario, registrar_uso, consultar_inventario,
                                productos_bajo_stock, consumo_por_cultivo, agregar_stock_desde_factura)
from vacaciones_manager import (crear_hojas_vacaciones, agregar_trabajador, registrar_vacacion,
                                listar_personal, vacaciones_pendientes, actualizar_dias_mensuales,
                                ultimas_vacaciones)
from chat_inteligente import responder_chat

_LOG_FMT = "%(asctime)s - %(name)s - %(levelname)s - %(message)s"
logging.basicConfig(format=_LOG_FMT, level=logging.INFO)
logger = logging.getLogger(__name__)

# FileHandler además del stdout para que warnings/errors persistan
# (ej: fallas al renombrar archivos por handles abiertos en Windows).
try:
    _LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bot.log")
    _fh = logging.FileHandler(_LOG_FILE, encoding="utf-8")
    _fh.setLevel(logging.INFO)
    _fh.setFormatter(logging.Formatter(_LOG_FMT))
    logging.getLogger().addHandler(_fh)
except Exception as _e:
    logger.warning(f"No se pudo crear FileHandler de log: {_e}")

CAMPOS_EDITABLES = {
    "edit_proveedor": ("Nombre Factura / Proveedor",      "🏢 Nombre del proveedor"),
    "edit_rut":       ("Rut",                             "🪪 RUT del proveedor"),
    "edit_fecha":     ("Fecha Emision",                   "📅 Fecha de emisión (YYYY-MM-DD)"),
    "edit_vence":     ("Fecha Vencimiento",               "⏰ Fecha de vencimiento (YYYY-MM-DD)"),
    "edit_nro":       ("Numero Factura / Nro Documento",  "📄 Número de documento"),
    "edit_ref":       ("Referencia Factura",              "🔗 Nº factura referenciada (para NC/ND)"),
    "edit_glosa":     ("Detalle / Glosa",                 "📦 Glosa / descripción corta"),
    "edit_glosa2":    ("Glosa II",                        "📝 Detalle completo"),
    "edit_cantidad":  ("Cantidad",                        "🔢 Cantidad"),
    "edit_unitario":  ("Valor unitario",                  "💲 Valor unitario neto"),
    "edit_total":     ("TOTAL NETO",                      "💰 Total ítem sin IVA"),
}
# "edit_total_factura" se maneja aparte (no está en CAMPOS_EDITABLES)

def _save_path(filename):
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    return os.path.join(DOWNLOAD_DIR, filename)

def _save_path_boleta(filename):
    os.makedirs(BOLETAS_DIR, exist_ok=True)
    return os.path.join(BOLETAS_DIR, filename)

def _renombrar_archivo(file_path: str, items: list) -> str:
    """Renombra el archivo usando Proveedor + Nº Documento.
    Devuelve la nueva ruta (o la original si algo falla)."""
    try:
        item = items[0]
        proveedor = str(item.get("Nombre Factura / Proveedor") or "").strip()
        nro       = str(item.get("Numero Factura / Nro Documento") or "").strip()
        if not proveedor and not nro:
            return file_path

        def _limpiar(s):
            s = re.sub(r'[\\/:*?"<>|]', '', s)   # chars inválidos en Windows
            s = re.sub(r'\s+', '_', s.strip())    # espacios → guión bajo
            return s[:60]

        partes = []
        if proveedor:
            partes.append(_limpiar(proveedor))
        if nro:
            partes.append(nro)

        ext      = os.path.splitext(file_path)[1]
        nombre   = "_".join(partes) + ext
        dir_path = os.path.dirname(file_path)
        nuevo    = os.path.join(dir_path, nombre)

        # Si ya existe un archivo con ese nombre, agregar timestamp
        if os.path.exists(nuevo) and os.path.abspath(nuevo) != os.path.abspath(file_path):
            ts     = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre = "_".join(partes) + f"_{ts}" + ext
            nuevo  = os.path.join(dir_path, nombre)

        os.rename(file_path, nuevo)
        logger.info(f"Archivo renombrado: {os.path.basename(file_path)} → {nombre}")

        # Limpiar archivos derivados creados por el extractor (_resized.jpg,
        # _scan.png) que tienen el basename viejo. Ya no son necesarios
        # después del OCR y dejarlos huérfanos llena la carpeta de basura.
        base_viejo = os.path.splitext(file_path)[0]
        for sufijo in ("_resized.jpg", "_scan.png"):
            derivado = base_viejo + sufijo
            if os.path.exists(derivado):
                try:
                    os.remove(derivado)
                    logger.info(f"Derivado eliminado: {os.path.basename(derivado)}")
                except Exception as e:
                    logger.warning(f"No se pudo eliminar {derivado}: {e}")
        return nuevo
    except Exception as e:
        logger.error(f"No se pudo renombrar archivo {file_path}: {e}")
        return file_path

def _es_boleta(items):
    """Detecta si los ítems corresponden a una boleta de compra (caja chica).
    Las Boletas de Honorarios NO son caja chica — van a la hoja Facturas."""
    if not items: return False
    doc = str(items[0].get("Documento") or "").lower()
    return "boleta" in doc and "boleta de honorario" not in doc

def _esc(text):
    """Escapa caracteres especiales de Telegram Markdown."""
    if not text: return text
    for ch in ('*', '_', '`', '[', ']'):
        text = str(text).replace(ch, '')
    return text

def _format_date(val):
    if val is None: return "—"
    try:
        if isinstance(val, str): return val[:10]
        return val.strftime("%d/%m/%Y")
    except: return str(val)

def _calc_vencimiento(fecha_emision):
    """Calcula fecha de vencimiento = emisión + 30 días si no viene."""
    try:
        from datetime import datetime
        from dateutil.relativedelta import relativedelta
        if not fecha_emision:
            return None
        dt = datetime.strptime(str(fecha_emision)[:10], "%Y-%m-%d")
        return (dt + relativedelta(months=1)).strftime("%Y-%m-%d")
    except Exception:
        return None


def _registrar_correccion(item: dict, campo: str, valor_original, valor_nuevo):
    """Guarda la corrección del usuario para aprendizaje futuro en correcciones_log.json."""
    try:
        import json
        log_path = os.path.join(DOWNLOAD_DIR, "correcciones_log.json")
        log = []
        if os.path.exists(log_path):
            with open(log_path, "r", encoding="utf-8") as f:
                log = json.load(f)
        entrada = {
            "timestamp": datetime.now().isoformat(),
            "rut":        item.get("Rut"),
            "proveedor":  item.get("Nombre Factura / Proveedor"),
            "nro_factura":item.get("Numero Factura / Nro Documento"),
            "campo":      campo,
            "valor_claude": valor_original,
            "valor_usuario": valor_nuevo,
        }
        # Para campos numéricos calcular factor de corrección (útil para detectar errores sistemáticos)
        if campo in ("Monto / TOTAL", "Total Factura", "Valor unitario"):
            try:
                orig = float(valor_original or 0)
                nuevo = float(valor_nuevo or 0)
                if orig > 0 and nuevo > 0:
                    entrada["factor"] = round(nuevo / orig, 6)
            except Exception:
                pass
        log.append(entrada)
        with open(log_path, "w", encoding="utf-8") as f:
            json.dump(log, f, ensure_ascii=False, indent=2)
        logger.info(f"Corrección registrada: {campo} | '{valor_original}' → '{valor_nuevo}' ({item.get('Rut')})")
    except Exception as e:
        logger.warning(f"No se pudo registrar corrección: {e}")


def _build_preview(items):
    lines = ["📋 *Datos extraídos — revisa antes de guardar:*\n"]
    first = items[0]
    fecha_venc = first.get("Fecha Vencimiento") or _calc_vencimiento(first.get("Fecha Emision"))

    # Cabecera (datos comunes)
    lines.append(f"🏢 *Proveedor:* {_esc(first.get('Nombre Factura / Proveedor') or '? no detectado')}")
    lines.append(f"🪪 *RUT:* {_esc(first.get('Rut') or '? no detectado')}")
    doc_tipo = _esc(first.get('Documento') or '—')
    doc_nro = _esc(first.get('Numero Factura / Nro Documento') or '—')
    lines.append(f"📄 *Documento:* {doc_tipo}  Nº {doc_nro}")
    # Referencia para NC/ND
    ref = first.get("Referencia Factura")
    if ref:
        lines.append(f"🔗 *Ref. Factura:* Nº {_esc(ref)}")
    lines.append(f"📅 *Emisión:* {_format_date(first.get('Fecha Emision'))}   ⏰ *Vence:* {_format_date(fecha_venc)}\n")

    # Determinar régimen de IVA
    doc = str(first.get('Documento') or '').lower()
    es_honorario = "boleta de honorario" in doc
    exenta = any(k in doc for k in ('exenta', 'exento', 'no afecta', 'no afecto')) or es_honorario

    # Pasada 1: acumular pesos (unit×qty) e impuesto específico
    pesos = []
    total_imp_esp = 0.0
    for item in items:
        unitario = float(item.get('Valor unitario') or 0)
        cantidad = float(item.get('Cantidad') or 1)
        pesos.append(unitario * cantidad)
        total_imp_esp += float(item.get('Impuesto Especifico') or 0)
    total_neto_raw = sum(pesos)

    # NETO anchor: derivado del Total Factura (hacia atrás) o fallback al raw
    total_factura = round(float(first.get('Total Factura') or 0))
    if total_factura > 0:
        base_iva = total_factura - round(total_imp_esp)
        if exenta:
            iva_total = 0
            neto_anchor = base_iva
        else:
            iva_total = round(base_iva / 1.19 * 0.19)
            neto_anchor = base_iva - iva_total
        total_con_iva = total_factura
    else:
        neto_anchor = round(total_neto_raw)
        iva_total = 0 if exenta else round(total_neto_raw * 0.19)
        total_con_iva = round(neto_anchor * (1.0 if exenta else 1.19) + total_imp_esp)

    # Distribuir neto_anchor entre ítems — la suma SIEMPRE calza exactamente
    n = len(items)
    netos_linea = [0] * n
    if n > 0:
        if total_neto_raw > 0:
            acumulado = 0
            for i, peso in enumerate(pesos):
                if i == n - 1:
                    netos_linea[i] = neto_anchor - acumulado  # último absorbe el residuo
                else:
                    val = round(neto_anchor * peso / total_neto_raw)
                    netos_linea[i] = val
                    acumulado += val
        else:
            # Fallback uniforme si no hay pesos
            base = neto_anchor // n
            for i in range(n):
                netos_linea[i] = base
            netos_linea[-1] += neto_anchor - base * n

    # Pasada 2: renderizar ítems con netos ajustados
    for i, item in enumerate(items, 1):
        unitario = float(item.get('Valor unitario') or 0)
        cantidad = float(item.get('Cantidad') or 1)
        neto_linea = netos_linea[i - 1]

        if n > 1:
            lines.append(f"*— Ítem {i} —*")
        lines.append(f"📦 *Glosa:* {_esc(item.get('Detalle / Glosa') or '? no detectado')}")
        if item.get('Glosa II'):
            lines.append(f"📝 *Detalle:* {_esc(item.get('Glosa II'))}")
        lines.append(f"🔢 *Cantidad:* {cantidad:g}   💲 *Unit neto:* ${unitario:,.3f}")
        lines.append(f"💵 *Neto línea:* ${neto_linea:,.0f}\n")

    # Totales
    if es_honorario:
        # Boleta de Honorarios: neto_anchor = pago profesional; imp_esp = retención SII
        pago_profesional = neto_anchor
        retencion = round(total_imp_esp)
        costo_total = total_con_iva if total_factura > 0 else pago_profesional + retencion
        lines.append(f"💵 *Pago al profesional:* ${pago_profesional:,.0f}")
        if retencion:
            lines.append(f"🏦 *Impto. Retenido (ASE→SII):* ${retencion:,.0f}")
        lines.append(f"💰 *COSTO TOTAL:* ${costo_total:,.0f}\n")
    else:
        lines.append(f"📊 *NETO:* ${neto_anchor:,.0f}")
        if not exenta:
            lines.append(f"📊 *IVA 19%:* ${iva_total:,.0f}")
        if total_imp_esp:
            lines.append(f"⛽ *Imp. Específico:* ${total_imp_esp:,.0f}")
        lines.append(f"💰 *TOTAL:* ${total_con_iva:,.0f}\n")
    lines.append("¿Qué deseas hacer?")
    return "\n".join(lines)

def _main_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Guardar en Excel", callback_data="confirm_save"),
         InlineKeyboardButton("✏️ Editar campo",    callback_data="edit_menu")],
        [InlineKeyboardButton("❌ Cancelar",         callback_data="cancel_save")],
    ])

CAMPOS_COMUNES = {"edit_proveedor", "edit_rut", "edit_fecha", "edit_vence", "edit_nro", "edit_ref"}
CAMPOS_POR_ITEM = {"edit_glosa", "edit_glosa2", "edit_cantidad", "edit_unitario", "edit_total"}

def _edit_keyboard(items=None):
    n_items = len(items) if items else 1
    kb = [
        [InlineKeyboardButton("🏢 Proveedor",  callback_data="edit_proveedor"),
         InlineKeyboardButton("🪪 RUT",         callback_data="edit_rut")],
        [InlineKeyboardButton("📅 Fecha",       callback_data="edit_fecha"),
         InlineKeyboardButton("⏰ Vencimiento", callback_data="edit_vence")],
        [InlineKeyboardButton("📄 Nº Documento", callback_data="edit_nro"),
         InlineKeyboardButton("🔗 Ref. Factura", callback_data="edit_ref")],
        [InlineKeyboardButton("📦 Glosa",       callback_data="edit_glosa")],
        [InlineKeyboardButton("📝 Detalle",     callback_data="edit_glosa2"),
         InlineKeyboardButton("🔢 Cantidad",    callback_data="edit_cantidad")],
        [InlineKeyboardButton("💲 Unit neto",   callback_data="edit_unitario"),
         InlineKeyboardButton("💰 Total ítem",  callback_data="edit_total")],
        [InlineKeyboardButton("💰 TOTAL FACTURA", callback_data="edit_total_factura")],
    ]
    # Gestión de ítems (agregar / eliminar)
    item_row = [InlineKeyboardButton("➕ Agregar ítem", callback_data="add_item")]
    if n_items > 1:
        item_row.append(InlineKeyboardButton("🗑️ Eliminar ítem", callback_data="del_item"))
    kb.append(item_row)
    kb.append([InlineKeyboardButton("« Volver", callback_data="back_preview")])
    return InlineKeyboardMarkup(kb)


def _proveedor_nuevo_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Sí, agregar a la lista", callback_data="add_proveedor_yes"),
         InlineKeyboardButton("❌ No",                    callback_data="add_proveedor_no")],
    ])

def _rut_existe(rut):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb["Proveedores"]
        rut_n = rut.replace(".", "").replace("-", "").replace(" ", "").upper()
        for row in ws.iter_rows(min_row=3, values_only=True):
            r = str(row[2]).replace(".", "").replace("-", "").replace(" ", "").upper() if row[2] else ""
            if rut_n and rut_n == r:
                wb.close(); return True
        wb.close()
    except Exception as e:
        logger.warning(f"Error verificando RUT: {e}")
    return False

def _agregar_proveedor(nombre, rut):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Proveedores"]
        ws.append([None, nombre, rut])
        wb.save(EXCEL_PATH)
        logger.info(f"Proveedor agregado: {nombre} — {rut}")
        return True
    except Exception as e:
        logger.error(f"Error agregando proveedor: {e}")
        return False

# ── COMANDOS ──────────────────────────────────
async def cmd_start(update, context):
    # Guardar chat_id para jobs automáticos
    context.bot_data["banco_chat_id"] = update.effective_chat.id
    await update.message.reply_text(
        "👋 ¡Hola! Soy el bot de *Agrícola Santa Elisa*.\n\n"
        "📲 Envíame una *foto* o *PDF* de una factura.\n"
        "💬 También puedes escribirme consultas agrícolas.\n\n"
        "*Finanzas:*\n"
        "  /pagado — Registrar pago de factura\n"
        "  /deposito — Depositar en caja chica\n"
        "  /saldo — Saldo caja chica\n"
        "  /banco — Sincronizar Scotiabank\n"
        "  /reporte — Reportes diario/semanal/mensual\n\n"
        "*Inventario:*\n"
        "  /inventario — Ver stock de insumos\n"
        "  /uso — Registrar uso de insumo\n\n"
        "*Tareas y Bitácora:*\n"
        "  /tarea — Crear tarea nueva\n"
        "  /tareas — Ver tareas pendientes\n"
        "  /hecho — Marcar tarea completada\n"
        "  /bitacora — Registrar actividad diaria\n\n"
        "*Personal:*\n"
        "  /personal — Ver personal y vacaciones\n"
        "  /vacaciones — Registrar vacaciones\n"
        "  /agregar\\_trabajador — Agregar trabajador\n\n"
        "*Dashboard:*\n"
        "  /dashboard — Abrir panel de control web\n\n"
        "*Otros:*\n"
        "  /deshacer — Eliminar ultima factura\n"
        "  /cancelar — Cancelar operacion\n"
        "  /ayuda — Esta ayuda",
        parse_mode="Markdown")

async def cmd_ayuda(update, context): await cmd_start(update, context)

async def cmd_pagado(update, context):
    context.user_data["pagado_state"] = "esperando_nro"
    context.user_data["editing_field"] = None  # limpiar edición pendiente
    await update.message.reply_text(
        "💳 *Registrar pago de factura*\n\n"
        "Escribe el *número de factura* que fue pagada\n"
        "(o /cancelar para salir):",
        parse_mode="Markdown")

async def cmd_deposito(update, context):
    context.user_data["deposito_state"] = "esperando_monto"
    context.user_data["editing_field"] = None
    context.user_data["pagado_state"] = None
    await update.message.reply_text(
        "💰 *Depósito a Caja Chica*\n\n"
        "Escribe el *monto* del depósito (ej: 700000):\n"
        "(o /cancelar para salir)",
        parse_mode="Markdown")

async def cmd_saldo(update, context):
    info = await asyncio.to_thread(consultar_saldo_caja)
    saldo = info["saldo"]
    alerta = ""
    if saldo < 100000:
        alerta = "\n⚠️ *Saldo bajo, considerar nuevo depósito*"
    await update.message.reply_text(
        f"💰 *Estado Caja Chica*\n\n"
        f"📊 Saldo actual: *${saldo:,.0f}*{alerta}\n\n"
        f"📥 Total depositado: ${info['total_ingresos']:,.0f}\n"
        f"📤 Total gastado: ${info['total_egresos']:,.0f}\n"
        f"🧾 Gastos registrados: {info['n_gastos']}\n"
        f"📅 Último depósito: {info.get('ultimo_deposito') or '—'}",
        parse_mode="Markdown")

async def cmd_dashboard(update, context):
    """Inicia el dashboard web y envía el link."""
    import subprocess
    port = 5000
    # Verificar si ya está corriendo
    try:
        import urllib.request
        urllib.request.urlopen(f"http://localhost:{port}/", timeout=2)
        running = True
    except Exception:
        running = False

    if not running:
        # Lanzar dashboard en background
        dashboard_script = os.path.join(os.path.dirname(__file__), "src", "dashboard.py")
        subprocess.Popen([sys.executable, dashboard_script],
                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                         creationflags=0x00000008 if os.name == 'nt' else 0)  # DETACHED_PROCESS on Windows
        await asyncio.sleep(2)

    await update.message.reply_text(
        f"📊 *Dashboard Agricola Santa Elisa*\n\n"
        f"🌐 Abrir en el navegador:\n"
        f"`http://localhost:{port}`\n\n"
        f"Pestanas disponibles:\n"
        f"• General — Resumen del campo\n"
        f"• Facturas — Estado de pagos\n"
        f"• Costos por Cultivo — Nogales, Cerezos, Avellanos\n"
        f"• Exportaciones — Espana, comparacion anual\n"
        f"• Finanzas — Banco y Caja Chica\n"
        f"• Inventario — Stock de insumos\n"
        f"• Personal — Vacaciones pendientes\n"
        f"• Tareas — Bitacora y seguimiento",
        parse_mode="Markdown")

async def cmd_reporte(update, context):
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📋 Diario", callback_data="rep_diario"),
         InlineKeyboardButton("📊 Semanal", callback_data="rep_semanal")],
        [InlineKeyboardButton("📈 Mensual", callback_data="rep_mensual")]])
    await update.message.reply_text(
        "📊 *Reportes*\n\nSelecciona el tipo de reporte:",
        parse_mode="Markdown", reply_markup=kb)

async def cb_reporte(update, context):
    query = update.callback_query; await query.answer()
    tipo = query.data  # rep_diario, rep_semanal, rep_mensual

    if tipo == "rep_diario":
        await query.edit_message_text("⏳ Generando reporte diario...")
        r = await asyncio.to_thread(reporte_diario)
        texto = f"📋 *REPORTE DIARIO — {r['fecha']}*\n\n"
        if r["vencen_hoy"]:
            texto += f"⚠️ *Vencen HOY ({len(r['vencen_hoy'])}):*\n"
            for f in r["vencen_hoy"][:10]:
                texto += f"  {_esc(f.get('proveedor'))} — Nº {_esc(f.get('nro_factura'))} — ${float(f.get('total') or 0):,.0f}\n"
            texto += f"  *Total: ${r['total_hoy']:,.0f}*\n\n"
        else:
            texto += "✅ No hay facturas que venzan hoy.\n\n"
        if r["vencidas"]:
            texto += f"🔴 *Vencidas sin pagar ({len(r['vencidas'])}):*\n"
            for f in r["vencidas"][:15]:
                texto += f"  {_esc(f.get('proveedor'))} — Nº {_esc(f.get('nro_factura'))} — ${float(f.get('total') or 0):,.0f}\n"
            texto += f"  *Total vencido: ${r['total_vencido']:,.0f}*\n"
        else:
            texto += "✅ No hay facturas vencidas pendientes."

    elif tipo == "rep_semanal":
        await query.edit_message_text("⏳ Generando reporte semanal...")
        r = await asyncio.to_thread(reporte_semanal)
        texto = f"📊 *REPORTE SEMANAL — {r['periodo']}*\n\n"
        texto += f"✅ *Pagadas esta semana:* {len(r['pagadas'])} — ${r['total_pagado']:,.0f}\n"
        if r["pagadas"]:
            for f in r["pagadas"][:10]:
                texto += f"  {_esc(f.get('proveedor'))} — ${float(f.get('total') or 0):,.0f}\n"
        texto += f"\n🔴 *Vencidas sin pagar:* {len(r['vencidas'])} — ${r['total_vencido']:,.0f}\n"
        if r["vencidas"]:
            for f in r["vencidas"][:10]:
                texto += f"  {_esc(f.get('proveedor'))} — Nº {_esc(f.get('nro_factura'))}\n"
        texto += f"\n💵 *Caja Chica esta semana:* {len(r['gastos_caja'])} gastos — ${r['total_caja']:,.0f}\n"
        if r["gastos_caja"]:
            for g in r["gastos_caja"][:10]:
                texto += f"  {_esc(g.get('comercio') or g.get('detalle'))} — ${g.get('monto', 0):,.0f}\n"

    elif tipo == "rep_mensual":
        await query.edit_message_text("⏳ Generando reporte mensual...")
        r = await asyncio.to_thread(reporte_mensual)
        texto = f"📈 *REPORTE MENSUAL — {r['periodo']}*\n\n"
        texto += f"📄 Facturas recibidas: *{r['total_emitidas']}* — ${r['monto_emitido']:,.0f}\n"
        texto += f"✅ Facturas pagadas: *{r['total_pagadas']}* — ${r['monto_pagado']:,.0f}\n"
        texto += f"🔴 Vencidas sin pagar: *{r['vencidas_sin_pago']}* — ${r['monto_vencido']:,.0f}\n"
        texto += f"💵 Caja chica: *{r['n_gastos_caja']}* gastos — ${r['total_caja_mes']:,.0f}\n"
        if r["top_proveedores"]:
            texto += f"\n🏢 *Top proveedores por gasto:*\n"
            for i, (prov, monto) in enumerate(r["top_proveedores"], 1):
                texto += f"  {i}. {_esc(prov)} — ${monto:,.0f}\n"
    else:
        texto = "❌ Tipo de reporte no reconocido."

    try:
        await query.edit_message_text(texto, parse_mode="Markdown")
    except Exception:
        await query.edit_message_text(texto)

async def _sync_banco_core() -> str:
    """Lógica compartida de sincronización bancaria. Retorna texto del resultado."""
    from scotiabank_scraper import sync_scotiabank_movements

    movimientos = await asyncio.to_thread(sync_scotiabank_movements)

    if not movimientos:
        return "🏦 Conexión exitosa pero no se encontraron movimientos nuevos."

    result = await asyncio.to_thread(guardar_movimientos_banco, movimientos)
    resumen = await asyncio.to_thread(obtener_resumen_banco)

    texto = f"🏦 *Scotiabank — Sincronización completa*\n\n"
    texto += f"📥 Movimientos obtenidos: *{result['total']}*\n"
    texto += f"✅ Nuevos guardados: *{result['nuevos']}*\n"
    texto += f"⏭️ Duplicados omitidos: *{result['duplicados']}*\n\n"
    if resumen["ultimo_saldo"] is not None:
        texto += f"💰 Último saldo: *${resumen['ultimo_saldo']:,.0f}*\n"
    texto += f"📊 Total en hoja: *{resumen['n_movimientos']}* movimientos\n"
    texto += f"   Cargos: ${resumen['total_cargos']:,.0f}\n"
    texto += f"   Abonos: ${resumen['total_abonos']:,.0f}"
    return texto


async def cmd_banco(update, context):
    """Sincroniza movimientos de Scotiabank y los guarda en Cuenta Banco."""
    # Guardar chat_id para las notificaciones automáticas
    chat_id = update.effective_chat.id
    context.bot_data["banco_chat_id"] = chat_id
    msg = await update.message.reply_text("🏦 Conectando con Scotiabank...\nEsto puede tomar 30-60 segundos.")
    try:
        texto = await _sync_banco_core()
        try:
            await msg.edit_text(texto, parse_mode="Markdown")
        except Exception:
            await msg.edit_text(texto)
    except ImportError:
        await msg.edit_text("❌ Playwright no está instalado.\nEjecuta: pip install playwright && playwright install chromium")
    except RuntimeError as e:
        await msg.edit_text(f"❌ Error: {str(e)}")
    except Exception as e:
        logger.error(f"Error sync banco: {e}")
        await msg.edit_text(f"❌ Error al sincronizar con el banco:\n{str(e)[:200]}")


async def job_sync_banco(context: ContextTypes.DEFAULT_TYPE):
    """Job programado: sincroniza banco y envía notificación."""
    chat_id = context.bot_data.get("banco_chat_id") or TELEGRAM_CHAT_ID
    if not chat_id:
        logger.warning("Job banco: no hay chat_id configurado. Usa /banco al menos una vez o agrega TELEGRAM_CHAT_ID al .env")
        return
    chat_id = int(chat_id)
    hora = datetime.now().strftime("%H:%M")
    try:
        texto = await _sync_banco_core()
        texto = f"⏰ *Sincronización automática ({hora})*\n\n" + texto
        try:
            await context.bot.send_message(chat_id=chat_id, text=texto, parse_mode="Markdown")
        except Exception:
            await context.bot.send_message(chat_id=chat_id, text=texto)
    except Exception as e:
        logger.error(f"Job banco falló: {e}")
        try:
            await context.bot.send_message(
                chat_id=chat_id,
                text=f"⏰ Sync automático banco ({hora}) falló:\n{str(e)[:200]}")
        except Exception:
            pass


async def job_vacaciones_mensuales(context: ContextTypes.DEFAULT_TYPE):
    """Job mensual: acumula días de vacaciones a cada trabajador."""
    try:
        result = await asyncio.to_thread(actualizar_dias_mensuales)
        chat_id = context.bot_data.get("banco_chat_id") or TELEGRAM_CHAT_ID
        if chat_id:
            await context.bot.send_message(
                chat_id=int(chat_id),
                text=f"🏖️ *Vacaciones actualizadas*\n\n"
                     f"📊 {result['actualizados']} trabajadores\n"
                     f"📅 +{result['incremento']} días acumulados por persona",
                parse_mode="Markdown")
    except Exception as e:
        logger.error(f"Job vacaciones falló: {e}")


async def cmd_cancelar(update, context):
    for key in ("pagado_state", "pagado_nro", "deposito_state", "editing_field",
                "editing_item_idx", "tarea_state", "uso_state", "uso_data",
                "vacacion_state", "vacacion_data", "trabajador_state", "bitacora_state"):
        context.user_data[key] = None
    await update.message.reply_text("🚫 Operación cancelada.")

async def cmd_deshacer(update, context):
    last_file = context.user_data.get("last_invoice_file")
    last_rows = context.user_data.get("last_invoice_rows", 0)
    if not last_file or last_rows <= 0:
        await update.message.reply_text("⚠️ No hay factura reciente para deshacer."); return
    msg = await update.message.reply_text("⏳ Deshaciendo…")
    try:
        fue_boleta = context.user_data.get("last_invoice_boleta", False)
        if fue_boleta:
            success = await asyncio.to_thread(delete_last_boletas, last_rows)
        else:
            success = await asyncio.to_thread(delete_last_rows, last_rows)
    except Exception as e: logger.error(e); success = False
    if success:
        try:
            if os.path.exists(last_file): os.remove(last_file)
        except: pass
        context.user_data["last_invoice_file"] = None
        context.user_data["last_invoice_rows"]  = 0
        await msg.edit_text(f"🗑️ Listo. {last_rows} fila(s) eliminadas del Excel.")
    else:
        await msg.edit_text("❌ Error al borrar del Excel.")

# ── TAREAS Y BITÁCORA ────────────────────────
async def cmd_tarea(update, context):
    context.user_data["tarea_state"] = "esperando_desc"
    await update.message.reply_text(
        "📝 *Nueva tarea*\n\nEscribe la *descripción* de la tarea:\n(o /cancelar)",
        parse_mode="Markdown")

async def cmd_tareas(update, context):
    tareas = await asyncio.to_thread(listar_tareas)
    if not tareas:
        await update.message.reply_text("✅ No hay tareas pendientes."); return
    texto = "📋 *Tareas pendientes:*\n\n"
    for t in tareas[:20]:
        icon = "🔴" if t["prioridad"] == "Alta" else "🟡" if t["prioridad"] == "Media" else "🟢"
        estado_icon = "🔄" if t["estado"] == "En Progreso" else "⏳"
        texto += f"{estado_icon} {icon} *#{t['id']}* — {_esc(t['descripcion'])}\n"
        if t["responsable"]:
            texto += f"    👤 {_esc(t['responsable'])}"
        if t["fecha_limite"]:
            texto += f"  📅 {t['fecha_limite']}"
        texto += "\n"
    try:
        await update.message.reply_text(texto, parse_mode="Markdown")
    except Exception:
        await update.message.reply_text(texto)

async def cmd_hecho(update, context):
    context.user_data["tarea_state"] = "esperando_id_hecho"
    await update.message.reply_text(
        "✅ *Completar tarea*\n\nEscribe el *número de tarea* (ej: 3):\n(o /cancelar)",
        parse_mode="Markdown")

async def cmd_bitacora(update, context):
    context.user_data["bitacora_state"] = "esperando_registro"
    await update.message.reply_text(
        "📓 *Bitácora diaria*\n\nEscribe lo que quieras registrar:\n(o /cancelar)",
        parse_mode="Markdown")

# ── INVENTARIO ───────────────────────────────
async def cmd_inventario(update, context):
    items = await asyncio.to_thread(consultar_inventario)
    if not items:
        await update.message.reply_text("📦 Inventario vacío. Los insumos se agregan automáticamente al procesar facturas."); return
    texto = "📦 *Inventario de insumos:*\n\n"
    for it in items[:30]:
        alerta = " ⚠️" if it["alerta"] else ""
        texto += f"• *{_esc(it['producto'])}*\n"
        texto += f"  {it['categoria']} — Stock: {it['stock']:g} {it['unidad']}{alerta}\n"
    bajo = [i for i in items if i["alerta"]]
    if bajo:
        texto += f"\n⚠️ *{len(bajo)} producto(s) con stock bajo*"
    try:
        await update.message.reply_text(texto, parse_mode="Markdown")
    except Exception:
        await update.message.reply_text(texto)

async def cmd_uso(update, context):
    context.user_data["uso_state"] = "esperando_producto"
    context.user_data["uso_data"] = {}
    await update.message.reply_text(
        "🧪 *Registrar uso de insumo*\n\nEscribe el *nombre del producto*:\n(o /cancelar)",
        parse_mode="Markdown")

# ── VACACIONES Y PERSONAL ────────────────────
async def cmd_personal(update, context):
    personal = await asyncio.to_thread(listar_personal)
    if not personal:
        await update.message.reply_text("👥 No hay personal registrado.\nUsa /agregar\\_trabajador para agregar.", parse_mode="Markdown"); return
    texto = "👥 *Personal — Agrícola Santa Elisa:*\n\n"
    for p in personal:
        texto += f"• *{_esc(p['nombre'])}*"
        if p['cargo']:
            texto += f" — {_esc(p['cargo'])}"
        texto += f"\n  📅 Pendientes: *{p['dias_pendientes']:.0f} días*"
        if p['ultima_vacacion']:
            texto += f" | Última: {p['ultima_vacacion'][:10]}"
        texto += "\n"
    try:
        await update.message.reply_text(texto, parse_mode="Markdown")
    except Exception:
        await update.message.reply_text(texto)

async def cmd_vacaciones(update, context):
    context.user_data["vacacion_state"] = "esperando_nombre"
    context.user_data["vacacion_data"] = {}
    personal = await asyncio.to_thread(listar_personal)
    if personal:
        nombres = "\n".join(f"  • {p['nombre']}" for p in personal)
        await update.message.reply_text(
            f"🏖️ *Registrar vacaciones*\n\nPersonal:\n{nombres}\n\n"
            "Escribe el *nombre del trabajador*:\n(o /cancelar)",
            parse_mode="Markdown")
    else:
        await update.message.reply_text(
            "🏖️ *Registrar vacaciones*\n\nNo hay personal registrado.\n"
            "Usa /agregar\\_trabajador primero.", parse_mode="Markdown")

async def cmd_agregar_trabajador(update, context):
    context.user_data["trabajador_state"] = "esperando_nombre"
    await update.message.reply_text(
        "👤 *Agregar trabajador*\n\nEscribe el *nombre completo*:\n(o /cancelar)",
        parse_mode="Markdown")

async def cb_tarea_prioridad(update, context):
    """Callback para seleccionar prioridad de tarea."""
    query = update.callback_query; await query.answer()
    prioridad = query.data.replace("prior_", "")
    desc = context.user_data.get("tarea_desc", "")
    result = await asyncio.to_thread(crear_tarea, desc, prioridad)
    context.user_data["tarea_state"] = None
    await query.edit_message_text(
        f"✅ *Tarea #{result['id']} creada*\n\n"
        f"📝 {_esc(result['descripcion'])}\n"
        f"{'🔴' if prioridad == 'Alta' else '🟡' if prioridad == 'Media' else '🟢'} Prioridad: {prioridad}",
        parse_mode="Markdown")

async def cb_cultivo(update, context):
    """Callback para seleccionar cultivo en uso de insumo."""
    query = update.callback_query; await query.answer()
    cultivo = query.data.replace("cult_", "")
    data = context.user_data.get("uso_data", {})
    result = await asyncio.to_thread(
        registrar_uso, data.get("producto", ""), data.get("cantidad", 0),
        cultivo, data.get("sector", ""))
    context.user_data["uso_state"] = None
    context.user_data["uso_data"] = {}
    alerta = "\n⚠️ *Stock bajo!*" if result.get("alerta_bajo") else ""
    await query.edit_message_text(
        f"✅ *Uso registrado*\n\n"
        f"🧪 {_esc(result['producto'])} — {result['cantidad']:g} {result['unidad']}\n"
        f"🌳 Cultivo: {cultivo}\n"
        f"📦 Stock restante: {result['stock_restante']:g} {result['unidad']}{alerta}",
        parse_mode="Markdown")

# ── ARCHIVOS ──────────────────────────────────
async def _download_with_retry(f, path, retries=3):
    """Descarga archivo de Telegram con reintentos."""
    for attempt in range(1, retries + 1):
        try:
            await f.download_to_drive(path)
            return True
        except Exception as e:
            logger.warning(f"Descarga intento {attempt}/{retries} falló: {e}")
            if attempt < retries:
                await asyncio.sleep(2 * attempt)
    return False

async def handle_document(update, context):
    doc = update.message.document
    if doc.mime_type != "application/pdf":
        await update.message.reply_text("❌ Solo acepto PDFs. Para imágenes, envíalas como foto."); return
    status = await update.message.reply_text("📥 Recibí tu PDF. Descargándolo...")
    try:
        f = await context.bot.get_file(doc.file_id)
        path = _save_path(doc.file_name)
        if not await _download_with_retry(f, path):
            await status.edit_text("❌ No pude descargar el PDF. Intenta enviarlo de nuevo."); return
        await status.edit_text("🔍 Leyendo documento con IA… (puede tardar hasta 1 minuto)")
        await _process_and_reply(update, context, status, path)
    except Exception as e:
        logger.error(f"Error en handle_document: {e}")
        try: await status.edit_text("❌ Error al procesar el PDF. Intenta de nuevo.")
        except: pass

async def handle_photo(update, context):
    photo = update.message.photo[-1]
    status = await update.message.reply_text("📥 Recibí tu imagen. Descargándola...")
    try:
        f = await context.bot.get_file(photo.file_id)
        filename = f"factura_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{photo.file_unique_id}.jpg"
        path = _save_path(filename)
        if not await _download_with_retry(f, path):
            await status.edit_text("❌ No pude descargar la imagen. Intenta enviarla de nuevo."); return
        await status.edit_text("🔍 Leyendo documento con IA… (puede tardar hasta 1 minuto)")
        await _process_and_reply(update, context, status, path)
    except Exception as e:
        logger.error(f"Error en handle_photo: {e}")
        try: await status.edit_text("❌ Error al procesar la imagen. Intenta de nuevo.")
        except: pass

# ── PROCESAMIENTO ─────────────────────────────
async def _process_and_reply(update, context, status_msg, file_path):
    result = await asyncio.to_thread(process_file, file_path)
    if result.get("status") == "error":
        await status_msg.edit_text(f"❌ {result.get('message')}", parse_mode="Markdown"); return
    items = result.get("items", [])
    if not items:
        await status_msg.edit_text("⚠️ No se identificaron datos. Intenta con foto más nítida."); return

    # Renombrar el archivo con Proveedor + Nº Documento para fácil identificación
    file_path = _renombrar_archivo(file_path, items)

    context.user_data["pending_items"]     = items
    context.user_data["pending_file_path"] = file_path
    context.user_data["editing_field"]     = None

    preview = _build_preview(items)

    # Advertir si es factura duplicada
    if result.get("duplicado"):
        preview = "⚠️ *ADVERTENCIA: Esta factura parece ya estar registrada.*\nRevisa bien antes de guardar.\n\n" + preview

    try:
        await status_msg.edit_text(preview, parse_mode="Markdown", reply_markup=_main_keyboard())
    except Exception:
        # Fallback sin Markdown si hay caracteres problemáticos
        await status_msg.edit_text(preview, reply_markup=_main_keyboard())

async def _show_preview(query, context):
    items = context.user_data.get("pending_items", [])
    await query.edit_message_text(_build_preview(items), parse_mode="Markdown", reply_markup=_main_keyboard())

# ── CALLBACKS ─────────────────────────────────
async def cb_confirm_save(update, context):
    query = update.callback_query; await query.answer()
    items     = context.user_data.get("pending_items", [])
    file_path = context.user_data.get("pending_file_path")
    if not items: await query.edit_message_text("⚠️ No hay datos pendientes."); return

    # ¿Proveedor nuevo?
    first  = items[0]
    nombre = first.get("Nombre Factura / Proveedor")
    rut    = first.get("Rut")
    if nombre and rut and not _rut_existe(rut):
        context.user_data["nuevo_proveedor_nombre"] = nombre
        context.user_data["nuevo_proveedor_rut"]    = rut
        await query.edit_message_text(
            f"🆕 *{nombre}* (RUT: {rut}) no está en tu lista de proveedores.\n\n¿Lo agregamos?",
            parse_mode="Markdown", reply_markup=_proveedor_nuevo_keyboard())
        return
    await _guardar_excel(query, context, items, file_path)

async def cb_add_proveedor_yes(update, context):
    query = update.callback_query; await query.answer()
    nombre = context.user_data.get("nuevo_proveedor_nombre")
    rut    = context.user_data.get("nuevo_proveedor_rut")
    if nombre and rut:
        ok = await asyncio.to_thread(_agregar_proveedor, nombre, rut)
        if not ok: await query.answer("⚠️ No se pudo agregar", show_alert=True)
    await _guardar_excel(query, context,
        context.user_data.get("pending_items", []),
        context.user_data.get("pending_file_path"))

async def cb_add_proveedor_no(update, context):
    query = update.callback_query; await query.answer()
    await _guardar_excel(query, context,
        context.user_data.get("pending_items", []),
        context.user_data.get("pending_file_path"))

async def _guardar_excel(query, context, items, file_path):
    es_boleta = _es_boleta(items)
    await query.edit_message_text("💾 Guardando en el Excel…")
    try:
        if es_boleta:
            success = await asyncio.to_thread(append_boleta, items)
            # Mover archivo a carpeta de boletas
            if success and file_path and os.path.exists(file_path):
                import shutil
                dest = _save_path_boleta(os.path.basename(file_path))
                shutil.move(file_path, dest)
                file_path = dest
        else:
            success = await asyncio.to_thread(append_to_excel, items)
            # Intentar agregar insumos al inventario automáticamente
            if success:
                try:
                    agregados = await asyncio.to_thread(agregar_stock_desde_factura, items)
                    if agregados:
                        logger.info(f"Inventario: {len(agregados)} insumos agregados desde factura")
                except Exception as e:
                    logger.warning(f"Error auto-inventario: {e}")
    except Exception as e: logger.error(e); success = False
    if success:
        context.user_data["last_invoice_file"]   = file_path
        context.user_data["last_invoice_rows"]   = len(items)
        context.user_data["last_invoice_boleta"] = es_boleta
        context.user_data["pending_items"]       = []
        first     = items[0]
        proveedor = first.get("Nombre Factura / Proveedor") or "Desconocido"
        nro       = first.get("Numero Factura / Nro Documento") or "S/N"
        # Preferir Total Factura (grand total del documento) sobre suma de líneas
        total_factura_anchor = float(first.get("Total Factura") or 0)
        total = round(total_factura_anchor) if total_factura_anchor > 0 \
                else sum(float(i.get("Monto / TOTAL") or 0) for i in items)
        if es_boleta:
            info_caja = await asyncio.to_thread(consultar_saldo_caja)
            saldo_txt = f"💰 Saldo caja chica: *${info_caja['saldo']:,.0f}*"
            alerta = "\n⚠️ Saldo bajo!" if info_caja['saldo'] < 100000 else ""
            await query.edit_message_text(
                f"✅ *Boleta guardada (Caja Chica)*\n\n🏪 {_esc(proveedor)}\n🧾 Nº {_esc(nro)}  —  📤 ${total:,.0f}\n"
                f"{saldo_txt}{alerta}\n\n_Usa /deshacer si hay algún error_",
                parse_mode="Markdown")
        else:
            await query.edit_message_text(
                f"✅ *Factura guardada*\n\n🏢 {_esc(proveedor)}\n📄 Nº {_esc(nro)}  —  💰 ${total:,.0f}\n"
                f"📊 {len(items)} fila(s) en el Excel\n\n_Usa /deshacer si hay algún error_",
                parse_mode="Markdown")
    else:
        await query.edit_message_text("❌ Error al guardar. ¿Está el Excel abierto en otro programa?")

async def cb_cancel_save(update, context):
    query = update.callback_query; await query.answer()
    context.user_data["pending_items"] = []
    await query.edit_message_text("🚫 Factura descartada. Mándame otra cuando quieras.")

async def cb_edit_menu(update, context):
    query = update.callback_query; await query.answer()
    # Limpiar estado de edición pendiente al volver
    context.user_data["editing_field"] = None
    context.user_data["editing_item_idx"] = None
    items = context.user_data.get("pending_items", [])
    texto = f"✏️ *¿Qué campo quieres corregir?* ({len(items)} ítem{'s' if len(items) != 1 else ''})"
    await query.edit_message_text(texto,
        parse_mode="Markdown", reply_markup=_edit_keyboard(items))

async def cb_edit_field(update, context):
    query = update.callback_query; await query.answer()
    campo_key = query.data
    campo_excel, campo_label = CAMPOS_EDITABLES[campo_key]
    items = context.user_data.get("pending_items", [])

    # Si hay varios ítems y es un campo por línea, preguntar cuál ítem
    if len(items) > 1 and campo_key in CAMPOS_POR_ITEM:
        buttons = []
        for i, item in enumerate(items):
            glosa = _esc(item.get("Detalle / Glosa") or f"Ítem {i+1}")[:30]
            buttons.append([InlineKeyboardButton(
                f"Ítem {i+1}: {glosa}", callback_data=f"selitem_{i}_{campo_key}")])
        buttons.append([InlineKeyboardButton("« Volver", callback_data="edit_menu")])
        await query.edit_message_text(
            f"✏️ *{campo_label}* — ¿En qué ítem?",
            parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(buttons))
        return

    # Campo común o un solo ítem: editar en todos
    context.user_data["editing_field"]       = campo_excel
    context.user_data["editing_field_label"] = campo_label
    context.user_data["editing_item_idx"]    = None  # todos
    _back_btn = InlineKeyboardMarkup([[InlineKeyboardButton("« Volver", callback_data="edit_menu")]])
    await query.edit_message_text(
        f"✏️ Escribe el nuevo valor para *{campo_label}*:",
        parse_mode="Markdown", reply_markup=_back_btn)

async def cb_select_item(update, context):
    """Callback para selitem_{idx}_{campo_key}"""
    query = update.callback_query; await query.answer()
    parts = query.data.split("_")  # selitem_0_edit_glosa
    idx = int(parts[1])
    campo_key = "_".join(parts[2:])  # edit_glosa
    campo_excel, campo_label = CAMPOS_EDITABLES[campo_key]
    items = context.user_data.get("pending_items", [])
    glosa = _esc(items[idx].get("Detalle / Glosa") or f"Ítem {idx+1}")[:30]

    context.user_data["editing_field"]       = campo_excel
    context.user_data["editing_field_label"] = campo_label
    context.user_data["editing_item_idx"]    = idx
    _back_btn = InlineKeyboardMarkup([[InlineKeyboardButton("« Volver", callback_data="edit_menu")]])
    await query.edit_message_text(
        f"✏️ Escribe el nuevo valor para *{campo_label}* (Ítem {idx+1}: {glosa}):",
        parse_mode="Markdown", reply_markup=_back_btn)

async def cb_edit_total_factura(update, context):
    """Permite editar el total general de la factura."""
    query = update.callback_query; await query.answer()
    items = context.user_data.get("pending_items", [])
    # Mostrar total actual calculado
    total_neto = sum(float(i.get("Valor unitario") or 0) * float(i.get("Cantidad") or 1) for i in items)
    total_actual = round(total_neto * 1.19)
    context.user_data["editing_field"] = "_total_factura"
    context.user_data["editing_field_label"] = "💰 TOTAL FACTURA"
    context.user_data["editing_item_idx"] = None
    _back_btn = InlineKeyboardMarkup([[InlineKeyboardButton("« Volver", callback_data="edit_menu")]])
    await query.edit_message_text(
        f"💰 *Total factura actual:* ${total_actual:,.0f}\n\n"
        f"Escribe el nuevo *total con IVA* de toda la factura:",
        parse_mode="Markdown", reply_markup=_back_btn)

async def cb_medio_pago(update, context):
    """Callback para seleccionar medio de pago (Banco/Caja Chica)."""
    query = update.callback_query; await query.answer()
    medio = "Banco" if query.data == "pago_banco" else "Caja Chica"
    nro = context.user_data.get("pagado_nro")
    fecha = context.user_data.get("pagado_fecha")
    resultado = await asyncio.to_thread(registrar_pago, nro, fecha, medio)
    context.user_data["pagado_state"] = None
    context.user_data["pagado_nro"] = None
    context.user_data["pagado_fecha"] = None

    actualizadas = resultado.get("actualizadas", 0) if isinstance(resultado, dict) else resultado
    calce = resultado.get("calce") if isinstance(resultado, dict) else None

    if actualizadas > 0:
        icono = "🏦" if medio == "Banco" else "💵"
        texto = (
            f"✅ *Pago registrado*\n\n"
            f"📄 Factura Nº {_esc(nro)}\n"
            f"📅 Fecha: {fecha}\n"
            f"{icono} Medio: {medio}\n"
            f"📊 {actualizadas} fila(s) actualizada(s)")

        if calce:
            if calce["exacto"]:
                texto += (
                    f"\n\n🔗 *Calce bancario encontrado:*\n"
                    f"📅 {calce['fecha_banco']}\n"
                    f"📝 {_esc(calce['descripcion_banco'])}\n"
                    f"💰 ${calce['cargo_banco']:,.0f} ✅ Monto exacto")
            else:
                texto += (
                    f"\n\n🔗 *Posible calce bancario:*\n"
                    f"📅 {calce['fecha_banco']}\n"
                    f"📝 {_esc(calce['descripcion_banco'])}\n"
                    f"💰 ${calce['cargo_banco']:,.0f}\n"
                    f"⚠️ Calce parcial — verifica manualmente")
                if calce["total_candidatos"] > 1:
                    texto += f"\n📊 {calce['total_candidatos']} movimientos similares encontrados"
            # Guardar calce para verificación
            context.user_data["ultimo_calce"] = calce
            context.user_data["ultimo_calce_nro"] = nro
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Calce correcto", callback_data="calce_ok"),
                 InlineKeyboardButton("❌ No es este", callback_data="calce_no")]])
            try:
                await query.edit_message_text(texto, parse_mode="Markdown", reply_markup=kb)
            except Exception:
                await query.edit_message_text(texto, reply_markup=kb)
        else:
            if medio == "Banco":
                texto += "\n\n🔍 No se encontró calce bancario automático."
            try:
                await query.edit_message_text(texto, parse_mode="Markdown")
            except Exception:
                await query.edit_message_text(texto)
    else:
        await query.edit_message_text("❌ Error al actualizar el Excel. ¿Está abierto?")


async def cb_calce_verificacion(update, context):
    """Callback para confirmar o rechazar calce bancario."""
    query = update.callback_query; await query.answer()
    calce = context.user_data.get("ultimo_calce")
    nro = context.user_data.get("ultimo_calce_nro", "?")

    if query.data == "calce_ok":
        await query.edit_message_text(
            f"✅ *Calce confirmado*\n\n"
            f"📄 Factura Nº {_esc(nro)} vinculada a:\n"
            f"🏦 {_esc(calce['descripcion_banco'] if calce else '—')}\n"
            f"💰 ${calce['cargo_banco']:,.0f}" if calce else "✅ Calce confirmado",
            parse_mode="Markdown")
    else:
        await query.edit_message_text(
            f"📝 *Calce descartado*\n\n"
            f"Factura Nº {_esc(nro)} marcada como pagada.\n"
            f"El calce bancario no se aplicó — puedes revisarlo manualmente en el Excel.",
            parse_mode="Markdown")

    context.user_data["ultimo_calce"] = None
    context.user_data["ultimo_calce_nro"] = None

async def cb_back_preview(update, context):
    query = update.callback_query; await query.answer()
    await _show_preview(query, context)

async def cb_add_item(update, context):
    """Agrega un ítem vacío a la factura pendiente (cuando el OCR omite alguno)."""
    query = update.callback_query; await query.answer()
    items = context.user_data.get("pending_items", [])
    if not items:
        await query.edit_message_text("⚠️ No hay factura pendiente."); return
    # Copiar cabecera del primer ítem (datos comunes del documento)
    first = items[0]
    COMUNES = ("Fecha Emision", "Fecha Vencimiento", "Fecha Pago",
               "Nombre Factura / Proveedor", "Rut", "Documento",
               "Numero Factura / Nro Documento", "Referencia Factura",
               "Total Factura")
    nuevo = {k: first.get(k) for k in COMUNES}
    nuevo["Detalle / Glosa"] = ""
    nuevo["Glosa II"] = ""
    nuevo["Valor unitario"] = 0
    nuevo["Cantidad"] = 1
    nuevo["Impuesto Especifico"] = 0
    nuevo["Monto / TOTAL"] = 0
    items.append(nuevo)
    context.user_data["pending_items"] = items
    idx = len(items)
    # Preseleccionar edición de Glosa del nuevo ítem
    context.user_data["editing_field"] = "Detalle / Glosa"
    context.user_data["editing_field_label"] = "📦 Glosa / descripción corta"
    context.user_data["editing_item_idx"] = idx - 1
    _back_btn = InlineKeyboardMarkup([[InlineKeyboardButton("« Volver", callback_data="edit_menu")]])
    await query.edit_message_text(
        f"➕ *Ítem {idx} agregado*\n\n"
        f"Escribe la *glosa / descripción* del nuevo ítem\n"
        f"(luego podrás editar Cantidad, Unit neto y Total ítem con el menú de edición):",
        parse_mode="Markdown", reply_markup=_back_btn)

async def cb_del_item_menu(update, context):
    """Muestra lista de ítems para elegir cuál eliminar."""
    query = update.callback_query
    items = context.user_data.get("pending_items", [])
    if len(items) <= 1:
        await query.answer("⚠️ No puedes eliminar el único ítem.", show_alert=True); return
    await query.answer()
    buttons = []
    for i, item in enumerate(items):
        glosa = _esc(item.get("Detalle / Glosa") or f"Ítem {i+1}")[:30]
        buttons.append([InlineKeyboardButton(
            f"🗑️ Ítem {i+1}: {glosa}", callback_data=f"delitem_{i}")])
    buttons.append([InlineKeyboardButton("« Volver", callback_data="edit_menu")])
    await query.edit_message_text(
        "🗑️ *¿Qué ítem quieres eliminar?*",
        parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(buttons))

async def cb_del_item_confirm(update, context):
    """Elimina el ítem seleccionado y vuelve al preview."""
    query = update.callback_query; await query.answer()
    try:
        idx = int(query.data.split("_")[1])
    except (ValueError, IndexError):
        await query.edit_message_text("❌ Índice de ítem inválido."); return
    items = context.user_data.get("pending_items", [])
    if 0 <= idx < len(items):
        items.pop(idx)
        context.user_data["pending_items"] = items
        # Recalcular Total Factura = suma de Monto/TOTAL restantes
        nuevo_tf = round(sum(float(it.get("Monto / TOTAL") or 0) for it in items))
        if nuevo_tf > 0:
            for it in items:
                it["Total Factura"] = nuevo_tf
    await _show_preview(query, context)

async def handle_text_edit(update, context):
    # ── Flujo /deposito ──
    dep_state = context.user_data.get("deposito_state")
    if dep_state:
        texto = update.message.text.strip()
        if dep_state == "esperando_monto":
            try:
                n = texto.replace("$", "").replace(" ", "")
                if "," in n:
                    n = n.replace(".", "").replace(",", ".")
                elif n.count(".") > 1:
                    n = n.replace(".", "")
                monto = float(n)
            except ValueError:
                await update.message.reply_text("❌ No es un monto válido. Intenta de nuevo:"); return
            context.user_data["deposito_monto"] = monto
            context.user_data["deposito_state"] = "esperando_fecha_dep"
            await update.message.reply_text(
                f"💰 Monto: *${monto:,.0f}*\n\n"
                f"📅 Escribe la *fecha del depósito* (DD/MM/YYYY)\n"
                f"o escribe *hoy* para usar la fecha de hoy:",
                parse_mode="Markdown")
            return
        if dep_state == "esperando_fecha_dep":
            from dateutil import parser as date_parser
            if texto.lower() == "hoy":
                fecha = datetime.now().strftime("%Y-%m-%d")
            else:
                try:
                    if "/" in texto:
                        parts = texto.split("/")
                        if len(parts) == 3 and len(parts[0]) <= 2:
                            fecha = datetime.strptime(texto, "%d/%m/%Y").strftime("%Y-%m-%d")
                        else:
                            fecha = date_parser.parse(texto).strftime("%Y-%m-%d")
                    else:
                        fecha = date_parser.parse(texto).strftime("%Y-%m-%d")
                except Exception:
                    await update.message.reply_text("❌ Formato no válido. Usa DD/MM/YYYY o escribe *hoy*:", parse_mode="Markdown"); return
            monto = context.user_data.get("deposito_monto", 0)
            nuevo_saldo = await asyncio.to_thread(registrar_deposito_caja, fecha, monto)
            context.user_data["deposito_state"] = None
            context.user_data["deposito_monto"] = None
            if nuevo_saldo >= 0:
                await update.message.reply_text(
                    f"✅ *Depósito registrado*\n\n"
                    f"💰 Monto: ${monto:,.0f}\n"
                    f"📅 Fecha: {fecha}\n"
                    f"📊 *Nuevo saldo caja chica: ${nuevo_saldo:,.0f}*",
                    parse_mode="Markdown")
            else:
                await update.message.reply_text("❌ Error al registrar. ¿Está el Excel abierto?")
            return

    # ── Flujo /pagado ──
    pagado_state = context.user_data.get("pagado_state")
    if pagado_state:
        texto = update.message.text.strip()

        if pagado_state == "esperando_nro":
            # Buscar factura en Excel
            resultados = await asyncio.to_thread(buscar_factura, texto)
            if not resultados:
                await update.message.reply_text(
                    f"❌ No encontré factura Nº *{_esc(texto)}* en el Excel.\n"
                    "Verifica el número e intenta de nuevo, o escribe /cancelar",
                    parse_mode="Markdown")
                return
            # Mostrar lo encontrado
            context.user_data["pagado_nro"] = texto
            context.user_data["pagado_state"] = "esperando_fecha"
            r = resultados[0]
            ya_pagada = f"\n⚠️ *Ya tiene fecha de pago:* {r['fecha_pago']}" if r.get("fecha_pago") else ""
            await update.message.reply_text(
                f"📄 *Factura encontrada:*\n\n"
                f"🏢 {_esc(r.get('proveedor') or '—')}\n"
                f"🪪 RUT: {_esc(r.get('rut') or '—')}\n"
                f"📄 {_esc(r.get('documento') or '—')} Nº {_esc(r.get('nro_factura'))}\n"
                f"📦 {_esc(r.get('glosa') or '—')}\n"
                f"💰 Total: ${float(r.get('total') or 0):,.0f}\n"
                f"({len(resultados)} fila(s) en Excel){ya_pagada}\n\n"
                f"📅 Escribe la *fecha de pago* (DD/MM/YYYY o YYYY-MM-DD):",
                parse_mode="Markdown")
            return

        if pagado_state == "esperando_fecha":
            # Parsear fecha
            from dateutil import parser as date_parser
            try:
                if "/" in texto:
                    parts = texto.split("/")
                    if len(parts) == 3 and len(parts[0]) <= 2:
                        fecha = datetime.strptime(texto, "%d/%m/%Y").strftime("%Y-%m-%d")
                    else:
                        fecha = date_parser.parse(texto).strftime("%Y-%m-%d")
                else:
                    fecha = date_parser.parse(texto).strftime("%Y-%m-%d")
            except Exception:
                await update.message.reply_text(
                    "❌ No entendí la fecha. Escribe en formato *DD/MM/YYYY* (ej: 15/03/2026):",
                    parse_mode="Markdown")
                return

            context.user_data["pagado_fecha"] = fecha
            context.user_data["pagado_state"] = "esperando_medio"
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("🏦 Banco", callback_data="pago_banco"),
                 InlineKeyboardButton("💵 Caja Chica", callback_data="pago_caja_chica")]])
            await update.message.reply_text(
                f"📅 Fecha: *{fecha}*\n\n💳 *¿Medio de pago?*",
                parse_mode="Markdown", reply_markup=kb)
            return

    # ── Flujo /tarea ──
    tarea_state = context.user_data.get("tarea_state")
    if tarea_state:
        texto = update.message.text.strip()
        if tarea_state == "esperando_desc":
            context.user_data["tarea_desc"] = texto
            context.user_data["tarea_state"] = "esperando_prioridad"
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("🔴 Alta", callback_data="prior_Alta"),
                 InlineKeyboardButton("🟡 Media", callback_data="prior_Media"),
                 InlineKeyboardButton("🟢 Baja", callback_data="prior_Baja")]])
            await update.message.reply_text(
                f"📝 *Tarea:* {_esc(texto)}\n\n¿Qué *prioridad*?",
                parse_mode="Markdown", reply_markup=kb)
            return
        if tarea_state == "esperando_id_hecho":
            try:
                task_id = int(texto)
            except ValueError:
                await update.message.reply_text("❌ Escribe solo el número de tarea."); return
            context.user_data["tarea_state"] = "esperando_obs_hecho"
            context.user_data["tarea_id_hecho"] = task_id
            await update.message.reply_text(
                f"✅ Tarea *#{task_id}*\n\nEscribe una *observación* de lo que se hizo\n(o escribe *ok* para omitir):",
                parse_mode="Markdown")
            return
        if tarea_state == "esperando_obs_hecho":
            task_id = context.user_data.get("tarea_id_hecho", 0)
            obs = texto if texto.lower() != "ok" else ""
            ok = await asyncio.to_thread(actualizar_tarea, task_id, "Hecho", obs)
            context.user_data["tarea_state"] = None
            if ok:
                await update.message.reply_text(f"✅ Tarea *#{task_id}* completada.", parse_mode="Markdown")
            else:
                await update.message.reply_text(f"❌ No encontré tarea #{task_id}.")
            return

    # ── Flujo /bitacora ──
    bitacora_state = context.user_data.get("bitacora_state")
    if bitacora_state == "esperando_registro":
        texto = update.message.text.strip()
        await asyncio.to_thread(registrar_bitacora, texto)
        context.user_data["bitacora_state"] = None
        await update.message.reply_text(f"📓 *Registrado en bitácora:*\n{_esc(texto)}", parse_mode="Markdown")
        return

    # ── Flujo /uso (inventario) ──
    uso_state = context.user_data.get("uso_state")
    if uso_state:
        texto = update.message.text.strip()
        data = context.user_data.get("uso_data", {})
        if uso_state == "esperando_producto":
            data["producto"] = texto
            context.user_data["uso_data"] = data
            context.user_data["uso_state"] = "esperando_cantidad"
            await update.message.reply_text(
                f"🧪 *Producto:* {_esc(texto)}\n\nEscribe la *cantidad* utilizada (ej: 10):",
                parse_mode="Markdown")
            return
        if uso_state == "esperando_cantidad":
            try:
                n = texto.replace(",", ".")
                data["cantidad"] = float(n)
            except ValueError:
                await update.message.reply_text("❌ No es un número válido."); return
            context.user_data["uso_data"] = data
            context.user_data["uso_state"] = "esperando_sector"
            await update.message.reply_text(
                f"📍 ¿En qué *sector* del campo? (o escribe *-* para omitir):",
                parse_mode="Markdown")
            return
        if uso_state == "esperando_sector":
            data["sector"] = texto if texto != "-" else ""
            context.user_data["uso_data"] = data
            context.user_data["uso_state"] = "esperando_cultivo"
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("🌰 Nogales", callback_data="cult_Nogales"),
                 InlineKeyboardButton("🍒 Cerezos", callback_data="cult_Cerezos"),
                 InlineKeyboardButton("🌿 Avellanos", callback_data="cult_Avellanos")]])
            await update.message.reply_text(
                f"🌳 ¿En qué *cultivo*?", parse_mode="Markdown", reply_markup=kb)
            return

    # ── Flujo /vacaciones ──
    vac_state = context.user_data.get("vacacion_state")
    if vac_state:
        texto = update.message.text.strip()
        data = context.user_data.get("vacacion_data", {})
        if vac_state == "esperando_nombre":
            data["nombre"] = texto
            context.user_data["vacacion_data"] = data
            context.user_data["vacacion_state"] = "esperando_inicio"
            await update.message.reply_text(
                f"🏖️ *Trabajador:* {_esc(texto)}\n\n📅 Fecha *inicio* vacaciones (DD/MM/YYYY):",
                parse_mode="Markdown")
            return
        if vac_state == "esperando_inicio":
            from dateutil import parser as date_parser
            try:
                if "/" in texto:
                    parts = texto.split("/")
                    if len(parts) == 3 and len(parts[0]) <= 2:
                        fecha = datetime.strptime(texto, "%d/%m/%Y").strftime("%Y-%m-%d")
                    else:
                        fecha = date_parser.parse(texto).strftime("%Y-%m-%d")
                else:
                    fecha = date_parser.parse(texto).strftime("%Y-%m-%d")
            except Exception:
                await update.message.reply_text("❌ Formato no válido. Usa DD/MM/YYYY"); return
            data["inicio"] = fecha
            context.user_data["vacacion_data"] = data
            context.user_data["vacacion_state"] = "esperando_fin"
            await update.message.reply_text(
                f"📅 Inicio: *{fecha}*\n\n📅 Fecha *término* vacaciones (DD/MM/YYYY):",
                parse_mode="Markdown")
            return
        if vac_state == "esperando_fin":
            from dateutil import parser as date_parser
            try:
                if "/" in texto:
                    parts = texto.split("/")
                    if len(parts) == 3 and len(parts[0]) <= 2:
                        fecha = datetime.strptime(texto, "%d/%m/%Y").strftime("%Y-%m-%d")
                    else:
                        fecha = date_parser.parse(texto).strftime("%Y-%m-%d")
                else:
                    fecha = date_parser.parse(texto).strftime("%Y-%m-%d")
            except Exception:
                await update.message.reply_text("❌ Formato no válido. Usa DD/MM/YYYY"); return
            result = await asyncio.to_thread(
                registrar_vacacion, data.get("nombre", ""), data.get("inicio", ""), fecha)
            context.user_data["vacacion_state"] = None
            context.user_data["vacacion_data"] = {}
            await update.message.reply_text(
                f"🏖️ *Vacaciones registradas*\n\n"
                f"👤 {_esc(result['nombre'])}\n"
                f"📅 {result['inicio']} al {result['fin']}\n"
                f"📊 *{result['dias_habiles']} días hábiles* ({result['dias_corridos']} corridos)",
                parse_mode="Markdown")
            return

    # ── Flujo /agregar_trabajador ──
    trab_state = context.user_data.get("trabajador_state")
    if trab_state:
        texto = update.message.text.strip()
        if trab_state == "esperando_nombre":
            context.user_data["trabajador_nombre"] = texto
            context.user_data["trabajador_state"] = "esperando_cargo"
            await update.message.reply_text(
                f"👤 *{_esc(texto)}*\n\nEscribe el *cargo* (o *-* para omitir):",
                parse_mode="Markdown")
            return
        if trab_state == "esperando_cargo":
            nombre = context.user_data.get("trabajador_nombre", "")
            cargo = texto if texto != "-" else ""
            ok = await asyncio.to_thread(agregar_trabajador, nombre, "", cargo)
            context.user_data["trabajador_state"] = None
            if ok:
                await update.message.reply_text(
                    f"✅ *{_esc(nombre)}* agregado al personal.", parse_mode="Markdown")
            else:
                await update.message.reply_text(f"⚠️ *{_esc(nombre)}* ya existe.", parse_mode="Markdown")
            return

    # ── Edición de factura pendiente ──
    campo = context.user_data.get("editing_field")
    if not campo:
        # ── Chat inteligente (fallback) ──
        texto = update.message.text.strip()
        if len(texto) > 2:
            msg = await update.message.reply_text("💬 Pensando...")
            respuesta = await responder_chat(texto)
            try:
                await msg.edit_text(respuesta, parse_mode="Markdown")
            except Exception:
                await msg.edit_text(respuesta)
        return
    nuevo = update.message.text.strip()
    items = context.user_data.get("pending_items", [])
    if not items: await update.message.reply_text("⚠️ No hay factura pendiente."); return

    idx = context.user_data.get("editing_item_idx")  # None = todos, int = solo ese

    # Caso especial: editar total de factura
    if campo == "_total_factura":
        try:
            n = nuevo.replace("$", "").replace(" ", "")
            if "," in n:
                n = n.replace(".", "").replace(",", ".")
            elif n.count(".") > 1:
                n = n.replace(".", "")
            total_nuevo = float(n)
        except ValueError:
            await update.message.reply_text(f"❌ '{nuevo}' no es un número válido. Intenta de nuevo."); return
        # Registrar corrección del total factura
        valor_orig_tf = items[0].get("Total Factura") or items[0].get("Monto / TOTAL")
        _registrar_correccion(items[0], "Total Factura", valor_orig_tf, total_nuevo)
        # Detectar si es exenta (sin IVA)
        doc = str(items[0].get("Documento") or "").lower()
        exenta = any(k in doc for k in ("exenta", "exento", "no afecta", "no afecto"))
        iva_factor = 1.0 if exenta else 1.19

        # Cálculo inverso: Total → Neto → Unitario
        neto_nuevo = total_nuevo / iva_factor
        total_neto_actual = sum(float(i.get("Valor unitario") or 0) * float(i.get("Cantidad") or 1) for i in items)
        if total_neto_actual > 0:
            factor = neto_nuevo / total_neto_actual
            for item in items:
                unit = float(item.get("Valor unitario") or 0)
                qty = float(item.get("Cantidad") or 1)
                nuevo_unit = round(unit * factor, 2)
                item["Valor unitario"] = nuevo_unit
                item["Monto / TOTAL"] = round(nuevo_unit * qty * iva_factor)
        else:
            # No hay unitarios: poner todo el total
            items[0]["Monto / TOTAL"] = total_nuevo
            items[0]["Valor unitario"] = round(neto_nuevo)
        # Actualizar Total Factura en todos los ítems para que _build_preview lo refleje
        for item in items:
            item["Total Factura"] = round(total_nuevo)
        context.user_data["total_override"] = total_nuevo
        context.user_data["editing_field"] = None
        context.user_data["editing_item_idx"] = None
        await update.message.reply_text(f"✅ *Total factura* actualizado a `${total_nuevo:,.0f}`", parse_mode="Markdown")
        try:
            await update.message.reply_text(_build_preview(items), parse_mode="Markdown", reply_markup=_main_keyboard())
        except Exception:
            await update.message.reply_text(_build_preview(items), reply_markup=_main_keyboard())
        return

    targets = [items[idx]] if idx is not None else items

    NUMERICOS = {"Valor unitario", "Cantidad", "Monto / TOTAL", "TOTAL NETO"}
    for item in targets:
        valor_original = item.get(campo)  # capturar antes de sobrescribir
        if campo in NUMERICOS:
            try:
                # Formato chileno: punto = miles, coma = decimal
                # "7.164,32" → 7164.32 | "105,612" → 105.612 | "500000" → 500000
                n = nuevo.replace("$", "").replace(" ", "")
                if "," in n:
                    n = n.replace(".", "").replace(",", ".")
                else:
                    if n.count(".") > 1:
                        n = n.replace(".", "")
                val = float(n)
                item[campo] = val
                if campo == "Monto / TOTAL":
                    # Cálculo inverso: Total ítem (con IVA) → Valor unitario
                    qty = float(item.get("Cantidad") or 1)
                    imp_esp = float(item.get("Impuesto Especifico") or 0)
                    doc = str(item.get("Documento") or "").lower()
                    sin_iva = any(k in doc for k in ("exenta", "exento", "no afecta", "no afecto", "boleta de honorario"))
                    iva_factor = 1.0 if sin_iva else 1.19
                    neto_nuevo = (val - imp_esp) / iva_factor
                    item["Valor unitario"] = neto_nuevo / qty if qty else 0
                elif campo == "TOTAL NETO":
                    # El usuario ingresó el neto de línea (precio × cantidad, sin IVA)
                    qty = float(item.get("Cantidad") or 1)
                    imp_esp = float(item.get("Impuesto Especifico") or 0)
                    doc = str(item.get("Documento") or "").lower()
                    sin_iva = any(k in doc for k in ("exenta", "exento", "no afecta", "no afecto", "boleta de honorario"))
                    iva_factor = 1.0 if sin_iva else 1.19
                    item["Valor unitario"] = val / qty if qty else 0
                    item["Monto / TOTAL"] = round(val * iva_factor + imp_esp)
            except ValueError:
                await update.message.reply_text(f"❌ '{nuevo}' no es un número válido. Intenta de nuevo."); return
        else:
            item[campo] = nuevo
        # Registrar corrección solo si el valor cambió
        if str(valor_original) != str(item.get(campo)):
            _registrar_correccion(item, campo, valor_original, item.get(campo))

    # Si fue edición numérica, propagar a Monto/TOTAL y Total Factura para que el preview calce
    if campo in NUMERICOS:
        doc0 = str(items[0].get("Documento") or "").lower()
        sin_iva0 = any(k in doc0 for k in ("exenta", "exento", "no afecta", "no afecto", "boleta de honorario"))
        iva_factor0 = 1.0 if sin_iva0 else 1.19
        # Si el usuario editó Valor unitario o Cantidad del/los ítem(s), refrescar su Monto/TOTAL
        if campo not in ("Monto / TOTAL", "TOTAL NETO"):
            for it in targets:
                u = float(it.get("Valor unitario") or 0)
                q = float(it.get("Cantidad") or 1)
                imp = float(it.get("Impuesto Especifico") or 0)
                it["Monto / TOTAL"] = round(u * q * iva_factor0 + imp)
        # Sumar Monto/TOTAL de todos los ítems → nuevo Total Factura
        nuevo_tf = round(sum(float(it.get("Monto / TOTAL") or 0) for it in items))
        if nuevo_tf > 0:
            for it in items:
                it["Total Factura"] = nuevo_tf

    context.user_data["editing_field"] = None
    context.user_data["editing_item_idx"] = None
    label = context.user_data.get("editing_field_label", campo)
    suffix = f" (Ítem {idx+1})" if idx is not None else ""
    await update.message.reply_text(f"✅ *{label}*{suffix} actualizado.", parse_mode="Markdown")
    try:
        await update.message.reply_text(_build_preview(items), parse_mode="Markdown", reply_markup=_main_keyboard())
    except Exception:
        await update.message.reply_text(_build_preview(items), reply_markup=_main_keyboard())

# ── ARRANQUE ──────────────────────────────────
def main():
    if not TELEGRAM_TOKEN:
        logger.error("TELEGRAM_TOKEN no configurado"); return
    app = (ApplicationBuilder().token(TELEGRAM_TOKEN)
           .read_timeout(120).write_timeout(120).connect_timeout(120).pool_timeout(120).build())

    # Comandos
    app.add_handler(CommandHandler("start",    cmd_start))
    app.add_handler(CommandHandler("ayuda",    cmd_ayuda))
    app.add_handler(CommandHandler("deshacer", cmd_deshacer))
    app.add_handler(CommandHandler("pagado",   cmd_pagado))
    app.add_handler(CommandHandler("deposito", cmd_deposito))
    app.add_handler(CommandHandler("saldo",    cmd_saldo))
    app.add_handler(CommandHandler("reporte",  cmd_reporte))
    app.add_handler(CommandHandler("dashboard", cmd_dashboard))
    app.add_handler(CommandHandler("banco",    cmd_banco))
    app.add_handler(CommandHandler("cancelar", cmd_cancelar))
    # Cash Flow (Fase 1)
    from handlers.cash_flow_cmds import cmd_proyeccion, cmd_categoria
    app.add_handler(CommandHandler("proyeccion", cmd_proyeccion))
    app.add_handler(CommandHandler("categoria",  cmd_categoria))
    # Tareas y Bitácora
    app.add_handler(CommandHandler("tarea",    cmd_tarea))
    app.add_handler(CommandHandler("tareas",   cmd_tareas))
    app.add_handler(CommandHandler("hecho",    cmd_hecho))
    app.add_handler(CommandHandler("bitacora", cmd_bitacora))
    # Inventario
    app.add_handler(CommandHandler("inventario", cmd_inventario))
    app.add_handler(CommandHandler("uso",      cmd_uso))
    # Personal y Vacaciones
    app.add_handler(CommandHandler("personal",  cmd_personal))
    app.add_handler(CommandHandler("vacaciones", cmd_vacaciones))
    app.add_handler(CommandHandler("agregar_trabajador", cmd_agregar_trabajador))
    # Mensajes
    app.add_handler(MessageHandler(filters.Document.ALL,            handle_document))
    app.add_handler(MessageHandler(filters.PHOTO,                   handle_photo))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text_edit))
    # Callbacks
    app.add_handler(CallbackQueryHandler(cb_confirm_save,      pattern="^confirm_save$"))
    app.add_handler(CallbackQueryHandler(cb_cancel_save,       pattern="^cancel_save$"))
    app.add_handler(CallbackQueryHandler(cb_edit_menu,         pattern="^edit_menu$"))
    app.add_handler(CallbackQueryHandler(cb_back_preview,      pattern="^back_preview$"))
    app.add_handler(CallbackQueryHandler(cb_add_proveedor_yes, pattern="^add_proveedor_yes$"))
    app.add_handler(CallbackQueryHandler(cb_add_proveedor_no,  pattern="^add_proveedor_no$"))
    app.add_handler(CallbackQueryHandler(cb_add_item,          pattern="^add_item$"))
    app.add_handler(CallbackQueryHandler(cb_del_item_menu,     pattern="^del_item$"))
    app.add_handler(CallbackQueryHandler(cb_del_item_confirm,  pattern="^delitem_"))
    app.add_handler(CallbackQueryHandler(cb_edit_total_factura, pattern="^edit_total_factura$"))
    app.add_handler(CallbackQueryHandler(cb_medio_pago,        pattern="^pago_"))
    app.add_handler(CallbackQueryHandler(cb_calce_verificacion, pattern="^calce_"))
    app.add_handler(CallbackQueryHandler(cb_reporte,           pattern="^rep_"))
    app.add_handler(CallbackQueryHandler(cb_tarea_prioridad,   pattern="^prior_"))
    app.add_handler(CallbackQueryHandler(cb_cultivo,           pattern="^cult_"))
    app.add_handler(CallbackQueryHandler(cb_select_item,       pattern="^selitem_"))
    app.add_handler(CallbackQueryHandler(cb_edit_field,        pattern="^edit_"))

    # Crear hojas si no existen
    try: crear_hoja_banco()
    except: pass
    try: crear_hojas_tareas()
    except: pass
    try: crear_hojas_inventario()
    except: pass
    try: crear_hojas_vacaciones()
    except: pass

    # Programar sincronización bancaria automática: 8:00 y 18:00 hora Chile
    try:
        from zoneinfo import ZoneInfo
    except ImportError:
        from backports.zoneinfo import ZoneInfo
    tz_chile = ZoneInfo("America/Santiago")
    app.job_queue.run_daily(job_sync_banco, time=dtime(hour=8,  minute=0, tzinfo=tz_chile), name="banco_8am")
    app.job_queue.run_daily(job_sync_banco, time=dtime(hour=18, minute=0, tzinfo=tz_chile), name="banco_6pm")
    # Actualizar días de vacaciones el 1ro de cada mes
    app.job_queue.run_monthly(job_vacaciones_mensuales, when=dtime(hour=7, minute=0, tzinfo=tz_chile),
                              day=1, name="vacaciones_mensuales")
    # Resumen semanal cash flow: lunes 08:00
    from handlers.cash_flow_jobs import job_resumen_semanal
    app.job_queue.run_daily(job_resumen_semanal,
                              time=dtime(hour=8, minute=0, tzinfo=tz_chile),
                              days=(0,), name="resumen_semanal")
    logger.info("⏰ Jobs programados: banco 08:00/18:00, vacaciones día 1, resumen lunes 08:00")

    logger.info("✅ Bot iniciado. Esperando mensajes...")
    app.run_polling()

if __name__ == "__main__":
    main()
