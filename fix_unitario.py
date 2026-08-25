"""
fix_unitario.py  —  Ejecutar UNA SOLA VEZ

Corrige todas las filas existentes en la hoja "Facturas":
  1. Lee el valor actual de TOTAL NETO (col L) y Cantidad (col K).
  2. Calcula  Valor unitario = TOTAL NETO / Cantidad  y lo escribe como número en col J.
  3. Reemplaza col L con la fórmula dinámica  =J{fila}*K{fila}

Así queda igual al comportamiento del bot actualizado:
  - J = valor numérico (fuente de verdad)
  - L = fórmula  =J*K  (se recalcula si se edita J o K en Excel)
"""

from openpyxl import load_workbook

EXCEL_PATH = r"C:\Users\Windows\Desktop\Workflow\Agricola Santa Elisa\MASTER Agricola Santa Elisa.xlsx"
SHEET_NAME = "Facturas"

COL_UNITARIO = 10   # J  — Valor unitario
COL_CANTIDAD = 11   # K  — Cantidad
COL_NETO     = 12   # L  — TOTAL NETO

wb = load_workbook(EXCEL_PATH)
ws = wb[SHEET_NAME]

fixed = 0

for row in range(2, ws.max_row + 1):
    cantidad_raw = ws.cell(row=row, column=COL_CANTIDAD).value
    neto_raw     = ws.cell(row=row, column=COL_NETO).value

    # Saltar filas completamente vacías
    if cantidad_raw is None and neto_raw is None:
        continue

    try:
        qty  = float(cantidad_raw or 0)
        neto = float(neto_raw or 0)
    except (TypeError, ValueError):
        print(f"  Fila {row}: no se pudo convertir cantidad={cantidad_raw!r} o neto={neto_raw!r}, se omite.")
        continue

    # 1. Escribir Valor unitario como número
    cell_j = ws.cell(row=row, column=COL_UNITARIO)
    cell_j.value = round(neto / qty, 6) if qty else 0
    cell_j.number_format = '#,##0.000'

    # 2. Reemplazar TOTAL NETO con fórmula dinámica
    cell_l = ws.cell(row=row, column=COL_NETO)
    cell_l.value = f"=J{row}*K{row}"
    cell_l.number_format = '#,##0'

    fixed += 1

wb.save(EXCEL_PATH)
print(f"\nOK — {fixed} filas corregidas.")
print("  Col J (Valor unitario): ahora es un valor numérico.")
print("  Col L (TOTAL NETO):     ahora es fórmula =J*K")
