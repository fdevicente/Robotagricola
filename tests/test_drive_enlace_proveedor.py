# -*- coding: utf-8 -*-
"""El enlace tiene que caer en la fila del PROVEEDOR correcto, no solo del número.

BUG REAL, medido contra el Master el 26-ago-2026: hay 19 números de factura que
usan dos o tres proveedores distintos (el Nº1034 lo usan AGROAVELLANO y LUCIA
GARCIA; el Nº264, ECOSMART e INV. SANTA VICTORIA). `guardar_enlace` filtraba
SOLO por la columna 7, así que el enlace de un proveedor terminaba en la fila
del otro. Y como no pisa lo ya escrito, el primero en subir ganaba y el segundo
se quedaba sin enlace. 21 de los 826 archivos a migrar caen en este caso.

Perder un enlace se repara con otra pasada; escribirlo en la factura equivocada
es dato malo en el Master, en silencio.

OJO: ruta EXPLÍCITA al Excel en todos los tests. Un test destruyó el Master
real por confiar en el default de _save_wb.
"""
import openpyxl
import pytest

from modules.drive.enlaces import guardar_enlace

COL_PROV, COL_NUM, COL_DRIVE = 4, 7, 22


def _libro(tmp_path, filas):
    """filas = [(proveedor, numero), ...]"""
    ruta = tmp_path / "master.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"
    ws.append([None] * 21 + ["Archivo Drive"])
    for prov, num in filas:
        f = [None] * 22
        f[COL_PROV - 1] = prov
        f[COL_NUM - 1] = num
        ws.append(f)
    wb.save(ruta)
    wb.close()
    return str(ruta)


def _drive(ruta, fila):
    wb = openpyxl.load_workbook(ruta)
    v = wb["Facturas"].cell(fila, COL_DRIVE).value
    wb.close()
    return v


def test_no_enlaza_la_factura_de_otro_proveedor_con_el_mismo_numero(tmp_path):
    """El caso exacto del Nº264: ECOSMART e INV. SANTA VICTORIA."""
    ruta = _libro(tmp_path, [("ECOSMART", 264), ("INV. SANTA VICTORIA", 264)])

    guardar_enlace(ruta, "264", "abc123", proveedor="ECOSMART")

    assert "abc123" in str(_drive(ruta, 2))          # la de ECOSMART
    assert _drive(ruta, 3) is None                    # la otra, intacta


def test_si_el_proveedor_no_calza_no_escribe_nada(tmp_path):
    """Mejor sin enlace que con el enlace equivocado."""
    ruta = _libro(tmp_path, [("ECOSMART", 264)])

    assert guardar_enlace(ruta, "264", "abc123",
                          proveedor="UN PROVEEDOR QUE NO ESTA") is False
    assert _drive(ruta, 2) is None


def test_sin_proveedor_sigue_enlazando_solo_por_numero(tmp_path):
    """Los llamados viejos no cambian de comportamiento."""
    ruta = _libro(tmp_path, [("ECOSMART", 264)])

    assert guardar_enlace(ruta, "264", "abc123") is True
    assert "abc123" in str(_drive(ruta, 2))


def test_el_proveedor_calza_sin_distinguir_mayusculas(tmp_path):
    """En el Master está 'Easy' y el archivo dice 'EASY'."""
    ruta = _libro(tmp_path, [("Easy", 34809729)])

    assert guardar_enlace(ruta, "34809729", "abc123", proveedor="EASY") is True
    assert "abc123" in str(_drive(ruta, 2))


def test_el_proveedor_calza_con_espacios_o_guiones_bajos(tmp_path):
    """El nombre del archivo trae '_' donde el Master tiene espacios."""
    ruta = _libro(tmp_path, [("Silpa Sur Spa", 952)])

    assert guardar_enlace(ruta, "952", "abc",
                          proveedor="Silpa_Sur_Spa") is True
    assert "abc" in str(_drive(ruta, 2))


def test_enlaza_todas_las_lineas_del_proveedor_correcto(tmp_path):
    """Una factura de varios ítems son varias filas; la del otro no se toca."""
    ruta = _libro(tmp_path, [("ECOSMART", 264), ("ECOSMART", 264),
                             ("INV. SANTA VICTORIA", 264)])

    guardar_enlace(ruta, "264", "abc123", proveedor="ECOSMART")

    assert "abc123" in str(_drive(ruta, 2))
    assert "abc123" in str(_drive(ruta, 3))
    assert _drive(ruta, 4) is None
