# -*- coding: utf-8 -*-
"""La caja y la proyección tienen que usar EL MISMO tipo de cambio.

El valor vivía en dos lados que podían separarse:

  hoja `Config` del Master  -> la lee `cuentas.caja_total()`, pega AL TOQUE
  `CASH_FLOW_CONFIG` de config.py -> la leía el proyector, solo al REINICIAR

O sea: editabas la hoja `Config` y el dashboard pasaba a mostrar la caja a un
tipo de cambio y los ingresos proyectados a otro, sin que nada avisara. Y
`config.py` se lee al importar, así que el proyector se quedaba con el valor
viejo hasta el próximo reinicio del bot.

Al 1-sep-2026 los dos decían 910, así que no había daño hecho — el bug era
latente y se abría en cuanto alguien tocara la hoja. Ahora el proyector lee la
misma fuente que la caja, y config.py queda solo como respaldo.

OJO: todos los tests que escriben Excel pasan la ruta EXPLICITA. Un test
destruyo el Master real por confiar en el default de _save_wb.
"""
import openpyxl
import pytest

from modules.cash_flow.projector import load_expected_ingresos
from modules.cuentas import tipo_cambio

HEADERS = ["Año", "Cultivo", "Kg total", "Exportadora", "Kg asignados",
           "Precio USD/kg", "N° cuotas", "Cuota #", "Fecha estimada",
           "Monto USD estimado", "Tipo cuota", "Estado", "Fecha real recibido",
           "Monto real recibido", "Moneda recibida", "Notas", "Aplica IVA"]

# una cuota de US$1.000 en mayo-2027, sin IVA, para que el monto sea el
# tipo de cambio puro y se lea de un vistazo
FILA = [2027, "NOGALES", 1000, "X", 1000, 1.0, 1, 1, "2027-05-15", 1000,
        "adelanto", "esperado", None, None, None, "", "NO"]


def _libro(tmp_path, usd_clp=None):
    """Master de juguete. `usd_clp=None` deja la hoja Config sin el valor."""
    ruta = tmp_path / "master.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cosechas"
    ws.append(HEADERS)
    ws.append(FILA)
    cfg = wb.create_sheet("Config")
    cfg.append(["clave", "valor"])
    if usd_clp is not None:
        cfg.append(["usd_clp_estimado", usd_clp])
    wb.save(ruta)
    wb.close()
    return str(ruta)


def test_el_proyector_usa_el_valor_de_la_hoja_Config(tmp_path):
    """El caso del bug: la hoja dice 950 y config.py dice otra cosa."""
    ruta = _libro(tmp_path, 950)
    ingresos = load_expected_ingresos(ruta)
    assert len(ingresos) == 1
    assert ingresos[0]["monto_clp"] == pytest.approx(1000 * 950)


def test_si_la_hoja_no_lo_trae_cae_en_config_py(tmp_path):
    from config import CASH_FLOW_CONFIG
    ruta = _libro(tmp_path, None)
    ingresos = load_expected_ingresos(ruta)
    esperado = 1000 * CASH_FLOW_CONFIG.get("usd_clp_estimado", 1000)
    assert ingresos[0]["monto_clp"] == pytest.approx(esperado)


def test_la_caja_y_la_proyeccion_leen_LO_MISMO(tmp_path):
    """El test que impide que vuelvan a separarse."""
    ruta = _libro(tmp_path, 877)
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    try:
        de_la_caja = tipo_cambio(wb)
    finally:
        wb.close()
    de_la_proyeccion = load_expected_ingresos(ruta)[0]["monto_clp"] / 1000
    assert de_la_caja == pytest.approx(de_la_proyeccion) == 877


def test_un_valor_ilegible_en_la_hoja_no_rompe(tmp_path):
    """Si alguien escribe 'novecientos', se cae al respaldo en vez de reventar."""
    from config import CASH_FLOW_CONFIG
    ruta = _libro(tmp_path, "novecientos")
    ingresos = load_expected_ingresos(ruta)
    esperado = 1000 * CASH_FLOW_CONFIG.get("usd_clp_estimado", 1000)
    assert ingresos[0]["monto_clp"] == pytest.approx(esperado)


def test_el_proyector_ya_no_lee_CASH_FLOW_CONFIG_directo():
    """Leerlo directo es justo lo que permitia que divergieran."""
    import re
    from pathlib import Path
    src = Path(__file__).resolve().parent.parent / "modules" / "cash_flow" / "projector.py"
    texto = src.read_text(encoding="utf-8")
    cuerpo = texto[texto.index("def load_expected_ingresos"):]
    cuerpo = cuerpo[:cuerpo.index("\ndef ")] if "\ndef " in cuerpo else cuerpo
    assert not re.search(r"CASH_FLOW_CONFIG.*usd_clp_estimado", cuerpo), \
        "load_expected_ingresos volvio a leer CASH_FLOW_CONFIG directo"
