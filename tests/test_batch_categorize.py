import os, shutil, pytest
from unittest.mock import patch
from openpyxl import load_workbook
from excel_manager import SHEET_NAME, COL_CATEGORIA, ensure_new_columns


@pytest.fixture
def test_master(tmp_path):
    src = os.path.join(os.path.dirname(__file__), "..", "..",
                        "MASTER Agricola Santa Elisa.xlsx")
    dst = tmp_path / "test_master.xlsx"
    shutil.copy2(src, dst)
    ensure_new_columns(str(dst))
    return str(dst)


def test_batch_skips_already_categorized(test_master, tmp_path):
    from modules.cash_flow.categorizer import batch_categorize_history
    wb = load_workbook(test_master)
    ws = wb[SHEET_NAME]
    for r in (2, 3):
        ws.cell(r, COL_CATEGORIA, "Fertilizantes")
    wb.save(test_master)
    wb.close()

    cache = str(tmp_path / "c.json")
    with patch("modules.cash_flow.categorizer.categorize_raw",
                return_value={"categoria": "Riego", "cultivo": "GENERAL",
                              "confianza": 0.9, "razon": ""}):
        report = batch_categorize_history(excel_path=test_master, limit=5,
                                            cache_path=cache)

    assert report["skipped"] >= 2


def test_batch_limit_respected(test_master, tmp_path):
    from modules.cash_flow.categorizer import batch_categorize_history
    cache = str(tmp_path / "c.json")
    with patch("modules.cash_flow.categorizer.categorize_raw",
                return_value={"categoria": "Riego", "cultivo": "GENERAL",
                              "confianza": 0.9, "razon": ""}):
        report = batch_categorize_history(excel_path=test_master, limit=3,
                                            cache_path=cache)
    assert report["processed"] <= 3


def test_batch_low_confidence_count(test_master, tmp_path):
    from modules.cash_flow.categorizer import batch_categorize_history
    cache = str(tmp_path / "c.json")
    with patch("modules.cash_flow.categorizer.categorize_raw",
                return_value={"categoria": "Riego", "cultivo": "GENERAL",
                              "confianza": 0.5, "razon": ""}):
        report = batch_categorize_history(excel_path=test_master, limit=3,
                                            cache_path=cache)
    assert report["low_confidence"] >= 1
