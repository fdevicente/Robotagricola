import os, shutil, pytest
from openpyxl import load_workbook
from excel_manager import (
    SHEET_NAME, CUENTA_BANCO_SHEET, ensure_new_columns,
)


@pytest.fixture
def m(tmp_path):
    src = os.path.join(os.path.dirname(__file__), "..", "..",
                        "MASTER Agricola Santa Elisa.xlsx")
    dst = tmp_path / "test.xlsx"
    shutil.copy2(src, dst)
    ensure_new_columns(str(dst))
    return str(dst)


def test_orchestrator_auto_matches_clear_case(m):
    from modules.cash_flow.matcher import match_new_bank_movements
    wb = load_workbook(m)
    ws_f = wb[SHEET_NAME]
    # Limpiar fila 2 y agregar datos predecibles
    ws_f.cell(2, 1).value = "2025-09-01"
    ws_f.cell(2, 3).value = None  # sin fecha pago
    ws_f.cell(2, 4).value = "PROVEEDOR_TEST_UNICO"
    ws_f.cell(2, 7).value = "FAC-TEST-999"
    ws_f.cell(2, 15).value = 1234567

    # Sobrescribir fila 2 del banco (en el primer lote del scan)
    ws_b = wb[CUENTA_BANCO_SHEET]
    br = 2
    ws_b.cell(br, 1).value = "2025-09-03"
    ws_b.cell(br, 2).value = "PAGO PROVEEDOR_TEST_UNICO FAC-TEST-999"
    ws_b.cell(br, 3).value = ""
    ws_b.cell(br, 4).value = 1234567
    ws_b.cell(br, 5).value = None
    # Limpiar Factura_linkeada (col J = 10) por si tenia algo
    ws_b.cell(br, 10).value = None
    ws_b.cell(br, 7).value = None  # Tipo
    wb.save(m); wb.close()

    report = match_new_bank_movements(excel_path=m, limit=50)
    assert report["auto_matched"] >= 1


def test_orchestrator_reports_categories(m):
    from modules.cash_flow.matcher import match_new_bank_movements
    report = match_new_bank_movements(excel_path=m, limit=10)
    for k in ("scanned", "auto_matched", "ambiguous", "no_match"):
        assert k in report
