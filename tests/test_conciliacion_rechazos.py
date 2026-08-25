"""Fase 3 del conciliador: rechazos y motivo legible de cada sugerencia.

Sin el registro de rechazos, el conciliador vuelve a proponer el mismo par en
cada corrida y la vista se llena de ruido que el usuario ya descartó.
"""
from datetime import date

import pytest
from openpyxl import Workbook

import modules.conciliacion_rechazos as rech
from modules.conciliador import MOTIVOS, explicar


@pytest.fixture
def master(tmp_path):
    """Master de prueba. Se pasa SIEMPRE explícito: nunca tocar el real."""
    wb = Workbook()
    wb.active.title = "Cuenta Banco"
    p = tmp_path / "master.xlsx"
    wb.save(p); wb.close()
    return str(p)


def _par(fila=100, nro="F123", prov="Copeval", monto=500_000):
    return {"fila_banco": fila, "fecha_mov": date(2026, 7, 1),
            "desc_mov": "TEF Copeval", "monto_mov": monto,
            "nro": nro, "prov": prov, "total": monto, "criterio": "monto+proveedor"}


# ── Clave del par ────────────────────────────────────────────────────────

def test_la_clave_ignora_mayusculas_y_espacios():
    assert rech.clave(10, "f123", "  copeval  ") == rech.clave(10, "F123", "COPEVAL")


def test_la_clave_distingue_proveedores_con_el_mismo_numero():
    """Dos proveedores repiten numeración de facturas: ya nos mordió antes."""
    assert rech.clave(10, "F94", "Misael") != rech.clave(10, "F94", "Contreras")


def test_la_clave_distingue_movimientos_distintos():
    assert rech.clave(10, "F1", "X") != rech.clave(11, "F1", "X")


# ── Registro ─────────────────────────────────────────────────────────────

def test_registrar_y_recordar(master):
    assert rech.registrar([_par()], excel_path=master) == 1
    assert rech.clave(100, "F123", "Copeval") in rech.rechazados(master)


def test_no_duplica_el_mismo_par(master):
    rech.registrar([_par()], excel_path=master)
    assert rech.registrar([_par()], excel_path=master) == 0
    assert len(rech.listar(master)) == 1


def test_registra_varios_de_una_vez(master):
    n = rech.registrar([_par(nro="F1"), _par(nro="F2"), _par(nro="F3")],
                       excel_path=master)
    assert n == 3
    assert len(rech.rechazados(master)) == 3


def test_sin_hoja_no_hay_rechazos(master):
    assert rech.rechazados(master) == set()
    assert rech.listar(master) == []


def test_deshacer_devuelve_la_sugerencia(master):
    rech.registrar([_par()], excel_path=master)
    rid = rech.listar(master)[0]["id"]
    assert rech.deshacer(rid, excel_path=master) is True
    assert rech.rechazados(master) == set()


def test_deshacer_un_id_inexistente_no_revienta(master):
    rech.registrar([_par()], excel_path=master)
    assert rech.deshacer(9999, excel_path=master) is False
    assert len(rech.listar(master)) == 1


def test_listar_trae_los_datos_para_mostrarlos(master):
    rech.registrar([_par(monto=1_234_567)], usuario="admin",
                   motivo="No es este pago", excel_path=master)
    r = rech.listar(master)[0]
    assert r["monto_mov"] == 1_234_567
    assert r["prov"] == "Copeval"
    assert r["usuario"] == "admin"
    assert r["motivo"] == "No es este pago"


def test_crear_hoja_es_idempotente(master):
    rech.crear_hoja(excel_path=master)
    rech.crear_hoja(excel_path=master)
    rech.registrar([_par()], excel_path=master)
    assert len(rech.listar(master)) == 1


def test_escribir_en_otro_archivo_no_toca_el_master_real(master, tmp_path):
    """Blindaje: `_save_wb` resolvía el Master real como default, así que un
    guardado sin path explícito lo sobrescribía con el libro equivocado.
    Este test lo destruyó una vez de verdad — que no vuelva a pasar.
    """
    from openpyxl import load_workbook

    from config import EXCEL_PATH
    from excel_manager import _save_wb

    antes = load_workbook(EXCEL_PATH, read_only=True)
    hojas_antes = set(antes.sheetnames)
    antes.close()

    wb = load_workbook(master)
    _save_wb(wb, master)          # destino explícito
    wb.close()

    despues = load_workbook(EXCEL_PATH, read_only=True)
    hojas_despues = set(despues.sheetnames)
    despues.close()
    assert hojas_despues == hojas_antes
    assert len(hojas_despues) > 5, "el Master real quedó vacío"


# ── Motivo legible ───────────────────────────────────────────────────────

@pytest.mark.parametrize("criterio", sorted(MOTIVOS))
def test_cada_criterio_tiene_explicacion_en_castellano(criterio):
    txt = explicar(criterio)
    assert txt and txt != criterio
    assert "+" not in txt          # no se filtra el nombre técnico


def test_explica_la_confianza_de_la_ia():
    txt = explicar("IA (92%)")
    assert "92%" in txt and "IA" in txt


def test_criterio_desconocido_no_deja_la_tarjeta_muda():
    assert explicar("") == "Sugerencia automática"
    assert explicar("algo-nuevo") == "algo-nuevo"
