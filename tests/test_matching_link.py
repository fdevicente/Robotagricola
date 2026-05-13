import os, shutil, pytest
from openpyxl import load_workbook
from excel_manager import (
    SHEET_NAME, CUENTA_BANCO_SHEET, ensure_new_columns,
    COL_BANCO_TIPO, COL_BANCO_FACTURA_LINK,
    apply_bank_factura_link,
)


@pytest.fixture
def m(tmp_path):
    src = os.path.join(os.path.dirname(__file__), "..", "..",
                        "MASTER Agricola Santa Elisa.xlsx")
    dst = tmp_path / "test.xlsx"
    shutil.copy2(src, dst)
    ensure_new_columns(str(dst))
    wb = load_workbook(str(dst))
    # Limpiar Fecha Pago de fila 2 para test pendiente
    wb[SHEET_NAME].cell(2, 3).value = None
    ws = wb[CUENTA_BANCO_SHEET]
    r = ws.max_row + 1
    ws.cell(r, 1, "2025-09-05"); ws.cell(r, 2, "PAGO X"); ws.cell(r, 4, 999000)
    wb.save(str(dst)); wb.close()
    return str(dst), r


def test_link_writes_both_sides(m):
    path, bank_row = m
    apply_bank_factura_link(
        bank_row=bank_row, factura_row=2,
        nro_factura="FAC-555", fecha_pago="2025-09-05",
        excel_path=path,
    )
    wb = load_workbook(path, read_only=True)
    ws_f = wb[SHEET_NAME]
    ws_b = wb[CUENTA_BANCO_SHEET]
    assert "2025-09-05" in str(ws_f.cell(2, 3).value)
    assert ws_b.cell(bank_row, COL_BANCO_TIPO).value == "factura"
    assert ws_b.cell(bank_row, COL_BANCO_FACTURA_LINK).value == "FAC-555"
    wb.close()


def test_link_preserves_existing_fecha_pago(m):
    path, bank_row = m
    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    ws.cell(2, 3, "2025-08-01 (Manual)")
    wb.save(path); wb.close()

    result = apply_bank_factura_link(
        bank_row=bank_row, factura_row=2,
        nro_factura="FAC-555", fecha_pago="2025-09-05",
        excel_path=path,
    )
    assert result["fecha_pago_skipped"] is True
    wb = load_workbook(path, read_only=True)
    ws = wb[SHEET_NAME]
    assert "Manual" in str(ws.cell(2, 3).value)
    wb.close()
