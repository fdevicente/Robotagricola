import os, shutil, pytest
from unittest.mock import patch
from openpyxl import load_workbook
from excel_manager import (
    CUENTA_BANCO_SHEET, COL_BANCO_TIPO, COL_BANCO_CATEGORIA,
    COL_BANCO_CULTIVO, ensure_new_columns,
)


@pytest.fixture
def test_master(tmp_path):
    src = os.path.join(os.path.dirname(__file__), "..", "..",
                        "MASTER Agricola Santa Elisa.xlsx")
    dst = tmp_path / "test_master.xlsx"
    shutil.copy2(src, dst)
    ensure_new_columns(str(dst))
    return str(dst)


def test_bank_charge_categorized(test_master, tmp_path):
    from modules.cash_flow.categorizer import categorize_bank_movement
    wb = load_workbook(test_master)
    ws = wb[CUENTA_BANCO_SHEET]
    r = ws.max_row + 1
    ws.cell(r, 1, "2025-09-01")
    ws.cell(r, 2, "PAGO PROVEEDOR XYZ")
    ws.cell(r, 3, "TRF")
    ws.cell(r, 4, 500000)
    wb.save(test_master)
    wb.close()

    with patch("modules.cash_flow.categorizer.categorize_raw",
                return_value={"categoria": "Fertilizantes",
                              "cultivo": "NOGALES",
                              "confianza": 0.9, "razon": ""}):
        result = categorize_bank_movement(r, excel_path=test_master,
                                            cache_path=str(tmp_path / "c.json"))

    assert result["categoria"] == "Fertilizantes"
    wb = load_workbook(test_master, read_only=True)
    ws = wb[CUENTA_BANCO_SHEET]
    assert ws.cell(r, COL_BANCO_CATEGORIA).value == "Fertilizantes"
    assert ws.cell(r, COL_BANCO_CULTIVO).value == "NOGALES"
    wb.close()


def test_bank_abono_marked_ingreso(test_master, tmp_path):
    from modules.cash_flow.categorizer import categorize_bank_movement
    wb = load_workbook(test_master)
    ws = wb[CUENTA_BANCO_SHEET]
    r = ws.max_row + 1
    ws.cell(r, 1, "2025-09-01")
    ws.cell(r, 2, "DEPOSITO VALBIFRUT")
    ws.cell(r, 4, None)
    ws.cell(r, 5, 8000000)
    wb.save(test_master)
    wb.close()

    with patch("modules.cash_flow.categorizer.categorize_raw") as m:
        result = categorize_bank_movement(r, excel_path=test_master,
                                            cache_path=str(tmp_path / "c.json"))
    assert m.called is False
    assert result["tipo"] == "ingreso"

    wb = load_workbook(test_master, read_only=True)
    ws = wb[CUENTA_BANCO_SHEET]
    assert ws.cell(r, COL_BANCO_TIPO).value == "ingreso"
    wb.close()
