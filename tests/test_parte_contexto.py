# -*- coding: utf-8 -*-
"""El contexto es lo que el bot ya sabe: a quien conoce y que maquinas tiene.

OJO: los trabajadores NO salen solo de la hoja Personal. Medido el 2-sep-2026,
Personal tiene 6 filas con el nombre legal completo ("Felicito Amigo Soto") y no
incluye a Richard Padilla ni a su hijo, mientras la columna Trabajadores de la
bitacora usa los 8 nombres canonicos que el bot viene usando hace meses.
Armar el contexto solo con Personal dejaria a la IA peor informada que hoy.
"""
from openpyxl import Workbook

from modules.parte_contexto import construir


def _excel(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bitácora"
    ws.append(["Fecha", "Hora", "Tipo", "Actividad", "Cultivo", "Sector",
               "Jornadas Hombre", "Trabajadores", "Insumo", "Cantidad",
               "Unidad", "Registro", "Registrado por", "Máquina", "Odómetro",
               "Horas Día", "Superficie ha", "Días Cubiertos"])
    ws.append(["2026-08-20", "14:09", "LABOR", "Poda", "NOGALES", "", 2,
               "Richard Padilla, Richard Padilla Crespo", "", None, "",
               "texto", "Juan Parada", "", None, None, None, None])
    per = wb.create_sheet("Personal")
    per.append(["Nombre", "RUT", "Cargo", "Fecha Ingreso"])
    per.append(["Felicito Amigo Soto", "9.850.887-2", None, None])
    ruta = tmp_path / "master.xlsx"
    wb.save(ruta)
    return str(ruta)


def test_trae_los_nombres_de_la_bitacora(tmp_path):
    ctx = construir(_excel(tmp_path))
    assert "Richard Padilla" in ctx["trabajadores"]
    assert "Richard Padilla Crespo" in ctx["trabajadores"]


def test_tambien_trae_los_de_personal(tmp_path):
    ctx = construir(_excel(tmp_path))
    assert "Felicito Amigo Soto" in ctx["trabajadores"]


def test_trae_los_canonicos_de_siempre_aunque_no_esten_en_el_excel(tmp_path):
    """Sin esto se perderian los apodos y la regla del padre/hijo."""
    ctx = construir(_excel(tmp_path))
    assert "Patricio Mora" in ctx["trabajadores"]
    assert ctx["alias"]["pato"] == "Patricio Mora"
    assert ctx["alias"]["richard"] == "Richard Padilla"


def test_no_repite_nombres(tmp_path):
    ctx = construir(_excel(tmp_path))
    assert len(ctx["trabajadores"]) == len(set(ctx["trabajadores"]))


def test_las_maquinas_traen_unidad_y_ultima_lectura(tmp_path):
    ctx = construir(_excel(tmp_path))
    assert isinstance(ctx["maquinas"], list)
    for m in ctx["maquinas"]:
        assert set(m) >= {"maquina", "ultimo_odometro", "fecha", "unidad"}


def test_un_excel_sin_hojas_no_revienta(tmp_path):
    wb = Workbook()
    ruta = tmp_path / "vacio.xlsx"
    wb.save(ruta)
    ctx = construir(str(ruta))
    assert ctx["trabajadores"]          # quedan los canónicos de siempre
    assert ctx["maquinas"] == []
