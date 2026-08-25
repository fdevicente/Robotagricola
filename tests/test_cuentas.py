"""La caja son DOS cuentas: corriente en pesos y cuenta dólar.

Mirar solo la de pesos subestima el efectivo — ese error llevó una vez a
concluir que un exportador no estaba pagando cuando pagaba en USD.
"""
from datetime import date

import pytest
from openpyxl import Workbook

from modules.cuentas import (DOLAR_HEADERS, DOLAR_SHEET, caja_total,
                              desactualizadas, formato)

BANCO_HEADERS = ["Fecha", "Descripcion", "Referencia", "Cargo", "Abono",
                 "Saldo", "Tipo", "Categoria", "Cultivo", "Factura_linkeada"]


@pytest.fixture
def libro(tmp_path):
    """Master mínimo con las dos cuentas y el tipo de cambio."""
    def _crear(clp_rows, usd_rows, tipo_cambio=1000):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cuenta Banco"
        ws.append(BANCO_HEADERS)
        for r in clp_rows:
            ws.append(r)
        wsd = wb.create_sheet(DOLAR_SHEET)
        wsd.append(DOLAR_HEADERS)
        for r in usd_rows:
            wsd.append(r)
        wsc = wb.create_sheet("Config")
        wsc.append(["Parámetro", "Valor"])
        wsc.append(["usd_clp_estimado", tipo_cambio])
        p = tmp_path / "master.xlsx"
        wb.save(p); wb.close()
        return str(p)
    return _crear


def _mov(fecha, saldo, abono=0):
    return [fecha, "mov", "", None, abono, saldo, "", "", "", ""]


def test_suma_las_dos_cuentas(libro):
    p = libro([_mov(date(2026, 8, 3), 80_000_000)],
              [_mov(date(2026, 8, 5), 141_701.84)])
    c = caja_total(excel_path=p)
    assert c["clp"] == 80_000_000
    assert c["usd"] == pytest.approx(141_701.84)
    assert c["usd_en_clp"] == pytest.approx(141_701_840)
    assert c["total"] == pytest.approx(221_701_840)


def test_toma_el_saldo_mas_reciente_no_el_ultimo_de_la_hoja(libro):
    """Las filas pueden venir desordenadas: manda la fecha, no la posición."""
    p = libro([_mov(date(2026, 8, 3), 80_000_000),
               _mov(date(2026, 7, 1), 999_999_999)],
              [_mov(date(2026, 8, 5), 1_000)])
    assert caja_total(excel_path=p)["clp"] == 80_000_000


def test_usa_el_tipo_de_cambio_de_la_hoja_config(libro):
    p = libro([_mov(date(2026, 8, 3), 0)],
              [_mov(date(2026, 8, 5), 100)], tipo_cambio=904)
    c = caja_total(excel_path=p)
    assert c["tipo_cambio"] == 904
    assert c["total"] == pytest.approx(90_400)


def test_sin_hoja_dolar_no_revienta(tmp_path):
    """Un Master viejo sin la hoja debe seguir funcionando (solo pesos)."""
    wb = Workbook()
    ws = wb.active; ws.title = "Cuenta Banco"
    ws.append(BANCO_HEADERS)
    ws.append(_mov(date(2026, 8, 3), 50_000_000))
    p = tmp_path / "viejo.xlsx"
    wb.save(p); wb.close()

    c = caja_total(excel_path=str(p))
    assert c["usd"] == 0
    assert c["total"] == 50_000_000


def test_avisa_cuando_una_cuenta_quedo_vieja(libro):
    p = libro([_mov(date(2026, 6, 1), 80_000_000)],
              [_mov(date(2026, 8, 5), 100)])
    c = caja_total(excel_path=p)
    avisos = desactualizadas(c, hoy=date(2026, 8, 5))
    assert len(avisos) == 1
    assert "corriente" in avisos[0]


def test_no_avisa_si_ambas_estan_al_dia(libro):
    p = libro([_mov(date(2026, 8, 4), 80_000_000)],
              [_mov(date(2026, 8, 5), 100)])
    c = caja_total(excel_path=p)
    assert desactualizadas(c, hoy=date(2026, 8, 5)) == []


def test_formato_muestra_las_dos_cuentas(libro):
    p = libro([_mov(date(2026, 8, 4), 80_703_080)],
              [_mov(date(2026, 8, 5), 141_701.84)])
    txt = formato(caja_total(excel_path=p))
    assert "222,404,920" in txt      # 80.703.080 + 141.701,84 × 1.000
    assert "141,701.84" in txt
    assert "Cuenta dólar" in txt
