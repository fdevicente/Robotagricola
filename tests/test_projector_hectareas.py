import os, shutil, pytest
from openpyxl import load_workbook
from excel_manager import ensure_cash_flow_sheets, AJUSTES_SHEET


@pytest.fixture
def m(tmp_path):
    src = os.path.join(os.path.dirname(__file__), "..", "..",
                        "MASTER Agricola Santa Elisa.xlsx")
    dst = tmp_path / "test.xlsx"
    shutil.copy2(src, dst)
    ensure_cash_flow_sheets(str(dst))
    return str(dst)


def _vaciar(path, hoja):
    """Deja la hoja solo con su encabezado (el MASTER real trae datos)."""
    wb = load_workbook(path)
    ws = wb[hoja]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    wb.save(path); wb.close()


def test_load_hectareas_returns_dict(m):
    """El loader devuelve lo que dice la hoja (no valores hardcodeados)."""
    from modules.cash_flow.projector import load_hectareas
    hc = load_hectareas(excel_path=m)
    wb = load_workbook(m, read_only=True, data_only=True)
    ws = wb["Hectareas"]
    filas = [r for r in ws.iter_rows(min_row=2, max_col=4, values_only=True) if r[0]]
    wb.close()

    assert filas, "la hoja Hectareas quedó vacía"
    for anio, nog, cer, ave in filas:
        assert anio in hc
        assert hc[anio]["NOGALES"] == nog
        assert hc[anio]["CEREZOS"] == cer
        assert hc[anio]["AVELLANOS"] == ave


def test_load_ajustes_empty_if_no_data(m):
    from modules.cash_flow.projector import load_ajustes_manuales
    _vaciar(m, AJUSTES_SHEET)
    assert load_ajustes_manuales(excel_path=m) == []


def test_load_ajustes_filters_inactive(m):
    from modules.cash_flow.projector import load_ajustes_manuales
    _vaciar(m, AJUSTES_SHEET)
    wb = load_workbook(m)
    ws = wb[AJUSTES_SHEET]
    ws.append(["2026-05-01", "2026-07", "Riego", "GENERAL", 5000000, "Bomba nueva", True])
    ws.append(["2026-05-01", "2026-08", "Fertilizantes", "NOGALES", 2000000, "Test", False])
    wb.save(m); wb.close()

    ajustes = load_ajustes_manuales(excel_path=m)
    assert len(ajustes) == 1
    assert ajustes[0]["categoria"] == "Riego"
    assert ajustes[0]["monto"] == 5000000
