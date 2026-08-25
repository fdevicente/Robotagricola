# tests/test_excel_sheets.py
import os, shutil, pytest
from openpyxl import load_workbook
from excel_manager import ensure_cash_flow_sheets, COSECHAS_SHEET, GUIAS_SHEET
from excel_manager import FLUJO_CAJA_SHEET, AJUSTES_SHEET, CONFIG_SHEET, HECTAREAS_SHEET
from config import CASH_FLOW_CONFIG


@pytest.fixture
def test_excel(tmp_path):
    src = os.path.join(os.path.dirname(__file__), "..", "..", "MASTER Agricola Santa Elisa.xlsx")
    dst = tmp_path / "test_master.xlsx"
    shutil.copy2(src, dst)
    return str(dst)


def test_creates_missing_sheets(test_excel):
    ensure_cash_flow_sheets(test_excel)
    wb = load_workbook(test_excel, read_only=True)
    names = wb.sheetnames
    assert COSECHAS_SHEET in names
    assert GUIAS_SHEET in names
    assert CONFIG_SHEET in names
    assert HECTAREAS_SHEET in names
    assert AJUSTES_SHEET in names
    wb.close()


def test_idempotent(test_excel):
    ensure_cash_flow_sheets(test_excel)
    ensure_cash_flow_sheets(test_excel)  # no error on second run
    wb = load_workbook(test_excel, read_only=True)
    count = wb.sheetnames.count(CONFIG_SHEET)
    assert count == 1
    wb.close()


def test_config_has_defaults(test_excel):
    ensure_cash_flow_sheets(test_excel)
    wb = load_workbook(test_excel, read_only=True)
    ws = wb[CONFIG_SHEET]
    params = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(2, ws.max_row + 1)}
    assert params['saldo_minimo_pct'] == 0.10
    # se compara contra la config, no contra un literal: asi no queda obsoleto
    # cada vez que el dueno ajusta el tipo de cambio
    assert params['usd_clp_estimado'] == CASH_FLOW_CONFIG['usd_clp_estimado']
    wb.close()


def test_hectareas_has_data(test_excel):
    """Hay una fila por año con superficie numérica en los 3 cultivos.

    No se fijan valores: las hectáreas cambian con el replante.
    """
    ensure_cash_flow_sheets(test_excel)
    wb = load_workbook(test_excel, read_only=True)
    ws = wb[HECTAREAS_SHEET]
    filas = [r for r in ws.iter_rows(min_row=2, max_col=4, values_only=True) if r[0]]
    wb.close()

    assert len(filas) >= 3, "faltan años en la hoja Hectareas"
    anios = [f[0] for f in filas]
    assert anios == sorted(anios), "los años no están en orden"
    for anio, nog, cer, ave in filas:
        assert isinstance(anio, int)
        for sup in (nog, cer, ave):
            assert isinstance(sup, (int, float)) and sup >= 0
