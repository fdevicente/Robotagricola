import os, shutil, pytest
from openpyxl import load_workbook
from excel_manager import (
    SHEET_NAME, CUENTA_BANCO_SHEET, ensure_new_columns,
    COL_BANCO_FACTURA_LINK,
    read_facturas_pendientes, read_bank_movements_unlinked,
)


@pytest.fixture
def m(tmp_path):
    src = os.path.join(os.path.dirname(__file__), "..", "..",
                        "MASTER Agricola Santa Elisa.xlsx")
    dst = tmp_path / "test.xlsx"
    shutil.copy2(src, dst)
    ensure_new_columns(str(dst))
    return str(dst)


def test_pendientes_excludes_paid(m):
    wb = load_workbook(m)
    ws = wb[SHEET_NAME]
    ws.cell(2, 3, "2025-09-01 (Banco)")
    wb.save(m); wb.close()
    pendientes = read_facturas_pendientes(excel_path=m)
    assert all(p["fila"] != 2 for p in pendientes)


def test_pendientes_has_required_fields(m):
    pendientes = read_facturas_pendientes(excel_path=m)
    if pendientes:
        p = pendientes[0]
        assert "fila" in p and "total" in p and "proveedor" in p
        assert "fecha_emision" in p and "nro_factura" in p


def test_bank_unlinked_excludes_linked(m):
    wb = load_workbook(m)
    ws = wb[CUENTA_BANCO_SHEET]
    r = ws.max_row + 1
    ws.cell(r, 1, "2025-09-01"); ws.cell(r, 2, "X"); ws.cell(r, 4, 1000)
    ws.cell(r, COL_BANCO_FACTURA_LINK, "FAC-123")
    wb.save(m); wb.close()
    movs = read_bank_movements_unlinked(excel_path=m)
    assert all(mv["fila"] != r for mv in movs)


def test_bank_unlinked_only_cargos(m):
    wb = load_workbook(m)
    ws = wb[CUENTA_BANCO_SHEET]
    r = ws.max_row + 1
    ws.cell(r, 1, "2025-09-01"); ws.cell(r, 2, "DEP"); ws.cell(r, 5, 5000)
    wb.save(m); wb.close()
    movs = read_bank_movements_unlinked(excel_path=m)
    assert all(mv["fila"] != r for mv in movs)
