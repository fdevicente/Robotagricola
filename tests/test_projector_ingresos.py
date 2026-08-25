import os, shutil, pytest
from openpyxl import load_workbook
from excel_manager import ensure_cash_flow_sheets, COSECHAS_SHEET
from config import CASH_FLOW_CONFIG


@pytest.fixture
def m(tmp_path):
    src = os.path.join(os.path.dirname(__file__), "..", "..",
                        "MASTER Agricola Santa Elisa.xlsx")
    dst = tmp_path / "test.xlsx"
    shutil.copy2(src, dst)
    ensure_cash_flow_sheets(str(dst))
    return str(dst)


def test_load_ingresos_uses_real_if_present(m):
    from modules.cash_flow.projector import load_expected_ingresos
    wb = load_workbook(m)
    ws = wb[COSECHAS_SHEET]
    # Limpiar primero
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    ws.append([2026, "NOGALES", 240000, "Valbifrut", 140000, 1.8, 2, 1,
                "2026-06-15", 126000, "adelanto", "recibido",
                "2026-06-20", 120000000, "CLP", ""])
    wb.save(m); wb.close()

    ingresos = load_expected_ingresos(excel_path=m)
    assert len(ingresos) == 1
    assert ingresos[0]["monto_clp"] == 120000000


def test_load_ingresos_estimates_usd_when_pending(m):
    from modules.cash_flow.projector import load_expected_ingresos
    wb = load_workbook(m)
    ws = wb[COSECHAS_SHEET]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    ws.append([2026, "NOGALES", 240000, "Valbifrut", 140000, 1.8, 2, 1,
                "2026-06-15", 126000, "adelanto", "esperado",
                None, None, "", ""])
    wb.save(m); wb.close()

    ingresos = load_expected_ingresos(excel_path=m)
    # sin columna "Aplica IVA" (fila de 16 columnas) -> queda en el neto.
    # Se compara contra el tipo de cambio configurado, no contra un literal.
    assert ingresos[0]["monto_clp"] == 126_000 * CASH_FLOW_CONFIG["usd_clp_estimado"]


def test_load_ingresos_empty_returns_empty(m):
    from modules.cash_flow.projector import load_expected_ingresos
    wb = load_workbook(m)
    ws = wb[COSECHAS_SHEET]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    wb.save(m); wb.close()
    ingresos = load_expected_ingresos(excel_path=m)
    assert ingresos == []
