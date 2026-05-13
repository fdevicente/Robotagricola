import os, shutil, pytest
from openpyxl import load_workbook
from excel_manager import (
    CUENTA_BANCO_SHEET, COL_BANCO_TIPO, ensure_new_columns,
)


@pytest.fixture
def test_master(tmp_path):
    src = os.path.join(os.path.dirname(__file__), "..", "..",
                        "MASTER Agricola Santa Elisa.xlsx")
    dst = tmp_path / "test_master.xlsx"
    shutil.copy2(src, dst)
    ensure_new_columns(str(dst))
    return str(dst)


def _add_bank_row(path, fecha, descripcion, cargo=None, abono=None):
    wb = load_workbook(path)
    ws = wb[CUENTA_BANCO_SHEET]
    r = ws.max_row + 1
    ws.cell(r, 1, fecha)
    ws.cell(r, 2, descripcion)
    ws.cell(r, 3, "REF")
    ws.cell(r, 4, cargo)
    ws.cell(r, 5, abono)
    wb.save(path)
    wb.close()
    return r


def test_detects_venta_dolares(test_master):
    from modules.cash_flow.historical_importer import detect_income_patterns
    r = _add_bank_row(test_master, "2025-09-01",
                       "VENTA DOLARES MISMA EMPRESA", abono=15000000)
    detect_income_patterns(excel_path=test_master)
    wb = load_workbook(test_master, read_only=True)
    ws = wb[CUENTA_BANCO_SHEET]
    assert ws.cell(r, COL_BANCO_TIPO).value == "venta_dolares"
    wb.close()


def test_detects_vitakai_ingreso_clp(test_master):
    from modules.cash_flow.historical_importer import detect_income_patterns
    r = _add_bank_row(test_master, "2025-09-02",
                       "DEPOSITO VITAKAI SPA", abono=5000000)
    detect_income_patterns(excel_path=test_master)
    wb = load_workbook(test_master, read_only=True)
    ws = wb[CUENTA_BANCO_SHEET]
    assert ws.cell(r, COL_BANCO_TIPO).value == "ingreso_clp"
    wb.close()


def test_detects_sueldo(test_master):
    from modules.cash_flow.historical_importer import detect_income_patterns
    r = _add_bank_row(test_master, "2025-09-03",
                       "PAGO REMUNERACIONES PERSONAL", cargo=8000000)
    detect_income_patterns(excel_path=test_master)
    wb = load_workbook(test_master, read_only=True)
    ws = wb[CUENTA_BANCO_SHEET]
    assert ws.cell(r, COL_BANCO_TIPO).value == "sueldo"
    wb.close()


def test_idempotent_skips_tagged(test_master):
    from modules.cash_flow.historical_importer import detect_income_patterns
    r = _add_bank_row(test_master, "2025-09-04",
                       "VENTA DOLARES", abono=1000000)
    detect_income_patterns(excel_path=test_master)
    wb = load_workbook(test_master)
    ws = wb[CUENTA_BANCO_SHEET]
    ws.cell(r, COL_BANCO_TIPO, "manual_override")
    wb.save(test_master)
    wb.close()
    detect_income_patterns(excel_path=test_master)
    wb = load_workbook(test_master, read_only=True)
    ws = wb[CUENTA_BANCO_SHEET]
    assert ws.cell(r, COL_BANCO_TIPO).value == "manual_override"
    wb.close()
