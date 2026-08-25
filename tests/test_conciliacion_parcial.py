"""Fase 4: conciliación parcial y N:M.

Dos casos reales que el modelo tiene que soportar:
  · un cargo paga VARIAS facturas (Copeval "ND506808 al ND506813")
  · una factura se paga en VARIAS cuotas (S-Invest, plantas de avellano)

Los tests pasan `excel_path` explícito: nunca tocar el Master real.
"""
import pytest
from openpyxl import Workbook

from modules.conciliacion_store import (HEADERS, SHEET, _clave_doc,
                                         asignado_por_documento,
                                         estado_documento, saldo_por_asignar,
                                         vinculos_de_documento)

BANCO = ["Fecha", "Descripcion", "Referencia", "Cargo", "Abono", "Saldo",
         "Tipo", "Categoria", "Cultivo", "Factura_linkeada"]


@pytest.fixture
def libro(tmp_path):
    """Master de prueba: banco + conciliaciones."""
    def _crear(movimientos, vinculos):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cuenta Banco"
        ws.append(BANCO)
        for m in movimientos:                      # (fecha, desc, cargo, abono)
            ws.append([m[0], m[1], "", m[2] or None, m[3] or None,
                       None, "", "", "", ""])
        wc = wb.create_sheet(SHEET)
        wc.append(HEADERS)
        for i, v in enumerate(vinculos, 1):        # (fila_banco, nro, prov, asignado)
            wc.append([i, "2026-08-01", v[0], "2026-07-01", "mov", 0,
                       "FACTURA", None, v[1], v[2], v[3], "manual", "", ""])
        p = tmp_path / "master.xlsx"
        wb.save(p); wb.close()
        return str(p)
    return _crear


# ── Clave del documento ──────────────────────────────────────────────────

def test_la_clave_normaliza_mayusculas_y_el_sufijo_punto_cero():
    """Excel devuelve los n° numéricos como '6231521.0'."""
    assert _clave_doc("6231521.0", "copeval") == _clave_doc("6231521", "COPEVAL")


def test_la_clave_separa_proveedores_con_el_mismo_numero():
    assert _clave_doc("F94", "Misael") != _clave_doc("F94", "Contreras")


# ── Un cargo paga VARIAS facturas (Copeval agrupadas) ────────────────────

def test_un_movimiento_repartido_entre_varias_facturas(libro):
    p = libro([("2026-06-09", "Copeval ND506808 al ND506813", 713_590, 0)],
              [(2, "ND506808", "Copeval", 300_000),
               (2, "ND506810", "Copeval", 213_590),
               (2, "ND506813", "Copeval", 200_000)])
    d = saldo_por_asignar(2, excel_path=p)
    assert d["monto"] == 713_590
    assert d["asignado"] == 713_590
    assert d["saldo"] == 0
    assert d["estado"] == "conciliado"


def test_movimiento_a_medio_asignar_queda_parcial(libro):
    p = libro([("2026-06-09", "Copeval agrupadas", 713_590, 0)],
              [(2, "ND506808", "Copeval", 300_000)])
    d = saldo_por_asignar(2, excel_path=p)
    assert d["asignado"] == 300_000
    assert d["saldo"] == 413_590
    assert d["estado"] == "parcial"


def test_movimiento_sin_vinculos_queda_por_conciliar(libro):
    p = libro([("2026-06-09", "Algo", 100_000, 0)], [])
    d = saldo_por_asignar(2, excel_path=p)
    assert d["asignado"] == 0
    assert d["saldo"] == 100_000
    assert d["estado"] == "por conciliar"


def test_toma_el_abono_cuando_no_hay_cargo(libro):
    p = libro([("2026-05-04", "Valbifrut adelanto", 0, 223_596_523)], [])
    assert saldo_por_asignar(2, excel_path=p)["monto"] == 223_596_523


# ── Una factura pagada en VARIAS cuotas (S-Invest) ───────────────────────

def test_factura_pagada_en_cuotas_acumula(libro):
    p = libro([("2026-05-12", "S-Invest cuota 1", 5_000_000, 0),
               ("2026-06-03", "S-Invest cuota 2", 5_000_000, 0)],
              [(2, "1928", "S-Invest 2", 5_000_000),
               (3, "1928", "S-Invest 2", 5_000_000)])
    assert asignado_por_documento(p)[_clave_doc("1928", "S-Invest 2")] == 10_000_000

    d = estado_documento("1928", "S-Invest 2", total=10_000_000, excel_path=p)
    assert d["asignado"] == 10_000_000
    assert d["saldo"] == 0
    assert d["estado"] == "pagado"


def test_factura_a_medias_queda_parcial(libro):
    p = libro([("2026-05-12", "S-Invest cuota 1", 5_000_000, 0)],
              [(2, "1928", "S-Invest 2", 5_000_000)])
    d = estado_documento("1928", "S-Invest 2", total=10_000_000, excel_path=p)
    assert d["saldo"] == 5_000_000
    assert d["estado"] == "parcial"


def test_factura_sin_pagos_queda_pendiente(libro):
    p = libro([], [])
    d = estado_documento("9999", "Nadie", total=1_000_000, excel_path=p)
    assert d["asignado"] == 0
    assert d["estado"] == "pendiente"


def test_las_cuotas_salen_ordenadas_por_fecha(libro):
    p = libro([("2026-05-12", "cuota 1", 5_000_000, 0),
               ("2026-06-03", "cuota 2", 5_000_000, 0)],
              [(3, "1928", "S-Invest 2", 5_000_000),
               (2, "1928", "S-Invest 2", 5_000_000)])
    cuotas = vinculos_de_documento("1928", "S-Invest 2", excel_path=p)
    assert len(cuotas) == 2
    assert [c["fila_banco"] for c in cuotas] == [3, 2]   # ambas fecha_mov igual → orden estable
    assert sum(c["monto_asignado"] for c in cuotas) == 10_000_000


def test_no_mezcla_documentos_de_proveedores_distintos(libro):
    """Mismo n° de factura, proveedor distinto: son documentos separados."""
    p = libro([("2026-05-12", "pago A", 100_000, 0),
               ("2026-05-13", "pago B", 200_000, 0)],
              [(2, "F94", "Misael Henriquez", 100_000),
               (3, "F94", "Herramientas Contreras", 200_000)])
    a = estado_documento("F94", "Misael Henriquez", total=100_000, excel_path=p)
    b = estado_documento("F94", "Herramientas Contreras", total=200_000, excel_path=p)
    assert a["asignado"] == 100_000
    assert b["asignado"] == 200_000


def test_sin_hoja_de_conciliaciones_no_revienta(tmp_path):
    wb = Workbook()
    wb.active.title = "Cuenta Banco"
    wb.active.append(BANCO)
    wb.active.append(["2026-01-01", "x", "", 100, None, None, "", "", "", ""])
    p = tmp_path / "viejo.xlsx"
    wb.save(p); wb.close()

    assert asignado_por_documento(str(p)) == {}
    assert vinculos_de_documento("1", "x", excel_path=str(p)) == []
    assert saldo_por_asignar(2, excel_path=str(p))["estado"] == "por conciliar"
