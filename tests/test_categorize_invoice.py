import os, shutil, pytest
from unittest.mock import patch
from openpyxl import load_workbook
from excel_manager import (
    SHEET_NAME, COL_CATEGORIA, COL_CULTIVO, COL_CONFIANZA, COL_CATEGORIZADO_POR,
    ensure_new_columns,
)


@pytest.fixture
def test_master(tmp_path):
    src = os.path.join(os.path.dirname(__file__), "..", "..",
                        "MASTER Agricola Santa Elisa.xlsx")
    dst = tmp_path / "test_master.xlsx"
    shutil.copy2(src, dst)
    ensure_new_columns(str(dst))
    return str(dst)


def _fake_result(cat="Fertilizantes", cultivo="NOGALES", conf=0.9):
    return {"categoria": cat, "cultivo": cultivo,
            "confianza": conf, "razon": "ok"}


def test_categorize_invoice_writes_to_master(test_master, tmp_path):
    from modules.cash_flow.categorizer import categorize_invoice
    wb = load_workbook(test_master)
    ws = wb[SHEET_NAME]
    target_row = None
    for r in range(2, min(ws.max_row + 1, 20)):
        if ws.cell(r, 4).value:
            target_row = r
            break
    wb.close()
    assert target_row is not None

    cache_path = str(tmp_path / "cache.json")
    with patch("modules.cash_flow.categorizer.categorize_raw",
                return_value=_fake_result()):
        result = categorize_invoice(target_row, excel_path=test_master,
                                      cache_path=cache_path)

    assert result["categoria"] == "Fertilizantes"

    wb = load_workbook(test_master, read_only=True)
    ws = wb[SHEET_NAME]
    assert ws.cell(target_row, COL_CATEGORIA).value == "Fertilizantes"
    assert ws.cell(target_row, COL_CULTIVO).value == "NOGALES"
    assert ws.cell(target_row, COL_CONFIANZA).value == 0.9
    assert ws.cell(target_row, COL_CATEGORIZADO_POR).value == "claude"
    wb.close()


def test_categorize_invoice_low_confidence_marks_revisar(test_master, tmp_path):
    from modules.cash_flow.categorizer import categorize_invoice
    wb = load_workbook(test_master)
    ws = wb[SHEET_NAME]
    target_row = None
    for r in range(2, min(ws.max_row + 1, 20)):
        if ws.cell(r, 4).value:
            target_row = r
            break
    wb.close()

    cache_path = str(tmp_path / "cache.json")
    with patch("modules.cash_flow.categorizer.categorize_raw",
                return_value=_fake_result(cat="Riego", conf=0.5)):
        categorize_invoice(target_row, excel_path=test_master,
                            cache_path=cache_path)

    wb = load_workbook(test_master, read_only=True)
    ws = wb[SHEET_NAME]
    assert ws.cell(target_row, COL_CATEGORIA).value == "REVISAR"
    wb.close()


def test_categorize_invoice_uses_cache_on_second_call(test_master, tmp_path):
    from modules.cash_flow.categorizer import categorize_invoice
    wb = load_workbook(test_master)
    ws = wb[SHEET_NAME]
    r = next((r for r in range(2, 20) if ws.cell(r, 4).value), None)
    wb.close()

    cache_path = str(tmp_path / "cache.json")
    with patch("modules.cash_flow.categorizer.categorize_raw",
                return_value=_fake_result()) as mock_raw:
        categorize_invoice(r, excel_path=test_master, cache_path=cache_path)
        categorize_invoice(r, excel_path=test_master, cache_path=cache_path)
    assert mock_raw.call_count == 1
