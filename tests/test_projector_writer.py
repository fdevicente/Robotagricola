import os, shutil, pytest
from openpyxl import load_workbook
from excel_manager import ensure_cash_flow_sheets, FLUJO_CAJA_SHEET


@pytest.fixture
def m(tmp_path):
    src = os.path.join(os.path.dirname(__file__), "..", "..",
                        "MASTER Agricola Santa Elisa.xlsx")
    dst = tmp_path / "test.xlsx"
    shutil.copy2(src, dst)
    ensure_cash_flow_sheets(str(dst))
    return str(dst)


def test_write_flujo_caja_creates_header(m):
    from modules.cash_flow.projector import write_flujo_caja
    saldo_data = {
        (2026, 5): {"saldo_inicio": 100, "ingresos": 200, "egresos": 50, "saldo_cierre": 250},
        (2026, 6): {"saldo_inicio": 250, "ingresos": 0, "egresos": 30, "saldo_cierre": 220},
    }
    egresos = {
        (2026, 5, "Fertilizantes", "NOGALES"): 30,
        (2026, 5, "Riego", "GENERAL"): 20,
        (2026, 6, "Combustible", "NOGALES"): 30,
    }
    ingresos = [
        {"year": 2026, "month": 5, "monto_clp": 200, "exportadora": "Valbifrut",
         "estado": "recibido"},
    ]
    write_flujo_caja(saldo_data, egresos, ingresos,
                       months=[(2026, 5), (2026, 6)], excel_path=m)

    wb = load_workbook(m, read_only=True)
    ws = wb[FLUJO_CAJA_SHEET]
    assert ws.cell(1, 1).value in ("SECCION", "Seccion", "Sección")
    assert ws.cell(1, 2).value is not None
    assert ws.cell(1, 3).value is not None
    wb.close()


def test_write_flujo_caja_has_saldo_rows(m):
    from modules.cash_flow.projector import write_flujo_caja
    saldo_data = {
        (2026, 5): {"saldo_inicio": 100, "ingresos": 200, "egresos": 50, "saldo_cierre": 250},
    }
    write_flujo_caja(saldo_data, {}, [], months=[(2026, 5)], excel_path=m)

    wb = load_workbook(m, read_only=True)
    ws = wb[FLUJO_CAJA_SHEET]
    labels = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
    assert any("SALDO INICIAL" in str(l or "") for l in labels)
    assert any("SALDO CIERRE" in str(l or "") for l in labels)
    assert any("TOTAL EGRESOS" in str(l or "") for l in labels)
    wb.close()
