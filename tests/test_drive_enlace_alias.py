# -*- coding: utf-8 -*-
"""El nombre del ARCHIVO queda congelado; el del Master cambia. Deben calzar igual.

Corrigiendo los proveedores del Master (31-ago-2026) quedó a la vista algo que
el cruce por nombre no aguanta: los archivos ya subidos conservan la grafía que
tenía el Master EL DÍA que se capturó la foto. Al unificar
'FERRETERIAINDUTRIAL TALCA LIMITADA' -> 'Ferreteria Industrial Talca Limitada'
se arreglaron 7 archivos y se rompieron 8, porque esos 8 se llaman con la
grafía vieja. Renombrar el Master NUNCA va a alcanzar por sí solo.

Dos reglas, las dos con evidencia medida:

1. ALIAS. Grafías distintas del mismo proveedor. Se comprobó con el RUT:
   FERRETERIAINDUTRIAL y Ferreteria Industrial Talca comparten 78.045.980-8.

2. TRUNCADO. `handlers.facturas._limpiar` corta el nombre a 60 caracteres, así
   que un proveedor de nombre largo queda cortado EN EL ARCHIVO y completo en
   el Master: 'CONFECCION DE MAXISACOS ... CABRERA E.I.' contra
   '... CABRERA E.I.R.L.'. Un prefijo de exactamente 60 es un nombre cortado,
   no otro proveedor.
"""
import openpyxl
import pytest

from modules.drive.enlaces import guardar_enlace

COL_PROV, COL_NUM, COL_DRIVE = 4, 7, 22


def _libro(tmp_path, proveedor, numero=100):
    ruta = tmp_path / "m.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"
    ws.append([None] * 21 + ["Archivo Drive"])
    f = [None] * 22
    f[COL_PROV - 1], f[COL_NUM - 1] = proveedor, numero
    ws.append(f)
    wb.save(ruta)
    wb.close()
    return str(ruta)


def _enlace(ruta):
    wb = openpyxl.load_workbook(ruta)
    v = wb["Facturas"].cell(2, COL_DRIVE).value
    wb.close()
    return v


class TestAlias:
    @pytest.mark.parametrize("en_master,en_archivo", [
        ("Ferreteria Industrial Talca Limitada", "FERRETERIAINDUTRIAL_TALCA_LIMITADA"),
        ("FERRETERIAINDUTRIAL TALCA LIMITADA", "Ferreteria_Industrial_Talca_Limitada"),
        ("Ferreteria Industrial Pachita SPA", "FERRETERIA_INDUSTRIAL_PAGHITA_SPA"),
        ("ROTORTEC", "Servicios_Y_Arriendos_Rotortec_SPA"),
        ("Irrifor", "IRRIFER"),
        ("Salinas y Fabres", "SALINAS_Y_FABRES"),
    ])
    def test_las_dos_grafias_son_el_mismo_proveedor(self, tmp_path, en_master,
                                                     en_archivo):
        ruta = _libro(tmp_path, en_master)
        assert guardar_enlace(ruta, "100", "abc", proveedor=en_archivo) is True
        assert "abc" in str(_enlace(ruta))

    def test_el_alias_NO_junta_proveedores_distintos(self, tmp_path):
        """ALBINO FUENTEALBA y Serc. Mantencion C. Bustos comparten el Nº303
        pero son dos empresas: tienen RUT distintos y no deben enlazarse."""
        ruta = _libro(tmp_path, "Serc. Mantencion C. Bustos", 303)
        assert guardar_enlace(ruta, "303", "abc",
                              proveedor="ALBINO_FUENTEALBA") is False
        assert _enlace(ruta) is None


class TestNombreTruncado:
    def test_el_nombre_cortado_a_60_calza_con_el_completo(self, tmp_path):
        completo = ("CONFECCION DE MAXISACOS CARMEN GLORIA MANRIQUEZ "
                    "CABRERA E.I.R.L.")
        cortado = "CONFECCION_DE_MAXISACOS_CARMEN_GLORIA_MANRIQUEZ_CABRERA_E.I."
        assert len(cortado) == 60, len(cortado)
        ruta = _libro(tmp_path, completo)
        assert guardar_enlace(ruta, "100", "abc", proveedor=cortado) is True

    def test_un_prefijo_CORTO_no_alcanza_para_calzar(self, tmp_path):
        """'COPEVAL' es prefijo de 'COPEVAL SA' pero no viene de un truncado:
        solo un nombre de exactamente 60 esta cortado."""
        ruta = _libro(tmp_path, "COPEVALSA DISTRIBUIDORA")
        assert guardar_enlace(ruta, "100", "abc", proveedor="COPEVAL") is False
