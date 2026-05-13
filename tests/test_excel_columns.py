import os, shutil, pytest
from openpyxl import load_workbook
from excel_manager import (ensure_new_columns, COL_CATEGORIA, COL_CULTIVO,
    COL_CONFIANZA, COL_CATEGORIZADO_POR, COL_BANCO_TIPO,
    SHEET_NAME, CUENTA_BANCO_SHEET)


@pytest.fixture
def test_excel(tmp_path):
    src = os.path.join(os.path.dirname(__file__), "..", "..", "MASTER Agricola Santa Elisa.xlsx")
    dst = tmp_path / "test_master.xlsx"
    shutil.copy2(src, dst)
    return str(dst)


def test_facturas_new_headers(test_excel):
    ensure_new_columns(test_excel)
    wb = load_workbook(test_excel, read_only=True)
    ws = wb[SHEET_NAME]
    assert ws.cell(1, COL_CATEGORIA).value == "Categoria"
    assert ws.cell(1, COL_CULTIVO).value == "Cultivo"
    assert ws.cell(1, COL_CONFIANZA).value == "Confianza"
    assert ws.cell(1, COL_CATEGORIZADO_POR).value == "Categorizado_por"
    wb.close()


def test_banco_new_headers(test_excel):
    ensure_new_columns(test_excel)
    wb = load_workbook(test_excel, read_only=True)
    ws = wb[CUENTA_BANCO_SHEET]
    assert ws.cell(1, COL_BANCO_TIPO).value == "Tipo"
    assert ws.cell(1, 8).value == "Categoria"
    assert ws.cell(1, 9).value == "Cultivo"
    assert ws.cell(1, 10).value == "Factura_linkeada"
    wb.close()


def test_idempotent(test_excel):
    ensure_new_columns(test_excel)
    ensure_new_columns(test_excel)
    wb = load_workbook(test_excel, read_only=True)
    ws = wb[SHEET_NAME]
    assert ws.cell(1, COL_CATEGORIA).value == "Categoria"
    wb.close()
