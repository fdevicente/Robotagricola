# -*- coding: utf-8 -*-
"""El enlace de Drive vuelve a la fila de la factura.

OJO: ruta EXPLÍCITA al Excel en todos los tests. Un test destruyó el Master
real por confiar en el default de _save_wb.
"""
import openpyxl
import pytest

from modules.drive.enlaces import guardar_enlace

COL_NUM, COL_DRIVE = 7, 22


@pytest.fixture
def libro(tmp_path):
    ruta = tmp_path / "master.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"
    ws.append([None] * 21 + ["Archivo Drive"])
    fila = [None] * 22
    fila[COL_NUM - 1] = 2777
    ws.append(fila)
    wb.save(ruta)
    wb.close()
    return str(ruta)


def test_guarda_el_enlace_en_la_fila_del_numero(libro):
    guardar_enlace(libro, "2777", "abc123")
    wb = openpyxl.load_workbook(libro)
    assert "abc123" in str(wb["Facturas"].cell(2, COL_DRIVE).value)
    wb.close()


def test_el_enlace_es_una_url_abrible(libro):
    guardar_enlace(libro, "2777", "abc123")
    wb = openpyxl.load_workbook(libro)
    assert str(wb["Facturas"].cell(2, COL_DRIVE).value).startswith("https://")
    wb.close()


def test_un_numero_que_no_existe_no_rompe(libro):
    guardar_enlace(libro, "9999", "abc123")
    wb = openpyxl.load_workbook(libro)
    assert wb["Facturas"].cell(2, COL_DRIVE).value is None
    wb.close()


def test_no_pisa_un_enlace_ya_puesto(libro):
    guardar_enlace(libro, "2777", "primero")
    guardar_enlace(libro, "2777", "segundo")
    wb = openpyxl.load_workbook(libro)
    assert "primero" in str(wb["Facturas"].cell(2, COL_DRIVE).value)
    wb.close()


def test_pone_el_enlace_en_TODAS_las_lineas_de_la_factura(tmp_path):
    """Una factura con varios ítems son varias filas con el mismo número."""
    ruta = tmp_path / "m.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"
    ws.append([None] * 21 + ["Archivo Drive"])
    for _ in range(3):
        f = [None] * 22
        f[COL_NUM - 1] = 2777
        ws.append(f)
    wb.save(ruta)
    wb.close()

    guardar_enlace(str(ruta), "2777", "abc123")
    wb = openpyxl.load_workbook(ruta)
    for fila in (2, 3, 4):
        assert "abc123" in str(wb["Facturas"].cell(fila, COL_DRIVE).value)
    wb.close()


def test_no_toca_las_filas_de_otras_facturas(tmp_path):
    ruta = tmp_path / "m.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"
    ws.append([None] * 21 + ["Archivo Drive"])
    for nro in (2777, 2888):
        f = [None] * 22
        f[COL_NUM - 1] = nro
        ws.append(f)
    wb.save(ruta)
    wb.close()

    guardar_enlace(str(ruta), "2777", "abc123")
    wb = openpyxl.load_workbook(ruta)
    assert wb["Facturas"].cell(3, COL_DRIVE).value is None   # la 2888 intacta
    wb.close()


def test_devuelve_si_encontro_la_factura(libro):
    assert guardar_enlace(libro, "2777", "abc123") is True
    assert guardar_enlace(libro, "9999", "abc123") is False
