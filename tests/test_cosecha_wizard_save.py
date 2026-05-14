import os, shutil, pytest
from openpyxl import load_workbook
from excel_manager import ensure_cash_flow_sheets, COSECHAS_SHEET


@pytest.fixture
def m(tmp_path):
    src = os.path.join(os.path.dirname(__file__), "..", "..",
                        "MASTER Agricola Santa Elisa.xlsx")
    dst = tmp_path / "test.xlsx"
    shutil.copy2(src, dst)
    ensure_cash_flow_sheets(str(dst))
    # Limpiar Cosechas (datos reales preexistentes interfieren)
    wb = load_workbook(str(dst))
    ws = wb[COSECHAS_SHEET]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    wb.save(str(dst))
    wb.close()
    return str(dst)


def test_save_wizard_data_writes_rows(m):
    from modules.cash_flow.cosecha_wizard import save_to_cosechas
    data = {
        "cultivo": "NOGALES",
        "kg_total": 240000,
        "exportadoras": [{
            "nombre": "Valbifrut", "kg": 140000,
            "precio_usd_kg": 1.8, "n_cuotas": 1,
            "cuotas": [{"fecha": "2026-06-15", "usd": 252000}],
        }],
        "liquidacion": {"fecha": "2026-12-15", "usd": 50000},
    }
    rows_added = save_to_cosechas(data, year=2026, excel_path=m)
    assert rows_added == 2

    wb = load_workbook(m, read_only=True)
    ws = wb[COSECHAS_SHEET]
    found = False
    for r in range(2, ws.max_row + 1):
        if (ws.cell(r, 4).value == "Valbifrut"
            and ws.cell(r, 8).value == 1):
            assert ws.cell(r, 11).value == "adelanto"
            assert ws.cell(r, 10).value == 252000
            found = True
            break
    wb.close()
    assert found
