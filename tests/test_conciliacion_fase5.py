"""Fase 5: notas por movimiento, export a Excel y % conciliado.

Los tests pasan `excel_path` explícito: nunca tocar el Master real.
"""
from datetime import date

import pytest
from openpyxl import Workbook, load_workbook

import modules.conciliacion_comentarios as com
from modules.conciliacion_export import (a_excel, pct_conciliado_mes,
                                          recolectar, resumen)
from modules.conciliacion_store import HEADERS, SHEET

BANCO = ["Fecha", "Descripcion", "Referencia", "Cargo", "Abono", "Saldo",
         "Tipo", "Categoria", "Cultivo", "Factura_linkeada"]


@pytest.fixture
def libro(tmp_path):
    def _crear(movimientos=(), vinculos=()):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cuenta Banco"
        ws.append(BANCO)
        for m in movimientos:                 # (fecha, desc, cargo, abono, categoria)
            ws.append([m[0], m[1], "", m[2] or None, m[3] or None, None,
                       "", (m[4] if len(m) > 4 else ""), "", ""])
        wc = wb.create_sheet(SHEET)
        wc.append(HEADERS)
        for i, v in enumerate(vinculos, 1):   # (fila_banco, nro, prov, asignado)
            wc.append([i, "2026-08-01", v[0], "2026-07-01", "mov", 0,
                       "FACTURA", None, v[1], v[2], v[3], "manual", "", ""])
        p = tmp_path / "master.xlsx"
        wb.save(p); wb.close()
        return str(p)
    return _crear


# ── Notas por movimiento ─────────────────────────────────────────────────

def test_guardar_y_leer_una_nota(libro):
    p = libro([("2026-07-01", "TEF Don Antonio", 500_000, 0)])
    r = com.agregar(2, "Adelanto convenido", usuario="felix", excel_path=p)
    assert r["id"] == 1
    notas = com.de_movimiento(2, excel_path=p)
    assert len(notas) == 1
    assert notas[0]["texto"] == "Adelanto convenido"
    assert notas[0]["usuario"] == "felix"


def test_varias_notas_quedan_en_orden(libro):
    p = libro([("2026-07-01", "x", 1, 0)])
    for t in ("primera", "segunda", "tercera"):
        com.agregar(2, t, excel_path=p)
    assert [n["texto"] for n in com.de_movimiento(2, excel_path=p)] == [
        "primera", "segunda", "tercera"]


def test_las_notas_no_se_mezclan_entre_movimientos(libro):
    p = libro([("2026-07-01", "a", 1, 0), ("2026-07-02", "b", 2, 0)])
    com.agregar(2, "de la fila 2", excel_path=p)
    com.agregar(3, "de la fila 3", excel_path=p)
    assert com.de_movimiento(2, excel_path=p)[0]["texto"] == "de la fila 2"
    assert com.conteo_por_fila(excel_path=p) == {2: 1, 3: 1}


def test_nota_vacia_se_rechaza(libro):
    p = libro([("2026-07-01", "x", 1, 0)])
    for vacio in ("", "   ", None):
        with pytest.raises(ValueError):
            com.agregar(2, vacio, excel_path=p)


def test_nota_muy_larga_se_recorta(libro):
    p = libro([("2026-07-01", "x", 1, 0)])
    com.agregar(2, "a" * 900, excel_path=p)
    assert len(com.de_movimiento(2, excel_path=p)[0]["texto"]) == com.MAX_LARGO


def test_borrar_nota(libro):
    p = libro([("2026-07-01", "x", 1, 0)])
    com.agregar(2, "se va", excel_path=p)
    cid = com.de_movimiento(2, excel_path=p)[0]["id"]
    assert com.eliminar(cid, excel_path=p) is True
    assert com.de_movimiento(2, excel_path=p) == []


def test_sin_hoja_de_notas_no_revienta(libro):
    p = libro([("2026-07-01", "x", 1, 0)])
    assert com.de_movimiento(2, excel_path=p) == []
    assert com.conteo_por_fila(excel_path=p) == {}


# ── Estado y % conciliado ────────────────────────────────────────────────

def test_recolectar_marca_los_tres_estados(libro):
    p = libro([("2026-07-01", "conciliado", 100_000, 0),
               ("2026-07-02", "parcial", 200_000, 0),
               ("2026-07-03", "pendiente", 300_000, 0)],
              [(2, "F1", "X", 100_000), (3, "F2", "Y", 50_000)])
    estados = {f["desc"]: f["estado"] for f in recolectar("todos", excel_path=p)}
    assert estados == {"conciliado": "conciliado", "parcial": "parcial",
                       "pendiente": "pendiente"}


def test_filtrar_por_estado(libro):
    p = libro([("2026-07-01", "a", 100_000, 0), ("2026-07-02", "b", 200_000, 0)],
              [(2, "F1", "X", 100_000)])
    assert len(recolectar("conciliado", excel_path=p)) == 1
    assert len(recolectar("pendiente", excel_path=p)) == 1


def test_el_pct_cuenta_lo_asignado_parcial(libro):
    """Un movimiento a medias aporta lo asignado, no cero ni el total."""
    p = libro([("2026-07-01", "a", 100_000, 0), ("2026-07-02", "b", 100_000, 0)],
              [(2, "F1", "X", 100_000), (3, "F2", "Y", 50_000)])
    r = resumen(recolectar("todos", excel_path=p))
    assert r["monto_total"] == 200_000
    assert r["monto_conciliado"] == 150_000
    assert r["monto_pendiente"] == 50_000
    assert r["pct"] == 75.0


def test_pct_sin_movimientos_no_divide_por_cero(libro):
    assert resumen([])["pct"] == 0.0


def test_pct_del_mes_solo_toma_ese_mes(libro):
    p = libro([("2026-07-15", "de julio", 100_000, 0),
               ("2026-08-15", "de agosto", 999_999, 0)],
              [(2, "F1", "X", 100_000)])
    r = pct_conciliado_mes(2026, 7, excel_path=p)
    assert r["movimientos"] == 1
    assert r["pct"] == 100.0


def test_filtra_por_rango_de_fechas(libro):
    p = libro([("2026-06-15", "junio", 1_000, 0),
               ("2026-07-15", "julio", 2_000, 0),
               ("2026-08-15", "agosto", 3_000, 0)])
    filas = recolectar("todos", date(2026, 7, 1), date(2026, 7, 31), excel_path=p)
    assert [f["desc"] for f in filas] == ["julio"]


def test_ignora_movimientos_sin_monto(libro):
    p = libro([("2026-07-01", "sin monto", 0, 0), ("2026-07-02", "con monto", 500, 0)])
    assert [f["desc"] for f in recolectar("todos", excel_path=p)] == ["con monto"]


# ── Export ───────────────────────────────────────────────────────────────

def test_el_excel_sale_legible_y_con_los_datos(libro, tmp_path):
    p = libro([("2026-07-01", "Copeval ND506808 al ND506813", 713_590, 0, "INSUMOS")],
              [(2, "ND506808", "COPEVAL", 713_590)])
    buf = a_excel(recolectar("todos", excel_path=p), "Prueba")
    salida = tmp_path / "export.xlsx"
    salida.write_bytes(buf.getvalue())

    wb = load_workbook(salida)
    ws = wb.active
    assert ws["A1"].value == "Prueba"
    assert "100.0% conciliado" in ws["A2"].value
    encabezados = [c.value for c in ws[4]]
    assert "Estado" in encabezados and "Documentos" in encabezados
    fila = [c.value for c in ws[5]]
    assert "Copeval" in str(fila[1])
    assert "conciliado" in fila
    assert "ND506808" in str(fila[9])
    wb.close()


def test_el_excel_vacio_no_revienta(tmp_path):
    buf = a_excel([], "Sin datos")
    salida = tmp_path / "vacio.xlsx"
    salida.write_bytes(buf.getvalue())
    wb = load_workbook(salida)
    assert wb.active["A1"].value == "Sin datos"
    wb.close()
