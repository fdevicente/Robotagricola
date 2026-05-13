"""
vacaciones_manager.py — Gestión de Vacaciones y Personal
Hojas: Personal, Vacaciones
"""
import logging
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from config import EXCEL_PATH

logger = logging.getLogger(__name__)

PERSONAL_SHEET = "Personal"
VACACIONES_SHEET = "Vacaciones"

PERSONAL_HEADERS = [
    "Nombre", "RUT", "Cargo", "Fecha Ingreso",
    "Días Pendientes", "Días Tomados Total", "Última Vacación"
]
VACACIONES_HEADERS = [
    "Trabajador", "Fecha Inicio", "Fecha Fin", "Días",
    "Estado", "Observaciones"
]

# Chile: 15 días hábiles = ~21 corridos por año
DIAS_ANUALES = 15


def _open_wb():
    return load_workbook(EXCEL_PATH)


def _create_sheet(wb, name, headers, widths, color):
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=11)
        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.fill = fill
            c.font = font
            c.alignment = Alignment(horizontal="center")
            c.border = thin
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = w
    return wb[name]


def _ensure_personal(wb):
    return _create_sheet(wb, PERSONAL_SHEET, PERSONAL_HEADERS,
                         [28, 14, 22, 12, 16, 16, 12], "4A148C")


def _ensure_vacaciones(wb):
    return _create_sheet(wb, VACACIONES_SHEET, VACACIONES_HEADERS,
                         [28, 12, 12, 8, 12, 35], "6A1B9A")


def crear_hojas_vacaciones():
    """Crea las hojas Personal y Vacaciones."""
    wb = _open_wb()
    _ensure_personal(wb)
    _ensure_vacaciones(wb)
    wb.save(EXCEL_PATH)


def agregar_trabajador(nombre: str, rut: str = "", cargo: str = "",
                       fecha_ingreso: str = "") -> bool:
    """Agrega un trabajador a la hoja Personal."""
    wb = _open_wb()
    ws = _ensure_personal(wb)

    # Verificar que no exista
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and str(row[0]).strip().lower() == nombre.strip().lower():
            wb.close()
            return False

    if not fecha_ingreso:
        fecha_ingreso = date.today().strftime("%Y-%m-%d")

    # Calcular días pendientes según antigüedad
    try:
        fi = datetime.strptime(fecha_ingreso[:10], "%Y-%m-%d").date()
        meses = (date.today().year - fi.year) * 12 + date.today().month - fi.month
        dias_acumulados = round(DIAS_ANUALES * meses / 12)
    except Exception:
        dias_acumulados = 0

    ws.append([nombre, rut, cargo, fecha_ingreso, dias_acumulados, 0, ""])
    wb.save(EXCEL_PATH)
    logger.info(f"Trabajador agregado: {nombre}, {dias_acumulados} días pendientes")
    return True


def registrar_vacacion(nombre: str, fecha_inicio: str, fecha_fin: str,
                       observaciones: str = "") -> dict:
    """Registra vacaciones para un trabajador."""
    wb = _open_wb()
    ws_per = _ensure_personal(wb)
    ws_vac = _ensure_vacaciones(wb)

    # Calcular días
    try:
        fi = datetime.strptime(fecha_inicio[:10], "%Y-%m-%d").date()
        ff = datetime.strptime(fecha_fin[:10], "%Y-%m-%d").date()
        dias = (ff - fi).days + 1
        # Descontar fines de semana (aprox)
        dias_habiles = sum(1 for i in range(dias) if (fi.toordinal() + i) % 7 not in (5, 6))
    except Exception:
        dias = 0
        dias_habiles = 0

    # Registrar en hoja Vacaciones
    ws_vac.append([nombre, fecha_inicio, fecha_fin, dias_habiles, "Aprobado", observaciones])

    # Actualizar hoja Personal
    for row_idx in range(2, ws_per.max_row + 1):
        cell_nombre = ws_per.cell(row=row_idx, column=1).value
        if cell_nombre and str(cell_nombre).strip().lower() == nombre.strip().lower():
            pendientes = float(ws_per.cell(row=row_idx, column=5).value or 0)
            tomados = float(ws_per.cell(row=row_idx, column=6).value or 0)
            ws_per.cell(row=row_idx, column=5).value = max(0, pendientes - dias_habiles)
            ws_per.cell(row=row_idx, column=6).value = tomados + dias_habiles
            ws_per.cell(row=row_idx, column=7).value = fecha_inicio
            break

    wb.save(EXCEL_PATH)
    logger.info(f"Vacación registrada: {nombre}, {dias_habiles} días hábiles")
    return {"nombre": nombre, "inicio": fecha_inicio, "fin": fecha_fin,
            "dias_habiles": dias_habiles, "dias_corridos": dias}


def listar_personal() -> list[dict]:
    """Lista todo el personal con días pendientes."""
    wb = _open_wb()
    ws = _ensure_personal(wb)
    personal = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        personal.append({
            "nombre": str(row[0]),
            "rut": str(row[1] or ""),
            "cargo": str(row[2] or ""),
            "fecha_ingreso": str(row[3] or ""),
            "dias_pendientes": float(row[4] or 0),
            "dias_tomados": float(row[5] or 0),
            "ultima_vacacion": str(row[6] or ""),
        })
    wb.close()
    return personal


def vacaciones_pendientes() -> list[dict]:
    """Retorna trabajadores con vacaciones pendientes."""
    return [p for p in listar_personal() if p["dias_pendientes"] > 0]


def actualizar_dias_mensuales():
    """Acumula 1.25 días por trabajador (15/12 = 1.25 por mes).
    Debe ejecutarse mensualmente."""
    wb = _open_wb()
    ws = _ensure_personal(wb)
    actualizados = 0
    incremento = round(DIAS_ANUALES / 12, 2)  # 1.25

    for row_idx in range(2, ws.max_row + 1):
        nombre = ws.cell(row=row_idx, column=1).value
        if not nombre:
            continue
        pendientes = float(ws.cell(row=row_idx, column=5).value or 0)
        ws.cell(row=row_idx, column=5).value = round(pendientes + incremento, 1)
        actualizados += 1

    wb.save(EXCEL_PATH)
    logger.info(f"Vacaciones actualizadas: +{incremento} días a {actualizados} trabajadores")
    return {"actualizados": actualizados, "incremento": incremento}


def ultimas_vacaciones(n: int = 10) -> list[dict]:
    """Retorna las últimas N vacaciones registradas."""
    wb = _open_wb()
    ws = _ensure_vacaciones(wb)
    vacaciones = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        vacaciones.append({
            "trabajador": str(row[0]),
            "inicio": str(row[1] or ""),
            "fin": str(row[2] or ""),
            "dias": int(row[3] or 0),
            "estado": str(row[4] or ""),
            "observaciones": str(row[5] or ""),
        })
    wb.close()
    return vacaciones[-n:]
