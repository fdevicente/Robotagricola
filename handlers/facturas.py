"""handlers/facturas.py — Procesamiento de facturas: helpers internos.

Paso 1: solo helpers/constantes. Pasos siguientes agregan comandos y callbacks.
"""
import asyncio
import json
import logging
import os
import re
from datetime import datetime

from config import DOWNLOAD_DIR, BOLETAS_DIR, EXCEL_PATH, AUTO_SAVE_USERS

logger = logging.getLogger(__name__)


CAMPOS_EDITABLES = {
    "edit_proveedor": ("Nombre Factura / Proveedor",      "🏢 Nombre del proveedor"),
    "edit_rut":       ("Rut",                             "🪪 RUT del proveedor"),
    "edit_fecha":     ("Fecha Emision",                   "📅 Fecha de emisión (YYYY-MM-DD)"),
    "edit_vence":     ("Fecha Vencimiento",               "⏰ Fecha de vencimiento (YYYY-MM-DD)"),
    "edit_nro":       ("Numero Factura / Nro Documento",  "📄 Número de documento"),
    "edit_ref":       ("Referencia Factura",              "🔗 Nº factura referenciada (para NC/ND)"),
    "edit_glosa":     ("Detalle / Glosa",                 "📦 Glosa / descripción corta"),
    "edit_glosa2":    ("Glosa II",                        "📝 Detalle completo"),
    "edit_cantidad":  ("Cantidad",                        "🔢 Cantidad"),
    "edit_unitario":  ("Valor unitario",                  "💲 Valor unitario neto"),
    "edit_total":     ("TOTAL NETO",                      "💰 Total ítem sin IVA"),
}
# "edit_total_factura" se maneja aparte
CAMPOS_COMUNES = {"edit_proveedor", "edit_rut", "edit_fecha", "edit_vence", "edit_nro", "edit_ref"}
CAMPOS_POR_ITEM = {"edit_glosa", "edit_glosa2", "edit_cantidad", "edit_unitario", "edit_total"}


def _save_path(filename):
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    return os.path.join(DOWNLOAD_DIR, filename)


def _save_path_boleta(filename):
    os.makedirs(BOLETAS_DIR, exist_ok=True)
    return os.path.join(BOLETAS_DIR, filename)


def _es_boleta(items):
    """Detecta si los items corresponden a una boleta (caja chica).
    Boletas de Honorarios NO son caja chica — van a Facturas."""
    if not items:
        return False
    doc = str(items[0].get("Documento") or "").lower()
    return "boleta" in doc and "boleta de honorario" not in doc


def _renombrar_archivo(file_path: str, items: list) -> str:
    """Renombra el archivo usando Proveedor + Nº Documento."""
    try:
        item = items[0]
        proveedor = str(item.get("Nombre Factura / Proveedor") or "").strip()
        nro = str(item.get("Numero Factura / Nro Documento") or "").strip()
        if not proveedor and not nro:
            return file_path

        def _limpiar(s):
            s = re.sub(r'[\\/:*?"<>|]', '', s)
            s = re.sub(r'\s+', '_', s.strip())
            return s[:60]

        partes = []
        if proveedor:
            partes.append(_limpiar(proveedor))
        if nro:
            partes.append(nro)

        ext = os.path.splitext(file_path)[1]
        nombre = "_".join(partes) + ext
        dir_path = os.path.dirname(file_path)
        nuevo = os.path.join(dir_path, nombre)

        if os.path.exists(nuevo) and os.path.abspath(nuevo) != os.path.abspath(file_path):
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre = "_".join(partes) + f"_{ts}" + ext
            nuevo = os.path.join(dir_path, nombre)

        os.rename(file_path, nuevo)
        logger.info(f"Archivo renombrado: {os.path.basename(file_path)} → {nombre}")

        # Limpiar archivos derivados (_resized.jpg, _scan.png)
        base_viejo = os.path.splitext(file_path)[0]
        for sufijo in ("_resized.jpg", "_scan.png"):
            derivado = base_viejo + sufijo
            if os.path.exists(derivado):
                try:
                    os.remove(derivado)
                    logger.info(f"Derivado eliminado: {os.path.basename(derivado)}")
                except Exception as e:
                    logger.warning(f"No se pudo eliminar {derivado}: {e}")
        return nuevo
    except Exception as e:
        logger.error(f"No se pudo renombrar archivo {file_path}: {e}")
        return file_path


def encolar_documento(file_path: str, fecha_emision=None, cola=None,
                       tipo: str = "factura") -> None:
    """Encola la subida del documento a Drive. Nunca lanza.

    El archivo ya está guardado en disco cuando esto corre: si encolar falla,
    se pierde el enlace, no el documento.
    """
    import os
    from datetime import date
    try:
        if cola is None:
            from config import DRIVE_COLA_PATH, DRIVE_MAX_INTENTOS
            from modules.drive.cola import Cola
            cola = Cola(DRIVE_COLA_PATH, DRIVE_MAX_INTENTOS)
        if tipo == "boleta":
            carpeta = "Boletas Honorarios"
        else:
            anio = str(fecha_emision or "")[:4]
            if not anio.isdigit():
                anio = str(date.today().year)
            carpeta = "Facturas Recibidas/%s" % anio
        cola.encolar(file_path, carpeta, os.path.basename(file_path))
    except Exception as e:
        logger.warning("No pude encolar %s para Drive: %s", file_path, e)


def _registrar_correccion(item: dict, campo: str, valor_original, valor_nuevo):
    """Guarda la corrección del usuario para aprendizaje futuro."""
    try:
        log_path = os.path.join(DOWNLOAD_DIR, "correcciones_log.json")
        log = []
        if os.path.exists(log_path):
            with open(log_path, "r", encoding="utf-8") as f:
                log = json.load(f)
        entrada = {
            "timestamp": datetime.now().isoformat(),
            "rut":        item.get("Rut"),
            "proveedor":  item.get("Nombre Factura / Proveedor"),
            "nro_factura": item.get("Numero Factura / Nro Documento"),
            "campo":      campo,
            "valor_claude": valor_original,
            "valor_usuario": valor_nuevo,
        }
        if campo in ("Monto / TOTAL", "Total Factura", "Valor unitario"):
            try:
                orig = float(valor_original or 0)
                nuevo = float(valor_nuevo or 0)
                if orig > 0 and nuevo > 0:
                    entrada["factor"] = round(nuevo / orig, 6)
            except Exception:
                pass
        log.append(entrada)
        with open(log_path, "w", encoding="utf-8") as f:
            json.dump(log, f, ensure_ascii=False, indent=2)
        logger.info(f"Corrección registrada: {campo} | '{valor_original}' → '{valor_nuevo}' "
                    f"({item.get('Rut')})")
    except Exception as e:
        logger.warning(f"No se pudo registrar corrección: {e}")


def _rut_existe(rut):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb["Proveedores"]
        rut_n = rut.replace(".", "").replace("-", "").replace(" ", "").upper()
        for row in ws.iter_rows(min_row=3, values_only=True):
            r = str(row[2]).replace(".", "").replace("-", "").replace(" ", "").upper() if row[2] else ""
            if rut_n and rut_n == r:
                wb.close()
                return True
        wb.close()
    except Exception as e:
        logger.warning(f"Error verificando RUT: {e}")
    return False


def _agregar_proveedor(nombre, rut):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Proveedores"]
        ws.append([None, nombre, rut])
        wb.save(EXCEL_PATH)
        logger.info(f"Proveedor agregado: {nombre} — {rut}")
        return True
    except Exception as e:
        logger.error(f"Error agregando proveedor: {e}")
        return False


async def _download_with_retry(f, path, retries=3):
    """Descarga archivo de Telegram con reintentos."""
    for attempt in range(1, retries + 1):
        try:
            await f.download_to_drive(path)
            return True
        except Exception as e:
            logger.warning(f"Descarga intento {attempt}/{retries} falló: {e}")
            if attempt < retries:
                await asyncio.sleep(2 * attempt)
    return False


# ── Preview builder ──────────────────────────────

from utils.formatting import esc as _esc, format_date as _format_date, calc_vencimiento as _calc_vencimiento


def _build_preview(items):
    lines = ["📋 *Datos extraídos — revisa antes de guardar:*\n"]
    first = items[0]
    fecha_venc = first.get("Fecha Vencimiento") or _calc_vencimiento(first.get("Fecha Emision"))

    lines.append(f"🏢 *Proveedor:* {_esc(first.get('Nombre Factura / Proveedor') or '? no detectado')}")
    lines.append(f"🪪 *RUT:* {_esc(first.get('Rut') or '? no detectado')}")
    doc_tipo = _esc(first.get('Documento') or '—')
    doc_nro = _esc(first.get('Numero Factura / Nro Documento') or '—')
    lines.append(f"📄 *Documento:* {doc_tipo}  Nº {doc_nro}")
    ref = first.get("Referencia Factura")
    if ref:
        lines.append(f"🔗 *Ref. Factura:* Nº {_esc(ref)}")
    lines.append(f"📅 *Emisión:* {_format_date(first.get('Fecha Emision'))}   "
                  f"⏰ *Vence:* {_format_date(fecha_venc)}\n")

    doc = str(first.get('Documento') or '').lower()
    es_honorario = "boleta de honorario" in doc
    exenta = any(k in doc for k in ('exenta', 'exento', 'no afecta', 'no afecto')) or es_honorario

    pesos = []
    total_imp_esp = 0.0
    for item in items:
        unitario = float(item.get('Valor unitario') or 0)
        cantidad = float(item.get('Cantidad') or 1)
        pesos.append(unitario * cantidad)
        total_imp_esp += float(item.get('Impuesto Especifico') or 0)
    total_neto_raw = sum(pesos)

    total_factura = round(float(first.get('Total Factura') or 0))
    if total_factura > 0:
        base_iva = total_factura - round(total_imp_esp)
        if exenta:
            iva_total = 0
            neto_anchor = base_iva
        else:
            iva_total = round(base_iva / 1.19 * 0.19)
            neto_anchor = base_iva - iva_total
        total_con_iva = total_factura
    else:
        neto_anchor = round(total_neto_raw)
        iva_total = 0 if exenta else round(total_neto_raw * 0.19)
        total_con_iva = round(neto_anchor * (1.0 if exenta else 1.19) + total_imp_esp)

    n = len(items)
    netos_linea = [0] * n
    if n > 0:
        if total_neto_raw > 0:
            acumulado = 0
            for i, peso in enumerate(pesos):
                if i == n - 1:
                    netos_linea[i] = neto_anchor - acumulado
                else:
                    val = round(neto_anchor * peso / total_neto_raw)
                    netos_linea[i] = val
                    acumulado += val
        else:
            base = neto_anchor // n
            for i in range(n):
                netos_linea[i] = base
            netos_linea[-1] += neto_anchor - base * n

    for i, item in enumerate(items, 1):
        unitario = float(item.get('Valor unitario') or 0)
        cantidad = float(item.get('Cantidad') or 1)
        neto_linea = netos_linea[i - 1]
        if n > 1:
            lines.append(f"*— Ítem {i} —*")
        lines.append(f"📦 *Glosa:* {_esc(item.get('Detalle / Glosa') or '? no detectado')}")
        if item.get('Glosa II'):
            lines.append(f"📝 *Detalle:* {_esc(item.get('Glosa II'))}")
        lines.append(f"🔢 *Cantidad:* {cantidad:g}   💲 *Unit neto:* ${unitario:,.3f}")
        lines.append(f"💵 *Neto línea:* ${neto_linea:,.0f}\n")

    if es_honorario:
        pago_profesional = neto_anchor
        retencion = round(total_imp_esp)
        costo_total = total_con_iva if total_factura > 0 else pago_profesional + retencion
        lines.append(f"💵 *Pago al profesional:* ${pago_profesional:,.0f}")
        if retencion:
            lines.append(f"🏦 *Impto. Retenido (ASE→SII):* ${retencion:,.0f}")
        lines.append(f"💰 *COSTO TOTAL:* ${costo_total:,.0f}\n")
    else:
        lines.append(f"📊 *NETO:* ${neto_anchor:,.0f}")
        if not exenta:
            lines.append(f"📊 *IVA 19%:* ${iva_total:,.0f}")
        if total_imp_esp:
            lines.append(f"⛽ *Imp. Específico:* ${total_imp_esp:,.0f}")
        lines.append(f"💰 *TOTAL:* ${total_con_iva:,.0f}\n")
    lines.append("¿Qué deseas hacer?")
    return "\n".join(lines)


async def _show_preview(query, context):
    from utils.keyboards import main_keyboard
    items = context.user_data.get("pending_items", [])
    await query.edit_message_text(_build_preview(items), parse_mode="Markdown",
                                    reply_markup=main_keyboard())


# ── Handlers de archivos + procesamiento ─────────

from datetime import datetime
from processors.extractor import process_file
from excel_manager import (
    append_to_excel, append_boleta, consultar_saldo_caja,
)
from inventario_manager import agregar_stock_desde_factura
from utils.keyboards import main_keyboard, proveedor_nuevo_keyboard


def _hay_factura_en_proceso(context, ud=None) -> bool:
    """True si hay una factura esperando confirmación."""
    ud = context.user_data if ud is None else ud
    return bool(ud.get("pending_items"))


def _destino_factura(update, context):
    """Decide quién revisa la factura.

    Si la manda un usuario en AUTO_SAVE_USERS (Juan) y hay dueño registrado,
    la revisión se delega al dueño: el preview y los botones van a SU chat y
    los datos quedan en SU user_data (para que su callback los encuentre).

    Devuelve (chat_destino, user_data_destino, delegada, nombre_remitente).
    """
    uid = update.effective_user.id if update.effective_user else None
    nombre = update.effective_user.full_name if update.effective_user else "?"
    owner_chat = context.bot_data.get("owner_chat_id")
    owner_uid = context.bot_data.get("owner_user_id") or owner_chat

    if uid in AUTO_SAVE_USERS and owner_chat and uid != owner_uid:
        try:
            ud_owner = context.application.user_data[owner_uid]
            return owner_chat, ud_owner, True, nombre
        except Exception as e:
            logger.warning(f"No pude delegar factura al dueño: {e}")
    return update.effective_chat.id, context.user_data, False, nombre


async def _encolar_factura(context, chat_id, path, ud=None, remitente=""):
    """Agrega una factura ya descargada a la cola y avisa la posición."""
    ud = context.user_data if ud is None else ud
    cola = ud.setdefault("cola_facturas", [])
    cola.append(path)
    de = f" de {remitente}" if remitente else ""
    await context.bot.send_message(
        chat_id=chat_id,
        text=f"📥 Recibí otra factura{de}. Está en cola (#{len(cola)} en espera). "
             f"La procesaré apenas termines con la actual.")


async def _procesar_siguiente_de_cola(context, chat_id):
    """Si hay facturas en cola, toma la siguiente y la procesa.

    Se llama al confirmar o cancelar la factura actual.
    """
    cola = context.user_data.get("cola_facturas", [])
    if not cola:
        return
    path = cola.pop(0)
    restantes = len(cola)
    status = await context.bot.send_message(
        chat_id=chat_id,
        text=(f"🔍 Procesando siguiente factura de la cola"
              f"{f' ({restantes} más en espera)' if restantes else ''}… "
              f"(puede tardar hasta 1 minuto)"))
    try:
        await _process_and_reply(None, context, status, path)
    except Exception as e:
        logger.error(f"Error procesando factura en cola: {e}")
        try:
            await status.edit_text("❌ Error al procesar la factura en cola. "
                                    "Reenvíala manualmente.")
        except Exception:
            pass

    # Si la factura NO quedó en preview (falló extracción IA, sin items),
    # no habrá confirmación que dispare la siguiente: avanzar la cola ahora.
    if not context.user_data.get("pending_items"):
        await _procesar_siguiente_de_cola(context, chat_id)


async def handle_document(update, context):
    doc = update.message.document

    # Cartola del banco (CSV/TXT/Excel) → importador de movimientos
    from handlers.banco_upload import (es_archivo_cartola,
                                        parece_cartola_por_nombre,
                                        procesar_cartola)

    # Una cartola en PDF se leería como factura y crearía un registro falso:
    # se ataja antes de llegar al extractor.
    if (doc.mime_type == "application/pdf"
            and parece_cartola_por_nombre(doc.file_name)):
        await update.message.reply_text(
            "🏦 Esto parece una *cartola del banco en PDF*, y de ahí no puedo "
            "leer los movimientos de forma confiable.\n\n"
            "En el portal, al descargar la cartola elige **Excel, CSV o TXT** "
            "y mándamela así: te muestro el preview y la importo sin duplicar.\n\n"
            "_Si en realidad es una factura, cámbiale el nombre y reenvíala._",
            parse_mode="Markdown")
        return

    if es_archivo_cartola(doc.file_name):
        status = await update.message.reply_text("🏦 Recibí la cartola. Revisándola…")
        try:
            f = await context.bot.get_file(doc.file_id)
            path = _save_path(doc.file_name)
            if not await _download_with_retry(f, path):
                await status.edit_text("❌ No pude descargar el archivo.")
                return
            await procesar_cartola(update, context, path, status)
        except Exception as e:
            logger.error(f"Error en cartola: {e}")
            await status.edit_text(f"❌ Error al procesar la cartola: {str(e)[:150]}")
        return

    if doc.mime_type != "application/pdf":
        await update.message.reply_text(
            "❌ Solo acepto PDFs (facturas) o cartolas del banco (CSV/TXT/Excel).\n"
            "Para imágenes, envíalas como foto.")
        return
    chat_destino, ud, delegada, remitente = _destino_factura(update, context)
    ud["auto_mode"] = bool(
        not delegada and update.effective_user
        and update.effective_user.id in AUTO_SAVE_USERS)
    status = await update.message.reply_text("📥 Recibí tu PDF. Descargándolo...")
    try:
        f = await context.bot.get_file(doc.file_id)
        path = _save_path(doc.file_name)
        if not await _download_with_retry(f, path):
            await status.edit_text("❌ No pude descargar el PDF. Intenta enviarlo de nuevo.")
            return
        # Si hay una factura en proceso, encolar (ya descargada)
        if _hay_factura_en_proceso(context, ud):
            if delegada:
                await status.edit_text("📥 Factura recibida")
            else:
                await status.delete()
            await _encolar_factura(context, chat_destino, path, ud,
                                    remitente if delegada else "")
            return
        if delegada:
            await status.edit_text("📥 Factura recibida")
            status = await context.bot.send_message(
                chat_destino, f"🔍 Leyendo factura enviada por {remitente}… "
                              f"(puede tardar hasta 1 minuto)")
        else:
            await status.edit_text("🔍 Leyendo documento con IA… (puede tardar hasta 1 minuto)")
        await _process_and_reply(update, context, status, path, ud,
                                  prefijo=f"📨 *Factura enviada por {remitente}*\n\n" if delegada else "")
    except Exception as e:
        logger.error(f"Error en handle_document: {e}")
        try:
            await status.edit_text("❌ Error al procesar el PDF. Intenta de nuevo.")
        except Exception:
            pass


async def handle_photo(update, context):
    photo = update.message.photo[-1]

    # Una foto de horómetro NO es una factura: si se cuela al extractor, crea
    # un registro basura y gasta una llamada a la IA.
    from handlers.maquinaria import (modo_activo, parece_maquinaria,
                                      procesar_foto_horometro)
    if modo_activo(context) or parece_maquinaria(update.message.caption or ""):
        if await procesar_foto_horometro(update, context):
            return

    chat_destino, ud, delegada, remitente = _destino_factura(update, context)
    # Auto-guardado solo si NO hay dueño a quien delegar la revisión
    ud["auto_mode"] = bool(
        not delegada and update.effective_user
        and update.effective_user.id in AUTO_SAVE_USERS)
    status = await update.message.reply_text("📥 Recibí tu imagen. Descargándola...")
    try:
        f = await context.bot.get_file(photo.file_id)
        filename = f"factura_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{photo.file_unique_id}.jpg"
        path = _save_path(filename)
        if not await _download_with_retry(f, path):
            await status.edit_text("❌ No pude descargar la imagen. Intenta enviarla de nuevo.")
            return
        if _hay_factura_en_proceso(context, ud):
            if delegada:
                await status.edit_text("📥 Factura recibida")
            else:
                await status.delete()
            await _encolar_factura(context, chat_destino, path, ud,
                                    remitente if delegada else "")
            return
        if delegada:
            await status.edit_text("📥 Factura recibida")
            status = await context.bot.send_message(
                chat_destino, f"🔍 Leyendo factura enviada por {remitente}… "
                              f"(puede tardar hasta 1 minuto)")
        else:
            await status.edit_text("🔍 Leyendo documento con IA… (puede tardar hasta 1 minuto)")
        await _process_and_reply(update, context, status, path, ud,
                                  prefijo=f"📨 *Factura enviada por {remitente}*\n\n" if delegada else "")
    except Exception as e:
        logger.error(f"Error en handle_photo: {e}")
        try:
            await status.edit_text("❌ Error al procesar la imagen. Intenta de nuevo.")
        except Exception:
            pass


class _MsgAsQuery:
    """Adapta un Message para reusar _guardar_excel (que espera un callback query)."""
    def __init__(self, msg):
        self.message = msg

    async def edit_message_text(self, *args, **kwargs):
        return await self.message.edit_text(*args, **kwargs)


async def _process_and_reply(update, context, status_msg, file_path, ud=None, prefijo=""):
    """Extrae la factura y deja el preview listo para confirmar.

    `ud` permite dejar los datos en el user_data de OTRO usuario (el dueño),
    cuando la factura la mandó Juan y la revisión se delega.
    """
    ud = context.user_data if ud is None else ud
    result = await asyncio.to_thread(process_file, file_path)
    if result.get("status") == "error":
        await status_msg.edit_text(f"❌ {result.get('message')}", parse_mode="Markdown")
        return
    items = result.get("items", [])
    if not items:
        await status_msg.edit_text("⚠️ No se identificaron datos. Intenta con foto más nítida.")
        return

    file_path = _renombrar_archivo(file_path, items)
    encolar_documento(file_path,
                      fecha_emision=(items[0].get("Fecha Emision")
                                      if items else None))
    ud["pending_items"] = items
    ud["pending_file_path"] = file_path
    ud["editing_field"] = None

    # ── MODO CAPATAZ: guardar directo sin preview/confirmación (la cola no se atora) ──
    if ud.get("auto_mode"):
        first = items[0]
        nombre = first.get("Nombre Factura / Proveedor")
        rut = first.get("Rut")
        try:
            if nombre and rut and not await asyncio.to_thread(_rut_existe, rut):
                await asyncio.to_thread(_agregar_proveedor, nombre, rut)
        except Exception as e:
            logger.warning(f"Auto-factura: no pude verificar/agregar proveedor: {e}")
        await _guardar_excel(_MsgAsQuery(status_msg), context, items, file_path)
        return

    preview = prefijo + _build_preview(items)
    if result.get("duplicado"):
        preview = ("⚠️ *ADVERTENCIA: Esta factura parece ya estar registrada.*\n"
                    "Revisa bien antes de guardar.\n\n" + preview)

    try:
        await status_msg.edit_text(preview, parse_mode="Markdown",
                                     reply_markup=main_keyboard())
    except Exception:
        await status_msg.edit_text(preview, reply_markup=main_keyboard())


async def _guardar_excel(query, context, items, file_path):
    es_boleta = _es_boleta(items)
    await query.edit_message_text("💾 Guardando en el Excel…")
    try:
        if es_boleta:
            success = await asyncio.to_thread(append_boleta, items)
            if success and file_path and os.path.exists(file_path):
                import shutil
                dest = _save_path_boleta(os.path.basename(file_path))
                shutil.move(file_path, dest)
                file_path = dest
        else:
            success = await asyncio.to_thread(append_to_excel, items)
            if success:
                try:
                    agregados = await asyncio.to_thread(agregar_stock_desde_factura, items)
                    if agregados:
                        logger.info(f"Inventario: {len(agregados)} insumos agregados desde factura")
                except Exception as e:
                    logger.warning(f"Error auto-inventario: {e}")
    except Exception as e:
        logger.error(e)
        success = False

    if success:
        context.user_data["last_invoice_file"] = file_path
        context.user_data["last_invoice_rows"] = len(items)
        context.user_data["last_invoice_boleta"] = es_boleta
        context.user_data["pending_items"] = []
        first = items[0]
        proveedor = first.get("Nombre Factura / Proveedor") or "Desconocido"
        nro = first.get("Numero Factura / Nro Documento") or "S/N"
        total_factura_anchor = float(first.get("Total Factura") or 0)
        total = (round(total_factura_anchor) if total_factura_anchor > 0
                  else sum(float(i.get("Monto / TOTAL") or 0) for i in items))
        if es_boleta:
            info_caja = await asyncio.to_thread(consultar_saldo_caja)
            saldo_txt = f"💰 Saldo caja chica: *${info_caja['saldo']:,.0f}*"
            alerta = "\n⚠️ Saldo bajo!" if info_caja['saldo'] < 100000 else ""
            await query.edit_message_text(
                f"✅ *Boleta guardada (Caja Chica)*\n\n"
                f"🏪 {_esc(proveedor)}\n🧾 Nº {_esc(nro)}  —  📤 ${total:,.0f}\n"
                f"{saldo_txt}{alerta}\n\n_Usa /deshacer si hay algún error_",
                parse_mode="Markdown")
        else:
            # El tipo de documento real (boleta de honorarios, NC, ND…) va en el mensaje
            doc = str(first.get("Documento") or "").strip()
            doc_l = doc.lower()
            if "boleta de honorario" in doc_l:
                titulo, icono = "Boleta de honorarios guardada", "🧾"
            elif "nota de credito" in doc_l or "nota de crédito" in doc_l:
                titulo, icono = "Nota de crédito guardada", "↩️"
            elif "nota de debito" in doc_l or "nota de débito" in doc_l:
                titulo, icono = "Nota de débito guardada", "↪️"
            elif doc and "factura" not in doc_l:
                titulo, icono = f"{doc.capitalize()} guardado", "📄"
            else:
                titulo, icono = "Factura guardada", "📄"

            # N° de archivo físico, para buscar el papel impreso
            n_arch = ""
            try:
                from modules.correlativo import COL_CORRELATIVO
                import openpyxl
                _wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
                _ws = _wb["Facturas"]
                _v = _ws.cell(_ws.max_row, COL_CORRELATIVO).value
                _wb.close()
                if _v:
                    n_arch = f"🗂️ N° de archivo: *{_v}*\n"
            except Exception as e:
                logger.warning(f"No pude leer el N° de archivo: {e}")

            await query.edit_message_text(
                f"✅ *{titulo}*\n\n"
                f"🏢 {_esc(proveedor)}\n{icono} Nº {_esc(nro)}  —  💰 ${total:,.0f}\n"
                f"{n_arch}"
                f"📊 {len(items)} fila(s) en el Excel\n\n_Usa /deshacer si hay algún error_",
                parse_mode="Markdown")
            # Si es proveedor de insumos, encolar productos para fecha de vencimiento
            try:
                from vencimientos_manager import es_proveedor_insumo, agregar_pendiente
                if es_proveedor_insumo(str(proveedor)):
                    fecha_compra = first.get("Fecha Emision")
                    for it in items:
                        prod = str(it.get("Detalle / Glosa") or "").strip()
                        if prod:
                            await asyncio.to_thread(
                                agregar_pendiente, prod, str(proveedor),
                                str(nro), fecha_compra)
            except Exception as e:
                logger.warning(f"No pude registrar pendientes de vencimiento: {e}")
    else:
        await query.edit_message_text(
            "❌ Error al guardar. ¿Está el Excel abierto en otro programa?")

    # Tras guardar (o fallar), procesar la siguiente factura de la cola
    try:
        chat_id = query.message.chat_id
        if context.user_data.get("cola_facturas"):
            await _procesar_siguiente_de_cola(context, chat_id)
        else:
            # Lote terminado: recordar pendientes de fecha de vencimiento
            from handlers.vencimientos import recordatorio_pendientes
            await recordatorio_pendientes(context, chat_id)
    except Exception as e:
        logger.warning(f"No pude procesar siguiente de cola: {e}")


# ── Callbacks de edición ─────────────────────────

from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from utils.keyboards import edit_keyboard


async def cb_confirm_save(update, context):
    query = update.callback_query
    await query.answer()
    items = context.user_data.get("pending_items", [])
    file_path = context.user_data.get("pending_file_path")
    if not items:
        await query.edit_message_text("⚠️ No hay datos pendientes.")
        return

    first = items[0]
    nombre = first.get("Nombre Factura / Proveedor")
    rut = first.get("Rut")
    if nombre and rut and not _rut_existe(rut):
        context.user_data["nuevo_proveedor_nombre"] = nombre
        context.user_data["nuevo_proveedor_rut"] = rut
        await query.edit_message_text(
            f"🆕 *{nombre}* (RUT: {rut}) no está en tu lista de proveedores.\n\n¿Lo agregamos?",
            parse_mode="Markdown", reply_markup=proveedor_nuevo_keyboard())
        return
    await _guardar_excel(query, context, items, file_path)


async def cb_add_proveedor_yes(update, context):
    query = update.callback_query
    await query.answer()
    nombre = context.user_data.get("nuevo_proveedor_nombre")
    rut = context.user_data.get("nuevo_proveedor_rut")
    if nombre and rut:
        ok = await asyncio.to_thread(_agregar_proveedor, nombre, rut)
        if not ok:
            await query.answer("⚠️ No se pudo agregar", show_alert=True)
    await _guardar_excel(query, context,
                          context.user_data.get("pending_items", []),
                          context.user_data.get("pending_file_path"))


async def cb_add_proveedor_no(update, context):
    query = update.callback_query
    await query.answer()
    await _guardar_excel(query, context,
                          context.user_data.get("pending_items", []),
                          context.user_data.get("pending_file_path"))


async def cb_cancel_save(update, context):
    query = update.callback_query
    await query.answer()
    context.user_data["pending_items"] = []
    await query.edit_message_text("🚫 Factura descartada. Mándame otra cuando quieras.")
    # Procesar siguiente de la cola si hay
    await _procesar_siguiente_de_cola(context, update.effective_chat.id)


async def cb_edit_menu(update, context):
    query = update.callback_query
    await query.answer()
    context.user_data["editing_field"] = None
    context.user_data["editing_item_idx"] = None
    items = context.user_data.get("pending_items", [])
    sufijo = "s" if len(items) != 1 else ""
    texto = f"✏️ *¿Qué campo quieres corregir?* ({len(items)} ítem{sufijo})"
    await query.edit_message_text(texto, parse_mode="Markdown",
                                    reply_markup=edit_keyboard(items))


async def cb_edit_field(update, context):
    query = update.callback_query
    await query.answer()
    campo_key = query.data
    campo_excel, campo_label = CAMPOS_EDITABLES[campo_key]
    items = context.user_data.get("pending_items", [])

    if len(items) > 1 and campo_key in CAMPOS_POR_ITEM:
        buttons = []
        for i, item in enumerate(items):
            glosa = _esc(item.get("Detalle / Glosa") or f"Ítem {i+1}")[:30]
            buttons.append([InlineKeyboardButton(
                f"Ítem {i+1}: {glosa}", callback_data=f"selitem_{i}_{campo_key}")])
        buttons.append([InlineKeyboardButton("« Volver", callback_data="edit_menu")])
        await query.edit_message_text(
            f"✏️ *{campo_label}* — ¿En qué ítem?",
            parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(buttons))
        return

    context.user_data["editing_field"] = campo_excel
    context.user_data["editing_field_label"] = campo_label
    context.user_data["editing_item_idx"] = None
    _back_btn = InlineKeyboardMarkup([[InlineKeyboardButton("« Volver", callback_data="edit_menu")]])
    await query.edit_message_text(
        f"✏️ Escribe el nuevo valor para *{campo_label}*:",
        parse_mode="Markdown", reply_markup=_back_btn)


async def cb_select_item(update, context):
    query = update.callback_query
    await query.answer()
    parts = query.data.split("_")  # selitem_0_edit_glosa
    idx = int(parts[1])
    campo_key = "_".join(parts[2:])
    campo_excel, campo_label = CAMPOS_EDITABLES[campo_key]
    items = context.user_data.get("pending_items", [])
    glosa = _esc(items[idx].get("Detalle / Glosa") or f"Ítem {idx+1}")[:30]

    context.user_data["editing_field"] = campo_excel
    context.user_data["editing_field_label"] = campo_label
    context.user_data["editing_item_idx"] = idx
    _back_btn = InlineKeyboardMarkup([[InlineKeyboardButton("« Volver", callback_data="edit_menu")]])
    await query.edit_message_text(
        f"✏️ Escribe el nuevo valor para *{campo_label}* (Ítem {idx+1}: {glosa}):",
        parse_mode="Markdown", reply_markup=_back_btn)


async def cb_edit_total_factura(update, context):
    query = update.callback_query
    await query.answer()
    items = context.user_data.get("pending_items", [])
    total_neto = sum(float(i.get("Valor unitario") or 0) * float(i.get("Cantidad") or 1)
                      for i in items)
    total_actual = round(total_neto * 1.19)
    context.user_data["editing_field"] = "_total_factura"
    context.user_data["editing_field_label"] = "💰 TOTAL FACTURA"
    context.user_data["editing_item_idx"] = None
    _back_btn = InlineKeyboardMarkup([[InlineKeyboardButton("« Volver", callback_data="edit_menu")]])
    await query.edit_message_text(
        f"💰 *Total factura actual:* ${total_actual:,.0f}\n\n"
        f"Escribe el nuevo *total con IVA* de toda la factura:",
        parse_mode="Markdown", reply_markup=_back_btn)


async def cb_back_preview(update, context):
    query = update.callback_query
    await query.answer()
    await _show_preview(query, context)


async def cb_add_item(update, context):
    """Agrega un item vacio a la factura pendiente."""
    query = update.callback_query
    await query.answer()
    items = context.user_data.get("pending_items", [])
    if not items:
        await query.edit_message_text("⚠️ No hay factura pendiente.")
        return
    first = items[0]
    COMUNES = ("Fecha Emision", "Fecha Vencimiento", "Fecha Pago",
               "Nombre Factura / Proveedor", "Rut", "Documento",
               "Numero Factura / Nro Documento", "Referencia Factura",
               "Total Factura")
    nuevo = {k: first.get(k) for k in COMUNES}
    nuevo["Detalle / Glosa"] = ""
    nuevo["Glosa II"] = ""
    nuevo["Valor unitario"] = 0
    nuevo["Cantidad"] = 1
    nuevo["Impuesto Especifico"] = 0
    nuevo["Monto / TOTAL"] = 0
    items.append(nuevo)
    context.user_data["pending_items"] = items
    idx = len(items)
    context.user_data["editing_field"] = "Detalle / Glosa"
    context.user_data["editing_field_label"] = "📦 Glosa / descripción corta"
    context.user_data["editing_item_idx"] = idx - 1
    _back_btn = InlineKeyboardMarkup([[InlineKeyboardButton("« Volver", callback_data="edit_menu")]])
    await query.edit_message_text(
        f"➕ *Ítem {idx} agregado*\n\n"
        f"Escribe la *glosa / descripción* del nuevo ítem\n"
        f"(luego podrás editar Cantidad, Unit neto y Total ítem con el menú de edición):",
        parse_mode="Markdown", reply_markup=_back_btn)


async def cb_del_item_menu(update, context):
    query = update.callback_query
    items = context.user_data.get("pending_items", [])
    if len(items) <= 1:
        await query.answer("⚠️ No puedes eliminar el único ítem.", show_alert=True)
        return
    await query.answer()
    buttons = []
    for i, item in enumerate(items):
        glosa = _esc(item.get("Detalle / Glosa") or f"Ítem {i+1}")[:30]
        buttons.append([InlineKeyboardButton(
            f"🗑️ Ítem {i+1}: {glosa}", callback_data=f"delitem_{i}")])
    buttons.append([InlineKeyboardButton("« Volver", callback_data="edit_menu")])
    await query.edit_message_text(
        "🗑️ *¿Qué ítem quieres eliminar?*",
        parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(buttons))


async def cb_del_item_confirm(update, context):
    query = update.callback_query
    await query.answer()
    try:
        idx = int(query.data.split("_")[1])
    except (ValueError, IndexError):
        await query.edit_message_text("❌ Índice de ítem inválido.")
        return
    items = context.user_data.get("pending_items", [])
    if 0 <= idx < len(items):
        items.pop(idx)
        context.user_data["pending_items"] = items
        nuevo_tf = round(sum(float(it.get("Monto / TOTAL") or 0) for it in items))
        if nuevo_tf > 0:
            for it in items:
                it["Total Factura"] = nuevo_tf
    await _show_preview(query, context)


# ── Flujo de edición por texto ───────────────────


async def handle_text_edit_factura(update, context) -> bool:
    """Procesa texto si hay edicion de factura pendiente. True si lo manejó."""
    campo = context.user_data.get("editing_field")
    if not campo:
        return False

    nuevo = update.message.text.strip()
    items = context.user_data.get("pending_items", [])
    if not items:
        await update.message.reply_text("⚠️ No hay factura pendiente.")
        return True

    idx = context.user_data.get("editing_item_idx")

    if campo == "_total_factura":
        try:
            n = nuevo.replace("$", "").replace(" ", "")
            if "," in n:
                n = n.replace(".", "").replace(",", ".")
            elif n.count(".") > 1:
                n = n.replace(".", "")
            total_nuevo = float(n)
        except ValueError:
            await update.message.reply_text(
                f"❌ '{nuevo}' no es un número válido. Intenta de nuevo.")
            return True
        valor_orig_tf = items[0].get("Total Factura") or items[0].get("Monto / TOTAL")
        _registrar_correccion(items[0], "Total Factura", valor_orig_tf, total_nuevo)
        doc = str(items[0].get("Documento") or "").lower()
        exenta = any(k in doc for k in ("exenta", "exento", "no afecta", "no afecto"))
        iva_factor = 1.0 if exenta else 1.19
        neto_nuevo = total_nuevo / iva_factor
        total_neto_actual = sum(float(i.get("Valor unitario") or 0) * float(i.get("Cantidad") or 1)
                                  for i in items)
        if total_neto_actual > 0:
            factor = neto_nuevo / total_neto_actual
            for item in items:
                unit = float(item.get("Valor unitario") or 0)
                qty = float(item.get("Cantidad") or 1)
                nuevo_unit = round(unit * factor, 2)
                item["Valor unitario"] = nuevo_unit
                item["Monto / TOTAL"] = round(nuevo_unit * qty * iva_factor)
        else:
            items[0]["Monto / TOTAL"] = total_nuevo
            items[0]["Valor unitario"] = round(neto_nuevo)
        for item in items:
            item["Total Factura"] = round(total_nuevo)
        context.user_data["total_override"] = total_nuevo
        context.user_data["editing_field"] = None
        context.user_data["editing_item_idx"] = None
        await update.message.reply_text(
            f"✅ *Total factura* actualizado a `${total_nuevo:,.0f}`",
            parse_mode="Markdown")
        try:
            await update.message.reply_text(_build_preview(items),
                                              parse_mode="Markdown",
                                              reply_markup=main_keyboard())
        except Exception:
            await update.message.reply_text(_build_preview(items),
                                              reply_markup=main_keyboard())
        return True

    targets = [items[idx]] if idx is not None else items

    NUMERICOS = {"Valor unitario", "Cantidad", "Monto / TOTAL", "TOTAL NETO"}
    for item in targets:
        valor_original = item.get(campo)
        if campo in NUMERICOS:
            try:
                n = nuevo.replace("$", "").replace(" ", "")
                if "," in n:
                    n = n.replace(".", "").replace(",", ".")
                else:
                    if n.count(".") > 1:
                        n = n.replace(".", "")
                val = float(n)
                item[campo] = val
                if campo == "Monto / TOTAL":
                    qty = float(item.get("Cantidad") or 1)
                    imp_esp = float(item.get("Impuesto Especifico") or 0)
                    doc = str(item.get("Documento") or "").lower()
                    sin_iva = any(k in doc for k in ("exenta", "exento", "no afecta",
                                                       "no afecto", "boleta de honorario"))
                    iva_factor = 1.0 if sin_iva else 1.19
                    neto_nuevo = (val - imp_esp) / iva_factor
                    item["Valor unitario"] = neto_nuevo / qty if qty else 0
                elif campo == "TOTAL NETO":
                    qty = float(item.get("Cantidad") or 1)
                    imp_esp = float(item.get("Impuesto Especifico") or 0)
                    doc = str(item.get("Documento") or "").lower()
                    sin_iva = any(k in doc for k in ("exenta", "exento", "no afecta",
                                                       "no afecto", "boleta de honorario"))
                    iva_factor = 1.0 if sin_iva else 1.19
                    item["Valor unitario"] = val / qty if qty else 0
                    item["Monto / TOTAL"] = round(val * iva_factor + imp_esp)
            except ValueError:
                await update.message.reply_text(
                    f"❌ '{nuevo}' no es un número válido. Intenta de nuevo.")
                return True
        else:
            item[campo] = nuevo
        if str(valor_original) != str(item.get(campo)):
            _registrar_correccion(item, campo, valor_original, item.get(campo))

    if campo in NUMERICOS:
        doc0 = str(items[0].get("Documento") or "").lower()
        sin_iva0 = any(k in doc0 for k in ("exenta", "exento", "no afecta",
                                            "no afecto", "boleta de honorario"))
        iva_factor0 = 1.0 if sin_iva0 else 1.19
        if campo not in ("Monto / TOTAL", "TOTAL NETO"):
            for it in targets:
                u = float(it.get("Valor unitario") or 0)
                q = float(it.get("Cantidad") or 1)
                imp = float(it.get("Impuesto Especifico") or 0)
                it["Monto / TOTAL"] = round(u * q * iva_factor0 + imp)
        nuevo_tf = round(sum(float(it.get("Monto / TOTAL") or 0) for it in items))
        if nuevo_tf > 0:
            for it in items:
                it["Total Factura"] = nuevo_tf

    context.user_data["editing_field"] = None
    context.user_data["editing_item_idx"] = None
    label = context.user_data.get("editing_field_label", campo)
    suffix = f" (Ítem {idx+1})" if idx is not None else ""
    await update.message.reply_text(f"✅ *{label}*{suffix} actualizado.",
                                      parse_mode="Markdown")
    try:
        await update.message.reply_text(_build_preview(items),
                                          parse_mode="Markdown",
                                          reply_markup=main_keyboard())
    except Exception:
        await update.message.reply_text(_build_preview(items),
                                          reply_markup=main_keyboard())
    return True
