"""handlers/monitoreo.py — Monitoreo del bot y visibilidad de actividad.

- /soydueno : el dueño lo corre UNA vez desde su chat → fija owner_chat_id.
- Mirror    : reenvía al dueño todo lo que mandan los demás (texto/fotos/PDF).
- Heartbeat : mensaje diario "sigo vivo" con resumen (si deja de llegar = algo se cayó).
- /estado   : estado on-demand.
"""
import asyncio
import logging
from datetime import date, datetime, timezone

logger = logging.getLogger(__name__)


def _pd(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for f in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(v[:10], f).date()
            except Exception:
                pass
    return None


def _resumen_dia() -> dict:
    """Lee bitácora de hoy + última factura del Excel. SYNC (usar en to_thread)."""
    from openpyxl import load_workbook
    from config import EXCEL_PATH
    hoy = date.today()
    out = {"bita_hoy": 0, "ultima_bita": "", "ult_fact": "", "fact_total": 0}
    try:
        wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        if "Bitácora" in wb.sheetnames:
            for row in wb["Bitácora"].iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                if _pd(row[0]) == hoy:
                    out["bita_hoy"] += 1
                    act = str(row[3] or row[2] or "").strip()
                    out["ultima_bita"] = f"{act} ({row[1] or ''})"
        if "Facturas" in wb.sheetnames:
            last = None
            n = 0
            for row in wb["Facturas"].iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    last = row
                    n += 1
            out["fact_total"] = n
            if last:
                out["ult_fact"] = f"{last[3]} F{last[6]}"
        wb.close()
    except Exception as e:
        logger.warning(f"_resumen_dia: {e}")
    return out


async def _build_estado(context) -> str:
    import asyncio
    from infrastructure import bot_state
    r = await asyncio.to_thread(_resumen_dia)

    # Heartbeat / última actividad
    act_txt = "—"
    try:
        st = bot_state.cargar_estado()
        ts = st.get("ultima_actividad_utc")
        if ts:
            t = datetime.fromisoformat(ts)
            delta = datetime.now(timezone.utc) - t
            horas = delta.total_seconds() / 3600
            if horas < 1:
                cuando = f"hace {int(delta.total_seconds()/60)} min"
            elif horas < 48:
                cuando = f"hace {int(horas)} h"
            else:
                cuando = f"hace {int(horas/24)} días"
            res = str(st.get("ultimo_resumen") or "").replace("\n", " ")[:50]
            act_txt = f"{cuando}" + (f" — {res}" if res else "")
    except Exception as e:
        logger.warning(f"estado actividad: {e}")

    ahora = datetime.now().strftime("%d-%m-%Y %H:%M")
    lines = [f"✅ *Bot funcionando* — {ahora}",
             f"🕐 Última actividad: {act_txt}"]
    if r["bita_hoy"]:
        lines.append(f"📓 Bitácora hoy: {r['bita_hoy']} registro(s) · última: {r['ultima_bita']}")
    else:
        lines.append("📓 Bitácora hoy: sin registros")
    if r["ult_fact"]:
        lines.append(f"📄 Última factura: {r['ult_fact']}  (total {r['fact_total']})")
    return "\n".join(lines)


async def cmd_soydueno(update, context):
    """El dueño lo corre desde SU chat para recibir heartbeat + mirror."""
    context.bot_data["owner_chat_id"] = update.effective_chat.id
    if update.effective_user:
        context.bot_data["owner_user_id"] = update.effective_user.id
    await update.message.reply_text(
        "✅ *Listo, te registré como dueño.*\n\n"
        "Desde ahora recibirás:\n"
        "• 🔁 *Mirror*: todo lo que escriban/manden los demás (Juan, etc.)\n"
        "• 📄 *Facturas de Juan*: él solo la manda, y la revisión/confirmación te llega a ti\n"
        "• ❤️ *Heartbeat diario*: un aviso de que sigo funcionando (si deja de llegar, algo se cayó)\n\n"
        "Usa /estado cuando quieras ver el estado al instante.",
        parse_mode="Markdown")
    logger.info(f"owner_chat_id fijado: {update.effective_chat.id}")


async def _send_md(bot, chat_id, texto):
    """Envía con Markdown; si el parseo falla (nombres con _ o *), manda plano."""
    try:
        await bot.send_message(chat_id, texto, parse_mode="Markdown")
    except Exception:
        await bot.send_message(chat_id, texto)


async def cmd_estado(update, context):
    txt = await _build_estado(context)
    owner = context.bot_data.get("owner_chat_id")
    extra = "" if owner else "\n\n⚠️ Aún no corres /soydueno — el heartbeat y el mirror no tienen destino."
    await _send_md(context.bot, update.effective_chat.id, txt + extra)


async def mirror_update(update, context):
    """TypeHandler (group propio): reenvía al dueño lo que mandan los demás.

    Solo mensajes NUEVOS reales (update.message) — no callbacks ni ediciones,
    que traerían mensajes del propio bot como effective_message.
    """
    owner = context.bot_data.get("owner_chat_id")
    if not owner:
        return
    if not update.effective_chat or update.effective_chat.id == owner:
        return  # no espejar el propio chat del dueño
    msg = update.message
    if not msg:
        return
    nombre = update.effective_user.full_name if update.effective_user else "?"
    try:
        if msg.text:
            await context.bot.send_message(owner, f"👤 {nombre}:\n{msg.text}")
        else:
            cap = f"👤 {nombre}" + (f": {msg.caption}" if msg.caption else " envió esto")
            await context.bot.copy_message(chat_id=owner,
                                           from_chat_id=update.effective_chat.id,
                                           message_id=msg.message_id, caption=cap)
    except Exception as e:
        logger.warning(f"mirror: {e}")


async def cmd_bodega(update, context):
    """/bodega — contrasta el Excel de bodega del equipo contra el Master."""
    status = await update.message.reply_text("📦 Revisando el Excel de bodega…")
    from modules.bodega_check import comparar, formato_alerta
    try:
        res = await asyncio.to_thread(comparar)
    except FileNotFoundError as e:
        await status.edit_text(f"❌ {e}")
        return
    except Exception as e:
        logger.error(f"Chequeo bodega: {e}")
        await status.edit_text(f"❌ No pude revisar la bodega: {str(e)[:150]}")
        return
    txt = formato_alerta(res)
    await status.edit_text(txt[:4000])


async def cmd_correlativos(update, context):
    """/correlativos — toma de FXP los N° de archivo y los pone en el Master."""
    status = await update.message.reply_text("🗂️ Sincronizando N° de archivo desde FXP…")
    from modules.correlativo import sincronizar_desde_fxp
    try:
        r = await asyncio.to_thread(sincronizar_desde_fxp)
    except Exception as e:
        logger.error(f"Correlativos: {e}")
        await status.edit_text(f"❌ No pude sincronizar: {str(e)[:150]}")
        return
    lineas = [f"🗂️ *N° de archivo sincronizados*",
              f"➕ Nuevos asignados: {r['nuevos']}",
              f"✏️ Corregidos: {r['corregidos']}",
              f"⏳ Aún sin número (no están en FXP): {r['sin_numero']}"]
    if r["detalle"]:
        lineas.append("")
        for n, prov, nro in r["detalle"][:15]:
            lineas.append(f"  N°{n} — {prov} F{nro}")
        if len(r["detalle"]) > 15:
            lineas.append(f"  … y {len(r['detalle'])-15} más")
    await _send_md(context.bot, update.effective_chat.id, "\n".join(lineas)[:4000])


async def cmd_basedatos(update, context):
    """/basedatos — sincroniza el Excel a la base y verifica que calcen."""
    status = await update.message.reply_text("🗄️ Sincronizando Excel → base de datos…")
    try:
        from modules.db.sync_excel import sincronizar
        from modules.db.verificar import comparar, formato
        r = await asyncio.to_thread(sincronizar, None, False)
        v = await asyncio.to_thread(comparar)
    except Exception as e:
        logger.error(f"Sync DB: {e}")
        await status.edit_text(f"❌ Error: {str(e)[:200]}")
        return
    filas = " · ".join(f"{k}: {n}" for k, n in r.items())
    await status.edit_text(f"🗄️ *Base de datos actualizada*\n{filas}\n\n"
                            f"```\n{formato(v)[:2600]}\n```",
                            parse_mode="Markdown")


async def job_sync_db(context):
    """Job diario: sincroniza la base y avisa SOLO si deja de calzar."""
    try:
        from modules.db.sync_excel import sincronizar
        from modules.db.verificar import comparar, formato
        await asyncio.to_thread(sincronizar, None, False)
        v = await asyncio.to_thread(comparar)
    except Exception as e:
        logger.warning(f"Job sync DB: {e}")
        owner = context.bot_data.get("owner_chat_id")
        if owner:
            await context.bot.send_message(
                owner, f"⚠️ Falló la sincronización con la base: {str(e)[:150]}")
        return
    if v["ok"]:
        logger.info("Sync DB diaria: todo calza.")
        return
    owner = context.bot_data.get("owner_chat_id")
    if owner:
        await _send_md(context.bot, owner,
                       "⚠️ *La base de datos no calza con el Excel*\n\n"
                       f"```\n{formato(v)[:2600]}\n```")


async def job_bodega_check(context):
    """Job semanal: avisa al dueño solo si la data NO calza."""
    owner = context.bot_data.get("owner_chat_id")
    if not owner:
        logger.info("Chequeo bodega sin owner_chat_id (correr /soydueno).")
        return
    from modules.bodega_check import comparar, formato_alerta
    try:
        res = await asyncio.to_thread(comparar)
    except Exception as e:
        logger.warning(f"Job bodega: {e}")
        try:
            await context.bot.send_message(
                owner, f"⚠️ No pude revisar el Excel de bodega: {str(e)[:150]}")
        except Exception:
            pass
        return

    hay_problemas = bool(res["diferencias"] or res["solo_bodega"] or res["solo_master"])
    if not hay_problemas:
        logger.info("Chequeo bodega semanal: todo calza, sin aviso.")
        return
    txt = "🔔 *Revisión semanal de inventario*\n\n" + formato_alerta(res)
    txt += "\n\nRevisa con /bodega o actualiza el Master."
    await _send_md(context.bot, owner, txt[:4000])


async def job_heartbeat(context):
    """Job diario: manda 'sigo vivo' + resumen al dueño."""
    owner = context.bot_data.get("owner_chat_id")
    if not owner:
        logger.info("Heartbeat sin owner_chat_id (correr /soydueno).")
        return
    try:
        txt = await _build_estado(context)
        await _send_md(context.bot, owner, "❤️ " + txt)
    except Exception as e:
        logger.warning(f"job_heartbeat: {e}")
