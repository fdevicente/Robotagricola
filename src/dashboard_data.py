"""
dashboard_data.py — Extrae datos del Excel y archivos de exportación para el dashboard.
"""
import os
import logging
from datetime import date, datetime, timedelta
from collections import defaultdict
from openpyxl import load_workbook
from config import EXCEL_PATH

logger = logging.getLogger(__name__)

# Claude/ -> Agricola Santa Elisa/
AGRICOLA_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
EXPORT_2024 = os.path.join(AGRICOLA_ROOT, "Exportacion Espana 2024")
EXPORT_2025 = os.path.join(AGRICOLA_ROOT, "Exportacion Espana 2025")
CAMARICO = os.path.join(AGRICOLA_ROOT, "CAMARICO 2023")


def _open_wb(path=None):
    return load_workbook(path or EXCEL_PATH, read_only=True, data_only=True)


def _parse_date(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val)[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def _safe_float(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        s = str(val).replace("$", "").replace(" ", "")
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except Exception:
            return 0.0


# ─── FACTURAS ───────────────────────────────────────────────

def get_facturas_summary():
    """Resumen general de facturas."""
    wb = _open_wb()
    ws = wb["Facturas"]
    total_facturas = 0
    total_monto = 0
    pagadas = 0
    vencidas = 0
    por_pagar = 0
    hoy = date.today()
    por_mes = defaultdict(float)
    por_proveedor = defaultdict(float)
    por_tipo_doc = defaultdict(int)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        total_facturas += 1
        monto = _safe_float(row[14])  # col O = Monto/TOTAL
        total_monto += monto
        fecha_emision = _parse_date(row[0])
        fecha_venc = _parse_date(row[1])
        fecha_pago = str(row[2] or "") if row[2] else ""
        proveedor = str(row[3] or "Sin proveedor")
        tipo_doc = str(row[5] or "Factura")

        if fecha_emision:
            key = f"{fecha_emision.year}-{fecha_emision.month:02d}"
            por_mes[key] += monto

        por_proveedor[proveedor] += abs(monto)
        por_tipo_doc[tipo_doc] += 1

        if fecha_pago and fecha_pago.strip():
            pagadas += 1
        elif fecha_venc and fecha_venc < hoy:
            vencidas += 1
        else:
            por_pagar += 1

    wb.close()

    top_proveedores = sorted(por_proveedor.items(), key=lambda x: x[1], reverse=True)[:10]

    return {
        "total_facturas": total_facturas,
        "total_monto": total_monto,
        "pagadas": pagadas,
        "vencidas": vencidas,
        "por_pagar": por_pagar,
        "por_mes": dict(sorted(por_mes.items())),
        "top_proveedores": [{"nombre": n, "monto": m} for n, m in top_proveedores],
        "por_tipo_doc": dict(por_tipo_doc),
    }


def get_facturas_detalle(filtro: str = "todas"):
    """Retorna filas detalladas de facturas filtradas por estado."""
    wb = _open_wb()
    ws = wb["Facturas"]
    hoy = date.today()
    filas = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        fecha_emision = _parse_date(row[0])
        fecha_venc = _parse_date(row[1])
        fecha_pago = str(row[2] or "").strip() if row[2] else ""
        proveedor = str(row[3] or "")
        rut = str(row[4] or "")
        tipo_doc = str(row[5] or "")
        nro = str(row[6] or "")
        glosa = str(row[7] or "")
        monto = _safe_float(row[14])

        # Determinar estado
        if fecha_pago:
            estado = "pagada"
        elif fecha_venc and fecha_venc < hoy:
            estado = "vencida"
        else:
            estado = "por_pagar"

        if filtro != "todas" and estado != filtro:
            continue

        dias_venc = (hoy - fecha_venc).days if fecha_venc else 0

        filas.append({
            "fecha_emision": str(fecha_emision or ""),
            "fecha_venc": str(fecha_venc or ""),
            "fecha_pago": fecha_pago,
            "proveedor": proveedor,
            "rut": rut,
            "documento": tipo_doc,
            "nro": nro,
            "glosa": glosa,
            "monto": monto,
            "estado": estado,
            "dias_vencida": dias_venc if estado == "vencida" else 0,
        })

    wb.close()
    filas.sort(key=lambda x: x.get("dias_vencida", 0), reverse=True)
    return filas


def get_movimientos_banco():
    """Retorna todas las filas de Cuenta Banco."""
    wb = _open_wb()
    if "Cuenta Banco" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Cuenta Banco"]
    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        filas.append({
            "fecha": str(row[0] or ""),
            "descripcion": str(row[1] or ""),
            "referencia": str(row[2] or ""),
            "cargo": _safe_float(row[3]),
            "abono": _safe_float(row[4]),
            "saldo": _safe_float(row[5]),
        })
    wb.close()
    return filas


def get_banco_revisar():
    """Retorna cargos con Categoria=REVISAR para revision manual.

    Columnas Cuenta Banco: A=fecha, B=descripcion, C=referencia, D=cargo,
    E=abono, F=saldo, G=tipo, H=categoria, I=cultivo
    """
    wb = _open_wb()
    if "Cuenta Banco" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Cuenta Banco"]
    filas = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        # Categoria col H = index 7
        categoria = row[7] if len(row) > 7 else None
        if categoria != "REVISAR":
            continue
        filas.append({
            "fila": idx,
            "fecha": str(row[0] or "")[:10],
            "descripcion": str(row[1] or ""),
            "referencia": str(row[2] or ""),
            "cargo": _safe_float(row[3]),
            "tipo": str(row[6] or "") if len(row) > 6 else "",
            "categoria": str(categoria or ""),
            "cultivo": str(row[8] or "") if len(row) > 8 else "",
        })
    wb.close()
    filas.sort(key=lambda x: x["cargo"], reverse=True)
    return filas


BANCO_CATEGORIAS_VALIDAS = [
    "Fertilizantes", "Fitosanitarios", "Maquinaria - mantencion",
    "Mano de obra temporal", "Mano de obra planta", "Combustible",
    "Riego", "Inversion / Replante", "Servicios profesionales",
    "Arriendos / Patentes / Seguros", "Caja chica / Imprevistos",
    "SERVICIOS DE EXPORTACION", "INSUMOS AGRICOLAS", "TRANSPORTE",
    "ENERGIA", "COSTO ENERGETICO", "HERRAMIENTAS", "SUMINISTROS",
    "MATERIALES", "SEGURIDAD", "MANTENIMIENTO", "CAMBIO DIVISA",
    "AGUA", "ALIMENTACION", "ARRIENDOS", "SERVICIOS",
]


def update_banco_categoria(fila: int, categoria: str, cultivo: str = "GENERAL"):
    """Actualiza Categoria/Cultivo de una fila de Cuenta Banco."""
    from openpyxl import load_workbook as _lw
    wb = _lw(EXCEL_PATH)
    ws = wb["Cuenta Banco"]
    ws.cell(fila, 8).value = categoria
    ws.cell(fila, 9).value = cultivo
    wb.save(EXCEL_PATH)
    wb.close()
    return {"ok": True, "fila": fila, "categoria": categoria}


# ─── COSTOS POR CULTIVO ────────────────────────────────────

def get_costos_por_cultivo():
    """Analiza costos por cultivo basado en facturas y aplicaciones."""
    wb = _open_wb()
    cultivos = {"Nogales": 0, "Cerezos": 0, "Avellanos": 0, "General": 0}

    # Desde Aplicaciones (uso de insumos)
    if "Aplicaciones" in wb.sheetnames:
        ws = wb["Aplicaciones"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            cultivo = str(row[4] or "General")
            cantidad = _safe_float(row[2])
            # Estimar costo basado en inventario
            if cultivo in cultivos:
                cultivos[cultivo] += cantidad * 100  # estimación base
            else:
                cultivos["General"] += cantidad * 100

    # Desde Facturas - clasificar por glosa
    ws_f = wb["Facturas"]
    keywords_cultivo = {
        "nogal": "Nogales", "nuez": "Nogales", "nueces": "Nogales",
        "cerezo": "Cerezos", "cereza": "Cerezos",
        "avellano": "Avellanos", "avellana": "Avellanos",
    }
    facturas_por_cultivo = {"Nogales": 0, "Cerezos": 0, "Avellanos": 0, "General": 0}

    for row in ws_f.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        glosa = str(row[7] or "").lower() + " " + str(row[8] or "").lower()
        monto = _safe_float(row[14])
        asignado = False
        for kw, cult in keywords_cultivo.items():
            if kw in glosa:
                facturas_por_cultivo[cult] += abs(monto)
                asignado = True
                break
        if not asignado:
            facturas_por_cultivo["General"] += abs(monto)

    wb.close()
    return {
        "aplicaciones": cultivos,
        "facturas": facturas_por_cultivo,
    }


# ─── COMPARACIÓN ANUAL ─────────────────────────────────────

def get_comparacion_anual():
    """Gastos por año y mes para comparación."""
    wb = _open_wb()
    ws = wb["Facturas"]
    por_anio = defaultdict(float)
    por_anio_mes = defaultdict(lambda: defaultdict(float))

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        fecha = _parse_date(row[0])
        monto = _safe_float(row[14])
        if fecha:
            por_anio[fecha.year] += abs(monto)
            por_anio_mes[fecha.year][fecha.month] += abs(monto)

    wb.close()

    # También intentar leer planilla de gastos histórica
    try:
        gastos_path = os.path.join(CAMARICO, "PLANILLA GASTOS CAMARICO 2023-2024.xlsx")
        if os.path.exists(gastos_path):
            wb2 = load_workbook(gastos_path, read_only=True, data_only=True)
            if "DATOS" in wb2.sheetnames:
                ws2 = wb2["DATOS"]
                for row in ws2.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    # Intentar extraer fecha y monto de la planilla
                    fecha = _parse_date(row[0] if len(row) > 0 else None)
                    monto = _safe_float(row[-1] if row else None)
                    if fecha and monto:
                        if fecha.year not in por_anio or por_anio[fecha.year] == 0:
                            por_anio[fecha.year] += abs(monto)
                            por_anio_mes[fecha.year][fecha.month] += abs(monto)
            wb2.close()
    except Exception as e:
        logger.warning(f"No se pudo leer planilla gastos: {e}")

    result = {}
    for year in sorted(por_anio.keys()):
        meses = [0.0] * 12
        for m, v in por_anio_mes[year].items():
            meses[m - 1] = v
        result[year] = {"total": por_anio[year], "meses": meses}

    return result


# ─── EXPORTACIONES A ESPAÑA ─────────────────────────────────

def get_exportaciones():
    """Extrae datos de exportación desde las carpetas de Exportación España."""
    exports = []

    for year, folder in [(2024, EXPORT_2024), (2025, EXPORT_2025)]:
        if not os.path.isdir(folder):
            continue

        # Proformas
        for fname in os.listdir(folder):
            if "proforma" in fname.lower() and fname.endswith(".xlsx"):
                try:
                    wb = load_workbook(os.path.join(folder, fname), read_only=True, data_only=True)
                    ws = wb[wb.sheetnames[0]]
                    export_data = {"year": year, "tipo": "Proforma", "archivo": fname}
                    found_header = False
                    for row in ws.iter_rows(min_row=1, max_row=40, values_only=True):
                        vals = [str(c or "").lower() for c in row]
                        joined = " ".join(vals)
                        # Buscar fila header "Producto, Item, ..., Valor US$"
                        if "valor" in joined and "us" in joined:
                            found_header = True
                            continue
                        # Filas de datos después del header
                        if found_header:
                            # Última columna = Valor US$
                            last_val = _safe_float(row[-1]) if row else 0
                            if last_val > 1000:
                                export_data["valor_usd"] = export_data.get("valor_usd", 0) + last_val
                            # Columna cantidad (index 5 normalmente)
                            for c in row:
                                v = _safe_float(c)
                                if 1000 < v < 500000 and "cantidad_kg" not in export_data:
                                    export_data["cantidad_kg"] = v
                            # "Total" marca fin
                            if "total" in joined:
                                if last_val > 1000:
                                    export_data["valor_usd"] = last_val  # usar Total como definitivo
                                found_header = False
                    wb.close()
                    exports.append(export_data)
                except Exception as e:
                    logger.warning(f"Error leyendo {fname}: {e}")

        # Packing Lists
        for fname in os.listdir(folder):
            if "packing" in fname.lower() and fname.endswith(".xlsx"):
                try:
                    wb = load_workbook(os.path.join(folder, fname), read_only=True, data_only=True)
                    ws = wb[wb.sheetnames[0]]
                    pack_data = {"year": year, "tipo": "Packing", "archivo": fname}
                    for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
                        vals = [str(c or "") for c in row]
                        for c in row:
                            v = _safe_float(c)
                            if 10000 < v < 200000:
                                if "peso" not in " ".join(vals).lower():
                                    pack_data.setdefault("peso_total_kg", v)
                    wb.close()
                    exports.append(pack_data)
                except Exception as e:
                    logger.warning(f"Error leyendo {fname}: {e}")

        # Tarjas (contenedores)
        tarjas = []
        for fname in os.listdir(folder):
            if "tarja" in fname.lower() and fname.endswith(".xlsx"):
                try:
                    wb = load_workbook(os.path.join(folder, fname), read_only=True, data_only=True)
                    n_contenedores = len(wb.sheetnames)
                    peso_total = 0
                    for sname in wb.sheetnames:
                        ws = wb[sname]
                        for row in ws.iter_rows(min_row=7, max_row=7, values_only=True):
                            for c in row:
                                v = _safe_float(c)
                                if 200 < v < 2000:
                                    peso_total += v
                    wb.close()
                    tarjas.append({
                        "year": year, "tipo": "Tarja", "archivo": fname,
                        "contenedores": n_contenedores, "peso_neto_kg": peso_total
                    })
                except Exception as e:
                    logger.warning(f"Error leyendo tarja {fname}: {e}")
        exports.extend(tarjas)

    # Resumen por año
    resumen = {}
    for year in [2024, 2025]:
        items = [e for e in exports if e.get("year") == year]
        total_kg = sum(e.get("cantidad_kg", 0) or e.get("peso_neto_kg", 0) or e.get("peso_total_kg", 0) for e in items)
        total_usd = sum(e.get("valor_usd", 0) for e in items)
        contenedores = sum(e.get("contenedores", 0) for e in items if e.get("tipo") == "Tarja")
        resumen[year] = {
            "total_kg": total_kg,
            "total_usd": total_usd,
            "contenedores": contenedores,
            "archivos": len(items),
        }

    return {"detalle": exports, "resumen": resumen}


# ─── CAJA CHICA ─────────────────────────────────────────────

def get_caja_chica():
    """Resumen de caja chica."""
    wb = _open_wb()
    if "Caja Chica" not in wb.sheetnames:
        wb.close()
        return {"saldo": 0, "ingresos": 0, "egresos": 0, "movimientos": []}

    ws = wb["Caja Chica"]
    ingresos = 0
    egresos = 0
    movimientos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        tipo = str(row[1] or "")
        monto = _safe_float(row[2])
        if tipo.lower() in ("ingreso", "depósito", "deposito"):
            ingresos += monto
        else:
            egresos += monto
        movimientos.append({
            "fecha": str(row[0] or ""),
            "tipo": tipo,
            "monto": monto,
            "detalle": str(row[3] or ""),
        })
    wb.close()
    return {
        "saldo": ingresos - egresos,
        "ingresos": ingresos,
        "egresos": egresos,
        "ultimos": movimientos[-10:],
    }


# ─── CUENTA BANCO ───────────────────────────────────────────

def get_cuenta_banco():
    """Resumen de cuenta banco."""
    wb = _open_wb()
    if "Cuenta Banco" not in wb.sheetnames:
        wb.close()
        return {"n_movimientos": 0, "ultimo_saldo": 0, "cargos": 0, "abonos": 0}

    ws = wb["Cuenta Banco"]
    cargos = 0
    abonos = 0
    ultimo_saldo = 0
    n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        n += 1
        cargos += _safe_float(row[3])
        abonos += _safe_float(row[4])
        s = _safe_float(row[5])
        if s:
            ultimo_saldo = s
    wb.close()
    return {
        "n_movimientos": n,
        "ultimo_saldo": ultimo_saldo,
        "cargos": cargos,
        "abonos": abonos,
    }


# ─── INVENTARIO ─────────────────────────────────────────────

def get_inventario_resumen():
    """Resumen de inventario."""
    wb = _open_wb()
    if "Inventario" not in wb.sheetnames:
        wb.close()
        return {"productos": [], "alertas": 0}

    ws = wb["Inventario"]
    productos = []
    alertas = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        stock = _safe_float(row[3])
        minimo = _safe_float(row[4])
        alerta = stock <= minimo and minimo > 0
        if alerta:
            alertas += 1
        productos.append({
            "producto": str(row[0]),
            "categoria": str(row[1] or ""),
            "unidad": str(row[2] or ""),
            "stock": stock,
            "minimo": minimo,
            "alerta": alerta,
        })
    wb.close()
    return {"productos": productos, "alertas": alertas}


# ─── PERSONAL Y VACACIONES ──────────────────────────────────

def get_personal_resumen():
    """Resumen de personal y vacaciones."""
    wb = _open_wb()
    personal = []
    if "Personal" in wb.sheetnames:
        ws = wb["Personal"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            personal.append({
                "nombre": str(row[0]),
                "cargo": str(row[2] or ""),
                "dias_pendientes": _safe_float(row[4]),
                "dias_tomados": _safe_float(row[5]),
                "ultima_vacacion": str(row[6] or ""),
            })

    vacaciones = []
    if "Vacaciones" in wb.sheetnames:
        ws = wb["Vacaciones"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            vacaciones.append({
                "trabajador": str(row[0]),
                "inicio": str(row[1] or ""),
                "fin": str(row[2] or ""),
                "dias": int(_safe_float(row[3])),
                "estado": str(row[4] or ""),
            })
    wb.close()
    return {"personal": personal, "vacaciones": vacaciones[-10:]}


# ─── TAREAS ─────────────────────────────────────────────────

def get_tareas_resumen():
    """Resumen de tareas."""
    wb = _open_wb()
    if "Tareas" not in wb.sheetnames:
        wb.close()
        return {"pendientes": 0, "en_progreso": 0, "completadas": 0, "tareas": []}

    ws = wb["Tareas"]
    pendientes = 0
    en_progreso = 0
    completadas = 0
    tareas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        estado = str(row[4] or "Pendiente")
        if estado == "Pendiente":
            pendientes += 1
        elif estado == "En Progreso":
            en_progreso += 1
        elif estado == "Hecho":
            completadas += 1
        tareas.append({
            "id": int(_safe_float(row[0])),
            "descripcion": str(row[2] or ""),
            "prioridad": str(row[3] or "Media"),
            "estado": estado,
            "responsable": str(row[5] or ""),
        })
    wb.close()
    return {
        "pendientes": pendientes,
        "en_progreso": en_progreso,
        "completadas": completadas,
        "tareas": tareas,
    }


# ─── DATOS COMPLETOS PARA DASHBOARD ────────────────────────

def get_dashboard_data():
    """Retorna todos los datos necesarios para el dashboard."""
    try:
        facturas = get_facturas_summary()
    except Exception as e:
        logger.error(f"Error facturas: {e}")
        facturas = {}

    try:
        costos = get_costos_por_cultivo()
    except Exception as e:
        logger.error(f"Error costos cultivo: {e}")
        costos = {}

    try:
        anual = get_comparacion_anual()
    except Exception as e:
        logger.error(f"Error comparacion anual: {e}")
        anual = {}

    try:
        exports = get_exportaciones()
    except Exception as e:
        logger.error(f"Error exportaciones: {e}")
        exports = {}

    try:
        caja = get_caja_chica()
    except Exception as e:
        logger.error(f"Error caja chica: {e}")
        caja = {}

    try:
        banco = get_cuenta_banco()
    except Exception as e:
        logger.error(f"Error banco: {e}")
        banco = {}

    try:
        inventario = get_inventario_resumen()
    except Exception as e:
        logger.error(f"Error inventario: {e}")
        inventario = {}

    try:
        personal = get_personal_resumen()
    except Exception as e:
        logger.error(f"Error personal: {e}")
        personal = {}

    try:
        tareas = get_tareas_resumen()
    except Exception as e:
        logger.error(f"Error tareas: {e}")
        tareas = {}

    return {
        "fecha": date.today().strftime("%Y-%m-%d"),
        "facturas": facturas,
        "costos_cultivo": costos,
        "comparacion_anual": anual,
        "exportaciones": exports,
        "caja_chica": caja,
        "cuenta_banco": banco,
        "inventario": inventario,
        "personal": personal,
        "tareas": tareas,
    }


# ─── TEMPORADAS (mayo → abril) ─────────────────────────────────────

def temporada_de_fecha(fecha):
    """Devuelve 'TEMP YY/YY+1' para una fecha.

    Corte: 1 de junio. Todo lo que sucede entre jun-año-X y may-año-(X+1) cae
    en TEMP X/(X+1). Esto encaja con el ciclo agrícola: la cosecha se realiza
    feb-may, los pagos finales llegan hasta may, y a partir de junio empiezan
    los gastos de la siguiente temporada (preparación post-cosecha).
    """
    if not fecha:
        return None
    if isinstance(fecha, datetime):
        fecha = fecha.date()
    if not isinstance(fecha, date):
        return None
    if fecha.month >= 6:
        return f"TEMP {fecha.year % 100:02d}/{(fecha.year + 1) % 100:02d}"
    return f"TEMP {(fecha.year - 1) % 100:02d}/{fecha.year % 100:02d}"


def get_temporadas_disponibles():
    """Lista temporadas presentes en facturas + banco, ordenadas desc."""
    wb = _open_wb()
    temps = set()
    if "Facturas" in wb.sheetnames:
        for row in wb["Facturas"].iter_rows(min_row=2, max_col=1, values_only=True):
            t = temporada_de_fecha(_parse_date(row[0]))
            if t: temps.add(t)
    if "Cuenta Banco" in wb.sheetnames:
        for row in wb["Cuenta Banco"].iter_rows(min_row=2, max_col=1, values_only=True):
            t = temporada_de_fecha(_parse_date(row[0]))
            if t: temps.add(t)
    wb.close()
    return sorted(temps, reverse=True)


def get_resumen_temporada(temporada: str):
    """Resumen agregado de facturas y banco para una temporada (mayo→abril)."""
    wb = _open_wb()
    egresos_por_cat = defaultdict(float)
    egresos_por_cultivo = defaultdict(float)
    ingresos_por_cat = defaultdict(float)
    total_facturas = 0
    total_facturas_monto = 0
    total_cargos_banco = 0
    total_abonos_banco = 0
    cargos_count = 0
    abonos_count = 0

    # Facturas (egresos)
    if "Facturas" in wb.sheetnames:
        ws = wb["Facturas"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            t = temporada_de_fecha(_parse_date(row[0]))
            if t != temporada: continue
            monto = _safe_float(row[14])
            cat = str(row[16] or "SIN CATEGORIA").upper()
            cult = str(row[17] or "GENERAL").upper()
            total_facturas += 1
            total_facturas_monto += monto
            egresos_por_cat[cat] += monto
            egresos_por_cultivo[cult] += monto

    # Banco (ingresos y egresos)
    cosecha_dedup = set()  # (fecha_iso, monto_int) ya contados en Cosechas
    USD_CLP = 904

    # Cosechas primero (fuente de verdad para ingresos de ventas)
    if "Cosechas" in wb.sheetnames:
        ws = wb["Cosechas"]
        for row in ws.iter_rows(min_row=2, max_col=16, values_only=True):
            if not row[0]: continue
            estado = row[11]
            cultivo = str(row[1] or "GENERAL").upper()
            if estado == "recibido":
                fecha = _parse_date(row[12])
                monto = row[13] or 0
                moneda = (row[14] or "CLP").upper()
                monto_clp = float(monto) if moneda == "CLP" else float(monto) * USD_CLP
            else:
                fecha = _parse_date(row[8])
                monto_usd = row[9] or 0
                monto_clp = float(monto_usd) * USD_CLP
            if not fecha or monto_clp <= 0: continue
            t = temporada_de_fecha(fecha)
            if t != temporada: continue
            cat = f"INGRESO {cultivo}"
            ingresos_por_cat[cat] += monto_clp
            abonos_count += 1
            total_abonos_banco += monto_clp
            cosecha_dedup.add((fecha.isoformat(), int(round(monto_clp))))

    if "Cuenta Banco" in wb.sheetnames:
        ws = wb["Cuenta Banco"]
        for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
            if not row[0]: continue
            fecha = _parse_date(row[0])
            t = temporada_de_fecha(fecha)
            if t != temporada: continue
            try:
                cargo = float(row[3] or 0)
                abono = float(row[4] or 0)
            except (TypeError, ValueError):
                continue
            cat = str(row[7] or "SIN CATEGORIA").upper()
            if cargo > 0:
                total_cargos_banco += cargo
                cargos_count += 1
                if "INGRESO" not in cat and "TRANSFERENCIA" not in cat and "PRE-2021" not in cat:
                    egresos_por_cat[cat] += cargo
            if abono > 0 and cargo == 0:
                # Dedup con Cosechas
                if (fecha.isoformat(), int(round(abono))) in cosecha_dedup:
                    continue
                total_abonos_banco += abono
                abonos_count += 1
                ingresos_por_cat[cat] += abono

    wb.close()

    top_egresos = sorted(egresos_por_cat.items(), key=lambda x: -x[1])[:15]
    top_ingresos = sorted(ingresos_por_cat.items(), key=lambda x: -x[1])[:10]
    top_cultivos = sorted(egresos_por_cultivo.items(), key=lambda x: -x[1])

    total_egresos = sum(egresos_por_cat.values())
    total_ingresos = sum(ingresos_por_cat.values())

    return {
        "temporada": temporada,
        "total_facturas": total_facturas,
        "total_facturas_monto": total_facturas_monto,
        "cargos_banco_count": cargos_count,
        "cargos_banco_monto": total_cargos_banco,
        "abonos_banco_count": abonos_count,
        "abonos_banco_monto": total_abonos_banco,
        "total_egresos": total_egresos,
        "total_ingresos": total_ingresos,
        "saldo_neto": total_ingresos - total_egresos,
        "top_egresos_cat": [{"categoria": c, "monto": m} for c, m in top_egresos],
        "top_ingresos_cat": [{"categoria": c, "monto": m} for c, m in top_ingresos],
        "egresos_por_cultivo": [{"cultivo": c, "monto": m} for c, m in top_cultivos],
    }


def get_reporte_mensual(year: int, month: int):
    """Genera datos para reporte mensual del mes especificado.

    Incluye:
    - Resumen ejecutivo: ingresos real, egresos real, flujo neto, saldo banco
    - Detalle ingresos del mes
    - Detalle egresos por categoría
    - Comparación proyectado vs real
    - Facturas vencidas globales
    - Vacaciones pendientes alertas
    - Próximos 3 meses proyectados
    """
    from datetime import date as _date
    wb = _open_wb()
    inicio = _date(year, month, 1)
    if month == 12:
        fin = _date(year + 1, 1, 1)
    else:
        fin = _date(year, month + 1, 1)

    # Dedup banco+factura: cada categoría se cuenta de UNA sola fuente
    # (misma lógica que el proyector).
    from modules.cash_flow.projector import CATS_SOLO_BANCO, EXCLUIR_CATS_EGRESO

    ingresos_lista = []
    egresos_lista = []
    ingresos_por_cat = defaultdict(float)
    egresos_por_cat = defaultdict(float)
    egresos_por_cultivo = defaultdict(float)

    if "Cuenta Banco" in wb.sheetnames:
        ws = wb["Cuenta Banco"]
        for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
            if not row[0]: continue
            f = _parse_date(row[0])
            if not f or f < inicio or f >= fin: continue
            try:
                cargo = float(row[3] or 0)
                abono = float(row[4] or 0)
            except: continue
            cat = str(row[7] or "SIN CATEGORIA").upper()
            desc = str(row[1] or "")[:60]
            if abono > 0 and cargo == 0:
                if "TRANSFERENCIA" in cat or "PRE-2021" in cat: continue
                ingresos_lista.append({"fecha": f.isoformat(), "desc": desc, "monto": abono, "cat": cat})
                ingresos_por_cat[cat] += abono
            elif cargo > 0:
                # Egresos del banco: solo categorías que viven en el banco
                if cat in EXCLUIR_CATS_EGRESO: continue
                if cat not in CATS_SOLO_BANCO: continue  # estas vienen de facturas
                egresos_lista.append({"fecha": f.isoformat(), "desc": desc, "monto": cargo, "cat": cat})
                egresos_por_cat[cat] += cargo

    # Facturas pagadas del mes (categorías que NO son solo-banco)
    if "Facturas" in wb.sheetnames:
        ws_f = wb["Facturas"]
        for row in ws_f.iter_rows(min_row=2, max_col=20, values_only=True):
            if not row[0]: continue
            f_pago = _parse_date(row[2]) if row[2] else None
            if not f_pago or f_pago < inicio or f_pago >= fin: continue
            cat = str(row[16] or "SIN CATEGORIA").upper()
            if cat in EXCLUIR_CATS_EGRESO: continue
            if cat in CATS_SOLO_BANCO: continue  # estas se cuentan del banco
            cat_por = str(row[19] or "") if len(row) > 19 else ""
            if "NN-no-pagar" in cat_por: continue
            cult = str(row[17] or "GENERAL").upper()
            monto = _safe_float(row[14])
            prov = str(row[3] or "")
            egresos_lista.append({"fecha": f_pago.isoformat(), "desc": prov[:60], "monto": monto, "cat": cat, "fuente": "factura"})
            egresos_por_cat[cat] += monto
            egresos_por_cultivo[cult] += monto

    # Saldo banco al final del mes
    saldo_fin_mes = 0
    saldo_fecha = None
    if "Cuenta Banco" in wb.sheetnames:
        ws = wb["Cuenta Banco"]
        for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
            if not row[0]: continue
            f = _parse_date(row[0])
            if not f or f >= fin: continue
            try: saldo = float(row[5] or 0)
            except: continue
            if saldo == 0: continue
            if not saldo_fecha or f >= saldo_fecha:
                saldo_fecha = f
                saldo_fin_mes = saldo

    # Facturas vencidas globales (al cierre del mes)
    hoy_ref = fin
    vencidas = []
    if "Facturas" in wb.sheetnames:
        ws_f = wb["Facturas"]
        for r_idx, row in enumerate(ws_f.iter_rows(min_row=2, max_col=20, values_only=True), start=2):
            if not row[0]: continue
            cat_por = str(row[19] or "") if len(row) > 19 else ""
            if "NN-no-pagar" in cat_por: continue
            f_pago = row[2]
            if f_pago and str(f_pago).strip(): continue  # pagada
            f_venc = _parse_date(row[1])
            if not f_venc or f_venc >= hoy_ref: continue
            dias_v = (hoy_ref - f_venc).days
            vencidas.append({
                "proveedor": str(row[3] or "")[:35],
                "nro": str(row[6] or ""),
                "monto": _safe_float(row[14]),
                "dias": dias_v,
                "vence": f_venc.isoformat(),
            })
    vencidas.sort(key=lambda x: -x["dias"])

    total_ingresos = sum(ingresos_por_cat.values())
    total_egresos = sum(egresos_por_cat.values())
    flujo_neto = total_ingresos - total_egresos

    egresos_lista.sort(key=lambda x: -x["monto"])
    ingresos_lista.sort(key=lambda x: -x["monto"])

    # ─── Egresos por categoría: top 10 + "OTROS" agrupado ───
    cats_ordenadas = sorted(egresos_por_cat.items(), key=lambda x: -x[1])
    top10 = cats_ordenadas[:10]
    resto = cats_ordenadas[10:]
    egresos_cat_final = [{"cat": k, "monto": v} for k, v in top10]
    if resto:
        otros_monto = sum(v for _, v in resto)
        otros_detalle = ", ".join(f"{k} ${v:,.0f}" for k, v in resto)
        egresos_cat_final.append({
            "cat": "OTROS", "monto": otros_monto,
            "nota": f"{len(resto)} categorías menores: {otros_detalle}",
        })

    # ─── Top 5 egresos con nota explicativa ───
    NOTAS_CAT = {
        "MANO DE OBRA TEMPORAL": "Jornaleros / contratistas para labores de temporada (cosecha, poda).",
        "MANO DE OBRA PLANTA": "Sueldos del personal fijo + cotizaciones (Previred).",
        "INSUMOS AGRICOLAS": "Compra de insumos generales del campo (Copeval, Martínez y Valdivieso, CALS).",
        "FERTILIZANTES": "Fertilizantes para nutrición de los cultivos.",
        "FITOSANITARIOS": "Pesticidas, fungicidas y herbicidas para control sanitario.",
        "COMBUSTIBLE": "Diésel, bencina y gas (incluye gas de secado en cosecha).",
        "MAQUINARIA - MANTENCION": "Reparación y mantención de tractores y equipos.",
        "ENERGIA": "Electricidad CGE.",
        "COSTO ENERGETICO": "Energía generada por paneles solares (S-Invest).",
        "RIEGO": "Mantención y operación del sistema de riego.",
        "SERVICIOS PROFESIONALES": "Asesorías (agrónomo, contabilidad).",
        "INVERSION / REPLANTE": "Inversión en plantación nueva (avellanos) y activos.",
        "SERVICIOS DE EXPORTACION": "Gastos de exportación a España (EcoSmart, aduanas, fletes).",
        "GASTOS VEHICULOS": "TAG, permiso de circulación, SOAP, mantención vehículos.",
        "IMPUESTOS": "Pago de F29 (IVA) y contribuciones.",
        "BONO VENTA NUECES": "Bono del 8% sobre la venta de nueces.",
        "MATERIALES": "Ferretería, herramientas y materiales varios.",
    }
    top5_egresos = []
    for e in egresos_lista[:5]:
        cat = e["cat"]
        top5_egresos.append({
            **e,
            "nota": NOTAS_CAT.get(cat, "Gasto registrado en el período."),
        })

    # ─── Egresos por cultivo (solo si hay diferenciación real) ───
    cultivos_no_general = {k: v for k, v in egresos_por_cultivo.items()
                           if k and k.upper() != "GENERAL" and v > 0}
    mostrar_cultivos = len(cultivos_no_general) >= 2  # al menos 2 cultivos para diferenciar
    egresos_cultivo_final = []
    if mostrar_cultivos:
        egresos_cultivo_final = sorted(
            [{"cultivo": k, "monto": v} for k, v in egresos_por_cultivo.items() if v > 0],
            key=lambda x: -x["monto"])

    # ─── Tareas y bitácora del mes ───
    tareas_mes = []
    if "Tareas" in wb.sheetnames:
        ws_t = wb["Tareas"]
        for row in ws_t.iter_rows(min_row=2, values_only=True):
            if not row or not row[2]: continue
            # Col 8 = Fecha Completada, col 1 = Fecha Creación
            f_comp = _parse_date(row[7]) if len(row) > 7 and row[7] else None
            f_crea = _parse_date(row[1]) if len(row) > 1 and row[1] else None
            estado = str(row[4] or "") if len(row) > 4 else ""
            # Incluir si se completó en el mes O si se creó en el mes
            relevante = False
            if f_comp and inicio <= f_comp < fin:
                relevante = True
            elif f_crea and inicio <= f_crea < fin:
                relevante = True
            if not relevante: continue
            tareas_mes.append({
                "descripcion": str(row[2] or ""),
                "estado": estado,
                "responsable": str(row[5] or "") if len(row) > 5 else "",
                "observaciones": str(row[8] or "") if len(row) > 8 else "",
                "fecha": (f_comp or f_crea).isoformat() if (f_comp or f_crea) else "",
            })

    bitacora_mes = []
    jh_por_labor = defaultdict(float)  # "Actividad · Cultivo" -> JH del mes
    if "Bitácora" in wb.sheetnames:
        ws_b = wb["Bitácora"]
        # Detectar esquema: nuevo (col3=Tipo) vs viejo (col3=Registro)
        header3 = str(ws_b.cell(1, 3).value or "").strip().lower()
        esquema_nuevo = header3 == "tipo"
        for row in ws_b.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            f = _parse_date(row[0])
            if not f or f < inicio or f >= fin: continue
            if esquema_nuevo:
                tipo = str(row[2] or "")
                actividad = str(row[3] or "")
                cultivo = str(row[4] or "")
                sector = str(row[5] or "")
                jh = row[6]
                trabajadores = str(row[7] or "")
                insumo = str(row[8] or "")
                cantidad = row[9]
                unidad = str(row[10] or "")
                registro = str(row[11] or "") if len(row) > 11 else ""
                # texto resumido para el reporte
                partes = []
                if actividad: partes.append(actividad)
                if cultivo and cultivo != "GENERAL": partes.append(cultivo)
                if sector: partes.append(f"({sector})")
                if jh: partes.append(f"· {jh} JH")
                if insumo:
                    c = f" {cantidad:g}{unidad}" if cantidad else ""
                    partes.append(f"· {insumo}{c}")
                texto = " ".join(partes) or registro
                bitacora_mes.append({
                    "fecha": f.isoformat(),
                    "registro": texto,
                    "categoria": tipo,
                    "trabajadores": trabajadores,
                })
                try:
                    jhv = float(jh or 0)
                    if jhv > 0:
                        key = f"{actividad} · {cultivo}" if cultivo != "GENERAL" else actividad
                        jh_por_labor[key] += jhv
                except (TypeError, ValueError):
                    pass
            else:
                bitacora_mes.append({
                    "fecha": f.isoformat(),
                    "registro": str(row[2] or "") if len(row) > 2 else "",
                    "categoria": str(row[3] or "") if len(row) > 3 else "",
                })

    # ─── Flujo acumulado vs temporada anterior ───
    flujo_comparativo = _flujo_acumulado_temporadas(wb, year, month)

    # ─── Vacaciones pendientes ───
    vacaciones = []
    try:
        wb.close()
        vacaciones = get_vacaciones_pendientes()
        wb = _open_wb()  # reabrir por si algo más lo necesita (ya no, pero seguro)
        wb.close()
    except Exception:
        pass

    # ─── Alertas de vencimiento de insumos ───
    alertas_venc = {"vencidos": [], "alerta_10": [], "alerta_50": []}
    try:
        from vencimientos_manager import listar_alertas
        alertas_venc = listar_alertas()
    except Exception as e:
        logger.warning(f"alertas vencimiento: {e}")

    # Cuánto del movimiento bancario del mes quedó respaldado con documentos
    conciliacion = None
    try:
        from modules.conciliacion_export import pct_conciliado_mes
        conciliacion = pct_conciliado_mes(year, month)
    except Exception as e:
        logger.warning(f"% conciliado del mes: {e}")

    return {
        "year": year,
        "month": month,
        "mes_label": _MESES[month] if 1 <= month <= 12 else str(month),
        "periodo": f"{inicio.isoformat()} a {(fin - _date.resolution).isoformat()}",
        "conciliacion": conciliacion,
        "total_ingresos": total_ingresos,
        "total_egresos": total_egresos,
        "flujo_neto": flujo_neto,
        "saldo_banco_fin_mes": saldo_fin_mes,
        "saldo_banco_fecha": saldo_fecha.isoformat() if saldo_fecha else None,
        "ingresos_lista": ingresos_lista[:10],
        "ingresos_count": len(ingresos_lista),
        "egresos_count": len(egresos_lista),
        "egresos_por_cat": egresos_cat_final,
        "top5_egresos": top5_egresos,
        "egresos_por_cultivo": egresos_cultivo_final,
        "mostrar_cultivos": mostrar_cultivos,
        "tareas_mes": tareas_mes,
        "bitacora_mes": bitacora_mes,
        "jh_por_labor": sorted([{"labor": k, "jh": v} for k, v in jh_por_labor.items()],
                                key=lambda x: -x["jh"]),
        "alertas_vencimiento": alertas_venc,
        "flujo_comparativo": flujo_comparativo,
        "vacaciones": vacaciones,
        "vencidas_top": vencidas[:10],
        "vencidas_count": len(vencidas),
        "vencidas_monto": sum(v["monto"] for v in vencidas),
    }


def _flujo_acumulado_temporadas(wb, year, month):
    """Flujo neto acumulado mes a mes: temporada actual vs anterior.

    Devuelve {labels, temp_actual, temp_anterior, nombre_actual, nombre_anterior}
    con el saldo neto (ingresos-egresos) acumulado por mes dentro de cada temporada.
    """
    from datetime import date as _date
    # Temporada del mes de reporte
    temp_actual = temporada_de_fecha(_date(year, month, 1))
    if not temp_actual:
        return None
    # Temporada anterior
    yy = int(temp_actual.split()[1].split("/")[0])
    temp_anterior = f"TEMP {yy-1:02d}/{yy:02d}"

    # Meses de una temporada en orden (jun→may)
    orden_meses = [6, 7, 8, 9, 10, 11, 12, 1, 2, 3, 4, 5]
    labels = ["Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic",
              "Ene", "Feb", "Mar", "Abr", "May"]

    # Acumular ingresos y egresos por (temporada, mes_idx)
    flujo = {temp_actual: defaultdict(float), temp_anterior: defaultdict(float)}

    def _proc_banco():
        if "Cuenta Banco" not in wb.sheetnames: return
        for row in wb["Cuenta Banco"].iter_rows(min_row=2, max_col=9, values_only=True):
            if not row[0]: continue
            f = _parse_date(row[0])
            if not f: continue
            t = temporada_de_fecha(f)
            if t not in flujo: continue
            try:
                cargo = float(row[3] or 0); abono = float(row[4] or 0)
            except: continue
            cat = str(row[7] or "").upper()
            if "TRANSFERENCIA" in cat or "PRE-2021" in cat: continue
            mi = orden_meses.index(f.month)
            if abono > 0 and cargo == 0 and "INGRESO" not in cat:
                # abono real (ingresos catalogados como INGRESO también suman)
                flujo[t][mi] += abono
            if abono > 0 and "INGRESO" in cat:
                flujo[t][mi] += abono
            if cargo > 0 and "INGRESO" not in cat:
                flujo[t][mi] -= cargo

    _proc_banco()

    # Convertir a acumulado
    def acum(temp):
        out = []
        running = 0
        for mi in range(12):
            running += flujo[temp].get(mi, 0)
            out.append(round(running))
        return out

    return {
        "labels": labels,
        "nombre_actual": temp_actual,
        "nombre_anterior": temp_anterior,
        "temp_actual": acum(temp_actual),
        "temp_anterior": acum(temp_anterior),
        "mes_actual_idx": orden_meses.index(month),
    }


_MESES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
           "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


def get_vacaciones_pendientes():
    """Calcula vacaciones pendientes en tiempo real para personal fijo.

    Base: saldo último conocido + acumulación 1.25 días/mes desde esa fecha.
    Descuenta vacaciones tomadas registradas en hoja 'Vacaciones'.
    """
    wb = _open_wb()
    if "Vacaciones Pendientes" not in wb.sheetnames:
        wb.close()
        return []

    DIAS_X_MES = 15.0 / 12.0
    hoy = date.today()

    # Cargar vacaciones tomadas (de hoja "Vacaciones")
    tomadas = defaultdict(float)  # nombre_norm -> total días
    if "Vacaciones" in wb.sheetnames:
        ws_v = wb["Vacaciones"]
        for row in ws_v.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            nombre = str(row[0] or "").strip().upper()
            estado = str(row[4] or "").lower() if len(row) > 4 else ""
            if "aprobad" not in estado: continue
            try: dias = float(row[3] or 0)
            except: continue
            tomadas[nombre] += dias

    # Personal con saldos
    ws = wb["Vacaciones Pendientes"]
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        nombre = str(row[0])
        rut = str(row[1] or "")
        fc = _parse_date(row[2])
        try: saldo_base = float(row[3] or 0)
        except: saldo_base = 0
        fs = _parse_date(row[4])
        notas = str(row[5] or "")
        if not fc or not fs: continue

        meses = (hoy.year - fs.year) * 12 + (hoy.month - fs.month)
        dias_acumulados = meses * DIAS_X_MES
        dias_tomadas = tomadas.get(nombre.upper(), 0)
        total = saldo_base + dias_acumulados - dias_tomadas

        anos_trabajados = (hoy - fc).days / 365.25
        items.append({
            "nombre": nombre,
            "rut": rut,
            "fecha_contrato": fc.isoformat(),
            "anos_trabajados": round(anos_trabajados, 1),
            "saldo_base": round(saldo_base, 2),
            "fecha_saldo": fs.isoformat(),
            "meses_acumulados": meses,
            "dias_acumulados": round(dias_acumulados, 2),
            "dias_tomadas": round(dias_tomadas, 2),
            "total_pendiente": round(total, 2),
            "notas": notas,
        })

    wb.close()
    items.sort(key=lambda x: -x["total_pendiente"])
    return items


def compare_temporadas(temp1: str, temp2: str):
    """Devuelve dos resúmenes lado a lado para comparar."""
    return {
        "temp1": get_resumen_temporada(temp1),
        "temp2": get_resumen_temporada(temp2),
    }


def get_facturas_temporada(temporada: str):
    """Resumen de facturas + vencidas/próximas globales (no filtradas por temporada).

    - Conteos por temporada para los KPIs.
    - Vencidas/próximas globales (todo lo pendiente de pago en el sistema).
    """
    wb = _open_wb()
    if "Facturas" not in wb.sheetnames:
        wb.close()
        return {"total": 0, "monto": 0, "vencidas": [], "proximas": []}

    ws = wb["Facturas"]
    hoy = date.today()
    vencidas_global = []
    proximas_global = []
    total = 0
    monto_total = 0
    pagadas = 0

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]: continue
        fecha_emi = _parse_date(row[0])

        # Conteo de la temporada seleccionada
        en_temporada = temporada_de_fecha(fecha_emi) == temporada
        if en_temporada:
            total += 1
            monto = _safe_float(row[14])
            monto_total += monto

        fecha_venc = _parse_date(row[1])
        fecha_pago = row[2]
        if fecha_pago and str(fecha_pago).strip():
            if en_temporada: pagadas += 1
            continue

        # Excluir facturas marcadas NN (no se van a pagar)
        cat_por = str(row[19] or "") if len(row) > 19 else ""
        if "NN-no-pagar" in cat_por:
            continue

        # Vencidas/próximas: SIN filtro de temporada (globales)
        f = {
            "fila": r_idx,
            "fecha_emision": fecha_emi.isoformat() if fecha_emi else "",
            "fecha_vencimiento": fecha_venc.isoformat() if fecha_venc else "",
            "proveedor": str(row[3] or ""),
            "nro": str(row[6] or ""),
            "detalle": str(row[7] or "")[:80],
            "monto": _safe_float(row[14]),
            "categoria": str(row[16] or ""),
            "temporada": temporada_de_fecha(fecha_emi) or "",
        }
        if fecha_venc and fecha_venc < hoy:
            f["dias_vencido"] = (hoy - fecha_venc).days
            vencidas_global.append(f)
        elif fecha_venc:
            f["dias_para_vencer"] = (fecha_venc - hoy).days
            proximas_global.append(f)

    wb.close()
    vencidas_global.sort(key=lambda x: -x.get("dias_vencido", 0))
    proximas_global.sort(key=lambda x: x.get("dias_para_vencer", 9999))
    return {
        "total": total,
        "monto": monto_total,
        "pagadas": pagadas,
        "por_pagar": total - pagadas,
        "vencidas": vencidas_global[:50],
        "proximas": proximas_global[:10],
        "total_vencidas": len(vencidas_global),
        "total_vencidas_monto": sum(v["monto"] for v in vencidas_global),
        "total_proximas": len(proximas_global),
    }


def get_saldo_banco_actual():
    """Lee el último saldo del banco (col F = saldo, fila más reciente)."""
    wb = _open_wb()
    if "Cuenta Banco" not in wb.sheetnames:
        wb.close()
        return {"saldo": 0, "fecha": None}
    ws = wb["Cuenta Banco"]
    last_saldo = 0
    last_fecha = None
    for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
        if not row[0]: continue
        fecha = _parse_date(row[0])
        if not fecha: continue
        try:
            saldo = float(row[5] or 0)
        except (TypeError, ValueError):
            continue
        if saldo == 0: continue
        if not last_fecha or fecha >= last_fecha:
            last_fecha = fecha
            last_saldo = saldo
    wb.close()
    return {
        "saldo": last_saldo,
        "fecha": last_fecha.isoformat() if last_fecha else None,
    }


def get_flujo_mensual_proyectado(meses: int = 12):
    """Tabla flujo proyectado vs real: categorías × meses.
    Real = lo ya ejecutado en facturas + cargos banco para meses pasados/actual.
    """
    from modules.cash_flow.projector import (
        load_historical_egresos, load_ajustes_manuales, load_hectareas,
        load_expected_ingresos, compute_factor_hc, EXCLUIR_PROYECCION,
    )
    today = date.today()
    sy, sm = today.year, today.month

    months = []
    y, m = sy, sm
    for _ in range(meses):
        months.append((y, m))
        m += 1
        if m > 12: m = 1; y += 1

    historicos = load_historical_egresos()
    ajustes = load_ajustes_manuales()
    hc = load_hectareas()
    base_year = sy - 1

    # ─── Proyectado: cat × (y,m) -> monto ───
    egresos = defaultdict(lambda: defaultdict(float))
    for (y_h, m_h, cat, cul), monto in historicos.items():
        if y_h != base_year: continue
        if (cat or "").upper() in EXCLUIR_PROYECCION: continue
        factor = compute_factor_hc(hc, cul, base_year, sy)
        target_y = sy if m_h >= sm else sy + 1
        if (target_y, m_h) in months:
            egresos[cat][(target_y, m_h)] += monto * factor

    for a in ajustes:
        ym = a["mes_proyectado"]
        if ym in months:
            if (a["categoria"] or "").upper() in EXCLUIR_PROYECCION: continue
            egresos[a["categoria"]][ym] += a["monto"]

    # ─── Real: lo realmente ejecutado este año (mes actual + meses pasados del flujo) ───
    # Pero solo mes actual cuenta como "real" (los futuros aún no pasaron, los pasados ya están en histórico)
    real = defaultdict(lambda: defaultdict(float))
    wb = _open_wb()

    # Facturas reales
    if "Facturas" in wb.sheetnames:
        ws = wb["Facturas"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            f = _parse_date(row[0])
            if not f: continue
            ym = (f.year, f.month)
            if ym not in months: continue
            cat = str(row[16] or "SIN CAT").upper()
            if cat in EXCLUIR_PROYECCION: continue
            monto = _safe_float(row[14])
            real[cat][ym] += monto

    # Cargos banco reales (egresos)
    if "Cuenta Banco" in wb.sheetnames:
        ws = wb["Cuenta Banco"]
        for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
            if not row[0]: continue
            f = _parse_date(row[0])
            if not f: continue
            ym = (f.year, f.month)
            if ym not in months: continue
            try:
                cargo = float(row[3] or 0)
            except: continue
            if cargo <= 0: continue
            cat = str(row[7] or "SIN CAT").upper()
            if cat in EXCLUIR_PROYECCION: continue
            if "INGRESO" in cat or "TRANSFERENCIA" in cat or "PRE-2021" in cat: continue
            real[cat][ym] += cargo
    wb.close()

    # Ingresos proyectados
    ingresos_por_mes = defaultdict(float)
    for i in load_expected_ingresos():
        ym = (i["year"], i["month"])
        if ym in months:
            ingresos_por_mes[ym] += i["monto_clp"]

    # Ingresos reales del banco
    real_ingresos = defaultdict(float)
    wb = _open_wb()
    if "Cuenta Banco" in wb.sheetnames:
        for row in wb["Cuenta Banco"].iter_rows(min_row=2, max_col=9, values_only=True):
            if not row[0]: continue
            f = _parse_date(row[0])
            if not f: continue
            ym = (f.year, f.month)
            if ym not in months: continue
            try:
                cargo = float(row[3] or 0)
                abono = float(row[4] or 0)
            except: continue
            if abono <= 0 or cargo > 0: continue
            cat = str(row[7] or "").upper()
            if "TRANSFERENCIA" in cat or "PRE-2021" in cat: continue
            real_ingresos[ym] += abono
    wb.close()

    labels = [f"{['','Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic'][m]}-{str(y)[-2:]}" for y, m in months]

    # Unión de todas las categorías (proyectado + real)
    all_cats = set(egresos.keys()) | set(real.keys())
    cat_totals = []
    for cat in all_cats:
        tot_proy = sum(egresos[cat].values())
        tot_real = sum(real[cat].values())
        cat_totals.append((cat, tot_proy, tot_real))
    cat_totals.sort(key=lambda x: -(x[1] + x[2]))

    filas = []
    for cat, tot_proy, tot_real in cat_totals:
        if tot_proy == 0 and tot_real == 0: continue
        filas.append({
            "categoria": cat or "(SIN CAT)",
            "total_proy": tot_proy,
            "total_real": tot_real,
            "meses_proy": [egresos[cat].get(ym, 0) for ym in months],
            "meses_real": [real[cat].get(ym, 0) for ym in months],
        })

    return {
        "meses": months,
        "labels": labels,
        "mes_actual_idx": 0,
        "ingresos_por_mes_proy": [ingresos_por_mes.get(ym, 0) for ym in months],
        "ingresos_por_mes_real": [real_ingresos.get(ym, 0) for ym in months],
        "total_ingresos_proy": sum(ingresos_por_mes.values()),
        "total_ingresos_real": sum(real_ingresos.values()),
        "filas_egresos": filas,
        "total_egresos_proy": sum(f["total_proy"] for f in filas),
        "total_egresos_real": sum(f["total_real"] for f in filas),
    }
