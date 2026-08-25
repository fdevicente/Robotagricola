"""
dashboard.py — Servidor web Flask para el dashboard de Agrícola Santa Elisa.
Ejecutar: python src/dashboard.py
"""
import os
import sys
import json
import logging
from datetime import date

# Agregar directorio padre al path para importar config
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from flask import Flask, render_template, jsonify, request
from dashboard_data import get_dashboard_data, get_facturas_summary, get_comparacion_anual, \
    get_costos_por_cultivo, get_exportaciones, get_caja_chica, get_cuenta_banco, \
    get_inventario_resumen, get_personal_resumen, get_tareas_resumen, \
    get_facturas_detalle, get_movimientos_banco, \
    get_banco_revisar, update_banco_categoria, BANCO_CATEGORIAS_VALIDAS, \
    get_temporadas_disponibles, get_resumen_temporada, compare_temporadas, \
    get_facturas_temporada, get_saldo_banco_actual, get_flujo_mensual_proyectado, \
    get_vacaciones_pendientes, get_reporte_mensual
from modules.cash_flow.projector import get_cash_flow
from modules.cash_flow.replante import afford_check

_MESES_LBL = ["", "Ene", "Feb", "Mar", "Abr", "May", "Jun",
                "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


def _month_label(y, m):
    return f"{_MESES_LBL[m]}-{str(y)[-2:]}"

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__,
            template_folder=os.path.join(os.path.dirname(__file__), "templates"),
            static_folder=os.path.join(os.path.dirname(__file__), "static"))

# ── Autenticación y roles ─────────────────────────────────────────────
import secrets
from functools import wraps
from flask import redirect, url_for, request as _rq, session as _sess, flash
from flask_login import (LoginManager, UserMixin, login_user, logout_user,
                          login_required, current_user)
from modules.auth import (verificar, obtener, puede, auditar, ROLES,
                           listar_usuarios, crear_usuario, hay_usuarios)

_SECRET_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                            ".flask_secret")
if os.path.exists(_SECRET_FILE):
    with open(_SECRET_FILE) as _f:
        app.secret_key = _f.read().strip()
else:
    app.secret_key = secrets.token_hex(32)
    try:
        with open(_SECRET_FILE, "w") as _f:
            _f.write(app.secret_key)
    except OSError:
        pass

from datetime import timedelta as _timedelta

app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    PERMANENT_SESSION_LIFETIME=_timedelta(hours=12),
)
# En producción (HTTPS) marcar la cookie como segura
if os.getenv("HTTPS", "").lower() in ("1", "true", "si"):
    app.config["SESSION_COOKIE_SECURE"] = True

# Token interno: permite que el generador de PDF entre sin login
TOKEN_INTERNO = os.getenv("DASHBOARD_TOKEN", "")
if not TOKEN_INTERNO:
    _tf = os.path.join(os.path.dirname(_SECRET_FILE), ".dashboard_token")
    if os.path.exists(_tf):
        with open(_tf) as _f:
            TOKEN_INTERNO = _f.read().strip()
    else:
        TOKEN_INTERNO = secrets.token_urlsafe(24)
        try:
            with open(_tf, "w") as _f:
                _f.write(TOKEN_INTERNO)
        except OSError:
            pass

login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = "Inicia sesión para continuar."


class WebUser(UserMixin):
    def __init__(self, u):
        self.id = str(u.id)
        self.usuario = u.usuario
        self.nombre = u.nombre or u.usuario
        self.rol = u.rol

    @property
    def rol_label(self):
        return ROLES.get(self.rol, self.rol)


@login_manager.user_loader
def _cargar_usuario(uid):
    u = obtener(uid)
    return WebUser(u) if u and u.activo else None


def _es_interno():
    """Petición interna autorizada (generador de PDF) vía token."""
    tok = _rq.args.get("token") or _rq.headers.get("X-Dashboard-Token")
    return bool(tok) and secrets.compare_digest(tok, TOKEN_INTERNO)


def requiere(permiso: str):
    """Protege una ruta exigiendo un permiso del rol."""
    def deco(fn):
        @wraps(fn)
        def wrapper(*a, **kw):
            if _es_interno():
                return fn(*a, **kw)
            if not current_user.is_authenticated:
                return redirect(url_for("login", next=_rq.path))
            if not puede(current_user.rol, permiso):
                auditar(current_user.usuario, "denegado", _rq.path,
                        _rq.remote_addr, f"falta permiso {permiso}")
                return render_template("sin_permiso.html",
                                        permiso=permiso, user=current_user), 403
            auditar(current_user.usuario, "ver", _rq.path, _rq.remote_addr, "")
            return fn(*a, **kw)
        return wrapper
    return deco


# Reglas de acceso por prefijo de ruta. Se evalúan de la más específica a la
# más general; lo que no calce exige al menos sesión iniciada.
REGLAS_ACCESO = [
    ("/usuarios",        "usuarios"),
    ("/api/usuarios",    "usuarios"),
    ("/vacaciones",      "personal"),
    ("/api/vacaciones",  "personal"),
    ("/api/personal",    "personal"),
    ("/api/inventario",  "operacion"),
    ("/api/tareas",      "operacion"),
    ("/conciliacion",    "finanzas"),
    ("/api/conciliacion", "finanzas"),
    ("/banco",           "finanzas"),
    ("/api/banco",       "finanzas"),
    ("/cash-flow",       "finanzas"),
    ("/api/cash-flow",   "finanzas"),
    ("/temporadas",      "finanzas"),
    ("/api/temporada",   "finanzas"),
    ("/reporte",         "finanzas"),
    ("/api/reporte",     "finanzas"),
    ("/general",         "finanzas"),
    ("/api/facturas",    "finanzas"),
    ("/api/caja-chica",  "finanzas"),
    ("/api/costos",      "finanzas"),
    ("/api/exportaciones", "finanzas"),
    ("/api/comparacion", "finanzas"),
    ("/api/saldo",       "finanzas"),
    ("/api/flujo",       "finanzas"),
    ("/api/dashboard",   "finanzas"),
    ("/",                "finanzas"),
]
LIBRES = ("/login", "/logout", "/static", "/favicon.ico")


@app.before_request
def _guardian():
    """Exige sesión y permiso en TODA la app (menos login/estáticos)."""
    ruta = _rq.path
    if ruta.startswith(LIBRES):
        return None
    if _es_interno():           # generador de PDF con token
        return None
    if not current_user.is_authenticated:
        if ruta.startswith("/api/"):
            return jsonify({"error": "no autenticado"}), 401
        return redirect(url_for("login", next=ruta))
    permiso = next((p for pref, p in REGLAS_ACCESO if ruta.startswith(pref)), None)
    if permiso and not puede(current_user.rol, permiso):
        auditar(current_user.usuario, "denegado", ruta, _rq.remote_addr,
                f"falta permiso {permiso}")
        if ruta.startswith("/api/"):
            return jsonify({"error": "sin permiso"}), 403
        return render_template("sin_permiso.html", permiso=permiso), 403
    # auditar solo navegación (no cada llamada de API, para no inflar el log)
    if not ruta.startswith("/api/"):
        auditar(current_user.usuario, "ver", ruta, _rq.remote_addr, "")
    return None


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    if not hay_usuarios():
        return render_template("login.html",
                                error="No hay usuarios creados. Ejecuta: "
                                      "python scripts/crear_usuario.py")
    if _rq.method == "POST":
        u, msg = verificar(_rq.form.get("usuario", ""), _rq.form.get("password", ""),
                           _rq.remote_addr or "")
        if u:
            login_user(WebUser(u), remember=False)
            _sess.permanent = True
            destino = _rq.args.get("next") or _rq.form.get("next") or ""
            return redirect(destino if destino.startswith("/") else url_for("index"))
        return render_template("login.html", error=msg,
                                usuario=_rq.form.get("usuario", ""))
    return render_template("login.html", next=_rq.args.get("next", ""))


@app.route("/logout")
def logout():
    if current_user.is_authenticated:
        auditar(current_user.usuario, "logout", "/logout", _rq.remote_addr, "")
    logout_user()
    return redirect(url_for("login"))


@app.context_processor
def _inyectar_usuario():
    """Deja el usuario y sus permisos disponibles en todas las plantillas."""
    return {"user": current_user if current_user.is_authenticated else None,
            "puede": (lambda p: current_user.is_authenticated
                      and puede(current_user.rol, p))}


@app.route("/usuarios")
@requiere("usuarios")
def usuarios_page():
    return render_template("usuarios.html", usuarios=listar_usuarios(), roles=ROLES)


@app.route("/api/usuarios", methods=["POST"])
@requiere("usuarios")
def api_crear_usuario():
    b = request.get_json() or {}
    try:
        r = crear_usuario(b.get("usuario", ""), b.get("password", ""),
                          b.get("rol", ""), b.get("nombre", ""))
        auditar(current_user.usuario, "crear_usuario", "/api/usuarios",
                _rq.remote_addr, f"{r['usuario']} ({r['rol']})")
        return jsonify(r)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


class DateEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, date):
            return obj.isoformat()
        return super().default(obj)


app.json_encoder = DateEncoder


@app.route("/")
def index():
    data = get_dashboard_data()
    return render_template("dashboard.html", data=data)


@app.route("/api/dashboard")
def api_dashboard():
    return jsonify(get_dashboard_data())


@app.route("/api/facturas")
def api_facturas():
    return jsonify(get_facturas_summary())


@app.route("/api/costos-cultivo")
def api_costos():
    return jsonify(get_costos_por_cultivo())


@app.route("/api/comparacion-anual")
def api_anual():
    return jsonify(get_comparacion_anual())


@app.route("/api/exportaciones")
def api_exports():
    return jsonify(get_exportaciones())


@app.route("/api/caja-chica")
def api_caja():
    return jsonify(get_caja_chica())


@app.route("/api/banco")
def api_banco():
    return jsonify(get_cuenta_banco())


@app.route("/api/inventario")
def api_inventario():
    return jsonify(get_inventario_resumen())


@app.route("/api/personal")
def api_personal():
    return jsonify(get_personal_resumen())


@app.route("/api/tareas")
def api_tareas():
    return jsonify(get_tareas_resumen())


@app.route("/api/facturas/detalle")
def api_facturas_detalle():
    filtro = request.args.get("filtro", "todas")
    return jsonify(get_facturas_detalle(filtro))


@app.route("/api/banco/movimientos")
def api_banco_movimientos():
    return jsonify(get_movimientos_banco())


@app.route("/cash-flow")
def cash_flow_page():
    return render_template("cash_flow.html")


@app.route("/api/cash-flow")
def api_cash_flow():
    # Por defecto la caja REAL de las dos cuentas (corriente + dólar).
    from modules.cuentas import caja_total
    caja = caja_total()
    saldo = float(request.args.get("saldo", caja["total"]))
    meses = int(request.args.get("meses", 12))
    from datetime import date as _date
    today = _date.today()
    sy, sm = today.year, today.month
    ey, em = sy, sm
    for _ in range(meses - 1):
        em += 1
        if em > 12:
            em = 1
            ey += 1
    cf = get_cash_flow(start=(sy, sm), end=(ey, em),
                       saldo_inicial=saldo)
    labels = [_month_label(y, m) for y, m in cf["months"]]
    return jsonify({
        "labels": labels,
        "saldo_inicio": [cf["saldo"][ym]["saldo_inicio"] for ym in cf["months"]],
        "ingresos": [cf["saldo"][ym]["ingresos"] for ym in cf["months"]],
        "egresos": [cf["saldo"][ym]["egresos"] for ym in cf["months"]],
        "saldo_cierre": [cf["saldo"][ym]["saldo_cierre"] for ym in cf["months"]],
        "ingresos_detail": cf["ingresos"],
        "caja": {                      # desglose del punto de partida
            "clp": caja["clp"],
            "usd": caja["usd"],
            "usd_en_clp": caja["usd_en_clp"],
            "tipo_cambio": caja["tipo_cambio"],
            "total": caja["total"],
            "fecha_clp": caja["fecha_clp"].isoformat() if caja["fecha_clp"] else None,
            "fecha_usd": caja["fecha_usd"].isoformat() if caja["fecha_usd"] else None,
        },
    })


@app.route("/api/cash-flow/replante", methods=["POST"])
def api_replante():
    body = request.get_json() or {}
    result = afford_check(
        cultivo=body.get("cultivo", "AVELLANOS"),
        hc=float(body.get("hc", 0)),
        saldo_proyectado=float(body.get("saldo_proyectado", 0)),
        saldo_minimo=float(body.get("saldo_minimo", 0)),
        costo_por_hc=float(body.get("costo_por_hc", 5_000_000)),
    )
    return jsonify(result)


@app.route("/temporadas")
def temporadas_page():
    return render_template("temporadas.html",
                            temporadas=get_temporadas_disponibles())


@app.route("/general")
def general_page():
    return render_template("general.html",
                            temporadas=get_temporadas_disponibles())


@app.route("/api/facturas-temporada/<path:temporada>")
def api_facturas_temporada(temporada):
    return jsonify(get_facturas_temporada(temporada))


@app.route("/api/saldo-banco")
def api_saldo_banco():
    """Saldo de la cuenta corriente + la cuenta dólar.

    `saldo` se mantiene en CLP por compatibilidad; `caja` trae el total real.
    """
    from modules.cuentas import caja_total, desactualizadas
    base = get_saldo_banco_actual()
    caja = caja_total()
    base["caja"] = {
        "clp": caja["clp"],
        "usd": caja["usd"],
        "usd_en_clp": caja["usd_en_clp"],
        "tipo_cambio": caja["tipo_cambio"],
        "total": caja["total"],
        "fecha_usd": caja["fecha_usd"].isoformat() if caja["fecha_usd"] else None,
        "avisos": desactualizadas(caja),
    }
    return jsonify(base)


@app.route("/api/flujo-mensual")
def api_flujo_mensual():
    meses = int(request.args.get("meses", 12))
    return jsonify(get_flujo_mensual_proyectado(meses=meses))


@app.route("/vacaciones")
def vacaciones_page():
    return render_template("vacaciones.html")


@app.route("/api/vacaciones")
def api_vacaciones():
    return jsonify(get_vacaciones_pendientes())


@app.route("/reporte/<int:year>/<int:month>")
def reporte_mensual_page(year, month):
    data = get_reporte_mensual(year, month)
    return render_template("reporte_mensual.html", data=data)


@app.route("/reporte")
def reporte_mensual_index():
    # Por defecto último mes completo
    from datetime import date as _d
    hoy = _d.today()
    if hoy.month == 1:
        y, m = hoy.year - 1, 12
    else:
        y, m = hoy.year, hoy.month - 1
    return reporte_mensual_page(y, m)


@app.route("/api/reporte/<int:year>/<int:month>")
def api_reporte_mensual(year, month):
    return jsonify(get_reporte_mensual(year, month))


@app.route("/reporte/<int:year>/<int:month>/pdf")
def reporte_pdf_download(year, month):
    """Genera y descarga el PDF del reporte mensual."""
    from flask import send_file
    from modules.cash_flow.reporte_pdf import generar_reporte_pdf
    try:
        pdf_path = generar_reporte_pdf(year, month)
        _MES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"][month]
        return send_file(pdf_path, as_attachment=True,
                          download_name=f"Reporte_{_MES}_{year}.pdf",
                          mimetype="application/pdf")
    except Exception as e:
        return f"Error generando PDF: {e}", 500


@app.route("/api/temporadas")
def api_temporadas():
    return jsonify(get_temporadas_disponibles())


@app.route("/api/temporada/<path:temporada>")
def api_temporada(temporada):
    return jsonify(get_resumen_temporada(temporada))


@app.route("/api/temporadas/comparar")
def api_temporadas_comparar():
    t1 = request.args.get("t1", "")
    t2 = request.args.get("t2", "")
    if not t1 or not t2:
        return jsonify({"error": "t1 y t2 requeridos"}), 400
    return jsonify(compare_temporadas(t1, t2))


@app.route("/conciliacion")
def conciliacion_page():
    return render_template("conciliacion.html")


def _pd_flex(v):
    from datetime import datetime as _dt, date as _date
    if isinstance(v, _dt):
        return v.date()
    if isinstance(v, _date):
        return v
    if isinstance(v, str):
        for f in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return _dt.strptime(v[:10], f).date()
            except ValueError:
                continue
    return None


@app.route("/api/conciliacion/kpis")
def api_conc_kpis():
    """Por Pagar (facturas sin pago, últimos 365d) · Por Cobrar · Saldo."""
    from datetime import date as _date, timedelta
    from openpyxl import load_workbook
    from config import EXCEL_PATH
    hoy = _date.today()
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    # Por pagar: facturas únicas sin fecha de pago (emisión ≤ 365 días)
    ws = wb["Facturas"]
    fact = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        nro = str(row[6] or "").strip()
        if not nro:
            continue
        k = (str(row[3] or "").strip().lower(), nro)
        d = fact.setdefault(k, {"total": 0.0, "pagada": False, "emision": _pd_flex(row[0])})
        try:
            t = float(row[15] or 0)
            if t > d["total"]:
                d["total"] = t
        except (TypeError, ValueError):
            pass
        if row[2] and str(row[2]).strip():
            d["pagada"] = True
    por_pagar = sum(d["total"] for d in fact.values()
                    if not d["pagada"] and d["emision"]
                    and d["emision"] >= hoy - timedelta(days=365))

    # Saldo: último saldo de Cuenta Banco
    ws_b = wb["Cuenta Banco"]
    ult_fecha, saldo = None, None
    for row in ws_b.iter_rows(min_row=2, values_only=True):
        f = _pd_flex(row[0]) if row else None
        if f and (ult_fecha is None or f >= ult_fecha):
            ult_fecha = f
            if row[5] not in (None, ""):
                saldo = float(row[5])
    wb.close()

    # Por cobrar: ingresos esperados de Cosechas (no recibidos)
    por_cobrar = 0.0
    try:
        from modules.cash_flow.projector import load_expected_ingresos
        for ing in load_expected_ingresos():
            if ing.get("estado") != "recibido":
                por_cobrar += float(ing.get("monto_clp") or 0)
    except Exception as e:
        logger.warning(f"kpi por cobrar: {e}")

    return jsonify({"por_pagar": round(por_pagar), "por_cobrar": round(por_cobrar),
                    "saldo": round(saldo or 0),
                    "saldo_fecha": ult_fecha.isoformat() if ult_fecha else None})


@app.route("/api/conciliacion/movimientos")
def api_conc_movimientos():
    """Movimientos del banco con su estado de conciliación (con filtros)."""
    from openpyxl import load_workbook
    from config import EXCEL_PATH
    from modules.conciliacion_store import resumen_estados

    a = request.args
    f_desde = _pd_flex(a.get("desde"))
    f_hasta = _pd_flex(a.get("hasta"))
    q = (a.get("q") or "").strip().lower()
    doc = (a.get("doc") or "").strip().lower()
    cat_f = (a.get("categoria") or "").strip().upper()
    monto_f = a.get("monto")
    limit = min(int(a.get("limit", 300)), 1000)

    try:
        monto_val = float(str(monto_f).replace(".", "").replace(",", ".")) if monto_f else None
    except ValueError:
        monto_val = None

    estados = resumen_estados()
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["Cuenta Banco"]
    movs, cats = [], set()
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]:
            continue
        f = _pd_flex(row[0])
        if not f:
            continue
        try:
            cargo = float(row[3] or 0)
            abono = float(row[4] or 0)
        except (TypeError, ValueError):
            cargo = abono = 0
        if cargo == 0 and abono == 0:
            continue
        cat = str(row[7] or "").strip()
        if cat:
            cats.add(cat)
        # filtros
        if f_desde and f < f_desde:
            continue
        if f_hasta and f > f_hasta:
            continue
        desc = str(row[1] or "")
        if q and q not in desc.lower():
            continue
        ref = str(row[2] or "").strip()
        if doc and doc not in ref.lower():
            continue
        if cat_f and cat_f != cat.upper():
            continue
        monto = cargo if cargo > 0 else abono
        if monto_val is not None and abs(monto - monto_val) > max(1, monto_val * 0.005):
            continue

        est = estados.get(i)
        movs.append({
            "fila": i, "fecha": f.isoformat(), "desc": desc, "doc": ref,
            "categoria": cat, "cargo": cargo, "abono": abono,
            "estado": est["estado"] if est else "por conciliar",
            "saldo": est["saldo"] if est else monto,
            "vinculos": str(row[9] or "") if len(row) > 9 else "",
        })
    wb.close()

    movs.sort(key=lambda m: (m["fecha"], m["fila"]), reverse=True)
    conteos = {
        "todos": len(movs),
        "cargos": sum(1 for m in movs if m["cargo"] > 0),
        "abonos": sum(1 for m in movs if m["abono"] > 0),
        "pendientes": sum(1 for m in movs if m["estado"] != "conciliado"),
    }
    return jsonify({"movimientos": movs[:limit], "conteos": conteos,
                    "categorias": sorted(cats), "truncado": len(movs) > limit})


@app.route("/api/conciliacion/buscar-docs")
def api_conc_buscar_docs():
    """Busca facturas por nº, proveedor o monto (para el modal de vincular)."""
    from collections import defaultdict
    from openpyxl import load_workbook
    from config import EXCEL_PATH
    q = (request.args.get("q") or "").strip().lower()
    if len(q) < 2:
        return jsonify({"docs": []})
    try:
        q_monto = float(q.replace(".", "").replace(",", "."))
    except ValueError:
        q_monto = None

    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["Facturas"]
    grupos = defaultdict(lambda: {"filas": [], "total": 0.0, "pagada": False,
                                    "emision": None, "prov": "", "nro": "", "narch": None})
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]:
            continue
        nro = str(row[6] or "").strip()
        if nro.endswith(".0"):
            nro = nro[:-2]
        prov = str(row[3] or "").strip()
        if not nro:
            continue
        k = (prov.lower(), nro)
        g = grupos[k]
        g["filas"].append(i)
        g["prov"], g["nro"] = prov, nro
        g["emision"] = g["emision"] or _pd_flex(row[0])
        try:
            t = float(row[15] or 0)
            if t > g["total"]:
                g["total"] = t
        except (TypeError, ValueError):
            pass
        if row[2] and str(row[2]).strip():
            g["pagada"] = True
        if len(row) >= 21 and isinstance(row[20], (int, float)):
            g["narch"] = int(row[20])
    wb.close()

    # Cuánto de cada documento ya está asignado (pagos en cuotas)
    from modules.conciliacion_store import _clave_doc, asignado_por_documento
    asignado = asignado_por_documento()

    out = []
    for g in grupos.values():
        hit = (q in g["nro"].lower() or q in g["prov"].lower()
               or (q_monto and g["total"] > 0
                   and abs(g["total"] - q_monto) <= max(1, q_monto * 0.01)))
        if hit:
            asig = asignado.get(_clave_doc(g["nro"], g["prov"]), 0.0)
            saldo = round(g["total"] - asig)
            out.append({"nro": g["nro"], "prov": g["prov"], "total": g["total"],
                         "emision": g["emision"].isoformat() if g["emision"] else "",
                         "pagada": g["pagada"], "filas": g["filas"], "narch": g["narch"],
                         "asignado": round(asig), "saldo_doc": saldo,
                         "estado_doc": ("pendiente" if asig <= 0
                                        else "pagado" if abs(saldo) <= 1 else "parcial")})
    out.sort(key=lambda x: (x["pagada"], -(len(x["emision"]) and 1),
                             x["emision"]), reverse=False)
    return jsonify({"docs": out[:25]})


@app.route("/api/conciliacion/vincular", methods=["POST"])
def api_conc_vincular():
    """Registra un vínculo manual (factura, balance, terceros, no conciliable)."""
    from modules.conciliacion_store import registrar_vinculos, TIPOS_DOC
    b = request.get_json() or {}
    tipo = str(b.get("tipo_doc") or "FACTURA").upper()
    if tipo not in TIPOS_DOC:
        return jsonify({"error": f"tipo inválido: {tipo}"}), 400
    try:
        v = {
            "fila_banco": int(b["fila_banco"]),
            "tipo_doc": tipo,
            "nro_doc": str(b.get("nro_doc") or ""),
            "proveedor": str(b.get("proveedor") or ""),
            "filas_doc": [int(x) for x in (b.get("filas_doc") or [])],
            "fila_doc": ([int(x) for x in (b.get("filas_doc") or [])] or [None])[0],
            "monto_asignado": float(b["monto"]) if b.get("monto") else None,
            "criterio": "manual",
            "nota": str(b.get("nota") or ""),
        }
    except (KeyError, TypeError, ValueError) as e:
        return jsonify({"error": f"payload inválido: {e}"}), 400
    try:
        r = registrar_vinculos([v], usuario="dashboard")
        return jsonify(r)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── FASE 4: conciliación parcial y N:M ───────────────────────────────────

@app.route("/api/conciliacion/detalle")
def api_conc_detalle():
    """Estado de un movimiento: monto, ya asignado y saldo por asignar."""
    from modules.conciliacion_store import saldo_por_asignar
    try:
        fila = int(request.args.get("fila", 0))
    except ValueError:
        return jsonify({"error": "fila inválida"}), 400
    if fila < 2:
        return jsonify({"error": "fila inválida"}), 400
    d = saldo_por_asignar(fila)
    f = d.pop("fecha", None)
    d["fecha"] = f.isoformat() if hasattr(f, "isoformat") else (str(f) if f else None)
    d["fila_banco"] = fila
    return jsonify(d)


@app.route("/api/conciliacion/estado-doc")
def api_conc_estado_doc():
    """Cobertura de un documento y las cuotas que lo pagaron."""
    from modules.conciliacion_store import estado_documento, vinculos_de_documento
    nro = (request.args.get("nro") or "").strip()
    prov = (request.args.get("prov") or "").strip()
    if not nro:
        return jsonify({"error": "falta el n° de documento"}), 400
    try:
        total = float(request.args.get("total") or 0)
    except ValueError:
        total = 0.0
    d = estado_documento(nro, prov, total)
    d["cuotas"] = [
        {**v, "fecha_mov": str(v["fecha_mov"] or "")}
        for v in vinculos_de_documento(nro, prov)
    ]
    return jsonify(d)


@app.route("/api/conciliacion/vincular-multiple", methods=["POST"])
def api_conc_vincular_multiple():
    """Vincula VARIOS documentos a un mismo movimiento (N:M y parcial).

    Ej: un cargo de Copeval que paga 6 notas de débito. Cada documento lleva
    su propio monto asignado; la suma no puede superar lo que queda por asignar.
    """
    from modules.conciliacion_store import (TIPOS_DOC, registrar_vinculos,
                                             saldo_por_asignar)
    b = request.get_json() or {}
    docs = b.get("docs") or []
    if not docs:
        return jsonify({"error": "sin documentos"}), 400
    try:
        fila = int(b["fila_banco"])
    except (KeyError, TypeError, ValueError):
        return jsonify({"error": "fila_banco inválida"}), 400

    estado = saldo_por_asignar(fila)
    disponible = estado["saldo"] if estado["asignado"] else estado["monto"]

    vinculos, suma = [], 0.0
    try:
        for d in docs:
            tipo = str(d.get("tipo_doc") or "FACTURA").upper()
            if tipo not in TIPOS_DOC:
                return jsonify({"error": f"tipo inválido: {tipo}"}), 400
            filas_doc = [int(x) for x in (d.get("filas_doc") or [])]
            monto = float(d.get("monto") or 0)
            if monto <= 0:
                return jsonify({"error": "cada documento necesita un monto > 0"}), 400
            suma += monto
            vinculos.append({
                "fila_banco": fila, "tipo_doc": tipo,
                "nro_doc": str(d.get("nro_doc") or ""),
                "proveedor": str(d.get("proveedor") or ""),
                "filas_doc": filas_doc,
                "fila_doc": (filas_doc or [None])[0],
                "monto_asignado": monto,
                "criterio": str(d.get("criterio") or "manual-multiple"),
                "nota": str(d.get("nota") or ""),
            })
    except (TypeError, ValueError) as e:
        return jsonify({"error": f"payload inválido: {e}"}), 400

    # Tolerancia de $1 por redondeos del IVA
    if suma - disponible > 1:
        return jsonify({
            "error": "La suma asignada supera el saldo del movimiento",
            "asignado_intentado": round(suma),
            "disponible": round(disponible),
        }), 400

    try:
        r = registrar_vinculos(vinculos, usuario=getattr(current_user, "usuario", "")
                                or "dashboard")
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    r["estado"] = saldo_por_asignar(fila)
    return jsonify(r)


@app.route("/api/conciliacion/vinculos")
def api_conc_vinculos():
    """Vínculos existentes de un movimiento (para verlos / deshacer)."""
    from openpyxl import load_workbook
    from config import EXCEL_PATH
    from modules.conciliacion_store import SHEET
    try:
        fila = int(request.args.get("fila", 0))
    except ValueError:
        return jsonify({"vinculos": []})
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    out = []
    if SHEET in wb.sheetnames:
        for row in wb[SHEET].iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            try:
                if int(row[2]) != fila:
                    continue
            except (TypeError, ValueError):
                continue
            out.append({"id": int(row[0]), "tipo": str(row[6] or ""),
                         "nro": str(row[8] or ""), "prov": str(row[9] or ""),
                         "monto": float(row[10] or 0), "criterio": str(row[11] or ""),
                         "nota": str(row[13] or "")})
    wb.close()
    return jsonify({"vinculos": out})


# ── FASE 5: comentarios, export y % conciliado ───────────────────────────

@app.route("/api/conciliacion/comentarios")
def api_conc_comentarios():
    """Notas de un movimiento, o el conteo por fila si no se pasa `fila`."""
    from modules.conciliacion_comentarios import conteo_por_fila, de_movimiento
    fila = request.args.get("fila")
    if fila is None:
        return jsonify({"conteo": {str(k): v for k, v in conteo_por_fila().items()}})
    try:
        return jsonify({"comentarios": de_movimiento(int(fila))})
    except ValueError:
        return jsonify({"error": "fila inválida"}), 400


@app.route("/api/conciliacion/comentar", methods=["POST"])
def api_conc_comentar():
    from modules.conciliacion_comentarios import agregar
    b = request.get_json() or {}
    try:
        r = agregar(int(b["fila_banco"]), b.get("texto", ""),
                    usuario=getattr(current_user, "usuario", "") or "dashboard")
    except (KeyError, TypeError, ValueError) as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify(r)


@app.route("/api/conciliacion/comentario/<int:cid>", methods=["DELETE"])
def api_conc_borrar_comentario(cid):
    from modules.conciliacion_comentarios import eliminar
    return jsonify({"ok": eliminar(cid)})


@app.route("/api/conciliacion/resumen-estado")
def api_conc_resumen_estado():
    """Cuánto está conciliado en un período (para el KPI de la página)."""
    from datetime import datetime as _dt

    from modules.conciliacion_export import recolectar, resumen

    def _d(p):
        v = request.args.get(p)
        try:
            return _dt.strptime(v[:10], "%Y-%m-%d").date() if v else None
        except ValueError:
            return None

    return jsonify(resumen(recolectar("todos", _d("desde"), _d("hasta"))))


@app.route("/api/conciliacion/exportar")
def api_conc_exportar():
    """Descarga un Excel con el estado de conciliación."""
    from datetime import datetime as _dt

    from flask import send_file

    from modules.conciliacion_export import a_excel, recolectar

    estado = (request.args.get("estado") or "todos").lower()
    if estado not in ("todos", "conciliado", "parcial", "pendiente"):
        return jsonify({"error": f"estado inválido: {estado}"}), 400

    def _d(p):
        v = request.args.get(p)
        try:
            return _dt.strptime(v[:10], "%Y-%m-%d").date() if v else None
        except ValueError:
            return None

    desde, hasta = _d("desde"), _d("hasta")
    filas = recolectar(estado, desde, hasta)
    rango = ""
    if desde or hasta:
        rango = f" {desde or '…'} a {hasta or '…'}"
    buf = a_excel(filas, f"Conciliación — {estado}{rango}")
    nombre = f"conciliacion_{estado}_{date.today():%Y%m%d}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=nombre,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/conciliacion/desconciliar", methods=["POST"])
def api_conc_desconciliar():
    from modules.conciliacion_store import desconciliar
    b = request.get_json() or {}
    try:
        ok = desconciliar(int(b.get("id", 0)))
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify({"ok": ok})


@app.route("/api/conciliacion")
def api_conciliacion():
    dias = max(7, min(400, int(request.args.get("dias", 90))))
    usar_ia = request.args.get("ia", "1") == "1"
    from modules.conciliador import analizar, resolver_dudosos_ia
    res = analizar(dias)
    ia = resolver_dudosos_ia(res["dudosos"]) if usar_ia else []

    def ser_match(m):
        c, f = m["cargo"], m["factura"]
        return {"fila_banco": c["fila"], "fecha": str(c["fecha"]),
                "desc": c["desc"], "monto": c["monto"],
                "nro": f["nro"], "prov": f["prov"], "total": f["total"],
                "filas_fact": f["filas"], "criterio": m.get("criterio", "")}

    return jsonify({
        "total_cargos": res["total_cargos"],
        "matches": [ser_match(m) for m in res["confirmados"] + ia],
        "dudosos_descartados": len(res["dudosos"]) - len(ia),
        "sin_factura": [{"fecha": str(c["fecha"]), "desc": c["desc"],
                          "monto": c["monto"], "categoria": c["categoria"]}
                         for c in sorted(res["sin_factura"], key=lambda x: -x["monto"])],
        "fact_sin_pago": [{"nro": d["nro"], "prov": d["prov"], "total": d["total"],
                            "emision": str(d["emision"])}
                           for d in sorted(res["fact_sin_pago"], key=lambda x: -x["total"])],
    })


@app.route("/api/conciliacion/aplicar", methods=["POST"])
def api_conciliacion_aplicar():
    from datetime import datetime as _dt
    from modules.conciliador import aplicar_conciliacion
    body = request.get_json() or {}
    matches = []
    try:
        for m in body.get("matches", []):
            matches.append({
                "cargo": {"fila": int(m["fila_banco"]),
                           "fecha": _dt.strptime(str(m["fecha"])[:10], "%Y-%m-%d").date()},
                "factura": {"nro": str(m["nro"]), "prov": str(m.get("prov", "")),
                             "filas": [int(x) for x in m.get("filas_fact", [])]},
            })
    except Exception as e:
        return jsonify({"error": f"payload inválido: {e}"}), 400
    if not matches:
        return jsonify({"error": "sin matches"}), 400
    try:
        return jsonify(aplicar_conciliacion(matches))
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── FASE 3: vista de sugerencias pareadas ────────────────────────────────

@app.route("/conciliacion/sugerencias")
def conciliacion_sugerencias_page():
    return render_template("conciliacion_sugerencias.html")


@app.route("/api/conciliacion/sugerencias")
def api_conciliacion_sugerencias():
    """Pares cargo ⇄ documento con el motivo de la sugerencia.

    Filtros opcionales: desde/hasta (fecha del movimiento), proveedor, monto
    mínimo. Los pares ya rechazados no vuelven a aparecer.
    """
    from datetime import datetime as _dt

    from modules.conciliador import analizar, explicar, resolver_dudosos_ia

    dias = max(7, min(400, int(request.args.get("dias", 90))))
    usar_ia = request.args.get("ia", "1") == "1"
    prov_q = (request.args.get("proveedor") or "").strip().upper()
    monto_min = float(request.args.get("monto_min") or 0)

    def _d(param):
        v = request.args.get(param)
        try:
            return _dt.strptime(v[:10], "%Y-%m-%d").date() if v else None
        except ValueError:
            return None

    desde, hasta = _d("desde"), _d("hasta")

    res = analizar(dias)
    pares = list(res["confirmados"])
    if usar_ia:
        pares += resolver_dudosos_ia(res["dudosos"])

    out = []
    for m in pares:
        c, f = m["cargo"], m["factura"]
        if desde and c["fecha"] < desde:
            continue
        if hasta and c["fecha"] > hasta:
            continue
        if monto_min and c["monto"] < monto_min:
            continue
        if prov_q and prov_q not in str(f["prov"]).upper() \
                and prov_q not in str(c["desc"]).upper():
            continue
        criterio = m.get("criterio", "")
        dif = float(c["monto"]) - float(f["total"] or 0)
        out.append({
            "fila_banco": c["fila"],
            "fecha": str(c["fecha"]),
            "desc": c["desc"],
            "monto": c["monto"],
            "categoria": c.get("categoria", ""),
            "nro": f["nro"],
            "prov": f["prov"],
            "total": f["total"],
            "emision": str(f["emision"]) if f.get("emision") else None,
            "filas_fact": f["filas"],
            "criterio": criterio,
            "motivo": explicar(criterio),
            "es_ia": criterio.startswith("IA"),
            "diferencia": round(dif, 2),
        })
    out.sort(key=lambda x: -x["monto"])
    return jsonify({
        "total": len(out),
        "dias": dias,
        "sugerencias": out,
        "rechazados_ocultos": len(_rechazos_listar()),
    })


def _rechazos_listar():
    from modules.conciliacion_rechazos import listar
    try:
        return listar()
    except Exception:
        return []


@app.route("/api/conciliacion/rechazar", methods=["POST"])
def api_conciliacion_rechazar():
    """Descarta pares sugeridos para que no se vuelvan a proponer."""
    from modules.conciliacion_rechazos import registrar
    body = request.get_json() or {}
    pares = body.get("pares") or []
    if not pares:
        return jsonify({"error": "sin pares"}), 400
    try:
        n = registrar(pares,
                      usuario=getattr(current_user, "usuario", "") or "",
                      motivo=body.get("motivo", ""))
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify({"rechazados": n})


@app.route("/api/conciliacion/rechazos")
def api_conciliacion_rechazos():
    return jsonify({"rechazos": _rechazos_listar()})


@app.route("/api/conciliacion/deshacer-rechazo", methods=["POST"])
def api_conciliacion_deshacer_rechazo():
    from modules.conciliacion_rechazos import deshacer
    body = request.get_json() or {}
    try:
        ok = deshacer(int(body.get("id", 0)))
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify({"ok": ok})


@app.route("/banco/revisar")
def banco_revisar_page():
    return render_template("banco_revisar.html",
                            categorias=BANCO_CATEGORIAS_VALIDAS)


@app.route("/api/banco/revisar")
def api_banco_revisar():
    return jsonify(get_banco_revisar())


@app.route("/api/banco/revisar/<int:fila>", methods=["POST"])
def api_banco_revisar_update(fila):
    body = request.get_json() or {}
    categoria = body.get("categoria", "")
    cultivo = body.get("cultivo", "GENERAL")
    if not categoria:
        return jsonify({"error": "categoria requerida"}), 400
    try:
        result = update_banco_categoria(fila, categoria, cultivo)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("DASHBOARD_PORT", 5000))
    print(f"\n{'='*50}")
    print(f"  Dashboard Agrícola Santa Elisa")
    print(f"  http://localhost:{port}")
    print(f"{'='*50}\n")
    app.run(host="0.0.0.0", port=port, debug=True)
