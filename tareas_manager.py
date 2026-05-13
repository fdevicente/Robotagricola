"""
tareas_manager.py — Gestión de Bitácora y Tareas en Excel
Hojas: Tareas, Bitácora
"""
import logging
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from config import EXCEL_PATH

logger = logging.getLogger(__name__)

TAREAS_SHEET = "Tareas"
BITACORA_SHEET = "Bitácora"

TAREAS_HEADERS = [
    "ID", "Fecha Creación", "Descripción", "Prioridad",
    "Estado", "Responsable", "Fecha Límite", "Fecha Completada", "Observaciones"
]
BITACORA_HEADERS = ["Fecha", "Hora", "Registro", "Categoría"]


def _open_wb():
    return load_workbook(EXCEL_PATH)


def _create_sheet(wb, name, headers, widths, color):
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = w
        logger.info(f"Hoja '{name}' creada")
    return wb[name]


def _ensure_tareas(wb):
    return _create_sheet(wb, TAREAS_SHEET, TAREAS_HEADERS,
                         [6, 12, 50, 10, 14, 20, 12, 12, 40], "FF6F00")


def _ensure_bitacora(wb):
    return _create_sheet(wb, BITACORA_SHEET, BITACORA_HEADERS,
                         [12, 8, 60, 18], "5D4037")


def crear_hojas_tareas():
    """Crea las hojas Tareas y Bitácora si no existen."""
    wb = _open_wb()
    _ensure_tareas(wb)
    _ensure_bitacora(wb)
    wb.save(EXCEL_PATH)


def _next_task_id(ws):
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and isinstance(row[0], (int, float)):
            max_id = max(max_id, int(row[0]))
    return max_id + 1


def crear_tarea(descripcion: str, prioridad: str = "Media",
                responsable: str = "", fecha_limite: str = "") -> dict:
    """Crea una nueva tarea. Retorna dict con ID y datos."""
    wb = _open_wb()
    ws = _ensure_tareas(wb)
    task_id = _next_task_id(ws)
    hoy = date.today().strftime("%Y-%m-%d")
    ws.append([
        task_id, hoy, descripcion, prioridad,
        "Pendiente", responsable, fecha_limite or "", "", ""
    ])
    wb.save(EXCEL_PATH)
    logger.info(f"Tarea #{task_id} creada: {descripcion}")
    return {"id": task_id, "descripcion": descripcion, "prioridad": prioridad,
            "estado": "Pendiente", "responsable": responsable, "fecha_limite": fecha_limite}


def listar_tareas(estado: str = None) -> list[dict]:
    """Lista tareas. Si estado=None, muestra Pendiente + En Progreso."""
    wb = _open_wb()
    ws = _ensure_tareas(wb)
    tareas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        t = {
            "id": int(row[0]),
            "fecha_creacion": str(row[1] or ""),
            "descripcion": str(row[2] or ""),
            "prioridad": str(row[3] or "Media"),
            "estado": str(row[4] or "Pendiente"),
            "responsable": str(row[5] or ""),
            "fecha_limite": str(row[6] or ""),
            "fecha_completada": str(row[7] or ""),
            "observaciones": str(row[8] or ""),
        }
        if estado:
            if t["estado"].lower() == estado.lower():
                tareas.append(t)
        else:
            if t["estado"] in ("Pendiente", "En Progreso"):
                tareas.append(t)
    wb.close()
    return tareas


def actualizar_tarea(task_id: int, estado: str = None, observaciones: str = None) -> bool:
    """Actualiza estado y/o observaciones de una tarea."""
    wb = _open_wb()
    ws = _ensure_tareas(wb)
    for row_idx in range(2, ws.max_row + 1):
        cell_id = ws.cell(row=row_idx, column=1).value
        if cell_id and int(cell_id) == task_id:
            if estado:
                ws.cell(row=row_idx, column=5).value = estado
                if estado == "Hecho":
                    ws.cell(row=row_idx, column=8).value = date.today().strftime("%Y-%m-%d")
            if observaciones:
                prev = ws.cell(row=row_idx, column=9).value or ""
                ws.cell(row=row_idx, column=9).value = (
                    f"{prev}\n{date.today().strftime('%d/%m')}: {observaciones}" if prev
                    else f"{date.today().strftime('%d/%m')}: {observaciones}")
            wb.save(EXCEL_PATH)
            logger.info(f"Tarea #{task_id} actualizada: estado={estado}")
            return True
    wb.close()
    return False


def registrar_bitacora(registro: str, categoria: str = "General") -> bool:
    """Registra una entrada en la bitácora."""
    wb = _open_wb()
    ws = _ensure_bitacora(wb)
    ahora = datetime.now()
    ws.append([
        ahora.strftime("%Y-%m-%d"),
        ahora.strftime("%H:%M"),
        registro,
        categoria,
    ])
    wb.save(EXCEL_PATH)
    logger.info(f"Bitácora: {registro[:50]}")
    return True


def listar_bitacora(dias: int = 7) -> list[dict]:
    """Lista entradas de bitácora de los últimos N días."""
    wb = _open_wb()
    ws = _ensure_bitacora(wb)
    hoy = date.today()
    entradas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        try:
            if isinstance(row[0], str):
                fecha = datetime.strptime(row[0][:10], "%Y-%m-%d").date()
            else:
                fecha = row[0] if isinstance(row[0], date) else row[0].date()
            if (hoy - fecha).days <= dias:
                entradas.append({
                    "fecha": str(fecha),
                    "hora": str(row[1] or ""),
                    "registro": str(row[2] or ""),
                    "categoria": str(row[3] or ""),
                })
        except Exception:
            continue
    wb.close()
    return entradas


def resumen_tareas_semana() -> dict:
    """Resumen semanal de tareas para el reporte."""
    tareas = listar_tareas()
    completadas = listar_tareas("Hecho")
    hoy = date.today()
    # Filtrar completadas esta semana
    completadas_semana = []
    for t in completadas:
        try:
            fc = datetime.strptime(t["fecha_completada"][:10], "%Y-%m-%d").date()
            if (hoy - fc).days <= 7:
                completadas_semana.append(t)
        except Exception:
            continue
    pendientes = [t for t in tareas if t["estado"] == "Pendiente"]
    en_progreso = [t for t in tareas if t["estado"] == "En Progreso"]
    return {
        "pendientes": len(pendientes),
        "en_progreso": len(en_progreso),
        "completadas_semana": len(completadas_semana),
        "tareas_pendientes": pendientes,
        "tareas_progreso": en_progreso,
    }
