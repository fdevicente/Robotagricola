"""
Cron 18:00 diario: trae movimientos del banco, matchea con facturas pendientes,
categoriza los no-factura via Claude, hace backup.

Uso: py -3.11 daily_banco_18h.py
     py -3.11 daily_banco_18h.py --skip-scrape
     py -3.11 daily_banco_18h.py --limit-match 100
"""
import argparse
import logging
import sys

from openpyxl import load_workbook

from config import EXCEL_PATH
from excel_manager import (
    CUENTA_BANCO_SHEET, guardar_movimientos_banco,
    COL_BANCO_TIPO,
)
from infrastructure.backups import backup_master
from modules.cash_flow.matcher import match_new_bank_movements
from modules.cash_flow.categorizer import categorize_bank_movement

logging.basicConfig(level=logging.INFO,
                     format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)


def _categorize_uncategorized_cargos(excel_path: str, limit: int | None = None) -> dict:
    """Categoriza cargos sin Tipo (los que matcher no resolvio)."""
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[CUENTA_BANCO_SHEET]
    target_rows = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=10,
                                              values_only=True), start=2):
        if not row[0]:  # fecha
            continue
        try:
            cargo = float(row[3] or 0)
        except (TypeError, ValueError):
            continue
        if cargo <= 0:
            continue
        if row[6]:  # Tipo
            continue
        target_rows.append(r_idx)
    wb.close()
    if limit is not None:
        target_rows = target_rows[:limit]

    counts = {"categorizados": 0, "errors": 0, "total": len(target_rows)}
    for r in target_rows:
        try:
            categorize_bank_movement(r, excel_path=excel_path)
            counts["categorizados"] += 1
        except Exception as e:
            logger.error(f"Cat banco fila {r}: {e}")
            counts["errors"] += 1
    return counts


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--skip-scrape", action="store_true",
                         help="No correr scraper, solo matcher + categorizer")
    parser.add_argument("--limit-match", type=int, default=None,
                         help="Limite de movs para matcher (debug)")
    parser.add_argument("--limit-cat", type=int, default=None,
                         help="Limite de cargos para categorizer (debug)")
    args = parser.parse_args()

    logger.info("=== Daily Banco 18:00 ===")

    logger.info("1/5 Backup preventivo")
    backup_master(reason="pre-daily-18h")

    if args.skip_scrape:
        logger.info("2/5 SKIP scraper")
        save_report = {"nuevos": 0, "duplicados": 0}
    else:
        logger.info("2/5 Scraping Scotiabank...")
        from scotiabank_scraper import sync_scotiabank_movements
        try:
            movs = sync_scotiabank_movements()
            save_report = guardar_movimientos_banco(movs)
            logger.info(f"   Banco: {save_report}")
        except Exception as e:
            logger.error(f"Scraper fallo: {e}")
            save_report = {"nuevos": 0, "duplicados": 0, "error": str(e)}

    logger.info(f"3/5 Matcheando facturas vs banco (limit={args.limit_match})...")
    match_report = match_new_bank_movements(limit=args.limit_match)
    logger.info(f"   Match: {match_report}")

    logger.info(f"4/5 Categorizando cargos no-factura (limit={args.limit_cat})...")
    cat_report = _categorize_uncategorized_cargos(EXCEL_PATH, limit=args.limit_cat)
    logger.info(f"   Categorizacion: {cat_report}")

    logger.info("5/5 Backup post-18h")
    backup_master(reason="post-daily-18h")

    logger.info("=== Daily completo ===")
    logger.info(f"Mov nuevos:   {save_report.get('nuevos', 0)}")
    logger.info(f"Auto-matched: {match_report.get('auto_matched', 0)}")
    logger.info(f"Ambiguos:     {match_report.get('ambiguous', 0)}")
    logger.info(f"Cargos cat:   {cat_report.get('categorizados', 0)}")


if __name__ == "__main__":
    sys.exit(main())
