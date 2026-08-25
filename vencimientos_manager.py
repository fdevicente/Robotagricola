"""vencimientos_manager.py — Control de vencimiento de insumos.

Hoja 'Vencimientos': registra productos vencibles (fitosanitarios,
fertilizantes) con fecha de compra y vencimiento. Calcula alertas según
vida útil consumida: 50%, 10% y vencido.
"""
import logging
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import EXCEL_PATH
from excel_manager import _save_wb  # guardado con reintentos si Excel está abierto

logger = logging.getLogger(__name__)

VENC_SHEET = "Vencimientos"
VENC_HEADERS = [
    "Producto", "Proveedor", "N° Factura", "Fecha Compra",
    "Fecha Vencimiento", "Vida Útil (días)", "% Restante", "Estado",
    "Fecha Registro",
]
_WIDTHS = [32, 24, 14, 14, 16, 14, 12, 14, 14]

# Estados
ST_PENDIENTE = "PENDIENTE"   # sin fecha de vencimiento aún
ST_VIGENTE = "VIGENTE"
ST_ALERTA_50 = "ALERTA 50%"
ST_ALERTA_10 = "ALERTA 10%"
ST_VENCIDO = "VENCIDO"
ST_NO_VENCE = "NO VENCE"

# Proveedores cuyos productos típicamente vencen (insumos agrícolas)
PROVEEDORES_INSUMOS = [
    "copeval", "martinez y valdivieso", "martinez valdivieso", "cals",
    "gmt", "cna", "vals bio", "agrokimun", "bayer", "anasac", "agrospec",
    "hidalga", "soquimich", "sqm", "basf", "syngenta", "agroadvance",
]


def _open_wb():
    return load_workbook(EXCEL_PATH)


def _pd(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(v[:10], fmt).date()
            except ValueError:
                pass
    return None


def _estilo_header(ws):
    fill = PatternFill("solid", fgColor="00695C")
    for i, h in enumerate(VENC_HEADERS, 1):
        c = ws.cell(1, i, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, w in enumerate(_WIDTHS, 1):
        ws.column_dimensions[chr(64 + i)].width = w


def crear_hoja_vencimientos():
    wb = _open_wb()
    if VENC_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(VENC_SHEET)
        _estilo_header(ws)
        _save_wb(wb)
        logger.info("Hoja Vencimientos creada")
    wb.close()


def calcular_estado(fecha_compra, fecha_venc, hoy=None) -> tuple[str, float | None, int | None]:
    """Devuelve (estado, pct_restante, vida_util_dias).

    pct_restante: fracción del tiempo de vida que queda (0..1).
    """
    hoy = hoy or date.today()
    fc = _pd(fecha_compra)
    fv = _pd(fecha_venc)
    if not fv:
        return ST_PENDIENTE, None, None
    if not fc:
        # sin fecha de compra, no se puede calcular vida útil; usar solo venc
        if hoy > fv:
            return ST_VENCIDO, 0.0, None
        return ST_VIGENTE, None, None
    vida = (fv - fc).days
    if vida <= 0:
        # vencimiento <= compra: tratar como vencido
        return (ST_VENCIDO if hoy >= fv else ST_VIGENTE), 0.0, vida
    restante = (fv - hoy).days
    pct = restante / vida
    if hoy > fv:
        return ST_VENCIDO, 0.0, vida
    if pct <= 0.10:
        return ST_ALERTA_10, pct, vida
    if pct <= 0.50:
        return ST_ALERTA_50, pct, vida
    return ST_VIGENTE, pct, vida


def es_proveedor_insumo(proveedor: str) -> bool:
    p = (proveedor or "").lower()
    return any(k in p for k in PROVEEDORES_INSUMOS)


def agregar_pendiente(producto: str, proveedor: str, nro_factura: str,
                       fecha_compra) -> bool:
    """Agrega un producto pendiente de fecha de vencimiento (sin duplicar)."""
    crear_hoja_vencimientos()
    wb = _open_wb()
    ws = wb[VENC_SHEET]
    fc = _pd(fecha_compra)
    # Evitar duplicado exacto (mismo producto + factura)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if (str(row[0] or "").strip().lower() == producto.strip().lower()
                and str(row[2] or "").strip() == str(nro_factura).strip()):
            wb.close()
            return False
    ws.append([
        producto.strip(), (proveedor or "").strip(), str(nro_factura or ""),
        fc.isoformat() if fc else "", "", None, None, ST_PENDIENTE,
        date.today().isoformat(),
    ])
    _save_wb(wb)
    wb.close()
    logger.info(f"Vencimiento pendiente: {producto} (F{nro_factura})")
    return True


def listar_pendientes() -> list[dict]:
    """Productos que aún no tienen fecha de vencimiento ni 'no vence'."""
    crear_hoja_vencimientos()
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[VENC_SHEET]
    out = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        estado = str(row[7] or "")
        if estado != ST_PENDIENTE:
            continue
        out.append({
            "fila": r_idx, "producto": str(row[0]),
            "proveedor": str(row[1] or ""), "nro_factura": str(row[2] or ""),
            "fecha_compra": str(row[3] or ""),
        })
    wb.close()
    return out


def registrar_vencimiento(fila: int, fecha_venc, no_vence: bool = False) -> bool:
    """Fija la fecha de vencimiento (o 'no vence') de un pendiente por fila."""
    crear_hoja_vencimientos()
    wb = _open_wb()
    ws = wb[VENC_SHEET]
    if no_vence:
        ws.cell(fila, 5).value = ""
        ws.cell(fila, 6).value = None
        ws.cell(fila, 7).value = None
        ws.cell(fila, 8).value = ST_NO_VENCE
    else:
        fv = _pd(fecha_venc)
        if not fv:
            wb.close()
            return False
        fc = ws.cell(fila, 4).value
        estado, pct, vida = calcular_estado(fc, fv)
        ws.cell(fila, 5).value = fv.isoformat()
        ws.cell(fila, 6).value = vida
        ws.cell(fila, 7).value = round(pct * 100, 1) if pct is not None else None
        ws.cell(fila, 8).value = estado
    _save_wb(wb)
    wb.close()
    return True


def actualizar_estados():
    """Recalcula estados/% de todos los registros con fecha (para cron/reporte)."""
    crear_hoja_vencimientos()
    wb = _open_wb()
    ws = wb[VENC_SHEET]
    cambios = 0
    for r in range(2, ws.max_row + 1):
        if not ws.cell(r, 1).value:
            continue
        estado_actual = str(ws.cell(r, 8).value or "")
        if estado_actual in (ST_NO_VENCE, ST_PENDIENTE):
            continue
        fc = ws.cell(r, 4).value
        fv = ws.cell(r, 5).value
        if not fv:
            continue
        estado, pct, vida = calcular_estado(fc, fv)
        ws.cell(r, 6).value = vida
        ws.cell(r, 7).value = round(pct * 100, 1) if pct is not None else None
        if estado != estado_actual:
            ws.cell(r, 8).value = estado
            cambios += 1
    _save_wb(wb)
    wb.close()
    return cambios


def listar_alertas() -> dict:
    """Productos en alerta (50%, 10%, vencidos) para reporte/comando."""
    actualizar_estados()
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[VENC_SHEET]
    alerta_50, alerta_10, vencidos = [], [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        estado = str(row[7] or "")
        item = {
            "producto": str(row[0]), "proveedor": str(row[1] or ""),
            "nro_factura": str(row[2] or ""),
            "fecha_vencimiento": str(row[4] or ""),
            "pct_restante": row[6],
        }
        if estado == ST_VENCIDO:
            vencidos.append(item)
        elif estado == ST_ALERTA_10:
            alerta_10.append(item)
        elif estado == ST_ALERTA_50:
            alerta_50.append(item)
    wb.close()
    return {"vencidos": vencidos, "alerta_10": alerta_10, "alerta_50": alerta_50}
