"""
fix_master_completo.py
Recalcula TODO el MASTER a partir de O (Total por Item) y K (Cantidad) como fuentes de verdad.

Para cada fila:
  - NETO = (O - Especifico) / 1.19   (o NETO = O - Especifico si exenta)
  - J (Unitario) = NETO / K
  - L (TOTAL NETO) = formula =J*K
  - M (IVA) = NETO * 0.19            (o 0 si exenta)
  - P (TOTAL FACTURA) = suma de O agrupado por (Proveedor + Nro Factura)
"""
from openpyxl import load_workbook
from datetime import datetime
import shutil
import os

BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(BASE, "..", "MASTER Agricola Santa Elisa.xlsx")

# 1. BACKUP
ts = datetime.now().strftime("%Y%m%d_%H%M%S")
backup = EXCEL.replace(".xlsx", f"_BACKUP_{ts}.xlsx")
shutil.copy2(EXCEL, backup)
print(f"Backup: {backup}")

# 2. ABRIR
wb = load_workbook(EXCEL)
ws = wb["Facturas"]
print(f"Filas: {ws.max_row}")

KEYWORDS_EXENTA = ("exenta", "exento", "no afecta", "no afecto", "boleta de honorario")

# 3. RECALCULAR J, L, M fila por fila
fixed = 0
skipped = 0
grupos = {}  # (proveedor, nro_factura) -> suma de O

for row in range(2, ws.max_row + 1):
    prov = ws.cell(row=row, column=4).value or ""
    nro = str(ws.cell(row=row, column=7).value or "").strip()
    doc = str(ws.cell(row=row, column=6).value or "").lower()
    k = ws.cell(row=row, column=11).value  # Cantidad
    n = ws.cell(row=row, column=14).value  # Especifico
    o = ws.cell(row=row, column=15).value  # Total por Item

    if o is None or k is None:
        skipped += 1
        continue

    try:
        of = float(o)
        kf = float(k)
        nf = float(n) if n else 0.0
    except (TypeError, ValueError):
        skipped += 1
        continue

    if kf == 0:
        skipped += 1
        continue

    es_exenta = any(x in doc for x in KEYWORDS_EXENTA)

    if es_exenta:
        neto = round(of - nf)
        iva = 0
    else:
        base_iva = of - nf
        neto = round(base_iva / 1.19)
        iva = round(base_iva - neto)

    unitario = neto / kf

    # Escribir J (Unitario)
    cell_j = ws.cell(row=row, column=10)
    if abs(unitario - round(unitario)) < 0.001:
        cell_j.value = int(round(unitario))
    else:
        cell_j.value = round(unitario, 6)
    cell_j.number_format = '#,##0.000'

    # Escribir L (TOTAL NETO) como formula
    cell_l = ws.cell(row=row, column=12)
    cell_l.value = f"=J{row}*K{row}"
    cell_l.number_format = '#,##0'

    # Escribir M (IVA)
    cell_m = ws.cell(row=row, column=13)
    cell_m.value = iva
    cell_m.number_format = '#,##0'

    # Acumular para TOTAL FACTURA
    key = (str(prov).strip(), nro)
    grupos.setdefault(key, {"total": 0.0, "filas": []})
    grupos[key]["total"] += of
    grupos[key]["filas"].append(row)

    fixed += 1

# 4. RECALCULAR P (TOTAL FACTURA) por grupo
for key, info in grupos.items():
    total_factura = round(info["total"])
    for row in info["filas"]:
        cell_p = ws.cell(row=row, column=16)
        cell_p.value = total_factura
        cell_p.number_format = '#,##0'

# 5. GUARDAR
wb.save(EXCEL)
print(f"\nListo: {fixed} filas corregidas, {skipped} saltadas")
print(f"Grupos (facturas): {len(grupos)}")
print("Guardado OK.")
