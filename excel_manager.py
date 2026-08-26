"""
excel_manager.py
Gestiona lectura y escritura en MASTER Agricola Santa Elisa.xlsx
Hoja: Facturas (columnas 1-15)
Hoja: Boletas  (columnas 1-7) — caja chica
"""
import logging
import time
from datetime import date, datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from config import EXCEL_PATH

logger = logging.getLogger(__name__)


def _save_wb(wb, path=None, intentos=5, espera=4):
    """Guarda el workbook con reintentos si el archivo está bloqueado por Excel.

    El destino se resuelve EN CADA LLAMADA, no al importar el módulo: con el
    default fijo, cualquier código que cargara otro archivo y llamara
    `_save_wb(wb)` sin argumento terminaba escribiéndolo encima del Master real.
    """
    path = path or EXCEL_PATH
    for intento in range(1, intentos + 1):
        try:
            wb.save(path)
            return True
        except PermissionError:
            if intento < intentos:
                logger.warning(
                    f"Excel bloqueado (intento {intento}/{intentos}). "
                    f"Cierra el archivo en Excel y espera {espera}s..."
                )
                time.sleep(espera)
            else:
                logger.error(
                    "No se pudo guardar: el archivo sigue abierto en Excel. "
                    "Ciérralo y vuelve a intentarlo."
                )
                raise

SHEET_NAME = "Facturas"
BOLETAS_SHEET = "Boletas"
CAJA_CHICA_SHEET = "Caja Chica"
BANCO_SHEET = "Cuenta Banco"
TOTAL_COLS = 16
IVA_RATE   = 0.19

# --- Hojas nuevas (Fase 1 Cash Flow) ---
COSECHAS_SHEET = "Cosechas"
GUIAS_SHEET = "Guias Despacho"
FLUJO_CAJA_SHEET = "Flujo Caja"
AJUSTES_SHEET = "Ajustes Manuales"
CONFIG_SHEET = "Config"
HECTAREAS_SHEET = "Hectareas"
CUENTA_BANCO_SHEET = BANCO_SHEET  # alias para Cash Flow

# Columnas nuevas en Facturas (despues de col 16)
COL_CATEGORIA = 17       # Q
COL_CULTIVO = 18          # R
COL_CONFIANZA = 19        # S
COL_CATEGORIZADO_POR = 20 # T
# Col 21 (U, "N° Archivo") la administra modules/correlativo.py
COL_ARCHIVO_DRIVE = 22    # V — enlace al PDF subido a Drive (integración Drive)

# Columnas nuevas en Cuenta Banco (despues de col 6)
COL_BANCO_TIPO = 7             # G
COL_BANCO_CATEGORIA = 8        # H
COL_BANCO_CULTIVO = 9           # I
COL_BANCO_FACTURA_LINK = 10     # J

# Categorias validas
CATEGORIAS = [
    "Mano de obra planta",
    "Mano de obra temporal",
    "Fertilizantes",
    "Fitosanitarios",
    "Combustible",
    "Maquinaria - mantencion",
    "Riego",
    "Servicios profesionales",
    "Arriendos / Patentes / Seguros",
    "Inversion / Replante",
    "Caja chica / Imprevistos",
]

CULTIVOS = ["NOGALES", "CEREZOS", "AVELLANOS", "GENERAL"]


def _ensure_sheet_with_headers(wb, name, headers):
    """Crea hoja con headers si no existe."""
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        ws.append(headers)


def ensure_cash_flow_sheets(path=None):
    """Crea hojas nuevas en Master si no existen. Idempotente."""
    from config import CASH_FLOW_CONFIG
    path = path or EXCEL_PATH
    wb = load_workbook(path)

    _ensure_sheet_with_headers(wb, COSECHAS_SHEET, [
        "Año", "Cultivo", "Kg total", "Exportadora", "Kg asignados",
        "Precio USD/kg", "N° cuotas", "Cuota #", "Fecha estimada",
        "Monto USD estimado", "Tipo cuota", "Estado",
        "Fecha real recibido", "Monto real recibido", "Moneda recibida", "Notas"
    ])

    _ensure_sheet_with_headers(wb, GUIAS_SHEET, [
        "Fecha", "N° Guía", "Cultivo", "Kg", "Exportadora destino",
        "Camión / Conductor", "Sector / Equipo", "Año cosecha",
        "Origen", "PDF_path", "Notas"
    ])

    _ensure_sheet_with_headers(wb, AJUSTES_SHEET, [
        "Fecha agregado", "Mes proyectado", "Categoria",
        "Cultivo", "Monto", "Razón", "Activo"
    ])

    # Cuenta en dólares: a los exportadores se les cobra en USD y ese efectivo
    # no pasa por la cuenta corriente. Ver modules/cuentas.py
    from modules.cuentas import DOLAR_SHEET, DOLAR_HEADERS
    _ensure_sheet_with_headers(wb, DOLAR_SHEET, DOLAR_HEADERS)

    if CONFIG_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(CONFIG_SHEET)
        ws.append(["Parámetro", "Valor"])
        for k, v in CASH_FLOW_CONFIG.items():
            ws.append([k, v])

    if HECTAREAS_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(HECTAREAS_SHEET)
        ws.append(["Año", "Nogales", "Cerezos", "Avellanos", "Notas"])
        ws.append([2024, 65, 1.8, 0, "Sin avellanos"])
        ws.append([2025, 54, 3.8, 11.5, "Inicio replante avellanos"])
        ws.append([2026, 46.82, 3.9, 16.6, "Replante avellanos en curso"])

    if FLUJO_CAJA_SHEET not in wb.sheetnames:
        wb.create_sheet(FLUJO_CAJA_SHEET)

    _save_wb(wb, path)
    wb.close()


def ensure_new_columns(path=None):
    """Agrega headers de columnas nuevas si no existen. Idempotente."""
    path = path or EXCEL_PATH
    wb = load_workbook(path)

    # Facturas: cols Q-T
    ws = wb[SHEET_NAME]
    new_fact = {COL_CATEGORIA: "Categoria", COL_CULTIVO: "Cultivo",
                COL_CONFIANZA: "Confianza", COL_CATEGORIZADO_POR: "Categorizado_por",
                COL_ARCHIVO_DRIVE: "Archivo Drive"}
    for col, header in new_fact.items():
        if ws.cell(1, col).value != header:
            ws.cell(1, col, header)

    # Cuenta Banco: cols G-J
    if CUENTA_BANCO_SHEET in wb.sheetnames:
        ws2 = wb[CUENTA_BANCO_SHEET]
        new_banco = {COL_BANCO_TIPO: "Tipo", COL_BANCO_CATEGORIA: "Categoria",
                     COL_BANCO_CULTIVO: "Cultivo", COL_BANCO_FACTURA_LINK: "Factura_linkeada"}
        for col, header in new_banco.items():
            if ws2.cell(1, col).value != header:
                ws2.cell(1, col, header)

    _save_wb(wb, path)
    wb.close()

# Columnas hoja Boletas (se mantiene para registro detallado):
BOLETAS_HEADERS = ["Fecha", "Comercio / Proveedor", "RUT", "Detalle", "Monto Total", "Nº Boleta", "Observaciones"]

# Columnas hoja Caja Chica (control de saldo):
CAJA_HEADERS = ["Fecha", "Tipo", "Detalle", "Comercio / Proveedor", "Nº Boleta", "Ingreso", "Egreso", "Saldo"]

# Columnas hoja Cuenta Banco:
BANCO_HEADERS = ["Fecha", "Descripcion", "Referencia", "Cargo", "Abono", "Saldo"]


def _open_wb():
    return load_workbook(EXCEL_PATH)


def _es_exenta(item: dict) -> bool:
    """Detecta si el documento es exento/no afecto (sin IVA).
    Boleta de Honorarios tampoco tiene IVA."""
    doc = str(item.get("Documento") or "").lower()
    return any(k in doc for k in ("exenta", "exento", "no afecta", "no afecto", "boleta de honorario"))


def _compute_totals(item: dict) -> dict:
    """Calcula TOTAL NETO, IVA y TOTAL.

    Dos modos:
    - Ítem único con Total Factura anchor: Monto/TOTAL = Total Factura → cálculo hacia atrás.
    - Multi-ítem o sin anchor: Monto/TOTAL = NETO de la línea → cálculo hacia adelante.
    Facturas exentas/no afectas → IVA = 0.
    """
    exenta   = _es_exenta(item)
    iva_rate = 0 if exenta else IVA_RATE

    total_anchor = float(item.get("Total Factura") or 0)
    total_linea  = float(item.get("Monto / TOTAL") or 0)

    # Para ítem único: si no hay total de línea, usar el anchor del documento
    if total_linea == 0 and total_anchor > 0:
        total_linea = total_anchor

    # Impuesto Específico (combustibles): debe restarse del total antes de calcular NETO/IVA
    imp_esp = float(item.get("Impuesto Especifico") or 0)

    if total_linea > 0:
        # Detectar si Monto/TOTAL es el Total Factura (con IVA) → cálculo hacia atrás
        # Esto ocurre para ítems únicos donde _limpiar_items fijó Monto/TOTAL = Total Factura
        es_total_con_iva = (total_anchor > 0 and abs(total_linea - round(total_anchor)) < 2)

        if es_total_con_iva:
            # Cálculo hacia atrás desde Total Factura (incluye IVA)
            total = total_linea
            if exenta:
                neto = round(total - imp_esp)
                iva  = 0
            else:
                base_iva = total - imp_esp
                neto = round(base_iva / (1 + IVA_RATE))
                iva  = round(base_iva - neto)
        else:
            # Cálculo hacia adelante: Monto/TOTAL = NETO de la línea (sin IVA)
            neto = round(total_linea)
            if exenta:
                iva = 0
            else:
                iva = round(neto * iva_rate)
            total = neto + iva + round(imp_esp)
    else:
        # Fallback: calcular hacia adelante desde precio unitario × cantidad
        try:
            qty   = float(item.get("Cantidad") or 1)
            unit  = float(item.get("Valor unitario") or 0)
            neto  = round(qty * unit)
            iva   = round(neto * iva_rate)
            total = neto + iva + round(imp_esp)
        except (TypeError, ValueError):
            neto = iva = total = 0

    item["TOTAL NETO"]    = neto
    item["IVA"]           = iva
    item["Monto / TOTAL"] = total
    return item


def _vencimiento(fecha_emision):
    """Si no hay fecha de vencimiento, calcula +1 mes."""
    if not fecha_emision:
        return None
    try:
        if isinstance(fecha_emision, (date, datetime)):
            dt = fecha_emision if isinstance(fecha_emision, date) else fecha_emision.date()
        else:
            dt = datetime.strptime(str(fecha_emision)[:10], "%Y-%m-%d").date()
        return dt + relativedelta(months=1)
    except Exception:
        return None


def _ensure_facturas_col16(ws):
    """Asegura que la columna 16 (P) tenga el encabezado 'TOTAL FACTURA'."""
    if ws.cell(row=1, column=16).value not in ("TOTAL FACTURA", "Total Factura"):
        cell = ws.cell(row=1, column=16, value="TOTAL FACTURA")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        ws.column_dimensions["P"].width = 16


def append_to_excel(items: list) -> bool:
    """Escribe una lista de ítems en el Excel. Devuelve True si OK."""
    try:
        wb = _open_wb()
        ws = wb[SHEET_NAME]
        _ensure_facturas_col16(ws)

        first_new_row = ws.max_row + 1

        # Anchor: Total Factura del documento (mismo en todos los ítems)
        total_factura_anchor = float((items[0].get("Total Factura") or 0)) if items else 0
        total_factura = 0.0

        for item in items:
            item = _compute_totals(item)
            total_factura += float(item.get("Monto / TOTAL") or 0)

            # Fecha vencimiento auto si no viene
            fecha_venc = item.get("Fecha Vencimiento") or _vencimiento(item.get("Fecha Emision"))

            row = [None] * TOTAL_COLS
            row[0]  = item.get("Fecha Emision")
            row[1]  = fecha_venc
            row[2]  = item.get("Fecha Pago")
            row[3]  = item.get("Nombre Factura / Proveedor")
            row[4]  = item.get("Rut")
            # Documento + referencia para NC/ND
            doc_tipo = item.get("Documento", "Factura")
            ref_fac = item.get("Referencia Factura")
            if ref_fac:
                doc_tipo = f"{doc_tipo} (Ref. Fac. {ref_fac})"
            row[5]  = doc_tipo
            nro_doc = item.get("Numero Factura / Nro Documento")
            if nro_doc is not None and str(nro_doc).strip().isdigit():
                row[6] = int(str(nro_doc).strip())
            else:
                row[6] = nro_doc
            row[7]  = item.get("Detalle / Glosa")
            row[8]  = item.get("Glosa II")
            row[9]  = item.get("Valor unitario")  # valor hardcodeado (fuente de verdad)
            row[10] = item.get("Cantidad")
            row[11] = None  # TOTAL NETO: fórmula después del append
            row[12] = item.get("IVA")
            row[13] = item.get("Impuesto Especifico")
            row[14] = item.get("Monto / TOTAL")
            # row[15] = Total Factura (se llena después del loop)

            ws.append(row)

            cur_row = ws.max_row

            # TOTAL NETO (col L) = Valor unitario × Cantidad — sin IVA, fórmula dinámica
            cell_neto = ws.cell(row=cur_row, column=12)
            cell_neto.value = f"=J{cur_row}*K{cur_row}"
            cell_neto.number_format = '#,##0'

        # Si Claude leyó el total del documento, tiene prioridad sobre la suma de líneas
        if total_factura_anchor > 0:
            total_factura = total_factura_anchor

        # Columna "TOTAL FACTURA" (col 16): mismo valor en cada fila del grupo
        total_factura = round(total_factura)
        last_new_row = ws.max_row
        for row_num in range(first_new_row, last_new_row + 1):
            cell = ws.cell(row=row_num, column=16)
            cell.value = total_factura
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")

        # N° de archivo físico (correlativo de FXP) — para ubicar el papel impreso
        try:
            from modules.correlativo import (correlativo_para, asegurar_columna,
                                              COL_CORRELATIVO)
            asegurar_columna(ws)
            n_arch = correlativo_para(items[0].get("Nombre Factura / Proveedor"),
                                      items[0].get("Numero Factura / Nro Documento"))
            if n_arch:
                for row_num in range(first_new_row, last_new_row + 1):
                    ws.cell(row=row_num, column=COL_CORRELATIVO).value = n_arch
                logger.info(f"N° de archivo asignado: {n_arch}")
        except Exception as e:
            logger.warning(f"No pude asignar el N° de archivo: {e}")

        _save_wb(wb)
        logger.info(f"Excel actualizado con {len(items)} fila(s). Total Factura: ${total_factura:,.0f}")
        return True

    except Exception as e:
        logger.error(f"Error al escribir en Excel: {e}")
        return False


def buscar_factura(nro_factura: str) -> list:
    """Busca facturas por número. Devuelve lista de dicts con fila y datos."""
    resultados = []
    try:
        wb = _open_wb()
        ws = wb[SHEET_NAME]
        nro_clean = str(nro_factura).strip()
        for row_idx in range(2, ws.max_row + 1):
            cell_nro = ws.cell(row=row_idx, column=7).value  # col 7 = Numero Factura
            if cell_nro is not None and str(cell_nro).strip() == nro_clean:
                resultados.append({
                    "fila": row_idx,
                    "fecha_emision": ws.cell(row=row_idx, column=1).value,
                    "proveedor": ws.cell(row=row_idx, column=4).value,
                    "rut": ws.cell(row=row_idx, column=5).value,
                    "documento": ws.cell(row=row_idx, column=6).value,
                    "nro_factura": str(cell_nro).strip(),
                    "glosa": ws.cell(row=row_idx, column=8).value,
                    "total": ws.cell(row=row_idx, column=15).value,
                    "fecha_pago": ws.cell(row=row_idx, column=3).value,
                })
        wb.close()
    except Exception as e:
        logger.error(f"Error buscando factura: {e}")
    return resultados


def registrar_pago(nro_factura: str, fecha_pago: str, medio: str = "") -> dict:
    """Actualiza Fecha Pago (col 3) en todas las filas con ese N° factura.
    medio puede ser 'Banco' o 'Caja Chica'.
    Si medio es 'Banco', intenta calzar automáticamente con un movimiento bancario.
    Devuelve dict con {actualizadas, calce}."""
    resultado = {"actualizadas": 0, "calce": None}
    try:
        wb = _open_wb()
        ws = wb[SHEET_NAME]
        nro_clean = str(nro_factura).strip()
        valor_pago = f"{fecha_pago} ({medio})" if medio else fecha_pago

        # Recopilar info de la factura para calce
        total_factura = 0
        proveedor_factura = ""
        for row_idx in range(2, ws.max_row + 1):
            cell_nro = ws.cell(row=row_idx, column=7).value
            if cell_nro is not None and str(cell_nro).strip() == nro_clean:
                ws.cell(row=row_idx, column=3).value = valor_pago
                resultado["actualizadas"] += 1
                total_factura += float(ws.cell(row=row_idx, column=15).value or 0)
                if not proveedor_factura:
                    proveedor_factura = str(ws.cell(row=row_idx, column=4).value or "")

        if resultado["actualizadas"] > 0:
            # Calce automático con banco
            if medio == "Banco" and BANCO_SHEET in wb.sheetnames:
                calce = _calzar_pago_banco(wb[BANCO_SHEET], total_factura, proveedor_factura, fecha_pago)
                resultado["calce"] = calce

            _save_wb(wb)
            logger.info(f"Pago registrado: Factura {nro_factura}, fecha {fecha_pago}, medio {medio}, "
                        f"{resultado['actualizadas']} fila(s), calce: {resultado['calce']}")
        wb.close()
    except Exception as e:
        logger.error(f"Error registrando pago: {e}")
    return resultado


def _calzar_pago_banco(ws_banco, total_factura: float, proveedor: str, fecha_pago: str) -> dict | None:
    """Busca en Cuenta Banco un movimiento que calce con la factura pagada.
    Criterios de calce (en orden de prioridad):
    1. Monto exacto + fecha cercana (+-5 días)
    2. Monto exacto en cualquier fecha
    3. Nombre proveedor similar + monto cercano (+-5%)
    Retorna dict con info del movimiento calzado o None."""
    if total_factura <= 0:
        return None

    total_int = round(total_factura)
    prov_lower = proveedor.lower().strip()
    # Extraer palabras clave del proveedor (>3 chars)
    prov_words = [w for w in prov_lower.split() if len(w) > 3]

    # Parsear fecha de pago
    try:
        fecha_dt = datetime.strptime(str(fecha_pago)[:10], "%Y-%m-%d").date()
    except Exception:
        fecha_dt = None

    candidatos = []
    for row_idx in range(2, ws_banco.max_row + 1):
        fecha_banco = ws_banco.cell(row=row_idx, column=1).value
        desc_banco = str(ws_banco.cell(row=row_idx, column=2).value or "").strip()
        cargo = ws_banco.cell(row=row_idx, column=4).value
        if not cargo:
            continue
        try:
            cargo_val = round(float(cargo))
        except (TypeError, ValueError):
            continue

        # Parsear fecha del banco
        fecha_banco_dt = None
        if isinstance(fecha_banco, (date, datetime)):
            fecha_banco_dt = fecha_banco if isinstance(fecha_banco, date) else fecha_banco.date()

        # Calcular score
        score = 0
        diff_monto = abs(cargo_val - total_int)
        diff_dias = None
        if fecha_dt and fecha_banco_dt:
            diff_dias = abs((fecha_banco_dt - fecha_dt).days)

        # Monto exacto
        if diff_monto == 0:
            score += 100
            if diff_dias is not None and diff_dias <= 5:
                score += 50  # fecha cercana bonus
        elif diff_monto <= total_int * 0.05:  # dentro del 5%
            score += 30

        # Nombre proveedor match
        desc_lower = desc_banco.lower()
        if prov_words:
            matches = sum(1 for w in prov_words if w in desc_lower)
            if matches > 0:
                score += 40 * matches / len(prov_words)

        if score >= 30:
            candidatos.append({
                "fila": row_idx,
                "fecha": str(fecha_banco_dt or fecha_banco),
                "descripcion": desc_banco,
                "cargo": cargo_val,
                "score": score,
                "diff_monto": diff_monto,
                "diff_dias": diff_dias,
            })

    if not candidatos:
        return None

    # Retornar el mejor candidato
    candidatos.sort(key=lambda x: x["score"], reverse=True)
    mejor = candidatos[0]
    return {
        "fila_banco": mejor["fila"],
        "fecha_banco": mejor["fecha"],
        "descripcion_banco": mejor["descripcion"],
        "cargo_banco": mejor["cargo"],
        "score": mejor["score"],
        "exacto": mejor["diff_monto"] == 0,
        "total_candidatos": len(candidatos),
    }


def _create_sheet_with_headers(wb, sheet_name, headers, widths, color="4472C4"):
    """Crea una hoja con encabezados formateados si no existe."""
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        logger.info(f"Hoja '{sheet_name}' creada en el Excel")
    return wb[sheet_name]


def _ensure_boletas_sheet(wb):
    widths = [12, 30, 14, 40, 15, 12, 30]
    return _create_sheet_with_headers(wb, BOLETAS_SHEET, BOLETAS_HEADERS, widths)


def _ensure_caja_chica_sheet(wb):
    widths = [12, 14, 40, 28, 12, 15, 15, 15]
    return _create_sheet_with_headers(wb, CAJA_CHICA_SHEET, CAJA_HEADERS, widths, color="2E7D32")


def _get_saldo_caja_chica(ws) -> float:
    """Calcula el saldo actual sumando ingresos y restando egresos."""
    saldo = 0.0
    for row_idx in range(2, ws.max_row + 1):
        ingreso = ws.cell(row=row_idx, column=6).value or 0
        egreso = ws.cell(row=row_idx, column=7).value or 0
        try:
            saldo += float(ingreso) - float(egreso)
        except (TypeError, ValueError):
            pass
    return saldo


def append_boleta(items: list) -> bool:
    """Escribe boletas en hoja Boletas Y registra gasto en Caja Chica."""
    try:
        wb = _open_wb()
        ws_bol = _ensure_boletas_sheet(wb)
        ws_caja = _ensure_caja_chica_sheet(wb)

        total_gasto = 0
        for item in items:
            monto = float(item.get("Monto / TOTAL") or 0)
            total_gasto += monto

            # Hoja Boletas (registro detallado)
            ws_bol.append([
                item.get("Fecha Emision"),
                item.get("Nombre Factura / Proveedor"),
                item.get("Rut"),
                item.get("Detalle / Glosa"),
                monto,
                item.get("Numero Factura / Nro Documento"),
                item.get("Glosa II"),
            ])

        # Hoja Caja Chica (una línea de egreso por el total de la boleta)
        saldo = _get_saldo_caja_chica(ws_caja) - total_gasto
        first = items[0]
        ws_caja.append([
            first.get("Fecha Emision"),
            "Gasto",
            first.get("Detalle / Glosa"),
            first.get("Nombre Factura / Proveedor"),
            first.get("Numero Factura / Nro Documento"),
            None,           # Ingreso
            total_gasto,    # Egreso
            saldo,          # Saldo actualizado
        ])

        _save_wb(wb)
        logger.info(f"Boleta guardada: {len(items)} fila(s) en Boletas + Caja Chica. Saldo: ${saldo:,.0f}")
        return True
    except Exception as e:
        logger.error(f"Error al guardar boleta: {e}")
        return False


def registrar_deposito_caja(fecha: str, monto: float, detalle: str = "Depósito caja chica") -> float:
    """Registra un depósito en Caja Chica. Devuelve el nuevo saldo."""
    try:
        wb = _open_wb()
        ws = _ensure_caja_chica_sheet(wb)
        saldo = _get_saldo_caja_chica(ws) + monto
        ws.append([
            fecha,
            "Depósito",
            detalle,
            None,       # Comercio
            None,       # Nº Boleta
            monto,      # Ingreso
            None,       # Egreso
            saldo,      # Saldo
        ])
        _save_wb(wb)
        logger.info(f"Depósito caja chica: ${monto:,.0f}. Nuevo saldo: ${saldo:,.0f}")
        return saldo
    except Exception as e:
        logger.error(f"Error registrando depósito: {e}")
        return -1


def consultar_saldo_caja() -> dict:
    """Consulta el estado actual de la caja chica."""
    try:
        wb = _open_wb()
        ws = _ensure_caja_chica_sheet(wb)
        total_ingresos = 0.0
        total_egresos = 0.0
        n_gastos = 0
        ultimo_deposito = None
        for row_idx in range(2, ws.max_row + 1):
            tipo = ws.cell(row=row_idx, column=2).value
            ingreso = float(ws.cell(row=row_idx, column=6).value or 0)
            egreso = float(ws.cell(row=row_idx, column=7).value or 0)
            total_ingresos += ingreso
            total_egresos += egreso
            if tipo == "Depósito":
                ultimo_deposito = ws.cell(row=row_idx, column=1).value
            if tipo == "Gasto":
                n_gastos += 1
        wb.close()
        return {
            "saldo": total_ingresos - total_egresos,
            "total_ingresos": total_ingresos,
            "total_egresos": total_egresos,
            "n_gastos": n_gastos,
            "ultimo_deposito": ultimo_deposito,
        }
    except Exception as e:
        logger.error(f"Error consultando saldo: {e}")
        return {"saldo": 0, "total_ingresos": 0, "total_egresos": 0, "n_gastos": 0, "ultimo_deposito": None}


def delete_last_boletas(n: int) -> bool:
    """Elimina las últimas n filas de la hoja Boletas y la última de Caja Chica."""
    try:
        wb = _open_wb()
        # Borrar de Boletas
        if BOLETAS_SHEET in wb.sheetnames:
            ws = wb[BOLETAS_SHEET]
            for row in range(ws.max_row, max(1, ws.max_row - n), -1):
                ws.delete_rows(row)
        # Borrar última línea de Caja Chica (el gasto correspondiente)
        if CAJA_CHICA_SHEET in wb.sheetnames:
            ws_caja = wb[CAJA_CHICA_SHEET]
            if ws_caja.max_row > 1:
                ws_caja.delete_rows(ws_caja.max_row)
        _save_wb(wb)
        logger.info(f"Eliminadas {n} fila(s) de Boletas + 1 de Caja Chica.")
        return True
    except Exception as e:
        logger.error(f"Error al eliminar boletas: {e}")
        return False


def delete_last_rows(n: int) -> bool:
    """Elimina las últimas n filas de la hoja Facturas."""
    try:
        wb = _open_wb()
        ws = wb[SHEET_NAME]
        max_row = ws.max_row
        for row in range(max_row, max_row - n, -1):
            ws.delete_rows(row)
        _save_wb(wb)
        logger.info(f"Eliminadas {n} fila(s) del Excel.")
        return True
    except Exception as e:
        logger.error(f"Error al eliminar filas del Excel: {e}")
        return False


# ─────────────────────────────────────────────
# CUENTA BANCO
# ─────────────────────────────────────────────
def _ensure_banco_sheet(wb):
    widths = [12, 40, 18, 15, 15, 15]
    return _create_sheet_with_headers(wb, BANCO_SHEET, BANCO_HEADERS, widths, color="1565C0")


def crear_hoja_banco():
    """Crea la hoja Cuenta Banco si no existe."""
    try:
        wb = _open_wb()
        _ensure_banco_sheet(wb)
        _save_wb(wb)
        return True
    except Exception as e:
        logger.error(f"Error creando hoja banco: {e}")
        return False


def guardar_movimientos_banco(movimientos: list[dict]) -> dict:
    """
    Guarda movimientos bancarios en la hoja Cuenta Banco.
    Evita duplicados comparando fecha+descripcion+cargo+abono.
    Retorna dict con {nuevos, duplicados, total}.
    """
    try:
        wb = _open_wb()
        ws = _ensure_banco_sheet(wb)

        def _fecha_iso(v):
            """Normaliza a 'YYYY-MM-DD' venga como date, datetime o texto."""
            if isinstance(v, datetime):
                return v.date().isoformat()
            if isinstance(v, date):
                return v.isoformat()
            s = str(v or "").strip()
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
                try:
                    return datetime.strptime(s[:10], fmt).date().isoformat()
                except ValueError:
                    continue
            return s

        def _doc(v):
            s = str(v or "").strip()
            return s[:-2] if s.endswith(".0") else s

        # Movimientos existentes: por N° de documento (único) y por
        # (fecha, cargo, abono). La fecha se normaliza en AMBOS lados: si se
        # compara texto contra fecha, nada calza y se reinsertan todos.
        docs_existentes = set()
        existentes = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            ref = _doc(row[2] if len(row) > 2 else "")
            if ref:
                docs_existentes.add(ref)
            existentes.add((_fecha_iso(row[0]),
                            int(round(float(row[3] or 0))),
                            int(round(float(row[4] or 0)))))

        nuevos = 0
        duplicados = 0
        for mov in movimientos:
            ref = _doc(mov.get("referencia"))
            if ref and ref in docs_existentes:
                duplicados += 1
                continue
            key = (_fecha_iso(mov.get("fecha")),
                   int(round(float(mov.get("cargo", 0) or 0))),
                   int(round(float(mov.get("abono", 0) or 0))))
            if key in existentes:
                duplicados += 1
                continue

            ws.append([
                mov.get("fecha", ""),
                mov.get("descripcion", ""),
                mov.get("referencia", ""),
                mov.get("cargo", 0) or None,
                mov.get("abono", 0) or None,
                mov.get("saldo", 0) or None,
            ])
            existentes.add(key)
            if ref:
                docs_existentes.add(ref)
            nuevos += 1

        _save_wb(wb)
        logger.info(f"Banco: {nuevos} nuevos, {duplicados} duplicados")
        return {"nuevos": nuevos, "duplicados": duplicados, "total": nuevos + duplicados}

    except Exception as e:
        logger.error(f"Error guardando movimientos banco: {e}")
        raise


def obtener_resumen_banco() -> dict:
    """Retorna resumen de la hoja Cuenta Banco."""
    try:
        wb = _open_wb()
        ws = _ensure_banco_sheet(wb)

        total_cargos = 0
        total_abonos = 0
        n_movimientos = 0
        ultimo_saldo = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            n_movimientos += 1
            total_cargos += float(row[3] or 0)
            total_abonos += float(row[4] or 0)
            if row[5] is not None:
                ultimo_saldo = float(row[5])

        return {
            "n_movimientos": n_movimientos,
            "total_cargos": total_cargos,
            "total_abonos": total_abonos,
            "ultimo_saldo": ultimo_saldo,
        }
    except Exception as e:
        logger.error(f"Error resumen banco: {e}")
        return {"n_movimientos": 0, "total_cargos": 0, "total_abonos": 0, "ultimo_saldo": None}


# ─────────────────────────────────────────────
# REPORTES
# ─────────────────────────────────────────────
def _parse_date(val) -> date:
    """Convierte un valor de celda a date."""
    if isinstance(val, (date, datetime)):
        return val if isinstance(val, date) else val.date()
    if isinstance(val, str):
        # Intentar varios formatos
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try: return datetime.strptime(val[:10], fmt).date()
            except: pass
        # Si tiene "(Banco)" o "(Caja Chica)" en fecha pago
        clean = val.split("(")[0].strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try: return datetime.strptime(clean[:10], fmt).date()
            except: pass
    return None


def _leer_facturas() -> list:
    """Lee todas las facturas del Excel como lista de dicts."""
    facturas = []
    try:
        wb = _open_wb()
        ws = wb[SHEET_NAME]
        for row_idx in range(2, ws.max_row + 1):
            proveedor = ws.cell(row=row_idx, column=4).value
            if not proveedor: continue
            facturas.append({
                "fila": row_idx,
                "fecha_emision": ws.cell(row=row_idx, column=1).value,
                "fecha_vencimiento": ws.cell(row=row_idx, column=2).value,
                "fecha_pago": ws.cell(row=row_idx, column=3).value,
                "proveedor": proveedor,
                "rut": ws.cell(row=row_idx, column=5).value,
                "documento": ws.cell(row=row_idx, column=6).value,
                "nro_factura": ws.cell(row=row_idx, column=7).value,
                "glosa": ws.cell(row=row_idx, column=8).value,
                "total": ws.cell(row=row_idx, column=15).value,
            })
        wb.close()
    except Exception as e:
        logger.error(f"Error leyendo facturas: {e}")
    return facturas


def reporte_diario() -> dict:
    """Genera reporte diario: vencen hoy + ya vencidas sin pagar."""
    hoy = date.today()
    facturas = _leer_facturas()
    vencen_hoy = []
    vencidas = []
    for f in facturas:
        if f.get("fecha_pago"): continue  # ya pagada
        venc = _parse_date(f.get("fecha_vencimiento"))
        if not venc: continue
        if venc == hoy:
            vencen_hoy.append(f)
        elif venc < hoy:
            vencidas.append(f)
    vencidas.sort(key=lambda x: _parse_date(x.get("fecha_vencimiento")) or hoy)
    return {
        "fecha": hoy.strftime("%d/%m/%Y"),
        "vencen_hoy": vencen_hoy,
        "vencidas": vencidas,
        "total_vencido": sum(float(f.get("total") or 0) for f in vencidas),
        "total_hoy": sum(float(f.get("total") or 0) for f in vencen_hoy),
    }


def reporte_semanal() -> dict:
    """Genera reporte semanal: pagadas esta semana + vencidas + caja chica."""
    hoy = date.today()
    inicio_semana = hoy - relativedelta(days=hoy.weekday())  # Lunes
    facturas = _leer_facturas()

    pagadas_semana = []
    vencidas = []
    for f in facturas:
        fp = f.get("fecha_pago")
        venc = _parse_date(f.get("fecha_vencimiento"))
        if fp:
            fecha_p = _parse_date(fp)
            if fecha_p and fecha_p >= inicio_semana:
                pagadas_semana.append(f)
        else:
            if venc and venc <= hoy:
                vencidas.append(f)

    # Gastos caja chica de la semana
    gastos_caja = []
    try:
        wb = _open_wb()
        if CAJA_CHICA_SHEET in wb.sheetnames:
            ws = wb[CAJA_CHICA_SHEET]
            for row_idx in range(2, ws.max_row + 1):
                tipo = ws.cell(row=row_idx, column=2).value
                if tipo != "Gasto": continue
                fecha_g = _parse_date(ws.cell(row=row_idx, column=1).value)
                if fecha_g and fecha_g >= inicio_semana:
                    gastos_caja.append({
                        "fecha": ws.cell(row=row_idx, column=1).value,
                        "detalle": ws.cell(row=row_idx, column=3).value,
                        "comercio": ws.cell(row=row_idx, column=4).value,
                        "monto": float(ws.cell(row=row_idx, column=7).value or 0),
                    })
        wb.close()
    except Exception as e:
        logger.error(f"Error leyendo caja chica semanal: {e}")

    return {
        "periodo": f"{inicio_semana.strftime('%d/%m')} - {hoy.strftime('%d/%m/%Y')}",
        "pagadas": pagadas_semana,
        "vencidas": vencidas,
        "gastos_caja": gastos_caja,
        "total_pagado": sum(float(f.get("total") or 0) for f in pagadas_semana),
        "total_vencido": sum(float(f.get("total") or 0) for f in vencidas),
        "total_caja": sum(g.get("monto", 0) for g in gastos_caja),
    }


def reporte_mensual(mes: int = None, anio: int = None) -> dict:
    """Genera reporte mensual completo."""
    hoy = date.today()
    if not mes: mes = hoy.month
    if not anio: anio = hoy.year
    facturas = _leer_facturas()

    emitidas_mes = []
    pagadas_mes = []
    vencidas_sin_pago = []
    gastos_por_proveedor = {}

    for f in facturas:
        fe = _parse_date(f.get("fecha_emision"))
        fp_raw = f.get("fecha_pago")
        fv = _parse_date(f.get("fecha_vencimiento"))
        total = float(f.get("total") or 0)
        prov = f.get("proveedor") or "Desconocido"

        # Emitidas en el mes
        if fe and fe.month == mes and fe.year == anio:
            emitidas_mes.append(f)
            gastos_por_proveedor[prov] = gastos_por_proveedor.get(prov, 0) + total

        # Pagadas en el mes
        if fp_raw:
            fp = _parse_date(fp_raw)
            if fp and fp.month == mes and fp.year == anio:
                pagadas_mes.append(f)

        # Vencidas sin pago
        if not fp_raw and fv and fv <= hoy:
            vencidas_sin_pago.append(f)

    # Top proveedores por gasto
    top_proveedores = sorted(gastos_por_proveedor.items(), key=lambda x: x[1], reverse=True)[:10]

    # Gastos caja chica del mes
    total_caja_mes = 0
    n_gastos_caja = 0
    try:
        wb = _open_wb()
        if CAJA_CHICA_SHEET in wb.sheetnames:
            ws = wb[CAJA_CHICA_SHEET]
            for row_idx in range(2, ws.max_row + 1):
                tipo = ws.cell(row=row_idx, column=2).value
                if tipo != "Gasto": continue
                fg = _parse_date(ws.cell(row=row_idx, column=1).value)
                if fg and fg.month == mes and fg.year == anio:
                    total_caja_mes += float(ws.cell(row=row_idx, column=7).value or 0)
                    n_gastos_caja += 1
        wb.close()
    except: pass

    MESES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    return {
        "periodo": f"{MESES[mes]} {anio}",
        "total_emitidas": len(emitidas_mes),
        "monto_emitido": sum(float(f.get("total") or 0) for f in emitidas_mes),
        "total_pagadas": len(pagadas_mes),
        "monto_pagado": sum(float(f.get("total") or 0) for f in pagadas_mes),
        "vencidas_sin_pago": len(vencidas_sin_pago),
        "monto_vencido": sum(float(f.get("total") or 0) for f in vencidas_sin_pago),
        "top_proveedores": top_proveedores,
        "total_caja_mes": total_caja_mes,
        "n_gastos_caja": n_gastos_caja,
    }


def _coerce_date(v):
    """Convierte valor de celda a date una sola vez."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(v[:10], fmt).date()
            except ValueError:
                pass
    return None


def read_facturas_pendientes(excel_path=None) -> list[dict]:
    """Lee Master.Facturas, devuelve facturas con Fecha Pago vacia.

    Usa iter_rows() para evitar slow random-access en read_only mode.
    """
    excel_path = excel_path or EXCEL_PATH
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    pendientes = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=15,
                                              values_only=True), start=2):
        # row index: 0=fecha emision, 1=fecha venc, 2=fecha pago, 3=proveedor,
        # 4=rut, 5=documento, 6=nro factura, 7=glosa, 14=total
        proveedor = row[3]
        if not proveedor:
            continue
        if row[2]:  # Fecha Pago
            continue
        pendientes.append({
            "fila": r_idx,
            "fecha_emision": _coerce_date(row[0]),
            "fecha_vencimiento": _coerce_date(row[1]),
            "proveedor": proveedor,
            "rut": row[4],
            "nro_factura": row[6],
            "glosa": row[7],
            "total": row[14],
        })
    wb.close()
    return pendientes


def read_bank_movements_unlinked(excel_path=None) -> list[dict]:
    """Lee Cuenta Banco, devuelve cargos sin Factura_linkeada.

    Usa iter_rows() para evitar slow random-access.
    """
    excel_path = excel_path or EXCEL_PATH
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[CUENTA_BANCO_SHEET]
    movs = []
    # Lee hasta col J (10) = Factura_linkeada
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=10,
                                              values_only=True), start=2):
        fecha = row[0]
        if not fecha:
            continue
        cargo_v = row[3]
        try:
            cargo = float(cargo_v or 0)
        except (TypeError, ValueError):
            continue
        if cargo <= 0:
            continue
        if row[9]:  # Factura_linkeada
            continue
        try:
            abono = float(row[4] or 0)
        except (TypeError, ValueError):
            abono = 0.0
        movs.append({
            "fila": r_idx,
            "fecha": _coerce_date(fecha) or fecha,
            "descripcion": row[1] or "",
            "referencia": row[2] or "",
            "cargo": cargo,
            "abono": abono,
        })
    wb.close()
    return movs


def apply_bank_factura_link(bank_row: int, factura_row: int,
                              nro_factura: str, fecha_pago: str,
                              excel_path=None) -> dict:
    """Escribe Fecha Pago en Facturas y Tipo+Factura_linkeada en Cuenta Banco.

    Si Fecha Pago ya tiene valor, NO sobrescribe (fecha_pago_skipped=True).
    """
    excel_path = excel_path or EXCEL_PATH
    result = {"fecha_pago_skipped": False, "linked": False}
    wb = load_workbook(excel_path)

    ws_f = wb[SHEET_NAME]
    if ws_f.cell(factura_row, 3).value:
        result["fecha_pago_skipped"] = True
    else:
        ws_f.cell(factura_row, 3, f"{fecha_pago} (Banco)")

    ws_b = wb[CUENTA_BANCO_SHEET]
    ws_b.cell(bank_row, COL_BANCO_TIPO, "factura")
    ws_b.cell(bank_row, COL_BANCO_FACTURA_LINK, str(nro_factura))
    result["linked"] = True

    _save_wb(wb, excel_path)
    wb.close()
    return result
