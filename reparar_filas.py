"""
reparar_filas.py
Script de reparación: detecta facturas con líneas faltantes desde fila 266
y las re-procesa usando las imágenes guardadas.

Uso:
    py -3.11 reparar_filas.py            → modo revisión (solo muestra qué falta)
    py -3.11 reparar_filas.py --reparar  → modo reparación (modifica el Excel)
"""
import os, sys, glob, shutil, logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment

logging.basicConfig(format="%(asctime)s %(levelname)s %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

# ── Configuración ──────────────────────────────────────────────────────────────
EXCEL_PATH    = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                             "MASTER Agricola Santa Elisa.xlsx")
DOWNLOAD_DIR  = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                             "Facturas Recibidas", "Facturas Recibidas por telegram")
SHEET_NAME    = "Facturas"
PRIMERA_FILA_AFECTADA = 266   # filas desde aquí en adelante pueden tener el bug
MODO_REPARAR  = "--reparar" in sys.argv

# ── Paso 1: Leer Excel y detectar facturas con líneas faltantes ────────────────
def leer_facturas_afectadas():
    """
    Lee las filas desde PRIMERA_FILA_AFECTADA y detecta facturas sospechosas:
    - Una sola fila por número de factura
    - Item (col 15) != Total Factura (col 16)
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # Leer TODAS las filas para contar cuántas líneas tiene cada factura
    facturas_todo = {}  # key = (nro_factura, rut) → lista de filas
    for row in range(2, ws.max_row + 1):
        nro  = ws.cell(row=row, column=7).value
        rut  = ws.cell(row=row, column=5).value
        if nro is None and rut is None:
            continue
        key = (str(nro or "").strip(), str(rut or "").strip())
        if key not in facturas_todo:
            facturas_todo[key] = []
        facturas_todo[key].append(row)

    # Filtrar: facturas afectadas (desde fila PRIMERA_FILA_AFECTADA, 1 sola línea, item ≠ total)
    afectadas = []
    vistas = set()
    for row in range(PRIMERA_FILA_AFECTADA, ws.max_row + 1):
        nro        = ws.cell(row=row, column=7).value
        rut        = ws.cell(row=row, column=5).value
        proveedor  = ws.cell(row=row, column=4).value
        glosa      = ws.cell(row=row, column=8).value
        total_item = ws.cell(row=row, column=15).value
        total_fac  = ws.cell(row=row, column=16).value
        fecha_em   = ws.cell(row=row, column=1).value

        if not proveedor:
            continue

        key = (str(nro or "").strip(), str(rut or "").strip())
        if key in vistas:
            continue
        vistas.add(key)

        filas_en_excel = facturas_todo.get(key, [row])
        n_filas = len(filas_en_excel)

        total_item_f = float(total_item or 0)
        total_fac_f  = float(total_fac  or 0)

        # Sospechosa: 1 sola línea y los totales no coinciden
        es_sospechosa = (n_filas == 1 and total_fac_f > 0 and
                         abs(total_item_f - total_fac_f) > 1.0)

        afectadas.append({
            "key":       key,
            "filas":     filas_en_excel,
            "nro":       str(nro or "").strip(),
            "rut":       str(rut or "").strip(),
            "proveedor": str(proveedor or "").strip(),
            "glosa":     str(glosa or ""),
            "fecha_em":  fecha_em,
            "total_item": total_item_f,
            "total_fac":  total_fac_f,
            "sospechosa": es_sospechosa,
        })

    wb.close()
    return afectadas


# ── Paso 2: Encontrar imágenes candidatas para cada factura ───────────────────
def _listar_imagenes():
    """Lista imágenes originales (sin _scan/_resized) ordenadas cronológicamente."""
    exts = ("*.jpg", "*.jpeg", "*.png", "*.pdf")
    archivos = []
    for ext in exts:
        archivos.extend(glob.glob(os.path.join(DOWNLOAD_DIR, ext)))
    # Excluir versiones procesadas
    archivos = [f for f in archivos
                if "_scan" not in os.path.basename(f).lower()
                and "_resized" not in os.path.basename(f).lower()]
    archivos.sort()  # orden cronológico por nombre (tiene timestamp)
    return archivos


# ── Paso 3: Re-procesar imagen con el extractor ───────────────────────────────
def _procesar_imagen(path):
    """Llama al extractor y devuelve la lista de items o None si falla."""
    try:
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        from processors.extractor import process_file
        result = process_file(path)
        if result.get("status") == "ok":
            return result.get("items", [])
        else:
            logger.warning(f"  Extractor error en {os.path.basename(path)}: {result.get('message')}")
            return None
    except Exception as e:
        logger.error(f"  Error procesando {os.path.basename(path)}: {e}")
        return None


# ── Paso 4: Reparar Excel ─────────────────────────────────────────────────────
def _backup_excel():
    backup = EXCEL_PATH.replace(".xlsx", f"_BACKUP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    shutil.copy2(EXCEL_PATH, backup)
    logger.info(f"Backup creado: {os.path.basename(backup)}")
    return backup


def _reemplazar_filas_en_excel(factura_info, items_nuevos):
    """
    Reemplaza las filas de una factura en el Excel con los nuevos items extraídos.
    Inserta filas adicionales si hay más items que los actuales.
    """
    from excel_manager import _compute_totals, _vencimiento, TOTAL_COLS, IVA_RATE

    wb  = openpyxl.load_workbook(EXCEL_PATH)
    ws  = wb[SHEET_NAME]

    filas_actuales = sorted(factura_info["filas"])
    primera_fila   = filas_actuales[0]
    n_actuales     = len(filas_actuales)
    n_nuevos       = len(items_nuevos)

    # Calcular total factura
    total_fac_anchor = float((items_nuevos[0].get("Total Factura") or 0)) if items_nuevos else 0
    total_fac_suma   = 0.0

    rows_data = []
    for item in items_nuevos:
        item = _compute_totals(item)
        total_fac_suma += float(item.get("Monto / TOTAL") or 0)

        fecha_venc = item.get("Fecha Vencimiento") or _vencimiento(item.get("Fecha Emision"))
        doc_tipo   = item.get("Documento", "Factura")
        ref_fac    = item.get("Referencia Factura")
        if ref_fac:
            doc_tipo = f"{doc_tipo} (Ref. Fac. {ref_fac})"

        row = [None] * TOTAL_COLS
        row[0]  = item.get("Fecha Emision")
        row[1]  = fecha_venc
        row[2]  = item.get("Fecha Pago")
        row[3]  = item.get("Nombre Factura / Proveedor")
        row[4]  = item.get("Rut")
        row[5]  = doc_tipo
        row[6]  = item.get("Numero Factura / Nro Documento")
        row[7]  = item.get("Detalle / Glosa")
        row[8]  = item.get("Glosa II")
        row[9]  = item.get("Valor unitario")
        row[10] = item.get("Cantidad")
        row[11] = item.get("TOTAL NETO")
        row[12] = item.get("IVA")
        row[13] = item.get("Impuesto Especifico")
        row[14] = item.get("Monto / TOTAL")
        row[15] = None  # se rellena después
        rows_data.append(row)

    total_fac = round(total_fac_anchor if total_fac_anchor > 0 else total_fac_suma)

    # Si necesitamos más filas que las actuales, insertar filas adicionales
    filas_extras = n_nuevos - n_actuales
    if filas_extras > 0:
        ws.insert_rows(primera_fila + n_actuales, amount=filas_extras)

    # Si necesitamos menos filas que las actuales, borrar las sobrantes
    if filas_extras < 0:
        for i in range(abs(filas_extras)):
            ws.delete_rows(primera_fila + n_nuevos)

    # Escribir los nuevos datos
    for i, row_data in enumerate(rows_data):
        fila_excel = primera_fila + i
        row_data[15] = total_fac
        for col_idx, valor in enumerate(row_data, 1):
            ws.cell(row=fila_excel, column=col_idx, value=valor)
        # Formato columna 16
        cell16 = ws.cell(row=fila_excel, column=16)
        cell16.number_format = '#,##0'
        cell16.alignment = Alignment(horizontal="right", vertical="center")

    wb.save(EXCEL_PATH)
    wb.close()
    logger.info(f"  [OK] Reemplazadas: {n_actuales} fila(s) -> {n_nuevos} fila(s) | Total: ${total_fac:,.0f}")


# ── Guardar progreso en JSON ──────────────────────────────────────────────────
_PROGRESO_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), "progreso_reparacion.json")

def _guardar_progreso(data):
    """Escribe el estado actual a un JSON para poder resumir si falla el proceso."""
    import json
    try:
        data["timestamp"] = datetime.now().isoformat()
        with open(_PROGRESO_JSON, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    except Exception as e:
        logger.warning(f"No se pudo guardar progreso: {e}")


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "="*70)
    print("  SCRIPT DE REPARACION - Facturas con lineas faltantes")
    print(f"  Excel: {os.path.basename(EXCEL_PATH)}")
    print(f"  Modo: {'[REPARAR] modificara el Excel' if MODO_REPARAR else '[REVISION] solo lectura'}")
    print("="*70 + "\n")

    # 1. Detectar facturas afectadas
    print("[1/4] Analizando Excel desde fila 266...\n")
    afectadas = leer_facturas_afectadas()

    sospechosas = [f for f in afectadas if f["sospechosa"]]
    normales    = [f for f in afectadas if not f["sospechosa"]]

    print(f"Total facturas unicas en rango: {len(afectadas)}")
    print(f"  [FALTA]  Sospechosas (lineas faltantes): {len(sospechosas)}")
    print(f"  [OK]     Parecen correctas:              {len(normales)}\n")

    if not sospechosas:
        print("[OK] No se detectaron facturas con lineas faltantes. Nada que reparar.")
        _guardar_progreso({"estado": "sin_sospechosas", "sospechosas": 0})
        return

    print("[FALTA] Facturas sospechosas:")
    lista_sosp = []
    for f in sospechosas:
        diff = f["total_fac"] - f["total_item"]
        print(f"   Fila {f['filas'][0]:>3} | {f['proveedor'][:35]:<35} | Nro:{f['nro']:<12} | "
              f"Item:${f['total_item']:>12,.0f} | Total:${f['total_fac']:>12,.0f} | Falta~${diff:,.0f}")
        lista_sosp.append({
            "fila": f["filas"][0], "nro": f["nro"], "proveedor": f["proveedor"],
            "total_item": f["total_item"], "total_fac": f["total_fac"],
        })
    _guardar_progreso({"estado": "revision", "sospechosas": len(sospechosas), "lista": lista_sosp})

    if not MODO_REPARAR:
        print("\n[INFO] Para reparar, ejecuta:")
        print("       py -3.11 reparar_filas.py --reparar\n")
        return

    # 2. Listar imagenes disponibles
    imagenes = _listar_imagenes()
    print(f"\n[2/4] Imagenes en carpeta: {len(imagenes)}")

    # 3. Hacer backup
    print("\n[3/4] Creando backup del Excel...")
    _backup_excel()

    # 4. Re-procesar imagenes y reparar
    print("\n[4/4] Iniciando reparacion...\n")
    reparadas  = 0

    # Mapear nro_factura -> datos de factura sospechosa
    mapa_sospechosas = {f["nro"]: f for f in sospechosas if f["nro"]}

    for idx_img, img_path in enumerate(imagenes, 1):
        nombre = os.path.basename(img_path)
        print(f"[{idx_img}/{len(imagenes)}] Procesando: {nombre}")

        items = _procesar_imagen(img_path)
        if not items:
            print("  [WARN] Sin resultado\n")
            continue

        nro_extraido = str(items[0].get("Numero Factura / Nro Documento") or "").strip()
        proveedor_extraido = str(items[0].get("Nombre Factura / Proveedor") or "").strip()

        if nro_extraido not in mapa_sospechosas:
            print(f"  [SKIP] Nro {nro_extraido} no esta en sospechosas.\n")
            continue

        factura_info = mapa_sospechosas[nro_extraido]
        n_extraidos  = len(items)
        n_en_excel   = len(factura_info["filas"])

        print(f"  Factura Nro {nro_extraido} - {proveedor_extraido}")
        print(f"  Items extraidos: {n_extraidos} | Items en Excel: {n_en_excel}")

        if n_extraidos <= n_en_excel:
            print(f"  [SKIP] No hay lineas faltantes (ya tiene {n_en_excel}).\n")
            continue

        print(f"  [FIX] Reemplazando {n_en_excel} fila(s) por {n_extraidos}...")
        try:
            _reemplazar_filas_en_excel(factura_info, items)
            reparadas += 1
            del mapa_sospechosas[nro_extraido]
            # Guardar progreso despues de cada reparacion
            _guardar_progreso({
                "estado": "reparando",
                "reparadas": reparadas,
                "pendientes": len(mapa_sospechosas),
                "ultima_reparada": {"nro": nro_extraido, "proveedor": proveedor_extraido},
            })
        except Exception as e:
            logger.error(f"  [ERROR] reparando factura {nro_extraido}: {e}")
        print()

    # Reporte final
    sin_imagen = len(mapa_sospechosas)
    print("="*70)
    print("[OK] REPARACION COMPLETA")
    print(f"   Facturas reparadas:          {reparadas}")
    print(f"   Sin imagen encontrada:       {sin_imagen}")
    pendientes_out = []
    if mapa_sospechosas:
        print("\n   [WARN] Estas facturas no se pudieron reparar automaticamente")
        print("          (re-envialas manualmente por Telegram):")
        for nro, f in mapa_sospechosas.items():
            print(f"   -> Nro {nro} | {f['proveedor'][:40]}")
            pendientes_out.append({"nro": nro, "proveedor": f["proveedor"], "fila": f["filas"][0]})
    print("="*70 + "\n")
    _guardar_progreso({
        "estado": "completado",
        "reparadas": reparadas,
        "sin_imagen": sin_imagen,
        "pendientes": pendientes_out,
    })


if __name__ == "__main__":
    main()
